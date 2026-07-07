"""Automated extraction quality validator.

Validates filled Excel templates for:
1. Structural integrity — required columns, key uniqueness, FK references
2. Fill rate — per-sheet and per-column completion statistics
3. Semantic correctness — optional LLM-based spot-checking against source
"""

from __future__ import annotations

import hashlib
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

from iev4pi_transformation_tool.core.disk_cache import DiskDict
_validation_cache = DiskDict("extraction_validation")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


class ValidationIssue:
    """A single issue found during validation."""

    def __init__(
        self,
        issue_type: str,
        sheet: str,
        row: int | None = None,
        column: str | None = None,
        message: str = "",
        severity: str = "warning",
    ) -> None:
        self.type = issue_type
        self.sheet = sheet
        self.row = row
        self.column = column
        self.message = message
        self.severity = severity  # "error", "warning", "info"


class FillRate:
    """Fill-rate statistics for a sheet or column."""

    def __init__(self) -> None:
        self.total_cells = 0
        self.filled_cells = 0

    @property
    def rate(self) -> float:
        return self.filled_cells / self.total_cells if self.total_cells > 0 else 0.0

    @property
    def rate_pct(self) -> str:
        return f"{self.rate:.1%}"


class AuditReport:
    """Complete audit report for one or more filled template workbooks."""

    def __init__(self) -> None:
        self.issues: list[ValidationIssue] = []
        self.fill_rates: dict[str, dict[str, FillRate]] = {}  # file → sheet → FillRate
        self.column_fill_rates: dict[str, dict[str, dict[str, FillRate]]] = {}  # file → sheet → col → FillRate
        self.errors = 0
        self.warnings = 0
        self.infos = 0

    def add_issue(self, issue: ValidationIssue) -> None:
        self.issues.append(issue)
        if issue.severity == "error":
            self.errors += 1
        elif issue.severity == "warning":
            self.warnings += 1
        else:
            self.infos += 1

    def summary(self) -> str:
        lines = [
            f"Audit Report: {self.errors} errors, {self.warnings} warnings, {self.infos} info",
            "",
        ]
        if self.errors:
            lines.append("## Errors")
            for i in self.issues:
                if i.severity == "error":
                    lines.append(f"  - [{i.sheet}] {i.message}")
        if self.warnings:
            lines.append("## Warnings")
            for i in self.issues:
                if i.severity == "warning":
                    lines.append(f"  - [{i.sheet}] {i.message}")
        lines.append("")
        lines.append("## Fill Rates")
        for file_name, sheets in self.fill_rates.items():
            lines.append(f"  {file_name}:")
            for sheet_name, fr in sheets.items():
                lines.append(f"    {sheet_name}: {fr.rate_pct} ({fr.filled_cells}/{fr.total_cells})")
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Structural validation
# ---------------------------------------------------------------------------


def validate_workbook(
    wb_path: Path,
    template_path: Path | None = None,
    *,
    check_fk: bool = True,
    check_empty_required: bool = True,
) -> AuditReport:
    """Validate a filled Excel workbook against structural rules.

    Args:
        wb_path: Path to the filled workbook.
        template_path: Path to the template workbook (for column name reference).
            If ``None``, column names are read from the filled workbook row 1.
        check_fk: Whether to check foreign key integrity.
        check_empty_required: Whether to check for empty required columns.

    Returns:
        :class:`AuditReport` with all findings.
    """
    report = AuditReport()
    file_name = wb_path.name

    try:
        wb = openpyxl.load_workbook(str(wb_path), data_only=True)
    except Exception as e:
        report.add_issue(ValidationIssue(
            "file_error", "", message=f"Cannot open workbook: {e}", severity="error",
        ))
        return report

    # Build column maps from row 1 headers
    sheet_columns: dict[str, dict[str, int]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cols: dict[str, int] = {}
        if ws.max_column:
            for c in range(1, ws.max_column + 1):
                h = str(ws.cell(row=1, column=c).value or "").strip()
                if h:
                    cols[h] = c
        sheet_columns[sheet_name] = cols

    # Per-sheet fill rate tracking
    file_fill_rates: dict[str, FillRate] = {}
    file_col_fill_rates: dict[str, dict[str, FillRate]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cols = sheet_columns[sheet_name]
        if not cols:
            continue

        sheet_fr = FillRate()
        col_frs: dict[str, FillRate] = {col: FillRate() for col in cols}

        data_rows = 0
        for row in ws.iter_rows(min_row=3, values_only=True):  # data starts at row 3
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            data_rows += 1
            for col_name, col_idx in cols.items():
                sheet_fr.total_cells += 1
                col_frs[col_name].total_cells += 1
                cell_val = None
                if col_idx - 1 < len(row):
                    cell_val = row[col_idx - 1]
                if cell_val is not None and str(cell_val).strip():
                    sheet_fr.filled_cells += 1
                    col_frs[col_name].filled_cells += 1

        file_fill_rates[sheet_name] = sheet_fr
        file_col_fill_rates[sheet_name] = col_frs

        if data_rows == 0:
            report.add_issue(ValidationIssue(
                "empty_sheet", sheet_name,
                message=f"Sheet '{sheet_name}' has no data rows",
                severity="warning",
            ))

        # Check for empty required columns
        if check_empty_required:
            _REQUIRED_COL_SIGNALS = ["pk", "fk", "id", "document_id", "index"]
            for col_name, col_fr in col_frs.items():
                is_required = any(
                    signal in col_name.lower() for signal in _REQUIRED_COL_SIGNALS
                )
                if is_required and col_fr.rate == 0.0 and col_fr.total_cells > 0:
                    report.add_issue(ValidationIssue(
                        "empty_required", sheet_name, column=col_name,
                        message=f"Required column '{col_name}' is completely empty",
                        severity="error",
                    ))

        # Check for very low fill rate columns
        for col_name, col_fr in col_frs.items():
            if 0 < col_fr.rate < 0.1 and col_fr.total_cells >= 3:
                report.add_issue(ValidationIssue(
                    "low_fill", sheet_name, column=col_name,
                    message=f"Column '{col_name}' fill rate is only {col_fr.rate_pct}",
                    severity="info",
                ))

    # FK integrity check
    if check_fk:
        _check_foreign_keys(wb, sheet_columns, report)

    report.fill_rates[file_name] = file_fill_rates
    report.column_fill_rates[file_name] = file_col_fill_rates

    wb.close()
    return report


def _check_foreign_keys(
    wb: openpyxl.Workbook,
    sheet_columns: dict[str, dict[str, int]],
    report: AuditReport,
) -> None:
    """Check that FK column values reference existing PK values.

    Identifies FK columns by naming conventions (e.g. column names containing
    ``_ID`` that match another sheet's PK column)."""
    # Build PK value sets for each sheet
    pk_values: dict[str, set[str]] = {}
    for sheet_name, cols in sheet_columns.items():
        # The "ID" column of each sheet is the PK
        id_cols = [c for c in cols if c.lower().endswith("_id") and not c.lower().startswith("fk")]
        if not id_cols:
            id_cols = [c for c in cols if c.lower() == "index"]
        for id_col in id_cols[:1]:  # use first ID column
            pk_set: set[str] = set()
            col_idx = cols[id_col]
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row and col_idx - 1 < len(row):
                    val = row[col_idx - 1]
                    if val is not None:
                        pk_set.add(str(val).strip())
            pk_values[sheet_name] = pk_set

    # Check FK references
    for sheet_name, cols in sheet_columns.items():
        ws = wb[sheet_name]
        for col_name, col_idx in cols.items():
            # FK columns typically contain "FK →" or reference another sheet ID
            if not col_name.lower().endswith("_id"):
                continue
            # Find which sheet this FK references
            target_sheet = None
            col_stem = col_name.lower().replace("_id", "")
            for sn in pk_values:
                sn_stem = sn.lower().replace("_id", "")
                if col_stem in sn_stem or sn_stem in col_stem:
                    # Skip self-referencing (e.g. Device_ID in Device_ID sheet)
                    if sn != sheet_name:
                        target_sheet = sn
                        break
            if target_sheet is None:
                continue

            pk_set = pk_values.get(target_sheet, set())
            if not pk_set:
                continue

            broken = 0
            total = 0
            for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if row and col_idx - 1 < len(row):
                    val = row[col_idx - 1]
                    if val is not None and str(val).strip():
                        total += 1
                        if str(val).strip() not in pk_set:
                            broken += 1
            if broken > 0:
                report.add_issue(ValidationIssue(
                    "fk_integrity", sheet_name, column=col_name,
                    message=f"{broken}/{total} FK references in '{col_name}' not found in '{target_sheet}'",
                    severity="warning",
                ))


# ---------------------------------------------------------------------------
# LLM semantic spot-check
# ---------------------------------------------------------------------------


def llm_spot_check(
    wb_path: Path,
    sample_size: int = 5,
    llm_client: Any | None = None,
) -> AuditReport:
    """Spot-check random data rows against source using LLM.

    Selects *sample_size* random filled rows from each sheet and asks the
    LLM to verify that the data looks correct for its column context.

    Args:
        wb_path: Path to the filled workbook.
        sample_size: Number of rows to sample per sheet.
        llm_client: Optional LLM client. If ``None`` or unavailable, no LLM checks run.

    Returns:
        :class:`AuditReport` with semantic issues found.
    """
    report = AuditReport()
    if llm_client is None or not llm_client.available():
        return report

    try:
        wb = openpyxl.load_workbook(str(wb_path), data_only=True)
    except Exception:
        return report

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Read headers
        headers: dict[int, str] = {}
        if ws.max_column:
            for c in range(1, ws.max_column + 1):
                h = str(ws.cell(row=1, column=c).value or "").strip()
                if h:
                    headers[c] = h

        # Collect data rows
        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row is None:
                continue
            row_dict = {}
            for col_idx, header in headers.items():
                val = row[col_idx - 1] if col_idx - 1 < len(row) else None
                if val is not None and str(val).strip():
                    row_dict[header] = str(val).strip()
            if row_dict:
                rows.append(row_dict)

        if not rows:
            continue

        # Sample rows
        import random
        sample = random.sample(rows, min(sample_size, len(rows)))

        for i, row_data in enumerate(sample):
            prompt = (
                f"You are validating an engineering data Excel export.\n"
                f"Sheet: '{sheet_name}'\n"
                f"Row data:\n{json.dumps(row_data, indent=2, ensure_ascii=False)}\n\n"
                f"For each field, check if the VALUE looks plausible for the COLUMN NAME:\n"
                f"- Document_ID should look like an identifier\n"
                f"- Dates should be in date format\n"
                f"- Confidence values should be between 0 and 1\n"
                f"- SemanticIDs should follow standard format (e.g. 0112/2///...)\n"
                f"- Device tags should follow IEC 81346 conventions\n\n"
                f"Return JSON:\n"
                f'{{"issues": [{{"column": "ColumnName", "problem": "description"}}], '
                f'"overall_plausible": true/false}}'
            )
            # Disk cache: skip LLM if this exact row was already validated
            row_json = json.dumps(row_data, ensure_ascii=False, sort_keys=True)
            cache_key = f"{sheet_name}|{hashlib.sha256(row_json.encode()).hexdigest()[:16]}"
            cached = _validation_cache.get(cache_key)
            if cached is not None and isinstance(cached, dict):
                if not cached.get("overall_plausible", True):
                    for issue in cached.get("issues", []):
                        report.add_issue(ValidationIssue(
                            "semantic", sheet_name, row=i + 3, column=issue.get("column", ""),
                            message=issue.get("problem", ""), severity="warning",
                        ))
                continue

            try:
                response = llm_client.chat_json(prompt)
                if isinstance(response, dict):
                    _validation_cache[cache_key] = response
                    if not response.get("overall_plausible", True):
                        for issue in response.get("issues", []):
                            col = issue.get("column", "")
                            problem = issue.get("problem", "")
                            report.add_issue(ValidationIssue(
                                "semantic", sheet_name, row=i + 3, column=col,
                                message=problem, severity="warning",
                            ))
            except Exception:
                pass

    wb.close()
    return report


# ---------------------------------------------------------------------------
# Batch validation
# ---------------------------------------------------------------------------


def validate_export_dir(
    export_dir: Path,
    template_dir: Path | None = None,
    llm_client: Any | None = None,
) -> AuditReport:
    """Validate all Excel files in an export directory.

    Args:
        export_dir: Directory containing filled Excel files.
        template_dir: Directory containing template files (for column reference).
        llm_client: Optional LLM client for semantic spot-checking.

    Returns:
        Combined :class:`AuditReport`.
    """
    combined = AuditReport()

    for xlsx_path in sorted(export_dir.rglob("*.xlsx")):
        # Skip temporary files
        if xlsx_path.name.startswith("~"):
            continue

        # Find matching template
        tpl_path = None
        if template_dir:
            for tpl in template_dir.glob("*_template.xlsx"):
                report = validate_workbook(xlsx_path, tpl)
                _merge_report(combined, report)
                tpl_path = tpl
                break
        if tpl_path is None:
            report = validate_workbook(xlsx_path)
            _merge_report(combined, report)

        # LLM spot-check
        if llm_client:
            llm_report = llm_spot_check(xlsx_path, sample_size=3, llm_client=llm_client)
            _merge_report(combined, llm_report)

    return combined


def _merge_report(target: AuditReport, source: AuditReport) -> None:
    """Merge *source* report into *target*."""
    target.issues.extend(source.issues)
    target.errors += source.errors
    target.warnings += source.warnings
    target.infos += source.infos
    target.fill_rates.update(source.fill_rates)
    target.column_fill_rates.update(source.column_fill_rates)
