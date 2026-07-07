from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from iev4pi_transformation_tool.core.utils import clean_cell, normalize_label
from iev4pi_transformation_tool.models import UC1CatalogCoverageReport, UC1CatalogCoverageRow, UC1CatalogEntry


UC1_SUPPORT_MAP: dict[str, dict[str, str]] = {
    "hasinstrumentationloopfunctionnumber": {
        "coverage_status": "guaranteed",
        "matched_field": "ri_devices.has_instrumentation_loop_function_number",
        "notes": "Directly exported from the canonical device identity row and mapped into SM_CoreIdentity.",
    },
    "deviceinformation": {
        "coverage_status": "guaranteed",
        "matched_field": "ri_devices.device_information",
        "notes": "Filled from the DEXPI loop/function description and carried into SM_FunctionAndVendor.",
    },
    "processinstrumentationfunctioncategory": {
        "coverage_status": "guaranteed",
        "matched_field": "ri_devices.process_instrumentation_function_category",
        "notes": "Derived from the function code and kept as a fixed UC1 identity property.",
    },
    "processinstrumentationfunctionmodifier": {
        "coverage_status": "guaranteed",
        "matched_field": "ri_devices.process_instrumentation_function_modifier",
        "notes": "Derived from the function code suffix and mapped into SM_CoreIdentity.",
    },
    "processinstrumentationfunctionnumber": {
        "coverage_status": "guaranteed",
        "matched_field": "ri_devices.process_instrumentation_function_number",
        "notes": "Exported as the canonical device number used across workbook, AAS, and ontology.",
    },
    "processinstrumentationfunctions": {
        "coverage_status": "guaranteed",
        "matched_field": "ri_devices.process_instrumentation_functions",
        "notes": "The function code is preserved end-to-end and exported into SM_CoreIdentity.",
    },
    "safetyrelevanceclass": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.safety_relevance_class",
        "notes": "The field is modeled end-to-end but depends on best-effort source availability in the current DEXPI/XML payload.",
    },
    "vendorcompanyname": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.vendor_company_name",
        "notes": "The field is part of the standardized workbook and AAS but is usually blank unless vendor data exists in the source package.",
    },
    "actuatingfunctions": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.actuating_function_number",
        "notes": "Actuation relations are modeled but only populated when DEXPI evidence exposes them explicitly.",
    },
    "actuatingfunctionnumber": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.actuating_function_number",
        "notes": "The column exists end-to-end, yet current extraction is best-effort because the sample documents expose limited actuation metadata.",
    },
    "actuatinglocation": {
        "coverage_status": "guaranteed",
        "matched_field": "ri_devices.actuating_location",
        "notes": "Anchored to the detected piping context and exported into SM_ActuationAndPiping.",
    },
    "systems": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.actuating_system_number",
        "notes": "Actuating systems are modeled in the workbook/AAS/Ontology but only filled when a concrete system identifier is found.",
    },
    "flowdirection": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.flow_direction",
        "notes": "Mapped as a fixed property, with values only when DEXPI piping attributes expose flow direction.",
    },
    "nominaldiameternumericalvaluerepresentation": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.nominal_diameter_numerical_value_representation",
        "notes": "Leaf XML elements are now collected, but nominal diameter values remain source-dependent.",
    },
    "nominaldiameterrepresentation": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.nominal_diameter_representation",
        "notes": "Modeled in the standardized workbook and AAS; populated only when the source contains nominal diameter text.",
    },
    "nominaldiameterstandard": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.nominal_diameter_standard",
        "notes": "Modeled end-to-end with best-effort extraction from DEXPI or IFC metadata.",
    },
    "nominaldiametertyperepresentation": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.nominal_diameter_type_representation",
        "notes": "Modeled end-to-end with best-effort extraction from piping metadata.",
    },
    "linenumber": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.line_number",
        "notes": "The field is supported structurally, but current project files expose line identifiers only sporadically.",
    },
    "actuatingsystemnumber": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.actuating_system_number",
        "notes": "Mapped into SM_ActuationAndPiping, with best-effort population from available piping/system data.",
    },
    "operatedvalvereference": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.operated_valve_reference",
        "notes": "The field is modeled across workbook, AAS, and ontology, but current project data does not always expose an operated valve reference explicitly.",
    },
    "subtagname": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.operated_valve_reference",
        "notes": "Captured as part of the operated-valve linkage when a sub-tag is available.",
    },
    "numberofports": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.has_ports",
        "notes": "Port-related details are modeled primarily through IFC connectivity rather than a dedicated DEXPI field.",
    },
    "pipingcomponentname": {
        "coverage_status": "guaranteed",
        "matched_field": "ri_devices.piping_component_name",
        "notes": "Anchored to the linked piping component / nozzle and exported as part of the fixed UC1 schema.",
    },
    "globalid": {
        "coverage_status": "guaranteed",
        "matched_field": "ifc_entries.global_id",
        "notes": "Ifc GlobalId is modeled as a first-class IFC identity field end-to-end.",
    },
    "workingpressure": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.working_pressure",
        "notes": "Modeled as a best-effort IFC property; populated only when the corresponding IFC property set exists.",
    },
    "nominaldiameter": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.nominal_diameter",
        "notes": "Supported as a normalized IFC property with best-effort extraction from property sets.",
    },
    "innerdiameter": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.inner_diameter",
        "notes": "Modeled in the IFC connectivity sheet but depends on IFC property-set availability.",
    },
    "outerdiameter": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.outer_diameter",
        "notes": "Modeled in the IFC connectivity sheet but depends on IFC property-set availability.",
    },
    "tag": {
        "coverage_status": "guaranteed",
        "matched_field": "ifc_entries.tag",
        "notes": "Tracked as a core IFC identity field and exported to SM_IFCConnectivity.",
    },
    "hasports": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.has_ports",
        "notes": "Port connectivity is modeled and can be derived from IFC relations when such relations are available.",
    },
    "valvepattern": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.valve_mechanism",
        "notes": "Mapped best-effort via generic IFC property-set harvesting.",
    },
    "size": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.size",
        "notes": "Modeled structurally and populated only when the IFC payload provides an explicit size attribute.",
    },
    "connectedto": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.connected_to",
        "notes": "Supported through IFC relation parsing and exported as ontology-ready graph links.",
    },
    "connectedfrom": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.connected_from",
        "notes": "Supported through IFC relation parsing and exported as ontology-ready graph links.",
    },
    "hascontrolelements": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.has_control_elements",
        "notes": "IfcRelFlowControlElements is now parsed, but availability depends on the IFC model content.",
    },
    "predefinedtype": {
        "coverage_status": "guaranteed",
        "matched_field": "ifc_entries.predefined_type",
        "notes": "Tracked as a core IFC classification field and propagated into SM_IFCConnectivity.",
    },
    "valvemechanism": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.valve_mechanism",
        "notes": "Modeled as a best-effort IFC technical property.",
    },
    "flowcoefficient": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.flow_coefficient",
        "notes": "Modeled as a best-effort IFC technical property.",
    },
    "failposition": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.fail_position",
        "notes": "Ifc actuator data is now represented in the schema, but current samples may not contain actuator instances.",
    },
    "manualoverride": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.manual_override",
        "notes": "Modeled as a best-effort actuator property for future IFC-rich UC2/UC3 scenarios.",
    },
    "actuatorapplication": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.actuator_application",
        "notes": "Modeled as a best-effort actuator property for future IFC-rich UC2/UC3 scenarios.",
    },
    "instrumentationloopfunction": {
        "coverage_status": "guaranteed",
        "matched_field": "ri_devices.class_name",
        "notes": "Represented as a fixed ontology/AAS concept in the device-centric UC1 pipeline.",
    },
    "processinstrumentationfunction": {
        "coverage_status": "guaranteed",
        "matched_field": "ri_devices.class_name",
        "notes": "Represented as a fixed ontology/AAS concept in the device-centric UC1 pipeline.",
    },
    "actuatingfunction": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.actuating_function_number",
        "notes": "Modeled as a target class, with instance population depending on explicit actuation evidence.",
    },
    "actuatingsystem": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.actuating_system_number",
        "notes": "Modeled as a target class, with instance population depending on explicit actuation evidence.",
    },
    "pipingnetworksegment": {
        "coverage_status": "guaranteed",
        "matched_field": "relations.anchored_to",
        "notes": "Represented via piping-anchor relations and exported as ontology-ready context nodes.",
    },
    "operatedvalve": {
        "coverage_status": "partial",
        "matched_field": "ri_devices.operated_valve_reference",
        "notes": "Modeled as a target class, with instance population depending on explicit valve linkage.",
    },
    "ifcpipesegment": {
        "coverage_status": "guaranteed",
        "matched_field": "ifc_entries.ifc_class",
        "notes": "Ifc pipe segments are represented as a protected target class in the fixed UC1 TBox.",
    },
    "ifcvalve": {
        "coverage_status": "guaranteed",
        "matched_field": "ifc_entries.ifc_class",
        "notes": "Ifc valves are represented as a protected target class in the fixed UC1 TBox.",
    },
    "ifcactuator": {
        "coverage_status": "partial",
        "matched_field": "ifc_entries.ifc_class",
        "notes": "Ifc actuators are now modeled structurally, pending richer IFC source coverage.",
    },
    "stellenplanentry": {
        "coverage_status": "guaranteed",
        "matched_field": "stellenplan_entries.entry_id",
        "notes": "Stellenplan rows are exported into a dedicated standardized sheet and ontology entry class.",
    },
    "wiringentry": {
        "coverage_status": "guaranteed",
        "matched_field": "wiring_entries.entry_id",
        "notes": "Wiring rows are exported into a dedicated standardized sheet and ontology entry class.",
    },
    "datasheetentry": {
        "coverage_status": "partial",
        "matched_field": "datasheet_entries.entry_id",
        "notes": "Datasheet rows are modeled structurally but depend on project-specific matching quality.",
    },
    "completionproposal": {
        "coverage_status": "guaranteed",
        "matched_field": "completion_candidates.proposal_status",
        "notes": "Completion proposals are now first-class AAS and ontology individuals.",
    },
}


class UC1CatalogService:
    def load_entries(self, catalog_path: Path) -> list[UC1CatalogEntry]:
        workbook = load_workbook(catalog_path, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]

        entries: list[UC1CatalogEntry] = []
        current_document = ""
        current_class_name = ""
        for row in sheet.iter_rows(min_row=2):
            values = [clean_cell(cell.value) for cell in row[:6]]
            if not any(values):
                continue
            document, class_name, data_property, data_type, example, notes = values
            if document:
                current_document = document
            if class_name:
                current_class_name = class_name
            priority = "guaranteed" if self._is_highlighted(row[:5]) else "best_effort"
            if priority != "guaranteed" and notes and "opt" in normalize_label(notes):
                priority = "partial"
            entries.append(
                UC1CatalogEntry(
                    row_number=row[0].row,
                    document=current_document,
                    class_name=current_class_name,
                    data_property=data_property,
                    data_type=data_type,
                    example=example,
                    notes=notes,
                    priority=priority,
                )
            )
        return entries

    def coverage_report(
        self,
        catalog_path: Path,
        support_map: dict[str, dict[str, str]] | None = None,
    ) -> UC1CatalogCoverageReport:
        resolved_support_map = support_map or UC1_SUPPORT_MAP
        entries = self.load_entries(catalog_path)
        rows: list[UC1CatalogCoverageRow] = []
        guaranteed_count = 0
        partial_count = 0
        missing_count = 0

        for entry in entries:
            support = self._resolve_support(entry, resolved_support_map)
            coverage_status = support.get("coverage_status", "missing")
            if coverage_status == "guaranteed":
                guaranteed_count += 1
            elif coverage_status == "partial":
                partial_count += 1
            else:
                missing_count += 1
            rows.append(
                UC1CatalogCoverageRow(
                    row_number=entry.row_number,
                    document=entry.document,
                    class_name=entry.class_name,
                    data_property=entry.data_property,
                    data_type=entry.data_type,
                    example=entry.example,
                    priority=entry.priority,
                    coverage_status=coverage_status,
                    matched_field=support.get("matched_field", ""),
                    notes=support.get("notes", entry.notes),
                )
            )

        highlighted_rows = sum(1 for entry in entries if entry.priority == "guaranteed")
        return UC1CatalogCoverageReport(
            catalog_path=catalog_path,
            total_rows=len(entries),
            highlighted_rows=highlighted_rows,
            guaranteed_count=guaranteed_count,
            partial_count=partial_count,
            missing_count=missing_count,
            rows=rows,
        )

    def _resolve_support(
        self,
        entry: UC1CatalogEntry,
        support_map: dict[str, dict[str, str]],
    ) -> dict[str, str]:
        candidates = [
            self._catalog_key(entry.data_property),
            self._catalog_key(entry.class_name),
        ]
        for candidate in candidates:
            if candidate and candidate in support_map:
                return support_map[candidate]
        if entry.priority == "best_effort":
            return {
                "coverage_status": "partial",
                "matched_field": "",
                "notes": entry.notes or "This non-highlighted catalog row remains a best-effort extension target.",
            }
        return {
            "coverage_status": "missing",
            "matched_field": "",
            "notes": entry.notes or "The highlighted catalog row is not implemented end-to-end yet.",
        }

    def _catalog_key(self, value: str) -> str:
        return normalize_label(value).replace(" ", "")

    def _is_highlighted(self, cells: tuple[Any, ...]) -> bool:
        for cell in cells:
            fill = getattr(cell, "fill", None)
            if fill is None:
                continue
            if getattr(fill, "fill_type", None) in {None, "none"} and getattr(fill, "patternType", None) in {None, "none"}:
                continue
            return True
        return False
