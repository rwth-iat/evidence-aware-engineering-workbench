"""UC1 checker: detect missing PLT-Stelle correspondence across documents.

Design: **Deterministic-first with LLM fallback**.

The deterministic path (Steps 1-3) handles ~95% of cases with zero LLM calls:
  1. Extract all AKZ from the R&I (DEXPI / standardized PID Excel).
  2. Extract all AKZ from Stellenplan, Klemmenplan, Datasheet, 3D standardised Excels.
  3. For each R&I AKZ: normalise → exact match → fuzzy match (edit dist 1→auto,
     edit dist 2→LLM verify).

The LLM path (Step 4) is only invoked for ~5% of boundary cases:
  4. Ambiguous: exactly one target document is missing (could be a merged entry,
     variant notation, or genuine missing correspondence).

Based on PDF slide 30 cardinality rules:
  - R&I PLT-Stelle → Stellenplan: min 1
  - R&I PLT-Stelle → Klemmenplan / Verschaltungsliste: min 1
  - R&I PLT-Stelle → Datasheet: some (0..*)
  - R&I PLT-Stelle → 3D data: some (0..*)
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from iev4pi_transformation_tool.core.akz_normalizer import (
    build_akz_index,
    fuzzy_match_akz,
    normalize_akz,
    strip_function_prefix,
)


# ---------------------------------------------------------------------------
# Cardinality rules from PDF slide 30
# ---------------------------------------------------------------------------

@dataclass
class CardinalityRule:
    source_doc: str       # "R&I"
    target_doc: str       # "Stellenplan"
    min_count: int        # 1 = must appear at least once
    max_count: int | None  # None = unbounded
    severity: str = "critical"  # "critical" | "warning" | "info"


UC1_RULE_SET: list[CardinalityRule] = [
    CardinalityRule("R&I", "Stellenplan", 1, None, "critical"),
    CardinalityRule("R&I", "Klemmenplan", 1, None, "critical"),
    CardinalityRule("R&I", "Datasheet", 0, None, "warning"),
    CardinalityRule("R&I", "3D_Daten", 0, None, "warning"),
]

# Document type → standardised Excel sheet name for AKZ lookup.
DOC_TO_STANDARDISED = {
    "R&I": None,  # source, not a target
    "Stellenplan": "Stellenplan_template.xlsx",
    "Klemmenplan": "Klemmenplan_template.xlsx",
    "Datasheet": "Datasheet_template.xlsx",
    "3D_Daten": "PID_template.xlsx",  # IFC data lands in PID Piping / Equipment
}


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------

@dataclass
class UC1Report:
    canonical_akz: str
    verdict: str  # "consistent" | "missing_correspondence" | "needs_review"
    missing_in: list[str] = field(default_factory=list)
    present_in: list[str] = field(default_factory=list)
    severity: str = "info"
    llm_reasoning: str = ""
    rule_reasoning: str = ""
    confidence: float = 0.0
    review_status: str = "auto"
    detected_at: str = field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def check_uc1(
    ri_workbook: Path,
    target_workbooks: dict[str, Path],
    *,
    llm_agent: Any | None = None,
    llm_verify: callable | None = None,
) -> list[UC1Report]:
    """Run UC1 detection on a set of standardized Excel workbooks.

    Args:
        ri_workbook: Path to the R&I standardized Excel (PID_template.xlsx).
        target_workbooks: Mapping of doc_type → workbook path, e.g.
            ``{"Stellenplan": Path(...), "Klemmenplan": Path(...)}``.
        llm_agent: Optional LLMAgent for edge-case judgment.
        llm_verify: Optional callback for AKZ match verification.

    Returns:
        List of UC1Report, one per R&I AKZ that triggers a rule violation.
    """
    # Step 1: Extract all AKZ from the R&I workbook.
    ri_index = build_akz_index(ri_workbook)
    ri_akz_set = set(ri_index.keys())

    if not ri_akz_set:
        return []

    # Step 2: Extract AKZ sets from all target documents.
    target_akz_sets: dict[str, set[str]] = {}
    target_akz_raw: dict[str, dict[str, str]] = {}  # canonical → original
    for doc_type, wb_path in target_workbooks.items():
        if not wb_path.is_file():
            continue
        idx = build_akz_index(wb_path)
        target_akz_sets[doc_type] = set(idx.keys())
        target_akz_raw[doc_type] = {
            canonical: occs[0]["original_akz"]
            for canonical, occs in idx.items()
        }

    # Step 3: For each R&I AKZ, check presence in each target document.
    reports: list[UC1Report] = []

    for canonical_akz in ri_akz_set:
        present_in: list[str] = []
        missing_in: list[str] = []
        needs_llm: list[tuple[str, str, float]] = []

        for rule in UC1_RULE_SET:
            target_set = target_akz_sets.get(rule.target_doc, set())
            if not target_set and rule.min_count > 0:
                missing_in.append(rule.target_doc)
                continue

            # 3a: Exact canonical match
            if canonical_akz in target_set:
                present_in.append(rule.target_doc)
                continue

            # 3b: Try without function prefix
            stripped = strip_function_prefix(canonical_akz)
            if stripped != canonical_akz and stripped in target_set:
                present_in.append(rule.target_doc)
                continue

            # 3c: Fuzzy match (edit distance)
            best, dist, ratio = fuzzy_match_akz(canonical_akz, target_set, max_distance=2)

            # 3c-ii: If stripped prefix gives a better match, use it instead.
            if stripped != canonical_akz:
                best_s, dist_s, ratio_s = fuzzy_match_akz(stripped, target_set, max_distance=2)
                if best_s is not None and (best is None or dist_s < dist):
                    best, dist, ratio = best_s, dist_s, ratio_s
            if best is None:
                if rule.min_count > 0:
                    missing_in.append(rule.target_doc)
                    continue
                else:
                    continue  # optional doc, no match → OK

            if dist <= 1:
                # Auto-accept
                present_in.append(rule.target_doc)
            elif dist == 2:
                # Needs LLM confirmation
                needs_llm.append((rule.target_doc, best, ratio))
            elif rule.min_count > 0:
                missing_in.append(rule.target_doc)

        # Step 4: LLM verification for ambiguous cases.
        for doc_type, best_match, ratio in needs_llm:
            rule = next((r for r in UC1_RULE_SET if r.target_doc == doc_type), None)
            min_count = rule.min_count if rule is not None else 1

            if llm_verify is not None:
                accepted = llm_verify(
                    canonical_akz, best_match,
                    {"ri": True},
                    {"doc_type": doc_type, "original": target_akz_raw.get(doc_type, {}).get(best_match, "")},
                )
                if accepted:
                    present_in.append(doc_type)
                elif min_count > 0:
                    missing_in.append(doc_type)
            elif llm_agent is not None:
                verdict = llm_agent.judge_akz_correspondence(
                    canonical_akz, best_match,
                    context_a={"source": "R&I", "canonical": canonical_akz},
                    context_b={"source": doc_type, "canonical": best_match},
                )
                if verdict.is_same and verdict.confidence >= 0.7:
                    present_in.append(doc_type)
                elif min_count > 0:
                    missing_in.append(doc_type)
            elif min_count > 0:
                # No LLM available → be conservative: flag as missing for mandatory docs
                missing_in.append(doc_type)
            # For optional documents (min_count==0) without LLM:
            # accept the fuzzy match silently — no missing_in entry.

        # Step 5: Build report if any mandatory document is missing.
        if not missing_in:
            continue  # all good

        severity = _determine_severity(missing_in)
        review_status = "auto" if not needs_llm else "needs_review"

        report = UC1Report(
            canonical_akz=canonical_akz,
            verdict="missing_correspondence",
            missing_in=missing_in,
            present_in=present_in,
            severity=severity,
            rule_reasoning=(
                f"Cardinality violation: must appear in {missing_in}. "
                f"Found in: {present_in or 'none'}."
            ),
            confidence=0.95 if not needs_llm else 0.75,
            review_status=review_status,
        )

        # LLM judge for edge cases (only when the pattern is unusual).
        if llm_agent is not None and (
            len(missing_in) == 1 or review_status == "needs_review"
        ):
            occurrences = {"R&I": ri_index.get(canonical_akz, [])}
            for doc_type in present_in + missing_in:
                idx = build_akz_index(target_workbooks.get(doc_type, Path("")))
                occurrences[doc_type] = idx.get(canonical_akz, [])

            llm_verdict = llm_agent.judge_uc1_inconsistency(
                canonical_akz,
                {k: v for k, v in occurrences.items() if v},
                {"rules": [r.__dict__ for r in UC1_RULE_SET]},
            )
            report.llm_reasoning = llm_verdict.reasoning
            if llm_verdict.confidence < 0.7:
                report.review_status = "needs_review"

        reports.append(report)

    return reports


def _determine_severity(missing_in: list[str]) -> str:
    critical_docs = {"Stellenplan", "Klemmenplan"}
    if any(d in critical_docs for d in missing_in):
        return "critical"
    return "warning"


# ---------------------------------------------------------------------------
# Write results to standardised Excel
# ---------------------------------------------------------------------------

def write_inconsistency_report(
    reports: list[UC1Report],
    output_path: Path,
) -> None:
    """Write UC1 reports into the Inconsistency_Report sheet of a PID workbook."""
    import openpyxl

    wb: openpyxl.Workbook
    if output_path.is_file():
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.Workbook()
        if "Inconsistency_Report" not in wb.sheetnames:
            wb.create_sheet("Inconsistency_Report")

    ws = wb["Inconsistency_Report"]

    # If sheet is empty, write headers.
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        headers = [
            "Index", "Inc_ID", "Canonical_AKZ", "Rule", "Severity",
            "Missing_In", "Present_In", "LLM_Verdict", "LLM_Reasoning",
            "Confidence", "Detected_At", "Reviewed_By", "Review_Status",
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

    next_row = ws.max_row + 1
    for idx, report in enumerate(reports, start=1):
        row = next_row + idx - 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=f"INC-{report.canonical_akz}")
        ws.cell(row=row, column=3, value=report.canonical_akz)
        ws.cell(row=row, column=4, value="UC1")
        ws.cell(row=row, column=5, value=report.severity)
        ws.cell(row=row, column=6, value=", ".join(report.missing_in))
        ws.cell(row=row, column=7, value=", ".join(report.present_in))
        ws.cell(row=row, column=8, value=report.verdict)
        ws.cell(row=row, column=9, value=report.llm_reasoning or report.rule_reasoning)
        ws.cell(row=row, column=10, value=report.confidence)
        ws.cell(row=row, column=11, value=report.detected_at)
        ws.cell(row=row, column=12, value="")
        ws.cell(row=row, column=13, value=report.review_status)

    wb.save(output_path)
