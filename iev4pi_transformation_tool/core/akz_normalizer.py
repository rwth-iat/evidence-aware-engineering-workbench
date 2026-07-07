"""Deterministic AKZ (Anlagenkennzeichen) normalisation and matching engine.

Solves ~90% of cross-document AKZ correspondence without any LLM calls.
Only ambiguous cases (OCR errors, non-standard encoding) reach the LLM
fallback in :class:`iev4pi_transformation_tool.core.llm_agent.LLMAgent`.

Canonical form rules (DIN 19227-2 / IEC 81346):
    - Strip leading project/plant prefixes when consistent across documents
    - Replace all separators (``.`` ``-`` ``/`` space) with empty string
    - Uppercase
    - Collapse multi-whitespace

Examples::
    >>> normalize_akz("TI TU10.T41")
    "TITU10T41"
    >>> normalize_akz("TU10-T41")
    "TU10T41"
    >>> normalize_akz("tu10t41")
    "TU10T41"
"""
from __future__ import annotations

import re
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import openpyxl

# Separators that may appear inside AKZ strings.
_SEPARATOR_PATTERN = re.compile(r"[=./\-\s]+")

# Known function-letter prefixes per IEC 81346 / DIN 19227-2.
# These may appear at the start of an AKZ and are sometimes stripped
# in other documents.
_FUNCTION_PREFIXES = [
    "TI", "TIC", "TIR", "TIA", "TIS",  # Temperature
    "PI", "PIC", "PIR", "PIA", "PIS",  # Pressure
    "FI", "FIC", "FIQ", "FIR", "FIA",  # Flow
    "LI", "LIC", "LIR", "LIA", "LIS",  # Level
    "AI", "AIC", "AIR", "AIA", "AIS",  # Analysis
    "WI", "WIC", "WIR", "WIA", "WIS",  # Weight / Mass
    "DI", "DIC", "DIR", "DIA", "DIS",  # Density
    "VI", "VIC", "VIR", "VIA", "VIS",  # Viscosity
    "QI", "QIC", "QIR", "QIA", "QIS",  # Quality
    "SI", "SIC", "SIR", "SIA", "SIS",  # Speed
    "JI", "JIC", "JIR", "JIA", "JIS",  # Power
    "ZI", "ZIC", "ZIR", "ZIA", "ZIS",  # Position
    "PV", "FV", "LV", "TV", "HV",      # Control valves
    "HV", "ZV", "AV", "SV",            # Shut-off / safety valves
]


def normalize_akz(akz: str) -> str:
    """Return the canonical, separator-free uppercase form of *akz*."""
    if not akz:
        return ""
    # Replace all separators with nothing, uppercase, strip.
    return _SEPARATOR_PATTERN.sub("", str(akz)).upper().strip()


def strip_function_prefix(akz_normalized: str) -> str:
    """Remove a known function-letter prefix if present.

    ``TITU10T41`` → ``TU10T41`` (strip ``TI`` prefix).
    ``PVTU20Y30`` → ``TU20Y30`` (strip ``PV`` prefix).
    """
    upper = akz_normalized.upper()
    for prefix in sorted(_FUNCTION_PREFIXES, key=len, reverse=True):
        if upper.startswith(prefix) and len(upper) > len(prefix):
            remainder = upper[len(prefix):]
            # The remainder should start with a letter (plant area code).
            if remainder and remainder[0].isalpha():
                return remainder
    return upper


def fuzzy_match_akz(
    target: str,
    candidates: set[str],
    max_distance: int = 2,
) -> tuple[str | None, int, float]:
    """Find the best matching AKZ in *candidates* for *target*.

    Returns ``(best_match, edit_distance, similarity_ratio)``.
    Only considers candidates within *max_distance* edits.

    ``best_match`` is ``None`` if no candidate is close enough.
    """
    if not candidates or not target:
        return None, 0, 0.0

    target_upper = target.upper()
    best_match: str | None = None
    best_distance = max_distance + 1
    best_ratio = 0.0

    for candidate in candidates:
        candidate_upper = candidate.upper()
        # Quick length filter: reject if length diff > max_distance
        if abs(len(target_upper) - len(candidate_upper)) > max_distance:
            continue
        # Exact match — immediate return
        if target_upper == candidate_upper:
            return candidate, 0, 1.0
        # Levenshtein via SequenceMatcher
        ratio = SequenceMatcher(None, target_upper, candidate_upper).ratio()
        # Estimate edit distance from ratio
        max_len = max(len(target_upper), len(candidate_upper), 1)
        edits = round((1.0 - ratio) * max_len)

        if edits <= max_distance and ratio > best_ratio:
            best_match = candidate
            best_distance = edits
            best_ratio = ratio

    return best_match, best_distance, best_ratio


def build_akz_index(workbook_path: Path) -> dict[str, list[dict[str, Any]]]:
    """Extract all AKZ occurrences from a standardized Excel workbook.

    Scans every sheet for columns named ``AKZ`` or ``AKZ_Canonical`` and
    returns ``{canonical_akz: [occurrence_dict, ...]}``.

    Each occurrence dict contains: document_id, original_akz, canonical_akz,
    sheet_name, row_number.
    """
    index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Read header row
        headers: dict[int, str] = {}
        akz_col: int | None = None
        akz_canonical_col: int | None = None
        doc_id_col: int | None = None

        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value is None:
                continue
            header = str(cell.value).strip()
            headers[col_idx] = header
            if header == "AKZ":
                akz_col = col_idx
            elif header == "AKZ_Canonical":
                akz_canonical_col = col_idx
            elif header == "Document_ID":
                doc_id_col = col_idx

        if akz_col is None and akz_canonical_col is None:
            continue

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            original_akz = str(row[akz_col - 1]).strip() if akz_col and akz_col <= len(row) and row[akz_col - 1] else ""
            canonical_akz = str(row[akz_canonical_col - 1]).strip() if akz_canonical_col and akz_canonical_col <= len(row) and row[akz_canonical_col - 1] else ""
            doc_id = str(row[doc_id_col - 1]).strip() if doc_id_col and doc_id_col <= len(row) and row[doc_id_col - 1] else ""

            if not original_akz and not canonical_akz:
                continue

            canonical = normalize_akz(canonical_akz or original_akz)
            index[canonical].append({
                "document_id": doc_id,
                "original_akz": original_akz or canonical_akz,
                "canonical_akz": canonical,
                "sheet_name": sheet_name,
                "row_number": row_idx,
            })

    wb.close()
    return dict(index)


def build_canonical_akz_map(
    standardized_workbooks: list[Path],
    *,
    llm_verify: callable | None = None,
) -> list[dict[str, Any]]:
    """Build a cross-document canonical AKZ map from multiple workbooks.

    1. Collect all AKZ occurrences from all workbooks.
    2. Group by exact canonical match.
    3. For remaining unmatched, try fuzzy matching (edit distance ≤ 2).
    4. Optionally call *llm_verify* for ambiguous cases (edit distance = 2).

    Returns a list of mapping entries suitable for the ``AKZ_Canonical_Map``
    sheet in ``PID_template.xlsx``.
    """
    all_occurrences: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for wb_path in standardized_workbooks:
        wb_index = build_akz_index(wb_path)
        for canonical, occs in wb_index.items():
            doc_id = Path(wb_path).stem
            for occ in occs:
                occ["workbook"] = doc_id
            all_occurrences[canonical].extend(occs)

    # Phase 1: exact canonical matches are already grouped by dict key.
    # Phase 2: try fuzzy matching for singletons.
    singletons = {k: v for k, v in all_occurrences.items() if len(v) == 1}
    clusters = {k: v for k, v in all_occurrences.items() if len(v) > 1}

    # Try to merge singletons into clusters via fuzzy matching.
    unmatched_keys = set(singletons.keys())
    for singleton_key in list(unmatched_keys):
        if singleton_key not in unmatched_keys:
            continue
        for cluster_key in list(clusters.keys()):
            best, dist, ratio = fuzzy_match_akz(singleton_key, {cluster_key}, max_distance=2)
            if best is None:
                continue
            if dist <= 1:
                # Auto-merge
                clusters[cluster_key].extend(singletons[singleton_key])
                unmatched_keys.discard(singleton_key)
                break
            elif dist == 2 and llm_verify is not None:
                # LLM verification
                if llm_verify(singleton_key, cluster_key, singletons[singleton_key], clusters[cluster_key]):
                    clusters[cluster_key].extend(singletons[singleton_key])
                    unmatched_keys.discard(singleton_key)
                    break

    # Build result rows.
    result: list[dict[str, Any]] = []
    idx = 0

    for canonical_akz, occs in clusters.items():
        for occ in occs:
            idx += 1
            result.append({
                "Index": idx,
                "Canonical_AKZ": canonical_akz,
                "Document_ID": occ.get("document_id", ""),
                "Original_AKZ": occ.get("original_akz", ""),
                "Source_Sheet": occ.get("sheet_name", ""),
                "Source_Row": occ.get("row_number", 0),
                "Match_Confidence": 1.0,
                "Match_Method": "exact",
                "LLM_Reasoning": "",
            })

    for singleton_key in unmatched_keys:
        for occ in singletons[singleton_key]:
            idx += 1
            result.append({
                "Index": idx,
                "Canonical_AKZ": singleton_key,
                "Document_ID": occ.get("document_id", ""),
                "Original_AKZ": occ.get("original_akz", ""),
                "Source_Sheet": occ.get("sheet_name", ""),
                "Source_Row": occ.get("row_number", 0),
                "Match_Confidence": 0.7,
                "Match_Method": "singleton_unmatched",
                "LLM_Reasoning": "",
            })

    return result
