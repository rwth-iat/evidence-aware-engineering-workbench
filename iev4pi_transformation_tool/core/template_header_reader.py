"""Dynamic Excel template header reader.

Reads template workbooks and returns column-name → column-index mappings
so that data writers can reference columns by name rather than hardcoded
position indices.  This single module eliminates ~50 positional-list
constructions across standardized_export.py.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


class TemplateStructure:
    """Complete structural description of one Excel template."""

    def __init__(self, template_path: Path) -> None:
        self.template_path: Path = template_path
        self.sheets: dict[str, SheetStructure] = {}

    def add_sheet(self, name: str, structure: SheetStructure) -> None:
        self.sheets[name] = structure

    def column_map(self, sheet_name: str) -> dict[str, int]:
        """Return {column_header: 1-based_column_index} for *sheet_name*."""
        s = self.sheets.get(sheet_name)
        return dict(s.column_map) if s else {}

    def column_names(self, sheet_name: str) -> list[str]:
        """Return ordered list of column header names for *sheet_name*."""
        s = self.sheets.get(sheet_name)
        return list(s.column_map.keys()) if s else []

    def required_columns(self, sheet_name: str) -> list[str]:
        """Return columns whose row-2 legend suggests they are required (PK/FK)."""
        s = self.sheets.get(sheet_name)
        return [c.name for c in s.columns if c.is_required] if s else []

    def data_start_row(self, sheet_name: str) -> int:
        s = self.sheets.get(sheet_name)
        return s.data_start_row if s else 3


class SheetStructure:
    """Structural description of a single sheet within a template."""

    def __init__(self, name: str) -> None:
        self.name: str = name
        self.columns: list[ColumnDef] = []
        self.column_map: dict[str, int] = {}  # normalized_name → 1-based col idx
        self.header_rows: int = 1
        self.data_start_row: int = 3

    def add_column(self, col_def: ColumnDef) -> None:
        self.columns.append(col_def)
        self.column_map[col_def.normalized_name] = col_def.index


class ColumnDef:
    """Definition of a single template column."""

    def __init__(self, index: int, name: str, legend: str = "") -> None:
        self.index: int = index  # 1-based column number
        self.name: str = name.strip() if name else ""
        self.legend: str = legend.strip() if legend else ""
        self.normalized_name: str = _normalize_header(name) if name else ""
        self.is_required: bool = _is_required_column(name, legend)


# ---------------------------------------------------------------------------
# Header cache (process-level — templates are immutable at runtime)
# ---------------------------------------------------------------------------

_header_cache: dict[str, TemplateStructure] = {}


def read_template(template_path: Path | str) -> TemplateStructure:
    """Read an Excel template and return its structural description.

    Cached in memory — repeated calls for the same path return the cached
    instance because template files never change at runtime.
    """
    path = Path(template_path)
    cache_key = str(path.resolve())
    if cache_key in _header_cache:
        return _header_cache[cache_key]

    structure = _parse_template(path)
    _header_cache[cache_key] = structure
    return structure


def get_column_map(template_path: Path | str, sheet_name: str) -> dict[str, int]:
    """Convenience: return {header → 1-based_col} for one sheet."""
    return read_template(template_path).column_map(sheet_name)


def get_column_names(template_path: Path | str, sheet_name: str) -> list[str]:
    """Convenience: return ordered list of column header names."""
    return read_template(template_path).column_names(sheet_name)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _normalize_header(name: str) -> str:
    """Normalize a column header name for reliable matching.

    - Strips whitespace
    - Replaces consecutive whitespace / underscores with a single underscore
    - Lowercases

    Known typo corrections are applied so that template misspellings
    (e.g. "Componenet_ID") map to their canonical form ("component_id").
    """
    import re

    n = name.strip()
    n = re.sub(r"[_\s]+", "_", n)
    n = n.lower()

    # Known typo corrections in current templates
    _TYPO_FIXES: dict[str, str] = {
        "componenet_id": "component_id",
        "describtion_entry": "description_entry",
    }
    return _TYPO_FIXES.get(n, n)


def _is_required_column(header: str, legend: str) -> bool:
    """Heuristic: a column is 'required' if its row-2 legend marks it as PK or FK."""
    combined = f"{header} {legend}".lower()
    signals = ["pk:", "pk ", "fk →", "fk→", "eindeutig", "unique", "primärschlüssel"]
    return any(signal in combined for signal in signals)


def _parse_template(path: Path) -> TemplateStructure:
    """Parse a template workbook into a :class:`TemplateStructure`."""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    structure = TemplateStructure(path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_struct = SheetStructure(sheet_name)

        # Determine how many header rows exist (rows before data).
        # Row 1 is always header.  Row 2 may be legend/description or data.
        # If row 2 contains FK/PK signals or is entirely non-numeric text,
        # treat it as a legend row.
        has_legend = False
        if ws.max_column and ws.max_column >= 1:
            row2_vals = [
                str(ws.cell(row=2, column=c).value or "")
                for c in range(1, ws.max_column + 1)
            ]
            row2_text = " ".join(row2_vals)
            legend_signals = [
                "fk", "pk", "fortlaufend", "fremdschlüssel", "eindeutig",
                "referenz", "normativ", "beschreibung", "1-based", "original",
                "normalised", "opt.", "e.g.",
            ]
            if any(sig in row2_text.lower() for sig in legend_signals):
                has_legend = True

        sheet_struct.header_rows = 2 if has_legend else 1
        sheet_struct.data_start_row = max(3, sheet_struct.header_rows + 1)

        # Read column headers from row 1
        if ws.max_column:
            for col_idx in range(1, ws.max_column + 1):
                header = str(ws.cell(row=1, column=col_idx).value or "").strip()
                if not header:
                    continue
                legend = ""
                if has_legend:
                    legend = str(ws.cell(row=2, column=col_idx).value or "").strip()
                col_def = ColumnDef(index=col_idx, name=header, legend=legend)
                sheet_struct.add_column(col_def)

        structure.add_sheet(sheet_name, sheet_struct)

    wb.close()
    return structure


# ---------------------------------------------------------------------------
# Row builder — construct named dict rows instead of positional lists
# ---------------------------------------------------------------------------


def named_row(**kwargs: Any) -> dict[str, Any]:
    """Build a named row dict, dropping ``None`` values.

    Usage::

        named_row(Index=1, Document_ID="HC10", Document="file.pdf")
    """
    return {k: v for k, v in kwargs.items() if v is not None}


def write_named_rows(
    ws,
    named_rows: list[dict[str, Any]],
    column_map: dict[str, int],
    start_row: int = 3,
) -> None:
    """Write data rows to a worksheet using column-header names.

    Args:
        ws: openpyxl worksheet.
        named_rows: list of ``{"ColumnName": value, ...}`` dicts.
        column_map: ``{"ColumnName": 1-based_column_index}`` from
            :func:`get_column_map` or :meth:`TemplateStructure.column_map`.
        start_row: first data row (default 3, below header rows 1-2).
    """
    # Clear old placeholder content below header
    if ws.max_row and ws.max_row >= start_row - 1:
        for row in ws.iter_rows(
            min_row=start_row - 1,
            max_row=ws.max_row,
            max_col=ws.max_column or 1,
        ):
            for cell in row:
                cell.value = None

    for r_offset, row_dict in enumerate(named_rows):
        row_num = start_row + r_offset
        for col_name, value in row_dict.items():
            col_idx = column_map.get(col_name)
            if col_idx is None:
                # Try normalized match
                norm = _normalize_header(col_name)
                for candidate, cidx in column_map.items():
                    if _normalize_header(candidate) == norm:
                        col_idx = cidx
                        break
            if col_idx is None:
                continue
            cell = ws.cell(row=row_num, column=col_idx)
            str_val = str(value) if value is not None else ""
            if str_val.startswith("="):
                cell.value = "'" + str_val
            else:
                cell.value = value
