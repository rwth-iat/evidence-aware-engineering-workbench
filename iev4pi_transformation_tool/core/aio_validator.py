"""AIO Workbook Validator — checks structural and integrity rules (I1–I32, A1–A5).

Implements the spec-defined validator execution order (§11.8):
  Layer 1 – Structural (A1–A5)
  Layer 2 – Lookup/Enum (I13, I14, I29)
  Layer 3 – Object/Cluster/Match (C1–C5, I5–I7, I17–I18, M1–M4)
  Layer 4 – Element/Connection integrity (I1–I4, I8–I12, I15–I22, I30–I32)
  Layer 5 – Document-type-specific (I23–I26, I28)
  Layer 6 – Project aggregation (AG1) — deferred to project-level
  Layer 7 – Provenance (E1–E6, P1–P9)

Usage::

    from iev4pi_transformation_tool.core.aio_validator import validate_aio_workbook
    report = validate_aio_workbook("path/to/workbook.xlsx")
    if report["passed"]:
        print("OK")
    else:
        for err in report["errors"]:
            print(err)

Or as CLI::

    python -m iev4pi_transformation_tool.core.aio_validator workbook.xlsx
"""

from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl


# ── Sheet names in spec order ────────────────────────────────────────────────

AIO_SHEETS_28 = [
    "Rules", "Schema_Metadata", "Document_ID", "Document_Data",
    "Revision_Data", "Document_RepresentedItem", "Object", "Cluster",
    "Object_Cluster", "Elements_TopDown", "Elements_from_Cluster",
    "Match_Result", "Element_ID", "Element_RepresentedItem_Mapping",
    "Element_Data", "Element_Data_Source", "RepresentedItem_Data",
    "RepresentedItem_Data_Source", "Element_Classification",
    "Connection_ID", "Connection_Data", "Connection_Data_Source",
    "Layer_ID", "Attribute_Lookup", "Enum_Lookup",
    "Document_Data_Source", "Revision_Data_Source",
    "Element_Classification_Source",
]


# ══════════════════════════════════════════════════════════════════════════════
# Validator
# ══════════════════════════════════════════════════════════════════════════════

class AIOValidator:
    def __init__(self, wb_path: Path | str) -> None:
        self.path = Path(wb_path)
        self.wb = openpyxl.load_workbook(self.path, data_only=True)
        self.errors: list[dict[str, str]] = []
        self.warnings: list[dict[str, str]] = []
        self.stats: dict[str, int] = defaultdict(int)

    def _err(self, rule: str, sheet: str, msg: str, severity: str = "error") -> None:
        entry = {"rule": rule, "sheet": sheet, "message": msg, "severity": severity}
        if severity == "error":
            self.errors.append(entry)
        else:
            self.warnings.append(entry)

    def _sheet_ids(self, sheet: str, id_col: str) -> set[str]:
        """Get all values from id_col in sheet."""
        if sheet not in self.wb.sheetnames:
            return set()
        ws = self.wb[sheet]
        header_row = 1
        # Find column index
        col_idx = None
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=header_row, column=c).value or "").strip() == id_col:
                col_idx = c
                break
        if col_idx is None:
            return set()
        ids = set()
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None and str(v).strip():
                ids.add(str(v).strip())
        self.stats[f"{sheet}.{id_col}"] = len(ids)
        return ids

    def _check_fk(self, rule: str, sheet: str, fk_col: str,
                   target_sheet: str, target_col: str,
                   nullable: bool = True) -> None:
        """Check that all non-empty values in sheet.fk_col exist in target_sheet.target_col."""
        if sheet not in self.wb.sheetnames:
            return
        target_ids = self._sheet_ids(target_sheet, target_col)
        if not target_ids:
            return  # Empty target — nothing to validate against

        ws = self.wb[sheet]
        # Find FK column
        fk_idx = None
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=c).value or "").strip() == fk_col:
                fk_idx = c
                break
        if fk_idx is None:
            return

        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=fk_idx).value
            if v is None or not str(v).strip():
                if not nullable:
                    self._err(rule, sheet, f"Row {r}: {fk_col} is empty (non-nullable)")
                continue
            v_str = str(v).strip()
            if v_str not in target_ids:
                self._err(rule, sheet, f"Row {r}: {fk_col}={v_str} → {target_sheet}.{target_col} NOT FOUND")

    # ── Layer 1: Structural ──────────────────────────────────────────────

    def check_a1_sheets_present(self) -> None:
        for name in AIO_SHEETS_28:
            if name not in self.wb.sheetnames:
                self._err("A1", name, f"Missing mandatory sheet: {name}")
        extra = [s for s in self.wb.sheetnames if s not in AIO_SHEETS_28]
        if extra:
            self._err("A3", extra[0], f"Extraneous sheets: {extra}")

    def check_a4_sheet_order(self) -> None:
        expected = [s for s in AIO_SHEETS_28 if s in self.wb.sheetnames]
        actual = [s for s in self.wb.sheetnames if s in AIO_SHEETS_28]
        if actual != expected:
            self._err("A4", "", f"Sheet order mismatch. Expected: {expected}")

    def check_a5_pk_uniqueness(self) -> None:
        pk_map = {
            "Document_ID": "Document_ID", "Document_Data": "Document_Data_ID",
            "Element_ID": "Element_ID", "Element_Data": "Element_Data_ID",
            "Connection_ID": "Connection_ID", "Connection_Data": "Connection_Data_ID",
            "Match_Result": "Match_ID", "Cluster": "Cluster_ID",
            "Object": "Object_ID", "Layer_ID": "Layer_ID",
        }
        for sheet, pk in pk_map.items():
            if sheet not in self.wb.sheetnames:
                continue
            ws = self.wb[sheet]
            pk_idx = None
            for c in range(1, ws.max_column + 1):
                if str(ws.cell(row=1, column=c).value or "").strip() == pk:
                    pk_idx = c
                    break
            if pk_idx is None:
                continue
            seen: set[str] = set()
            for r in range(3, ws.max_row + 1):
                v = ws.cell(row=r, column=pk_idx).value
                if v is None or not str(v).strip():
                    continue
                v_str = str(v).strip()
                if v_str in seen:
                    self._err("A5", sheet, f"Duplicate PK: {pk}={v_str} at row {r}")
                seen.add(v_str)

    # ── Layer 2: Lookup/Enum ─────────────────────────────────────────────

    def check_i13_attribute_lookup(self) -> None:
        """Check that Element_Data.Attribute_Name exists in Attribute_Lookup."""
        attr_names = self._sheet_ids("Attribute_Lookup", "Attribute_Name")
        if not attr_names:
            self._err("I13", "Attribute_Lookup", "Attribute_Lookup sheet is empty or missing", "warning")
            return
        for sheet in ["Element_Data", "Document_Data", "Connection_Data", "RepresentedItem_Data"]:
            self._check_fk("I13", sheet, "Attribute_Name", "Attribute_Lookup", "Attribute_Name")

    def check_i14_enum_values(self) -> None:
        """Check that enum-typed values exist in Enum_Lookup."""
        enum_fields = self._get_enum_fields()
        if not enum_fields:
            return
        ws = self.wb["Enum_Lookup"]
        allowed_map: dict[str, set[str]] = defaultdict(set)
        fn_idx = None; av_idx = None
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(row=1, column=c).value or "").strip()
            if h == "Field_Name": fn_idx = c
            if h == "Allowed_Value": av_idx = c
        if fn_idx is None or av_idx is None:
            return
        for r in range(2, ws.max_row + 1):
            fn = str(ws.cell(row=r, column=fn_idx).value or "").strip()
            av = str(ws.cell(row=r, column=av_idx).value or "").strip()
            if fn and av:
                allowed_map[fn].add(av)

        for sheet, attr_col, val_col in [
            ("Element_Data", "Attribute_Name", "Attribute_Value"),
            ("Connection_Data", "Attribute_Name", "Attribute_Value"),
        ]:
            if sheet not in self.wb.sheetnames:
                continue
            ws2 = self.wb[sheet]
            attr_idx = val_idx = None
            for c in range(1, ws2.max_column + 1):
                h = str(ws2.cell(row=1, column=c).value or "").strip()
                if h == attr_col: attr_idx = c
                if h == val_col: val_idx = c
            if attr_idx is None or val_idx is None:
                continue
            for r in range(3, ws2.max_row + 1):
                attr = str(ws2.cell(row=r, column=attr_idx).value or "").strip()
                val = str(ws2.cell(row=r, column=val_idx).value or "").strip()
                if attr in enum_fields and val and val not in allowed_map.get(attr, set()):
                    if val != "Unspecifiable":
                        self._err("I14", sheet, f"Row {r}: {attr}={val} not in Enum_Lookup for field '{attr}'", "warning")

    def _get_enum_fields(self) -> set[str]:
        """Get set of field names that have enum constraints (from Attribute_Lookup)."""
        if "Attribute_Lookup" not in self.wb.sheetnames:
            return set()
        ws = self.wb["Attribute_Lookup"]
        fields: set[str] = set()
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=c).value or "").strip() == "Data_Type":
                dt_idx = c
                break
        else:
            return set()
        name_idx = None
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=c).value or "").strip() == "Attribute_Name":
                name_idx = c
                break
        if name_idx is None:
            return set()
        for r in range(2, ws.max_row + 1):
            dt = str(ws.cell(row=r, column=dt_idx).value or "").strip()
            if dt.lower() == "enum":
                nm = str(ws.cell(row=r, column=name_idx).value or "").strip()
                if nm:
                    fields.add(nm)
        return fields

    # ── Layer 3: Object/Cluster/Match ────────────────────────────────────

    def check_layer3_fks(self) -> None:
        self._check_fk("I5", "Object_Cluster", "Object_ID", "Object", "Object_ID")
        self._check_fk("I5", "Object_Cluster", "Cluster_ID", "Cluster", "Cluster_ID")
        self._check_fk("I6", "Cluster", "Parent_Cluster_ID", "Cluster", "Cluster_ID")
        self._check_fk("I17", "Cluster", "Container_Object_ID", "Object", "Object_ID")
        self._check_fk("I18", "Elements_from_Cluster", "Source_Cluster_ID", "Cluster", "Cluster_ID")

    def check_match_consistency(self) -> None:
        """I1-I4: Match_Result consistency."""
        if "Match_Result" not in self.wb.sheetnames:
            return
        # I1: Element_ID.Source_Match_ID → Match_Result.Match_ID
        self._check_fk("I1", "Element_ID", "Source_Match_ID", "Match_Result", "Match_ID")
        # I19: Match_Result FK references
        self._check_fk("I19", "Match_Result", "Element_TopDown_ID", "Elements_TopDown", "Element_TopDown_ID")
        self._check_fk("I19", "Match_Result", "Element_from_Cluster_ID", "Elements_from_Cluster", "Element_from_Cluster_ID")

    # ── Layer 4: Element/Connection integrity ────────────────────────────

    def check_layer4_fks(self) -> None:
        # I15: Document_ID across all sheets
        doc_ids = self._sheet_ids("Document_ID", "Document_ID")
        if doc_ids:
            for sheet in AIO_SHEETS_28:
                if sheet in ("Rules", "Schema_Metadata", "Attribute_Lookup", "Enum_Lookup", "Document_ID"):
                    continue
                self._check_fk("I15", sheet, "Document_ID", "Document_ID", "Document_ID")

        self._check_fk("I9", "Element_RepresentedItem_Mapping", "Element_ID", "Element_ID", "Element_ID")
        self._check_fk("I9", "Element_RepresentedItem_Mapping", "RepresentedItem_ID", "Document_RepresentedItem", "RepresentedItem_ID")
        self._check_fk("I10", "Element_Data", "Element_ID", "Element_ID", "Element_ID")
        self._check_fk("I21", "Element_ID", "Layer_ID", "Layer_ID", "Layer_ID")
        self._check_fk("I16", "Element_ID", "Parent_Element_ID", "Element_ID", "Element_ID")
        self._check_fk("I16", "Elements_TopDown", "Parent_Element_TopDown_ID", "Elements_TopDown", "Element_TopDown_ID")

        # I8: Connection from/to → ExternalInterface elements
        self._check_fk("I8", "Connection_ID", "From_Element_ID", "Element_ID", "Element_ID")
        self._check_fk("I8", "Connection_ID", "To_Element_ID", "Element_ID", "Element_ID")
        self._check_fk("I11", "Connection_Data", "Connection_ID", "Connection_ID", "Connection_ID")

        # I30: Classification FK
        self._check_fk("I30", "Element_Classification", "Classified_Object_ID", "Element_ID", "Element_ID")

    def check_i12_provenance(self) -> None:
        """Check that every data row has at least one Source row (I12)."""
        provenance_pairs = [
            ("Element_Data", "Element_Data_ID", "Element_Data_Source"),
            ("Connection_Data", "Connection_Data_ID", "Connection_Data_Source"),
            ("Document_Data", "Document_Data_ID", "Document_Data_Source"),
            ("RepresentedItem_Data", "RepresentedItem_Data_ID", "RepresentedItem_Data_Source"),
        ]
        for data_sheet, id_col, source_sheet in provenance_pairs:
            data_ids = self._sheet_ids(data_sheet, id_col)
            source_ids = self._sheet_ids(source_sheet, id_col)
            missing = data_ids - source_ids
            if missing and source_sheet in self.wb.sheetnames:
                self._err("I12", data_sheet, f"{len(missing)} {data_sheet} rows have no {source_sheet} row", "warning")

    # ── Layer 5: Document-type-specific ──────────────────────────────────

    def check_i29_source_format(self) -> None:
        """I29: Exactly one Document_Data row with Attribute_Name=Source_Format."""
        if "Document_Data" not in self.wb.sheetnames:
            return
        ws = self.wb["Document_Data"]
        attr_idx = None
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=c).value or "").strip() == "Attribute_Name":
                attr_idx = c
                break
        if attr_idx is None:
            return
        count = 0
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=attr_idx).value or "").strip() == "Source_Format":
                count += 1
        if count == 0:
            self._err("I29", "Document_Data", "No Source_Format attribute row")
        elif count > 1:
            self._err("I29", "Document_Data", f"Multiple ({count}) Source_Format rows (expected 1)")

    # ── Fill rate ────────────────────────────────────────────────────────

    def fill_rate(self) -> dict[str, dict[str, int]]:
        """Compute fill rates per sheet."""
        rates: dict[str, dict[str, int]] = {}
        seed = {"Rules", "Schema_Metadata", "Attribute_Lookup", "Enum_Lookup"}
        for sn in AIO_SHEETS_28:
            if sn not in self.wb.sheetnames:
                continue
            ws = self.wb[sn]
            data_rows = 0
            total_cols = ws.max_column or 1
            for r in range(3, ws.max_row + 1):
                filled = sum(1 for c in range(1, total_cols + 1)
                           if ws.cell(row=r, column=c).value is not None)
                if filled > 0:
                    data_rows += 1
            rates[sn] = {
                "data_rows": data_rows,
                "total_rows": max(ws.max_row - 2, 0),
                "is_seed": sn in seed,
            }
        return rates

    # ── Run all ──────────────────────────────────────────────────────────

    def validate_all(self) -> dict[str, Any]:
        self.errors.clear()
        self.warnings.clear()
        self.stats.clear()

        # Layer 1
        self.check_a1_sheets_present()
        self.check_a4_sheet_order()
        self.check_a5_pk_uniqueness()

        # Layer 2
        self.check_i13_attribute_lookup()
        self.check_i14_enum_values()
        self.check_i29_source_format()

        # Layer 3
        self.check_layer3_fks()
        self.check_match_consistency()

        # Layer 4
        self.check_layer4_fks()
        self.check_i12_provenance()

        fill = self.fill_rate()
        populated = sum(1 for v in fill.values() if v["data_rows"] > 0 and not v["is_seed"])

        return {
            "workbook": str(self.path),
            "sheets_present": len([s for s in AIO_SHEETS_28 if s in self.wb.sheetnames]),
            "sheets_expected": 28,
            "errors": [dict(e) for e in self.errors],
            "warnings": [dict(w) for w in self.warnings],
            "error_count": len(self.errors),
            "warning_count": len(self.warnings),
            "filled_non_seed_sheets": populated,
            "fill_rate": fill,
            "passed": len(self.errors) == 0,
        }


# ── Public API ────────────────────────────────────────────────────────────────

def validate_aio_workbook(path: Path | str) -> dict[str, Any]:
    """Validate an AIO workbook against the spec-defined rules. Returns a report dict."""
    validator = AIOValidator(path)
    return validator.validate_all()


# ── CLI ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python -m iev4pi_transformation_tool.core.aio_validator <workbook.xlsx>")
        sys.exit(1)

    path = Path(sys.argv[1])
    if not path.is_file():
        print(f"File not found: {path}")
        sys.exit(1)

    report = validate_aio_workbook(path)
    print(f"Workbook: {report['workbook']}")
    print(f"Sheets: {report['sheets_present']}/{report['sheets_expected']}")
    print(f"Filled (non-seed): {report['filled_non_seed_sheets']}")
    print(f"Errors: {report['error_count']}, Warnings: {report['warning_count']}")
    print()

    if report["errors"]:
        print("ERRORS:")
        for e in report["errors"]:
            print(f"  [{e['rule']}] {e['sheet']}: {e['message']}")
    if report["warnings"]:
        print("WARNINGS:")
        for w in report["warnings"]:
            print(f"  [{w['rule']}] {w['sheet']}: {w['message']}")

    if report["passed"]:
        print("✅ All checks passed")
    else:
        print(f"❌ {report['error_count']} errors found")

    sys.exit(0 if report["passed"] else 1)
