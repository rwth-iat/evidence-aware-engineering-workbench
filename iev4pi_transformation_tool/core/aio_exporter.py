"""AIO Workbook exporter — fills the 28-sheet v0.8 schema from ExtractedRecords.

Implements the spec-defined P0–P9 pipeline (§11.4).  For V1, P2–P7 are
simplified (1:1 mapping from records to elements) since the existing extraction
pipeline already produces structured records grouped by family.

Design:
  - All IDs are deterministically generated (sheet_prefix.integer counter)
  - Classification uses ``LLMElementClassifier`` with keyword/lookup fallback
  - Enum encoding uses ``aio_schema_mapping`` lookup tables
  - Provenance rows link every data row to its source Object
"""

from __future__ import annotations

import copy
import json
import re
from concurrent.futures import as_completed

from iev4pi_transformation_tool.core.qos_helpers import QoSAwareThreadPoolExecutor
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

import openpyxl
from openpyxl.utils import get_column_letter

from iev4pi_transformation_tool.core.aio_schema_mapping import (
    DOCUMENT_TYPE_MAP,
    ELEMENT_TYPE_MAP,
    WIRE_COLOR_MAP,
    get_aio_element_info,
    get_wire_color_code,
    get_document_attribute_name,
    get_element_attribute_name,
    get_connection_attribute_name,
    get_polarity_code,
    get_document_attribute_name_rag,
    get_element_attribute_name_rag,
)
from iev4pi_transformation_tool.core.standardized_templates import (
    AIO_TEMPLATE,
    FILLED_TEMPLATES_DIR,
    STANDARDIZED_TEMPLATE_DIR,
)
from iev4pi_transformation_tool.core.source_artifacts import (
    SourceArtifactIndex,
    build_source_artifact_objects,
    field_source_object_id,
    record_source_object_id,
)
from iev4pi_transformation_tool.core.ml_evidence_linker import (
    MLEvidenceLinker,
)
from iev4pi_transformation_tool.models import ExtractedRecord, ExtractedFieldResult


# ── Sheet prefix convention (§3.7.2) ────────────────────────────────────────
SHEET_PREFIX: dict[str, str] = {
    "Document_ID": "D",
    "Document_Data": "DD",
    "Revision_Data": "R",
    "Document_RepresentedItem": "RI",
    "Object": "O",
    "Cluster": "CL",
    "Object_Cluster": "OC",
    "Elements_TopDown": "ETD",
    "Elements_from_Cluster": "EFC",
    "Match_Result": "M",
    "Element_ID": "E",
    "Element_RepresentedItem_Mapping": "MAP",
    "Element_Data": "ED",
    "Element_Data_Source": "EDS",
    "RepresentedItem_Data": "RID",
    "RepresentedItem_Data_Source": "RIDS",
    "Element_Classification": "EC",
    "Element_Classification_Source": "ECS",
    "Connection_ID": "C",
    "Connection_Data": "CD",
    "Connection_Data_Source": "CDS",
    "Layer_ID": "LYR",
    "Document_Data_Source": "DDS",
    "Revision_Data_Source": "RDS",
}


# ── Template sheet names (in spec order) ─────────────────────────────────────
AIO_SHEET_NAMES = [
    "Rules", "Schema_Metadata", "Document_ID", "Document_Data",
    "Revision_Data", "Document_RepresentedItem", "Object", "Cluster",
    "Object_Cluster", "Elements_TopDown", "Elements_from_Cluster",
    "Match_Result", "Element_ID", "Element_RepresentedItem_Mapping",
    "Element_Data", "Element_Data_Source", "RepresentedItem_Data",
    "RepresentedItem_Data_Source", "Element_Classification",
    "Connection_ID", "Connection_Data", "Connection_Data_Source",
    "Layer_ID", "Attribute_Lookup", "Enum_Lookup",
    "Document_Data_Source", "Revision_Data_Source",
    "Element_Classification_Source",
]

# Sheets with pre-populated seed data that must be preserved
SEED_SHEETS = {"Rules", "Schema_Metadata", "Attribute_Lookup", "Enum_Lookup"}

# Mapping from DocumentFamily → AIO Document_Type
FAMILY_TO_DOC_TYPE: dict[str, str] = {
    "stellen_overview_record":  "Instrument_Loop_Diagram",
    "klemmenplan_row":          "Terminal_Diagram",
    "verschaltungsliste_row":   "Terminal_Diagram",
    "cabinet_reference_row":    "Terminal_Diagram",
    "stromlauf_component_group": "Circuit_Diagram",
    "stromlauf_component":      "Circuit_Diagram",
    "stromlauf_connection":     "Circuit_Diagram",
}


# ══════════════════════════════════════════════════════════════════════════════
# Semantic field classifier — reduces hardcoded field-name matching
# ══════════════════════════════════════════════════════════════════════════════

# Regex patterns for connection-related fields (language-neutral, no hardcoding)
_CONNECTION_FIELD_PATTERNS = [
    re.compile(r'leiterfarbe', re.I),       # German: wire color
    re.compile(r'wire.?color', re.I),       # English
    re.compile(r'aderfarbe', re.I),         # German: strand color
    re.compile(r'from_.*', re.I),           # from_component_id, from_element, etc.
    re.compile(r'to_.*', re.I),             # to_component_id, to_element, etc.
    re.compile(r'querschnitt', re.I),       # German: cross-section
    re.compile(r'cross.?section', re.I),    # English
    re.compile(r'betriebsmittel_zugang', re.I),   # German: equipment access (from)
    re.compile(r'betriebsmittel_abgang', re.I),   # German: equipment exit (to)
    re.compile(r'connection.*', re.I),      # connection_id, connection_type
    re.compile(r'polarit', re.I),           # polarity / Polarität
    re.compile(r'kabel', re.I),             # cable / Kabel
    re.compile(r'cable', re.I),
    re.compile(r'verschaltung', re.I),      # German: wiring/interconnection
    re.compile(r'^schirm', re.I),           # German: shielding
    re.compile(r'shielding', re.I),
]

# Regex patterns for RKZ-like values (IEC 81346 designations)
_RKZ_VALUE_PATTERNS = [
    re.compile(r'^-?[A-Z]\d{1,4}[:./][\w/]+'),   # -X1:11, -K1:13/14, X1:11/L3
    re.compile(r'^=[\dA-Z][\d.A-Z]*'),              # =0.H1.T1.TU10.F17
    re.compile(r'^[A-Z]{2,4}\d+[.-][A-Z]\d+'),     # TU10.F17, HC10-N12
    re.compile(r'^[A-Z]\d{1,3}$'),                  # F10, K1, Q2 (simple designation)
    re.compile(r'^\d+/\d+$'),                       # 1/2, 13/14 (contact designations)
]

# Polarity extraction patterns (language-neutral, standard electrical codes)
_POLARITY_PATTERN = re.compile(
    r'\b(L1|L2|L3|L\+|L\-|N|PE|PEN|G\+|G\-|V\+|V\-|FE|AC|DC)\b',
    re.IGNORECASE
)

def _extract_polarity_from_text(text: str) -> str:
    """Extract polarity designation from raw context text.

    Returns the first found polarity code (L1/L2/L3/N/PE/etc.), or empty string.
    Uses standard IEC polarity codes — no language-specific hardcoding.
    """
    if not text:
        return ""
    matches = _POLARITY_PATTERN.findall(text)
    if not matches:
        return ""
    # Return the most common match (e.g., if "L1" appears multiple times)
    from collections import Counter
    counts = Counter(m.upper() for m in matches)
    return counts.most_common(1)[0][0]


def _looks_like_rkz(value: str) -> bool:
    """Check if a value looks like an IEC 81346 reference designation."""
    if not value or len(value) < 2:
        return False
    for pat in _RKZ_VALUE_PATTERNS:
        if pat.match(value.strip()):
            return True
    return False

def _is_connection_field(field_name: str) -> bool:
    """Check if a field name semantically belongs to connection data (not element data)."""
    for pat in _CONNECTION_FIELD_PATTERNS:
        if pat.match(field_name.strip()):
            return True
    return False

def _classify_field_scope(
    field_names: list[str],
    llm_client: Any = None,
) -> dict[str, str]:
    """Classify each field as 'document', 'element', or 'connection'.

    Uses fast regex patterns first, LLM as fallback for ambiguous fields.
    Results are cached in-memory (per field name).
    """
    result: dict[str, str] = {}
    ambiguous: list[str] = []

    for fn in field_names:
        fn_clean = fn.strip().lower()
        # Fast path: connection patterns
        if _is_connection_field(fn_clean):
            result[fn] = "connection"
            continue
        # Fast path: known document-level fields
        mapped = get_document_attribute_name(fn)
        if mapped != fn:  # In DOCUMENT_ATTRIBUTE_MAP
            result[fn] = "document"
            continue
        # Fast path: known element-level fields
        mapped = get_element_attribute_name(fn)
        if mapped != fn:  # In ELEMENT_ATTRIBUTE_MAP
            result[fn] = "element"
            continue
        ambiguous.append(fn)

    # LLM fallback for ambiguous fields
    if ambiguous and llm_client:
        try:
            llm_result = _llm_classify_fields(llm_client, ambiguous)
            for fn in ambiguous:
                result[fn] = llm_result.get(fn, "element")  # default: element
        except Exception:
            for fn in ambiguous:
                result[fn] = "element"

    # Default for anything still unclassified
    for fn in field_names:
        if fn not in result:
            result[fn] = "element"

    return result

def _llm_classify_fields(llm_client: Any, fields: list[str]) -> dict[str, str]:
    """Use LLM to classify fields as document/element/connection."""
    system = (
        "You are classifying industrial engineering data fields. "
        "For each field name, classify it as:\n"
        '  "document" — document-level metadata (project name, plant, date, author, sheet number, version)\n'
        '  "connection" — wire/connection data (wire color, from/to component, cross-section, polarity, cable)\n'
        '  "element" — individual device/component data (function, description, manufacturer, type, terminal)\n'
        "Return JSON: {\"classifications\": {\"field_name\": \"scope\", ...}}"
    )
    user = "Classify these fields:\n" + "\n".join(fields)
    try:
        raw = llm_client.chat_json(system, user)
        return raw.get("classifications", {})
    except Exception:
        return {}


def _map_field_name_with_llm(
    source_field_name: str,
    scope: str,
    allowed_attribute_names: list[str],
    deterministic_map: dict[str, str],
    llm_client: Any = None,
) -> str:
    """Map a source field name to a valid Attribute_Lookup name using LLM fallback.

    Fast path: check deterministic_map (hardcoded mapping table).
    Fallback: if the raw name is already valid in Attribute_Lookup, use it.
    LLM path: for unknown names, use LLM to find the closest valid match.
    Results are cached on disk via the unified LLM cache.
    """
    # 1. Deterministic map (fast path)
    if source_field_name in deterministic_map:
        return deterministic_map[source_field_name]

    # 2. Exact match in allowed names
    if source_field_name in allowed_attribute_names:
        return source_field_name

    # 3. LLM semantic mapping (with disk cache)
    if llm_client and allowed_attribute_names:
        import hashlib
        content_hash = hashlib.sha256(
            f"{source_field_name}|{scope}|{','.join(sorted(allowed_attribute_names)[:50])}".encode()
        ).hexdigest()[:16]
        cache_key = f"attr_map:{content_hash}"

        cached = _cache_get(cache_key)
        if cached is not None:
            result = cached.get(source_field_name, "")
            if result in allowed_attribute_names:
                return result

        # LLM call — only for genuinely unknown fields
        system = (
            f"You are mapping a source field name to a standard attribute name for "
            f"industrial engineering documentation. The scope is '{scope}'.\n"
            f"Available standard attribute names for this scope:\n"
            + "\n".join(f"  - {n}" for n in sorted(allowed_attribute_names)[:50])
            + "\n\nReturn the SINGLE closest matching standard name. "
            "If none matches, return 'Unspecifiable'.\n"
            'Return JSON: {"mapped_name": "Terminal_Number"}'
        )
        user = f"Source field name: {source_field_name}"
        try:
            raw = llm_client.chat_json(system, user)
            result = str(raw.get("mapped_name", "")).strip()
        except Exception:
            result = "Unspecifiable"

        # Cache both positive and negative results
        _cache_put(cache_key, {source_field_name: result})

        if result in allowed_attribute_names:
            return result
        return source_field_name  # Keep original if LLM can't find match

    # 4. No LLM — keep original field name (may trigger I13 warning)
    return source_field_name


# ══════════════════════════════════════════════════════════════════════════════
# VLM connection attribute extraction from PDF circuit diagrams
# ══════════════════════════════════════════════════════════════════════════════

# ── Persistent unified disk cache (.iev4pi/llm_cache.json) ──

_CACHE_PATH = Path(__file__).resolve().parents[2] / ".iev4pi" / "llm_cache.json"

def _load_llm_cache() -> dict:
    """Load the unified LLM/VLM cache from disk.

    Handles two formats:
    - Flat: {'vlm:path': {...}, 'llm_conn:hash': {...}} (direct write, legacy)
    - Nested: {'aio_exporter': {...}} (llm_cache module save, current)
    """
    if _CACHE_PATH.is_file():
        try:
            data = json.loads(_CACHE_PATH.read_text())
            # Current format: nested via llm_cache module save()
            if "aio_exporter" in data and isinstance(data["aio_exporter"], dict):
                return dict(data["aio_exporter"])
            # Legacy format: flat keys directly in root (v0.8 migration)
            has_vlm = any(k.startswith("vlm:") for k in data)
            if has_vlm:
                return dict(data)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}

def _save_llm_cache(data: dict) -> None:
    """Save the unified LLM/VLM cache to disk."""
    _CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        _CACHE_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2))
    except IOError:
        pass

# In-memory caches (lazy-loaded from disk, registered with llm_cache module)
_cache: dict | None = None

def _get_cache() -> dict:
    global _cache
    if _cache is None:
        _cache = _load_llm_cache()
    # Always ensure registration with central llm_cache for persistence
    try:
        from iev4pi_transformation_tool.core import llm_cache as _lc
        if "aio_exporter" not in _lc._registry:
            _lc.register("aio_exporter", _cache)
    except Exception:
        pass
    return _cache

def _cache_get(key: str) -> dict | None:
    v = _get_cache().get(key)
    return v if isinstance(v, dict) else None

def _cache_put(key: str, value: dict) -> None:
    """Store in cache dict. Disk persistence via llm_cache.save() at pipeline end."""
    _get_cache()[key] = value

def _llm_derive_rkzs_batch(
    llm_client: Any,
    items: list[tuple[int, dict[str, str], str]],  # (idx, fields_dict, current_rkz)
) -> dict[int, str]:
    """Use LLM to derive IEC 81346 designations for ALL records in a document.

    Results are cached to .iev4pi/llm_cache.json keyed by document content hash
    for instant reuse on subsequent runs.
    """
    if not llm_client or not items:
        return {}

    # Build content hash for cache key
    import hashlib
    content_str = "||".join(
        f"{idx}:" + "|".join(f"{k}={v[:40]}" for k, v in sorted(fields.items()) if v)
        for idx, fields, _ in items[:30]  # First 30 rows enough for uniqueness
    )
    cache_key = f"rkz_batch:{hashlib.sha256(content_str.encode()).hexdigest()[:16]}"

    # Check disk cache
    cached = _cache_get(cache_key)
    if cached is not None and isinstance(cached, dict):
        try:
            return {int(k): str(v) for k, v in cached.items()}
        except (ValueError, TypeError):
            pass

    # Build a table-like view of all records
    table_lines = []
    for idx, fields, current_rkz in items:
        row = " | ".join(f"{k}={v[:60]}" for k, v in sorted(fields.items()) if v)
        table_lines.append(f"  Row {idx}: {row[:400]}")
    table_text = "\n".join(table_lines[:50])

    system = (
        "You are given rows from an industrial electrical engineering table "
        "(wiring list, terminal list, or equipment list). Your task: for EACH row, "
        "derive the CORRECT IEC 81346-1/2 reference designation based on the field values.\n\n"
        "STRICT RULES:\n"
        "- Terminal: '-X{strip}:{num}' or '-X{strip}:{num}/{polarity}'\n"
        "  e.g. row with klemmleiste_x01='1.0' terminal='3' → '-X01:3'\n"
        "  e.g. row with klemmleiste_x01='4.0' terminal='N' → '-X01:N'\n"
        "- Terminal strip: '-X{num}' e.g. '-X01'\n"
        "- Contactor/relay: '-K{num}' based on Gerät/Bezeichnung field\n"
        "- Fuse: '-F{num}' based on Gerät field (extract number from '1 F 0/63A' → '-F0')\n"
        "- Circuit breaker: '-F{num}' or '-Q{num}'\n"
        "- Switch/E-stop: '-S{num}'\n"
        "- Power supply/rectifier: '-T{num}'\n"
        "- Motor/pump: '-M{num}'\n"
        "- Sensor: '-B{num}' or '-L{num}'\n"
        "- Cabinet: '-A{num}'\n"
        "- PLC module: '-A{num}-M{num}'\n"
        "- Valve actuator: '-Y{num}'\n"
        "- Heater: '-E{num}'\n"
        "- Indicator lamp: '-P{num}'\n\n"
        "For each row, extract the BEST designation from fields like Gerät, Bezeichnung, "
        "Beschreibung, Funktion, PLT-Stelle, klemmleiste_*, Terminal_ID, etc.\n"
        "If a Gerät field says '1 F 0/ 63A FI - 63 A 30 m A 1', extract '-F0' (the fuse designation).\n"
        "If klemmleiste_x01 says '1.0' and the row has a terminal number '3', use '-X01:3'.\n"
        "Return JSON: {\"designations\": {\"0\": \"-F0\", \"1\": \"-X01:3\", ...}} "
        "with the row INDEX as key and the derived designation as value."
    )
    user = f"Table rows:\n{table_text}"

    try:
        raw = llm_client.chat_json(system, user)
        result: dict[int, str] = {}
        raw_designations = raw.get("designations", {})
        for k, v in raw_designations.items():
            try:
                result[int(k)] = str(v).strip()
            except (ValueError, TypeError):
                pass
        # Persist to disk cache
        _cache_put(cache_key, {str(k): v for k, v in result.items()})
        return result
    except Exception:
        return {}


def _llm_derive_rkz(llm_client: Any, record_fields: dict[str, str]) -> str:
    """Use LLM to derive a meaningful designation from record field values.

    Called when the record has no explicit RKZ field and the fallback
    produces a row-reference like 'Tabelle1 row 3'.  The LLM analyzes
    available fields (Gerät, Bezeichnung, Beschreibung, Funktion) and
    extracts or constructs the best designation.
    """
    if not llm_client:
        return ""
    fields_text = "\n".join(f"  {k}: {v[:150]}" for k, v in sorted(record_fields.items()) if v)
    if not fields_text:
        return ""
    system = (
        "You extract STRICT IEC 81346-1/2 reference designations from industrial engineering data. "
        "Given field values from a row in a terminal/wiring/equipment list, return the PROPER "
        "designation following these STRICT rules:\n"
        "- Terminal: '-X{strip}:{number}' e.g. '-X1:11' or '-X1:11/L3'\n"
        "- Terminal strip: '-X{number}' e.g. '-X1'\n"
        "- Contactor/relay/coil: '-K{number}' e.g. '-K1', '-K1:A1/A2'\n"
        "- Fuse: '-F{number}' e.g. '-F10'\n"
        "- Circuit breaker/MCB: '-F{number}' or '-Q{number}' e.g. '-F1'\n"
        "- Switch: '-S{number}' e.g. '-S1', for E-stop '-S{number}' e.g. '-S2'\n"
        "- Power supply/rectifier/transformer: '-T{number}' e.g. '-T1'\n"
        "- Motor/pump: '-M{number}' e.g. '-M1'\n"
        "- Sensor/transmitter: '-B{number}' or '-L{number}' e.g. '-B10'\n"
        "- PLC module: '-A{number}-M{number}' e.g. '-A1-M01'\n"
        "- Heater: '-E{number}' e.g. '-E1'\n"
        "- Cabinet: '-A{number}' e.g. '-A1'\n"
        "- Valve actuator: '-Y{number}' or '-M{number}' e.g. '-Y17'\n"
        "NEVER return descriptive text like '1 F 0/ 63A FI' or 'Gleichrichtergerät'. "
        "If the description says 'Gleichrichter', return '-T1'. "
        "If it says 'Not-Aus', return '-S1'. "
        "If it says 'Schütz', return '-K1'. "
        "Extract terminal numbers from 'Klemmleiste' or terminal fields. "
        "Create the MOST PLAUSIBLE designation from available fields. "
        "Return JSON: {\"designation\": \"-F10\"} with a valid IEC 81346 designation."
    )
    user = f"Fields:\n{fields_text}"
    try:
        raw = llm_client.chat_json(system, user)
        result = str(raw.get("designation", "")).strip()
        return result if result else ""
    except Exception:
        return ""


def _relpath(path: str) -> str:
    """Convert a path to repo-relative form for stable cache keys."""
    try:
        repo = Path(__file__).resolve().parents[2]
        return str(Path(path).resolve().relative_to(repo))
    except (ValueError, OSError):
        return path

def _resolve_relpath(rel: str) -> Path:
    """Resolve a repo-relative path back to absolute."""
    repo = Path(__file__).resolve().parents[2]
    return repo / rel

def _extract_connections_vlm(
    pdf_path: str,
    llm_client: Any,
) -> dict[tuple[str, str], dict[str, str]]:
    """Use VLM to extract connection topology + visible text labels from a PDF.

    VLM reads RAW text labels on/near each wire (no interpretation).
    Code deterministically classifies labels: L1/L2/L3/N/PE → Polarity,
    "2.5"/"6" → Cross_Section.  Wire colors are NOT extracted (no text
    labels exist in typical Stromlaufplan — colors are only graphical).

    Returns: {(from_id, to_id): {"Polarity": "L1", "Cross_Section": "2.5 mm²"}}
    Results cached by relative PDF path for portability.
    """
    if not llm_client or not pdf_path:
        return {}

    # Cache key: "vlm2:" + relative path for portability (vlm2 = text-labels version)
    rel = _relpath(pdf_path)
    cache_key = f"vlm2:{rel}"
    cached = _cache_get(cache_key)
    if cached is not None:
        result: dict[tuple[str, str], dict[str, str]] = {}
        for k, v in cached.items():
            parts = tuple(k.split("||"))
            if len(parts) == 2:
                result[(parts[0], parts[1])] = v
        return result

    try:
        import fitz
    except ImportError:
        try:
            import pymupdf as fitz
        except ImportError:
            return {}

    repo = Path(__file__).resolve().parents[2]
    pdf_file = repo / pdf_path if not Path(pdf_path).is_absolute() else Path(pdf_path)
    if not pdf_file.is_file():
        return {}

    try:
        doc = fitz.open(str(pdf_file))
        result: dict[tuple[str, str], dict[str, str]] = {}

        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            pix = page.get_pixmap(dpi=300)
            import base64
            img_b64 = base64.b64encode(pix.tobytes("png")).decode()

            system = (
                "You are analyzing a German electrical circuit diagram (Stromlaufplan). "
                "For EVERY wire connection on this page, report:\n"
                "- from: source component label (e.g., 'L1', '-F1', '-X24')\n"
                "- to: target component label\n"
                "- visible_text: list ANY text labels you can READ on or immediately "
                "next to the wire. Report the EXACT text (e.g. 'L1', 'N', 'PE', "
                "'2.5', '6', 'BK', 'BN', 'gnge'). Do NOT interpret or classify — "
                "just report what TEXT you can READ.\n"
                "- connection_type: Wire or Bridge\n\n"
                "IMPORTANT: Do NOT assign polarity or wire color based on visual "
                "appearance. Only report TEXT that you can READ with your eyes. "
                "If no text is visible near a wire, visible_text is an empty list.\n"
                "Return JSON: {\"connections\": [{\"from\": \"...\", \"to\": \"...\", "
                "\"visible_text\": [\"L1\", \"2.5\"], \"connection_type\": \"Wire\"}]}"
            )

            messages = [
                {"role": "system", "content": system},
                {"role": "user", "content": [
                    {"type": "text", "text": "List ALL connections. Only report TEXT you can READ."},
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_b64}"}},
                ]},
            ]

            try:
                raw = llm_client.chat_json_messages(messages)
            except Exception:
                continue

            for conn in raw.get("connections", []):
                if not isinstance(conn, dict):
                    continue
                from_id = str(conn.get("from", "") or "").strip()
                to_id = str(conn.get("to", "") or "").strip()
                if not from_id or not to_id:
                    continue

                # Deterministic classification of visible text labels
                # (no VLM interpretation — code classifies raw text)
                visible_texts = conn.get("visible_text", [])
                if not isinstance(visible_texts, list):
                    visible_texts = []
                attrs: dict[str, str] = {}

                for label in visible_texts:
                    if not isinstance(label, str):
                        continue
                    label_upper = label.strip().upper()

                    # Classify as Polarity
                    if label_upper in {"L1", "L2", "L3", "N", "PE", "PEN",
                                        "L+", "L-", "G+", "G-", "V+", "V-",
                                        "FE", "AC", "DC"}:
                        if "Polarity" not in attrs:
                            attrs["Polarity"] = label_upper

                    # Classify as Cross_Section (e.g. "2.5", "6", "1.5", "2.5 mm²")
                    elif re.match(r'^[\d.]+$', label.strip()):
                        if "Cross_Section" not in attrs:
                            attrs["Cross_Section"] = f"{label.strip()} mm²"

                    elif "mm" in label.lower() or "mm²" in label.lower() or "mm2" in label.lower():
                        if "Cross_Section" not in attrs:
                            attrs["Cross_Section"] = label.strip()

                # Connection type
                ct = str(conn.get("connection_type", "") or "").strip()
                if ct in {"Wire", "Bridge", "Bridge_Longitudinal",
                          "Bridge_Cross_Fixed", "Bridge_Cross_Pluggable"}:
                    attrs["Connection_Type"] = ct

                if attrs:
                    from_base = re.sub(r'[-.]\d+$', '', from_id).lower()
                    to_base = re.sub(r'[-.]\d+$', '', to_id).lower()
                    key = (from_id.lower(), to_id.lower())
                    key_base = (from_base, to_base)
                    result[key] = attrs
                    if key_base != key:
                        result[key_base] = attrs

        doc.close()
        # Save to cache with string keys
        serializable: dict[str, dict[str, str]] = {}
        for (fid, tid), attrs in result.items():
            serializable[f"{fid}||{tid}"] = attrs
        _cache_put(cache_key, serializable)
        return result

    except Exception:
        return {}


# ══════════════════════════════════════════════════════════════════════════════
# LLM-based connection attribute extraction from text context
# ══════════════════════════════════════════════════════════════════════════════

# ── Attribute value normalization ──────────────────────────────────────────
# Electrical engineering units and their standard forms
_UNIT_PATTERNS = [
    (re.compile(r'(\d+[.,]?\d*)\s*(mm²|mm2|mm\^2|mm)', re.I), 'mm²'),
    (re.compile(r'(\d+[.,]?\d*)\s*(k?A|Ampere|Amp|A)\b', re.I), None),  # keep original
    (re.compile(r'(\d+[.,]?\d*)\s*(k?V|Volt|V)\b', re.I), None),
    (re.compile(r'(\d+[.,]?\d*)\s*(k?W|Watt|W)\b', re.I), None),
    (re.compile(r'(\d+[.,]?\d*)\s*(k?Hz|Hertz)\b', re.I), None),
    (re.compile(r'(\d+[.,]?\d*)\s*(k?VA)\b', re.I), None),
    (re.compile(r'(\d+[.,]?\d*)\s*(°C|Celsius|°F|Fahrenheit|K|Kelvin)\b', re.I), None),
    (re.compile(r'(\d+[.,]?\d*)\s*(bar|mbar|Pa|kPa|MPa)\b', re.I), None),
    (re.compile(r'(\d+[.,]?\d*)\s*(l|L|ml|mL|m³|m3)\b'), None),
    (re.compile(r'(\d+[.,]?\d*)\s*(kg|g|mg|t)\b'), None),
    (re.compile(r'(\d+[.,]?\d*)\s*(m|cm|mm|km)\b(?!²|2|\^)'), None),
    (re.compile(r'(\d+[.,]?\d*)\s*(min|s|h|sec)\b'), None),
    (re.compile(r'(\d+[.,]?\d*)\s*%\b'), '%'),
]

# Qualifier patterns (DC, AC, 3-phase, etc.)
_QUALIFIER_PATTERNS = [
    re.compile(r'\b(DC|AC|3~|3-phase|1~|1-phase)\b', re.I),
    re.compile(r'\b(\d+)\s*Ph(?:ase)?\b', re.I),
]

def _parse_attribute_value(raw_value: str) -> dict[str, str]:
    """Parse an attribute value string into Normalized_Value, Unit, and Quantity_Qualifier.

    Deterministic parsing for common electrical engineering units.
    Returns dict with keys 'Normalized_Value', 'Unit', 'Quantity_Qualifier' (only populated keys).
    Empty dict if value cannot be parsed.
    """
    if not raw_value or not raw_value.strip():
        return {}

    val = raw_value.strip()

    result: dict[str, str] = {}

    # Try each unit pattern
    for pattern, standard_unit in _UNIT_PATTERNS:
        m = pattern.search(val)
        if m:
            num_str = m.group(1).replace(',', '.')
            try:
                num = float(num_str)
                # Format without trailing zeros
                if num == int(num):
                    result["Normalized_Value"] = str(int(num))
                else:
                    result["Normalized_Value"] = str(num)
            except ValueError:
                result["Normalized_Value"] = num_str

            unit = m.group(2) if m.lastindex and m.lastindex >= 2 else m.group(1)
            if standard_unit:
                result["Unit"] = standard_unit
            elif unit:
                # Normalize common unit variants
                unit_map = {
                    'volt': 'V', 'volts': 'V', 'ampere': 'A', 'amp': 'A', 'amps': 'A',
                    'watt': 'W', 'watts': 'W', 'hertz': 'Hz',
                    'celsius': '°C', 'fahrenheit': '°F', 'kelvin': 'K',
                }
                unit_lower = unit.lower()
                result["Unit"] = unit_map.get(unit_lower, unit)
            break

    # Extract qualifier
    for pat in _QUALIFIER_PATTERNS:
        qm = pat.search(val)
        if qm:
            result["Quantity_Qualifier"] = qm.group(0)
            break

    return result


def _extract_connection_attrs_llm(
    raw_context: str,
    wire_label: str,
    from_id: str,
    to_id: str,
    llm_client: Any,
) -> dict[str, str]:
    """Use LLM to extract connection attributes from raw text context.

    Sends the raw_context and wire_label from a connection record to the LLM
    and asks it to identify polarity and cross-section from text.
    Wire_Color is NOT extracted — it cannot be verified without Polarity text labels.

    Returns dict with keys like 'Polarity', 'Cross_Section', 'Connection_Type'.
    Empty dict if LLM unavailable.
    """
    if not llm_client or not (raw_context or wire_label):
        return {}

    user = f"Connection from '{from_id}' to '{to_id}'.\n"
    if raw_context:
        user += f"Raw context text: {raw_context}\n"
    if wire_label:
        user += f"Wire label: {wire_label}\n"

    system = (
        "Extract connection attributes from the text describing an electrical wire. "
        "Return ONLY attributes you can confidently identify from the TEXT:\n"
        "- polarity: L1, L2, L3, N, PE, PEN, L+, L- (standard IEC codes)\n"
        "- cross_section: e.g. '1.5 mm²', '2.5 mm²', '6 mm²'\n"
        "- connection_type: Wire, Bridge, Bridge_Longitudinal\n"
        "If an attribute is not explicitly mentioned in the text, omit it. Do not guess.\n"
        'Return JSON: {"polarity": "L1", '
        '"cross_section": "2.5 mm²", "connection_type": "Wire"}'
    )

    # Persistent disk cache keyed by content hash
    import hashlib
    content_hash = hashlib.sha256(
        (raw_context[:300] + wire_label[:200]).encode()
    ).hexdigest()[:16]
    cache_key = f"llm_conn:{content_hash}"

    cached = _cache_get(cache_key)
    if cached is not None:
        if isinstance(cached, dict):
            return {str(k): str(v) for k, v in cached.items() if k != "Wire_Color"}

    try:
        raw = llm_client.chat_json(system, user)
    except Exception:
        _cache_put(cache_key, {})
        return {}

    result: dict[str, str] = {}
    pol = str(raw.get("polarity", "") or "").strip().upper()
    if pol in {"L1","L2","L3","N","PE","PEN","L+","L-","G+","G-","V+","V-","FE","AC","DC"}:
        result["Polarity"] = pol
    cs = str(raw.get("cross_section", "") or "").strip()
    if cs and cs not in ("None", "none", ""):
        result["Cross_Section"] = cs
    ct = str(raw.get("connection_type", "") or "").strip()
    if ct and ct not in ("None", "none", ""):
        result["Connection_Type"] = ct

    _cache_put(cache_key, result)
    return result


# ══════════════════════════════════════════════════════════════════════════════
# ID Generator
# ══════════════════════════════════════════════════════════════════════════════

class IDGen:
    def __init__(self) -> None:
        self._counters: dict[str, int] = defaultdict(int)

    def next(self, sheet: str) -> str:
        self._counters[sheet] += 1
        prefix = SHEET_PREFIX.get(sheet, sheet[:3].upper())
        return f"{prefix}.{self._counters[sheet]}"

    def peek(self, sheet: str) -> int:
        return self._counters[sheet] + 1


# ══════════════════════════════════════════════════════════════════════════════
# Row builder helpers
# ══════════════════════════════════════════════════════════════════════════════

def _safe_text(val: Any) -> str:
    """Convert value to string, stripping None."""
    if val is None:
        return ""
    return str(val).strip()


_AIO_INTERNAL_FIELD_RE = re.compile(
    r"(^|_)(bbox|raw_context|trace_path|graph|geometry|object_type|component_(role|type)|"
    r"confidence|decision_confidence|status|review|evidence|provenance|source|record|"
    r"page_number|page|index|row|column|cluster|match|element_type)($|_)",
    re.IGNORECASE,
)
_AIO_DERIVED_VALUE_RE = re.compile(
    r"^\s*-?\d+(?:\.\d+)?\s*,\s*-?\d+(?:\.\d+)?\s*,\s*-?\d+(?:\.\d+)?\s*,\s*-?\d+(?:\.\d+)?\s*$"
)


def _is_aio_internal_element_field(field_name: str, value: Any) -> bool:
    """Return true for extraction/control fields that are not v0.8 Element_Data."""
    name = _safe_text(field_name)
    text = _safe_text(value)
    if not name or not text:
        return True
    if _AIO_INTERNAL_FIELD_RE.search(name):
        return True
    if _AIO_DERIVED_VALUE_RE.match(text):
        return True
    if len(text) > 400 and text.count(",") > text.count(" "):
        return True
    return False


def _element_attribute_lookup_names(wb: openpyxl.Workbook) -> set[str]:
    """Read allowed Element-scope Attribute_Name values from Attribute_Lookup."""
    rows = _safe_read_sheet(wb, "Attribute_Lookup")
    allowed: set[str] = set()
    for row in rows:
        scope = _safe_text(row.get("Scope", "")).lower()
        name = _safe_text(row.get("Attribute_Name", ""))
        if name and (not scope or scope == "element"):
            allowed.add(name)
    return allowed


def _is_valid_element_attribute(attr_name: str, allowed: set[str]) -> bool:
    if not attr_name:
        return False
    return not allowed or attr_name in allowed


def _element_attribute_name_for_field(field_name: str, allowed: set[str]) -> str:
    mapped = get_element_attribute_name_rag(field_name)
    if _is_valid_element_attribute(mapped, allowed):
        return mapped
    raw = _safe_text(field_name)
    if _is_valid_element_attribute(raw, allowed):
        return raw
    return ""


def _source_review_status(confidence: float | None, *, has_source: bool) -> str:
    if not has_source:
        return "Requires_Review"
    if confidence is None:
        return "Unreviewed"
    return "Auto_Approved" if confidence >= 0.8 else "Requires_Review"


def _source_confidence_for_field(
    source_artifact_index: SourceArtifactIndex | None,
    obj_id: str,
    field: ExtractedFieldResult,
) -> float | None:
    artifact = source_artifact_index.artifact_for_object(obj_id) if source_artifact_index and obj_id else None
    raw_conf = _field_confidence(field)
    if artifact is not None:
        if artifact.source_operation in {"VL_Row", "Cell"} and artifact.method == "Native_Table_Row":
            return 1.0 if artifact.confidence is None else max(float(artifact.confidence), 0.95)
        if artifact.source_operation == "Manual_Entry":
            return None
        method = artifact.method.lower()
        artifact_conf = round(float(artifact.confidence), 4) if artifact.confidence is not None else None
        if raw_conf is not None and artifact_conf is not None:
            return min(raw_conf, artifact_conf)
        if raw_conf is not None:
            return raw_conf
        if artifact.method in {"Native_Text", "pymupdf"} or "pymupdf" in method:
            return None
        if artifact.confidence is not None:
            return artifact_conf
    return raw_conf


def _field_confidence(field: ExtractedFieldResult) -> float | None:
    for value in (getattr(field, "decision_confidence", None), getattr(field, "confidence", None)):
        try:
            confidence = float(value)
        except (TypeError, ValueError):
            continue
        if confidence > 0:
            return round(max(0.0, min(1.0, confidence)), 4)
    return None


def _calibrated_field_source_confidence(
    source_artifact_index: SourceArtifactIndex | None,
    obj_id: str,
    field: ExtractedFieldResult,
    *,
    link_confidence: float | None = None,
) -> float | None:
    components: list[float] = []
    field_confidence = _field_confidence(field)
    source_confidence = _source_confidence_for_field(source_artifact_index, obj_id, field)
    for value in (field_confidence, source_confidence, link_confidence):
        try:
            confidence = float(value)
        except (TypeError, ValueError):
            continue
        if confidence > 0:
            components.append(max(0.0, min(1.0, confidence)))
    if not components:
        return None
    return round(min(components), 4)


def _source_method_for_object(source_artifact_index: SourceArtifactIndex | None, obj_id: str) -> str:
    artifact = source_artifact_index.artifact_for_object(obj_id) if source_artifact_index and obj_id else None
    if artifact is None:
        return "Unspecifiable"
    if artifact.source_operation in {"VL_Row", "Cell"}:
        return "Native_Text"
    if artifact.source_operation == "Manual_Entry":
        return "Manual_Entry"
    if artifact.method and "ocr" in artifact.method.lower():
        return "OCR"
    if artifact.method:
        return "Native_Text"
    return "Unspecifiable"


def _is_deterministic_native_source(source_artifact_index: SourceArtifactIndex | None, obj_id: str) -> bool:
    artifact = source_artifact_index.artifact_for_object(obj_id) if source_artifact_index and obj_id else None
    return bool(artifact is not None and artifact.source_operation in {"VL_Row", "Cell"})


def _build_ml_evidence_linker(
    llm_client: Any,
    *,
    enabled: bool,
    benchmark_report_path: Path | str | None,
) -> MLEvidenceLinker | None:
    _ = benchmark_report_path
    if not enabled or llm_client is None:
        return None
    if not (
        getattr(llm_client, "available", lambda: False)()
        and getattr(llm_client, "embedding_available", lambda: False)()
    ):
        return None
    return MLEvidenceLinker(llm_client)


def _source_link_for_field(
    source_artifact_index: SourceArtifactIndex | None,
    ml_evidence_linker: MLEvidenceLinker | None,
    record: ExtractedRecord,
    field: ExtractedFieldResult,
    fallback_object_id: str = "",
) -> tuple[str, float | None, str, str]:
    obj_id = field_source_object_id(source_artifact_index, record, field)
    if not obj_id:
        obj_id = fallback_object_id

    if _is_deterministic_native_source(source_artifact_index, obj_id):
        confidence = _source_confidence_for_field(source_artifact_index, obj_id, field)
        return (
            obj_id,
            confidence,
            _source_method_for_object(source_artifact_index, obj_id),
            _source_review_status(confidence, has_source=True),
        )

    if ml_evidence_linker is not None and source_artifact_index is not None:
        decision = ml_evidence_linker.link(
            field,
            source_artifact_index.artifacts(),
            source_path=getattr(record, "source_path", "") or "",
        )
        selected = decision.selected_artifact_ids[0] if decision.selected_artifact_ids else ""
        if selected and source_artifact_index.artifact_for_object(selected) is not None:
            ml_conf = _calibrated_field_source_confidence(
                source_artifact_index,
                selected,
                field,
                link_confidence=round(float(decision.confidence), 4),
            )
            review_status = _source_review_status(ml_conf, has_source=True)
            method = (
                "Exact_Source_Match"
                if "LLM verifier skipped" in str(decision.rule_basis)
                else "LLM_Classification"
            )
            return selected, ml_conf, method, review_status

    confidence = _source_confidence_for_field(source_artifact_index, obj_id, field) if obj_id else None
    return (
        obj_id,
        confidence,
        _source_method_for_object(source_artifact_index, obj_id) if obj_id else "Unspecifiable",
        _source_review_status(confidence, has_source=bool(obj_id)),
    )


def _aio_ml_evidence_linking_worker_count(
    ml_evidence_linker: MLEvidenceLinker | None,
    item_count: int,
) -> int:
    from iev4pi_transformation_tool.core.qos_helpers import io_worker_count

    if ml_evidence_linker is None or item_count <= 1:
        return 1
    llm_client = getattr(ml_evidence_linker, "llm_client", None)
    config = getattr(llm_client, "config", None)
    try:
        configured = int(getattr(config, "parallel_workers", 0) or 0)
    except (TypeError, ValueError):
        configured = 0
    if configured > 0:
        return max(1, min(8, item_count, configured))
    return max(1, min(8, item_count, io_worker_count(cap=8)))


def _source_links_for_element_data(
    source_artifact_index: SourceArtifactIndex | None,
    ml_evidence_linker: MLEvidenceLinker | None,
    tasks: list[dict[str, Any]],
    progress: Callable[[int, str], None] | None = None,
) -> dict[str, tuple[str, float | None, str, str]]:
    if not tasks:
        return {}

    def link_one(task: dict[str, Any]) -> tuple[str, tuple[str, float | None, str, str]]:
        ed_id = str(task["ed_id"])
        try:
            result = _source_link_for_field(
                source_artifact_index,
                ml_evidence_linker,
                task["record"],
                task["field"],
                str(task.get("fallback_object_id", "")),
            )
        except Exception:
            result = _source_link_for_field(
                source_artifact_index,
                None,
                task["record"],
                task["field"],
                str(task.get("fallback_object_id", "")),
            )
        return ed_id, result

    worker_count = _aio_ml_evidence_linking_worker_count(ml_evidence_linker, len(tasks))
    if worker_count <= 1:
        results = []
        for index, task in enumerate(tasks, start=1):
            results.append(link_one(task))
            if progress is not None:
                progress(round(index * 100 / len(tasks)), f"AIO evidence linking {index}/{len(tasks)}")
    else:
        with QoSAwareThreadPoolExecutor(max_workers=worker_count) as executor:
            future_map = {executor.submit(link_one, task): idx for idx, task in enumerate(tasks)}
            ordered: list[tuple[str, tuple[str, float | None, str, str]] | None] = [None] * len(tasks)
            completed = 0
            for future in as_completed(future_map):
                ordered[future_map[future]] = future.result()
                completed += 1
                if progress is not None:
                    progress(round(completed * 100 / len(tasks)), f"AIO evidence linking {completed}/{len(tasks)}")
            results = [item for item in ordered if item is not None]
    return {ed_id: result for ed_id, result in results}


def _report_export_progress(progress: Callable[[int, str], None] | None, value: int, message: str) -> None:
    if progress is not None:
        progress(max(0, min(100, int(value))), message)


def _sub_export_progress(
    progress: Callable[[int, str], None] | None,
    start: int,
    end: int,
) -> Callable[[int, str], None] | None:
    if progress is None:
        return None

    def nested(value: int, message: str) -> None:
        scaled = start + round((end - start) * max(0, min(100, int(value))) / 100)
        _report_export_progress(progress, scaled, message)

    return nested


def _family_value(record: ExtractedRecord) -> str:
    family = getattr(record, "family", "")
    if hasattr(family, "value"):
        return str(family.value)
    return str(family)


def _record_fields(record: ExtractedRecord) -> dict[str, str]:
    return {
        field.field_name.lower().strip(): _safe_text(field.value)
        for field in (record.results or [])
        if field.field_name and _safe_text(field.value)
    }


def _derive_source_format(source_file: str, records: list[ExtractedRecord]) -> str:
    """Derive AIO Source_Format from source kind rather than a fixed fallback."""
    source_lower = (source_file or "").lower()
    suffix = Path(source_file or "").suffix.lower()
    family_values = {_family_value(record).lower() for record in records}
    source_hint = " ".join([source_lower, *family_values])
    looks_like_excel = suffix in {".xls", ".xlsx", ".xlsm", ".xlsb"} or bool(
        re.search(r"\.xls[xmb]?", source_lower)
    )

    if suffix == ".pdf":
        return "PDF_Drawing"
    if "verschaltungsliste_row" in family_values or "verschaltung" in source_hint:
        return "Verschaltungsliste"
    if looks_like_excel:
        return "Excel_Sheet"
    return "Manual_Entry"


def _clean_designation_token(value: str) -> str:
    value = _safe_text(value)
    if re.fullmatch(r"\d+\.0+", value):
        return value.split(".", 1)[0]
    value = re.sub(r"\s+", "", value)
    return value.upper()


def _canonical_rkz(value: str) -> str:
    value = _clean_designation_token(value)
    if not value:
        return ""
    if value.startswith(("=", "-")):
        return value
    if re.match(r"^[A-Z]\d", value):
        return f"-{value}"
    return value


def _rkz_lookup_keys(value: str) -> list[str]:
    canonical = _canonical_rkz(value)
    if not canonical:
        return []
    variants = {canonical, canonical.replace(".", ""), canonical.replace("_", "")}
    variants.update(v[1:] for v in list(variants) if v.startswith("-"))
    return [v.lower().strip() for v in variants if v]


def _rkz_base_key(value: str) -> str:
    canonical = _canonical_rkz(value).lower().strip()
    canonical = re.sub(r"^-", "", canonical)
    canonical = re.sub(r"[:/].*$", "", canonical)
    return canonical


def _build_element_rkz_lookup(
    element_rows: list[dict[str, Any]],
) -> tuple[dict[str, str], dict[str, str]]:
    exact: dict[str, str] = {}
    base_candidates: dict[str, set[str]] = defaultdict(set)
    for elem in element_rows:
        eid = _safe_text(elem.get("Element_ID"))
        rkz = _safe_text(elem.get("Primary_RKZ"))
        if not eid or not rkz:
            continue
        for key in _rkz_lookup_keys(rkz):
            exact.setdefault(key, eid)
        base = _rkz_base_key(rkz)
        if base:
            base_candidates[base].add(eid)
    base = {key: next(iter(eids)) for key, eids in base_candidates.items() if len(eids) == 1}
    return exact, base


def _resolve_element_id(
    candidates: list[str],
    exact_lookup: dict[str, str],
    base_lookup: dict[str, str],
) -> str:
    for candidate in candidates:
        for key in _rkz_lookup_keys(candidate):
            if key in exact_lookup:
                return exact_lookup[key]
    for candidate in candidates:
        canonical = _canonical_rkz(candidate)
        if not canonical or ":" in canonical:
            continue
        base = _rkz_base_key(canonical)
        if base in base_lookup:
            return base_lookup[base]
    return ""


def _terminal_designation_candidates(strip_field: str, terminal_value: str) -> list[str]:
    suffix = strip_field.lower().strip()
    if suffix.startswith("klemmleiste_"):
        suffix = suffix[len("klemmleiste_"):]
    suffix = suffix.strip("_")
    suffix_dot = suffix.replace("_", ".").upper()
    suffix_compact = suffix_dot.replace(".", "")
    bases: list[str] = []
    for raw_base in (suffix_dot, suffix_compact):
        if not raw_base:
            continue
        base = raw_base if raw_base.startswith("X") else f"X{raw_base}"
        bases.append(base)
        bases.append(base.replace("X.", "X", 1))

    terminal = _clean_designation_token(terminal_value)
    terminals = [terminal]
    if re.fullmatch(r"\d+", terminal):
        terminals.append(str(int(terminal)))

    candidates: list[str] = []
    for base in dict.fromkeys(bases):
        for term in dict.fromkeys(terminals):
            if base and term:
                candidates.append(f"-{base}:{term}")
    return candidates


def _device_designation_candidates(fields: dict[str, str]) -> list[str]:
    raw_values: list[str] = []
    for key in ("component_id", "bezeichnung_im_stromlaufplan", "logical_tag", "terminal_id", "gerat"):
        value = fields.get(key, "")
        if value:
            raw_values.append(value)

    candidates: list[str] = []
    for raw in raw_values:
        cleaned = _safe_text(raw)
        if not cleaned:
            continue
        if _looks_like_rkz(cleaned) or re.match(r"^-?[A-Za-z]\s*\d", cleaned):
            candidates.append(_canonical_rkz(cleaned))

        for match in re.finditer(r"[A-Z]\s*\d+(?:\s*[:/.]\s*\d+)*", cleaned.upper()):
            token = _canonical_rkz(match.group(0))
            if not token:
                continue
            candidates.append(token)
            if "/" in token:
                candidates.append(token.split("/", 1)[0])
            if ":" in token:
                candidates.append(token.split(":", 1)[0])

        for match in re.finditer(r"\d+\s*([A-Z])\s*(\d+)", cleaned.upper()):
            candidates.append(f"-{match.group(1)}{match.group(2)}")

    return [candidate for candidate in dict.fromkeys(candidates) if candidate]


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _write_data_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    rows: list[dict[str, Any]],
    column_order: list[str] | None = None,
    start_row: int = 2,
) -> None:
    """Write dict rows to a worksheet, matching dict keys to column headers in row 1.

    Clears existing content below row 1 first.
    """
    # Clear existing content (delete all data rows to avoid phantom rows)
    if ws.max_row and ws.max_row >= 2:
        for row_num in range(ws.max_row, 1, -1):
            # Check if row is completely empty
            is_empty = True
            for col_num in range(1, (ws.max_column or 1) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None and str(cell.value).strip():
                    is_empty = False
                    break
            if not is_empty:
                ws.delete_rows(row_num)
        # After deleting, max_row should be 1 (header only)

    if not rows:
        return

    # Build header → column index map from row 1
    header_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = _safe_text(ws.cell(row=1, column=col).value)
        if header:
            header_map[header] = col

    # Determine column order
    if column_order is None:
        column_order = list(rows[0].keys())

    for r_offset, row_dict in enumerate(rows):
        row_num = start_row + r_offset
        for key, value in row_dict.items():
            col = header_map.get(key)
            if col is None:
                continue  # Column not in template — skip
            str_val = str(value) if value is not None else ""
            if str_val.startswith("="):
                ws.cell(row=row_num, column=col).value = "'" + str_val
            else:
                ws.cell(row=row_num, column=col).value = value


# ══════════════════════════════════════════════════════════════════════════════
# Main exporter
# ══════════════════════════════════════════════════════════════════════════════

def export_aio_workbook(
    result_dir: Path,
    family: str,
    records: list[ExtractedRecord],
    *,
    llm_client: Any = None,
    aio_ml_evidence_linking_enabled: bool = False,
    aio_ml_benchmark_report_path: Path | str | None = None,
    progress: Callable[[int, str], None] | None = None,
) -> list[Path]:
    """Fill AIO workbooks for source documents — one workbook per document.

    Groups records by Document_ID and creates one workbook per unique document
    (single-document invariant per spec §1.1).  Each workbook is saved to
    ``FILLED_TEMPLATES_DIR/{doc_key}_AIO.xlsx`` and also to *result_dir*.

    Returns the list of all output paths (may be empty if no records).
    """
    if not records:
        return []

    # Group records by source document
    docs: dict[str, list[ExtractedRecord]] = defaultdict(list)
    for rec in records:
        doc_id = _derive_document_id(rec)
        docs[doc_id].append(rec)

    output_paths: list[Path] = []
    total_docs = max(1, len(docs))
    for doc_index, (doc_key, doc_records) in enumerate(docs.items()):
        doc_progress = _sub_export_progress(
            progress,
            round(doc_index * 100 / total_docs),
            round((doc_index + 1) * 100 / total_docs),
        )
        path = _export_single_document(
            result_dir,
            family,
            doc_key,
            doc_records,
            llm_client,
            aio_ml_evidence_linking_enabled=aio_ml_evidence_linking_enabled,
            aio_ml_benchmark_report_path=aio_ml_benchmark_report_path,
            progress=doc_progress,
        )
        if path:
            output_paths.append(path)

    return output_paths


def _derive_document_id(record: ExtractedRecord) -> str:
    """Derive a stable document identifier from a record."""
    source = getattr(record, "source_path", "") or getattr(record, "source_file", "") or ""
    if source:
        return Path(source).stem
    return getattr(record, "record_key", "") or str(hash(str(record)))[:8]


def _export_single_document(
    result_dir: Path,
    family: str,
    doc_key: str,
    records: list[ExtractedRecord],
    llm_client: Any = None,
    *,
    aio_ml_evidence_linking_enabled: bool = False,
    aio_ml_benchmark_report_path: Path | str | None = None,
    progress: Callable[[int, str], None] | None = None,
) -> Path | None:
    """Export one AIO workbook for one source document."""
    _report_export_progress(progress, 0, f"AIO {doc_key}: loading template")

    # Determine document type from the records' actual families (not primary_family)
    # Multiple families may be batched into one AIO export; use the majority family
    family_votes: dict[str, int] = defaultdict(int)
    for rec in records:
        rf = str(getattr(rec, "family", ""))
        if hasattr(rec.family, 'value'):
            rf = str(rec.family.value)
        family_votes[rf] += 1
    dominant_family = max(family_votes, key=family_votes.get) if family_votes else family
    doc_type = FAMILY_TO_DOC_TYPE.get(dominant_family, "Terminal_Diagram")

    # Load template
    template_path = STANDARDIZED_TEMPLATE_DIR / AIO_TEMPLATE
    if not template_path.is_file():
        return None
    wb = openpyxl.load_workbook(template_path)

    # Verify all 28 sheets exist
    for name in AIO_SHEET_NAMES:
        if name not in wb.sheetnames:
            return None

    g = IDGen()

    # ── P0: Document metadata ──
    _report_export_progress(progress, 5, f"AIO {doc_key}: metadata")
    doc_id, meta_obj_id = _p0_metadata(wb, g, doc_key, doc_type, records, llm_client)

    # ── P1: Objects (v0.8 source artifacts, not extracted field values) ──
    _report_export_progress(progress, 12, f"AIO {doc_key}: source artifacts")
    object_rows, source_artifact_index = _p1_objects(g, doc_id, records)
    ml_evidence_linker = _build_ml_evidence_linker(
        llm_client,
        enabled=aio_ml_evidence_linking_enabled,
        benchmark_report_path=aio_ml_benchmark_report_path,
    )
    # Prepend metadata anchor object for provenance (I12)
    source_artifact_index.metadata_object_id = meta_obj_id
    object_rows.insert(0, {
        "Index": 0, "Object_ID": meta_obj_id, "Document_ID": doc_id,
        "Page_Number": 1, "Object_Type": "Text", "Source_Operation": "Manual_Entry",
        "Content_Text": f"Document metadata for {doc_key}", "Object_Role": "Label",
    })

    # ── P2: Top-down elements (from record grouping) ──
    _report_export_progress(progress, 24, f"AIO {doc_key}: top-down elements")
    td_rows, td_map = _p2_top_down(g, doc_id, records, llm_client)

    # ── P3: Clusters (for V1: one cluster per logical group) ──
    _report_export_progress(progress, 36, f"AIO {doc_key}: clusters")
    cluster_rows, oc_rows = _p3_clusters(g, doc_id, records, object_rows, source_artifact_index)

    # ── P4: Elements from cluster ──
    _report_export_progress(progress, 48, f"AIO {doc_key}: elements from clusters")
    efc_rows, efc_map = _p4_elements_from_cluster(g, doc_id, cluster_rows, records, llm_client)

    # ── P5-P7: Match + Element_IDs ──
    match_rows, element_rows, eid_map = _p5_p7_match_and_consolidate(
        g, doc_id, td_rows, efc_rows, td_map, efc_map, records
    )

    # ── LLM re-classify remaining Consumer elements with enriched context ──
    if llm_client:
        _llm_reclassify_consumers(element_rows, records, object_rows, llm_client)

    # ── P8: Classification + attributes ──
    _report_export_progress(progress, 62, f"AIO {doc_key}: classification and attributes")
    _p8_classification_attributes(
        wb,
        g,
        doc_id,
        element_rows,
        records,
        object_rows,
        source_artifact_index,
        ml_evidence_linker,
        progress=_sub_export_progress(progress, 62, 82),
    )

    # ── P9: Connections ──
    _report_export_progress(progress, 84, f"AIO {doc_key}: connections")
    _p9_connections(wb, g, doc_id, element_rows, records, object_rows, llm_client, source_artifact_index)

    # ── Write all sheets ──
    _report_export_progress(progress, 90, f"AIO {doc_key}: writing sheets")
    _write_sheet(wb, "Object", object_rows)
    _write_sheet(wb, "Elements_TopDown", td_rows)
    _write_sheet(wb, "Cluster", cluster_rows)
    _write_sheet(wb, "Object_Cluster", oc_rows)
    _write_sheet(wb, "Elements_from_Cluster", efc_rows)
    _write_sheet(wb, "Match_Result", match_rows)

    # Element rows need special handling for RKZ values starting with '='
    element_data_rows = []
    for row in element_rows:
        cleaned = {}
        for k, v in row.items():
            if isinstance(v, str) and v.startswith("="):
                cleaned[k] = "'" + v
            else:
                cleaned[k] = v
        element_data_rows.append(cleaned)
    _write_sheet(wb, "Element_ID", element_data_rows)

    # ── RepresentedItem_Data: populated only when source provides RI-level attributes ──
    # (Terminal_Strip: Function, Voltage_Level, Terminal_Count, etc. per Attribute_Lookup)
    rid_rows: list[dict[str, Any]] = []
    rids_rows: list[dict[str, Any]] = []
    ri_sheet_rows = _safe_read_sheet(wb, "Document_RepresentedItem")
    for ri_row in ri_sheet_rows:
        ri_key = ri_row.get("RepresentedItem_ID", "")
        ri_type = ri_row.get("RepresentedItem_Type", "")
        if ri_key and ri_type:
            # Write Function as the only universal RepresentedItem attribute
            rid_id = g.next("RepresentedItem_Data")
            rid_rows.append({
                "Index": len(rid_rows) + 1,
                "RepresentedItem_Data_ID": rid_id,
                "RepresentedItem_ID": ri_key,
                "Attribute_Name": "Function",
                "Attribute_Value": ri_type.replace("_", " "),
                "Parsing_Status": "Parsed_OK",
            })
            rids_rows.append({
                "Index": len(rids_rows) + 1,
                "RepresentedItem_Data_ID": rid_id,
                "Source_Object_ID": meta_obj_id,
                "Source_Role": "Label",
                "Extraction_Method": "Rule_Based_Parser",
                "Confidence": 0.9,
                "Review_Status": "Auto_Approved",
                "Extraction_Timestamp": _now_iso(),
            })
    _write_sheet(wb, "RepresentedItem_Data", rid_rows)
    _write_sheet(wb, "RepresentedItem_Data_Source", rids_rows)

    # ── Populate Revision_Data (minimal: one P0 extraction revision) ──
    rev_rows_existing = _safe_read_sheet(wb, "Revision_Data")
    if not rev_rows_existing:
        rev_id = g.next("Revision_Data")
        rev_rows = [{
            "Index": 1, "Revision_ID": rev_id, "Document_ID": doc_id,
            "Revision_Index": "0", "Revision_Date": _now_iso()[:10],
            "Revision_Author": "IEV4PI_Transformation",
            "Revision_Description": "Automated extraction from source document",
        }]
        _write_sheet(wb, "Revision_Data", rev_rows)
        # Revision_Data_Source
        rev_src_rows = [{
            "Index": 1, "Revision_ID": rev_id,
            "Source_Object_ID": meta_obj_id,
            "Source_Role": "Label",
            "Extraction_Method": "Rule_Based_Parser",
            "Confidence": 0.9,
            "Review_Status": "Auto_Approved",
            "Extraction_Timestamp": _now_iso(),
        }]
        _write_sheet(wb, "Revision_Data_Source", rev_src_rows)

    # ── LLM-based attribute name normalization (I13 compliance) ──
    if llm_client:
        _report_export_progress(progress, 93, f"AIO {doc_key}: normalizing attributes")
        _normalize_all_attribute_names(wb, llm_client)

    # ── Fill Normalized_Value/Unit/Quantity_Qualifier from Attribute_Value ──
    _report_export_progress(progress, 96, f"AIO {doc_key}: normalized values")
    _fill_normalized_values(wb)

    # ── Persist LLM/VLM cache to disk ──
    try:
        from iev4pi_transformation_tool.core import llm_cache as _lc
        _lc.save()
    except Exception:
        pass

    # Save to result_dir (temp dir used by pipeline)
    safe_key = re.sub(r'[<>:"/\\|?*]', '_', doc_key)[:100]
    output_name = f"{safe_key}_AIO.xlsx"
    output_path = result_dir / output_name
    _report_export_progress(progress, 98, f"AIO {doc_key}: saving workbook")
    wb.save(str(output_path))

    # Also save to filled_templates for GUI "Save" button and collect_filled_templates()
    filled_path = FILLED_TEMPLATES_DIR / output_name
    wb.save(str(filled_path))
    _report_export_progress(progress, 100, f"AIO {doc_key}: complete")

    return output_path


# ══════════════════════════════════════════════════════════════════════════════
# P0 — Document metadata
# ══════════════════════════════════════════════════════════════════════════════

def _p0_metadata(
    wb: openpyxl.Workbook,
    g: IDGen,
    doc_key: str,
    doc_type: str,
    records: list[ExtractedRecord],
    llm_client: Any = None,
) -> tuple[str, str]:
    """Populate Document_ID, Document_Data, Document_RepresentedItem.

    Returns (Document_ID, meta_object_id) — the FK used by downstream sheets
    and the metadata anchor Object for provenance (I12).
    """
    doc_id = g.next("Document_ID")

    # Extract metadata from records
    source_file = ""
    page_count = 1
    for rec in records:
        src = getattr(rec, "source_path", "") or getattr(rec, "source_file", "") or ""
        if src:
            source_file = Path(src).name
        pages = getattr(rec, "pages", 0) or 0
        if pages:
            page_count = max(page_count, pages)

    # Document_ID row
    _write_sheet(wb, "Document_ID", [{
        "Index": 1,
        "Document_ID": doc_id,
        "Document_Type": doc_type,
        "Document_Filename": source_file,
        "Page_Count": page_count,
        "Schema_Version": "v0.8",
        "Lookup_Version": "v0.8.0",
        "Created_Timestamp": _now_iso(),
        "Created_By": "IEV4PI_Transformation",
        "SemanticID": doc_key,
    }])

    # Document_Data rows (EAV) — document-level attributes
    dd_rows: list[dict[str, Any]] = []
    seen_attrs: set[str] = set()

    # 1. Collect all unique field names across records for scope classification
    all_field_names: list[str] = list({f.field_name for rec in records for f in (rec.results or []) if f.field_name})
    field_scope = _classify_field_scope(all_field_names, llm_client=None)  # regex only, no LLM for speed

    # 2. Extract document-level fields from records
    for rec in records:
        for field in (rec.results or []):
            fn = field.field_name
            scope = field_scope.get(fn, "element")
            if scope != "document":
                continue
            mapped_name = get_document_attribute_name_rag(fn)
            if mapped_name == fn:
                mapped_name = fn  # Use as-is if not in mapping
            if mapped_name in seen_attrs:
                continue
            seen_attrs.add(mapped_name)
            dd_id = g.next("Document_Data")
            dd_rows.append({
                "Index": len(dd_rows) + 1,
                "Document_Data_ID": dd_id,
                "Document_ID": doc_id,
                "Attribute_Name": mapped_name,
                "Attribute_Value": _safe_text(field.value),
                "Raw_Value": _safe_text(field.value),
                "Parsing_Status": "Parsed_OK",
                "SemanticID": f"doc_attr_{mapped_name}",
                "_source": "Native_Text",
            })

    # 3. LLM-driven: extract document metadata from filename and record content
    if llm_client and len(dd_rows) <= 2 and source_file:
        # Collect sample field values for context
        sample_fields: list[str] = []
        for rec in records[:5]:
            for field in (rec.results or [])[:10]:
                if field.value and str(field.value).strip():
                    sample_fields.append(f"{field.field_name}={str(field.value)[:80]}")
        context = "\n".join(sample_fields[:20])

        # Cache key: source_file + content hash
        import hashlib
        meta_cache_key = f"doc_meta:{hashlib.sha256((source_file + context[:500]).encode()).hexdigest()[:16]}"
        cached_meta = _cache_get(meta_cache_key)

        if cached_meta is not None and isinstance(cached_meta, dict):
            # Use cached metadata
            for key, val in cached_meta.items():
                attr_name = get_document_attribute_name_rag(str(key))
                if val and str(val).strip() and attr_name not in seen_attrs:
                    seen_attrs.add(attr_name)
                    dd_id = g.next("Document_Data")
                    dd_rows.append({
                        "Index": len(dd_rows) + 1, "Document_Data_ID": dd_id,
                        "Document_ID": doc_id, "Attribute_Name": attr_name,
                        "Attribute_Value": str(val).strip()[:500],
                        "Raw_Value": str(val).strip()[:500], "Parsing_Status": "Parsed_OK",
                        "_source": "LLM_Classification",
                    })
        else:
            try:
                system = (
                    "Extract document-level metadata from an industrial engineering file. "
                    "Given the filename and sample data rows, identify: "
                    "Project_Name, Plant, Cabinet, Revision, Creation_Date. "
                    'Return JSON: {"metadata": {"Project_Name": "...", "Plant": "...", '
                    '"Cabinet": "...", "Revision": "..."}}'
                )
                user = f"Filename: {source_file}\nSample data:\n{context[:1500]}"
                raw = llm_client.chat_json(system, user)
                metadata = raw.get("metadata", {})
                # Persist to disk cache (normalize keys through attribute map)
                clean_meta = {}
                for k, v in metadata.items():
                    if v and str(v).strip():
                        norm_k = get_document_attribute_name_rag(str(k))
                        clean_meta[norm_k] = str(v).strip()
                if clean_meta:
                    _cache_put(meta_cache_key, clean_meta)
                for key, val in metadata.items():
                    attr_name = get_document_attribute_name_rag(str(key))
                    if val and str(val).strip() and attr_name not in seen_attrs:
                        seen_attrs.add(attr_name)
                        dd_id = g.next("Document_Data")
                        dd_rows.append({
                            "Index": len(dd_rows) + 1,
                            "Document_Data_ID": dd_id,
                            "Document_ID": doc_id,
                            "Attribute_Name": attr_name,
                            "Attribute_Value": str(val).strip()[:500],
                            "Raw_Value": str(val).strip()[:500],
                            "Parsing_Status": "Parsed_OK",
                            "_source": "LLM_Classification",
                        })
            except Exception:
                pass

    # 4. Extract metadata from source filename (fallback)
    if source_file and "Document_Name" not in seen_attrs:
        dd_id = g.next("Document_Data")
        stem = Path(source_file).stem
        dd_rows.append({
            "Index": len(dd_rows) + 1,
            "Document_Data_ID": dd_id,
            "Document_ID": doc_id,
            "Attribute_Name": "Document_Name",
            "Attribute_Value": stem[:200],
            "Raw_Value": stem[:200],
            "Parsing_Status": "Parsed_OK",
            "_source": "Rule_Based_Parser",
        })
        seen_attrs.add("Document_Name")

    # 4. Ensure Source_Format is always present and consistent with source kind (I29 rule)
    source_format = _derive_source_format(source_file, records)
    source_format_rows = [
        row for row in dd_rows
        if row.get("Attribute_Name") == "Source_Format"
    ]
    if source_format_rows:
        for row in source_format_rows:
            if row.get("Attribute_Value") != source_format:
                row["Attribute_Value"] = source_format
                row["Raw_Value"] = source_format
                row["Parsing_Status"] = "Parsed_OK"
                row["_source"] = "Rule_Based_Parser"
    else:
        dd_id = g.next("Document_Data")
        dd_rows.append({
            "Index": len(dd_rows) + 1,
            "Document_Data_ID": dd_id,
            "Document_ID": doc_id,
            "Attribute_Name": "Source_Format",
            "Attribute_Value": source_format,
            "Raw_Value": source_format,
            "Parsing_Status": "Parsed_OK",
            "_source": "Rule_Based_Parser",
        })
        seen_attrs.add("Source_Format")

    _write_sheet(wb, "Document_Data", dd_rows)

    # Document_Data_Source — provenance for every Document_Data row (I12)
    meta_obj_id = g.next("Object")
    dds_rows: list[dict[str, Any]] = []
    for i, dd_row in enumerate(dd_rows):
        src_method = dd_row.pop("_source", "Native_Text")
        src_conf = 0.95 if src_method == "Native_Text" else (0.70 if src_method == "LLM_Classification" else 0.85)
        dds_rows.append({
            "Index": i + 1,
            "Document_Data_ID": dd_row["Document_Data_ID"],
            "Source_Object_ID": meta_obj_id,
            "Source_Role": "Label",
            "Extraction_Method": src_method,
            "Confidence": src_conf,
            "Review_Status": "Auto_Approved" if src_method == "Native_Text" else "Requires_Review",
            "Extraction_Timestamp": _now_iso(),
        })
    _write_sheet(wb, "Document_Data_Source", dds_rows)

    # Document_RepresentedItem (one per PCE loop / terminal strip / logical sheet)
    ri_rows: list[dict[str, Any]] = []
    for rec in records:
        ri_id = g.next("Document_RepresentedItem")
        primary_rkz = ""
        for field in (rec.results or []):
            if field.field_name in ("Position_Entry", "Primary_RKZ", "Instrument_ID",
                                     "AKZ", "CanonicalTag", "Terminal_Name"):
                primary_rkz = _safe_text(field.value)
                break

        ri_type = "PCE_Request"
        fam_str = str(getattr(rec, "family", "")).lower()
        if "klemmenplan" in fam_str or "verschaltungsliste" in fam_str or "cabinet_reference" in fam_str:
            ri_type = "Terminal_Strip"
        elif "stromlauf" in fam_str:
            ri_type = "Circuit"

        ri_rows.append({
            "Index": len(ri_rows) + 1,
            "RepresentedItem_ID": ri_id,
            "Document_ID": doc_id,
            "RepresentedItem_Type": ri_type,
            "Primary_RKZ": primary_rkz,
            "Topic_Identification_Status": "Confirmed" if primary_rkz else "Inferred",
        })
    # ── Derive Primary_RKZ for RepresentedItems from source metadata only ──
    # Only runs if the document has meaningful identifying fields (Project_Name,
    # Cabinet, Plant, Position).  Uses LLM to combine source fields into an
    # IEC 81346 designation WITHOUT adding conventional prefixes or hierarchy
    # levels not present in the source.  Result is marked Inferred (not Confirmed)
    # because the designation is constructed, not directly extracted.
    if llm_client:
        empty_rkz_items = [
            (i, r) for i, r in enumerate(ri_rows)
            if not r.get("Primary_RKZ", "")
        ]
        if empty_rkz_items:
            doc_context = {r["Attribute_Name"]: r["Attribute_Value"] for r in dd_rows
                          if r.get("Attribute_Name") and r.get("Attribute_Value")}
            doc_context["source_file"] = source_file

            # Only proceed if there are REAL identifying fields (not filename-derived)
            # Plant, Cabinet, Position are explicit source fields.
            # Project_Name is only used if it looks like a code (short, contains digits/dots).
            real_fields = {}
            for k in ("Plant", "Cabinet", "Position"):
                v = doc_context.get(k, "")
                if v and len(v) < 30:
                    real_fields[k] = v
            pn = doc_context.get("Project_Name", "")
            if pn and len(pn) < 20 and re.search(r'[\d.]', pn):
                real_fields["Project_Name"] = pn
            if not real_fields:
                # Not enough real source data — leave RKZ empty rather than fabricate
                pass
            else:
                context_text = "\n".join(f"{k}: {v}" for k, v in sorted(real_fields.items()))

                import hashlib
                ri_cache_key = f"ri_rkz_v3:{hashlib.sha256(context_text[:600].encode()).hexdigest()[:16]}"
                cached_rkz = _cache_get(ri_cache_key)

                if cached_rkz is not None and isinstance(cached_rkz, dict):
                    base_rkz = cached_rkz.get("base", "")
                    if base_rkz:
                        for i, r in empty_rkz_items:
                            r["Primary_RKZ"] = base_rkz
                            r["Topic_Identification_Status"] = "Inferred"
                else:
                    system = (
                        "Construct a concise reference designation from these source "
                        "document fields. Use ONLY the values provided — do NOT add "
                        "standard prefixes like =0, H1, T1, or plant hierarchy levels "
                        "that are not in the data. Simply combine the available fields "
                        "with '.' as separator, IEC 81346 style.\n"
                        "Example: Plant=TU10, Cabinet=N → 'TU10.N'\n"
                        "Example: Project_Name=HC10, Cabinet=X01 → 'HC10.X01'\n"
                        "If there is only one meaningful field, return it as-is.\n"
                        'Return JSON: {"designation": "TU10.N"}'
                    )
                    user = f"Source fields:\n{context_text[:1000]}"
                    try:
                        raw = llm_client.chat_json(system, user)
                        derived = str(raw.get("designation", "")).strip()
                        if derived and not re.match(r'^=[\d.]+', derived):
                            _cache_put(ri_cache_key, {"base": derived})
                            for i, r in empty_rkz_items:
                                r["Primary_RKZ"] = derived
                                r["Topic_Identification_Status"] = "Inferred"
                    except Exception:
                        pass

    _write_sheet(wb, "Document_RepresentedItem", ri_rows)

    # Return the metadata anchor object — caller prepends it to P1 Object rows
    return doc_id, meta_obj_id


# ══════════════════════════════════════════════════════════════════════════════
# P1 — Object extraction
# ══════════════════════════════════════════════════════════════════════════════

def _p1_objects(
    g: IDGen,
    doc_id: str,
    records: list[ExtractedRecord],
) -> tuple[list[dict[str, Any]], SourceArtifactIndex]:
    """Create v0.8 Object rows from source-verbatim artifacts.

    ``Object.Content_Text`` is an evidence layer, so it may contain parser/OCR/
    Excel row text or a Manual_Entry rationale, but never derived bbox strings
    or element classification values copied from extracted fields.
    """
    return build_source_artifact_objects(g, doc_id, records)


# ══════════════════════════════════════════════════════════════════════════════
# P2 — Top-Down Elements
# ══════════════════════════════════════════════════════════════════════════════

def _p2_top_down(
    g: IDGen,
    doc_id: str,
    records: list[ExtractedRecord],
    llm_client: Any = None,
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    """Create Elements_TopDown rows from record structure.

    Uses LLM classifier when available for Element_Type inference.
    Returns (rows, {record_key → Element_TopDown_ID}) for later matching.
    """
    rows: list[dict[str, Any]] = []
    key_map: dict[str, str] = {}

    # Lazy-init LLM classifier
    classifier = None
    if llm_client:
        try:
            from iev4pi_transformation_tool.core.llm_element_classifier import LLMElementClassifier
            classifier = LLMElementClassifier(llm_client)
        except Exception:
            pass

    # First pass: extract field data from all records
    rec_data: list[dict[str, Any]] = []
    for i, rec in enumerate(records):
        data: dict[str, Any] = {"name": "", "desc": "", "etype": "Consumer", "rkz": "", "idx": i}

        # Collect all field values for RKZ scanning
        all_values: list[str] = []
        for field in (rec.results or []):
            fn = field.field_name.lower()
            fv = _safe_text(field.value)
            if not fv:
                continue
            all_values.append(fv)
            if fn in ("terminal_name", "element_name", "component_name", "name",
                       "display_label", "display_name", "logical_tag",
                       "bezeichnung_im_stromlaufplan"):
                data["name"] = fv
            if fn in ("beschreibung", "description", "funktion", "component_role",
                       "funktion_und_bestelldaten"):
                data["desc"] = fv
            # Known RKZ-bearing fields
            if fn in ("terminal_id", "element_id", "component_key", "primary_rkz",
                       "canonicaltag", "akz", "canonical_tag", "component_id",
                       "logical_tag", "gerat"):
                data["rkz"] = fv
            if fn in ("element_type", "component_role", "object_type", "component_type",
                       "component_role"):
                info = get_aio_element_info(fv)
                if info:
                    data["etype"] = info["element_type"]

        # Smart RKZ detection: scan all field values for designation-like patterns
        if not data["rkz"]:
            for fv in all_values:
                if _looks_like_rkz(fv):
                    data["rkz"] = fv
                    break
        # Try display_name if it looks like an RKZ (e.g. "O1", "-X1:11", "X24 -> B1")
        if not data["rkz"]:
            disp = _safe_text(getattr(rec, "display_name", ""))
            if disp:
                # Extract first token that looks like a designation
                for token in re.split(r'[ ,→\->]+', disp):
                    token = token.strip()
                    if _looks_like_rkz(token):
                        data["rkz"] = token
                        break
                if not data["rkz"] and len(disp) < 30:
                    data["rkz"] = disp  # Short enough to be a designation

        if not data["name"]:
            data["name"] = data["rkz"]
        if not data["rkz"]:
            data["rkz"] = f"UNKNOWN_{getattr(rec, 'record_key', str(i))}"

        rec_data.append(data)

    # Filter out records that are source header rows (field values = field names)
    # These are extraction artifacts where column headers were treated as data
    valid_indices = []
    for i, d in enumerate(rec_data):
        rec = records[d["idx"]]
        field_values = set()
        for field in (rec.results or []):
            val = _safe_text(field.value)
            if val:
                field_values.add(val.lower())
        field_names_lower = {f.field_name.lower() for f in (rec.results or []) if f.field_name}
        # If all field values are also field names, this is a header row → skip
        if field_values and field_names_lower and field_values.issubset(field_names_lower):
            continue
        valid_indices.append(i)

    if valid_indices:
        rec_data = [rec_data[i] for i in valid_indices]

    # Per-document batch LLM: derive designations for all records at once
    # with full table context (one LLM call per document, not per record)
    if classifier and llm_client:
        batch_items = []
        for i, d in enumerate(rec_data):
            rk = d["rkz"]
            # Only re-derive if RKZ doesn't look like a proper IEC 81346 designation
            if not _looks_like_rkz(rk):
                rec_fields = {f.field_name: _safe_text(f.value)
                              for f in (records[i].results or [])
                              if f.value and str(f.value).strip()}
                batch_items.append((i, rec_fields, rk))

        if batch_items:
            try:
                batch_rkzs = _llm_derive_rkzs_batch(llm_client, batch_items)
                for i in range(len(rec_data)):
                    if i in batch_rkzs:
                        new_rkz = batch_rkzs[i]
                        if new_rkz and len(new_rkz) > 2:
                            rec_data[i]["rkz"] = new_rkz
                            rec_data[i]["name"] = new_rkz
            except Exception:
                pass  # LLM unavailable — fall back to current RKZ values

    # Batch LLM classification for Consumer-fallback elements
    # classify_batch sends one LLM call per uncached item (with lookup/keyword
    # fallback hitting the disk cache), so the first run populates the cache
    # and subsequent runs are nearly instant.
    if classifier:
        uncached = [
            (d["rkz"] or d["name"], d["desc"] or "",
             _safe_text(getattr(records[d["idx"]], "raw_context", ""))[:200] if hasattr(records[d["idx"]], "raw_context") else "")
            for d in rec_data
            if d["etype"] == "Consumer" and (d["name"] or d["desc"] or d["rkz"])
        ]
        if uncached:
            try:
                batch_results = classifier.classify_batch(uncached)
                result_idx = 0
                for d in rec_data:
                    if d["etype"] == "Consumer" and result_idx < len(batch_results):
                        r = batch_results[result_idx]
                        if r.is_valid and r.element_type != "Consumer":
                            d["etype"] = r.element_type
                        result_idx += 1
            except Exception:
                pass  # Fall back to Consumer on batch LLM error

    # Deduplicate RKZs: use LLM to derive unique designations for duplicates
    seen_rkzs: dict[str, list[int]] = {}  # rkz → [indices]
    for i, d in enumerate(rec_data):
        rk = d["rkz"]
        if rk not in seen_rkzs:
            seen_rkzs[rk] = []
        seen_rkzs[rk].append(i)

    duplicates = {rk: indices for rk, indices in seen_rkzs.items() if len(indices) > 1}

    if duplicates and llm_client and classifier:
        import hashlib
        for rk, dup_indices in duplicates.items():
            # Collect context for each duplicate element
            dup_contexts = []
            for idx in dup_indices:
                rec = records[idx]
                ctx_parts = []
                for field in (rec.results or []):
                    fv = _safe_text(field.value)
                    if fv:
                        ctx_parts.append(f"{field.field_name}={fv[:80]}")
                desc = rec_data[idx].get("desc", "")
                if desc:
                    ctx_parts.append(f"description={desc[:100]}")
                dup_contexts.append(" | ".join(ctx_parts[:10]))

            context_str = "||".join(dup_contexts[:15])[:800]
            dup_cache_key = f"rkz_dedup:{hashlib.sha256(context_str.encode()).hexdigest()[:16]}"
            cached = _cache_get(dup_cache_key)

            if cached is not None and isinstance(cached, dict):
                for idx_str, new_rkz in cached.items():
                    try:
                        rec_data[int(idx_str)]["rkz"] = new_rkz
                    except (ValueError, KeyError, IndexError):
                        pass
            else:
                dup_lines = "\n".join(
                    f"  Row {i}: {ctx[:200]}" for i, ctx in zip(dup_indices, dup_contexts)
                )
                system = (
                    "You are given rows that share the same reference designation. "
                    "Each row represents a DISTINCT element (e.g., different contacts of "
                    "a relay, different circuit breaker poles, different terminal positions). "
                    "Derive a UNIQUE designation for EACH row by adding proper IEC 81346 "
                    "sub-element suffixes: contact numbers (13/14, 21/22), terminal numbers "
                    "(1, 2, 3), pole markers (L1, L2, L3), or position indices.\n"
                    "Examples: '-K1:13/14', '-K1:21/22', '-F1/L1', '-F1/L2', "
                    "'-X1:1', '-X1:2'\n"
                    'Return JSON: {"designations": {"<row_index>": "<unique_designation>", ...}}'
                )
                user = f"Duplicate designation '{rk}':\n{dup_lines[:1500]}"
                try:
                    raw = llm_client.chat_json(system, user)
                    result_map = raw.get("designations", {})
                    clean_map = {}
                    for k, v in result_map.items():
                        try:
                            idx = int(k)
                            if idx in dup_indices and v and str(v).strip():
                                clean_map[str(k)] = str(v).strip()
                                rec_data[idx]["rkz"] = str(v).strip()
                        except (ValueError, TypeError):
                            pass
                    if clean_map:
                        _cache_put(dup_cache_key, clean_map)
                except Exception:
                    pass  # Fall back to simple suffix on LLM error

    # Fallback: simple suffix for any remaining duplicates that LLM couldn't resolve
    final_counts: dict[str, int] = {}
    for d in rec_data:
        rk = d["rkz"]
        if rk in final_counts:
            final_counts[rk] += 1
            # Only use numeric suffix as last resort — append contact-style
            if ":" not in rk:
                d["rkz"] = f"{rk}:{final_counts[rk]}"
            else:
                d["rkz"] = f"{rk}-{final_counts[rk]}"
        else:
            final_counts[rk] = 1

    # Second pass: build rows
    for d in rec_data:
        i = d["idx"]
        key_map[getattr(records[i], "record_key", str(i))] = g.peek("Elements_TopDown")
        td_id = g.next("Elements_TopDown")

        rows.append({
            "Index": len(rows) + 1,
            "Element_TopDown_ID": td_id,
            "Document_ID": doc_id,
            "Element_Name": d["name"][:200],
            "Primary_RKZ": d["rkz"][:200],
            "Element_Type": d["etype"],
        })
    return rows, key_map


# ══════════════════════════════════════════════════════════════════════════════
# P3 — Clusters
# ══════════════════════════════════════════════════════════════════════════════

def _p3_clusters(
    g: IDGen,
    doc_id: str,
    records: list[ExtractedRecord],
    object_rows: list[dict[str, Any]],
    source_artifact_index: SourceArtifactIndex | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Create Cluster + Object_Cluster rows.

    For V1: one Pre_Existing_Structural cluster per record (1:1 with elements).
    """
    cluster_rows: list[dict[str, Any]] = []
    oc_rows: list[dict[str, Any]] = []

    for i, rec in enumerate(records):
        cl_id = g.next("Cluster")
        cluster_rows.append({
            "Index": len(cluster_rows) + 1,
            "Cluster_ID": cl_id,
            "Document_ID": doc_id,
            "Cluster_Type": "Pre_Existing_Structural",
            "Cluster_Method": "Record_Grouping",
        })

        # Link the record's source artifact to this cluster (if available).
        record_obj_id = record_source_object_id(source_artifact_index, rec)
        if not record_obj_id and i < len(object_rows):
            record_obj_id = object_rows[i]["Object_ID"]
        if record_obj_id:
            oc_rows.append({
                "Index": len(oc_rows) + 1,
                "Object_ID": record_obj_id,
                "Cluster_ID": cl_id,
                "Membership_Reason": "Pre_Existing_Structural",
            })

    return cluster_rows, oc_rows


# ══════════════════════════════════════════════════════════════════════════════
# P4 — Elements from Cluster
# ══════════════════════════════════════════════════════════════════════════════

def _p4_elements_from_cluster(
    g: IDGen,
    doc_id: str,
    cluster_rows: list[dict[str, Any]],
    records: list[ExtractedRecord],
    llm_client: Any = None,
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    """Derive elements from clusters (one per cluster for V1)."""
    rows: list[dict[str, Any]] = []
    key_map: dict[str, str] = {}

    for i, (cluster, rec) in enumerate(zip(cluster_rows, records)):
        efc_id = g.next("Elements_from_Cluster")
        cl_id = cluster["Cluster_ID"]

        # Derive element from record
        rkz = ""
        etype = "Consumer"
        for field in (rec.results or []):
            fv = _safe_text(field.value)
            if not fv:
                continue
            fn = field.field_name.lower()
            if fn in ("primary_rkz", "terminal_id", "element_id", "canonicaltag", "akz"):
                rkz = fv
        if not rkz:
            rkz = getattr(rec, "record_key", str(i))

        # LLM classification (if available)
        if llm_client:
            name = ""
            context = ""
            for field in (rec.results or []):
                fn = field.field_name.lower()
                fv = _safe_text(field.value)
                if fn in ("terminal_name", "element_name", "name"):
                    name = fv
                elif fn == "raw_context":
                    context = fv[:200]
            try:
                from iev4pi_transformation_tool.core.llm_element_classifier import LLMElementClassifier
                classifier = LLMElementClassifier(llm_client)
                result = classifier.classify(rkz, name, context_text=context)
                if result.is_valid:
                    etype = result.element_type
            except Exception:
                pass

        key_map[getattr(rec, "record_key", str(i))] = efc_id

        rows.append({
            "Index": len(rows) + 1,
            "Element_from_Cluster_ID": efc_id,
            "Document_ID": doc_id,
            "Source_Cluster_ID": cl_id,
            "Element_Name": rkz[:200],
            "Primary_RKZ_Extracted": rkz[:200],
            "Element_Type_Inferred": etype,
            "Derivation_Status": "Element_Derived" if rkz else "No_Element_Derivable",
        })
    return rows, key_map


# ══════════════════════════════════════════════════════════════════════════════
# P5-P7 — Matching + Consolidation → Element_ID
# ══════════════════════════════════════════════════════════════════════════════

def _p5_p7_match_and_consolidate(
    g: IDGen,
    doc_id: str,
    td_rows: list[dict[str, Any]],
    efc_rows: list[dict[str, Any]],
    td_map: dict[str, str],
    efc_map: dict[str, str],
    records: list[ExtractedRecord],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, str]]:
    """P5: Match TopDown ↔ from_Cluster. P6: Resolve. P7: Consolidate → Element_ID.

    For V1: simple 1:1 match by position (each TD element matches the
    corresponding EFC element at the same index).
    """
    match_rows: list[dict[str, Any]] = []
    element_rows: list[dict[str, Any]] = []
    eid_map: dict[str, str] = {}

    for i in range(max(len(td_rows), len(efc_rows))):
        mr_id = g.next("Match_Result")
        td_id = td_rows[i]["Element_TopDown_ID"] if i < len(td_rows) else ""
        efc_id = efc_rows[i]["Element_from_Cluster_ID"] if i < len(efc_rows) else ""

        # Determine match status
        if td_id and efc_id:
            match_status = "Matched"
            resolution = "Resolved_AutoMatch"
            match_rule = "M1_Primary_RKZ"  # Positional for V1
        elif td_id:
            match_status = "Only_TopDown"
            resolution = "Open"
            match_rule = "Not_Applicable"
        else:
            match_status = "Only_Cluster"
            resolution = "Open"
            match_rule = "Not_Applicable"

        rev_status = "Auto_Approved" if match_status == "Matched" else "Requires_Review"
        match_rows.append({
            "Index": len(match_rows) + 1,
            "Match_ID": mr_id,
            "Document_ID": doc_id,
            "Element_TopDown_ID": td_id,
            "Element_from_Cluster_ID": efc_id,
            "Match_Status": match_status,
            "Match_Rule": match_rule,
            "Resolution_Status": resolution,
            "Review_Status": rev_status,
            "Review_Timestamp": _now_iso(),
        })

        # Create Element_ID from match
        e_id = g.next("Element_ID")
        rkz = td_rows[i].get("Primary_RKZ", "") if i < len(td_rows) else ""
        etype = td_rows[i].get("Element_Type", "Consumer") if i < len(td_rows) else "Consumer"

        # Connection-point sub-elements are Terminals (ExternalInterface per §5.13)
        # Patterns: "-X1:11", "-B6:1", "-K1:13/14", "-A0:L1", "-O12:U1", "4F1/3", "-A2:p1"
        if etype == "Consumer" and rkz:
            if re.search(r'[:/]\d+', rkz) or re.search(r'[:/](L[123]|N|PE|U\d+|V\d+|W\d+|p\d+)', rkz):
                etype = "Terminal"

        # Determine CAEX type
        info = get_aio_element_info(etype.lower().replace(" ", "_"))
        caex_type = info["caex_type"] if info else "InternalElement"
        iec_class = info["iec_class"] if info else "M"

        element_rows.append({
            "Index": len(element_rows) + 1,
            "Element_ID": e_id,
            "Document_ID": doc_id,
            "Source_Match_ID": mr_id,
            "Source": "Matched" if match_status == "Matched" else "TopDown",
            "Element_Type": etype,
            "Primary_RKZ": rkz[:200],
            "CAEX_Type": caex_type,
        })

        # Map record → element
        if i < len(records):
            rid = getattr(records[i], "record_id", str(i))
            eid_map[rid] = e_id

    return match_rows, element_rows, eid_map


# ══════════════════════════════════════════════════════════════════════════════
# P8 — Classification + Attributes
# ══════════════════════════════════════════════════════════════════════════════

def _llm_reclassify_consumers(
    element_rows: list[dict[str, Any]],
    records: list[ExtractedRecord],
    object_rows: list[dict[str, Any]],
    llm_client: Any,
) -> None:
    """Re-classify remaining Consumer elements with enriched LLM context.

    Collects raw_context, component_id, trace_path, and sibling element info
    for each Consumer element, then sends a batch LLM call. Results are cached
    on disk keyed by content hash.
    """
    import hashlib

    consumer_indices = [
        i for i, row in enumerate(element_rows)
        if row.get("Element_Type") == "Consumer"
    ]
    if not consumer_indices:
        return

    # Build context for each Consumer
    items: list[tuple[int, str, str]] = []  # (idx, rkz, enriched_context)
    for idx in consumer_indices:
        row = element_rows[idx]
        rkz = str(row.get("Primary_RKZ", ""))

        # Gather context from linked record
        rec_idx = idx if idx < len(records) else -1
        context_parts = []
        if rec_idx >= 0:
            rec = records[rec_idx]
            for field in (rec.results or []):
                fn = field.field_name.lower()
                fv = _safe_text(field.value)
                if not fv:
                    continue
                if fn in ("raw_context", "component_id", "trace_path",
                          "display_label", "logical_tag"):
                    context_parts.append(f"{fn}={fv[:120]}")
        context_str = " | ".join(context_parts[:6])

        # Add sibling context (elements sharing same page/region)
        sibling_types = []
        for j, other in enumerate(element_rows):
            if j != idx and other.get("Element_Type") != "Consumer":
                other_rkz = str(other.get("Primary_RKZ", ""))
                if other_rkz and abs(j - idx) <= 5:
                    sibling_types.append(f"{other_rkz}({other.get('Element_Type','')})")
        if sibling_types:
            context_str += f" | siblings: {', '.join(sibling_types[:8])}"

        items.append((idx, rkz, context_str))

    # Cache key from all items
    content_str = "||".join(f"{rkz}:{ctx[:120]}" for _, rkz, ctx in items)
    cache_key = f"consumer_reclassify:{hashlib.sha256(content_str.encode()).hexdigest()[:16]}"
    cached = _cache_get(cache_key)
    if cached is not None and isinstance(cached, dict):
        for idx_str, new_type in cached.items():
            try:
                element_rows[int(idx_str)]["Element_Type"] = str(new_type)
            except (ValueError, IndexError, KeyError):
                pass
        return

    # Build LLM prompt
    item_lines = []
    for item_idx, (elem_idx, rkz, ctx) in enumerate(items):
        item_lines.append(f'{item_idx}. RKZ="{rkz}", context="{ctx[:200]}"')

    system = (
        'Re-classify electrical engineering elements from their context. '
        'Available Element_Types: Terminal, Terminal_Strip, Contactor, Auxiliary_Contactor, '
        'Fuse, Circuit_Breaker, Switch, Socket_Outlet, Power_Supply, PLC_Module, Motor, '
        'Valve_Actuator, Sensor, Heater, Transducer, Actuator, Consumer, Coil, '
        'Main_Contact, Auxiliary_Contact, Indicator_Lamp, Cabinet_Aggregate\n\n'
        'Classification guidelines:\n'
        '- context mentioning SNT/Schaltnetzteil/Bemessungsleistung → Power_Supply\n'
        '- context with U/V/W + PE → motor winding terminals → Terminal\n'
        '- context with L1L2L3NPE → three-phase input block → Terminal\n'
        '- designation starting with -F → Fuse (unless CB/RCBO context)\n'
        '- designation starting with -A → Cabinet_Aggregate (unless SNT/PLC context)\n'
        '- designation starting with O (not IEC standard) → Consumer if unclear\n'
        '- trace_path or bbox context → likely Terminal (connection point)\n'
        'Return JSON: {"classifications": [{"index": 0, "element_type": "Fuse"}, ...]}'
    )
    user = "Re-classify:\n" + "\n".join(item_lines)

    try:
        raw = llm_client.chat_json(system, user)
        classifications = raw.get("classifications", [])
        updates: dict[str, str] = {}
        for cls in classifications:
            item_idx = int(cls.get("index", -1))
            new_type = str(cls.get("element_type", "")).strip()
            if 0 <= item_idx < len(items) and new_type in {
                "Terminal", "Terminal_Strip", "Contactor", "Auxiliary_Contactor",
                "Fuse", "Circuit_Breaker", "Switch", "Socket_Outlet", "Power_Supply",
                "PLC_Module", "Motor", "Valve_Actuator", "Sensor", "Heater",
                "Transducer", "Actuator", "Consumer", "Coil", "Main_Contact",
                "Auxiliary_Contact", "Indicator_Lamp", "Cabinet_Aggregate",
            }:
                elem_idx = items[item_idx][0]
                if new_type != "Consumer":
                    element_rows[elem_idx]["Element_Type"] = new_type
                    # Also update CAEX_Type
                    info = get_aio_element_info(new_type.lower().replace(" ", "_"))
                    if info:
                        element_rows[elem_idx]["CAEX_Type"] = info["caex_type"]
                    updates[str(elem_idx)] = new_type
        if updates:
            _cache_put(cache_key, updates)
    except Exception:
        pass  # LLM unavailable — keep Consumer


def _p8_classification_attributes(
    wb: openpyxl.Workbook,
    g: IDGen,
    doc_id: str,
    element_rows: list[dict[str, Any]],
    records: list[ExtractedRecord],
    object_rows: list[dict[str, Any]],
    source_artifact_index: SourceArtifactIndex | None = None,
    ml_evidence_linker: MLEvidenceLinker | None = None,
    progress: Callable[[int, str], None] | None = None,
) -> None:
    """Populate Element_Data, Element_Data_Source, Element_Classification,
    Element_Classification_Source, Element_RepresentedItem_Mapping, Layer_ID.
    """
    ed_rows: list[dict[str, Any]] = []
    eds_rows: list[dict[str, Any]] = []
    ec_rows: list[dict[str, Any]] = []
    ecs_rows: list[dict[str, Any]] = []
    erm_rows: list[dict[str, Any]] = []
    allowed_element_attrs = _element_attribute_lookup_names(wb)
    source_link_tasks: list[dict[str, Any]] = []

    # Layer_ID — create one default layer
    layer_id = g.next("Layer_ID")
    _write_sheet(wb, "Layer_ID", [{
        "Index": 1, "Layer_ID": layer_id, "Document_ID": doc_id,
        "Layer_Description": "Default", "Layer_Type": "Signal_Group",
    }])

    for i, (elem, rec) in enumerate(zip(element_rows, records)):
        e_id = elem["Element_ID"]
        etype = elem["Element_Type"]

        # Add layer reference to element
        elem["Layer_ID"] = layer_id

        # Element_Data rows — element-level only; connection fields → P9
        for field in (rec.results or []):
            if not field.value or not str(field.value).strip():
                continue
            fn = field.field_name

            if _is_aio_internal_element_field(fn, field.value):
                continue

            # Smart filtering: use regex patterns instead of hardcoded field lists
            if _is_connection_field(fn):
                continue  # Connection-level → collected and passed to P9

            attr_name = _element_attribute_name_for_field(field.field_name, allowed_element_attrs)
            if not attr_name:
                continue
            raw_val = _safe_text(field.value)

            ed_id = g.next("Element_Data")
            ed_rows.append({
                "Index": len(ed_rows) + 1,
                "Element_Data_ID": ed_id,
                "Element_ID": e_id,
                "Attribute_Name": attr_name,
                "Attribute_Value": raw_val[:500],
                "Raw_Value": raw_val[:500],
                "Parsing_Status": "Parsed_OK",
            })

            # Provenance: link to the actual source artifact selected for this field.
            fallback_obj_id = object_rows[min(i, len(object_rows) - 1)]["Object_ID"] if object_rows else ""
            source_link_tasks.append({
                "ed_id": ed_id,
                "record": rec,
                "field": field,
                "fallback_object_id": fallback_obj_id,
            })

        # Element_Classification — IEC 81346-2 class
        info = get_aio_element_info(etype.lower().replace(" ", "_"))
        if info:
            ec_id = g.next("Element_Classification")
            ec_rows.append({
                "Index": len(ec_rows) + 1,
                "Classification_ID": ec_id,
                "Document_ID": doc_id,
                "Classified_Object_Type": "Element",
                "Classified_Object_ID": e_id,
                "Classification_System": "IEC 81346-2",
                "Classification_Code": info["iec_class"],
            })
            # Element_Classification_Source (I12 provenance) — lookup table → Rule_Based_Parser
            obj_id = record_source_object_id(source_artifact_index, rec)
            if not obj_id and object_rows:
                obj_id = object_rows[min(i, len(object_rows) - 1)]["Object_ID"]
            ecs_rows.append({
                "Index": len(ecs_rows) + 1,
                "Classification_ID": ec_id,
                "Source_Object_ID": obj_id,
                "Source_Role": "Label",
                "Extraction_Method": "Rule_Based_Parser",
                "Confidence": 0.85,
                "Review_Status": "Auto_Approved",
                "Extraction_Timestamp": _now_iso(),
            })

        # Element_RepresentedItem_Mapping — link to first RepresentedItem
        ri_rows_list = _safe_read_sheet(wb, "Document_RepresentedItem")
        if ri_rows_list:
            erm_id = g.next("Element_RepresentedItem_Mapping")
            erm_rows.append({
                "Index": len(erm_rows) + 1,
                "Mapping_ID": erm_id,
                "Element_ID": e_id,
                "RepresentedItem_ID": ri_rows_list[0].get("RepresentedItem_ID", ""),
                "Relationship_Type": "Primary",
            })

    source_links = _source_links_for_element_data(
        source_artifact_index,
        ml_evidence_linker,
        source_link_tasks,
        progress=progress,
    )
    for task in source_link_tasks:
        ed_id = str(task["ed_id"])
        obj_id, field_conf, extraction_method, review_status = source_links.get(
            ed_id,
            ("", None, "Unspecifiable", "Requires_Review"),
        )
        if obj_id:
            eds_rows.append({
                "Index": len(eds_rows) + 1,
                "Element_Data_ID": ed_id,
                "Source_Object_ID": obj_id,
                "Source_Role": "Value",
                "Extraction_Method": extraction_method,
                "Confidence": field_conf,
                "Review_Status": review_status,
                "Extraction_Timestamp": _now_iso(),
            })

    _write_sheet(wb, "Element_Data", ed_rows)
    _write_sheet(wb, "Element_Data_Source", eds_rows)
    _write_sheet(wb, "Element_Classification", ec_rows)
    _write_sheet(wb, "Element_Classification_Source", ecs_rows)
    _write_sheet(wb, "Element_RepresentedItem_Mapping", erm_rows)


# ══════════════════════════════════════════════════════════════════════════════
# P9 — Connections
# ══════════════════════════════════════════════════════════════════════════════

def _p9_connections(
    wb: openpyxl.Workbook,
    g: IDGen,
    doc_id: str,
    element_rows: list[dict[str, Any]],
    records: list[ExtractedRecord],
    object_rows: list[dict[str, Any]],
    llm_client: Any = None,
    source_artifact_index: SourceArtifactIndex | None = None,
) -> None:
    """Create Connection_ID + Connection_Data rows.

    For V1: looks for wire/connection attributes in records and creates
    connection rows where from/to can be determined.
    """
    conn_rows: list[dict[str, Any]] = []
    cd_rows: list[dict[str, Any]] = []
    cds_rows: list[dict[str, Any]] = []

    # Build element lookup by RKZ (needed by both Excel and PDF connection paths)
    e_by_rkz: dict[str, str] = {}
    for elem in element_rows:
        rkz = elem.get("Primary_RKZ", "")
        if rkz:
            e_by_rkz[rkz.lower().strip()] = elem["Element_ID"]
    e_by_exact, e_by_base = _build_element_rkz_lookup(element_rows)
    existing_conns: set[tuple[str, str]] = set()

    # ── Excel Verschaltungsliste: infer terminal-to-device connections per row ──
    # Rows carry both a terminal strip cell (klemmleiste_*) and device/context
    # fields.  We only emit a connection when both sides resolve to real,
    # distinct Element_ID values; unresolved rows stay reviewable data rather
    # than becoming misleading self-connections.
    for rec in records:
        strip_id = ""
        terminal_num = ""
        wire_color_raw = ""
        fields = _record_fields(rec)
        for field in (rec.results or []):
            fn = field.field_name.lower()
            fv = _safe_text(field.value)
            if not fv: continue
            if fn.startswith("klemmleiste_"):
                strip_id = fn  # e.g., klemmleiste_x01
                terminal_num = fv
            if re.search(r'leiterfarbe|wire.color|aderfarbe', fn):
                wire_color_raw = fv
        if strip_id and terminal_num:
            terminal_candidates = _terminal_designation_candidates(strip_id, terminal_num)
            device_candidates = _device_designation_candidates(fields)
            from_eid = _resolve_element_id(terminal_candidates, e_by_exact, e_by_base)
            to_eid = _resolve_element_id(device_candidates, e_by_exact, e_by_base)

            if not from_eid or not to_eid or from_eid == to_eid:
                continue
            if (from_eid, to_eid) in existing_conns or (to_eid, from_eid) in existing_conns:
                continue

            c_id = g.next("Connection_ID")
            obj_ref = record_source_object_id(source_artifact_index, rec)
            if not obj_ref:
                obj_ref = object_rows[0]["Object_ID"] if object_rows else ""

            conn_rows.append({
                "Index": len(conn_rows) + 1,
                "Connection_ID": c_id,
                "Document_ID": doc_id,
                "From_Element_ID": from_eid,
                "To_Element_ID": to_eid,
                "Source_Topology_Object_ID": obj_ref,
                "Connection_Status": "Resolved",
            })
            existing_conns.add((from_eid, to_eid))

            # Connection_Data from terminal strip attributes
            # Try wire color first, then polarity from Verschaltung/Brücke columns
            attrs_added = set()
            for raw_color in [wire_color_raw]:
                if raw_color and "Wire_Color" not in attrs_added:
                    wc = get_wire_color_code(raw_color)
                    if wc and wc != "Unspecifiable":
                        cd_id = g.next("Connection_Data")
                        cd_rows.append({"Index": len(cd_rows)+1, "Connection_Data_ID": cd_id, "Connection_ID": c_id, "Attribute_Name": "Wire_Color", "Attribute_Value": wc, "Parsing_Status": "Parsed_OK"})
                        if obj_ref: cds_rows.append({"Index": len(cds_rows)+1, "Connection_Data_ID": cd_id, "Source_Object_ID": obj_ref, "Source_Role": "Value", "Extraction_Method": "Native_Text", "Confidence": 0.85, "Review_Status": "Auto_Approved", "Extraction_Timestamp": _now_iso()})
                        attrs_added.add("Wire_Color")
                        break

            # Extract polarity from Verschaltung column (e.g., "N", "L1", "PE")
            for field in (rec.results or []):
                fn = field.field_name.lower()
                if fn == "verschaltung":
                    pol = _extract_polarity_from_text(_safe_text(field.value))
                    if pol:
                        cd_id = g.next("Connection_Data")
                        cd_rows.append({"Index": len(cd_rows)+1, "Connection_Data_ID": cd_id, "Connection_ID": c_id, "Attribute_Name": "Polarity", "Attribute_Value": pol, "Parsing_Status": "Parsed_OK"})
                        if obj_ref: cds_rows.append({"Index": len(cds_rows)+1, "Connection_Data_ID": cd_id, "Source_Object_ID": obj_ref, "Source_Role": "Value", "Extraction_Method": "Native_Text", "Confidence": 0.85, "Review_Status": "Auto_Approved", "Extraction_Timestamp": _now_iso()})
                        attrs_added.add("Polarity")
                        break

    # Build cross-reference: component_id → {attribute: value} from ALL records
    # Needed because connection attributes (wire color, etc.) are often stored
    # in component records, not connection records.
    comp_attrs: dict[str, dict[str, str]] = {}
    for rec in records:
        comp_id = ""
        attrs: dict[str, str] = {}
        for field in (rec.results or []):
            fn = field.field_name.lower()
            fv = _safe_text(field.value)
            if not fv:
                continue
            if fn in ("component_id", "logical_tag", "display_label", "gerat",
                       "bezeichnung_im_stromlaufplan"):
                comp_id = comp_id or fv
            if re.search(r'leiterfarbe|wire.color|aderfarbe|farbe', fn):
                attrs["wire_color"] = get_wire_color_code(fv)
            if re.search(r'querschnitt|cross.section', fn):
                attrs["cross_section"] = fv
            if re.search(r'polarit', fn):
                attrs["polarity"] = get_polarity_code(fv)
        if comp_id:
            comp_attrs[comp_id.lower().strip()] = attrs

    # Get PDF path from first record (for VLM)
    pdf_path = ""
    for rec in records:
        sp = getattr(rec, "source_path", "") or ""
        if sp and sp.lower().endswith(".pdf"):
            pdf_path = sp
            break

    for rec in records:
        from_id = ""
        to_id = ""
        wire_color = ""
        cross_section = ""
        polarity = ""

        for field in (rec.results or []):
            fn = field.field_name.lower()
            fv = _safe_text(field.value)
            if not fv:
                continue
            # Smart from/to detection: any field starting with "from_" or "to_"
            # plus known German patterns (betriebsmittel_zugang/abgang, verschaltung)
            if re.match(r'from_', fn) or fn == "betriebsmittel_zugang":
                from_id = fv
            elif re.match(r'to_', fn) or fn == "betriebsmittel_abgang":
                to_id = fv
            # Wire color: leiterfarbe* (DE), wire_color (EN), aderfarbe (DE)
            elif re.search(r'leiterfarbe|wire.color|aderfarbe', fn):
                wire_color = get_wire_color_code(fv)
            elif re.search(r'querschnitt|cross.section', fn):
                cross_section = fv
            elif re.search(r'polarit', fn):
                polarity = get_polarity_code(fv)

        if not from_id and not to_id:
            continue

        # Cross-reference: look up wire color etc. from component records
        if not wire_color and from_id:
            attrs = comp_attrs.get(from_id.lower().strip(), {})
            wire_color = wire_color or attrs.get("wire_color", "")
            cross_section = cross_section or attrs.get("cross_section", "")
            polarity = polarity or attrs.get("polarity", "")
        if not wire_color and to_id:
            attrs = comp_attrs.get(to_id.lower().strip(), {})
            wire_color = wire_color or attrs.get("wire_color", "")
            cross_section = cross_section or attrs.get("cross_section", "")
            polarity = polarity or attrs.get("polarity", "")

        # Resolve to Element_IDs.  Do not synthesize endpoints from row position:
        # that creates plausible-looking but wrong connections for unresolved data.
        from_eid = _resolve_element_id([from_id], e_by_exact, e_by_base)
        to_eid = _resolve_element_id([to_id], e_by_exact, e_by_base)

        if not from_eid or not to_eid or from_eid == to_eid:
            continue
        if (from_eid, to_eid) in existing_conns or (to_eid, from_eid) in existing_conns:
            continue

        c_id = g.next("Connection_ID")
        obj_ref = record_source_object_id(source_artifact_index, rec)
        if not obj_ref:
            obj_ref = object_rows[0]["Object_ID"] if object_rows else ""

        conn_rows.append({
            "Index": len(conn_rows) + 1,
            "Connection_ID": c_id,
            "Document_ID": doc_id,
            "From_Element_ID": from_eid,
            "To_Element_ID": to_eid,
            "Source_Topology_Object_ID": obj_ref,
            "Connection_Status": "Resolved",
        })
        existing_conns.add((from_eid, to_eid))

        # ── LLM-based attribute extraction (replaces hardcoded regex) ──
        # Collect raw_context and wire_label for LLM semantic parsing
        ctx_text = ""
        wl_text = ""
        for field in (rec.results or []):
            if field.field_name == "raw_context":
                ctx_text = _safe_text(field.value)
            elif field.field_name == "wire_label":
                wl_text = _safe_text(field.value)

        # LLM: extract polarity, wire_color, cross_section from text context
        if llm_client and (ctx_text or wl_text):
            llm_attrs = _extract_connection_attrs_llm(
                ctx_text, wl_text, from_id, to_id, llm_client
            )
            wire_color = wire_color or llm_attrs.get("Wire_Color", "")
            cross_section = cross_section or llm_attrs.get("Cross_Section", "")
            polarity = polarity or llm_attrs.get("Polarity", "")

        # VLM: look up connection attributes from full-page text-label extraction
        # Only returns Polarity and Cross_Section from text labels — no wire color inference
        if llm_client and pdf_path:
            vlm_conns = _extract_connections_vlm(pdf_path, llm_client)
            vlm_key = (from_id.lower().strip(), to_id.lower().strip())
            vlm_attrs = vlm_conns.get(vlm_key, {})
            if vlm_attrs:
                cross_section = cross_section or vlm_attrs.get("Cross_Section", "")
                polarity = polarity or vlm_attrs.get("Polarity", "")

        # Connection attributes
        for attr_name, attr_val in [("Wire_Color", wire_color),
                                      ("Cross_Section", cross_section),
                                      ("Polarity", polarity)]:
            if not attr_val:
                continue
            cd_id = g.next("Connection_Data")
            cd_rows.append({
                "Index": len(cd_rows) + 1,
                "Connection_Data_ID": cd_id,
                "Connection_ID": c_id,
                "Attribute_Name": attr_name,
                "Attribute_Value": attr_val,
                "Parsing_Status": "Parsed_OK",
            })
            if obj_ref:
                cds_rows.append({
                    "Index": len(cds_rows) + 1,
                    "Connection_Data_ID": cd_id,
                    "Source_Object_ID": obj_ref,
                    "Source_Role": "Value",
                    "Extraction_Method": "Rule_Based_Parser",
                    "Confidence": 0.8,
                    "Extraction_Timestamp": _now_iso(),
                })

    # ── VLM-primary connection creation for PDF sources (§A.7 PDF convention) ──
    # VLM extracts connections from the PDF image with terminal labels that may
    # not match record-based from/to IDs.  This pass creates connections directly
    # from VLM data, matching VLM labels to Element RKZs via normalized lookup.
    if llm_client and pdf_path:
        vlm_conns_all = _extract_connections_vlm(pdf_path, llm_client)
        if vlm_conns_all:
            # Build normalized element lookup: (base_rkz) → element_id
            # base_rkz strips leading -, trailing :terminal, and lowers case
            e_by_base: dict[str, str] = {}
            e_by_exact_lower: dict[str, str] = {}
            for elem in element_rows:
                rkz = str(elem.get("Primary_RKZ", ""))
                eid = elem.get("Element_ID", "")
                if not rkz or not eid:
                    continue
                rkz_lower = rkz.lower().strip()
                e_by_exact_lower[rkz_lower] = eid
                # Base: strip leading - and :suffix
                base = re.sub(r'^-', '', rkz_lower)
                base = re.sub(r'[:/].*$', '', base)
                if base and base not in e_by_base:
                    e_by_base[base] = eid

            # Track existing connections to avoid duplicates
            for cr in conn_rows:
                f = str(cr.get("From_Element_ID", ""))
                t = str(cr.get("To_Element_ID", ""))
                if f and t:
                    existing_conns.add((f, t))

            vlm_connections_added = 0
            for (vlm_from, vlm_to), vlm_attrs in vlm_conns_all.items():
                pol = vlm_attrs.get("Polarity", "")
                cs = vlm_attrs.get("Cross_Section", "")

                # Skip if no useful attributes (text-verified Polarity or Cross_Section)
                if not pol and not cs:
                    continue

                # Match VLM labels to Element IDs
                from_eid = _match_vlm_label_to_element(vlm_from, e_by_exact_lower, e_by_base, e_by_rkz)
                to_eid = _match_vlm_label_to_element(vlm_to, e_by_exact_lower, e_by_base, e_by_rkz)

                if not from_eid or not to_eid or from_eid == to_eid:
                    continue

                # Skip if connection already exists
                if (from_eid, to_eid) in existing_conns or (to_eid, from_eid) in existing_conns:
                    continue

                c_id = g.next("Connection_ID")
                obj_ref = object_rows[0]["Object_ID"] if object_rows else ""

                conn_rows.append({
                    "Index": len(conn_rows) + 1,
                    "Connection_ID": c_id,
                    "Document_ID": doc_id,
                    "From_Element_ID": from_eid,
                    "To_Element_ID": to_eid,
                    "Source_Topology_Object_ID": obj_ref,
                    "Connection_Status": "Resolved",
                })
                existing_conns.add((from_eid, to_eid))

                # Connection_Data from VLM (text-verified Polarity + Cross_Section only)
                for attr_name, attr_val in [("Polarity", pol),
                                              ("Cross_Section", cs)]:
                    if not attr_val:
                        continue
                    cd_id = g.next("Connection_Data")
                    cd_rows.append({
                        "Index": len(cd_rows) + 1,
                        "Connection_Data_ID": cd_id,
                        "Connection_ID": c_id,
                        "Attribute_Name": attr_name,
                        "Attribute_Value": attr_val,
                        "Parsing_Status": "Parsed_OK",
                    })
                    if obj_ref:
                        cds_rows.append({
                            "Index": len(cds_rows) + 1,
                            "Connection_Data_ID": cd_id,
                            "Source_Object_ID": obj_ref,
                            "Source_Role": "Value",
                            "Extraction_Method": "LLM_Classification",
                            "Confidence": 0.6,
                            "Extraction_Timestamp": _now_iso(),
                        })
                vlm_connections_added += 1

    _write_sheet(wb, "Connection_ID", conn_rows)
    _write_sheet(wb, "Connection_Data", cd_rows)
    _write_sheet(wb, "Connection_Data_Source", cds_rows)


def _match_vlm_label_to_element(
    vlm_label: str,
    e_by_exact_lower: dict[str, str],
    e_by_base: dict[str, str],
    e_by_rkz: dict[str, str],
) -> str:
    """Match a VLM connection label to an Element_ID via layered lookup.

    Layer 1: exact case-insensitive match
    Layer 2: base-name match (strip leading -, :suffix)
    Layer 3: fuzzy prefix match (first 2 chars of base name)
    """
    label_lower = vlm_label.lower().strip()

    # 1. Exact match
    if label_lower in e_by_exact_lower:
        return e_by_exact_lower[label_lower]

    # 2. Base-name match
    base = re.sub(r'^-', '', label_lower)
    base = re.sub(r'[:/].*$', '', base)
    if base in e_by_base:
        return e_by_base[base]

    # 3. Fuzzy: try matching by prefix (first 2-3 chars)
    if len(base) >= 2:
        prefix = base[:3]
        for rkz_lower, eid in sorted(e_by_exact_lower.items()):
            rkz_base = re.sub(r'^-', '', rkz_lower)
            rkz_base = re.sub(r'[:/].*$', '', rkz_base)
            if rkz_base.startswith(prefix) or prefix.startswith(rkz_base):
                return eid

    # 4. Try original e_by_rkz (full RKZ → element ID)
    if label_lower in e_by_rkz:
        return e_by_rkz[label_lower]
    if base in e_by_rkz:
        return e_by_rkz[base]

    return ""


# ══════════════════════════════════════════════════════════════════════════════
# Utility
# ══════════════════════════════════════════════════════════════════════════════

def _normalize_all_attribute_names(
    wb: openpyxl.Workbook,
    llm_client: Any,
) -> None:
    """Post-pass: normalize all Attribute_Name values against Attribute_Lookup using LLM.

    Reads allowed attribute names per scope from the template's Attribute_Lookup sheet,
    then checks and fixes every Attribute_Name in Document_Data, Element_Data,
    RepresentedItem_Data, and Connection_Data. Uses LLM semantic matching with
    disk caching for unknown field names.
    """
    if "Attribute_Lookup" not in wb.sheetnames:
        return

    # 1. Build allowed attribute names per scope from the template
    allowed_by_scope: dict[str, set[str]] = {}
    al_ws = wb["Attribute_Lookup"]
    for row_num in range(2, al_ws.max_row + 1):
        scope = _safe_text(al_ws.cell(row=row_num, column=3).value)  # Scope column
        attr_name = _safe_text(al_ws.cell(row=row_num, column=5).value)  # Attribute_Name column
        if scope and attr_name:
            if scope not in allowed_by_scope:
                allowed_by_scope[scope] = set()
            allowed_by_scope[scope].add(attr_name)

    # 2. Collect all unique attribute names per scope from data sheets
    scope_to_sheet = {
        "Document": "Document_Data",
        "Element": "Element_Data",
        "RepresentedItem": "RepresentedItem_Data",
        "Connection": "Connection_Data",
        # Also handle combined scope+constraint patterns like "RepresentedItem_Type=Terminal_Strip"
    }

    # Build combined allowed set per scope (include both plain scope and constraint-prefixed entries)
    combined_allowed: dict[str, set[str]] = {}
    for scope_key, attrs in allowed_by_scope.items():
        # scope_key may be "Element" or "Element_Type=Terminal"
        base_scope = scope_key.split("_")[0] if "_" in scope_key and not scope_key.startswith("RepresentedItem_Type") else scope_key
        # Simple approach: collect ALL attributes for each scope regardless of Type_Constraint
        for s in ["Document", "Element", "RepresentedItem", "Connection"]:
            if scope_key.startswith(s):
                if s not in combined_allowed:
                    combined_allowed[s] = set()
                combined_allowed[s].update(attrs)

    # Gather fields needing normalization
    unknowns: list[tuple[str, str, str, int]] = []  # (sheet, attr_col_letter, attr, row)

    for scope, sheet_name in scope_to_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = {}
        for col in range(1, ws.max_column + 1):
            h = _safe_text(ws.cell(row=1, column=col).value)
            if h:
                headers[h] = col
        if "Attribute_Name" not in headers:
            continue
        attr_col = headers["Attribute_Name"]
        allowed = combined_allowed.get(scope, set())
        if not allowed:
            continue

        for row_num in range(2, ws.max_row + 1):
            attr_name = _safe_text(ws.cell(row=row_num, column=attr_col).value)
            if not attr_name:
                continue
            if attr_name not in allowed:
                from openpyxl.utils import get_column_letter
                unknowns.append((sheet_name, get_column_letter(attr_col), attr_name, row_num))

    if not unknowns:
        return

    # 3. LLM-based normalization (with caching)
    import hashlib
    for sheet_name, col_letter, attr_name, row_num in unknowns:
        scope = next((s for s, sn in scope_to_sheet.items() if sn == sheet_name), "Element")
        allowed_list = sorted(combined_allowed.get(scope, set()))

        # Cache key
        content_hash = hashlib.sha256(
            f"{attr_name}|{scope}|{','.join(allowed_list[:30])}".encode()
        ).hexdigest()[:16]
        cache_key = f"attr_norm:{content_hash}"
        cached = _cache_get(cache_key)
        if cached is not None and isinstance(cached, dict):
            new_name = cached.get(attr_name, "")
            if new_name and new_name in combined_allowed.get(scope, set()):
                ws = wb[sheet_name]
                ws[f"{col_letter}{row_num}"] = new_name
                continue

        # LLM call
        system = (
            f"Map a source field name to the closest standard attribute name for scope '{scope}'.\n"
            f"Available standard names:\n" + "\n".join(f"  - {n}" for n in allowed_list[:30])
            + "\n\nReturn the SINGLE closest match. If none, return 'Unspecifiable'.\n"
            'Return JSON: {"mapped_name": "Standard_Name"}'
        )
        try:
            raw = llm_client.chat_json(system, f"Field: {attr_name}")
            new_name = str(raw.get("mapped_name", "")).strip()
        except Exception:
            new_name = ""

        if new_name and new_name in combined_allowed.get(scope, set()):
            _cache_put(cache_key, {attr_name: new_name})
            ws = wb[sheet_name]
            ws[f"{col_letter}{row_num}"] = new_name


def _fill_normalized_values(wb: openpyxl.Workbook) -> None:
    """Post-process data sheets: fill Normalized_Value/Unit/Quantity_Qualifier
    from Attribute_Value where they are empty and the value is parseable."""
    data_sheets = ["Document_Data", "Element_Data", "RepresentedItem_Data", "Connection_Data"]
    for sn in data_sheets:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        headers = {}
        for col in range(1, ws.max_column + 1):
            h = _safe_text(ws.cell(row=1, column=col).value)
            if h:
                headers[h] = col

        needed = {"Attribute_Value", "Normalized_Value", "Unit"}
        if not needed.issubset(headers):
            continue

        av_col = headers["Attribute_Value"]
        nv_col = headers["Normalized_Value"]
        unit_col = headers["Unit"]
        qq_col = headers.get("Quantity_Qualifier")
        raw_col = headers.get("Raw_Value")

        for row_num in range(2, ws.max_row + 1):
            # Skip if Normalized_Value already filled
            nv_cell = ws.cell(row=row_num, column=nv_col)
            if nv_cell.value and str(nv_cell.value).strip():
                continue

            av = _safe_text(ws.cell(row=row_num, column=av_col).value)
            if not av:
                continue

            parsed = _parse_attribute_value(av)
            if not parsed:
                continue

            nv_cell.value = parsed.get("Normalized_Value", "")
            ws.cell(row=row_num, column=unit_col).value = parsed.get("Unit", "")
            if qq_col and "Quantity_Qualifier" in parsed:
                ws.cell(row=row_num, column=qq_col).value = parsed["Quantity_Qualifier"]
            # Fill Raw_Value if empty
            if raw_col:
                rv_cell = ws.cell(row=row_num, column=raw_col)
                if not rv_cell.value or not str(rv_cell.value).strip():
                    rv_cell.value = av


def _write_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    rows: list[dict[str, Any]],
) -> None:
    """Write rows to a named sheet in the workbook."""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    # Preserve seed sheets (Rules, Attribute_Lookup, Enum_Lookup, Schema_Metadata)
    if sheet_name in SEED_SHEETS:
        return  # These sheets already have pre-populated data

    _write_data_rows(ws, rows)


def _safe_read_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
) -> list[dict[str, str]]:
    """Read a sheet's data rows as list of dicts."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    # Read headers from row 1
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        headers.append(_safe_text(ws.cell(row=1, column=col).value))

    rows: list[dict[str, str]] = []
    for row_num in range(2, ws.max_row + 1):
        row_dict: dict[str, str] = {}
        for col, h in enumerate(headers, 1):
            val = ws.cell(row=row_num, column=col).value
            if val is not None:
                row_dict[h] = _safe_text(val)
        if row_dict:
            rows.append(row_dict)
    return rows
