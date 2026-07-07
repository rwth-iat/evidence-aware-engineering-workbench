"""Assembly 3D pipeline service — build Assembly template from FreeCAD source."""
from __future__ import annotations

import csv
import math
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

import ifcopenshell
import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

from iev4pi_transformation_tool.models import DocumentFamily

# ══════════════════════════════════════════════════════════════════════════
HF = Font(name="Calibri", size=11, bold=True)
HFILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
GFILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RFILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BD = Border(left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))

SHEET_ASSEMBLY_STEPS = [
    "assembly_no", "label", "type", "akz_tag", "vv_tag",
    "fcstd_link_name", "source_fcstd_file", "ifc_parts_library", "ifc_part_name", "ifc_global_id",
    "pos_x_mm", "pos_y_mm", "pos_z_mm",
    "rot_qw", "rot_qx", "rot_qy", "rot_qz",
    "rot_axis_x", "rot_axis_y", "rot_axis_z", "rot_angle_deg",
    "euler_roll_deg", "euler_pitch_deg", "euler_yaw_deg",
    "forward_x", "forward_y", "forward_z",
    "mapping_confidence", "ifc_coverage", "notes",
]

SHEET_CONN_TOPOLOGY = [
    "connection_id", "from_assembly_no", "from_label", "from_port",
    "to_assembly_no", "to_label", "to_port", "connection_type",
    "from_pos_x", "from_pos_y", "from_pos_z",
    "to_pos_x", "to_pos_y", "to_pos_z", "euclidean_distance_mm",
]

SHEET_CONN_ANGLES = [
    "connection_id", "from_label", "to_label",
    "from_forward_x", "from_forward_y", "from_forward_z",
    "to_forward_x", "to_forward_y", "to_forward_z",
    "connection_vector_x", "connection_vector_y", "connection_vector_z",
    "bend_angle_deg", "torsion_angle_deg", "from_to_distance_mm",
]

SHEET_POSITION = [
    "fcstd_link_name", "label", "type", "source_fcstd_file", "ifc_parts_library",
    "pos_x_mm", "pos_y_mm", "pos_z_mm",
    "qw", "qx", "qy", "qz", "axis_x", "axis_y", "axis_z", "angle_deg",
    "euler_roll_deg", "euler_pitch_deg", "euler_yaw_deg",
]

SHEET_PART_LIBRARY = [
    "part_id", "part_ref_name", "part_type", "fcstd_link_name",
    "source_fcstd_file", "ifc_parts_library", "ifc_part_name",
    "instance_count", "ifc_coverage", "description",
]

SHEET_TREE = ["tree_id", "parent_label", "child_label", "aggregation_type", "assembly_level", "notes"]


DEFAULT_AKZ_MAPPING = """tank_001, TU10.B1
tank_002, TU20.B2
tank_003, TU30.B3
blackbox_001, TU10.N18, PL002
blackbox_002, TU10.N13, PL001
blackbox_003, TU10.P14
blackbox_004, TU10.P19
blackbox_005, TU30.F40
blackbox_007, TU30.Y37, VV009
blackbox_008, TU30.Y39, VV010
blackbox_011, TU20.F31
blackbox_012, TU20.Y30, VV006
blackbox_013, TU20.N29, PL003
blackbox_014, HE002
blackbox_015, TU10.T15
blackbox_016, TU10.T20
blackbox_017, TU10.Y21, VV002
blackbox_018, TU10.F22
blackbox_019, TU10.Y6, VV001
blackbox_020, TU10.F17
blackbox_021, TU10.T23
blackbox_022, TU10.Y25
blackbox_023, TU10.Y24
blackbox_024, TU10.L10
blackbox_025, TU30.L35
blackbox_026, TU30.T33
blackbox_027, TU30.L34
blackbox_028, TU20.L26
blackbox_029, TU20.T27
blackbox_030, TU20.Q28
blackbox_031, TU30.L32
blackbox_032, TU30.N38
blackbox_033, TU30.N36
blackbox_034, VV007
blackbox_035, VV008
"""


class Assembly3DPipelineService:
    """Build the Assembly_3D_template_filled.xlsx from a FreeCAD assembly."""

    def __init__(self, workspace_root: Path) -> None:
        self.workspace_root = workspace_root.resolve()
        self.piping_dir = workspace_root / "Documents" / "Piping Diagram"

    # ── Smart IFC discovery ─────────────────────────────────────────────

    def discover_ifc_sources(self) -> tuple[Path | None, Path | None, str]:
        """Auto-discover primary and fallback IFC files.

        Scans ``Documents/Piping Diagram/`` for ``.ifc`` files and classifies
        them by their characteristics:
          - **primary**: entities with ObjectPlacement NOT all at origin
          - **fallback**: entities all at origin with a detectable suffix
            pattern (e.g. ``_locked``)

        Returns ``(primary_path, fallback_path, fallback_suffix)`` where
        *fallback_suffix* is the naming pattern to strip (e.g. ``_locked``).
        """
        primary: Path | None = None
        fallback: Path | None = None
        fallback_suffix = ""

        for ifc_path in sorted(self.piping_dir.glob("*.ifc")):
            try:
                f = ifcopenshell.open(str(ifc_path))
            except Exception:
                continue

            # Discover entity types dynamically
            placed = 0; at_origin = 0
            name_samples: list[str] = []
            for entity in f:
                etype = entity.is_a()
                # Skip non-component types
                if etype in ("IfcProject", "IfcSite", "IfcBuilding",
                             "IfcBuildingStorey", "IfcOwnerHistory",
                             "IfcPerson", "IfcOrganization", "IfcApplication",
                             "IfcUnitAssignment", "IfcSIUnit", "IfcDirection",
                             "IfcCartesianPoint", "IfcAxis2Placement3D",
                             "IfcLocalPlacement", "IfcShapeRepresentation",
                             "IfcProductDefinitionShape", "IfcTriangulatedFaceSet",
                             "IfcCartesianPointList3D", "IfcGeometricRepresentationContext",
                             "IfcGeometricRepresentationSubContext",
                             "IfcRelAggregates", "IfcRelContainedInSpatialStructure",
                             "IfcDimensionalExponents", "IfcMeasureWithUnit",
                             "IfcConversionBasedUnit", "IfcPersonAndOrganization"):
                    continue
                try:
                    name = entity.Name or ""
                except AttributeError:
                    continue
                if not name:
                    continue
                name_samples.append(name)
                try:
                    op = entity.ObjectPlacement
                except AttributeError:
                    continue
                if op and op.RelativePlacement and op.RelativePlacement.Location:
                    c = op.RelativePlacement.Location.Coordinates
                    placed += 1
                    if abs(c[0]) < 0.1 and abs(c[1]) < 0.1 and abs(c[2]) < 0.1:
                        at_origin += 1

            if placed == 0:
                continue

            # Classify
            if placed > 0 and at_origin < placed:  # has real placements
                primary = ifc_path
            elif placed > 0 and at_origin == placed:  # all at origin → fallback
                fallback = ifc_path
                fallback_suffix = self._guess_name_suffix(name_samples)

        return primary, fallback, fallback_suffix

    @staticmethod
    def _guess_name_suffix(samples: list[str]) -> str:
        """Detect naming suffix like ``_locked`` from entity name samples."""
        if not samples:
            return ""
        # Look for common trailing pattern after semantic base names
        from collections import Counter
        suffixes: Counter = Counter()
        for name in samples:
            # e.g. tank_001_locked → split on last '_' pattern
            parts = name.rsplit("_", 2)
            if len(parts) >= 2:
                # Check if stripping the last part gives a shorter base
                base = "_".join(parts[:-1])
                suffix = "_" + parts[-1]
                if len(base) < len(name) and len(suffix) > 1:
                    suffixes[suffix] += 1
        if suffixes:
            top = suffixes.most_common(1)[0]
            if top[1] >= len(samples) * 0.8:  # 80%+ consistency
                return top[0]
        return ""

    # ── Public API ────────────────────────────────────────────────────────

    def build_template_from_ifc(
        self,
        ifc_path: Path,
        csv_path: Path | None = None,
        akz_raw: str = "",
        output_path: str | Path | None = None,
        fallback_ifc: Path | None = None,
    ) -> Path:
        """Build Assembly_3D_template_filled.xlsx from a single assembled IFC.

        The IFC should have entities named by their semantic label (tank_001,
        blackbox_019, etc.).  Use ``scripts/prepare_assembled_ifc.py`` to
        rename entities from FreeCAD internal names if needed.

        When *fallback_ifc* is provided, entities not found in the primary
        IFC are looked up by fuzzy label matching in the fallback
        (auto-detected suffix such as ``_locked``).
        """
        csv_path = csv_path or (ifc_path.parent / "_legacy" / "final_result_components.csv")
        akz_map, vv_map = self._parse_akz(akz_raw or DEFAULT_AKZ_MAPPING)

        comps = self._parse_assembled_ifc(ifc_path)
        ifc_idx = self._build_ifc_index_from_single(ifc_path)

        # Build fallback index (label → GlobalId) from LOCKED-style IFC
        fallback_idx: dict[str, tuple[str, str]] = {}
        if fallback_ifc and fallback_ifc.exists():
            fallback_idx = self._build_fallback_ifc_index(fallback_ifc)

        # Compute assembly order (embedded sheet → CSV fallback)
        output_path = Path(output_path) if output_path else (self.piping_dir / "Assembly_3D_template_filled.xlsx")
        csv_order = self._read_component_order(csv_path=csv_path,
                                               existing_template=output_path)

        comps_with = [c for c in comps if c["label"] in csv_order]
        comps_without = [c for c in comps if c["label"] not in csv_order]
        comps_with.sort(key=lambda c: csv_order[c["label"]])
        sorted_comps = comps_with + comps_without
        label_to_ano = {c["label"]: i + 1 for i, c in enumerate(sorted_comps)}

        # Build steps
        steps = []
        for c in sorted_comps:
            label = c["label"]
            ctype = self._classify(label)
            euler = self._q2e(c["qw"], c["qx"], c["qy"], c["qz"])
            fwd = self._qrot(c["qw"], c["qx"], c["qy"], c["qz"], 0, 0, 1)

            ifc_file, ifc_part, ifc_gid = "", "", ""
            part_name = c.get("ifc_part_name", c["name"])
            if part_name and part_name in ifc_idx:
                ifc_file, ifc_gid = ifc_idx[part_name]
                ifc_part = part_name
            else:
                for try_name in (c["name"], c["label"]):
                    if try_name and try_name in ifc_idx:
                        ifc_file, ifc_gid = ifc_idx[try_name]
                        ifc_part = try_name
                        break
            # Fallback: lookup by label in LOCKED-style IFC
            if not ifc_file and fallback_idx:
                fb_key = self._match_fallback_label(c['label'], fallback_idx)
                if fb_key:
                    ifc_file, ifc_gid = fallback_idx[fb_key]
                    ifc_part = fb_key

            akz = akz_map.get(label, "")
            vv = vv_map.get(label, "")
            notes = ""
            if akz in ("TU10.Y6", "VV005"):
                notes = "uc1_expected_missing"
            elif akz in ("TU10.U41", "TU20.U42"):
                notes = "software_control_no_ifc_required"

            steps.append({
                "assembly_no": label_to_ano.get(label, ""),
                "label": label, "type": ctype,
                "akz_tag": akz, "vv_tag": vv,
                "fcstd_link_name": c["name"],
                "source_fcstd_file": c.get("src", ifc_path.name),
                "ifc_parts_library": ifc_file, "ifc_part_name": ifc_part, "ifc_global_id": ifc_gid,
                "pos_x_mm": round(c["px"], 4), "pos_y_mm": round(c["py"], 4), "pos_z_mm": round(c["pz"], 4),
                "rot_qw": round(c["qw"], 9), "rot_qx": round(c["qx"], 9),
                "rot_qy": round(c["qy"], 9), "rot_qz": round(c["qz"], 9),
                "rot_axis_x": round(c["ax"], 6), "rot_axis_y": round(c["ay"], 6), "rot_axis_z": round(c["az"], 6),
                "rot_angle_deg": round(math.degrees(c["angle"]), 4),
                "euler_roll_deg": round(euler[0], 4), "euler_pitch_deg": round(euler[1], 4), "euler_yaw_deg": round(euler[2], 4),
                "forward_x": round(fwd[0], 4), "forward_y": round(fwd[1], 4), "forward_z": round(fwd[2], 4),
                "mapping_confidence": "matched",
                "ifc_coverage": "YES" if ifc_file else "NO",
                "notes": notes,
            })

        # Append CSV entries resolved from fallback (LOCKED) IFC when
        # the primary IFC does not have a matching entity.  Only entries
        # that actually get a GlobalId from the fallback are added.
        # They are placed after all IFC-based entries to avoid
        # assembly_no collisions.
        seen_labels = {s["label"] for s in steps}
        max_ano = max(
            (int(s["assembly_no"]) for s in steps
             if isinstance(s["assembly_no"], (int, str))
             and str(s["assembly_no"]).isdigit()),
            default=0,
        )
        for label in csv_order:
            if label in seen_labels:
                continue
            ifc_file, ifc_part, ifc_gid = "", "", ""
            if fallback_idx:
                fb_key = self._match_fallback_label(label, fallback_idx)
                if fb_key:
                    ifc_file, ifc_gid = fallback_idx[fb_key]
                    ifc_part = fb_key
            if not ifc_file:
                continue  # Not in primary or fallback IFC → skip
            max_ano += 1
            ctype = self._classify(label)
            akz = akz_map.get(label, "")
            vv = vv_map.get(label, "")
            steps.append({
                "assembly_no": max_ano,
                "label": label, "type": ctype,
                "akz_tag": akz, "vv_tag": vv,
                "fcstd_link_name": label, "source_fcstd_file": "",
                "ifc_parts_library": ifc_file, "ifc_part_name": ifc_part, "ifc_global_id": ifc_gid,
                "pos_x_mm": 0.0, "pos_y_mm": 0.0, "pos_z_mm": 0.0,
                "rot_qw": 1.0, "rot_qx": 0.0, "rot_qy": 0.0, "rot_qz": 0.0,
                "rot_axis_x": 0.0, "rot_axis_y": 0.0, "rot_axis_z": 1.0, "rot_angle_deg": 0.0,
                "euler_roll_deg": 0.0, "euler_pitch_deg": 0.0, "euler_yaw_deg": 0.0,
                "forward_x": 0.0, "forward_y": 0.0, "forward_z": 1.0,
                "mapping_confidence": "fallback",
                "ifc_coverage": "YES",
                "notes": "fallback_ifc",
            })
            seen_labels.add(label)

        part_lib = self._build_part_library(steps)
        connections = self._build_connections(steps)
        angles = self._build_angles(connections, steps)
        positions = self._build_positions(steps)
        tree = self._build_tree(steps)

        self.piping_dir.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._write_sheet(wb, "Part_Library", SHEET_PART_LIBRARY, list(part_lib.values()))
        self._write_sheet(wb, "Assembly_Steps", SHEET_ASSEMBLY_STEPS, steps)
        self._write_sheet(wb, "Connection_Topology", SHEET_CONN_TOPOLOGY, connections)
        self._write_sheet(wb, "Connection_Angles", SHEET_CONN_ANGLES, angles)
        self._write_sheet(wb, "Position_Data", SHEET_POSITION, positions)
        self._write_sheet(wb, "Assembly_Tree", SHEET_TREE, tree)

        akz_summary = [s for s in steps if s["akz_tag"] or s["vv_tag"]]
        self._write_sheet(wb, "AKZ_Summary",
            ["label", "type", "akz_tag", "vv_tag", "pos_x_mm", "pos_y_mm", "pos_z_mm",
             "ifc_parts_library", "ifc_part_name", "ifc_global_id"],
            [{k: s.get(k, "") for k in ["label", "type", "akz_tag", "vv_tag",
                "pos_x_mm", "pos_y_mm", "pos_z_mm",
                "ifc_parts_library", "ifc_part_name", "ifc_global_id"]} for s in akz_summary])
        ifc_gaps = [s for s in steps if s["ifc_coverage"] == "NO"]
        self._write_sheet(wb, "IFC_Gaps",
            ["label", "type", "fcstd_link_name", "source_fcstd_file",
             "pos_x_mm", "pos_y_mm", "pos_z_mm"],
            [{k: s.get(k, "") for k in ["label", "type", "fcstd_link_name",
                "source_fcstd_file", "pos_x_mm", "pos_y_mm", "pos_z_mm"]} for s in ifc_gaps])

        self._write_component_order(wb, steps)
        wb.save(str(output_path))
        return output_path

    @staticmethod
    def _read_component_order(csv_path: Path | None = None,
                               existing_template: Path | None = None) -> dict[str, int]:
        """Read label → assembly_no order.

        Prefers the ``Component_Order`` sheet in an existing template;
        falls back to *csv_path* if provided.
        """
        if existing_template and existing_template.exists():
            try:
                wb = openpyxl.load_workbook(existing_template, data_only=True)
                if "Component_Order" in wb.sheetnames:
                    ws = wb["Component_Order"]
                    order: dict[str, int] = {}
                    for r in range(2, ws.max_row + 1):
                        label = ws.cell(row=r, column=1).value
                        ano = ws.cell(row=r, column=2).value
                        if label and ano is not None:
                            order[str(label)] = int(ano)
                    wb.close()
                    if order:
                        return order
                wb.close()
            except Exception:
                pass
        if csv_path and csv_path.exists():
            order: dict[str, int] = {}
            with open(csv_path, newline="", encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    order[row["label"]] = int(row["assembly_no"])
            return order
        return {}

    def _write_component_order(self, wb, steps: list[dict]) -> None:
        """Write ``Component_Order`` from the actual assembly steps."""
        ws = wb.create_sheet("Component_Order")
        for ci, h in enumerate(["label", "assembly_no"], 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = HF; c.fill = HFILL; c.border = BD
        for ri, s in enumerate(steps, 2):
            ws.cell(row=ri, column=1, value=s.get("label", "")).border = BD
            ws.cell(row=ri, column=2, value=s.get("assembly_no", "")).border = BD
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 14

    def build_template(
        self,
        fcstd_path: Path,
        v1_dir: Path | None = None,
        v0_dir: Path | None = None,
        csv_path: Path | None = None,
        akz_raw: str = "",
        output_path: str | Path | None = None,
    ) -> Path:
        """Build Assembly_3D_template_filled.xlsx. Returns output path."""
        v1_dir = v1_dir or fcstd_path.parent
        v0_dir = v0_dir or v1_dir
        csv_path = csv_path or (v1_dir / "final_result_components.csv")

        akz_map, vv_map = self._parse_akz(akz_raw or DEFAULT_AKZ_MAPPING)
        comps = self._parse_fcstd(fcstd_path)
        ifc_idx = self._build_ifc_index(v1_dir, v0_dir)
        xlink_map = self._build_xlink_map(comps, v1_dir)

        # Compute assembly order
        output_path = Path(output_path) if output_path else (self.piping_dir / "Assembly_3D_template_filled.xlsx")
        csv_order = self._read_component_order(csv_path=csv_path,
                                               existing_template=output_path)

        comps_with = [c for c in comps if c["label"] in csv_order]
        comps_without = [c for c in comps if c["label"] not in csv_order]
        comps_with.sort(key=lambda c: csv_order[c["label"]])
        sorted_comps = comps_with + comps_without
        label_to_ano = {c["label"]: i + 1 for i, c in enumerate(sorted_comps)}

        # Build steps
        steps = []
        for c in sorted_comps:
            label = c["label"]
            ctype = self._classify(label)
            euler = self._q2e(c["qw"], c["qx"], c["qy"], c["qz"])
            fwd = self._qrot(c["qw"], c["qx"], c["qy"], c["qz"], 0, 0, 1)

            ifc_file, ifc_part, ifc_gid = "", "", ""
            for try_name in (c["name"], c["internal"]):
                if try_name and try_name in ifc_idx:
                    ifc_file, ifc_gid = ifc_idx[try_name]
                    ifc_part = try_name
                    break
            if not ifc_file:
                key = f"{c['src']}::{c['name']}"
                resolved = xlink_map.get(key, "")
                if resolved and resolved in ifc_idx:
                    ifc_file, ifc_gid = ifc_idx[resolved]
                    ifc_part = resolved

            akz = akz_map.get(label, "")
            vv = vv_map.get(label, "")
            notes = ""
            if akz in ("TU10.Y6", "VV005"):
                notes = "uc1_expected_missing"
            elif akz in ("TU10.U41", "TU20.U42"):
                notes = "software_control_no_ifc_required"

            steps.append({
                "assembly_no": label_to_ano.get(label, ""),
                "label": label, "type": ctype,
                "akz_tag": akz, "vv_tag": vv,
                "fcstd_link_name": c["name"],
                "source_fcstd_file": c["src"],
                "ifc_parts_library": ifc_file, "ifc_part_name": ifc_part, "ifc_global_id": ifc_gid,
                "pos_x_mm": round(c["px"], 4), "pos_y_mm": round(c["py"], 4), "pos_z_mm": round(c["pz"], 4),
                "rot_qw": round(c["qw"], 9), "rot_qx": round(c["qx"], 9),
                "rot_qy": round(c["qy"], 9), "rot_qz": round(c["qz"], 9),
                "rot_axis_x": round(c["ax"], 6), "rot_axis_y": round(c["ay"], 6), "rot_axis_z": round(c["az"], 6),
                "rot_angle_deg": round(math.degrees(c["angle"]), 4),
                "euler_roll_deg": round(euler[0], 4), "euler_pitch_deg": round(euler[1], 4), "euler_yaw_deg": round(euler[2], 4),
                "forward_x": round(fwd[0], 4), "forward_y": round(fwd[1], 4), "forward_z": round(fwd[2], 4),
                "mapping_confidence": "matched",
                "ifc_coverage": "YES" if ifc_file else "NO",
                "notes": notes,
            })

        # Build auxiliary sheets
        part_lib = self._build_part_library(steps)
        connections = self._build_connections(steps)
        angles = self._build_angles(connections, steps)
        positions = self._build_positions(steps)
        tree = self._build_tree(steps)

        # Write
        self.piping_dir.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._write_sheet(wb, "Part_Library", SHEET_PART_LIBRARY, list(part_lib.values()))
        self._write_sheet(wb, "Assembly_Steps", SHEET_ASSEMBLY_STEPS, steps)
        self._write_sheet(wb, "Connection_Topology", SHEET_CONN_TOPOLOGY, connections)
        self._write_sheet(wb, "Connection_Angles", SHEET_CONN_ANGLES, angles)
        self._write_sheet(wb, "Position_Data", SHEET_POSITION, positions)
        self._write_sheet(wb, "Assembly_Tree", SHEET_TREE, tree)
        # Bonus sheets
        akz_summary = [s for s in steps if s["akz_tag"] or s["vv_tag"]]
        self._write_sheet(wb, "AKZ_Summary",
            ["label", "type", "akz_tag", "vv_tag", "pos_x_mm", "pos_y_mm", "pos_z_mm",
             "ifc_parts_library", "ifc_part_name", "ifc_global_id"],
            [{k: s.get(k, "") for k in ["label", "type", "akz_tag", "vv_tag",
                "pos_x_mm", "pos_y_mm", "pos_z_mm",
                "ifc_parts_library", "ifc_part_name", "ifc_global_id"]} for s in akz_summary])
        ifc_gaps = [s for s in steps if s["ifc_coverage"] == "NO"]
        self._write_sheet(wb, "IFC_Gaps",
            ["label", "type", "fcstd_link_name", "source_fcstd_file",
             "pos_x_mm", "pos_y_mm", "pos_z_mm"],
            [{k: s.get(k, "") for k in ["label", "type", "fcstd_link_name",
                "source_fcstd_file", "pos_x_mm", "pos_y_mm", "pos_z_mm"]} for s in ifc_gaps])

        self._write_component_order(wb, steps)
        wb.save(str(output_path))
        return output_path

    # ── Helpers ───────────────────────────────────────────────────────────

    def _parse_akz(self, raw: str) -> tuple[dict[str, str], dict[str, str]]:
        akz_m, vv_m = {}, {}
        for line in raw.strip().split("\n"):
            parts = [p.strip() for p in line.split(",") if p.strip()]
            if len(parts) < 2:
                continue
            label = parts[0]
            akz = vv = ""
            for p in parts[1:]:
                if p.startswith(("VV", "PL", "HE")):
                    vv = p
                else:
                    akz = p
            if akz:
                akz_m[label] = akz
            if vv:
                vv_m[label] = vv
        return akz_m, vv_m

    def _parse_fcstd(self, path: Path) -> list[dict]:
        with zipfile.ZipFile(path, "r") as z:
            with z.open("Document.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
        ntt = {}
        for obj in root.find("Objects").findall("Object"):
            ntt[obj.get("name", "")] = obj.get("type", "")
        comps = []
        for obj in root.find("ObjectData").findall("Object"):
            if ntt.get(obj.get("name", "")) != "App::Link":
                continue
            props = obj.find("Properties")
            if props is None:
                continue
            c = {"name": obj.get("name", ""), "label": "", "src": "", "internal": "",
                 "px": 0., "py": 0., "pz": 0.,
                 "qw": 0., "qx": 0., "qy": 0., "qz": 1.,
                 "ax": 0., "ay": 0., "az": 1., "angle": 0.}
            for prop in props.findall("Property"):
                pn = prop.get("name", "")
                if pn == "Label":
                    se = prop.find("String")
                    if se is not None:
                        c["label"] = se.get("value", "")
                elif pn == "LinkedObject":
                    xl = prop.find("XLink")
                    if xl is not None:
                        c["src"] = xl.get("file", "")
                        c["internal"] = xl.get("name", "")
                elif pn == "LinkPlacement":
                    pp = prop.find("PropertyPlacement")
                    if pp is not None:
                        c["px"] = float(pp.get("Px", "0"))
                        c["py"] = float(pp.get("Py", "0"))
                        c["pz"] = float(pp.get("Pz", "0"))
                        c["qw"] = float(pp.get("Q0", "0"))
                        c["qx"] = float(pp.get("Q1", "0"))
                        c["qy"] = float(pp.get("Q2", "0"))
                        c["qz"] = float(pp.get("Q3", "1"))
                        c["ax"] = float(pp.get("Ox", "0"))
                        c["ay"] = float(pp.get("Oy", "0"))
                        c["az"] = float(pp.get("Oz", "1"))
                        c["angle"] = float(pp.get("A", "0"))
            if c["label"]:
                comps.append(c)
        return comps

    def _build_ifc_index(self, v1_dir: Path, v0_dir: Path) -> dict[str, tuple[str, str]]:
        idx = {}
        for d in [v1_dir, v0_dir]:
            for ip in sorted(d.glob("*.ifc")):
                if any(x in ip.name for x in ["FLAT_EXPORT", "from_step", "geometry"]):
                    continue
                try:
                    f = ifcopenshell.open(str(ip))
                    for e in f.by_type("IfcBuildingElementProxy"):
                        if e.Name and e.Name not in idx:
                            idx[e.Name] = (ip.name, e.GlobalId)
                except Exception:
                    pass
        return idx

    @staticmethod
    def _discover_component_types(ifc_file) -> tuple[list[str], set[str]]:
        """Auto-discover component entity types and metadata names from an IFC.

        Scans all entity types present in the file and identifies which ones
        carry ``ObjectPlacement`` (components) vs metadata (axes, planes, etc.).

        Returns ``(component_types, skip_names)``.
        """
        from collections import Counter
        type_placement: Counter = Counter()
        type_total: Counter = Counter()
        skip_names: set[str] = set()

        # Metadata type patterns (IFC schema infrastructure)
        meta_types = {
            "IfcProject", "IfcSite", "IfcBuilding", "IfcBuildingStorey",
            "IfcOwnerHistory", "IfcPerson", "IfcOrganization",
            "IfcApplication", "IfcUnitAssignment", "IfcSIUnit",
            "IfcDirection", "IfcCartesianPoint", "IfcAxis2Placement3D",
            "IfcLocalPlacement", "IfcShapeRepresentation",
            "IfcProductDefinitionShape", "IfcTriangulatedFaceSet",
            "IfcCartesianPointList3D", "IfcGeometricRepresentationContext",
            "IfcGeometricRepresentationSubContext", "IfcRelAggregates",
            "IfcRelContainedInSpatialStructure", "IfcDimensionalExponents",
            "IfcMeasureWithUnit", "IfcConversionBasedUnit",
            "IfcPersonAndOrganization",
        }

        for entity in ifc_file:
            etype = entity.is_a()
            if etype in meta_types:
                continue
            type_total[etype] += 1
            try:
                name = entity.Name or ""
            except AttributeError:
                continue
            try:
                has_placement = entity.ObjectPlacement is not None
            except AttributeError:
                has_placement = False
            if has_placement:
                type_placement[etype] += 1
            else:
                if name:
                    skip_names.add(name)

        # Component types = those where >50% of entities have placements
        component_types = [
            t for t in type_total
            if type_placement.get(t, 0) >= type_total[t] * 0.5
        ]
        return component_types, skip_names

    def _parse_assembled_ifc(self, ifc_path: Path) -> list[dict]:
        """Parse a single assembled IFC file for component placements.

        Auto-discovers entity types from the IFC — no hardcoded type list.
        Deduplicates by GlobalId.
        """
        f = ifcopenshell.open(str(ifc_path))
        comps: list[dict] = []
        seen_gids: set[str] = set()

        component_types, skip_names = self._discover_component_types(f)
        # Always include these common IFC component types as a safe baseline
        for extra in ("IfcBuildingElementProxy", "IfcPipeSegment",
                      "IfcPipeFitting", "IfcFlowSegment", "IfcFlowFitting"):
            if extra not in component_types:
                try:
                    if f.by_type(extra):
                        component_types.append(extra)
                except Exception:
                    pass

        for etype in component_types:
            for entity in f.by_type(etype):
                if entity.GlobalId in seen_gids:
                    continue
                name = entity.Name or ""
                if not name or name in skip_names:
                    continue

                placement = entity.ObjectPlacement
                if placement is None:
                    continue

                seen_gids.add(entity.GlobalId)
                px, py, pz, qw, qx, qy, qz, ax, ay, az, angle = \
                    self._decompose_ifc_placement(placement, f)

                comps.append({
                    "name": name, "label": name, "src": ifc_path.name,
                    "internal": name,
                    "px": px, "py": py, "pz": pz,
                    "qw": qw, "qx": qx, "qy": qy, "qz": qz,
                    "ax": ax, "ay": ay, "az": az, "angle": angle,
                    "ifc_part_name": name,
                })

        return comps

    def _build_ifc_index_from_single(self, ifc_path: Path) -> dict[str, tuple[str, str]]:
        """Build part_name → (ifc_filename, GlobalId) index from a single IFC."""
        idx: dict[str, tuple[str, str]] = {}
        f = ifcopenshell.open(str(ifc_path))
        component_types, skip_names = self._discover_component_types(f)
        for extra in ("IfcBuildingElementProxy", "IfcPipeSegment",
                      "IfcPipeFitting", "IfcFlowSegment", "IfcFlowFitting"):
            if extra not in component_types:
                try:
                    if f.by_type(extra):
                        component_types.append(extra)
                except Exception:
                    pass
        for etype in component_types:
            for entity in f.by_type(etype):
                if entity.Name and entity.Name not in idx:
                    idx[entity.Name] = (ifc_path.name, entity.GlobalId)
        return idx

    @staticmethod
    def _build_fallback_ifc_index(ifc_path: Path) -> dict[str, tuple[str, str]]:
        """Build entity_name → (ifc_filename, GlobalId) index from fallback IFC.

        Auto-discovers entity types — no hardcoded type list.
        """
        idx: dict[str, tuple[str, str]] = {}
        f = ifcopenshell.open(str(ifc_path))
        from collections import Counter
        type_total: Counter = Counter()
        placement_counts: Counter = Counter()
        meta_types = {"IfcProject","IfcSite","IfcBuilding","IfcBuildingStorey",
                      "IfcOwnerHistory","IfcPerson","IfcOrganization","IfcApplication",
                      "IfcUnitAssignment","IfcSIUnit","IfcDirection","IfcCartesianPoint",
                      "IfcAxis2Placement3D","IfcLocalPlacement","IfcShapeRepresentation",
                      "IfcProductDefinitionShape","IfcTriangulatedFaceSet",
                      "IfcCartesianPointList3D","IfcGeometricRepresentationContext",
                      "IfcGeometricRepresentationSubContext","IfcRelAggregates",
                      "IfcRelContainedInSpatialStructure","IfcDimensionalExponents",
                      "IfcMeasureWithUnit","IfcConversionBasedUnit","IfcPersonAndOrganization"}
        for entity in f:
            et = entity.is_a()
            if et in meta_types: continue
            type_total[et] += 1
            try:
                has_pl = entity.ObjectPlacement is not None
            except AttributeError:
                has_pl = False
            if has_pl:
                placement_counts[et] += 1
        component_types = [t for t in type_total
                          if placement_counts.get(t,0) >= type_total[t]*0.5]
        # Fallback: include common types
        for extra in ("IfcBuildingElementProxy","IfcPipeSegment","IfcPipeFitting"):
            if extra not in component_types and extra in type_total:
                component_types.append(extra)
        for etype in component_types:
            for entity in f.by_type(etype):
                try:
                    name = entity.Name or ""
                except AttributeError:
                    continue
                if name and name not in idx:
                    idx[name] = (ifc_path.name, entity.GlobalId)
        return idx

    @staticmethod
    def _match_fallback_label(
        label: str, fallback_idx: dict[str, tuple[str, str]]
    ) -> str | None:
        """Find a fallback entity name matching *label*.

        Uses fuzzy prefix matching: ``tank_003`` matches
        ``tank_003_locked``, ``tank_003_flat``, etc.  No hardcoded suffix.
        """
        if not label:
            return None
        # Exact match
        if label in fallback_idx:
            return label
        # Prefix match: label + some suffix pattern
        for key in fallback_idx:
            if key.startswith(label) and len(key) > len(label):
                return key
        return None

    @staticmethod
    def _decompose_ifc_placement(
        placement, f
    ) -> tuple[float, float, float, float, float, float, float, float, float, float, float]:
        """Decompose an ``IfcLocalPlacement`` into position + rotation.

        Returns (px, py, pz, qw, qx, qy, qz, axis_x, axis_y, axis_z, angle_rad).
        Traverses PlacementRelTo for world-space coordinates.
        """
        px = py = pz = 0.0
        # Accumulate matrices for parent chain
        matrices: list[tuple[list[float], ...]] = []

        current = placement
        while current is not None:
            rel = current.RelativePlacement
            if rel is None:
                break
            loc = rel.Location
            lx, ly, lz = (loc.Coordinates[0], loc.Coordinates[1], loc.Coordinates[2]) if loc else (0.0, 0.0, 0.0)
            z_dir = list(rel.Axis.DirectionRatios) if rel.Axis else [0.0, 0.0, 1.0]
            x_dir = list(rel.RefDirection.DirectionRatios) if rel.RefDirection else [1.0, 0.0, 0.0]

            # Normalize
            z_mag = math.sqrt(z_dir[0]**2 + z_dir[1]**2 + z_dir[2]**2)
            if z_mag > 1e-15:
                z_dir = [v / z_mag for v in z_dir]
            x_mag = math.sqrt(x_dir[0]**2 + x_dir[1]**2 + x_dir[2]**2)
            if x_mag > 1e-15:
                x_dir = [v / x_mag for v in x_dir]

            # Y = Z × X
            y_dir = [
                z_dir[1] * x_dir[2] - z_dir[2] * x_dir[1],
                z_dir[2] * x_dir[0] - z_dir[0] * x_dir[2],
                z_dir[0] * x_dir[1] - z_dir[1] * x_dir[0],
            ]
            # Re-orthogonalize X: X' = Y × Z
            x_dir = [
                y_dir[1] * z_dir[2] - y_dir[2] * z_dir[1],
                y_dir[2] * z_dir[0] - y_dir[0] * z_dir[2],
                y_dir[0] * z_dir[1] - y_dir[1] * z_dir[0],
            ]

            matrices.insert(0, (x_dir, y_dir, z_dir, [lx, ly, lz]))
            current = current.PlacementRelTo

        # Compose matrices through the parent chain
        world_x = [1.0, 0.0, 0.0]
        world_y = [0.0, 1.0, 0.0]
        world_z = [0.0, 0.0, 1.0]
        world_pos = [0.0, 0.0, 0.0]

        for xd, yd, zd, trans in matrices:
            # Rotate accumulated position
            rx = (world_x[0] * trans[0] + world_x[1] * trans[1] + world_x[2] * trans[2])
            ry = (world_y[0] * trans[0] + world_y[1] * trans[1] + world_y[2] * trans[2])
            rz = (world_z[0] * trans[0] + world_z[1] * trans[1] + world_z[2] * trans[2])
            world_pos[0] += rx
            world_pos[1] += ry
            world_pos[2] += rz

            # Compose rotation
            nx = [xd[0]*world_x[0] + xd[1]*world_y[0] + xd[2]*world_z[0],
                  xd[0]*world_x[1] + xd[1]*world_y[1] + xd[2]*world_z[1],
                  xd[0]*world_x[2] + xd[1]*world_y[2] + xd[2]*world_z[2]]
            ny = [yd[0]*world_x[0] + yd[1]*world_y[0] + yd[2]*world_z[0],
                  yd[0]*world_x[1] + yd[1]*world_y[1] + yd[2]*world_z[1],
                  yd[0]*world_x[2] + yd[1]*world_y[2] + yd[2]*world_z[2]]
            nz = [zd[0]*world_x[0] + zd[1]*world_y[0] + zd[2]*world_z[0],
                  zd[0]*world_x[1] + zd[1]*world_y[1] + zd[2]*world_z[1],
                  zd[0]*world_x[2] + zd[1]*world_y[2] + zd[2]*world_z[2]]
            world_x, world_y, world_z = nx, ny, nz

        px, py, pz = world_pos[0], world_pos[1], world_pos[2]

        # Rotation matrix → quaternion
        # Matrix columns: [X] [Y] [Z]
        trace = world_x[0] + world_y[1] + world_z[2]
        if trace > 0:
            s = math.sqrt(trace + 1.0) * 2
            qw = 0.25 * s
            qx = (world_y[2] - world_z[1]) / s
            qy = (world_z[0] - world_x[2]) / s
            qz = (world_x[1] - world_y[0]) / s
        elif world_x[0] > world_y[1] and world_x[0] > world_z[2]:
            s = math.sqrt(1.0 + world_x[0] - world_y[1] - world_z[2]) * 2
            qw = (world_y[2] - world_z[1]) / s
            qx = 0.25 * s
            qy = (world_y[0] + world_x[1]) / s
            qz = (world_z[0] + world_x[2]) / s
        elif world_y[1] > world_z[2]:
            s = math.sqrt(1.0 + world_y[1] - world_x[0] - world_z[2]) * 2
            qw = (world_z[0] - world_x[2]) / s
            qx = (world_y[0] + world_x[1]) / s
            qy = 0.25 * s
            qz = (world_z[1] + world_y[2]) / s
        else:
            s = math.sqrt(1.0 + world_z[2] - world_x[0] - world_y[1]) * 2
            qw = (world_x[1] - world_y[0]) / s
            qx = (world_z[0] + world_x[2]) / s
            qy = (world_z[1] + world_y[2]) / s
            qz = 0.25 * s

        # Normalize quaternion
        qn = math.sqrt(qw*qw + qx*qx + qy*qy + qz*qz)
        if qn > 1e-15:
            qw, qx, qy, qz = qw/qn, qx/qn, qy/qn, qz/qn

        # Quaternion → axis-angle
        if abs(qw) > 0.999999:
            ax, ay, az, angle = 0.0, 0.0, 1.0, 0.0
        else:
            half_angle = math.acos(max(-1.0, min(1.0, qw)))
            sin_half = math.sin(half_angle)
            if abs(sin_half) > 1e-15:
                ax, ay, az = qx/sin_half, qy/sin_half, qz/sin_half
                angle = 2.0 * half_angle
            else:
                ax, ay, az, angle = 0.0, 0.0, 1.0, 0.0

        return px, py, pz, qw, qx, qy, qz, ax, ay, az, angle

    def _build_xlink_map(self, comps: list[dict], v1_dir: Path) -> dict[str, str]:
        xm = {}
        for sf in set(c["src"] for c in comps):
            fpath = v1_dir / sf
            if not fpath.exists() or sf == "Unnamed9.FCStd":
                continue
            try:
                with zipfile.ZipFile(fpath, "r") as z:
                    with z.open("Document.xml") as f:
                        t = ET.parse(f)
                        r = t.getroot()
                ntt = {}
                for obj in r.find("Objects").findall("Object"):
                    ntt[obj.get("name", "")] = obj.get("type", "")
                for obj in r.find("ObjectData").findall("Object"):
                    if ntt.get(obj.get("name", "")) != "App::Link":
                        continue
                    pr = obj.find("Properties")
                    if pr is None:
                        continue
                    for prop in pr.findall("Property"):
                        if prop.get("name") == "LinkedObject":
                            xl = prop.find("XLink")
                            if xl is not None:
                                xm[f"{sf}::{obj.get('name','')}"] = xl.get("name", "")
            except Exception:
                pass
        return xm

    @staticmethod
    def _classify(label: str) -> str:
        if label.startswith("tank"): return "tank"
        if label.startswith("blackbox"): return "blackbox"
        if label.startswith("elbow"): return "elbow"
        if label.startswith("pipeline"): return "pipeline"
        return "other"

    @staticmethod
    def _q2e(qw, qx, qy, qz):
        n = math.sqrt(qw*qw+qx*qx+qy*qy+qz*qz)
        if n < 1e-15: return (0, 0, 0)
        qw, qx, qy, qz = qw/n, qx/n, qy/n, qz/n
        r20 = 2*(qx*qz-qw*qy); r21 = 2*(qy*qz+qw*qx); r22 = 1-2*(qx*qx+qy*qy)
        r10 = 2*(qx*qy+qw*qz); r00 = 1-2*(qy*qy+qz*qz)
        p = -math.asin(max(-1, min(1, r20))); cp = math.cos(p)
        if abs(cp) > 1e-10:
            return (math.degrees(math.atan2(r21/cp, r22/cp)), math.degrees(p), math.degrees(math.atan2(r10/cp, r00/cp)))
        return (0, math.degrees(p), math.degrees(math.atan2(-2*(qy*qz-qw*qx), 1-2*(qx*qx+qz*qz))))

    @staticmethod
    def _qrot(qw, qx, qy, qz, vx, vy, vz):
        tw = -qx*vx-qy*vy-qz*vz; tx = qw*vx+qy*vz-qz*vy
        ty = qw*vy+qz*vx-qx*vz; tz = qw*vz+qx*vy-qy*vx
        return (tx*qw+tw*(-qx)+ty*(-qz)-tz*(-qy),
                ty*qw+tw*(-qy)+tz*(-qx)-tx*(-qz),
                tz*qw+tw*(-qz)+tx*(-qy)-ty*(-qx))

    def _build_part_library(self, steps: list[dict]) -> dict[str, dict]:
        parts = {}
        for s in steps:
            fn = s["fcstd_link_name"]
            if fn not in parts:
                parts[fn] = {
                    "part_id": f"PART_{len(parts)+1:04d}", "part_ref_name": fn,
                    "part_type": s["type"], "fcstd_link_name": fn,
                    "source_fcstd_file": s["source_fcstd_file"],
                    "ifc_parts_library": s["ifc_parts_library"], "ifc_part_name": s["ifc_part_name"],
                    "instance_count": 0, "ifc_coverage": s["ifc_coverage"], "description": "",
                }
            parts[fn]["instance_count"] += 1
        return parts

    def _build_connections(self, steps: list[dict]) -> list[dict]:
        valid = [s for s in steps if isinstance(s["assembly_no"], int)]
        conns = []
        for i in range(len(valid) - 1):
            frm, to = valid[i], valid[i + 1]
            dist = round(math.sqrt(
                (frm["pos_x_mm"]-to["pos_x_mm"])**2 +
                (frm["pos_y_mm"]-to["pos_y_mm"])**2 +
                (frm["pos_z_mm"]-to["pos_z_mm"])**2), 2)
            ctype = "flange" if "tank" in (frm["type"], to["type"]) else \
                "weld_or_threaded" if "elbow" in (frm["type"], to["type"]) and "pipeline" in (frm["type"], to["type"]) else \
                "flange_or_threaded" if "blackbox" in (frm["type"], to["type"]) else \
                "direct_fitting" if frm["type"] == "elbow" and to["type"] == "elbow" else \
                "weld_or_coupling" if frm["type"] == "pipeline" and to["type"] == "pipeline" else "unknown"
            conns.append({
                "connection_id": i + 1,
                "from_assembly_no": frm["assembly_no"], "from_label": frm["label"], "from_port": "out",
                "to_assembly_no": to["assembly_no"], "to_label": to["label"], "to_port": "in",
                "connection_type": ctype,
                "from_pos_x": frm["pos_x_mm"], "from_pos_y": frm["pos_y_mm"], "from_pos_z": frm["pos_z_mm"],
                "to_pos_x": to["pos_x_mm"], "to_pos_y": to["pos_y_mm"], "to_pos_z": to["pos_z_mm"],
                "euclidean_distance_mm": dist,
            })
        return conns

    def _build_angles(self, connections: list[dict], steps: list[dict]) -> list[dict]:
        by_label = {s["label"]: s for s in steps}
        angles = []
        for conn in connections:
            frm = by_label.get(conn["from_label"])
            to = by_label.get(conn["to_label"])
            if not frm or not to:
                continue
            fwd_f = (frm.get("forward_x", 0), frm.get("forward_y", 0), frm.get("forward_z", 0))
            fwd_t = (to.get("forward_x", 0), to.get("forward_y", 0), to.get("forward_z", 0))
            cv = (0., 0., 0.)
            if (d := conn["euclidean_distance_mm"]) and d > 0:
                cv = ((to["pos_x_mm"]-frm["pos_x_mm"])/d, (to["pos_y_mm"]-frm["pos_y_mm"])/d, (to["pos_z_mm"]-frm["pos_z_mm"])/d)
            dot = max(-1, min(1, fwd_f[0]*fwd_t[0]+fwd_f[1]*fwd_t[1]+fwd_f[2]*fwd_t[2]))
            bend = math.degrees(math.acos(dot))
            dot_cv = max(-1, min(1, abs(fwd_f[0]*cv[0]+fwd_f[1]*cv[1]+fwd_f[2]*cv[2])))
            torsion = math.degrees(math.acos(dot_cv))
            angles.append({
                "connection_id": conn["connection_id"], "from_label": conn["from_label"], "to_label": conn["to_label"],
                "from_forward_x": round(fwd_f[0], 6), "from_forward_y": round(fwd_f[1], 6), "from_forward_z": round(fwd_f[2], 6),
                "to_forward_x": round(fwd_t[0], 6), "to_forward_y": round(fwd_t[1], 6), "to_forward_z": round(fwd_t[2], 6),
                "connection_vector_x": round(cv[0], 6), "connection_vector_y": round(cv[1], 6), "connection_vector_z": round(cv[2], 6),
                "bend_angle_deg": round(bend, 2), "torsion_angle_deg": round(torsion, 2),
                "from_to_distance_mm": round(conn["euclidean_distance_mm"] or 0, 2),
            })
        return angles

    def _build_positions(self, steps: list[dict]) -> list[dict]:
        return [{
            "fcstd_link_name": s["fcstd_link_name"], "label": s["label"], "type": s["type"],
            "source_fcstd_file": s["source_fcstd_file"], "ifc_parts_library": s["ifc_parts_library"],
            "pos_x_mm": s["pos_x_mm"], "pos_y_mm": s["pos_y_mm"], "pos_z_mm": s["pos_z_mm"],
            "qw": s["rot_qw"], "qx": s["rot_qx"], "qy": s["rot_qy"], "qz": s["rot_qz"],
            "axis_x": s["rot_axis_x"], "axis_y": s["rot_axis_y"], "axis_z": s["rot_axis_z"],
            "angle_deg": s["rot_angle_deg"],
            "euler_roll_deg": s["euler_roll_deg"], "euler_pitch_deg": s["euler_pitch_deg"], "euler_yaw_deg": s["euler_yaw_deg"],
        } for s in steps if s["ifc_coverage"] == "YES"]

    def _build_tree(self, steps: list[dict]) -> list[dict]:
        tree = []
        tid, current_tank = 0, None
        for s in steps:
            if s["type"] == "tank":
                tid += 1; current_tank = s["label"]
                tree.append({"tree_id": tid, "parent_label": None, "child_label": s["label"],
                             "aggregation_type": "root_assembly", "assembly_level": "assembly",
                             "notes": "tank root"})
            elif current_tank and s["type"] in ("pipeline", "elbow", "blackbox"):
                tid += 1
                tree.append({"tree_id": tid, "parent_label": current_tank, "child_label": s["label"],
                             "aggregation_type": "connected_to", "assembly_level": "part",
                             "notes": f"part of {current_tank} run"})
        return tree

    def _write_sheet(self, wb, name, headers, rows):
        ws = wb.create_sheet(name)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h); c.font = HF; c.fill = HFILL; c.border = BD
        for ri, row in enumerate(rows, 2):
            for ci, h in enumerate(headers, 1):
                v = row.get(h, "")
                cell = ws.cell(row=ri, column=ci, value=v); cell.border = BD
                if h == "ifc_coverage" and v == "NO": cell.fill = RFILL
                if h == "akz_tag" and v: cell.fill = GFILL
        for ci in range(1, len(headers)+1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(ci)].width = min(22, max(len(str(headers[ci-1])), 8))
        return ws
