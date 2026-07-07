from __future__ import annotations

import json
import platform
import re
import shutil
import tempfile
from hashlib import sha1
from collections import Counter, defaultdict
from pathlib import Path
from typing import Callable

import pandas as pd
from pydantic import ValidationError

from iev4pi_transformation_tool.core.chunk_builder import ChunkBuilder
from iev4pi_transformation_tool.core.database import Database
from iev4pi_transformation_tool.core.dexpi import DexpiPackageAnalyzer
from iev4pi_transformation_tool.core.document_classifier import DocumentClassifier
from iev4pi_transformation_tool.core.document_reader import DocumentReader
from iev4pi_transformation_tool.core.evidence_resolver import EvidenceResolver
from iev4pi_transformation_tool.core.ocr_defaults import OCRPlatformDefaults, get_ocr_platform_defaults
from iev4pi_transformation_tool.core.export_service import ExportService
from iev4pi_transformation_tool.core.extractor import Extractor
from iev4pi_transformation_tool.core.llm_client import OpenAICompatibleLLMClient
from iev4pi_transformation_tool.core.retriever import Retriever
from iev4pi_transformation_tool.core.schema_miner import SchemaMiner
from iev4pi_transformation_tool.core.source_preview import SourcePreviewLoader
from iev4pi_transformation_tool.core.surya_prewarm import prewarm_surya_models, surya_prewarm_status
from iev4pi_transformation_tool.core.uc1_catalog import UC1_SUPPORT_MAP, UC1CatalogService
from iev4pi_transformation_tool.core.semantic_ids import get_irdi
from iev4pi_transformation_tool.core.utils import clean_cell, ensure_dir, extract_component_tokens, normalize_identifier, normalize_label
from iev4pi_transformation_tool.services.debug_log import DebugLogStore
from iev4pi_transformation_tool.services.aas_generation import AASGenerationService
from iev4pi_transformation_tool.services.ontology_export import OntologyExportService
from iev4pi_transformation_tool.t1t5 import T1T5Executor, T1T5RuleBundle, T1T5RuleStore, build_default_t1_t5_bundle, stage_source_type
from iev4pi_transformation_tool.tx import TxExecutor, TxRuleSet, TxRuleStore, TxRuleSuggester, TxValidationIssue, WeightedEntityResolver, build_default_uc1_rule_set
from iev4pi_transformation_tool.models import (
    AASGenerationRequest,
    ConsistencyDecision,
    DocumentDescriptor,
    DocumentFamily,
    EvidenceRef,
    EvidenceBundle,
    ExcelCellProvenance,
    ExcelCellTooltipContext,
    ExcelSheetPreview,
    ExcelWorkbookPreview,
    ExtractionStatus,
    ExtractedFieldResult,
    ExtractedRecord,
    ParsedDocument,
    PidInconsistencyRow,
    PidInconsistencySummary,
    PidJumpTarget,
    ProjectSettings,
    RiBundle,
    RunSummary,
    ScanSnapshot,
    SchemaFamily,
    SourceDocumentKind,
    UC1CatalogCoverageReport,
)

ProgressCallback = Callable[[int, str], None]
EXCEL_TOOLTIP_CONTEXT_VERSION = 9


class Workbench:
    def __init__(
        self,
        workspace_root: Path,
        external_log_sink: Callable[[dict[str, object]], None] | None = None,
    ) -> None:
        self.workspace_root = workspace_root.resolve()
        self._external_log_sink = external_log_sink
        self.state_dir = ensure_dir(self.workspace_root / ".iev4pi")
        self.cache_root = ensure_dir(self.state_dir / "cache")
        self.cache_dir = ensure_dir(self.cache_root / "ocr")
        self.embedding_cache_dir = ensure_dir(self.cache_root / "embeddings")
        self.llm_verifier_cache_dir = ensure_dir(self.cache_root / "llm_verifier")
        self.llm_normalizer_cache_dir = ensure_dir(self.cache_root / "llm_normalizer")
        self.vlm_verifier_cache_dir = ensure_dir(self.cache_root / "vlm_verifier")
        self.settings_path = ensure_dir(self.workspace_root / "Exports") / "settings.json"
        self.settings = self._load_settings()
        self.debug_log = DebugLogStore(
            max_entries=800,
            persist_path=None if external_log_sink is not None else (self.state_dir / "debug_log.jsonl")
        )

        self.classifier = DocumentClassifier(self.workspace_root)
        self.reader = self._build_reader()
        self.chunk_builder = ChunkBuilder()
        self.llm_client = OpenAICompatibleLLMClient(self.settings.llm, logger=self.log_debug)
        self.schema_miner = SchemaMiner()

        # Initialize disk-persistent LLM/VLM cache
        from iev4pi_transformation_tool.core import llm_cache
        llm_cache.init(self.workspace_root / ".iev4pi")

        self.retriever = Retriever(
            llm_client=self.llm_client,
            cache_dir=self.embedding_cache_dir,
            logger=self.log_debug,
        )
        self.evidence_resolver = EvidenceResolver(
            self.retriever,
            self.llm_client,
            cache_dir=self.llm_verifier_cache_dir,
            logger=self.log_debug,
        )
        self.extractor = Extractor(
            self.retriever,
            self.llm_client,
            self.evidence_resolver,
            cache_dir=self.llm_normalizer_cache_dir,
            logger=self.log_debug,
        )
        self.export_service = ExportService()
        self.dexpi_analyzer = DexpiPackageAnalyzer()
        self.database = Database(self.settings.database_path)
        self.aas_generation_service = AASGenerationService(self.workspace_root)
        self._uc1_evidence_cache: dict[tuple[str, str], EvidenceBundle] = {}
        self._last_uc1_generation_warnings: list[str] = []
        self.uc1_catalog_service = UC1CatalogService()
        self.ontology_export_service = OntologyExportService()
        self.t1_t5_executor = T1T5Executor()
        self.t1_t5_rule_store = T1T5RuleStore(self.state_dir)
        self.tx_executor = TxExecutor()
        self.tx_rule_store = TxRuleStore(self.state_dir)
        self.tx_rule_suggester = TxRuleSuggester(self.llm_client)
        self.entity_resolver = WeightedEntityResolver()

        self.documents: list[DocumentDescriptor] = []
        self.ri_bundles: list[RiBundle] = []
        self.parsed_cache: dict[str, ParsedDocument] = {}
        self.schemas: dict[DocumentFamily, SchemaFamily] = {}
        self.ri_bundle_schemas: dict[str, dict[DocumentFamily, SchemaFamily]] = {}
        self._latest_run_summary: RunSummary | None = None
        self.records = []
        self.log_debug(
            source="workbench",
            action="startup",
            message=f"Workspace initialized at {self.workspace_root}",
            details={"workspace_root": str(self.workspace_root)},
        )

    def log_debug(
        self,
        *,
        source: str,
        action: str,
        message: str,
        level: str = "INFO",
        details: dict | None = None,
    ) -> dict:
        entry = self.debug_log.add(
            source=source,
            action=action,
            message=message,
            level=level,
            details=details,
        )
        if self._external_log_sink is not None:
            self._external_log_sink(entry)
        return entry

    def debug_log_entries(self) -> list[dict]:
        return self.debug_log.entries()

    def debug_log_total_count(self) -> int:
        return self.debug_log.total_count()

    def debug_log_entry_at(self, index: int) -> dict | None:
        return self.debug_log.entry_at(index)

    def iter_debug_log_entries(self):
        return self.debug_log.iter_entries()

    def iter_debug_log_recent(self, limit: int = 5000):
        return self.debug_log.iter_recent_entries(limit)

    def clear_debug_log(self) -> None:
        self.debug_log.clear()

    def _load_settings(self) -> ProjectSettings:
        default = ProjectSettings(
            workspace_root=self.workspace_root,
            input_dirs=["Documents", "Documents-Others"],
            scan_root_dir="Documents",
            database_path=self.workspace_root / ".iev4pi" / "state.sqlite",
            export_dir=self.workspace_root / "Exports",
            results_export_dir=self.workspace_root / "Exports",
        )
        default = self._apply_platform_ocr_defaults(default, force=True)
        if not self.settings_path.exists():
            self.settings_path.write_text(default.model_dump_json(indent=2), encoding="utf-8")
            return default
        try:
            settings = ProjectSettings.model_validate_json(self.settings_path.read_text(encoding="utf-8"))
        except (ValidationError, json.JSONDecodeError):
            self.settings_path.write_text(default.model_dump_json(indent=2), encoding="utf-8")
            return default
        return self._normalize_settings(settings)

    def save_settings(self, settings: ProjectSettings) -> None:
        self.settings = self._normalize_settings(settings)
        self.settings_path.write_text(self.settings.model_dump_json(indent=2), encoding="utf-8")
        self.reader = self._build_reader()
        self.llm_client = OpenAICompatibleLLMClient(self.settings.llm, logger=self.log_debug)
        self.retriever = Retriever(
            llm_client=self.llm_client,
            cache_dir=self.embedding_cache_dir,
            logger=self.log_debug,
        )
        self.evidence_resolver = EvidenceResolver(
            self.retriever,
            self.llm_client,
            cache_dir=self.llm_verifier_cache_dir,
            logger=self.log_debug,
        )
        self.extractor = Extractor(
            self.retriever,
            self.llm_client,
            self.evidence_resolver,
            cache_dir=self.llm_normalizer_cache_dir,
            logger=self.log_debug,
        )
        self.tx_rule_suggester = TxRuleSuggester(self.llm_client)

    def _aio_ml_evidence_linking_benchmark_report_path(self) -> Path:
        configured = clean_cell(self.settings.aio_ml_evidence_linking_benchmark_report)
        if configured:
            path = Path(configured).expanduser()
            if not path.is_absolute():
                path = self.workspace_root / path
            return path
        return self.state_dir / "source_artifact_linking_benchmark.json"

    def _aio_ml_evidence_linking_enabled(self) -> bool:
        return bool(
            getattr(self.settings, "aio_ml_evidence_linking", False)
            or getattr(self.settings, "aio_ml_evidence_linking_enabled", False)
        )

    def _normalize_settings(self, settings: ProjectSettings) -> ProjectSettings:
        settings.workspace_root = Path(".")
        if not settings.scan_root_dir:
            settings.scan_root_dir = "Documents"
        # Preserve input_dirs from persistent settings; only fall back to
        # scan_root_dir when the list is genuinely empty.
        if not settings.input_dirs:
            settings.input_dirs = [settings.scan_root_dir]
        settings.database_path = self._normalize_project_relative_path(settings.database_path, Path(".iev4pi/state.sqlite"))
        settings.export_dir = self._normalize_project_relative_path(settings.export_dir, Path("Exports"))
        if settings.results_export_dir is None:
            settings.results_export_dir = Path("Exports")
        settings.results_export_dir = self._normalize_project_relative_path(
            settings.results_export_dir,
            Path("Exports"),
        )
        settings.apple_ocr_framework = "vision"
        settings.apple_ocr_recognition_level = "accurate"
        settings = self._apply_platform_ocr_defaults(settings)
        settings.apple_ocr_framework = "vision"
        settings.apple_ocr_recognition_level = "accurate"
        return settings

    def _apply_platform_ocr_defaults(self, settings: ProjectSettings, *, force: bool = False) -> ProjectSettings:
        system = platform.system()
        platform_defaults = get_ocr_platform_defaults(system)
        if system != "Darwin":
            valid_primary = {"paddle", "rapidocr", "surya", "easyocr"}
            valid_fallback = {"rapidocr", "surya", "easyocr", "none"}
            if force or settings.ocr_backend not in valid_primary:
                settings.ocr_backend = platform_defaults.ocr_backend
            if force or settings.ocr_fallback_backend not in valid_fallback or settings.ocr_fallback_backend == settings.ocr_backend:
                settings.ocr_fallback_backend = (
                    platform_defaults.ocr_fallback_backend
                    if platform_defaults.ocr_fallback_backend != settings.ocr_backend
                    else "rapidocr"
                )
            if force or not str(settings.ocr_device).strip():
                settings.ocr_device = platform_defaults.ocr_device
            return settings
        if force or self._uses_legacy_macos_ocr_defaults(settings):
            self._set_platform_ocr_defaults(settings, platform_defaults)
        return settings

    def _uses_legacy_macos_ocr_defaults(self, settings: ProjectSettings) -> bool:
        return (
            settings.ocr_backend == "paddle"
            and settings.ocr_fallback_backend in {"surya", "rapidocr"}
            and settings.ocr_device in {"cuda:0", "cpu"}
            and settings.apple_ocr_framework == "vision"
            and settings.apple_ocr_recognition_level == "accurate"
        )

    def _set_platform_ocr_defaults(
        self,
        settings: ProjectSettings,
        defaults: OCRPlatformDefaults,
    ) -> None:
        settings.ocr_backend = defaults.ocr_backend
        settings.ocr_fallback_backend = defaults.ocr_fallback_backend
        settings.ocr_device = defaults.ocr_device
        settings.apple_ocr_framework = defaults.apple_ocr_framework
        settings.apple_ocr_recognition_level = defaults.apple_ocr_recognition_level

    def _build_reader(self) -> DocumentReader:
        return self._build_reader_with_overrides()

    def _build_reader_with_overrides(self, *, ocr_enabled: bool | None = None) -> DocumentReader:
        return DocumentReader(
            ocr_enabled=self.settings.ocr_enabled if ocr_enabled is None else bool(ocr_enabled),
            ocr_zoom=self.settings.ocr_zoom,
            ocr_backend=self.settings.ocr_backend,
            ocr_fallback_backend=self.settings.ocr_fallback_backend,
            ocr_device=self.settings.ocr_device,
            apple_ocr_framework=self.settings.apple_ocr_framework,
            apple_ocr_recognition_level=self.settings.apple_ocr_recognition_level,
            ocr_dpi=self.settings.ocr_dpi,
            diagram_dpi=self.settings.diagram_dpi,
            diagram_analysis_mode=self.settings.diagram_analysis_mode,
            diagram_extraction_backend=self.settings.diagram_extraction_backend,
            ocr_min_confidence=self.settings.ocr_min_confidence,
            ocr_pipeline_mode=self.settings.ocr_pipeline_mode,
            ocr_ensemble_backends=self.settings.ocr_ensemble_backends,
            enable_diagram_relation_extraction=self.settings.enable_diagram_relation_extraction,
            enable_hard_page_fallback=self.settings.enable_hard_page_fallback,
            cache_dir=self.cache_dir,
            llm_config=self.settings.llm,
            logger=self.log_debug,
        )

    def _serialize_path(self, path: Path) -> str:
        resolved = path.resolve()
        try:
            return resolved.relative_to(self.workspace_root.resolve()).as_posix()
        except ValueError:
            return str(resolved)

    def _normalize_project_relative_path(self, path: Path, default: Path) -> Path:
        raw = str(path or "").strip()
        if not raw:
            return default
        normalized = raw.replace("\\", "/")
        if re.match(r"^[A-Za-z]:/", normalized):
            return default
        candidate = Path(normalized)
        if candidate.is_absolute():
            try:
                return Path(candidate.resolve().relative_to(self.workspace_root.resolve()).as_posix())
            except ValueError:
                return default
        return candidate

    def display_export_dir(self) -> str:
        return str(self.settings.export_dir or Path("Exports")).replace("\\", "/")

    def display_results_export_dir(self) -> str:
        return str(self.settings.results_export_dir or Path("Exports")).replace("\\", "/")

    def display_cache_dir(self) -> str:
        try:
            return self.cache_root.relative_to(self.workspace_root).as_posix()
        except ValueError:
            return str(self.cache_root)

    def display_ocr_cache_dir(self) -> str:
        try:
            return self.cache_dir.relative_to(self.workspace_root).as_posix()
        except ValueError:
            return str(self.cache_dir)

    def clear_cache(self) -> int:
        removed_entries = 0
        cache_targets = [
            self.cache_dir,
            self.embedding_cache_dir,
            self.llm_verifier_cache_dir,
            self.llm_normalizer_cache_dir,
            self.vlm_verifier_cache_dir,
            self.cache_root / "preview",
        ]
        seen_paths: set[Path] = set()
        for target_dir in cache_targets:
            resolved_target = target_dir.resolve()
            if resolved_target in seen_paths:
                continue
            seen_paths.add(resolved_target)
            if not target_dir.exists():
                ensure_dir(target_dir)
                continue
            for child in target_dir.iterdir():
                if child.is_symlink() or child.is_file():
                    child.unlink()
                else:
                    shutil.rmtree(child)
                removed_entries += 1
            ensure_dir(target_dir)
        self.parsed_cache = {}
        if hasattr(self.reader, "_task_ocr_result_cache"):
            self.reader._task_ocr_result_cache = {}
        return removed_entries

    def clear_ocr_cache(self) -> int:
        return self.clear_cache()

    def resolve_scan_root(self) -> Path:
        path = Path(self.settings.scan_root_dir)
        if not path.is_absolute():
            path = self.workspace_root / path
        return path

    def resolve_input_dirs(self) -> list[Path]:
        """Return resolved, existing input directories from settings."""
        resolved: list[Path] = []
        for d in self.settings.input_dirs:
            path = Path(d)
            if not path.is_absolute():
                path = self.workspace_root / path
            if path.is_dir():
                resolved.append(path)
        if not resolved:
            # Ultimate fallback: scan Documents
            fallback = self.workspace_root / "Documents"
            if fallback.is_dir():
                resolved.append(fallback)
        return resolved

    def resolve_export_dir(self) -> Path:
        path = self.settings.export_dir
        if not path.is_absolute():
            path = self.workspace_root / path
        return path

    def resolve_results_export_dir(self) -> Path:
        path = self.settings.results_export_dir or (self.settings.export_dir / "results")
        if not path.is_absolute():
            path = self.workspace_root / path
        return path

    def update_scan_root(self, path: Path) -> None:
        settings = ProjectSettings.model_validate(self.settings.model_dump())
        settings.scan_root_dir = self._serialize_path(path)
        settings.input_dirs = [settings.scan_root_dir]
        self.save_settings(settings)
        self.documents = []
        self.ri_bundles = []
        self.parsed_cache = {}

    def update_export_dir(self, path: Path) -> None:
        settings = ProjectSettings.model_validate(self.settings.model_dump())
        settings.export_dir = Path(self._serialize_path(path))
        self.save_settings(settings)

    def update_results_export_dir(self, path: Path) -> None:
        settings = ProjectSettings.model_validate(self.settings.model_dump())
        settings.results_export_dir = Path(self._serialize_path(path))
        self.save_settings(settings)

    def resolve_source_path(self, source_path: str) -> Path | None:
        candidate = Path(source_path)
        probes: list[Path] = []
        if candidate.is_absolute():
            probes.append(candidate)
        else:
            probes.append(self.resolve_scan_root() / candidate)
            probes.append(self.workspace_root / candidate)
            for document in self.documents:
                if document.relative_path == source_path:
                    probes.append(document.path)
                    break
            stored = self.database.document_path_for_relative_path(source_path)
            if stored is not None:
                probes.append(stored)
        for probe in probes:
            resolved = probe.resolve()
            if resolved.exists():
                return resolved
        return None

    def load_value_source_preview(
        self,
        source_path: str,
        evidences_payload: list[dict[str, object]] | list[EvidenceRef],
        *,
        evidence_index: int = 0,
        record_display_name: str = "",
        field_name: str = "",
        target_value: str = "",
    ) -> dict[str, object]:
        loader = SourcePreviewLoader(
            workspace_root=self.workspace_root,
            scan_root=self.resolve_scan_root(),
            documents=self.documents,
            classifier=self.classifier,
            reader=self.reader,
            cache_dir=self.state_dir / "cache",
            parsed_cache=self.parsed_cache,
            resolve_source_path=self.resolve_source_path,
        )
        return loader.load(
            source_path=source_path,
            evidences_payload=evidences_payload,
            evidence_index=evidence_index,
            record_display_name=record_display_name,
            field_name=field_name,
            target_value=target_value,
        )

    def filled_excel_workbook_items(self) -> list[dict[str, str]]:
        from iev4pi_transformation_tool.core.standardized_templates import (
            collect_filled_templates,
            get_export_category,
        )

        items: list[dict[str, str]] = []
        for workbook_name, path in sorted(collect_filled_templates().items()):
            items.append(
                {
                    "workbook_name": workbook_name,
                    "path": str(path),
                    "category": get_export_category(workbook_name) or "",
                    "provenance_path": str(self._excel_provenance_path(path)),
                }
            )
        return items

    def load_filled_excel_preview(self, workbook_name: str) -> ExcelWorkbookPreview | None:
        from openpyxl import load_workbook
        from iev4pi_transformation_tool.core.standardized_templates import collect_filled_templates

        filled_templates = collect_filled_templates()
        path = filled_templates.get(workbook_name)
        if path is None or not path.is_file():
            return None
        provenance_payload = self._load_or_build_excel_provenance(path, workbook_name)
        sheets_payload = provenance_payload.get("sheets", {}) if isinstance(provenance_payload, dict) else {}
        tooltip_payload = provenance_payload.get("tooltip_contexts", {}) if isinstance(provenance_payload, dict) else {}
        wb = load_workbook(str(path), read_only=True, data_only=True)
        try:
            sheets: list[ExcelSheetPreview] = []
            for ws in wb.worksheets:
                rows: list[list[str]] = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([self._excel_value_text(value) for value in row])
                sheet_provenance_payload = sheets_payload.get(ws.title, {})
                cell_provenance: dict[str, ExcelCellProvenance] = {}
                if isinstance(sheet_provenance_payload, dict):
                    for coord, item in sheet_provenance_payload.items():
                        if isinstance(item, dict):
                            try:
                                cell_provenance[str(coord)] = ExcelCellProvenance.model_validate(item)
                            except Exception:
                                continue
                sheet_tooltip_payload = tooltip_payload.get(ws.title, {})
                tooltip_contexts: dict[str, ExcelCellTooltipContext] = {}
                if isinstance(sheet_tooltip_payload, dict):
                    for coord, item in sheet_tooltip_payload.items():
                        if isinstance(item, dict):
                            try:
                                tooltip_contexts[str(coord)] = ExcelCellTooltipContext.model_validate(item)
                            except Exception:
                                continue
                sheets.append(
                    ExcelSheetPreview(
                        name=ws.title,
                        rows=rows,
                        cell_provenance=cell_provenance,
                        tooltip_contexts=tooltip_contexts,
                    )
                )
            return ExcelWorkbookPreview(workbook_name=workbook_name, path=str(path), sheets=sheets)
        finally:
            wb.close()

    def refresh_filled_excel_provenance(self) -> None:
        from iev4pi_transformation_tool.core.standardized_templates import collect_filled_templates

        if not self.records:
            self.reload_records()
        if not self.records:
            return
        for workbook_name, path in sorted(collect_filled_templates().items()):
            if path.is_file():
                self._write_excel_provenance(path, workbook_name)

    def update_filled_excel_cell(
        self,
        workbook_name: str,
        sheet_name: str,
        row: int,
        column: int,
        value: str,
    ) -> ExcelCellProvenance | None:
        updated = self.update_filled_excel_cells(workbook_name, sheet_name, {(row, column): value})
        coord = next(iter(updated), "")
        return updated.get(coord)

    def update_filled_excel_cells(
        self,
        workbook_name: str,
        sheet_name: str,
        edits: dict[tuple[int, int], str],
    ) -> dict[str, ExcelCellProvenance | None]:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from iev4pi_transformation_tool.core.standardized_templates import collect_filled_templates

        normalized_edits: dict[tuple[int, int], str] = {}
        for (row, column), value in edits.items():
            normalized_edits[(max(1, int(row)), max(1, int(column)))] = str(value)
        if not normalized_edits:
            return {}
        filled_templates = collect_filled_templates()
        workbook_path = filled_templates.get(workbook_name)
        if workbook_path is None or not workbook_path.is_file():
            raise FileNotFoundError(f"Filled workbook not found: {workbook_name}")
        wb = load_workbook(str(workbook_path))
        try:
            if sheet_name not in wb.sheetnames:
                raise KeyError(f"Sheet not found: {sheet_name}")
            ws = wb[sheet_name]
            for (row, column), value in normalized_edits.items():
                cell = ws.cell(row=row, column=column)
                cell.value = None if value == "" else value
            wb.save(str(workbook_path))
        finally:
            wb.close()

        updated: dict[str, ExcelCellProvenance | None] = {
            f"{get_column_letter(column)}{row}": None
            for row, column in normalized_edits
        }
        provenance_path = self._excel_provenance_path(workbook_path)
        if not provenance_path.is_file():
            return updated
        try:
            payload = json.loads(provenance_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return updated
        sheets_payload = payload.setdefault("sheets", {})
        if not isinstance(sheets_payload, dict):
            return updated
        sheet_payload = sheets_payload.get(sheet_name, {})
        if not isinstance(sheet_payload, dict):
            return updated
        changed_provenance = False
        for (row, column), value in normalized_edits.items():
            coord = f"{get_column_letter(column)}{row}"
            item_payload = sheet_payload.get(coord)
            if not isinstance(item_payload, dict):
                continue
            try:
                provenance = ExcelCellProvenance.model_validate(item_payload)
            except ValidationError:
                continue
            updated_provenance = provenance.model_copy(
                update={
                    "workbook_name": workbook_name,
                    "sheet_name": sheet_name,
                    "row": row,
                    "column": column,
                    "coord": coord,
                    "value": value,
                    "normalized_value": self._normalize_excel_key(value),
                    "confidence": 1.0,
                    "decision_confidence": 1.0,
                    "status": ExtractionStatus.FILLED,
                    "review_feedback_status": "corrected",
                }
            )
            sheet_payload[coord] = updated_provenance.model_dump(mode="json")
            updated[coord] = updated_provenance
            changed_provenance = True
        if changed_provenance:
            provenance_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return updated

    def _load_or_build_excel_provenance(self, workbook_path: Path, workbook_name: str | None = None) -> dict[str, object]:
        provenance_path = self._excel_provenance_path(workbook_path)
        if provenance_path.is_file():
            try:
                payload = json.loads(provenance_path.read_text(encoding="utf-8"))
                if (
                    isinstance(payload, dict)
                    and "tooltip_contexts" in payload
                    and int(payload.get("tooltip_contexts_version", 0) or 0) >= EXCEL_TOOLTIP_CONTEXT_VERSION
                ):
                    return payload
                if isinstance(payload, dict):
                    return self._ensure_excel_tooltip_contexts(workbook_path, workbook_name or workbook_path.name, payload)
            except json.JSONDecodeError:
                pass
        if workbook_path.is_file():
            return self._write_excel_provenance(workbook_path, workbook_name or workbook_path.name)
        return {"workbook_name": workbook_path.name, "path": str(workbook_path), "sheets": {}}

    def _ensure_excel_tooltip_contexts(
        self,
        workbook_path: Path,
        workbook_name: str,
        payload: dict[str, object],
    ) -> dict[str, object]:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        payload.setdefault("workbook_name", workbook_name)
        payload.setdefault("path", str(workbook_path))
        payload["tooltip_contexts_version"] = EXCEL_TOOLTIP_CONTEXT_VERSION
        sheets_payload = payload.setdefault("sheets", {})
        if not isinstance(sheets_payload, dict):
            sheets_payload = {}
            payload["sheets"] = sheets_payload
        tooltip_payload: dict[str, object] = {}
        wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
        try:
            document_sources = self._excel_document_source_map(wb)
            source_objects = self._excel_source_object_map(wb, document_sources)
            is_aio_workbook = self._is_aio_workbook(wb)
            aio_source_metadata = (
                self._excel_aio_source_metadata_map(wb, document_sources, source_objects)
                if is_aio_workbook
                else {}
            )
            for ws in wb.worksheets:
                sheet_tooltip_payload: dict[str, object] = {}
                sheet_provenance_payload = sheets_payload.get(ws.title, {})
                if not isinstance(sheet_provenance_payload, dict):
                    sheet_provenance_payload = {}
                headers: dict[int, str] = {}
                header_labels: dict[int, str] = {}
                for column_index, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1), []), start=1):
                    header_labels[column_index] = self._excel_value_text(getattr(cell, "value", None))
                    headers[column_index] = self._normalize_excel_key(header_labels[column_index])
                for row_index, row in enumerate(ws.iter_rows(), start=1):
                    if not row:
                        continue
                    row_values = {
                        column_index: self._excel_value_text(getattr(cell, "value", None))
                        for column_index, cell in enumerate(row, start=1)
                    }
                    row_metadata = (
                        {}
                        if row_index == 1
                        else self._excel_row_tooltip_metadata(
                            headers,
                            row_values,
                            document_sources,
                            source_objects,
                            sheet_name=ws.title,
                            aio_source_metadata=aio_source_metadata,
                            is_aio_workbook=is_aio_workbook,
                        )
                    )
                    for column_index, cell in enumerate(row, start=1):
                        value = self._excel_value_text(getattr(cell, "value", None))
                        if not self._normalize_excel_key(value):
                            continue
                        coord = f"{get_column_letter(column_index)}{row_index}"
                        provenance = None
                        item_payload = sheet_provenance_payload.get(coord)
                        if isinstance(item_payload, dict):
                            try:
                                provenance = ExcelCellProvenance.model_validate(item_payload)
                            except ValidationError:
                                provenance = None
                        tooltip_context = self._excel_tooltip_context_for_cell(
                            workbook_name=workbook_name,
                            sheet_name=ws.title,
                            row=row_index,
                            column=column_index,
                            coord=coord,
                            value=value,
                            field_name=header_labels.get(column_index, ""),
                            provenance=provenance,
                            row_metadata=row_metadata,
                        )
                        sheet_tooltip_payload[coord] = tooltip_context.model_dump(mode="json")
                if sheet_tooltip_payload:
                    tooltip_payload[ws.title] = sheet_tooltip_payload
        finally:
            wb.close()
        payload["tooltip_contexts"] = tooltip_payload
        provenance_path = self._excel_provenance_path(workbook_path)
        ensure_dir(provenance_path.parent)
        provenance_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return payload

    def _write_excel_provenance(self, workbook_path: Path, workbook_name: str) -> dict[str, object]:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        payload: dict[str, object] = {
            "workbook_name": workbook_name,
            "path": str(workbook_path),
            "sheets": {},
            "tooltip_contexts": {},
            "tooltip_contexts_version": EXCEL_TOOLTIP_CONTEXT_VERSION,
        }
        wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
        try:
            document_sources = self._excel_document_source_map(wb)
            source_objects = self._excel_source_object_map(wb, document_sources)
            is_aio_workbook = self._is_aio_workbook(wb)
            result_index = {} if is_aio_workbook else self._excel_result_index()
            pid_provenance = self._load_pid_provenance()
            aio_source_metadata = (
                self._excel_aio_source_metadata_map(wb, document_sources, source_objects)
                if is_aio_workbook
                else {}
            )
            for ws in wb.worksheets:
                sheet_payload: dict[str, object] = {}
                sheet_tooltip_payload: dict[str, object] = {}
                headers: dict[int, str] = {}
                header_labels: dict[int, str] = {}
                for column_index, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1), []), start=1):
                    header_labels[column_index] = self._excel_value_text(getattr(cell, "value", None))
                    headers[column_index] = self._normalize_excel_key(header_labels[column_index])
                for row_index, row in enumerate(ws.iter_rows(), start=1):
                    if not row:
                        continue
                    row_values = {
                        column_index: self._excel_value_text(getattr(cell, "value", None))
                        for column_index, cell in enumerate(row, start=1)
                    }
                    row_metadata = (
                        {}
                        if row_index == 1
                        else self._excel_row_tooltip_metadata(
                            headers,
                            row_values,
                            document_sources,
                            source_objects,
                            sheet_name=ws.title,
                            aio_source_metadata=aio_source_metadata,
                            is_aio_workbook=is_aio_workbook,
                        )
                    )
                    for column_index, cell in enumerate(row, start=1):
                        value = self._excel_value_text(getattr(cell, "value", None))
                        value_key = self._normalize_excel_key(value)
                        if not value_key:
                            continue
                        coord = f"{get_column_letter(column_index)}{row_index}"
                        field_key = headers.get(column_index, "")
                        item: ExcelCellProvenance | None = None
                        if row_index != 1 and not is_aio_workbook:
                            provenance = self._match_excel_cell_provenance(result_index, field_key, value_key)
                            if provenance is not None:
                                item = provenance.model_copy(
                                    update={
                                        "workbook_name": workbook_name,
                                        "sheet_name": ws.title,
                                        "row": row_index,
                                        "column": column_index,
                                        "coord": coord,
                                    }
                                )
                            elif pid_provenance is not None and ws.title.startswith("P&ID"):
                                pid_sheet = pid_provenance.get(ws.title, {})
                                pid_prov = pid_sheet.get(coord)
                                if pid_prov:
                                    try:
                                        item = ExcelCellProvenance(**pid_prov)
                                    except Exception:
                                        item = None
                            if item is not None:
                                sheet_payload[coord] = item.model_dump(mode="json")
                        tooltip_context = self._excel_tooltip_context_for_cell(
                            workbook_name=workbook_name,
                            sheet_name=ws.title,
                            row=row_index,
                            column=column_index,
                            coord=coord,
                            value=value,
                            field_name=header_labels.get(column_index, ""),
                            provenance=item,
                            row_metadata=row_metadata,
                        )
                        sheet_tooltip_payload[coord] = tooltip_context.model_dump(mode="json")
                if sheet_payload:
                    payload["sheets"][ws.title] = sheet_payload
                if sheet_tooltip_payload:
                    payload["tooltip_contexts"][ws.title] = sheet_tooltip_payload
        finally:
            wb.close()
        provenance_path = self._excel_provenance_path(workbook_path)
        ensure_dir(provenance_path.parent)
        provenance_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return payload

    def _excel_document_source_map(self, workbook) -> dict[str, str]:
        document_sources: dict[str, str] = {}
        for ws in workbook.worksheets:
            header_row = list(next(ws.iter_rows(min_row=1, max_row=1), []))
            headers = {
                self._normalize_excel_key(getattr(cell, "value", None)): index
                for index, cell in enumerate(header_row, start=1)
            }
            document_id_column = headers.get("document_id")
            filename_column = (
                headers.get("document_filename")
                or headers.get("source_file")
                or headers.get("sourcedocid")
                or headers.get("document")
            )
            if not document_id_column or not filename_column:
                continue
            for row in ws.iter_rows(min_row=2):
                row_values = {
                    column_index: self._excel_value_text(getattr(cell, "value", None))
                    for column_index, cell in enumerate(row, start=1)
                }
                document_id = self._normalize_excel_key(row_values.get(document_id_column, ""))
                filename = clean_cell(row_values.get(filename_column, ""))
                if document_id and filename:
                    document_sources.setdefault(document_id, filename)
        return document_sources

    @staticmethod
    def _is_aio_workbook(workbook) -> bool:
        required = {
            "Object",
            "Element_ID",
            "Element_Data",
            "Element_Data_Source",
            "Connection_Data_Source",
            "Attribute_Lookup",
        }
        return required.issubset(set(workbook.sheetnames))

    def _excel_source_object_map(
        self,
        workbook,
        document_sources: dict[str, str],
    ) -> dict[str, dict[str, str]]:
        source_objects: dict[str, dict[str, str]] = {}
        if "Object" not in workbook.sheetnames:
            return source_objects
        ws = workbook["Object"]
        header_row = list(next(ws.iter_rows(min_row=1, max_row=1), []))
        headers = {
            self._normalize_excel_key(getattr(cell, "value", None)): index
            for index, cell in enumerate(header_row, start=1)
        }
        object_column = headers.get("object_id")
        if not object_column:
            return source_objects

        def row_value(row_values: dict[int, str], *keys: str) -> str:
            for key in keys:
                column = headers.get(key)
                if column:
                    value = clean_cell(row_values.get(column, ""))
                    if value:
                        return value
            return ""

        for row in ws.iter_rows(min_row=2):
            row_values = {
                column_index: self._excel_value_text(getattr(cell, "value", None))
                for column_index, cell in enumerate(row, start=1)
            }
            object_id = clean_cell(row_values.get(object_column, ""))
            if not object_id:
                continue
            document_id = self._normalize_excel_key(row_value(row_values, "document_id"))
            page_number = row_value(row_values, "page_number")
            source_operation = row_value(row_values, "source_operation")
            content_text = row_value(row_values, "content_text")
            object_type = row_value(row_values, "object_type")
            bbox_values = [
                row_value(row_values, "bbox_x1"),
                row_value(row_values, "bbox_y1"),
                row_value(row_values, "bbox_x2"),
                row_value(row_values, "bbox_y2"),
            ]
            bbox = ", ".join(value for value in bbox_values if value)
            location_bits = []
            if page_number:
                location_bits.append(f"Page {page_number}")
            if bbox and bbox != "0, 0, 0, 0":
                location_bits.append(f"bbox({bbox})")
            location_bits.append(f"source object {object_id}")
            source_objects[self._normalize_excel_key(object_id)] = {
                "source_path": document_sources.get(document_id, ""),
                "location": " / ".join(location_bits),
                "content_text": content_text,
                "source_operation": source_operation,
                "object_type": object_type,
            }
        return source_objects

    def _excel_aio_source_metadata_map(
        self,
        workbook,
        document_sources: dict[str, str],
        source_objects: dict[str, dict[str, str]],
    ) -> dict[tuple[str, str], dict[str, object]]:
        source_specs = {
            "Document_Data_Source": ("Document_Data", "document_data_id"),
            "Element_Data_Source": ("Element_Data", "element_data_id"),
            "RepresentedItem_Data_Source": ("RepresentedItem_Data", "representeditem_data_id"),
            "Connection_Data_Source": ("Connection_Data", "connection_data_id"),
            "Revision_Data_Source": ("Revision_Data", "revision_id"),
            "Element_Classification_Source": ("Element_Classification", "classification_id"),
        }
        metadata: dict[tuple[str, str], dict[str, object]] = {}

        def merge_metadata(existing: dict[str, object] | None, new: dict[str, object]) -> dict[str, object]:
            if not existing:
                return dict(new)
            merged = dict(existing)
            old_conf = merged.get("confidence")
            new_conf = new.get("confidence")
            if old_conf is None:
                merged["confidence"] = new_conf
            elif new_conf is not None:
                try:
                    merged["confidence"] = min(float(old_conf), float(new_conf))
                except (TypeError, ValueError):
                    pass
            for key in ("decision_confidence", "source_path", "location", "extraction_method"):
                if not clean_cell(merged.get(key, "")) and clean_cell(new.get(key, "")):
                    merged[key] = new.get(key)
            notes = [clean_cell(merged.get("note", "")), clean_cell(new.get("note", ""))]
            unique_notes = []
            for note in notes:
                if note and note not in unique_notes:
                    unique_notes.append(note)
            merged["note"] = " | ".join(unique_notes)
            merged["has_real_metadata"] = True
            merged["source_path_from_document_only"] = False
            return merged

        def sheet_rows(sheet_name: str) -> list[dict[str, str]]:
            if sheet_name not in workbook.sheetnames:
                return []
            ws = workbook[sheet_name]
            header_row = list(next(ws.iter_rows(min_row=1, max_row=1), []))
            headers = {
                index: self._normalize_excel_key(getattr(cell, "value", None))
                for index, cell in enumerate(header_row, start=1)
                if self._normalize_excel_key(getattr(cell, "value", None))
            }
            rows: list[dict[str, str]] = []
            for row in ws.iter_rows(min_row=2):
                item = {
                    header: self._excel_value_text(getattr(cell, "value", None))
                    for index, cell in enumerate(row, start=1)
                    if (header := headers.get(index))
                }
                rows.append(item)
            return rows

        for source_sheet, (target_sheet, target_id_key) in source_specs.items():
            if source_sheet not in workbook.sheetnames:
                continue
            ws = workbook[source_sheet]
            header_row = list(next(ws.iter_rows(min_row=1, max_row=1), []))
            headers = {
                self._normalize_excel_key(getattr(cell, "value", None)): index
                for index, cell in enumerate(header_row, start=1)
            }
            target_col = headers.get(target_id_key)
            if not target_col:
                continue

            def source_value(row_values: dict[int, str], *keys: str) -> str:
                for key in keys:
                    column = headers.get(key)
                    if column:
                        value = clean_cell(row_values.get(column, ""))
                        if value:
                            return value
                return ""

            for row in ws.iter_rows(min_row=2):
                row_values = {
                    column_index: self._excel_value_text(getattr(cell, "value", None))
                    for column_index, cell in enumerate(row, start=1)
                }
                target_id = self._normalize_excel_key(row_values.get(target_col, ""))
                if not target_id:
                    continue
                source_object_id = self._normalize_excel_key(source_value(row_values, "source_object_id"))
                source_object = source_objects.get(source_object_id, {})
                confidence = self._excel_optional_float(source_value(row_values, "confidence"))
                extraction_method = source_value(row_values, "extraction_method")
                review_status = source_value(row_values, "review_status")
                correction_reason = source_value(row_values, "correction_reason")
                source_path = clean_cell(source_object.get("source_path", ""))
                location = clean_cell(source_object.get("location", ""))
                content_text = clean_cell(source_object.get("content_text", ""))
                note_parts = [f"Audit source: {source_sheet}"]
                if review_status:
                    note_parts.append(f"Review status: {review_status}")
                if correction_reason:
                    note_parts.append(f"Correction reason: {correction_reason}")
                if content_text:
                    note_parts.append(f"Source Object.Content_Text: {content_text[:160]}")
                row_metadata = {
                    "confidence": confidence,
                    "decision_confidence": None,
                    "source_path": source_path,
                    "location": location,
                    "extraction_method": extraction_method,
                    "has_real_metadata": True,
                    "source_path_from_document_only": False,
                    "note": " | ".join(note_parts),
                }
                metadata[(target_sheet, target_id)] = row_metadata
                if source_object_id:
                    object_note_parts = [f"Referenced by {source_sheet}"]
                    if review_status:
                        object_note_parts.append(f"Review status: {review_status}")
                    if content_text:
                        object_note_parts.append(f"Object.Content_Text: {content_text[:160]}")
                    object_metadata = dict(row_metadata)
                    object_metadata["note"] = " | ".join(object_note_parts)
                    metadata[("Object", source_object_id)] = merge_metadata(
                        metadata.get(("Object", source_object_id)),
                        object_metadata,
                    )
                    metadata[("Object_Cluster", source_object_id)] = merge_metadata(
                        metadata.get(("Object_Cluster", source_object_id)),
                        object_metadata,
                    )
        element_metadata: dict[str, dict[str, object]] = {}
        element_rows = sheet_rows("Element_Data")
        attribute_priority = {
            "component_id": 0,
            "device_id": 0,
            "terminal_id": 0,
            "element_id": 0,
            "display_label": 1,
            "logical_tag": 1,
            "element_name": 1,
            "primary_rkz": 1,
        }
        for row in element_rows:
            element_id = self._normalize_excel_key(row.get("element_id", ""))
            element_data_id = self._normalize_excel_key(row.get("element_data_id", ""))
            if not element_id or not element_data_id:
                continue
            row_metadata = metadata.get(("Element_Data", element_data_id))
            if not row_metadata:
                continue
            attr_name = self._normalize_excel_key(row.get("attribute_name", ""))
            priority = attribute_priority.get(attr_name, 5)
            candidate = dict(row_metadata)
            candidate["note"] = " | ".join(
                item
                for item in [
                    f"Derived row linked through Element_Data {element_data_id}",
                    clean_cell(row_metadata.get("note", "")),
                ]
                if item
            )
            existing = element_metadata.get(element_id)
            existing_priority = int(existing.get("_priority", 99)) if existing else 99
            if existing is None or priority < existing_priority:
                candidate["_priority"] = priority
                element_metadata[element_id] = candidate
            else:
                element_metadata[element_id] = merge_metadata(existing, candidate)
                element_metadata[element_id]["_priority"] = existing_priority

        match_to_element: dict[str, str] = {}
        topdown_to_element: dict[str, str] = {}
        cluster_element_to_element: dict[str, str] = {}
        for row in sheet_rows("Element_ID"):
            element_id = self._normalize_excel_key(row.get("element_id", ""))
            match_id = self._normalize_excel_key(row.get("source_match_id", ""))
            if element_id and match_id:
                match_to_element[match_id] = element_id
            if element_id and element_id in element_metadata:
                metadata[("Element_ID", element_id)] = dict(element_metadata[element_id])

        for row in sheet_rows("Match_Result"):
            match_id = self._normalize_excel_key(row.get("match_id", ""))
            element_id = match_to_element.get(match_id, "")
            if not element_id:
                continue
            topdown_id = self._normalize_excel_key(row.get("element_topdown_id", ""))
            cluster_element_id = self._normalize_excel_key(row.get("element_from_cluster_id", ""))
            if topdown_id:
                topdown_to_element[topdown_id] = element_id
            if cluster_element_id:
                cluster_element_to_element[cluster_element_id] = element_id
            row_metadata = element_metadata.get(element_id)
            if row_metadata and match_id:
                metadata[("Match_Result", match_id)] = dict(row_metadata)

        for topdown_id, element_id in topdown_to_element.items():
            row_metadata = element_metadata.get(element_id)
            if row_metadata:
                metadata[("Elements_TopDown", topdown_id)] = dict(row_metadata)
        for cluster_element_id, element_id in cluster_element_to_element.items():
            row_metadata = element_metadata.get(element_id)
            if row_metadata:
                metadata[("Elements_from_Cluster", cluster_element_id)] = dict(row_metadata)
        return metadata

    def _excel_row_tooltip_metadata(
        self,
        headers: dict[int, str],
        row_values: dict[int, str],
        document_sources: dict[str, str],
        source_objects: dict[str, dict[str, str]],
        *,
        sheet_name: str = "",
        aio_source_metadata: dict[tuple[str, str], dict[str, object]] | None = None,
        is_aio_workbook: bool = False,
    ) -> dict[str, object]:
        values_by_header = {
            header: clean_cell(row_values.get(column_index, ""))
            for column_index, header in headers.items()
            if header
        }

        def first_value(*keys: str) -> str:
            for key in keys:
                value = clean_cell(values_by_header.get(key, ""))
                if value:
                    return value
            return ""

        if aio_source_metadata:
            primary_key = first_value(
                "document_data_id",
                "element_data_id",
                "representeditem_data_id",
                "connection_data_id",
                "revision_id",
                "classification_id",
                "element_id",
                "element_topdown_id",
                "element_from_cluster_id",
                "match_id",
                "object_id",
                "cluster_id",
            )
            metadata = aio_source_metadata.get((sheet_name, self._normalize_excel_key(primary_key)))
            if metadata:
                return dict(metadata)

        confidence = self._excel_optional_float(
            first_value("confidence", "mapping_confidence", "matchconfidence", "min_confidence", "avg_confidence")
        )
        decision_confidence = self._excel_optional_float(first_value("decisionconfidence", "decision_confidence"))
        source_path = first_value(
            "source_file",
            "sourcefile",
            "source_doc",
            "source_document",
            "sourcedocid",
            "source_fcstd_file",
            "document_filename",
        )
        document_id = self._normalize_excel_key(first_value("document_id"))
        source_path_from_document_only = False
        if not source_path and document_id:
            source_path = document_sources.get(document_id, "")
            source_path_from_document_only = bool(source_path)
        location = first_value("sourcelocator", "source_locator", "source_row", "source_object_id")
        source_object_id = self._normalize_excel_key(first_value("source_object_id"))
        if sheet_name == "Object" and not source_object_id:
            source_object_id = self._normalize_excel_key(first_value("object_id"))
        if source_object_id:
            source_object = source_objects.get(source_object_id, {})
            if source_object:
                if not source_path:
                    source_path = clean_cell(source_object.get("source_path", ""))
                    source_path_from_document_only = False
                source_location = clean_cell(source_object.get("location", ""))
                if source_location:
                    location = source_location
        if not location:
            page_number = first_value("page_number")
            bbox_values = [
                first_value("bbox_x1"),
                first_value("bbox_y1"),
                first_value("bbox_x2"),
                first_value("bbox_y2"),
            ]
            bbox = ", ".join(value for value in bbox_values if value)
            if page_number or bbox:
                location_bits = []
                if page_number:
                    location_bits.append(f"Page {page_number}")
                if bbox and bbox != "0, 0, 0, 0":
                    location_bits.append(f"bbox({bbox})")
                location = " / ".join(location_bits)
        extraction_method = first_value(
            "extraction_method",
            "matchmethod",
            "match_method",
            "source_operation",
            "cluster_method",
            "match_rule",
        )
        note = ""
        if sheet_name == "Object" and source_object_id:
            content_text = clean_cell(source_objects.get(source_object_id, {}).get("content_text", ""))
            if content_text:
                note = f"Source artifact Object.Content_Text: {content_text[:160]}"
        has_real_metadata = any(
            [
                confidence is not None,
                decision_confidence is not None,
                location,
                extraction_method,
            ]
        ) or bool(source_path and not source_path_from_document_only)
        return {
            "confidence": confidence,
            "decision_confidence": decision_confidence,
            "source_path": source_path,
            "location": location,
            "extraction_method": extraction_method,
            "has_real_metadata": has_real_metadata,
            "source_path_from_document_only": source_path_from_document_only,
            "is_aio_workbook": is_aio_workbook,
            "note": note,
        }

    def _excel_tooltip_context_for_cell(
        self,
        *,
        workbook_name: str,
        sheet_name: str,
        row: int,
        column: int,
        coord: str,
        value: str,
        field_name: str,
        provenance: ExcelCellProvenance | None,
        row_metadata: dict[str, object],
    ) -> ExcelCellTooltipContext:
        current_location = f"{sheet_name}!{coord}"
        if provenance is not None:
            method = self._excel_provenance_method(provenance)
            location = self._excel_provenance_location(provenance) or current_location
            return ExcelCellTooltipContext(
                workbook_name=workbook_name,
                sheet_name=sheet_name,
                row=row,
                column=column,
                coord=coord,
                value=value,
                field_name=field_name or provenance.field_name,
                source_type="direct_extraction",
                confidence=float(provenance.confidence or 0.0),
                decision_confidence=provenance.decision_confidence,
                extraction_method=method,
                source_path=provenance.source_path,
                location=location,
                current_location=current_location,
                note="Direct extraction provenance.",
            )
        if bool(row_metadata.get("has_real_metadata")):
            return ExcelCellTooltipContext(
                workbook_name=workbook_name,
                sheet_name=sheet_name,
                row=row,
                column=column,
                coord=coord,
                value=value,
                field_name=field_name,
                source_type="row_metadata",
                confidence=row_metadata.get("confidence"),  # type: ignore[arg-type]
                decision_confidence=row_metadata.get("decision_confidence"),  # type: ignore[arg-type]
                extraction_method=clean_cell(row_metadata.get("extraction_method", "")),
                source_path=clean_cell(row_metadata.get("source_path", "")),
                location=clean_cell(row_metadata.get("location", "")),
                current_location=current_location,
                note=clean_cell(row_metadata.get("note", "")) or "Row-level source metadata from the exported workbook.",
            )
        source_type = "template" if row <= 2 or sheet_name in {"Rules", "Schema_Metadata"} else "exporter_generated"
        if source_type == "template":
            return ExcelCellTooltipContext(
                workbook_name=workbook_name,
                sheet_name=sheet_name,
                row=row,
                column=column,
                coord=coord,
                value=value,
                field_name=field_name,
                source_type=source_type,
                confidence=1.0,
                decision_confidence=None,
                extraction_method="Template_Static",
                source_path=workbook_name,
                location=current_location,
                current_location=current_location,
                note="Static template/header/rule cell from the workbook template.",
            )
        if source_type == "exporter_generated":
            return ExcelCellTooltipContext(
                workbook_name=workbook_name,
                sheet_name=sheet_name,
                row=row,
                column=column,
                coord=coord,
                value=value,
                field_name=field_name,
                source_type=source_type,
                confidence=None,
                decision_confidence=None,
                extraction_method="Deterministic_Exporter",
                source_path=workbook_name,
                location=current_location,
                current_location=current_location,
                note=(
                    "Generated by the deterministic exporter. "
                    "No direct parser/OCR/LLM source object is available for this specific cell; "
                    "source file and location refer to the exported workbook cell, not an original source document."
                ),
            )
        note = (
            "Template/header/rule cell; no direct extraction source."
            if source_type == "template"
            else "Generated or derived export cell; no direct extraction source."
        )
        return ExcelCellTooltipContext(
            workbook_name=workbook_name,
            sheet_name=sheet_name,
            row=row,
            column=column,
            coord=coord,
            value=value,
            field_name=field_name,
            source_type=source_type,
            confidence=None,
            decision_confidence=None,
            extraction_method="",
            source_path="",
            location="",
            current_location=current_location,
            note=note,
        )

    @staticmethod
    def _excel_optional_float(value: object) -> float | None:
        text = clean_cell(value)
        if not text:
            return None
        try:
            return float(text.replace(",", "."))
        except ValueError:
            return None

    @staticmethod
    def _excel_provenance_method(provenance: ExcelCellProvenance) -> str:
        if provenance.evidence_refs:
            evidence = provenance.evidence_refs[0]
            evidence_type = clean_cell(evidence.evidence_type)
            engine = clean_cell(evidence.engine)
            if evidence_type and engine:
                return f"{evidence_type} ({engine})"
            return evidence_type or engine
        if provenance.rule_support:
            return " | ".join(clean_cell(rule) for rule in provenance.rule_support if clean_cell(rule))
        return clean_cell(provenance.llm_verification_status)

    @staticmethod
    def _excel_provenance_location(provenance: ExcelCellProvenance) -> str:
        if not provenance.evidence_refs:
            return ""
        evidence = provenance.evidence_refs[0]
        return " / ".join(
            item
            for item in [clean_cell(evidence.page_or_sheet), clean_cell(evidence.cell_range_or_bbox)]
            if item
        )

    def _excel_result_index(self) -> dict[str, dict[tuple[str, str] | str, list[ExcelCellProvenance]]]:
        by_field_value: dict[tuple[str, str], list[ExcelCellProvenance]] = defaultdict(list)
        by_value: dict[str, list[ExcelCellProvenance]] = defaultdict(list)
        for record in self.records:
            for result in record.results:
                value_key = self._normalize_excel_key(result.value)
                if not value_key:
                    continue
                provenance = self._excel_provenance_for_result(record, result)
                field_key = self._normalize_excel_key(result.field_name)
                by_field_value[(field_key, value_key)].append(provenance)
                by_value[value_key].append(provenance)
        return {
            "by_field_value": by_field_value,
            "by_value": by_value,
        }

    def _match_excel_cell_provenance(
        self,
        result_index: dict[str, dict[tuple[str, str] | str, list[ExcelCellProvenance]]],
        field_key: str,
        value_key: str,
    ) -> ExcelCellProvenance | None:
        by_field_value = result_index.get("by_field_value", {})
        by_value = result_index.get("by_value", {})
        field_matches = by_field_value.get((field_key, value_key), []) if field_key else []
        if len(field_matches) == 1:
            return field_matches[0]
        value_matches = by_value.get(value_key, [])
        if len(value_matches) == 1 and len(value_key) >= 3:
            return value_matches[0]
        return None

    def _excel_provenance_for_result(
        self,
        record: ExtractedRecord,
        result: ExtractedFieldResult,
    ) -> ExcelCellProvenance:
        return ExcelCellProvenance(
            source_path=record.source_path,
            record_key=record.record_key,
            record_display_name=record.display_name,
            field_name=result.field_name,
            value=result.value,
            normalized_value=result.normalized_value,
            confidence=float(result.confidence or 0.0),
            decision_confidence=result.decision_confidence,
            status=result.status,
            notes=result.notes,
            evidence_refs=result.evidence_refs,
            evidence_bundle_id=result.evidence_bundle_id,
            uncertainty_reason=result.uncertainty_reason,
            llm_verification_status=result.llm_verification_status,
            rule_support=result.rule_support,
            review_feedback_status=result.review_feedback_status,
        )

    @staticmethod
    def _excel_provenance_path(workbook_path: Path) -> Path:
        return workbook_path.with_suffix(workbook_path.suffix + ".provenance.json")

    @staticmethod
    def _load_pid_provenance() -> dict[str, dict[str, dict[str, object]]] | None:
        """Load P&ID mapping provenance sidecar, if present."""
        try:
            from iev4pi_transformation_tool.core.standardized_templates import (
                ASSEMBLY_3D_TEMPLATE, FILLED_TEMPLATES_DIR,
            )
            import json
            pid_path = FILLED_TEMPLATES_DIR / (ASSEMBLY_3D_TEMPLATE + ".pid_provenance.json")
            if pid_path.is_file():
                return json.loads(pid_path.read_text(encoding="utf-8"))
        except Exception:
            pass
        return None

    @staticmethod
    def _excel_value_text(value: object) -> str:
        if value is None:
            return ""
        return str(value)

    @staticmethod
    def _normalize_excel_key(value: object) -> str:
        return clean_cell(value).casefold()

    def _report_progress(self, callback: ProgressCallback | None, value: int, message: str) -> None:
        if callback is not None:
            if value >= 0:
                callback(max(0, min(100, int(value))), message)
            else:
                callback(int(value), message)

    def _sub_progress(self, callback: ProgressCallback | None, start: int, end: int) -> ProgressCallback:
        last_scaled = [start]

        def nested(value: int, message: str) -> None:
            if value >= 0:
                scaled = start + round((end - start) * max(0, min(100, value)) / 100)
                if scaled < last_scaled[0]:
                    scaled = last_scaled[0]
                else:
                    last_scaled[0] = scaled
            else:
                scaled = last_scaled[0]
            self._report_progress(callback, scaled, message)

        return nested

    def _sub_text_progress(self, callback: ProgressCallback | None, suffix: str = "") -> ProgressCallback | None:
        if callback is None:
            return None

        def nested(value: int, message: str) -> None:
            self._report_progress(callback, -1, f"{message}{suffix}")

        return nested

    def scan(self, progress: ProgressCallback | None = None) -> ScanSnapshot:
        input_dirs = self.resolve_input_dirs()
        self.parsed_cache = {}
        dir_labels = ", ".join(str(d) for d in input_dirs)
        self._report_progress(progress, 0, f"Scanning {dir_labels}")
        all_files: list[Path] = []
        for d in input_dirs:
            all_files.extend(self.classifier.iter_supported_files([d]))
        # Deduplicate while preserving order
        seen: set[str] = set()
        files: list[Path] = []
        for f in all_files:
            key = f.resolve().as_posix()
            if key not in seen:
                seen.add(key)
                files.append(f)
        total = max(1, len(files))
        self.documents = []
        for index, path in enumerate(files, start=1):
            # Use workspace_root as relative_to so that paths keep their
            # top-level directory prefix (Documents/… or Documents-Others/…).
            self.documents.append(self.classifier.classify(path, relative_to=self.workspace_root))
            self._report_progress(progress, round(index * 90 / total), f"Classified {index}/{len(files)} files")
        self.ri_bundles = self.classifier.discover_ri_bundles(self.documents)
        self.database.upsert_documents(self.documents)
        bundle_suffix = f", {len(self.ri_bundles)} R&I bundles" if self.ri_bundles else ""
        self._report_progress(progress, 100, f"Scan complete: {len(self.documents)} files{bundle_suffix}")
        return self.current_snapshot()

    def current_snapshot(self) -> ScanSnapshot:
        family_counts = Counter(
            family.value
            for document in self.documents
            for family in document.output_families
        )
        source_kind_counts = Counter(document.source_kind.value for document in self.documents)
        return ScanSnapshot(
            documents=self.documents,
            family_counts=dict(family_counts),
            source_kind_counts=dict(source_kind_counts),
            scan_root=str(self.resolve_scan_root()),
            ri_bundles=self.ri_bundles,
        )

    def apply_snapshot_payload(self, payload: dict) -> ScanSnapshot:
        snapshot = ScanSnapshot.model_validate(payload)
        self.documents = snapshot.documents
        self.ri_bundles = snapshot.ri_bundles
        return snapshot

    def reload_schemas(self) -> None:
        self.schemas = {
            schema.family: schema
            for schema in self.database.load_latest_global_schemas().values()
        }
        scoped = self.database.load_latest_scoped_schemas()
        self.ri_bundle_schemas = {}
        for (scope_id, _family_name), schema in scoped.items():
            self.ri_bundle_schemas.setdefault(scope_id, {})[schema.family] = schema

    def reload_records(self) -> None:
        self.records = self.database.load_latest_records()

    def _parse_document(self, document: DocumentDescriptor, progress: ProgressCallback | None = None):
        if document.relative_path not in self.parsed_cache:
            self.parsed_cache[document.relative_path] = self.reader.read(document, progress)
        return self.parsed_cache[document.relative_path]

    def _build_parsed_ri_bundle(
        self,
        bundle: RiBundle,
        pdf_doc: DocumentDescriptor,
        parsed_pdf: ParsedDocument,
        reader: DocumentReader,
    ) -> ParsedDocument:
        cache_key = f"bundle::{bundle.bundle_id}::{reader.ocr_enabled}"
        if reader is self.reader and cache_key in self.parsed_cache:
            return self.parsed_cache[cache_key]
        ri_package = self.dexpi_analyzer.analyze(bundle)
        ri_package.pdf_pages = [page.page_number for page in parsed_pdf.pages]
        bundle_document = DocumentDescriptor.model_validate(
            {
                **pdf_doc.model_dump(mode="python"),
                "output_families": [
                    DocumentFamily.RI_EQUIPMENT_ROW,
                    DocumentFamily.RI_INSTRUMENT_FUNCTION_ROW,
                    DocumentFamily.RI_PIPING_COMPONENT_ROW,
                    DocumentFamily.RI_CONNECTION_ROW,
                ],
                "bundle_id": bundle.bundle_id,
                "bundle_role": "bundle",
            }
        )
        parsed_bundle = ParsedDocument(
            document=bundle_document,
            pages=parsed_pdf.pages,
            metadata={
                **parsed_pdf.metadata,
                "ri_bundle_id": bundle.bundle_id,
                "ri_display_name": bundle.display_name,
                "ri_pairing_status": bundle.pairing_status,
                "ri_pdf_path": pdf_doc.relative_path,
                "ri_xml_path": bundle.xml_path.as_posix(),
                "ri_xsd_path": bundle.xsd_path.as_posix() if bundle.xsd_path else "",
            },
            ri_package=ri_package,
        )
        if reader is self.reader:
            self.parsed_cache[cache_key] = parsed_bundle
        return parsed_bundle

    def _parse_ri_bundle(
        self,
        bundle: RiBundle,
        reader: DocumentReader,
        progress: ProgressCallback | None = None,
        *,
        parsed_pdf: ParsedDocument | None = None,
    ) -> ParsedDocument | None:
        if bundle.pdf_path is None or bundle.xml_path is None:
            return None
        pdf_doc = next((doc for doc in self.documents if doc.path == bundle.pdf_path), None)
        if pdf_doc is None:
            return None
        cache_key = f"bundle::{bundle.bundle_id}::{reader.ocr_enabled}"
        if reader is self.reader and cache_key in self.parsed_cache:
            return self.parsed_cache[cache_key]
        effective_parsed_pdf = parsed_pdf or reader.read(pdf_doc, progress)
        return self._build_parsed_ri_bundle(bundle, pdf_doc, effective_parsed_pdf, reader)

    def _parse_target_label(self, target: DocumentDescriptor | RiBundle) -> str:
        if isinstance(target, RiBundle):
            if target.display_name:
                return target.display_name
            if target.pdf_path is not None:
                return target.pdf_path.name
            return target.bundle_id
        return target.relative_path

    def parse_all(self, progress: ProgressCallback | None = None) -> list:
        return self._parse_all_with_reader(progress=progress, reader=None)

    def _parse_all_with_reader(self, progress: ProgressCallback | None = None, reader: DocumentReader | None = None) -> list:
        if not self.documents:
            self.scan(self._sub_progress(progress, 0, 20) if progress else None)
        parsed_documents = []
        active_reader = reader or self.reader
        use_cache = active_reader is self.reader
        parse_targets: list[DocumentDescriptor | RiBundle] = [
            document
            for document in self.documents
            if document.source_kind != SourceDocumentKind.RI_FLOWSHEET
        ] + [bundle for bundle in self.ri_bundles if bundle.pairing_status in {"paired", "missing_xsd"}]
        total = max(1, len(parse_targets))
        batched_pdf_documents: list[DocumentDescriptor] = []
        seen_pdf_paths: set[str] = set()
        for target in parse_targets:
            if isinstance(target, DocumentDescriptor) and target.extension == ".pdf":
                if target.relative_path not in seen_pdf_paths:
                    seen_pdf_paths.add(target.relative_path)
                    batched_pdf_documents.append(target)
                continue
            if isinstance(target, RiBundle) and target.pdf_path is not None:
                pdf_doc = next((doc for doc in self.documents if doc.path == target.pdf_path), None)
                if pdf_doc is not None and pdf_doc.relative_path not in seen_pdf_paths:
                    seen_pdf_paths.add(pdf_doc.relative_path)
                    batched_pdf_documents.append(pdf_doc)
        pending_batched_pdf_documents = [
            document
            for document in batched_pdf_documents
            if not (use_cache and document.relative_path in self.parsed_cache)
        ]
        batched_pdf_results: dict[str, ParsedDocument] = {}
        batched_phase_end = 0
        if pending_batched_pdf_documents:
            weighted_total = len(parse_targets) + len(pending_batched_pdf_documents)
            batched_phase_end = max(
                35,
                min(
                    85,
                    round(200 * len(pending_batched_pdf_documents) / max(1, weighted_total)),
                ),
            )
            self._report_progress(
                progress,
                0,
                f"Preparing deferred PDF analysis for {len(pending_batched_pdf_documents)} documents",
            )
            batched_pdf_results = active_reader.read_many(
                pending_batched_pdf_documents,
                self._sub_progress(progress, 0, batched_phase_end) if progress else None,
            )
            if use_cache:
                self.parsed_cache.update(batched_pdf_results)

        loop_progress_start = batched_phase_end
        loop_progress_span = max(0, 100 - loop_progress_start)
        for index, target in enumerate(parse_targets, start=1):
            label = self._parse_target_label(target)
            short_label = label.split('/')[-1] if '/' in label else label
            text_only_progress = self._sub_text_progress(progress, f" ({short_label}, {index}/{len(parse_targets)})")
            
            self._report_progress(
                progress,
                loop_progress_start + round((index - 1) * loop_progress_span / total),
                f"Reading {index}/{len(parse_targets)}: {label}",
            )
            if isinstance(target, RiBundle):
                parsed_pdf = None
                if target.pdf_path is not None:
                    pdf_doc = next((doc for doc in self.documents if doc.path == target.pdf_path), None)
                    if pdf_doc is not None:
                        parsed_pdf = batched_pdf_results.get(pdf_doc.relative_path)
                        if parsed_pdf is None and use_cache:
                            parsed_pdf = self.parsed_cache.get(pdf_doc.relative_path)
                parsed = self._parse_ri_bundle(
                    target,
                    active_reader,
                    text_only_progress,
                    parsed_pdf=parsed_pdf,
                )
                if parsed is not None:
                    parsed_documents.append(parsed)
            else:
                if target.extension == ".pdf":
                    parsed = batched_pdf_results.get(target.relative_path)
                    if parsed is None and use_cache:
                        parsed = self.parsed_cache.get(target.relative_path)
                    if parsed is not None:
                        parsed_documents.append(parsed)
                    elif use_cache:
                        parsed_documents.append(self._parse_document(target, text_only_progress))
                    else:
                        parsed_documents.append(active_reader.read(target, text_only_progress))
                    self._report_progress(
                        progress,
                        loop_progress_start + round(index * loop_progress_span / total),
                        f"Finished {index}/{len(parse_targets)}: {label}",
                    )
                    continue
                if use_cache:
                    parsed_documents.append(self._parse_document(target, text_only_progress))
                else:
                    parsed_documents.append(active_reader.read(target, text_only_progress))
            self._report_progress(
                progress,
                loop_progress_start + round(index * loop_progress_span / total),
                f"Finished {index}/{len(parse_targets)}: {label}",
            )
        return parsed_documents

    def build_chunks(self, parsed_documents: list | None = None, progress: ProgressCallback | None = None) -> list:
        chunk_progress = progress
        if parsed_documents is None:
            parsed_documents = self.parse_all(self._sub_progress(progress, 0, 40) if progress else None)
            chunk_progress = self._sub_progress(progress, 40, 100) if progress else None
        chunks = []
        total = max(1, len(parsed_documents))
        for index, parsed in enumerate(parsed_documents, start=1):
            label = parsed.document.relative_path
            short_label = label.split("/")[-1] if "/" in label else label
            chunks.extend(self.chunk_builder.build(parsed))
            self._report_progress(
                chunk_progress,
                round(index * 55 / total),
                f"Building retrieval chunks {short_label} ({index}/{len(parsed_documents)})",
            )
        self.retriever.build(
            chunks,
            progress=self._sub_progress(chunk_progress, 55, 95) if chunk_progress else None,
        )
        self._report_progress(chunk_progress, 96, f"Indexed {len(chunks)} chunks for RAG")
        self.database.replace_chunks(
            (
                {
                    "id": chunk.id,
                    "document_path": chunk.document_path,
                    "family": chunk.family.value,
                    "source_kind": chunk.source_kind.value,
                    "source_locator": chunk.source_locator,
                    "text": chunk.text,
                    "tokens": chunk.tokens,
                    "metadata_json": json.dumps(chunk.metadata, ensure_ascii=False),
                }
                for chunk in chunks
            )
        )
        self._report_progress(chunk_progress, 100, "Chunk index ready")
        return chunks

    def generate_schemas(
        self,
        progress: ProgressCallback | None = None,
        *,
        use_ocr: bool | None = None,
        prune_blank_fields: bool = True,
    ) -> dict[str, SchemaFamily]:
        reader = None
        if use_ocr is not None and use_ocr != self.settings.ocr_enabled:
            reader = self._build_reader_with_overrides(ocr_enabled=use_ocr)
        parsed_documents = self._parse_all_with_reader(
            self._sub_progress(progress, 0, 55) if progress else None,
            reader=reader,
        )
        family_map: dict[DocumentFamily, list] = {family: [] for family in DocumentFamily}
        bundle_map: dict[str, dict[DocumentFamily, list[ParsedDocument]]] = {}
        for parsed in parsed_documents:
            bundle_id = parsed.document.bundle_id or ""
            for family in parsed.document.output_families:
                if self._is_ri_family(family) and bundle_id:
                    bundle_map.setdefault(bundle_id, {}).setdefault(family, []).append(parsed)
                else:
                    family_map[family].append(parsed)

        candidate_schemas: dict[DocumentFamily, SchemaFamily] = {}
        candidate_ri_bundle_schemas: dict[str, dict[DocumentFamily, SchemaFamily]] = {}
        families_with_documents = [
            (family, documents)
            for family, documents in family_map.items()
            if documents and not self._is_ri_family(family)
        ]
        total = max(1, len(families_with_documents) + sum(len(items) for items in bundle_map.values()))
        completed = 0
        for index, (family, documents) in enumerate(families_with_documents, start=1):
            schema = self.schema_miner.mine_family(family, documents)
            candidate_schemas[family] = schema
            completed += 1
            self._report_progress(progress, 55 + round(completed * 45 / total), f"Generated schema {family.value}")
        for bundle_id, bundle_families in sorted(bundle_map.items()):
            bundle = self._bundle_by_id(bundle_id)
            bundle_name = bundle.display_name if bundle else bundle_id
            for family, documents in sorted(bundle_families.items(), key=lambda item: item[0].value):
                schema = self.schema_miner.mine_family(family, documents)
                schema.scope_id = bundle_id
                schema.source_root = bundle.source_root if bundle else "R&I-Fließbild"
                schema.bundle_name = bundle_name
                schema.sheet_name = self._ri_sheet_name(family)
                candidate_ri_bundle_schemas.setdefault(bundle_id, {})[family] = schema
                completed += 1
                self._report_progress(progress, 55 + round(completed * 45 / total), f"Generated schema {bundle_name} / {family.value}")
        if prune_blank_fields and (candidate_schemas or candidate_ri_bundle_schemas):
            self._report_progress(progress, 92, "Pruning schema fields that remain blank across the current result set")
            candidate_schemas, candidate_ri_bundle_schemas = self._prune_blank_schema_fields(
                parsed_documents,
                candidate_schemas,
                candidate_ri_bundle_schemas,
            )
        self.schemas = candidate_schemas
        self.ri_bundle_schemas = candidate_ri_bundle_schemas
        for schema in self.schemas.values():
            self.database.save_schema(schema)
        for bundle_schemas in self.ri_bundle_schemas.values():
            for schema in bundle_schemas.values():
                self.database.save_schema(schema)
        if not families_with_documents and not bundle_map:
            self._report_progress(progress, 100, "No schema families found")
        else:
            self._report_progress(progress, 100, f"Generated {len(self.schemas) + sum(len(items) for items in self.ri_bundle_schemas.values())} schemas")
        return self.all_schema_entries()

    def _prune_blank_schema_fields(
        self,
        parsed_documents: list[ParsedDocument],
        schemas: dict[DocumentFamily, SchemaFamily],
        ri_bundle_schemas: dict[str, dict[DocumentFamily, SchemaFamily]],
    ) -> tuple[dict[DocumentFamily, SchemaFamily], dict[str, dict[DocumentFamily, SchemaFamily]]]:
        if not parsed_documents or (not schemas and not ri_bundle_schemas):
            return schemas, ri_bundle_schemas

        chunks = []
        for parsed in parsed_documents:
            chunks.extend(self.chunk_builder.build(parsed))
        temp_retriever = Retriever(logger=self.log_debug)
        temp_retriever.build(chunks)
        temp_extractor = Extractor(temp_retriever, logger=self.log_debug)

        global_used_fields: defaultdict[DocumentFamily, set[str]] = defaultdict(set)
        scoped_used_fields: defaultdict[tuple[str, DocumentFamily], set[str]] = defaultdict(set)
        reference_tokens: set[str] = set()
        ordered_documents = sorted(
            parsed_documents,
            key=lambda parsed: 1
            if any(
                family in {
                    DocumentFamily.STROMLAUF_COMPONENT_GROUP,
                    DocumentFamily.STROMLAUF_COMPONENT,
                    DocumentFamily.STROMLAUF_CONNECTION,
                }
                for family in parsed.document.output_families
            )
            else 0,
        )

        for parsed in ordered_documents:
            active_schemas = {**schemas, **ri_bundle_schemas.get(parsed.document.bundle_id or "", {})}
            extracted = temp_extractor.extract(
                parsed,
                active_schemas,
                retrieval_top_k=self.settings.retrieval_top_k,
                reference_tokens=reference_tokens,
            )
            for record in extracted:
                for result in record.results:
                    if not result.value.strip():
                        continue
                    if record.scope_id:
                        scoped_used_fields[(record.scope_id, record.family)].add(result.field_name)
                    else:
                        global_used_fields[record.family].add(result.field_name)
                    if any(key in result.field_name for key in ("tag", "id", "component", "klemme")):
                        reference_tokens.add(result.value)

        pruned_schemas = {
            family: self._schema_with_pruned_blank_fields(schema, global_used_fields.get(family, set()))
            for family, schema in schemas.items()
        }
        pruned_ri_bundle_schemas: dict[str, dict[DocumentFamily, SchemaFamily]] = {}
        for bundle_id, bundle_schemas in ri_bundle_schemas.items():
            pruned_ri_bundle_schemas[bundle_id] = {
                family: self._schema_with_pruned_blank_fields(
                    schema,
                    scoped_used_fields.get((bundle_id, family), set()),
                )
                for family, schema in bundle_schemas.items()
            }
        return pruned_schemas, pruned_ri_bundle_schemas

    def _schema_with_pruned_blank_fields(self, schema: SchemaFamily, used_fields: set[str]) -> SchemaFamily:
        if not used_fields:
            return schema
        kept_fields = [field for field in schema.fields if field.name in used_fields]
        if not kept_fields or len(kept_fields) == len(schema.fields):
            return schema
        pruned_names = [field.name for field in schema.fields if field.name not in used_fields]
        pruned_schema = schema.model_copy(deep=True)
        pruned_schema.fields = kept_fields
        note = (
            f"Pruned {len(pruned_names)} fields with no extracted values across the current document set: "
            + ", ".join(pruned_names[:12])
        )
        if len(pruned_names) > 12:
            note += ", ..."
        if note not in pruned_schema.review_notes:
            pruned_schema.review_notes = [*pruned_schema.review_notes, note]
        return pruned_schema

    def update_schema(self, schema: SchemaFamily) -> None:
        if schema.scope_id:
            self.ri_bundle_schemas.setdefault(schema.scope_id, {})[schema.family] = schema
        else:
            self.schemas[schema.family] = schema
        self.database.save_schema(schema)

    def _reset_database_before_extraction(self, progress: ProgressCallback | None = None) -> None:
        self._report_progress(progress, 2, "Clearing previous database state")
        self.database.reset_state()
        self.records = []
        if self.documents:
            self.database.upsert_documents(self.documents)
            self._report_progress(progress, 5, f"Restored {len(self.documents)} scanned documents")
        if self.schemas:
            for schema in sorted(self.schemas.values(), key=lambda item: item.family.value):
                self.database.save_schema(schema)
            self._report_progress(progress, 8, f"Restored {len(self.schemas)} schema definitions")
        if self.ri_bundle_schemas:
            count = 0
            for bundle_schemas in self.ri_bundle_schemas.values():
                for schema in sorted(bundle_schemas.values(), key=lambda item: item.family.value):
                    self.database.save_schema(schema)
                    count += 1
            self._report_progress(progress, 9, f"Restored {count} scoped R&I schema definitions")
        self._report_progress(progress, 10, "Database reset complete")

    def run_extraction(
        self,
        progress: ProgressCallback | None = None,
        *,
        use_ocr: bool | None = None,
    ) -> RunSummary:
        schema_progress_start = 0
        if self.settings.clear_database_before_extraction:
            self._reset_database_before_extraction(progress)
            schema_progress_start = 10
        if not self.schemas and not self.ri_bundle_schemas:
            self.reload_schemas()
        if not self.schemas and not self.ri_bundle_schemas:
            self.generate_schemas(
                self._sub_progress(progress, schema_progress_start, 30) if progress else None,
                use_ocr=use_ocr,
                prune_blank_fields=False,
            )
        else:
            existing_count = len(self.schemas) + sum(len(items) for items in self.ri_bundle_schemas.values())
            self._report_progress(progress, max(10, schema_progress_start), f"Using {existing_count} existing schemas")
        reader = None
        extraction_ocr_enabled = self.settings.ocr_enabled if use_ocr is None else bool(use_ocr)
        if extraction_ocr_enabled != self.settings.ocr_enabled:
            reader = self._build_reader_with_overrides(ocr_enabled=extraction_ocr_enabled)
        parsed_documents = self._parse_all_with_reader(
            self._sub_progress(progress, 30, 45) if progress else None,
            reader=reader,
        )
        self.build_chunks(parsed_documents, self._sub_progress(progress, 45, 65) if progress else None)
        run_id = self.database.create_run(
            config={
                "retrieval_top_k": self.settings.retrieval_top_k,
                "ocr_enabled": extraction_ocr_enabled,
                "ocr_zoom": self.settings.ocr_zoom,
                "ocr_backend": self.settings.ocr_backend,
                "ocr_fallback_backend": self.settings.ocr_fallback_backend,
                "ocr_device": self.settings.ocr_device,
                "apple_ocr_framework": self.settings.apple_ocr_framework,
                "apple_ocr_recognition_level": self.settings.apple_ocr_recognition_level,
                "ocr_dpi": self.settings.ocr_dpi,
                "diagram_dpi": self.settings.diagram_dpi,
                "diagram_analysis_mode": self.settings.diagram_analysis_mode,
                "ocr_min_confidence": self.settings.ocr_min_confidence,
                "enable_diagram_relation_extraction": self.settings.enable_diagram_relation_extraction,
                "enable_hard_page_fallback": self.settings.enable_hard_page_fallback,
                "clear_database_before_extraction": self.settings.clear_database_before_extraction,
            }
        )

        def sort_key(parsed) -> int:
            return 1 if any(
                family in {
                    DocumentFamily.STROMLAUF_COMPONENT_GROUP,
                    DocumentFamily.STROMLAUF_COMPONENT,
                    DocumentFamily.STROMLAUF_CONNECTION,
                }
                for family in parsed.document.output_families
            ) else 0

        ordered_documents = sorted(parsed_documents, key=sort_key)
        total = max(1, len(ordered_documents))
        records: list[ExtractedRecord] = []

        # ---- Phase 1: pre-collect reference_tokens from raw document data ----
        reference_tokens: set[str] = set()
        for parsed in ordered_documents:
            for page in parsed.pages:
                for kv in page.kv_pairs:
                    if kv.value and isinstance(kv.value, str) and len(kv.value.strip()) > 1:
                        reference_tokens.add(kv.value.strip())
                for block in page.blocks:
                    for token in re.findall(r"\b[A-Z]{2,}\d{2,}[A-Z0-9.-]*\b", block.text or ""):
                        reference_tokens.add(token)

        # ---- Phase 2: parallel extraction ----
        import threading
        from concurrent.futures import as_completed

        records_map: dict[int, list[ExtractedRecord]] = {}
        progress_lock = threading.Lock()
        completed_count = [0]

        def _extract_one(index: int, parsed: ParsedDocument) -> tuple[int, list[ExtractedRecord]]:
            scoped_schemas = self.ri_bundle_schemas.get(parsed.document.bundle_id or "", {})
            active_schemas = {**self.schemas, **scoped_schemas}
            extracted = self.extractor.extract(
                parsed,
                active_schemas,
                retrieval_top_k=self.settings.retrieval_top_k,
                reference_tokens=reference_tokens,
                progress=None,  # per-document progress disabled in parallel mode
            )
            label = parsed.document.relative_path
            short_label = label.split("/")[-1] if "/" in label else label
            with progress_lock:
                completed_count[0] += 1
                self._report_progress(
                    progress,
                    65 + round(completed_count[0] * 25 / total),
                    f"Extracted {short_label} ({completed_count[0]}/{total})",
                )
            return index, extracted

        from iev4pi_transformation_tool.core.qos_helpers import QoSAwareThreadPoolExecutor, io_worker_count

        worker_count = min(io_worker_count(cap=10), total)
        with QoSAwareThreadPoolExecutor(max_workers=worker_count) as executor:
            futures = {
                executor.submit(_extract_one, idx, parsed): idx
                for idx, parsed in enumerate(ordered_documents)
            }
            for future in as_completed(futures):
                idx, extracted = future.result()
                records_map[idx] = extracted

        # ---- Phase 3: ordered reconstruction + token accumulation ----
        for idx, parsed in enumerate(ordered_documents):
            extracted = records_map.get(idx, [])
            records.extend(extracted)
            for record in extracted:
                for result in record.results:
                    if result.value and any(key in result.field_name for key in ("tag", "id", "component", "klemme")):
                        reference_tokens.add(result.value)
        self.records = records
        self._report_progress(progress, 94, f"Saving {len(records)} records")
        self.database.save_run_records(run_id, records)
        summary = RunSummary(
            run_id=run_id,
            status="completed",
            record_count=len(records),
            family_counts=dict(Counter(record.family.value for record in records)),
            output_dir=str(self.resolve_results_export_dir() / "Excel"),
        )
        self.database.finalize_run(summary)
        self._latest_run_summary = summary
        self._report_progress(progress, 100, f"Extraction complete: {len(records)} records")
        return summary

    def fill_standardized_templates(
        self,
        progress: ProgressCallback | None = None,
        *,
        use_ocr: bool | None = None,
    ) -> RunSummary:
        """Scan, generate schemas, extract, and fill standardized templates in-place."""
        import tempfile

        from iev4pi_transformation_tool.core.standardized_export import (
            export_standardized_workbook, set_title_block_llm,
        )
        # Wire the LLM client so title block extraction uses semantic parsing
        set_title_block_llm(self.llm_client)
        aio_ml_benchmark_report_path = self._aio_ml_evidence_linking_benchmark_report_path()
        aio_ml_evidence_linking_enabled = self._aio_ml_evidence_linking_enabled()
        from iev4pi_transformation_tool.core.standardized_templates import (
            AIO_TEMPLATE,
            ASSEMBLY_3D_TEMPLATE,
            FAMILY_TO_STANDARDIZED_TEMPLATE,
            get_template_output_path,
        )

        # 0. Ensure Assembly 3D template exists before scan/schema
        self._report_progress(progress, 0, "Ensuring Assembly 3D template")
        self._ensure_assembly_3d_template()

        # 1. Scan workspace
        self._report_progress(progress, 2, "Scanning workspace")
        if not self.documents:
            self.scan(self._sub_progress(progress, 2, 10) if progress else None)

        # 2. Generate schemas
        self._report_progress(progress, 10, "Generating schemas")
        if not self.schemas and not self.ri_bundle_schemas:
            self.reload_schemas()
        if not self.schemas and not self.ri_bundle_schemas:
            self.generate_schemas(
                self._sub_progress(progress, 10, 20) if progress else None,
                use_ocr=use_ocr,
                prune_blank_fields=False,
            )

        # 3. Run extraction
        self._report_progress(progress, 20, "Extracting records")
        summary = self.run_extraction(
            self._sub_progress(progress, 20, 70) if progress else None,
            use_ocr=use_ocr,
        )

        # 4. Fill standardized templates → saved to data/filled_templates/
        self._report_progress(progress, 70, "Filling standardized templates")

        records_by_template: dict[str, list[ExtractedRecord]] = defaultdict(list)
        for record in self.records:
            template_name = FAMILY_TO_STANDARDIZED_TEMPLATE.get(record.family.value)
            if template_name:
                records_by_template[template_name].append(record)

        filled = 0
        template_items: list[tuple[str, list[ExtractedRecord], str, int]] = []
        for template_name, template_records in records_by_template.items():
            families_for_template = [
                fam
                for fam, tpl in FAMILY_TO_STANDARDIZED_TEMPLATE.items()
                if tpl == template_name
            ]
            if not families_for_template:
                continue
            # AIO export expands records into per-document workbooks and may run
            # evidence linking, so its progress budget must reflect real work.
            weight = max(1, len(template_records)) * (4 if template_name == AIO_TEMPLATE else 1)
            template_items.append((template_name, template_records, families_for_template[0], weight))
        total_weight = max(1, sum(item[3] for item in template_items))
        template_progress_start = 70
        template_progress_end = 92
        template_progress_span = template_progress_end - template_progress_start
        completed_weight = 0
        last_template_progress = template_progress_start

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            for template_name, template_records, primary_family, weight in template_items:
                raw_start = template_progress_start + round(completed_weight * template_progress_span / total_weight)
                completed_weight += weight
                raw_end = template_progress_start + round(completed_weight * template_progress_span / total_weight)
                template_start = max(raw_start, last_template_progress)
                template_end = max(raw_end, template_start)
                if template_name == AIO_TEMPLATE and template_end <= template_start and template_start < template_progress_end:
                    template_end = template_start + 1

                self._report_progress(
                    progress,
                    template_start,
                    f"Filling {template_name}",
                )

                exported = export_standardized_workbook(
                    tmp_path,
                    primary_family,
                    template_records,
                    aio_ml_evidence_linking_enabled=aio_ml_evidence_linking_enabled,
                    aio_ml_benchmark_report_path=aio_ml_benchmark_report_path,
                    progress=self._sub_progress(progress, template_start, template_end) if progress else None,
                )
                if exported and exported.is_file():
                    dest = get_template_output_path(primary_family)
                    if dest and exported.resolve() != dest.resolve():
                        shutil.copy2(str(exported), str(dest))
                        filled += 1
                self._report_progress(
                    progress,
                    template_end,
                    f"Filled {template_name}",
                )
                last_template_progress = template_end

        # 4b. Assembly_3D does not depend on extracted records — it reads
        # from a pre-built file.  Export it even when no IFC family records
        # were produced.
        if ASSEMBLY_3D_TEMPLATE not in records_by_template:
            self._report_progress(progress, 92, "Filling Assembly_3D template")
            _asm_family = next(
                (fam for fam, tpl in FAMILY_TO_STANDARDIZED_TEMPLATE.items()
                 if tpl == ASSEMBLY_3D_TEMPLATE), None
            )
            if _asm_family:
                with tempfile.TemporaryDirectory() as _asm_tmp:
                    _asm_exported = export_standardized_workbook(
                        Path(_asm_tmp), _asm_family, [],
                        aio_ml_evidence_linking_enabled=aio_ml_evidence_linking_enabled,
                        aio_ml_benchmark_report_path=aio_ml_benchmark_report_path,
                    )
                    if _asm_exported and _asm_exported.is_file():
                        _asm_dest = get_template_output_path(_asm_family)
                        if _asm_dest and _asm_exported.resolve() != _asm_dest.resolve():
                            shutil.copy2(str(_asm_exported), str(_asm_dest))
                            filled += 1

        # 4c. Fill Stellenplan (instrument list) from ALL datasheet records.
        # Individual TU PDFs map to Datasheet template, but together they form
        # the instrument list — so we aggregate them into the Stellenplan template.
        # This runs HERE (not in save_extraction_results) so it's always fresh.
        from iev4pi_transformation_tool.models import DocumentFamily as _DF
        _ds_records = [r for r in self.records if r.family == _DF.STELLEN_TU_DATASHEET]
        if _ds_records:
            self._report_progress(progress, 93, "Filling Stellenplan aggregate")
            try:
                from iev4pi_transformation_tool.core.standardized_export import _export_stellenplan
                from iev4pi_transformation_tool.core.standardized_templates import (
                    FILLED_TEMPLATES_DIR as _FTD, STELLENPLAN_TEMPLATE as _ST,
                )
                import shutil as _shutil
                with tempfile.TemporaryDirectory() as _stmp:
                    _sp = _export_stellenplan(Path(_stmp), _DF.STELLEN_TU_DATASHEET.value, _ds_records)
                    if _sp and _sp.is_file():
                        _FTD.mkdir(parents=True, exist_ok=True)
                        _sd = _FTD / _ST
                        if _sp.resolve() != _sd.resolve():
                            _shutil.copy2(str(_sp), str(_sd))
                        filled += 1
            except Exception as _exc:
                self.log_debug("stellenplan_export_error", f"Stellenplan export failed: {_exc}")

        # 5. Enrich filled templates with DEXPI analysis (progress 94→99)
        enrichment_progress = self._sub_progress(progress, 94, 99) if progress else None
        self._enrich_filled_templates(enrichment_progress)
        self.refresh_filled_excel_provenance()
        self._report_progress(progress, 100, f"Extraction complete: {len(self.records)} records, {filled} templates filled")
        return summary

    def _enrich_filled_templates(self, progress: ProgressCallback | None = None) -> None:
        """Run DEXPI analysis and write results into the filled template Excel files."""
        from concurrent.futures import as_completed
        from hashlib import sha1

        from iev4pi_transformation_tool.core.standardized_templates import (
            FILLED_TEMPLATES_DIR,
        )

        # --- fingerprint check: skip if records haven't changed ---
        FILLED_TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
        fp_file = FILLED_TEMPLATES_DIR / ".enrichment_fingerprint"
        if self.records:
            payload = "|".join(
                f"{r.record_key}:{r.family.value}"
                for r in sorted(self.records, key=lambda r: r.record_key)
            )
            current_fp = sha1(payload.encode()).hexdigest()
            if fp_file.exists() and fp_file.read_text().strip() == current_fp:
                self._report_progress(progress, 100, "Enrichment skipped (fingerprint unchanged)")
                return
        else:
            current_fp = None

        if not self.documents:
            self.scan()
        if not self.records:
            self.reload_records()

        # Cross-document analysis (~60% of enrichment time, must be serial)
        self._report_progress(progress, 0, "Enriching templates: analysing DEXPI cross-document data")
        source_rows = self._uc1_build_source_rows(
            self._sub_progress(progress, 0, 55) if progress else None
        )

        # Collect enrichment tasks (each operates on a different file → thread-safe)
        enrich_defs = [
            ("PID_template.xlsx", "Instrumentation", "pid", "AKZ_Canonical"),
            ("PID_template.xlsx", "Piping", "piping", "AKZ_Canonical"),
            ("Stellenplan_template.xlsx", "Instrument_Data", "instrument_list", "CanonicalTag"),
            ("Klemmenplan_template.xlsx", "Terminal_ID", "wiring", "CanonicalTag"),
            ("Datasheet_template.xlsx", "Device_ID", "datasheet", "CanonicalTag"),
            ("Stromlaufplan_template.xlsx", "Object_ID", "stromlaufplan", "Object_ID"),
        ]
        tasks: list[tuple[int, str, str, Path, str, str]] = []
        for idx, (filename, sheet, key, key_field) in enumerate(enrich_defs):
            path = FILLED_TEMPLATES_DIR / filename
            if path.is_file() and source_rows.get(key):
                tasks.append((idx, filename, sheet, path, key, key_field))

        # Parallel write (each template is a separate file → no write conflicts)
        total = max(1, len(tasks))
        from iev4pi_transformation_tool.core.qos_helpers import pcore_worker_count, QoSAwareThreadPoolExecutor

        with QoSAwareThreadPoolExecutor(max_workers=min(pcore_worker_count(), total)) as executor:
            futures = {
                executor.submit(
                    self._enrich_template_sheet, path, sheet, source_rows[key], key_field
                ): (idx, filename, sheet)
                for idx, filename, sheet, path, key, key_field in tasks
            }
            for future in as_completed(futures):
                idx, filename, sheet = futures[future]
                try:
                    future.result()
                except Exception as exc:
                    self.log_debug(
                        source="enrich", action="enrich_template",
                        message=f"Failed to enrich {filename}/{sheet}: {exc}",
                        level="WARNING",
                    )
                pct = 55 + round((idx + 1) * 40 / total)
                self._report_progress(
                    progress, pct,
                    f"Enriching templates: {filename}/{sheet} ({idx + 1}/{total})",
                )

        # Write fingerprint so next run can skip
        if current_fp:
            fp_file.write_text(current_fp)
        self._report_progress(progress, 100, "Enriching templates: done")
        self.refresh_filled_excel_provenance()

    @staticmethod
    def _enrich_template_sheet(
        filepath: Path, sheet_name: str, source_rows: list[dict],
        key_field: str,
    ) -> None:
        """Write enriched source rows into an Excel template sheet.

        Matches source rows to existing template rows by *key_field*,
        then fills any additional columns present in source rows but
        not yet in the template header.

        Matching is case-insensitive and underscore-insensitive to handle
        snake_case ↔ CamelCase conventions between source rows and
        template headers.
        """

        def _norm(s: str) -> str:
            return s.lower().replace("_", "")

        import openpyxl

        wb = openpyxl.load_workbook(str(filepath))
        if sheet_name not in wb.sheetnames:
            wb.close()
            return
        ws = wb[sheet_name]

        # Build header map: {column_name: col_index} + normalized alias
        header_map: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val:
                name = str(val).strip()
                header_map[name] = col_idx
                header_map[_norm(name)] = col_idx

        # Build index of template rows by key_field (try exact + normalized)
        key_col = header_map.get(key_field) or header_map.get(_norm(key_field))
        if key_col is None:
            wb.close()
            return
        template_row_map: dict[str, int] = {}
        for row_idx in range(3, ws.max_row + 1):
            key_val = clean_cell(ws.cell(row=row_idx, column=key_col).value)
            if key_val:
                template_row_map[key_val] = row_idx

        # Normalize source row keys once for fast lookup
        for src_row in source_rows:
            norm_keys: dict[str, str] = {}
            for k in src_row:
                norm_keys[_norm(k)] = k

            # Resolve key_field to an actual source key
            src_key = norm_keys.get(_norm(key_field))
            if src_key is None:
                continue
            key_val = clean_cell(src_row.get(src_key, ""))
            if not key_val:
                continue
            target_row = template_row_map.get(key_val)
            if target_row is None:
                continue

            for field_name, field_value in src_row.items():
                if not field_name:
                    continue
                col = header_map.get(field_name) or header_map.get(_norm(field_name))
                if col is None:
                    continue
                existing = ws.cell(row=target_row, column=col).value
                if existing is None or str(existing).strip() == "":
                    ws.cell(row=target_row, column=col,
                            value=str(field_value) if field_value else "")

        wb.save(str(filepath))
        wb.close()

    def save_extraction_results(
        self,
        progress: ProgressCallback | None = None,
    ) -> dict[str, str]:
        """Save filled templates from data/filled_templates/ to Exports/Excel/.

        Also fills the Stellenplan (instrument list) template from aggregated
        datasheet records, since individual TU PDFs are mapped to the
        Datasheet template but together they form the instrument list.
        """
        # Wire LLM for title block extraction
        from iev4pi_transformation_tool.core.standardized_export import set_title_block_llm
        set_title_block_llm(self.llm_client)
        # Clean up root-level stray .standardized.xlsx left by previous
        # broken runs — only subdirs (instrument_list/, wiring/, etc.)
        # should contain export files.
        _export_root = self.resolve_results_export_dir() / "Excel"
        if _export_root.is_dir():
            for _stray in _export_root.glob("*.standardized.xlsx"):
                _stray.unlink(missing_ok=True)

        # Fill Instrument List (Stellenplan) template from ALL datasheet records.
        # Individual TU PDFs map to Datasheet template, but together they form
        # the instrument list — so we aggregate them into the Stellenplan template here.
        from iev4pi_transformation_tool.models import DocumentFamily
        from iev4pi_transformation_tool.core.standardized_templates import (
            collect_filled_templates,
            TEMPLATE_TO_EXPORT_CATEGORY,
            get_export_category,
        )
        if not self.records:
            self.reload_records()
        # Stellenplan is now filled during fill_standardized_templates() —
        # no need to re-export here.  Just copy the already-filled templates.
        import openpyxl

        def _has_data(ws) -> bool:
            if ws.max_row is None or ws.max_row < 3:
                return False
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column or 1):
                if any(cell.value is not None and str(cell.value).strip() for cell in row):
                    return True
            return False

        def _clear_stale_sheets(ws) -> None:
            """Remove data rows (row 3+) from sheets that were not populated
            during *this* export session. If a sheet still has placeholder
            legend text in row 2, it wasn't touched → clear its old data."""
            if ws.max_row is None or ws.max_row < 3:
                return
            # If row 2 still has legend/template text, the sheet was never
            # written by a family exporter → all rows 3+ are stale.
            _SIGNALS = {"FK →", "1-based row index", "Fortlaufend", "Unique ",
                        "Fremdschlüssel", "e.g.", "Original AKZ", "Normalised canonical"}
            row2_text = " ".join(str(ws.cell(row=2, column=c).value or "") for c in range(1, (ws.max_column or 1) + 1))
            if any(s in row2_text for s in _SIGNALS):
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column or 1):
                    for cell in row:
                        cell.value = None

        export_base = ensure_dir(self.resolve_results_export_dir() / "Excel")
        filled_templates = collect_filled_templates()
        saved: dict[str, str] = {}
        total = max(1, len(filled_templates))
        idx = 0

        for template_name, src_path in filled_templates.items():
            category = get_export_category(template_name)
            if not category:
                continue

            # Clear stale data from sheets that weren't touched by this export.
            wb_rw = openpyxl.load_workbook(str(src_path))
            for ws in wb_rw.worksheets:
                _clear_stale_sheets(ws)
            wb_rw.save(str(src_path))

            # Check if any real data remains after stale clearing.
            wb_ro = openpyxl.load_workbook(str(src_path), read_only=True, data_only=True)
            has_any_data = any(_has_data(ws) for ws in wb_ro.worksheets)
            wb_ro.close()
            if not has_any_data:
                continue

            self._report_progress(
                progress,
                round(idx * 100 / total),
                f"Saving {category}",
            )
            dest_dir = ensure_dir(export_base / category)
            dest = dest_dir / template_name
            if src_path.resolve() != dest.resolve():
                shutil.copy2(str(src_path), str(dest))
            provenance_src = self._excel_provenance_path(src_path)
            if not provenance_src.is_file():
                self._write_excel_provenance(src_path, template_name)
            provenance_src = self._excel_provenance_path(src_path)
            if provenance_src.is_file():
                provenance_dest = self._excel_provenance_path(dest)
                shutil.copy2(str(provenance_src), str(provenance_dest))
            saved[category] = str(dest)
            idx += 1

        self._report_progress(progress, 100, f"Saved {len(saved)} result files")

        # Persist LLM/VLM caches to disk
        try:
            from iev4pi_transformation_tool.core import llm_cache
            llm_cache.save()
        except Exception:
            pass

        return saved

    def export_all(self, progress: ProgressCallback | None = None) -> list[Path]:
        return self.export_results(progress)

    def export_results(self, progress: ProgressCallback | None = None) -> list[Path]:
        # Stale-data clearing is now handled by _clear_data_rows() in each
        # exporter.  Canonical templates in data/templates/ are never modified
        # at runtime — filled results go to data/filled_templates/ and
        # Exports/Excel/.
        result_dir = ensure_dir(self.resolve_results_export_dir() / "Excel")
        if not self.records:
            self.reload_records()
        written: list[Path] = []
        non_ri_record_families = sorted(
            {
                record.family.value
                for record in self.records
                if not self._is_ri_family(record.family)
            }
        )
        ri_bundle_record_groups = self._ri_record_groups()
        total = sum(
            len(self.export_service.record_output_paths(self._result_family_dir(result_dir, family_name), family_name))
            for family_name in non_ri_record_families
        ) + sum(
            len(self.export_service.ri_record_output_paths(self._ri_result_dir(result_dir, bundle_id), bundle_name))
            for bundle_id, bundle_name in (
                (bundle_id, self.bundle_export_name_for_scope(bundle_id))
                for bundle_id in ri_bundle_record_groups["names"]
            )
        )
        total = max(1, total)
        completed = 0

        for family_name in non_ri_record_families:
            family_records = [record for record in self.records if record.family.value == family_name]
            family_dir = self._result_family_dir(result_dir, family_name)
            self.export_service.export_record_family(family_dir, family_name, family_records)
            output_paths = self.export_service.record_output_paths(family_dir, family_name)
            written.extend(output_paths)
            completed += len(output_paths)
            self._report_progress(progress, round(completed * 100 / total), f"Exported results {family_name}")

        for bundle_id, grouped_records in sorted(ri_bundle_record_groups["records"].items()):
            bundle_name = self.bundle_export_name_for_scope(bundle_id)
            bundle_dir = self._ri_result_dir(result_dir, bundle_id)
            self.export_service.export_ri_record_workbook(bundle_dir, bundle_name, grouped_records)
            output_paths = self.export_service.ri_record_output_paths(bundle_dir, bundle_name)
            written.extend(output_paths)
            completed += len(output_paths)
            self._report_progress(
                progress,
                round(completed * 100 / total),
                f"Exported R&I results {self.bundle_display_name_for_scope(bundle_id)}",
            )

        if not written:
            self._report_progress(progress, 100, "No results to export")
        else:
            self._report_progress(progress, 100, f"Result export complete: {len(written)} files")
        return written

    def apply_run_summary_payload(self, payload: dict) -> RunSummary:
        summary = RunSummary.model_validate(payload)
        self._latest_run_summary = summary
        self.records = []
        return summary

    def review_page(
        self,
        family: str | None = None,
        *,
        keyword: str | None = None,
        offset: int = 0,
        limit: int = 200,
    ) -> dict[str, object]:
        total_count = self.database.review_row_count(family, keyword=keyword)
        rows = self.database.review_rows_page(family, keyword=keyword, limit=limit, offset=offset)
        return {
            "family": family,
            "keyword": (keyword or "").strip(),
            "offset": max(0, int(offset)),
            "limit": max(1, int(limit)),
            "total_count": total_count,
            "rows": [row.model_dump(mode="json") for row in rows],
        }

    def review_record_page(
        self,
        family: str | None = None,
        *,
        keyword: str | None = None,
        offset: int = 0,
        limit: int = 100,
    ) -> dict[str, object]:
        family_enum, scope_id = self.parse_family_selection(family)
        family_value = family_enum.value if family_enum is not None else None
        total_count = self.database.review_record_count(family_value, scope_id=scope_id or None, keyword=keyword)
        rows = self.database.review_records_page(
            family_value,
            scope_id=scope_id or None,
            keyword=keyword,
            limit=limit,
            offset=offset,
        )
        effective_limit = total_count if int(limit) <= 0 else max(1, int(limit))
        return {
            "family": family,
            "keyword": (keyword or "").strip(),
            "offset": max(0, int(offset)),
            "limit": effective_limit,
            "total_count": total_count,
            "rows": [row.model_dump(mode="json") for row in rows],
        }

    def review_rows(self, family: str | None = None, keyword: str | None = None):
        return self.database.review_rows(family, keyword=keyword)

    def save_review_feedback(
        self,
        record_key: str,
        field_name: str,
        feedback_status: str,
        comment: str = "",
    ) -> object | None:
        return self.database.save_review_feedback(record_key, field_name, feedback_status, comment=comment)

    def llm_runtime_probe(self) -> dict[str, object]:
        return self.llm_client.runtime_probe()

    def latest_run_summary(self) -> RunSummary | None:
        if self._latest_run_summary is None:
            self._latest_run_summary = self.database.latest_run_summary()
        return self._latest_run_summary

    def ocr_runtime_status(self) -> dict[str, object]:
        return self.reader.describe_runtime()

    def surya_model_cache_status(self) -> dict[str, object]:
        return surya_prewarm_status(self.workspace_root)

    def prewarm_surya_models(self, progress: ProgressCallback | None = None) -> dict[str, object]:
        return prewarm_surya_models(self.workspace_root, progress)

    def all_schema_entries(self) -> dict[str, SchemaFamily]:
        entries = {family.value: schema for family, schema in self.schemas.items()}
        for bundle_id, bundle_schemas in self.ri_bundle_schemas.items():
            for family, schema in bundle_schemas.items():
                entries[self.schema_token_for(family, bundle_id)] = schema
        return entries

    def schema_selection_items(self) -> list[tuple[str, str]]:
        items: list[tuple[str, str]] = []
        for family, schema in sorted(self.schemas.items(), key=lambda item: item[0].value):
            source_root = schema.source_root or self._family_source_root(family)
            items.append((family.value, f"{source_root} / {family.value}"))
        for bundle_id, bundle_schemas in sorted(self.ri_bundle_schemas.items()):
            bundle_name = self.bundle_display_name_for_scope(bundle_id)
            for family, schema in sorted(bundle_schemas.items(), key=lambda item: item[0].value):
                source_root = schema.source_root or self._family_source_root(family)
                selection = self.schema_token_for(family, bundle_id)
                label = f"{source_root} / {bundle_name} / {family.value}"
                items.append((selection, label))
        return items

    def schema_by_selection(self, selection: str) -> SchemaFamily | None:
        family, scope_id = self.parse_family_selection(selection)
        if family is None:
            return None
        if scope_id:
            return self.ri_bundle_schemas.get(scope_id, {}).get(family)
        return self.schemas.get(family)

    def schema_token_for(self, family: DocumentFamily, scope_id: str = "") -> str:
        if scope_id:
            return f"ri|{scope_id}|{family.value}"
        return family.value

    def parse_family_selection(self, selection: str | None) -> tuple[DocumentFamily | None, str]:
        if not selection or selection == "__all__":
            return None, ""
        if selection.startswith("ri|"):
            _, scope_id, family_value = selection.split("|", 2)
            return DocumentFamily(family_value), scope_id
        if selection.startswith("ri::"):
            body = selection[len("ri::"):]
            scope_id, family_value = body.rsplit("::", 1)
            return DocumentFamily(family_value), scope_id
        return DocumentFamily(selection), ""

    def schema_field_names_for_family(self, family: DocumentFamily) -> list[str]:
        ordered: list[str] = []
        schema = self.schemas.get(family)
        if schema:
            for field in schema.fields:
                if field.name not in ordered:
                    ordered.append(field.name)
        for bundle_schemas in self.ri_bundle_schemas.values():
            scoped = bundle_schemas.get(family)
            if not scoped:
                continue
            for field in scoped.fields:
                if field.name not in ordered:
                    ordered.append(field.name)
        return ordered

    def schema_field_names_for_selection(self, selection: str | None) -> list[str]:
        family, scope_id = self.parse_family_selection(selection)
        if family is None:
            ordered: list[str] = []
            for selection_token, _label in self.review_selection_items():
                for field_name in self.schema_field_names_for_selection(selection_token):
                    if field_name not in ordered:
                        ordered.append(field_name)
            return ordered
        if scope_id:
            schema = self.ri_bundle_schemas.get(scope_id, {}).get(family)
            if schema:
                return [field.name for field in schema.fields]
        return self.schema_field_names_for_family(family)

    def review_selection_items(self) -> list[tuple[str, str]]:
        rows = self.database.latest_review_selection_rows()
        items: list[tuple[str, str]] = []
        for row in rows:
            family = DocumentFamily(str(row["family"]))
            scope_id = str(row["scope_id"] or "")
            selection = self.schema_token_for(family, scope_id)
            source_root = str(row["source_root"] or self._family_source_root(family))
            if scope_id:
                bundle_name = self.bundle_display_name_for_scope(scope_id)
                label = f"{source_root} / {bundle_name} / {family.value}"
            else:
                label = f"{source_root} / {family.value}"
            items.append((selection, label))
        return items

    def _record_result_map(self, record: ExtractedRecord) -> dict[str, str]:
        return {
            result.field_name: clean_cell(result.value)
            for result in record.results
            if clean_cell(result.field_name)
        }

    def _record_value(self, record: ExtractedRecord, *field_names: str) -> str:
        result_map = self._record_result_map(record)
        for field_name in field_names:
            value = clean_cell(result_map.get(field_name, ""))
            if value:
                return value
        return ""

    def _record_trace_metadata(self, record: ExtractedRecord) -> dict[str, str]:
        confidence_values = [
            float(result.decision_confidence)
            for result in record.results
            if result.decision_confidence is not None
        ]
        uncertainty_values = sorted(
            {
                clean_cell(result.uncertainty_reason)
                for result in record.results
                if clean_cell(result.uncertainty_reason)
            }
        )
        llm_status_values = sorted(
            {
                clean_cell(result.llm_verification_status)
                for result in record.results
                if clean_cell(result.llm_verification_status)
            }
        )
        rule_support_values = sorted(
            {
                clean_cell(rule)
                for result in record.results
                for rule in result.rule_support
                if clean_cell(rule)
            }
        )
        feedback_values = sorted(
            {
                clean_cell(result.review_feedback_status)
                for result in record.results
                if clean_cell(result.review_feedback_status)
            }
        )
        return {
            "decision_confidence": self._uc1_score_text(min(confidence_values), 0.0) if confidence_values else "",
            "evidence_bundle_id": next(
                (
                    clean_cell(result.evidence_bundle_id)
                    for result in record.results
                    if clean_cell(result.evidence_bundle_id)
                ),
                "",
            ),
            "uncertainty_reason": " | ".join(uncertainty_values),
            "llm_verification_status": " | ".join(llm_status_values),
            "rule_support": " | ".join(rule_support_values),
            "review_feedback_status": " | ".join(feedback_values),
            "decision_trace_json": json.dumps(record.decision_trace, ensure_ascii=False, sort_keys=True) if record.decision_trace else "",
        }

    def _strict_external_match_index(self) -> dict[str, dict[str, set[str]]]:
        index: dict[str, dict[str, set[str]]] = {
            "Stellenplaene": defaultdict(set),
            "Verschaltungslisten": defaultdict(set),
            "Gerätedatenblätter": defaultdict(set),
        }
        for record in self.records:
            source_root = record.source_root or self.family_source_root(record.family)
            if source_root == "Stellenplaene":
                candidate = self._record_value(record, "tag", "messstelle")
                normalized = normalize_identifier(candidate)
                if normalized:
                    index[source_root][normalized].add(record.record_key)
            elif source_root == "Verschaltungslisten":
                candidate = self._record_value(record, "plt_stelle")
                normalized = normalize_identifier(candidate)
                if normalized:
                    index[source_root][normalized].add(record.record_key)
            else:
                # Any other source (Gerätedatenblätter, etc.): index by tag + model
                for field in ("device", "device_id", "model", "tag"):
                    candidate = self._record_value(record, field)
                    norm = normalize_identifier(candidate)
                    if norm and len(norm) > 3:
                        index.setdefault(source_root, defaultdict(set))[norm].add(record.record_key)
        # Cross-populate: Gerätedatenblätter model numbers → Stellenplaene tags
        _datasheet_tags = index.get("Gerätedatenblätter", {})
        for norm_tag, rec_keys in _datasheet_tags.items():
            if norm_tag not in index["Stellenplaene"]:
                # Allow datasheet-match as evidence for instrument presence
                index["Stellenplaene"][norm_tag] = rec_keys.copy()
        return index

    def _uc1_source_prefixes(self, source_type: str) -> tuple[str, ...]:
        return {
            "instrument_list": ("Stellenplaene/", "Gerätedatenblätter/"),
            "datasheet": ("Stellenplaene/", "Gerätedatenblätter/"),
            "wiring": ("Verschaltungslisten/",),
            "piping": ("IFC/", "Piping Diagram/"),
        }.get(source_type, ())

    def _uc1_evidence_bundle_for_source(
        self,
        source_type: str,
        query: str,
        *,
        top_k: int = 8,
    ) -> EvidenceBundle:
        cache_key = (source_type, query)
        if cache_key in self._uc1_evidence_cache:
            return self._uc1_evidence_cache[cache_key]
        raw_hits = self.retriever.search(query, top_k=max(4, top_k * 4))
        prefixes = self._uc1_source_prefixes(source_type)
        filtered_hits = [
            hit
            for hit in raw_hits
            if not prefixes or any(hit.chunk.document_path.startswith(prefix) for prefix in prefixes)
        ][:top_k]
        digest = sha1(
            json.dumps(
                {
                    "source_type": source_type,
                    "query": query,
                    "hits": [hit.chunk.id for hit in filtered_hits],
                },
                ensure_ascii=False,
                sort_keys=True,
            ).encode("utf-8")
        ).hexdigest()[:16]
        bundle = EvidenceBundle(
            id=f"bundle:{digest}",
            query=query,
            hits=filtered_hits,
            support_evidence_ids=[hit.chunk.id for hit in filtered_hits if hit.score >= 0.45],
            contradiction_evidence_ids=[hit.chunk.id for hit in filtered_hits if hit.score <= 0.15],
            metadata={"source_type": source_type},
        )
        self._uc1_evidence_cache[cache_key] = bundle
        return bundle

    def _uc1_presence_decision(
        self,
        source_type: str,
        *,
        canonical_tag: str,
        display_name: str = "",
        strict_record_keys: list[str] | None = None,
        recommended_action: str = "",
        extra_query_terms: list[str] | None = None,
        skip_rag_if_no_strict: bool = False,
    ) -> ConsistencyDecision:
        strict_keys = [clean_cell(item) for item in (strict_record_keys or []) if clean_cell(item)]
        query_parts = [canonical_tag, display_name, *(extra_query_terms or [])]
        query = " ".join(part for part in query_parts if clean_cell(part)).strip()
        if strict_keys:
            return ConsistencyDecision(
                decision="present",
                canonical_entity_id=self._uc1_device_id(canonical_tag),
                support_evidence_ids=strict_keys,
                contradiction_evidence_ids=[],
                confidence=1.0,
                uncertainty_reason="",
                recommended_action=recommended_action,
                needs_review=False,
                evidence_bundle_id="",
                rule_support=["strict_external_match"],
                llm_verification_status="not_needed",
            )
        if skip_rag_if_no_strict and not strict_keys:
            return ConsistencyDecision(
                decision="missing",
                canonical_entity_id=self._uc1_device_id(canonical_tag),
                support_evidence_ids=[],
                contradiction_evidence_ids=[],
                confidence=0.9,
                uncertainty_reason="no_strict_index_match_skipped_rag",
                recommended_action=recommended_action,
                needs_review=False,
                evidence_bundle_id="",
                rule_support=["strict_index_missing"],
                llm_verification_status="not_needed",
            )
        evidence_bundle = self._uc1_evidence_bundle_for_source(source_type, query or canonical_tag)
        return self.evidence_resolver.verify_document_presence(
            source_type=source_type,
            canonical_entity_id=self._uc1_device_id(canonical_tag),
            query=query or canonical_tag,
            evidence_bundle=evidence_bundle,
            strict_present=False,
            recommended_action=recommended_action,
            rule_support=[],
        )

    def _uc1_ifc_presence_decision(
        self,
        *,
        canonical_tag: str,
        display_name: str = "",
        ifc_matches: list[dict[str, object]],
        recommended_action: str = "",
    ) -> ConsistencyDecision:
        query = " ".join(part for part in [canonical_tag, display_name, "ifc piping flange"] if clean_cell(part))
        if ifc_matches:
            return ConsistencyDecision(
                decision="present",
                canonical_entity_id=self._uc1_device_id(canonical_tag),
                support_evidence_ids=[clean_cell(match.get("record_key", "")) for match in ifc_matches if clean_cell(match.get("record_key", ""))],
                contradiction_evidence_ids=[],
                confidence=1.0,
                uncertainty_reason="",
                recommended_action=recommended_action,
                needs_review=False,
                evidence_bundle_id="",
                rule_support=["ifc_match_keys"],
                llm_verification_status="not_needed",
            )
        evidence_bundle = self._uc1_evidence_bundle_for_source("piping", query)
        return self.evidence_resolver.verify_document_presence(
            source_type="piping",
            canonical_entity_id=self._uc1_device_id(canonical_tag),
            query=query,
            evidence_bundle=evidence_bundle,
            strict_present=False,
            recommended_action=recommended_action,
            rule_support=[],
        )

    def _uc1_decision_status(self, decision: ConsistencyDecision) -> str:
        if decision.decision == "present":
            return "present"
        if decision.decision == "missing":
            return "missing"
        if decision.decision == "ambiguous":
            return "ambiguous"
        return "needs_review"

    def _ri_instrument_record_index(self) -> dict[str, dict[str, set[str]]]:
        index: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
        for record in self.records:
            if record.family != DocumentFamily.RI_INSTRUMENT_FUNCTION_ROW:
                continue
            canonical_tag = self._record_value(record, "canonical_tag", "tag_name", "name")
            normalized = normalize_identifier(canonical_tag or record.display_name)
            if not normalized:
                continue
            index[normalized]["record_keys"].add(record.record_key)
            evidence_types = {
                evidence.evidence_type
                for result in record.results
                for evidence in result.evidence_refs
            }
            if any(
                evidence_type in {"native_text", "ocr_text", "table_cell", "kv_pair"}
                for evidence_type in evidence_types
            ):
                index[normalized]["pdf_record_keys"].add(record.record_key)
            if any((evidence_type or "").startswith("dexpi_") for evidence_type in evidence_types):
                index[normalized]["xml_record_keys"].add(record.record_key)
        return index

    def _ifc_match_index(self) -> tuple[bool, dict[str, list[dict[str, object]]]]:
        if not self.documents:
            return False, {}
        ifc_documents = [document for document in self.documents if document.source_kind == SourceDocumentKind.IFC_MODEL]
        if not ifc_documents:
            return False, {}
        index: dict[str, list[dict[str, object]]] = defaultdict(list)
        for document in ifc_documents:
            parsed = self._parse_document(document)
            package = parsed.ifc_package
            if package is None:
                continue
            for node in package.ifc_nodes:
                for match_key in node.match_keys:
                    if not match_key:
                        continue
                    index[match_key].append(
                        {
                            "record_key": f"{document.relative_path}::ifc::{node.node_id}",
                            "node_id": node.node_id,
                            "display_name": node.tag or node.name or node.node_id,
                            "flange_complete": node.flange_complete,
                        }
                    )
        return True, index

    def _pid_instrument_xsd_status(self, context: dict[str, object] | None) -> str:
        if context is None:
            return "missing"
        parsed = context.get("parsed")
        if not isinstance(parsed, ParsedDocument) or parsed.ri_package is None:
            return "missing"
        package = parsed.ri_package
        if package.validation_errors:
            return "conflict"
        available = {field.category for field in package.xsd_field_defs if field.category}
        if "instrument_function" in available or "common" in available:
            return "present"
        return "missing"

    def _recommended_action(
        self,
        *,
        missing_in_stellenplan: bool,
        missing_in_verschaltung: bool,
        ifc_match_status: str,
        flange_status: str,
    ) -> str:
        language = self._ui_language_code()
        actions: list[str] = []
        if missing_in_stellenplan:
            actions.append(
                {
                    "en": "Complete Stellenplan",
                    "de": "Stellenplan ergaenzen",
                    "zh": "补全 Stellenplan",
                }[language]
            )
        if missing_in_verschaltung:
            actions.append(
                {
                    "en": "Complete Verschaltungsliste",
                    "de": "Verschaltungsliste ergaenzen",
                    "zh": "补全 Verschaltungsliste",
                }[language]
            )
        if ifc_match_status == "missing":
            actions.append(
                {
                    "en": "Add IFC object",
                    "de": "IFC-Objekt ergaenzen",
                    "zh": "补充 IFC 对象",
                }[language]
            )
        if flange_status == "missing":
            actions.append(
                {
                    "en": "Complete flange attributes",
                    "de": "Flanschattribute ergaenzen",
                    "zh": "补充法兰属性",
                }[language]
            )
        if not actions:
            return {
                "en": "No completion required",
                "de": "Keine Ergaenzung erforderlich",
                "zh": "无需补全",
            }[language]
        return {
            "en": "; ".join(actions),
            "de": "; ".join(actions),
            "zh": "，".join(actions),
        }[language]

    def _aas_generation_status_for_tag(self, normalized_tag: str) -> str:
        aas_dir = self.resolve_results_export_dir() / "AAS"
        if not aas_dir.exists():
            return "not_generated"
        found_formats: set[str] = set()
        for path in aas_dir.rglob("*"):
            if not path.is_file():
                continue
            if normalize_identifier(path.stem) != normalized_tag:
                continue
            found_formats.add(path.suffix.lower().lstrip("."))
        if not found_formats:
            return "not_generated"
        return "generated:" + ",".join(sorted(found_formats))

    def _process_cross_check_instance(
        self,
        scope_id: str,
        bundle_name: str,
        pdf_identifiers: dict[str, object],
        instance: object,
        canonical_tag: str,
        normalized_key: str,
        display_name: str,
        record_keys: dict,
        stellen_keys: list[str],
        versch_keys: list[str],
        ifc_matches: list[dict],
        ifc_loaded: bool,
        xsd_status: str,
        ri_record_index: dict,
        external_index: dict,
        ifc_index: dict,
    ) -> PidInconsistencyRow:
        """Process a single instrument instance for cross-checking (thread-safe)."""
        pdf_status = "present" if normalized_key in pdf_identifiers or record_keys.get("pdf_record_keys") else "missing"
        xml_status = "present"

        stellen_decision = self._uc1_presence_decision(
            "instrument_list",
            canonical_tag=canonical_tag,
            display_name=display_name,
            strict_record_keys=stellen_keys,
            extra_query_terms=["stellenplan", "instrument list"],
        )
        wiring_decision = self._uc1_presence_decision(
            "wiring",
            canonical_tag=canonical_tag,
            display_name=display_name,
            strict_record_keys=versch_keys,
            extra_query_terms=["wiring", "verschaltungsliste"],
            skip_rag_if_no_strict=True,
        )
        _has_asm_evidence = any(
            m.get("record_key", "").startswith("assembly_3d::") for m in ifc_matches
        ) if ifc_matches else False
        _has_special_status = any(
            m.get("_status") in ("software_control", "ifc_missing") for m in ifc_matches
        ) if ifc_matches else False

        if ifc_matches:
            ifc_decision = self._uc1_ifc_presence_decision(
                canonical_tag=canonical_tag,
                display_name=display_name,
                ifc_matches=ifc_matches,
            )
            ifc_status = self._uc1_decision_status(ifc_decision)
            if _has_special_status:
                ifc_status = "not_required" if any(
                    m.get("_status") == "software_control" for m in ifc_matches
                ) else "missing"
            if _has_asm_evidence:
                flange_status = "complete"
            elif not _has_special_status:
                flange_status = "complete" if any(match.get("flange_complete") is True for match in ifc_matches) else "missing"
            else:
                flange_status = "unknown" if ifc_status in {"ambiguous", "needs_review"} else "unknown"
        elif ifc_loaded:
            ifc_decision = ConsistencyDecision(
                decision="missing",
                canonical_entity_id=self._uc1_device_id(canonical_tag),
                confidence=0.9,
                uncertainty_reason="no_ifc_index_match_skipped_rag",
                recommended_action="",
                needs_review=False,
                evidence_bundle_id="",
                rule_support=["ifc_index_missing"],
                llm_verification_status="not_needed",
            )
            ifc_status = "missing"
            flange_status = "unknown"
        else:
            ifc_decision = ConsistencyDecision(
                decision="needs_review",
                canonical_entity_id=self._uc1_device_id(canonical_tag),
                confidence=0.0,
                uncertainty_reason="ifc_not_loaded",
                recommended_action="",
                needs_review=True,
                llm_verification_status="not_applicable",
            )
            ifc_status = "deferred"
            flange_status = "unknown"

        stellen_status = self._uc1_decision_status(stellen_decision)
        versch_status = self._uc1_decision_status(wiring_decision)
        missing_in_stellenplan = stellen_status != "present"
        missing_in_verschaltung = versch_status != "present"
        is_uc1_candidate = missing_in_stellenplan or missing_in_verschaltung
        _is_uy_tag = canonical_tag in ("TU10.U41", "TU20.U42")
        _eff_ifc_status = "not_required" if _is_uy_tag else ifc_status

        recommendation = self._recommended_action(
            missing_in_stellenplan=missing_in_stellenplan,
            missing_in_verschaltung=missing_in_verschaltung,
            ifc_match_status=_eff_ifc_status,
            flange_status=flange_status,
        )

        issues: list[str] = []
        if pdf_status != "present":
            issues.append("missing_pdf")
        if missing_in_stellenplan:
            issues.append("missing_stellenplaene")
        if missing_in_verschaltung:
            issues.append("missing_verschaltungslisten")
        if _eff_ifc_status == "missing":
            issues.append("missing_ifc")
        if _eff_ifc_status in {"ambiguous", "needs_review"}:
            issues.append("ifc_needs_review")
        if flange_status == "missing":
            issues.append("missing_flange")

        decision_trace = {
            "instrument_list": stellen_decision,
            "wiring": wiring_decision,
            "piping": ifc_decision,
        }
        confidence_values = [
            decision.confidence
            for decision in decision_trace.values()
            if decision.confidence > 0
        ]
        combined_rule_support = sorted(
            {
                rule
                for decision in decision_trace.values()
                for rule in decision.rule_support
                if clean_cell(rule)
            }
        )
        combined_uncertainty = " | ".join(
            sorted(
                {
                    clean_cell(decision.uncertainty_reason)
                    for decision in decision_trace.values()
                    if clean_cell(decision.uncertainty_reason)
                }
            )
        )
        combined_llm_status = " | ".join(
            sorted(
                {
                    clean_cell(decision.llm_verification_status)
                    for decision in decision_trace.values()
                    if clean_cell(decision.llm_verification_status)
                }
            )
        )
        primary_bundle_id = next(
            (
                decision.evidence_bundle_id
                for decision in decision_trace.values()
                if clean_cell(decision.evidence_bundle_id)
            ),
            "",
        )

        jump_targets = {
            "pdf": self._pid_jump_target(
                canonical_tag,
                preferred_source_root="R&I-Fließbild",
                preferred_scope_id=scope_id,
                matching_record_keys=sorted(record_keys.get("pdf_record_keys", set()) or record_keys.get("record_keys", set())),
                preferred_record_key=self._first_record_key(record_keys.get("pdf_record_keys", set()) or record_keys.get("record_keys", set())),
            ),
            "xml": self._pid_jump_target(
                canonical_tag,
                preferred_source_root="R&I-Fließbild",
                preferred_scope_id=scope_id,
                matching_record_keys=sorted(record_keys.get("xml_record_keys", set()) or record_keys.get("record_keys", set())),
                preferred_record_key=self._first_record_key(record_keys.get("xml_record_keys", set()) or record_keys.get("record_keys", set())),
            ),
            "stellenplaene": self._pid_jump_target(
                canonical_tag,
                preferred_source_root="Stellenplaene",
                matching_record_keys=stellen_keys,
                preferred_record_key=self._first_record_key(stellen_keys),
            ),
            "verschaltungslisten": self._pid_jump_target(
                canonical_tag,
                preferred_source_root="Verschaltungslisten",
                matching_record_keys=versch_keys,
                preferred_record_key=self._first_record_key(versch_keys),
            ),
            "ifc": self._pid_jump_target(
                canonical_tag,
                preferred_source_root="IFC",
                matching_record_keys=[match["record_key"] for match in ifc_matches],
                preferred_record_key=self._first_record_key([match["record_key"] for match in ifc_matches]),
            ),
        }

        return PidInconsistencyRow(
            component_key=canonical_tag,
            display_name=display_name,
            normalized_key=normalized_key,
            canonical_tag=canonical_tag,
            primary_type=instance.function_code or "instrument_function",
            pdf_status=pdf_status,
            xml_status=xml_status,
            xsd_status=xsd_status,
            stellenplaene_status=stellen_status,
            verschaltungslisten_status=versch_status,
            missing_in_stellenplan=missing_in_stellenplan,
            missing_in_verschaltung=missing_in_verschaltung,
            ifc_match_status=ifc_status,
            flange_status=flange_status,
            ifc_match_key=ifc_matches[0]["node_id"] if ifc_matches else "",
            context_summary=instance.context_summary,
            proposal_status="ready" if is_uc1_candidate else "not_required",
            aas_generation_status=self._aas_generation_status_for_tag(normalized_key),
            recommended_action=recommendation,
            is_uc1_candidate=is_uc1_candidate,
            issue_count=len(issues),
            issues=issues,
            jump_targets=jump_targets,
            scope_id=scope_id,
            decision_confidence=min(confidence_values) if confidence_values else 0.0,
            evidence_bundle_id=primary_bundle_id,
            uncertainty_reason=combined_uncertainty,
            llm_verification_status=combined_llm_status,
            rule_support=combined_rule_support,
            decision_trace=decision_trace,
        )

    def pid_inconsistency_report(self, progress: ProgressCallback | None = None) -> PidInconsistencySummary:
        self._uc1_evidence_cache.clear()
        if not self.documents:
            self.scan(self._sub_progress(progress, 0, 15) if progress else None)
        if not self.schemas and not self.ri_bundle_schemas:
            self.reload_schemas()
        if not self.records:
            self.reload_records()

        if not self.ri_bundles:
            self._report_progress(progress, 100, "No R&I bundle data available")
            return PidInconsistencySummary(empty_reason="no_ri_data")

        self._report_progress(progress, 20, "Collecting latest extracted records")
        bundle_contexts = self._pid_bundle_contexts(progress)
        external_index = self._strict_external_match_index()
        ri_record_index = self._ri_instrument_record_index()
        ifc_loaded, ifc_index = self._ifc_match_index()
        # Enrich with Assembly/mapping evidence — special statuses WINS over raw IFC
        assembly_evidence = self._uc1_assembly_ifc_evidence_index()
        for akz_tag, ev in assembly_evidence.items():
            normalized = normalize_identifier(akz_tag)
            if not normalized:
                continue
            status = ev.get("_status", "")
            if status == "software_control_no_ifc_required":
                ifc_index[normalized] = [{"node_id": "", "record_key": f"mapping::{akz_tag}",
                    "_status": "software_control", "global_id": "", "tag": ""}]
            elif status == "ifc_missing":
                ifc_index[normalized] = [{"node_id": "", "record_key": f"mapping::{akz_tag}",
                    "_status": "ifc_missing", "global_id": "", "tag": ""}]
            elif ev.get("global_id") and not ifc_index.get(normalized):
                ifc_index[normalized] = [{
                    "node_id": ev.get("global_id", ""),
                    "record_key": f"assembly_3d::{ev.get('label', akz_tag)}",
                    "ifc_class": "IfcBuildingElementProxy",
                    "global_id": ev.get("global_id", ""),
                    "tag": ev.get("label", ""),
                    "display_name": ev.get("label", akz_tag),
                    "source_doc_id": f"urn:ievpi:document:assembly3d:{ev.get('label', '')}",
                    "source_locator": f"Assembly_Steps::{ev.get('label', '')}",
                }]

        # Count total instances for progress reporting
        total_instances = 0
        bundle_instance_counts: dict[str, int] = {}
        for scope_id, context in bundle_contexts.items():
            parsed = context.get("parsed")
            if isinstance(parsed, ParsedDocument) and parsed.ri_package is not None:
                count = len(parsed.ri_package.instrument_instances)
                bundle_instance_counts[scope_id] = count
                total_instances += count
        total_instances = max(1, total_instances)

        self._report_progress(progress, 55, f"Checking cross references ({total_instances} instruments, 6 workers)")
        rows: list[PidInconsistencyRow] = []
        all_pdf_identifiers: set[str] = set()

        # --- Build work items (pre-compute before parallel dispatch) ---
        work_items: list[dict[str, object]] = []
        for scope_id, context in sorted(bundle_contexts.items(), key=lambda item: self.bundle_display_name_for_scope(item[0])):
            parsed = context.get("parsed")
            if not isinstance(parsed, ParsedDocument) or parsed.ri_package is None:
                continue
            package = parsed.ri_package
            bundle_name = self.bundle_display_name_for_scope(scope_id)
            pdf_identifiers = context.get("pdf_identifiers", {})
            all_pdf_identifiers.update(pdf_identifiers.keys())
            xsd_status = self._pid_instrument_xsd_status(context)
            for instance in package.instrument_instances:
                canonical_tag = clean_cell(instance.canonical_tag)
                normalized_key = normalize_identifier(canonical_tag)
                if not normalized_key:
                    continue
                display_name = instance.label_text or canonical_tag
                record_keys = ri_record_index.get(normalized_key, {})
                stellen_keys = sorted(external_index["Stellenplaene"].get(normalized_key, set()))
                # Merge Gerätedatenblätter matches into Stellenplan evidence
                _ds_keys = external_index.get("Gerätedatenblätter", {}).get(normalized_key, set())
                if _ds_keys:
                    stellen_keys = sorted(set(stellen_keys) | _ds_keys)
                versch_keys = sorted(external_index["Verschaltungslisten"].get(normalized_key, set()))
                ifc_matches = ifc_index.get(normalized_key, [])
                work_items.append({
                    "scope_id": scope_id,
                    "bundle_name": bundle_name,
                    "pdf_identifiers": pdf_identifiers,
                    "instance": instance,
                    "canonical_tag": canonical_tag,
                    "normalized_key": normalized_key,
                    "display_name": display_name,
                    "record_keys": record_keys,
                    "stellen_keys": stellen_keys,
                    "versch_keys": versch_keys,
                    "ifc_matches": ifc_matches,
                    "xsd_status": xsd_status,
                })

        # --- Parallel processing ---
        from concurrent.futures import as_completed
        import threading
        from collections import OrderedDict

        processed_lock = threading.Lock()
        processed_count = [0]  # mutable container for closure

        def _process_one(item: dict) -> PidInconsistencyRow:
            row = self._process_cross_check_instance(
                scope_id=str(item["scope_id"]),
                bundle_name=str(item["bundle_name"]),
                pdf_identifiers=item["pdf_identifiers"],
                instance=item["instance"],
                canonical_tag=str(item["canonical_tag"]),
                normalized_key=str(item["normalized_key"]),
                display_name=str(item["display_name"]),
                record_keys=item["record_keys"],
                stellen_keys=item["stellen_keys"],
                versch_keys=item["versch_keys"],
                ifc_matches=item["ifc_matches"],
                ifc_loaded=ifc_loaded,
                xsd_status=str(item["xsd_status"]),
                ri_record_index=ri_record_index,
                external_index=external_index,
                ifc_index=ifc_index,
            )
            with processed_lock:
                processed_count[0] += 1
                if processed_count[0] % 10 == 0 or processed_count[0] == total_instances:
                    pct = 55 + round(processed_count[0] * 40 / total_instances)
                    self._report_progress(
                        progress, pct,
                        f"Checking cross references ({processed_count[0]}/{total_instances})",
                    )
            return row

        # Preserve original ordering: assign an index to each work item
        for i, item in enumerate(work_items):
            item["_idx"] = i

        results: dict[int, PidInconsistencyRow] = {}
        from iev4pi_transformation_tool.core.qos_helpers import io_worker_count, QoSAwareThreadPoolExecutor

        configured = getattr(self.settings.llm, "parallel_workers", 0) or 0
        _workers = max(4, configured) if configured > 0 else io_worker_count(cap=8)
        with QoSAwareThreadPoolExecutor(max_workers=_workers) as executor:
            futures = {executor.submit(_process_one, item): item for item in work_items}
            for future in as_completed(futures):
                item = futures[future]
                idx = item["_idx"]
                try:
                    results[idx] = future.result()
                except Exception:
                    pass  # skip individual failures

        # Reconstruct ordered rows
        rows = [results[i] for i in sorted(results)]

        # Inject P&ID component rows (VV, PL, HE) with real status from all sources
        self._uc1_inject_component_rows(rows, external_index, all_pdf_identifiers)

        summary = PidInconsistencySummary(
            total_components=len(rows),
            problem_component_count=sum(1 for row in rows if row.issue_count > 0),
            problem_item_count=sum(row.issue_count for row in rows),
            uc1_candidate_count=sum(1 for row in rows if row.is_uc1_candidate),
            rows=rows,
            empty_reason="" if rows else "no_ri_data",
        )
        self._report_progress(progress, 100, f"Built P&ID inconsistency report ({summary.total_components} components)")
        return summary

    def pid_jump_payload(self, component_key: str, source_column: str, scope_id: str = "") -> dict[str, object]:
        report = self.pid_inconsistency_report()
        normalized_key = normalize_identifier(component_key) or normalize_label(component_key)
        for row in report.rows:
            if row.normalized_key != normalized_key:
                continue
            if scope_id and row.scope_id != scope_id:
                continue
            target = row.jump_targets.get(source_column)
            if target is not None:
                return target.model_dump(mode="json")
        return {}

    def use_case_1_report(self, progress: ProgressCallback | None = None) -> PidInconsistencySummary:
        report = self.pid_inconsistency_report(progress)
        uc1_rows = [row for row in report.rows if row.is_uc1_candidate]
        return PidInconsistencySummary(
            total_components=len(uc1_rows),
            problem_component_count=sum(1 for row in uc1_rows if row.issue_count > 0),
            problem_item_count=sum(row.issue_count for row in uc1_rows),
            uc1_candidate_count=len(uc1_rows),
            rows=uc1_rows,
            empty_reason="" if uc1_rows else report.empty_reason,
        )

    def export_use_case_1_workbook(
        self,
        progress: ProgressCallback | None = None,
        *,
        output_dir: Path | None = None,
    ) -> Path:
        report = self.use_case_1_report(progress)
        target_dir = output_dir or ensure_dir(self.resolve_results_export_dir() / "UseCase1")
        overview_rows = [
            {
                "canonical_tag": row.canonical_tag,
                "display_name": row.display_name,
                "context_summary": row.context_summary,
                "stellenplaene_status": row.stellenplaene_status,
                "verschaltungslisten_status": row.verschaltungslisten_status,
                "ifc_match_status": row.ifc_match_status,
                "flange_status": row.flange_status,
                "recommended_action": row.recommended_action,
                "source_row_key": row.component_key,
                "scope_id": row.scope_id,
            }
            for row in report.rows
        ]
        stellenplaene_rows = [
            {
                "source_row_key": row.component_key,
                "canonical_tag": row.canonical_tag,
                "tag": row.canonical_tag,
                "function_code": row.primary_type,
                "context_summary": row.context_summary,
                "recommended_action": row.recommended_action,
            }
            for row in report.rows
            if row.missing_in_stellenplan
        ]
        verschaltung_rows = [
            {
                "source_row_key": row.component_key,
                "canonical_tag": row.canonical_tag,
                "plt_stelle": row.canonical_tag,
                "funktion": row.primary_type,
                "beschreibung": row.context_summary,
                "context_summary": row.context_summary,
                "recommended_action": row.recommended_action,
            }
            for row in report.rows
            if row.missing_in_verschaltung
        ]
        ifc_rows = [
            {
                "source_row_key": row.component_key,
                "canonical_tag": row.canonical_tag,
                "ifc_match_status": row.ifc_match_status,
                "ifc_match_key": row.ifc_match_key,
                "flange_status": row.flange_status,
                "context_summary": row.context_summary,
                "recommended_action": row.recommended_action,
            }
            for row in report.rows
        ]
        path = self.export_service.export_use_case_1_workbook(
            target_dir,
            overview_rows=overview_rows,
            stellenplaene_rows=stellenplaene_rows,
            verschaltung_rows=verschaltung_rows,
            ifc_rows=ifc_rows,
            language=self._ui_language_code(),
        )
        self._report_progress(progress, 100, f"Exported UC1 workbook to {path.name}")
        return path

    def uc1_catalog_path(self) -> Path:
        return self.workspace_root / "data" / "Datenpunkte_V1.xlsx"

    def uc1_catalog_coverage_report(self) -> UC1CatalogCoverageReport:
        catalog_path = self.uc1_catalog_path()
        return self.uc1_catalog_service.coverage_report(catalog_path, UC1_SUPPORT_MAP)

    def tx_rules_dir(self) -> Path:
        return self.tx_rule_store.root_dir

    def t1_t5_rules_dir(self) -> Path:
        return self.t1_t5_rule_store.root_dir

    def _use_saved_t1_t5_rules(self) -> bool:
        return bool(self.settings.use_custom_t1_t5_rules)

    def _use_saved_tx_rules(self) -> bool:
        return bool(self.settings.use_custom_tx_rules)

    def load_t1_t5_rule_bundle(
        self,
        stage_id: str,
        *,
        rule_path: Path | None = None,
        allow_saved_rules: bool = True,
    ) -> T1T5RuleBundle:
        explicit_rule_request = rule_path is not None
        if not allow_saved_rules and not explicit_rule_request:
            return build_default_t1_t5_bundle(stage_id)
        stored = self.t1_t5_rule_store.load_if_available(stage_id=stage_id, rule_path=rule_path)
        if stored is not None:
            return stored
        return build_default_t1_t5_bundle(stage_id)

    def validate_t1_t5_rules(
        self,
        stage_id: str,
        bundle_payload: dict[str, object],
    ) -> dict[str, object]:
        bundle = T1T5RuleBundle.model_validate(bundle_payload)
        issues = self.t1_t5_executor.validate_bundle(bundle)
        return {
            "bundle": bundle.model_dump(mode="json"),
            "issues": [issue.model_dump(mode="json") for issue in issues],
            "valid": not any(issue.severity == "error" for issue in issues),
        }

    def save_t1_t5_rules(
        self,
        stage_id: str,
        bundle_payload: dict[str, object],
        *,
        output_path: Path | None = None,
    ) -> dict[str, object]:
        bundle = T1T5RuleBundle.model_validate(bundle_payload)
        issues = self.t1_t5_executor.validate_bundle(bundle)
        if any(issue.severity == "error" for issue in issues):
            return {
                "rule_path": "",
                "bundle": bundle.model_dump(mode="json"),
                "issues": [issue.model_dump(mode="json") for issue in issues],
                "saved": False,
            }
        path = self.t1_t5_rule_store.save(bundle, rule_path=output_path)
        return {
            "rule_path": str(path),
            "bundle": bundle.model_dump(mode="json"),
            "issues": [issue.model_dump(mode="json") for issue in issues],
            "saved": True,
        }

    def resolve_t1_t5_profile(
        self,
        stage_id: str,
        *,
        workbook_path: Path | None = None,
        bundle_payload: dict[str, object] | None = None,
        profile_id: str = "",
    ) -> dict[str, object]:
        bundle = T1T5RuleBundle.model_validate(bundle_payload) if bundle_payload else self.load_t1_t5_rule_bundle(
            stage_id,
            allow_saved_rules=self._use_saved_t1_t5_rules(),
        )
        workbook_sheets = self._uc1_load_excel_rows(workbook_path) if workbook_path is not None else None
        profile, profile_match = self.t1_t5_executor.resolve_profile(
            bundle,
            workbook_sheets=workbook_sheets,
            requested_profile_id=profile_id,
        )
        return {
            "profile_id": profile.profile_id if profile is not None else "",
            "profile_match": profile_match.model_dump(mode="json") if profile_match is not None else {},
            "score": profile_match.score if profile_match is not None else 0.0,
            "matched_sheet_name": profile_match.matched_sheet_name if profile_match is not None else "",
        }

    def preview_t1_t5_rules(
        self,
        stage_id: str,
        *,
        workbook_path: Path | None = None,
        profile_id: str = "",
        bundle_payload: dict[str, object] | None = None,
    ) -> dict[str, object]:
        bundle = T1T5RuleBundle.model_validate(bundle_payload) if bundle_payload else self.load_t1_t5_rule_bundle(
            stage_id,
            allow_saved_rules=self._use_saved_t1_t5_rules(),
        )
        workbook_sheets = self._uc1_load_excel_rows(workbook_path) if workbook_path is not None else None
        input_rows: list[dict[str, str]] = []
        if workbook_sheets is None or not clean_cell(profile_id):
            input_rows = self._legacy_t1_t5_preview_rows(stage_id)
        else:
            resolved_profile, _profile_match = self.t1_t5_executor.resolve_profile(
                bundle,
                workbook_sheets=workbook_sheets,
                requested_profile_id=profile_id,
            )
            if resolved_profile is None or resolved_profile.input_mode != "custom_workbook":
                input_rows = self._legacy_t1_t5_preview_rows(stage_id)
        preview = self.t1_t5_executor.preview(
            bundle,
            workbook_path=workbook_path,
            workbook_sheets=workbook_sheets,
            input_rows=input_rows,
            requested_profile_id=profile_id,
        )
        return preview.model_dump(mode="json")

    def load_tx_rule_set(
        self,
        source_type: str,
        *,
        tx_rule_path: Path | None = None,
        tx_rule_set_id: str = "",
        allow_saved_rules: bool = True,
    ) -> TxRuleSet:
        explicit_rule_request = tx_rule_path is not None or bool(clean_cell(tx_rule_set_id))
        if not allow_saved_rules and not explicit_rule_request:
            return build_default_uc1_rule_set(source_type)
        stored = self.tx_rule_store.load_if_available(
            source_type=source_type,
            rule_path=tx_rule_path,
            rule_set_id=tx_rule_set_id,
        )
        if stored is not None:
            return stored
        return build_default_uc1_rule_set(source_type)

    def validate_tx_rules(
        self,
        rule_payload: dict[str, object],
    ) -> dict[str, object]:
        rule_set = TxRuleSet.model_validate(rule_payload)
        issues = self.tx_executor.validate(rule_set)
        return {
            "rule_set": rule_set.model_dump(mode="json"),
            "issues": [issue.model_dump(mode="json") for issue in issues],
            "valid": not any(issue.severity == "error" for issue in issues),
        }

    def save_tx_rules(
        self,
        rule_payload: dict[str, object],
        *,
        tx_rule_set_id: str = "",
        output_path: Path | None = None,
    ) -> dict[str, object]:
        rule_set = TxRuleSet.model_validate(rule_payload)
        issues = self.tx_executor.validate(rule_set)
        if any(issue.severity == "error" for issue in issues):
            return {
                "rule_path": "",
                "rule_set": rule_set.model_dump(mode="json"),
                "issues": [issue.model_dump(mode="json") for issue in issues],
                "saved": False,
            }
        path = self.tx_rule_store.save(rule_set, rule_path=output_path, rule_set_id=tx_rule_set_id)
        return {
            "rule_path": str(path),
            "rule_set": rule_set.model_dump(mode="json"),
            "issues": [issue.model_dump(mode="json") for issue in issues],
            "saved": True,
        }

    def preview_tx_rules(
        self,
        source_type: str,
        workbook_path: Path,
        *,
        identity_key: str = "",
        tx_rule_path: Path | None = None,
        tx_rule_set_id: str = "",
        rule_payload: dict[str, object] | None = None,
    ) -> dict[str, object]:
        rule_set = TxRuleSet.model_validate(rule_payload) if rule_payload else self.load_tx_rule_set(
            source_type,
            tx_rule_path=tx_rule_path,
            tx_rule_set_id=tx_rule_set_id,
            allow_saved_rules=self._use_saved_tx_rules(),
        )
        sheets = self._uc1_load_excel_rows(workbook_path)
        primary_sheet = clean_cell(rule_set.primary_sheet_name) or self._uc1_primary_sheet_name(source_type)
        grouped = self._uc1_group_rows_by_identity(sheets.get(primary_sheet, []))
        if identity_key and identity_key in grouped:
            chosen_identity = identity_key
        else:
            chosen_identity = next(iter(grouped.keys()), "")
        rows = grouped.get(chosen_identity, [])
        preview = self.tx_executor.preview(
            rule_set,
            rows,
            identity_value=chosen_identity,
            workbook_path=workbook_path,
            source_type=source_type,
        )
        return preview.model_dump(mode="json")

    def suggest_tx_rules(
        self,
        source_type: str,
        workbook_path: Path,
        *,
        target_properties: dict[str, list[str]] | None = None,
    ) -> dict[str, object]:
        suggestion = self.tx_rule_suggester.suggest(
            source_type,
            workbook_path,
            target_properties=target_properties,
        )
        return suggestion.model_dump(mode="json")

    def generate_uc1_aas_from_tx(
        self,
        source_type: str,
        workbook_path: Path,
        progress: ProgressCallback | None = None,
        *,
        output_dir: Path | None = None,
        target_formats: list[str] | None = None,
        tx_rule_path: Path | None = None,
        tx_rule_set_id: str = "",
        rule_payload: dict[str, object] | None = None,
    ) -> dict[str, object]:
        formats = target_formats or ["json", "xml"]
        rule_set = TxRuleSet.model_validate(rule_payload) if rule_payload else self.load_tx_rule_set(
            source_type,
            tx_rule_path=tx_rule_path,
            tx_rule_set_id=tx_rule_set_id,
            allow_saved_rules=self._use_saved_tx_rules(),
        )
        workbook_rows = self._uc1_load_excel_rows(workbook_path)
        grouped_rows = self._uc1_group_rows_by_identity(
            workbook_rows.get(clean_cell(rule_set.primary_sheet_name) or self._uc1_primary_sheet_name(source_type), [])
        )
        output_root = ensure_dir(output_dir or (self.resolve_results_export_dir() / "AAS" / "tx" / source_type))
        generated_paths: list[str] = []
        traces_by_identity: dict[str, list[dict[str, object]]] = {}
        total = max(1, len(grouped_rows))
        for index, (identity_key, rows) in enumerate(sorted(grouped_rows.items()), start=1):
            payload, traces, issues = self._uc1_build_tx_payload(
                source_type,
                rows,
                identity_key,
                workbook_path=workbook_path,
                tx_rule_path=tx_rule_path,
                tx_rule_set_id=tx_rule_set_id,
                rule_payload=rule_set.model_dump(mode="json"),
            )
            traces_by_identity[identity_key] = [trace.model_dump(mode="json") for trace in traces]
            for path in self._uc1_write_payload_formats(
                payload,
                output_root,
                self._uc1_payload_identity(payload),
                formats,
            ):
                generated_paths.append(str(path))
            self._report_progress(progress, round(index * 100 / total), f"Generated Tx AAS for {identity_key}")
            if any(issue.severity == "error" for issue in issues):
                self.log_debug(
                    source="tx",
                    action="generate_uc1_aas_from_tx",
                    message=f"Fallback used for {source_type}:{identity_key}",
                    level="WARNING",
                    details={"identity_key": identity_key, "issues": [issue.model_dump(mode="json") for issue in issues]},
                )
        return {
            "source_type": source_type,
            "generated_paths": generated_paths,
            "trace_count": sum(len(items) for items in traces_by_identity.values()),
            "traces_by_identity": traces_by_identity,
        }

    def _uc1_build_source_rows(
        self,
        progress: ProgressCallback | None = None,
    ) -> dict[str, object]:
        """Build enriched source rows from DEXPI analysis + cross-document report.

        Returns a dict with keys: pid, instrument_list, wiring, datasheet
        (each a list of row dicts), plus completion_map and aggregated_sheets.
        No Excel files are written.
        """
        if not self.documents:
            self.scan(self._sub_progress(progress, 0, 10) if progress else None)
        if not self.schemas and not self.ri_bundle_schemas:
            self.reload_schemas()
        if not self.records:
            self.reload_records()

        report = self.pid_inconsistency_report(self._sub_progress(progress, 10, 40) if progress else None)
        bundle_contexts = self._pid_bundle_contexts(self._sub_progress(progress, 40, 55) if progress else None)
        strict_index = self._strict_external_match_index()
        coverage = self.uc1_catalog_coverage_report()
        record_lookup = {record.record_key: record for record in self.records}
        ri_instance_index: dict[tuple[str, str], dict[str, object]] = {}
        for scope_id, context in bundle_contexts.items():
            parsed = context.get("parsed")
            if not isinstance(parsed, ParsedDocument) or parsed.ri_package is None:
                continue
            node_lookup = {node.node_id: node for node in parsed.ri_package.xml_nodes}
            for instance in parsed.ri_package.instrument_instances:
                normalized_key = normalize_identifier(instance.canonical_tag)
                if normalized_key:
                    ri_instance_index[(scope_id, normalized_key)] = {
                        "instance": instance,
                        "node_lookup": node_lookup,
                    }

        ifc_details: dict[str, list[dict[str, str]]] = {}
        if any(doc.source_kind == SourceDocumentKind.IFC_MODEL for doc in self.documents):
            ifc_details = self._uc1_ifc_details_index()
        documents_rows = [
            {
                "doc_id": self._uc1_document_id(document.relative_path, document.source_root),
                "source_family": document.source_kind.value,
                "original_path": document.relative_path,
                "format": document.extension.lstrip("."),
                "bundle_id": document.bundle_id or "",
                "source_root": document.source_root,
                "revision": "",
                "parser_status": "indexed",
                "unit_count": len(document.output_families),
            }
            for document in sorted(self.documents, key=lambda item: item.relative_path)
        ]

        ri_device_rows: list[dict[str, object]] = []
        stellenplan_rows: list[dict[str, object]] = []
        wiring_rows: list[dict[str, object]] = []
        datasheet_rows: list[dict[str, object]] = []
        ifc_rows: list[dict[str, object]] = []
        relation_rows: list[dict[str, object]] = []
        completion_rows: list[dict[str, object]] = []

        enrichment_total = max(1, len(report.rows))
        enrichment_processed = 0
        for row in sorted(report.rows, key=lambda item: (item.scope_id, item.canonical_tag)):
            enrichment_processed += 1
            enrichment_pct = 55 + round(enrichment_processed * 40 / enrichment_total)
            bundle_display = self.bundle_display_name_for_scope(row.scope_id)
            self._report_progress(
                progress, enrichment_pct,
                f"Building enriched data: {bundle_display}/{row.canonical_tag} ({enrichment_processed}/{enrichment_total})",
            )
            device_id = self._uc1_device_id(row.canonical_tag)
            ri_context = ri_instance_index.get((row.scope_id, row.normalized_key), {})
            instance = ri_context.get("instance")
            node_lookup = ri_context.get("node_lookup", {})
            loop_node = node_lookup.get(getattr(instance, "loop_node_id", ""), None) if isinstance(node_lookup, dict) else None
            function_node = node_lookup.get(getattr(instance, "function_node_id", ""), None) if isinstance(node_lookup, dict) else None
            anchor_node = node_lookup.get(getattr(instance, "piping_anchor_id", ""), None) if isinstance(node_lookup, dict) else None
            ri_doc_id = self._uc1_document_id(f"{row.scope_id}:{row.canonical_tag}", "R&I-Fließbild")
            ri_locator = self._uc1_evidence_locator(getattr(instance, "evidence_refs", []))

            function_code = clean_cell(getattr(instance, "function_code", "")) or row.primary_type
            process_number = row.canonical_tag
            process_category = function_code[:1] if function_code else ""
            process_modifier = function_code[1:] if len(function_code) > 1 else ""
            anchor_attributes = getattr(anchor_node, "attributes", {}) if anchor_node is not None else {}
            function_attributes = getattr(function_node, "attributes", {}) if function_node is not None else {}
            loop_attributes = getattr(loop_node, "attributes", {}) if loop_node is not None else {}
            missing_targets: list[str] = []
            if row.stellenplaene_status != "present":
                missing_targets.append("stellenplan")
            if row.verschaltungslisten_status != "present":
                missing_targets.append("wiring")
            _is_uy = (function_code == "UY" or row.canonical_tag in ("TU10.U41", "TU20.U42"))
            if row.ifc_match_status == "not_required" or _is_uy:
                pass  # software control — not a true IFC gap
            elif row.ifc_match_status != "present":
                missing_targets.append("ifc")
            if row.flange_status == "missing" and not _is_uy:
                missing_targets.append("ifc_flange")
            completion_context = {
                "recommended_action": row.recommended_action,
                "proposal_status": row.proposal_status,
                "missing_targets": " | ".join(missing_targets),
                "present_in_stellenplan": "true" if row.stellenplaene_status == "present" else "false",
                "present_in_wiring": "true" if row.verschaltungslisten_status == "present" else "false",
                "present_in_ifc": "true" if row.ifc_match_status == "present" else ("not_required" if _is_uy else "false"),
                "flange_complete": "true" if row.flange_status == "complete" else "false",
                "decision_confidence": self._uc1_score_text(row.decision_confidence, 0.0),
                "evidence_bundle_id": clean_cell(row.evidence_bundle_id),
                "uncertainty_reason": clean_cell(row.uncertainty_reason),
                "llm_verification_status": clean_cell(row.llm_verification_status),
                "rule_support": " | ".join(clean_cell(rule) for rule in row.rule_support if clean_cell(rule)),
                "review_feedback_status": clean_cell(row.review_feedback_status),
                "decision_trace_json": self._uc1_decision_trace_json(row.decision_trace),
            }

            ri_device_rows.append(
                self._uc1_finalize_standardized_row(
                    "pid",
                    {
                        "device_id": device_id,
                        "canonical_tag": row.canonical_tag,
                        "class_name": "ProcessInstrumentationFunction",
                        "has_instrumentation_loop_function_number": row.canonical_tag,
                        "process_instrumentation_function_number": process_number,
                        "process_instrumentation_function_category": process_category,
                        "process_instrumentation_function_modifier": process_modifier,
                        "process_instrumentation_functions": function_code,
                        "device_information": self._uc1_first_value(
                            getattr(instance, "description", ""),
                            self._uc1_attr_value(loop_attributes, "description"),
                            self._uc1_attr_value(function_attributes, "description"),
                            row.display_name,
                        ),
                        "vendor_company_name": self._uc1_attr_value(loop_attributes, "vendor_company_name", "vendor"),
                        "safety_relevance_class": self._uc1_attr_value(
                            function_attributes,
                            "safety_relevance_class",
                            "safety_class",
                        ),
                        "actuating_function_number": self._uc1_attr_value(
                            function_attributes,
                            "actuating_function_number",
                            "actuatingfunctionnumber",
                        ),
                        "actuating_location": self._uc1_first_value(
                            getattr(instance, "piping_anchor_id", ""),
                            getattr(anchor_node, "tag_name", "") if anchor_node is not None else "",
                        ),
                        "actuating_system_number": self._uc1_attr_value(
                            anchor_attributes,
                            "actuating_system_number",
                            "system_number",
                        ),
                        "operated_valve_reference": self._uc1_attr_value(
                            anchor_attributes,
                            "operated_valve_reference",
                            "operatedvalvereference",
                            "sub_tag_name",
                        ),
                        "flow_direction": self._uc1_attr_value(anchor_attributes, "flow_direction"),
                        "nominal_diameter_numerical_value_representation": self._uc1_attr_value(
                            anchor_attributes,
                            "nominal_diameter_numerical_value_representation",
                            "nominal_diameter_value",
                            "nominal_diameter",
                        ),
                        "nominal_diameter_representation": self._uc1_attr_value(
                            anchor_attributes,
                            "nominal_diameter_representation",
                            "nominal_diameter",
                        ),
                        "nominal_diameter_standard": self._uc1_attr_value(
                            anchor_attributes,
                            "nominal_diameter_standard",
                        ),
                        "nominal_diameter_type_representation": self._uc1_attr_value(
                            anchor_attributes,
                            "nominal_diameter_type_representation",
                        ),
                        "line_number": self._uc1_attr_value(anchor_attributes, "line_number", "linenumber"),
                        "piping_component_name": self._uc1_first_value(
                            self._uc1_attr_value(anchor_attributes, "piping_component_name"),
                            getattr(anchor_node, "tag_name", "") if anchor_node is not None else "",
                        ),
                        "label_text": getattr(instance, "label_text", ""),
                        "function_code": function_code,
                        "piping_anchor_id": getattr(instance, "piping_anchor_id", ""),
                        "from_equipment_id": getattr(instance, "from_equipment", ""),
                        "to_equipment_id": getattr(instance, "to_equipment", ""),
                        "context_summary": row.context_summary,
                        "source_doc_id": ri_doc_id,
                        "source_locator": ri_locator,
                        "xsd_status": row.xsd_status,
                        "confidence": 1.0,
                        "recommended_action": row.recommended_action,
                        "proposal_status": row.proposal_status,
                        "missing_targets": " | ".join(missing_targets),
                        "needs_review": "true" if missing_targets else "false",
                        "decision_confidence": self._uc1_score_text(row.decision_confidence, 0.0),
                        "evidence_bundle_id": clean_cell(row.evidence_bundle_id),
                        "uncertainty_reason": clean_cell(row.uncertainty_reason),
                        "llm_verification_status": clean_cell(row.llm_verification_status),
                        "rule_support": " | ".join(clean_cell(rule) for rule in row.rule_support if clean_cell(rule)),
                        "review_feedback_status": clean_cell(row.review_feedback_status),
                        "decision_trace_json": self._uc1_decision_trace_json(row.decision_trace),
                    },
                    completion_context,
                    canonical_entity_id=device_id,
                    match_confidence="1.0",
                    match_method="ri_ground_truth",
                    needs_review_reason="pid_cross_document_gap" if missing_targets else "",
                )
            )

            completion_rows.append(
                {
                    "device_id": device_id,
                    "canonical_tag": row.canonical_tag,
                    "present_in_ri": "true",
                    "present_in_stellenplan": completion_context["present_in_stellenplan"],
                    "present_in_wiring": completion_context["present_in_wiring"],
                    "present_in_datasheet": "false",
                    "present_in_ifc": completion_context["present_in_ifc"],
                    "flange_complete": completion_context["flange_complete"],
                    "uc1_candidate": "true" if row.is_uc1_candidate else "false",
                    "missing_targets": completion_context["missing_targets"],
                    "recommended_action": row.recommended_action,
                    "proposal_status": row.proposal_status,
                    "decision_confidence": completion_context["decision_confidence"],
                    "evidence_bundle_id": completion_context["evidence_bundle_id"],
                    "uncertainty_reason": completion_context["uncertainty_reason"],
                    "llm_verification_status": completion_context["llm_verification_status"],
                    "rule_support": completion_context["rule_support"],
                    "review_feedback_status": completion_context["review_feedback_status"],
                    "decision_trace_json": completion_context["decision_trace_json"],
                }
            )

            stellen_keys = sorted(strict_index["Stellenplaene"].get(row.normalized_key, set()))
            wiring_keys = sorted(strict_index["Verschaltungslisten"].get(row.normalized_key, set()))
            self._uc1_append_standardized_sheet_rows(
                stellenplan_rows,
                sheet_type="stellenplan",
                device_id=device_id,
                canonical_tag=row.canonical_tag,
                record_keys=stellen_keys,
                record_lookup=record_lookup,
            )
            self._uc1_append_standardized_sheet_rows(
                wiring_rows,
                sheet_type="wiring",
                device_id=device_id,
                canonical_tag=row.canonical_tag,
                record_keys=wiring_keys,
                record_lookup=record_lookup,
            )

            detail_rows = ifc_details.get(row.normalized_key, [])
            if detail_rows:
                for detail in detail_rows:
                    ifc_rows.append(
                        self._uc1_finalize_standardized_row(
                            "piping",
                            {
                                "entry_id": self._uc1_relation_id(device_id, detail.get("global_id", ""), "ifc"),
                                "device_id": device_id,
                                "canonical_tag": row.canonical_tag,
                                "ifc_class": detail.get("ifc_class", ""),
                                "global_id": detail.get("global_id", ""),
                                "tag": detail.get("tag", ""),
                                "has_ports": detail.get("has_ports", ""),
                                "connected_to": detail.get("connected_to", ""),
                                "connected_from": detail.get("connected_from", ""),
                                "has_control_elements": detail.get("has_control_elements", ""),
                                "predefined_type": detail.get("predefined_type", ""),
                                "size": detail.get("size", ""),
                                "valve_mechanism": detail.get("valve_mechanism", ""),
                                "flow_coefficient": detail.get("flow_coefficient", ""),
                                "fail_position": detail.get("fail_position", ""),
                                "manual_override": detail.get("manual_override", ""),
                                "actuator_application": detail.get("actuator_application", ""),
                                "source_doc_id": detail.get("source_doc_id", ""),
                                "source_locator": detail.get("source_locator", ""),
                                "confidence": 1.0,
                                "presence_status": "present",
                                "flange_complete": completion_context["flange_complete"],
                            },
                            completion_context,
                            canonical_entity_id=device_id,
                            match_confidence="1.0",
                            match_method="ifc_match_keys",
                            needs_review_reason="ifc_flange_incomplete" if completion_context["flange_complete"] == "false" else "",
                        )
                    )
            else:
                # Software-control (UY) instruments do not require physical IFC
                _func_code = getattr(instance, "function_code", "") or row.primary_type or ""
                _is_uy = (_func_code == "UY" or row.canonical_tag in ("TU10.U41", "TU20.U42"))
                if _is_uy:
                    ifc_rows.append(
                        self._uc1_finalize_standardized_row(
                            "piping",
                            {
                                "entry_id": self._uc1_relation_id(device_id, "ifc"),
                                "device_id": device_id,
                                "canonical_tag": row.canonical_tag,
                                "ifc_class": "",
                                "global_id": "",
                                "tag": "",
                                "has_ports": "",
                                "connected_to": "",
                                "connected_from": "",
                                "has_control_elements": "",
                                "predefined_type": "",
                                "size": "",
                                "valve_mechanism": "",
                                "flow_coefficient": "",
                                "fail_position": "",
                                "manual_override": "",
                                "actuator_application": "",
                                "source_doc_id": "",
                                "source_locator": "",
                                "confidence": "",
                                "presence_status": "not_required",
                                "flange_complete": completion_context["flange_complete"],
                            },
                            completion_context,
                            canonical_entity_id=device_id,
                            match_confidence="1.0",
                            match_method="software_control_no_ifc_required",
                            needs_review_reason="",
                        )
                    )
                else:
                    ifc_rows.append(
                        self._uc1_finalize_standardized_row(
                            "piping",
                            {
                                "entry_id": self._uc1_relation_id(device_id, "ifc"),
                                "device_id": device_id,
                                "canonical_tag": row.canonical_tag,
                                "ifc_class": "",
                                "global_id": "",
                                "tag": "",
                                "has_ports": "",
                                "connected_to": "",
                                "connected_from": "",
                                "has_control_elements": "",
                                "predefined_type": "",
                                "size": "",
                                "valve_mechanism": "",
                                "flow_coefficient": "",
                                "fail_position": "",
                                "manual_override": "",
                                "actuator_application": "",
                                "source_doc_id": "",
                                "source_locator": "",
                                "confidence": "",
                                "presence_status": "missing",
                                "flange_complete": completion_context["flange_complete"],
                            },
                            completion_context,
                            canonical_entity_id=device_id,
                            match_confidence="0.0",
                            match_method="missing_placeholder",
                            needs_review_reason="ifc_missing",
                        )
                    )

            if getattr(instance, "piping_anchor_id", ""):
                relation_rows.append(
                    {
                        "relation_id": self._uc1_relation_id(device_id, "anchored_to", getattr(instance, "piping_anchor_id", "")),
                        "subject_id": device_id,
                        "predicate": "anchored_to",
                        "object_id": getattr(instance, "piping_anchor_id", ""),
                        "object_type": "PipingNetworkSegment",
                        "source_doc_id": ri_doc_id,
                        "source_locator": ri_locator,
                        "confidence": 1.0,
                    }
                )
            for equipment_id in (getattr(instance, "from_equipment", ""), getattr(instance, "to_equipment", "")):
                if not clean_cell(equipment_id):
                    continue
                relation_rows.append(
                    {
                        "relation_id": self._uc1_relation_id(device_id, "connected_to_equipment", equipment_id),
                        "subject_id": device_id,
                        "predicate": "connected_to_equipment",
                        "object_id": equipment_id,
                        "object_type": "Equipment",
                        "source_doc_id": ri_doc_id,
                        "source_locator": ri_locator,
                        "confidence": 1.0,
                    }
                )

        completion_map = self._uc1_completion_map(
            [
                {clean_cell(key): clean_cell(value) for key, value in row.items()}
                for row in completion_rows
            ]
        )
        datasheet_rows = self._uc1_datasheet_source_rows(ri_device_rows, completion_map)
        datasheet_presence_by_device: dict[str, bool] = {}
        for datasheet_row in datasheet_rows:
            identity_key = self._uc1_identity_key(datasheet_row)
            if not identity_key:
                continue
            if clean_cell(datasheet_row.get("presence_status", "")) == "present":
                datasheet_presence_by_device[identity_key] = True
            else:
                datasheet_presence_by_device.setdefault(identity_key, False)
        for completion_row_item in completion_rows:
            identity_key = self._uc1_identity_key(completion_row_item)
            completion_row_item["present_in_datasheet"] = "true" if datasheet_presence_by_device.get(identity_key, False) else "false"

        coverage_rows = [item.model_dump(mode="json") for item in coverage.rows]
        if self._use_saved_t1_t5_rules():
            legacy_dataset = {
                "ri_device_rows": ri_device_rows,
                "stellenplan_rows": stellenplan_rows,
                "wiring_rows": wiring_rows,
                "datasheet_rows": datasheet_rows,
                "ifc_rows": ifc_rows,
                "completion_rows": completion_rows,
            }
            ri_device_rows = self._build_t1_t5_runtime_rows("t1", legacy_dataset)
            stellenplan_rows = self._build_t1_t5_runtime_rows("t2", legacy_dataset)
            wiring_rows = self._build_t1_t5_runtime_rows("t3", legacy_dataset)
            datasheet_rows = self._build_t1_t5_runtime_rows("t4", legacy_dataset)
            ifc_rows = self._build_t1_t5_runtime_rows("t5", legacy_dataset)

        self._report_progress(progress, 100, "Built enriched source rows")
        return {
            "pid": ri_device_rows,
            "instrument_list": stellenplan_rows,
            "wiring": wiring_rows,
            "datasheet": datasheet_rows,
            "piping": ifc_rows,
            "completion_rows": completion_rows,
            "relation_rows": relation_rows,
            "documents_rows": documents_rows,
            "coverage_rows": coverage_rows,
            "completion_map": completion_map,
        }

    def _uc1_resolve_ifc_evidence(self, canonical_tag: str) -> dict:
        """Unified IFC evidence resolution for a single instrument.
        Priority: mapping special status > Assembly evidence > raw IFC > missing.
        Returns {status, global_id, label, method, ...}
        """
        evidence = self._uc1_assembly_ifc_evidence_index()
        ev = evidence.get(canon_tag := canonical_tag, {})

        # Priority 1: mapping workbook special status
        if ev.get("_status") == "software_control_no_ifc_required":
            return {"status": "not_required", "global_id": "", "label": "",
                    "method": "software_control_no_ifc_required"}
        if ev.get("_status") == "ifc_missing":
            return {"status": "missing", "global_id": "", "label": "",
                    "method": "mapping_ifc_missing"}

        # Priority 2: Assembly workbook evidence
        if ev.get("global_id"):
            return {"status": "present", "global_id": ev["global_id"],
                    "label": ev.get("label", ""), "method": "assembly_mapping_evidence",
                    "ifc_class": "IfcBuildingElementProxy"}

        # Priority 3: fallback — no evidence
        return {"status": "unknown", "global_id": "", "label": "", "method": "no_evidence"}

    @staticmethod
    def _uc1_trace_json(source_type: str, decision: str, rule_support: str) -> str:
        import json
        return json.dumps({
            source_type: {
                "decision": decision,
                "rule_support": [rule_support] if rule_support else [],
                "confidence": 1.0 if decision != "missing" else 0.0,
            }
        }, ensure_ascii=False)

    def _ensure_assembly_3d_template(self) -> str:
        """Auto-generate Assembly_3D_template_filled.xlsx from project-local sources.

        Prefers a prepared assembled IFC (assembly_prepared.ifc) when available;
        falls back to the legacy FCStd + multi-IFC pipeline otherwise.

        Returns: 'generated' | 'skipped_missing_source' | 'failed'
        """
        try:
            from iev4pi_transformation_tool.services.assembly_3d_pipeline import Assembly3DPipelineService

            piping_dir = self.workspace_root / "Documents" / "Piping Diagram"
            svc = Assembly3DPipelineService(self.workspace_root)

            # Auto-discover IFC sources (no hardcoded filenames)
            primary_ifc, fallback_ifc, _ = svc.discover_ifc_sources()
            if primary_ifc:
                svc.build_template_from_ifc(
                    ifc_path=primary_ifc,
                    csv_path=piping_dir / "_legacy" / "final_result_components.csv",
                    fallback_ifc=fallback_ifc)
                self._generate_pid_ifc_mapping()
                return "generated"

            # Fallback: legacy FCStd path
            fcstd_path = piping_dir / "_legacy" / "final_result.FCStd"
            if not fcstd_path.exists():
                return "skipped_missing_source"

            svc.build_template(fcstd_path=fcstd_path,
                v1_dir=piping_dir / "_legacy", v0_dir=piping_dir / "_legacy",
                csv_path=piping_dir / "_legacy" / "final_result_components.csv")
            self._generate_pid_ifc_mapping()
            return "generated"
        except ImportError:
            return "skipped_missing_source"
        except Exception:
            return "failed"

    @staticmethod
    def _generate_pid_ifc_mapping() -> None:
        """Auto-generate P&ID→IFC mapping sheets in the Assembly workbook."""
        try:
            from scripts.build_pid_ifc_component_mapping import build_mapping_workbook
            from pathlib import Path
            build_mapping_workbook(
                Path("Documents/Piping Diagram/Assembly_3D_template_filled.xlsx")
            )
        except Exception:
            pass  # Mapping is optional; skip gracefully if sources unavailable



    def _uc1_component_evidence_index(self) -> dict[str, dict]:
        """Build shared component evidence index from Assembly workbook P&ID sheets.
        Returns {"comp_data": {component: {ms, gid, label, anchor, tags, pid_x, pid_y}},
                 "asm_by_vv": {vv_tag: {gid, label, akz}}}
        Used by both UI and export injection methods to avoid duplicate logic.
        """
        import openpyxl
        asm_path = self.workspace_root / "Documents" / "Piping Diagram" / "Assembly_3D_template_filled.xlsx"

        comp_data: dict[str, dict] = {}
        asm_by_vv: dict[str, dict] = {}

        # ── Read P&ID mapping from Assembly workbook ──
        if asm_path.exists():
            wb = openpyxl.load_workbook(asm_path)
            # Prefer merged sheet name; fall back to legacy file
            sheet_name = "P&ID_instance_mapping" if "P&ID_instance_mapping" in wb.sheetnames else (
                "instance_mapping" if "instance_mapping" in wb.sheetnames else None)
            if sheet_name is not None:
                ws = wb[sheet_name]
                headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                ci = {h: i + 1 for i, h in enumerate(headers)}
                for r in range(2, ws.max_row + 1):
                    comp = ws.cell(row=r, column=ci.get("pand_id_component", 1)).value or ""
                    if not comp or any(comp.startswith(p) for p in ("B", "DM", "R")):
                        continue
                    if "(no anchor)" in comp.lower() or comp.strip() == "":
                        continue
                    if comp not in comp_data:
                        comp_data[comp] = {"ms": "", "gid": "", "label": "", "anchor": "",
                            "tags": set(), "pid_x": None, "pid_y": None}
                    d = comp_data[comp]
                    d["ms"] = ws.cell(row=r, column=ci.get("match_status", 1)).value or d["ms"]
                    d["gid"] = ws.cell(row=r, column=ci.get("global_id", 1)).value or d["gid"]
                    d["label"] = ws.cell(row=r, column=ci.get("mapped_3d_label", 1)).value or d["label"]
                    d["anchor"] = ws.cell(row=r, column=ci.get("pand_id_anchor", 1)).value or d["anchor"]
                    d["pid_x"] = ws.cell(row=r, column=ci.get("pid_x", 1)).value
                    d["pid_y"] = ws.cell(row=r, column=ci.get("pid_y", 1)).value
                    tag = ws.cell(row=r, column=ci.get("canonical_tag", 1)).value or ""
                    if tag: d["tags"].add(tag)
            wb.close()

        # ── Read Assembly workbook ──
        if asm_path.exists():
            awb = openpyxl.load_workbook(asm_path)
            aws = awb["Assembly_Steps"]
            ah = [aws.cell(row=1, column=c).value for c in range(1, aws.max_column + 1)]
            aci = {h: i + 1 for i, h in enumerate(ah)}
            for ar in range(2, aws.max_row + 1):
                vv = aws.cell(row=ar, column=aci.get("vv_tag", 1)).value or ""
                akz = aws.cell(row=ar, column=aci.get("akz_tag", 1)).value or ""
                gid = aws.cell(row=ar, column=aci.get("ifc_global_id", 1)).value or ""
                cov = aws.cell(row=ar, column=aci.get("ifc_coverage", 1)).value or ""
                lbl = aws.cell(row=ar, column=aci.get("label", 1)).value or ""
                if cov != "YES" or not gid:
                    continue
                if vv and vv not in asm_by_vv:
                    asm_by_vv[vv] = {"gid": gid, "label": lbl, "akz": akz}
                if akz and akz not in asm_by_vv:
                    asm_by_vv.setdefault(akz, {"gid": gid, "label": lbl, "vv": vv})
            awb.close()

        return {"comp_data": comp_data, "asm_by_vv": asm_by_vv}

    def _uc1_inject_component_rows(
        self, rows: list, external_index: dict, all_pdf_identifiers: set
    ) -> None:
        """Inject P&ID component rows (VV, PL, HE) with REAL status from all sources.
        Also injects IFC-extra components (VV007, VV008) from Assembly workbook.
        """
        existing_tags = {row.canonical_tag for row in rows if row.canonical_tag}
        evidence = self._uc1_component_evidence_index()
        comp_data = evidence["comp_data"]
        asm_by_vv = evidence["asm_by_vv"]

        # ── Step 1: Inject mapping-derived component rows ──
        for comp, d in sorted(comp_data.items()):
            if comp in existing_tags:
                continue
            ms = d["ms"]
            gid = d["gid"] or asm_by_vv.get(comp, {}).get("gid", "")
            mapped_label = d["label"] or asm_by_vv.get(comp, {}).get("label", "")
            anchor = d["anchor"]
            px, py = d["pid_x"], d["pid_y"]

            norm_key = normalize_identifier(comp)
            # Also check related AKZ tags for document status
            related_keys = {norm_key} | {normalize_identifier(t) for t in (d.get("tags", set()) if isinstance(d.get("tags"), set) else set())}
            d_tags = d.get("tags", set()) if isinstance(d.get("tags"), set) else set()
            related_keys |= {normalize_identifier(t) for t in d_tags}
            # Check component, its tags, and anchor against external indexes
            anchor_key = normalize_identifier(d.get("anchor", "")) if d.get("anchor") else ""
            all_lookup_keys = related_keys | ({anchor_key} if anchor_key else set())

            # VV* components: Stellenplaene only from own tag (not related AKZ)
            is_vv = comp.startswith("VV")
            if is_vv:
                stellen_keys = external_index["Stellenplaene"].get(norm_key, set())
                stellen_hit_key = norm_key if stellen_keys else ""
            else:
                stellen_keys = set()
                stellen_hit_key = ""
                for lk in sorted(all_lookup_keys, key=lambda k: k != norm_key):
                    sk = external_index["Stellenplaene"].get(lk, set())
                    if sk and not stellen_keys:
                        stellen_hit_key = lk
                    stellen_keys |= sk

            versch_keys = set()
            versch_hit_key = ""
            for lk in sorted(all_lookup_keys, key=lambda k: k != norm_key):
                vk = external_index["Verschaltungslisten"].get(lk, set())
                if vk and not versch_keys:
                    versch_hit_key = lk
                versch_keys |= vk

            pdf_present = any(k in all_pdf_identifiers for k in all_lookup_keys)
            # Also check by stripped component anchor
            if not pdf_present and anchor_key:
                pdf_present = anchor_key in all_pdf_identifiers
            if not pdf_present:
                pdf_present = norm_key in all_pdf_identifiers

            # Resolve IFC status (no "deferred" in UI)
            if gid:
                ifc_status = "present"; ifc_key = gid
            elif ms == "ifc_missing":
                ifc_status = "missing"; ifc_key = ""
            elif ms == "software_control_no_ifc_required":
                ifc_status = "not_required"; ifc_key = ""
            else:
                ifc_status = "missing"; ifc_key = ""

            # Compute issues from all 5 sources
            issues = []
            if not pdf_present: issues.append("missing_pdf")
            issues.append("missing_xml")
            if not stellen_keys: issues.append("missing_stellenplaene")
            if not versch_keys: issues.append("missing_verschaltungslisten")
            if ifc_status == "missing": issues.append("missing_ifc")
            action = "补充 IFC 对象" if ifc_status == "missing" else ("无需补全" if not issues else "")

            # Build context WITHOUT blackbox/IFC labels
            ctx_parts = [f"P&ID {comp}"]
            if d.get("tags"):
                related_str = ", ".join(sorted(d["tags"]))
                ctx_parts.append(f"[tags: {related_str}]")

            # Build jump_targets for Verschaltung only (VV Stellenplaene disabled)
            jt: dict[str, PidJumpTarget] = {}
            if versch_hit_key and versch_keys:
                vkeys = sorted(versch_keys)
                jt["verschaltungslisten"] = PidJumpTarget(
                    keyword=versch_hit_key, preferred_source_root="Verschaltungslisten",
                    matching_record_keys=vkeys,
                    preferred_record_key=vkeys[0] if vkeys else "")

            rows.append(PidInconsistencyRow(
                component_key=comp, display_name=comp, normalized_key=norm_key,
                canonical_tag=comp, pdf_status="present" if pdf_present else "missing",
                xml_status="missing",
                stellenplaene_status="present" if stellen_keys else "missing",
                verschaltungslisten_status="present" if versch_keys else "missing",
                ifc_match_status=ifc_status, ifc_match_key=ifc_key, flange_status="unknown",
                is_uc1_candidate=(ifc_status == "missing" or not (stellen_keys or versch_keys)),
                issue_count=len(issues), issues=issues, recommended_action=action,
                context_summary=" ".join(ctx_parts),
                jump_targets=jt,
            ))

        # ── Step 4: Inject IFC-extra components (VV007, VV008, etc.) ──
        for vv_tag, asm in sorted(asm_by_vv.items()):
            if vv_tag in comp_data or vv_tag in existing_tags:
                continue  # already covered
            # Only inject standalone VV tags (not AKZ tags that happen to be in the dict)
            if not vv_tag.startswith(("VV", "PL", "HE")):
                continue
            norm_key = normalize_identifier(vv_tag)
            stellen_keys = external_index["Stellenplaene"].get(norm_key, set())
            versch_keys = external_index["Verschaltungslisten"].get(norm_key, set())
            pdf_present = norm_key in all_pdf_identifiers

            extra_ctx = f"3D/IFC extra {vv_tag}"

            rows.append(PidInconsistencyRow(
                component_key=vv_tag, display_name=vv_tag, normalized_key=norm_key,
                canonical_tag=vv_tag, pdf_status="present" if pdf_present else "missing",
                xml_status="missing",
                stellenplaene_status="present" if stellen_keys else "missing",
                verschaltungslisten_status="present" if versch_keys else "missing",
                ifc_match_status="present", ifc_match_key=asm["gid"], flange_status="unknown",
                is_uc1_candidate=False,
                issue_count=sum(1 for s in [
                    "present" if pdf_present else "missing",
                    "missing", "present" if stellen_keys else "missing",
                    "present" if versch_keys else "missing"] if s != "present"),
                issues=[i for i in (["missing_pdf"] if not pdf_present else []) + ["missing_xml"] +
                        (["missing_stellenplaene"] if not stellen_keys else []) +
                        (["missing_verschaltungslisten"] if not versch_keys else [])],
                recommended_action="补全文档来源",
                context_summary=extra_ctx,
            ))



    @staticmethod
    def _uc1_component_missing_targets(pdf_missing=False, xml_missing=True,
                                       stellen_missing=False, versch_missing=False,
                                       ifc_missing=False, ifc_present=False):
        """Compute missing_targets string and presence flags from 5-source status."""
        parts = []
        if pdf_missing: parts.append("pdf")
        if xml_missing: parts.append("xml")
        if stellen_missing: parts.append("stellenplan")
        if versch_missing: parts.append("wiring")
        if ifc_missing: parts.append("ifc")
        return {
            "missing_targets": " | ".join(parts),
            "present_in_stellenplan": "false" if stellen_missing else "true",
            "present_in_wiring": "false" if versch_missing else "true",
            "present_in_ifc": "false" if ifc_missing else ("true" if ifc_present else "false"),
            "recommended_action": "补充 IFC 对象" if ifc_missing else
                ("补全文档来源" if parts else "无需补全"),
        }

    def _uc1_inject_component_export_rows(self, source_data: dict) -> dict:

        """Inject VV/PL/HE component rows into export (ifc_entries + completion + relations).
        Uses same logic as UI method for consistency."""
        existing_tags = {r.get("canonical_tag", "") for r in source_data.get("piping", [])}
        evidence = self._uc1_component_evidence_index()
        comp_data = evidence["comp_data"]
        asm_by_vv = evidence["asm_by_vv"]
        external_index = self._strict_external_match_index()

        def _component_external_status(component: str, tags: object = None, anchor: str = "") -> tuple[bool, bool]:
            norm_key = normalize_identifier(component)
            tag_values = tags if isinstance(tags, set) else set()
            lookup_keys = {norm_key} | {normalize_identifier(tag) for tag in tag_values}
            anchor_key = normalize_identifier(anchor)
            if anchor_key:
                lookup_keys.add(anchor_key)

            if component.startswith("VV"):
                stellen_present = bool(external_index["Stellenplaene"].get(norm_key))
            else:
                stellen_present = any(external_index["Stellenplaene"].get(key) for key in lookup_keys if key)
            versch_present = any(external_index["Verschaltungslisten"].get(key) for key in lookup_keys if key)
            return stellen_present, versch_present

        for comp, d in sorted(comp_data.items()):
            gid = d["gid"] or asm_by_vv.get(comp, {}).get("gid", "")
            lbl = d["label"] or asm_by_vv.get(comp, {}).get("label", "")
            ms = d["ms"]
            display_tag = comp
            device_id = f"urn:ievpi:device:{normalize_identifier(display_tag)}"

            _stellen_present, _versch_present = _component_external_status(
                comp,
                d.get("tags", set()),
                d.get("anchor", ""),
            )

            if gid:
                tgt = self._uc1_component_missing_targets(
                    xml_missing=True, stellen_missing=not _stellen_present,
                    versch_missing=not _versch_present, ifc_present=True)
                mt = tgt["missing_targets"]
                ps, mm, nrr, pifc, conf = "present", "assembly_mapping_evidence", "", "true", "0.9"
                ra = tgt["recommended_action"]
            elif ms == "ifc_missing":
                tgt = self._uc1_component_missing_targets(
                    xml_missing=True, stellen_missing=not _stellen_present,
                    versch_missing=not _versch_present, ifc_missing=True)
                mt = tgt["missing_targets"]
                ps, mm, nrr, pifc, conf = "missing", "missing_placeholder", "ifc_missing", "false", "1.0"
                ra = "补充 IFC 对象"
            else:
                continue

            trace = self._uc1_trace_json("piping", "present" if gid else "missing",
                "assembly_mapping_evidence" if gid else "ifc_missing")

            # UPSERT: update existing row if present, otherwise append
            updated = False
            for row in source_data["piping"]:
                if row.get("canonical_tag") == display_tag:
                    row["global_id"] = gid
                    row["presence_status"] = ps
                    row["match_method"] = mm
                    row["needs_review_reason"] = nrr
                    row["flange_complete"] = "true" if gid else "false"
                    row["decision_trace_json"] = trace
                    row["tag"] = lbl
                    if gid: row["ifc_class"] = "IfcBuildingElementProxy"
                    updated = True
                    break
            if not updated:
                source_data["piping"].append(self._uc1_finalize_standardized_row(
                "piping",
                {"entry_id": self._uc1_relation_id(device_id, "ifc"), "device_id": device_id,
                 "canonical_tag": display_tag, "ifc_class": "IfcBuildingElementProxy" if gid else "",
                 "global_id": gid, "tag": lbl, "has_ports": "false",
                 "connected_to": "", "connected_from": "", "has_control_elements": "",
                 "predefined_type": "", "size": "", "valve_mechanism": "",
                 "flow_coefficient": "", "fail_position": "", "manual_override": "",
                 "actuator_application": "", "source_doc_id": "", "source_locator": "",
                 "confidence": conf, "presence_status": ps, "flange_complete": "true" if gid else "false",
                 "decision_trace_json": trace},
                {"recommended_action": ra, "proposal_status": "ready", "missing_targets": mt,
                 "present_in_ifc": pifc, "decision_trace_json": trace},
                canonical_entity_id=device_id, match_confidence=conf,
                match_method=mm, needs_review_reason=nrr))
            # UPSERT completion
            comp_updated = False
            for row in source_data.get("completion_rows", []):
                if row.get("canonical_tag") == display_tag:
                    row["present_in_ifc"] = pifc
                    row["missing_targets"] = mt
                    row["recommended_action"] = ra
                    row["decision_trace_json"] = trace
                    comp_updated = True
                    break
            if not comp_updated:
                source_data["completion_rows"].append({
                    "completion_id": f"{device_id}:completion", "device_id": device_id,
                    "canonical_tag": display_tag, "recommended_action": ra, "proposal_status": "ready",
                    "missing_targets": mt, "present_in_ifc": pifc, "decision_trace_json": trace})

            # UPSERT relation
            rel_exists = any(r.get("object_id") == comp for r in source_data.get("relation_rows", []))
            if not rel_exists:
                source_data["relation_rows"].append({
                "relation_id": self._uc1_relation_id(device_id, "anchored_to", comp),
                "subject_id": device_id, "predicate": "anchored_to", "object_id": comp,
                "object_type": "PipingNetworkSegment", "confidence": conf})

        

        # ── Inject IFC-extra components (VV007, VV008, HE002, PL001-003) ──
        for vv_tag, asm in sorted(asm_by_vv.items()):
            if vv_tag in comp_data:
                continue  # already handled by mapping component loop
            if not vv_tag.startswith(("VV", "PL", "HE")):
                continue
            display_tag = vv_tag
            device_id = f"urn:ievpi:device:{normalize_identifier(display_tag)}"
            gid = asm["gid"]
            lbl = asm["label"]
            _stellen_present, _versch_present = _component_external_status(vv_tag)
            x_tgt = self._uc1_component_missing_targets(
                xml_missing=True,
                stellen_missing=not _stellen_present,
                versch_missing=not _versch_present,
                ifc_present=True,
            )
            trace = self._uc1_trace_json("piping", "present", "assembly_mapping_evidence")
            updated = False
            for row in source_data["piping"]:
                if row.get("canonical_tag") == display_tag:
                    row["global_id"] = gid; row["presence_status"] = "present"
                    row["match_method"] = "assembly_mapping_evidence"; row["needs_review_reason"] = ""
                    row["flange_complete"] = "true"; row["decision_trace_json"] = trace
                    row["tag"] = lbl; row["ifc_class"] = "IfcBuildingElementProxy"
                    updated = True; break
            if not updated:
                source_data["piping"].append(self._uc1_finalize_standardized_row(
                    "piping",
                    {"entry_id": self._uc1_relation_id(device_id, "ifc"), "device_id": device_id,
                     "canonical_tag": display_tag, "ifc_class": "IfcBuildingElementProxy",
                     "global_id": gid, "tag": lbl, "has_ports": "false",
                     "connected_to": "", "connected_from": "", "has_control_elements": "",
                     "predefined_type": "", "size": "", "valve_mechanism": "",
                     "flow_coefficient": "", "fail_position": "", "manual_override": "",
                     "actuator_application": "", "source_doc_id": "", "source_locator": "",
                     "confidence": "0.9", "presence_status": "present", "flange_complete": "true",
                     "decision_trace_json": trace},
                    {"recommended_action": x_tgt["recommended_action"], "proposal_status": "ready",
                     "missing_targets": x_tgt["missing_targets"], "present_in_ifc": x_tgt["present_in_ifc"],
                     "decision_trace_json": trace},
                    canonical_entity_id=device_id, match_confidence="0.9",
                    match_method="assembly_mapping_evidence", needs_review_reason=""))
            # UPSERT completion
            comp_up = False
            for row in source_data.get("completion_rows", []):
                if row.get("canonical_tag") == display_tag:
                    row["present_in_ifc"] = x_tgt["present_in_ifc"]
                    row["missing_targets"] = x_tgt["missing_targets"]
                    row["recommended_action"] = x_tgt["recommended_action"]
                    row["decision_trace_json"] = trace; comp_up = True; break
            if not comp_up:
                source_data["completion_rows"].append({
                    "completion_id": f"{device_id}:completion", "device_id": device_id,
                    "canonical_tag": display_tag, "recommended_action": x_tgt["recommended_action"],
                    "proposal_status": "ready", "missing_targets": x_tgt["missing_targets"],
                    "present_in_ifc": x_tgt["present_in_ifc"],
                    "decision_trace_json": trace})
            rel_ex = any(r.get("object_id") == vv_tag for r in source_data.get("relation_rows", []))
            if not rel_ex:
                source_data["relation_rows"].append({
                    "relation_id": self._uc1_relation_id(device_id, "anchored_to", vv_tag),
                    "subject_id": device_id, "predicate": "anchored_to", "object_id": vv_tag,
                    "object_type": "PipingNetworkSegment", "confidence": "0.9"})

        return source_data

    def _uc1_assembly_ifc_evidence_index(self) -> dict[str, dict]:
        """Build evidence index: canonical_tag → {global_id, label, _status, ...}
        Reads Assembly_3D_template_filled.xlsx and pid_ifc_component_mapping.xlsx.
        """
        import openpyxl
        piping_dir = self.workspace_root / "Documents" / "Piping Diagram"
        evidence: dict[str, dict] = {}

        assembly_path = piping_dir / "Assembly_3D_template_filled.xlsx"
        if assembly_path.exists():
            wb = openpyxl.load_workbook(assembly_path)
            ws = wb["Assembly_Steps"]
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1):
                row = {h: ws.cell(row=r, column=c + 1).value for c, h in enumerate(headers)}
                akz = row.get("akz_tag", "")
                if not akz or row.get("ifc_coverage") != "YES":
                    continue
                evidence[akz] = {
                    "global_id": row.get("ifc_global_id", ""),
                    "ifc_part_name": row.get("ifc_part_name", ""),
                    "ifc_parts_library": row.get("ifc_parts_library", ""),
                    "label": row.get("label", ""),
                    "type": row.get("type", ""),
                    "ifc_class": "IfcBuildingElementProxy",
                }
            wb.close()

        # Read P&ID status flags from Assembly workbook (merged) or legacy file
        if assembly_path.exists():
            wb2 = openpyxl.load_workbook(assembly_path)
            sheet_name = "P&ID_instance_mapping" if "P&ID_instance_mapping" in wb2.sheetnames else (
                "instance_mapping" if "instance_mapping" in wb2.sheetnames else None)
            if sheet_name is not None:
                ws2 = wb2[sheet_name]
                headers2 = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
                ci2 = {h: i + 1 for i, h in enumerate(headers2)}
                for r in range(2, ws2.max_row + 1):
                    tag = ws2.cell(row=r, column=ci2.get("canonical_tag", 1)).value
                    if not tag: continue
                    ms = ws2.cell(row=r, column=ci2.get("match_status", 1)).value or ""
                    if ms in ("software_control_no_ifc_required", "ifc_missing"):
                        evidence.setdefault(tag, {})["_status"] = ms
            wb2.close()
        return evidence

    def _uc1_enrich_ifc_from_assembly_mapping(self, source_data: dict) -> dict:
        """Enrich ifc_entries and completion_rows with unified Assembly/mapping evidence."""
        # Apply evidence to existing ifc_entries
        new_piping = []
        for row in source_data.get("piping", []):
            tag = row.get("canonical_tag", "")
            resolved = self._uc1_resolve_ifc_evidence(tag)

            if resolved["status"] == "not_required":
                row["presence_status"] = "not_required"
                row["match_method"] = "software_control_no_ifc_required"
                row["global_id"] = ""
                row["needs_review_reason"] = ""
                row["confidence"] = "1.0"
                row["decision_trace_json"] = self._uc1_trace_json("piping", "not_required", "software_control")
            elif resolved["status"] == "missing":
                row["presence_status"] = "missing"
                row["match_method"] = "missing_placeholder"
                row["needs_review_reason"] = "ifc_missing"
                row["global_id"] = ""
                row["flange_complete"] = "false"
                row["decision_trace_json"] = self._uc1_trace_json("piping", "missing", "ifc_missing")
            elif resolved["status"] == "present":
                row["decision_trace_json"] = self._uc1_trace_json("piping", "present", "assembly_mapping_evidence")
                row["global_id"] = resolved["global_id"]
                row["ifc_class"] = resolved.get("ifc_class", "IfcBuildingElementProxy")
                row["tag"] = resolved.get("label", "")
                row["source_doc_id"] = f"urn:ievpi:document:assembly3d:{resolved.get('label', tag)}"
                row["source_locator"] = f"Assembly_Steps::{resolved.get('label', '')}"
                row["presence_status"] = "present"
                row["match_method"] = resolved["method"]
                row["match_confidence"] = "0.9"
                row["confidence"] = "0.9"
                row["needs_review_reason"] = ""
                row["flange_complete"] = "true"
                if not row.get("has_ports"): row["has_ports"] = "false"
                # Clean constraint_violations of ifc entries
                cv = (row.get("constraint_violations") or "")
                cv = " | ".join(v.strip() for v in cv.split("|") if "ifc" not in v.strip().lower() and v.strip())
                row["constraint_violations"] = cv
            new_piping.append(row)
        source_data["piping"] = new_piping

        # Update completion_rows
        tag_to_info = {}
        for r in new_piping:
            tag = r.get("canonical_tag", "")
            tag_to_info[tag] = {
                "present": r.get("presence_status") == "present",
                "not_required": r.get("match_method") == "software_control_no_ifc_required",
                "missing": r.get("needs_review_reason") == "ifc_missing",
            }
        new_completions = []
        for row in source_data.get("completion_rows", []):
            tag = row.get("canonical_tag", "")
            info = tag_to_info.get(tag, {})
            if info.get("present"):
                row["present_in_ifc"] = "true"
                row["decision_trace_json"] = self._uc1_trace_json("piping", "present", "assembly_mapping_evidence")
                mt = (row.get("missing_targets") or "")
                mt = " | ".join(p.strip() for p in mt.split("|") if p.strip() and "ifc" not in p.strip().lower())
                row["missing_targets"] = mt
                ra = (row.get("recommended_action") or "")
                for phrase in ("补充 IFC 对象", "Add IFC object", "IFC-Objekt ergaenzen"):
                    ra = ra.replace(f"，{phrase}", "").replace(f"，{phrase}，", "，").replace(phrase + "，", "").replace(phrase, "")
                for phrase in ("补充法兰属性", "补充法兰属性，", "，补充法兰属性",
                              "Complete flange attributes", "Flanschattribute ergaenzen"):
                    ra = ra.replace(phrase, "")
                ra = ra.strip("，").strip(",").strip()
                row["recommended_action"] = ra
                row["flange_complete"] = "true"
            elif info.get("missing"):
                row["present_in_ifc"] = "false"
                row["decision_trace_json"] = self._uc1_trace_json("piping", "missing", "ifc_missing")
                mt = (row.get("missing_targets") or "")
                if "ifc" not in mt.lower():
                    row["missing_targets"] = (" | ".join([p.strip() for p in mt.split("|") if p.strip()] + ["ifc"])).strip("|").strip()
            elif info.get("not_required"):
                row["present_in_ifc"] = "not_required"
                row["decision_trace_json"] = self._uc1_trace_json("piping", "not_required", "software_control")
            new_completions.append(row)
        source_data["completion_rows"] = new_completions

        return source_data

    def export_use_case_1_standardized_workbook(
        self,
        progress: ProgressCallback | None = None,
        *,
        output_dir: Path | None = None,
    ) -> Path:
        source_data = self._uc1_build_source_rows(
            self._sub_progress(progress, 0, 90) if progress else None
        )
        # ── Enrich IFC evidence from Assembly + Mapping workbooks ──
        source_data = self._uc1_enrich_ifc_from_assembly_mapping(source_data)
        # Inject P&ID component rows from mapping (VV, PL, HE) — present AND missing
        source_data = self._uc1_inject_component_export_rows(source_data)
        target_dir = output_dir or ensure_dir(self.resolve_results_export_dir() / "UseCase1")
        path = self.export_service.export_use_case_1_standardized_workbook(
            target_dir,
            documents_rows=source_data["documents_rows"],
            ri_device_rows=source_data["pid"],
            stellenplan_rows=source_data["instrument_list"],
            wiring_rows=source_data["wiring"],
            datasheet_rows=source_data["datasheet"],
            ifc_rows=source_data["piping"],
            relation_rows=source_data["relation_rows"],
            completion_rows=source_data["completion_rows"],
            coverage_rows=source_data["coverage_rows"],
            language=self._ui_language_code(),
            catalog_path=self.uc1_catalog_path().name,
        )
        self._report_progress(progress, 100, f"Exported standardized UC1 workbook to {path.name}")
        return path


    def _legacy_t1_t5_preview_rows(self, stage_id: str) -> list[dict[str, str]]:
        previous_flag = self.settings.use_custom_t1_t5_rules
        try:
            self.settings.use_custom_t1_t5_rules = False
            workbook_path = self.export_use_case_1_standardized_workbook()
        finally:
            self.settings.use_custom_t1_t5_rules = previous_flag
        sheets = self._uc1_load_excel_rows(workbook_path)
        source_type = stage_source_type(stage_id)
        sheet_map = {
            "pid": "ri_devices",
            "instrument_list": "instrument_list_entries",
            "wiring": "wiring_entries",
            "datasheet": "datasheet_entries",
            "piping": "ifc_entries",
        }
        return sheets.get(sheet_map[source_type], [])

    def _stage_rows_from_standardized_dataset(
        self,
        dataset: dict[str, list[dict[str, object]]],
        stage_id: str,
    ) -> list[dict[str, str]]:
        mapping = {
            "t1": "ri_device_rows",
            "t2": "stellenplan_rows",
            "t3": "wiring_rows",
            "t4": "datasheet_rows",
            "t5": "ifc_rows",
        }
        rows = dataset.get(mapping[normalize_identifier(stage_id)], [])
        return [
            {clean_cell(key): clean_cell(value) for key, value in row.items()}
            for row in rows
        ]

    def _t1_t5_candidate_workbooks(self) -> list[DocumentDescriptor]:
        return [
            document
            for document in self.documents
            if document.extension.lower() in {".xls", ".xlsx", ".xlsm"}
        ]

    def _build_t1_t5_runtime_rows(
        self,
        stage_id: str,
        dataset: dict[str, list[dict[str, object]]],
    ) -> list[dict[str, str]]:
        legacy_rows = self._stage_rows_from_standardized_dataset(dataset, stage_id)
        completion_map = self._uc1_completion_map(
            [
                {clean_cell(key): clean_cell(value) for key, value in row.items()}
                for row in dataset.get("completion_rows", [])
            ]
        )
        bundle = self.load_t1_t5_rule_bundle(
            stage_id,
            allow_saved_rules=self._use_saved_t1_t5_rules(),
        )
        matched_rows: list[dict[str, str]] = []
        for document in self._t1_t5_candidate_workbooks():
            try:
                workbook_sheets = self._uc1_load_excel_rows(document.path)
            except Exception:
                continue
            profile, profile_match = self.t1_t5_executor.resolve_profile(bundle, workbook_sheets=workbook_sheets)
            if profile is None or profile.input_mode != "custom_workbook" or profile_match is None or profile_match.score <= 0.0:
                continue
            generated_rows = self.t1_t5_executor.execute_profile(
                profile,
                workbook_path=document.path,
                workbook_sheets=workbook_sheets,
                profile_match=profile_match,
            )
            if not generated_rows:
                continue
            self.log_debug(
                source="t1t5",
                action="resolve_profile",
                message=f"Matched {stage_id} profile {profile.profile_id} for {document.relative_path}",
                details={
                    "stage_id": stage_id,
                    "profile_id": profile.profile_id,
                    "score": profile_match.score,
                    "matched_sheet_name": profile_match.matched_sheet_name,
                },
            )
            for row in generated_rows:
                matched_rows.append(
                    self._finalize_t1_t5_runtime_row(
                        stage_id,
                        row,
                        completion_map,
                        document=document,
                    )
                )
        if matched_rows:
            return matched_rows

        preview = self.t1_t5_executor.preview(
            bundle,
            input_rows=legacy_rows,
        )
        if preview.output_rows:
            return [
                self._finalize_t1_t5_runtime_row(stage_id, row, completion_map)
                for row in preview.output_rows
            ]
        return legacy_rows

    def _finalize_t1_t5_runtime_row(
        self,
        stage_id: str,
        row: dict[str, str],
        completion_map: dict[str, dict[str, str]],
        *,
        document: DocumentDescriptor | None = None,
    ) -> dict[str, str]:
        source_type = stage_source_type(stage_id)
        merged = {
            clean_cell(key): clean_cell(value)
            for key, value in row.items()
            if not clean_cell(key).startswith("__")
        }
        if document is not None:
            source_root = clean_cell(document.source_root) or self._uc1_source_root_from_path(document.relative_path)
            if not clean_cell(merged.get("source_doc_id", "")):
                merged["source_doc_id"] = self._uc1_document_id(document.relative_path, source_root)
            if not clean_cell(merged.get("source_locator", "")):
                sheet_name = clean_cell(row.get("__sheet_name", ""))
                row_number = clean_cell(row.get("__row_number", ""))
                if sheet_name and row_number:
                    merged["source_locator"] = f"{sheet_name}!row {row_number}"
        if source_type == "pid":
            completion = completion_map.get(self._uc1_identity_key(merged), {})
            if not clean_cell(merged.get("device_id", "")) and clean_cell(merged.get("canonical_tag", "")):
                merged["device_id"] = self._uc1_device_id(clean_cell(merged.get("canonical_tag", "")))
            return self._uc1_finalize_standardized_row(
                "pid",
                merged,
                completion,
                canonical_entity_id=clean_cell(merged.get("device_id", "")) or clean_cell(merged.get("canonical_tag", "")),
                match_method=clean_cell(merged.get("match_method", "")) or ("custom_workbook_profile" if document is not None else ""),
            )
        if not clean_cell(merged.get("presence_status", "")):
            merged["presence_status"] = "present"
        return self._uc1_with_completion_context(source_type, merged, completion_map)

    def generate_aas_batch(
        self,
        workbook_path: Path,
        output_dir: Path | None = None,
        target_format: str = "json",
        progress: ProgressCallback | None = None,
    ) -> list[Path]:
        """Generate AAS for every device row using the batch API.

        Convenience wrapper that delegates to AASGenerationService.generate_batch.
        """
        target_dir = ensure_dir(output_dir or (self.resolve_results_export_dir() / "AAS" / "batch"))
        results = self.aas_generation_service.generate_batch(
            workbook_path,
            target_dir,
            target_format=target_format,
            progress_callback=lambda current, total: self._report_progress(
                progress, round(current * 100 / max(1, total)), f"Generating AAS {current}/{total}"
            ) if progress else None,
        )
        return [r.generated_path for r in results if r.generated_path]

    def generate_aas_from_excel(
        self,
        excel_path: Path,
        progress: ProgressCallback | None = None,
        *,
        output_dir: Path | None = None,
        target_formats: list[str] | None = None,
    ) -> list[Path]:
        target_dir = ensure_dir(output_dir or (self.resolve_results_export_dir() / "AAS"))
        formats = target_formats or ["json", "xml"]
        generated: list[Path] = []
        tasks: list[AASGenerationRequest] = []
        with pd.ExcelFile(excel_path) as workbook:
            if {"ri_devices", "completion_candidates"}.issubset(set(workbook.sheet_names)):
                device_rows = workbook.parse("ri_devices").fillna("").to_dict(orient="records")
                for row in device_rows:
                    source_row_key = clean_cell(row.get("device_id", "") or row.get("canonical_tag", ""))
                    if not source_row_key:
                        continue
                    for target_format in formats:
                        tasks.append(
                            AASGenerationRequest(
                                excel_path=excel_path,
                                output_dir=target_dir / "uc1_standardized",
                                excel_template_type="uc1_standardized_device",
                                source_row_key=source_row_key,
                                target_format=target_format,
                            )
                        )
            else:
                if "stellenplaene_proposal" in workbook.sheet_names:
                    stellen_rows = workbook.parse("stellenplaene_proposal").fillna("").to_dict(orient="records")
                    for row in stellen_rows:
                        source_row_key = clean_cell(row.get("source_row_key", "") or row.get("canonical_tag", "") or row.get("tag", ""))
                        if not source_row_key:
                            continue
                        for target_format in formats:
                            tasks.append(
                                AASGenerationRequest(
                                    excel_path=excel_path,
                                    output_dir=target_dir / "stellenplaene",
                                    excel_template_type="stellenplaene",
                                    source_row_key=source_row_key,
                                    target_format=target_format,
                                )
                            )
                if "verschaltung_proposal" in workbook.sheet_names:
                    versch_rows = workbook.parse("verschaltung_proposal").fillna("").to_dict(orient="records")
                    for row in versch_rows:
                        source_row_key = clean_cell(row.get("source_row_key", "") or row.get("canonical_tag", "") or row.get("plt_stelle", ""))
                        if not source_row_key:
                            continue
                        for target_format in formats:
                            tasks.append(
                                AASGenerationRequest(
                                    excel_path=excel_path,
                                    output_dir=target_dir / "verschaltung",
                                    excel_template_type="verschaltung",
                                    source_row_key=source_row_key,
                                    target_format=target_format,
                                )
                            )
        total = max(1, len(tasks))
        for index, task in enumerate(tasks, start=1):
            result = self.aas_generation_service.generate(task)
            generated.append(result.generated_path)
            self._report_progress(progress, round(index * 100 / total), f"Generated {result.generated_path.name}")
        return generated

    def export_ontology_from_aas_files(
        self,
        aas_paths: list[Path],
        *,
        output_path: Path | None = None,
    ) -> Path:
        json_paths = [path for path in aas_paths if path.suffix.lower() == ".json"]
        if not json_paths:
            raise ValueError("Ontology export requires at least one JSON AAS file.")
        target_path = output_path or ensure_dir(self.resolve_results_export_dir() / "OWL") / "uc1_transformation.owl"
        return self.ontology_export_service.export_from_aas_json(json_paths, target_path)

    def export_use_case_1_ontology_bundle(
        self,
        progress: ProgressCallback | None = None,
    ) -> dict[str, object]:
        workbook_path = self.export_use_case_1_standardized_workbook(
            self._sub_progress(progress, 0, 30) if progress else None,
        )
        aas_bundle = self.generate_use_case_1_aas_models(
            self._sub_progress(progress, 30, 80) if progress else None,
            target_formats=["json"],
        )
        _KNOWN_KEYS = {"pid", "instrument_list", "wiring", "datasheet", "piping", "stromlaufplan"}
        all_aas_paths: list[str] = []
        for k in _KNOWN_KEYS:
            all_aas_paths.extend(aas_bundle.get(k, []))
        aas_paths = [Path(p) for p in all_aas_paths if Path(p).suffix == ".json"]
        ontology_path = self.export_ontology_from_aas_files(aas_paths)
        self._report_progress(progress, 100, f"Exported UC1 ontology bundle to {ontology_path.name}")
        return {
            "workbook_path": str(workbook_path),
            "aas_paths": [str(path) for path in aas_paths],
            "ontology_path": str(ontology_path),
        }

    def export_use_case_1_standardized_workbooks(
        self,
        progress: ProgressCallback | None = None,
        *,
        output_dir: Path | None = None,
    ) -> dict[str, str]:
        aggregated_path = self.export_use_case_1_standardized_workbook(
            self._sub_progress(progress, 0, 55) if progress else None
        )
        sheets = self._uc1_load_excel_rows(aggregated_path)
        completion_map = self._uc1_completion_map(sheets.get("completion_candidates", []))
        ri_device_rows = sheets.get("ri_devices", [])
        split_dir = ensure_dir(output_dir or (self.resolve_results_export_dir() / "UseCase1" / "Standardized"))
        catalog_name = self.uc1_catalog_path().name
        documents_rows = sheets.get("documents", [])
        coverage_rows = sheets.get("catalog_coverage", [])

        instrument_list_rows = [
            self._uc1_with_completion_context("instrument_list", row, completion_map)
            for row in sheets.get("stellenplan_entries", [])
        ]
        wiring_rows = [
            self._uc1_with_completion_context("wiring", row, completion_map)
            for row in sheets.get("wiring_entries", [])
        ]
        datasheet_rows = self._uc1_datasheet_source_rows(ri_device_rows, completion_map)

        piping_rows = [
            self._uc1_with_completion_context("piping", row, completion_map)
            for row in sheets.get("ifc_entries", [])
        ]

        outputs = {
            "aggregated_workbook_path": str(aggregated_path),
            "instrument_list": str(
                self.export_service.export_uc1_source_standardized_workbook(
                    split_dir,
                    workbook_name="standardized_instrument_list.xlsx",
                    workbook_kind="uc1_instrument_list",
                    primary_sheet_name="instrument_list_entries",
                    primary_rows=instrument_list_rows,
                    documents_rows=self._uc1_documents_for_source(documents_rows, "instrument_list"),
                    coverage_rows=self._uc1_coverage_for_source(coverage_rows, "instrument_list"),
                    language=self._ui_language_code(),
                    catalog_path=catalog_name,
                )
            ),
            "wiring": str(
                self.export_service.export_uc1_source_standardized_workbook(
                    split_dir,
                    workbook_name="standardized_wiring.xlsx",
                    workbook_kind="uc1_wiring",
                    primary_sheet_name="wiring_entries",
                    primary_rows=wiring_rows,
                    documents_rows=self._uc1_documents_for_source(documents_rows, "wiring"),
                    coverage_rows=self._uc1_coverage_for_source(coverage_rows, "wiring"),
                    language=self._ui_language_code(),
                    catalog_path=catalog_name,
                )
            ),
            "datasheet": str(
                self.export_service.export_uc1_source_standardized_workbook(
                    split_dir,
                    workbook_name="standardized_datasheet.xlsx",
                    workbook_kind="uc1_datasheet",
                    primary_sheet_name="datasheet_entries",
                    primary_rows=datasheet_rows,
                    documents_rows=self._uc1_documents_for_source(documents_rows, "datasheet"),
                    coverage_rows=self._uc1_coverage_for_source(coverage_rows, "datasheet"),
                    language=self._ui_language_code(),
                    catalog_path=catalog_name,
                )
            ),
            "piping": str(
                self.export_service.export_uc1_source_standardized_workbook(
                    split_dir,
                    workbook_name="standardized_piping.xlsx",
                    workbook_kind="uc1_piping",
                    primary_sheet_name="piping_entries",
                    primary_rows=piping_rows,
                    documents_rows=self._uc1_documents_for_source(documents_rows, "piping"),
                    coverage_rows=self._uc1_coverage_for_source(coverage_rows, "piping"),
                    language=self._ui_language_code(),
                    catalog_path=catalog_name,
                )
            ),
        }
        self._report_progress(progress, 100, f"Exported 5 UC1 standardized workbooks to {split_dir.name}")
        return outputs

    _SOURCE_TYPE_TEMPLATE_SHEET: dict[str, tuple[str, str]] = {
        "pid": ("PID_template.xlsx", "Instrumentation"),
        "instrument_list": ("Stellenplan_template.xlsx", "Instrument_Data"),
        "wiring": ("Klemmenplan_template.xlsx", "Terminal_ID"),
        "datasheet": ("Datasheet_template.xlsx", "Device_ID"),
        "piping": ("PID_template.xlsx", "Piping"),
    }

    # source_type → primary sheet name in exported standardized workbook.
    _SOURCE_TYPE_PRIMARY_SHEET: dict[str, str] = {
        "pid": "ri_devices",
        "instrument_list": "instrument_list_entries",
        "wiring": "wiring_entries",
        "datasheet": "datasheet_entries",
        "piping": "piping_entries",
    }

    def _uc1_rows_from_standardized_workbook(
        self,
        source_type: str,
        workbook_path: Path,
    ) -> list[dict[str, str]]:
        """Extract rows from an exported standardized workbook for AAS generation."""
        sheet_name = self._SOURCE_TYPE_PRIMARY_SHEET.get(source_type)
        if not sheet_name:
            return []
        sheets = self._uc1_load_excel_rows(workbook_path)
        result: list[dict[str, str]] = []
        for row in sheets.get(sheet_name, []):
            cleaned = {clean_cell(k): clean_cell(v) for k, v in row.items()}
            for new_name, old_name in self._COLUMN_ALIASES.items():
                if new_name in cleaned and old_name not in cleaned:
                    cleaned[old_name] = cleaned[new_name]
            result.append(cleaned)
        return result

    # Column name aliases: template → old standardized workbook names.
    # Ensures both Tx rules and AAS fallback builders can find their fields.
    _COLUMN_ALIASES: dict[str, str] = {
        "AKZ_Canonical": "canonical_tag",
        "Instrument_ID": "device_id",
        "Device_ID": "device_id",
        "Terminal_ID": "device_id",
        "CanonicalTag": "canonical_tag",
        "TagName": "label_text",
        "Tag": "tag",
        "Function": "function_code",
        "Source_File": "source_doc_id",
        "Source_Vendor": "vendor_company_name",
        "ManufacturerName": "vendor_company_name",
        "DeviceInformation": "device_information",
        "LabelText": "label_text",
        "FunctionCode": "function_code",
        "SafetyRelevanceClass": "safety_relevance_class",
        "ProcessInstrumentationFunctionNumber": "process_instrumentation_function_number",
        "ProcessInstrumentationFunctionCategory": "process_instrumentation_function_category",
        "ProcessInstrumentationFunctionModifier": "process_instrumentation_function_modifier",
        "ProcessInstrumentationFunctions": "process_instrumentation_functions",
        "HasInstrumentationLoopFunctionNumber": "has_instrumentation_loop_function_number",
        "ActuatingFunctionNumber": "actuating_function_number",
        "ActuatingLocation": "actuating_location",
        "ActuatingSystemNumber": "actuating_system_number",
        "OperatedValveRef": "operated_valve_reference",
        "PipingAnchorId": "piping_anchor_id",
        "PipingComponentName": "piping_component_name",
        "FlowDirection": "flow_direction",
        "LineNumber": "line_number",
        "NominalDiameterValue": "nominal_diameter_numerical_value_representation",
        "NominalDiameterRepr": "nominal_diameter_representation",
        "NominalDiameterStandard": "nominal_diameter_standard",
        "NominalDiameterType": "nominal_diameter_type_representation",
        "FromEquipmentId": "from_equipment_id",
        "ToEquipmentId": "to_equipment_id",
        "SourceDocId": "source_doc_id",
        "SourceLocator": "source_locator",
        "ContextSummary": "context_summary",
        "XSDStatus": "xsd_status",
        "RecommendedAction": "recommended_action",
        "ProposalStatus": "proposal_status",
        "MissingTargets": "missing_targets",
        "NeedsReview": "needs_review",
        "DecisionConfidence": "decision_confidence",
        "EvidenceBundleId": "evidence_bundle_id",
        "UncertaintyReason": "uncertainty_reason",
        "LLMVerificationStatus": "llm_verification_status",
        "RuleSupport": "rule_support",
        "ReviewFeedbackStatus": "review_feedback_status",
        "DecisionTraceJson": "decision_trace_json",
        "Confidence": "confidence",
        "EntryId": "entry_id",
        "PLTStelle": "plt_stelle",
        "Funktion": "funktion",
        "Beschreibung": "beschreibung",
        "ESchrank": "e_schrank",
        "WireLabel": "wire_label",
        "PresenceStatus": "presence_status",
        "DisplayName": "display_name",
        "GlobalId": "global_id",
        "IfcClass": "ifc_class",
        "CabinetId": "cabinet_id",
        # Datasheet-specific template columns
        "Address": "address",
        "Art": "art",
        "Kanal": "kanal",
        "Position": "position",
        "Project": "project",
        "YP": "yp",
    }

    def _uc1_rows_from_template(self, template_path: Path, sheet_name: str) -> list[dict[str, str]]:
        """Read rows from a filled standard template, enriching with UC1 data.

        Column aliases map template PascalCase names to Tx-rule snake_case names.
        UC1 extended columns (canonical_tag, device_id, entry_id, etc.) are
        supplemented from record data when the template exporter hasn't filled them.
        """
        sheets = self._uc1_load_excel_rows(template_path)
        # Build a record lookup keyed by document-level identity
        record_lookup: dict[str, dict[str, str]] = {}
        for record in self.records:
            for res in record.results:
                fn = clean_cell(res.field_name or "")
                if fn in ("tag", "akz", "device", "position") and clean_cell(res.value or ""):
                    key = normalize_identifier(clean_cell(res.value or ""))
                    if key not in record_lookup:
                        record_lookup[key] = {}
                    record_lookup[key][fn] = clean_cell(res.value or "")
                    record_lookup[key]["record_source_path"] = record.source_path or ""

        result: list[dict[str, str]] = []
        _PLACEHOLDER_SIGNALS = {"FK →", "1-based row index", "Fortlaufend", "Fremdschlussel",
                                "e.g.", "Original AKZ", "Normalised canonical", "Unique "}
        for row in sheets.get(sheet_name, []):
            cleaned = {clean_cell(k): clean_cell(v) for k, v in row.items()}
            # Skip placeholder / legend rows
            row_text = " ".join(str(v) for v in cleaned.values())
            if any(s in row_text for s in _PLACEHOLDER_SIGNALS):
                continue
            # Skip completely empty rows
            if not any(v.strip() for v in cleaned.values() if v):
                continue
            # Apply column aliases
            for new_name, old_name in self._COLUMN_ALIASES.items():
                if new_name in cleaned and old_name not in cleaned:
                    cleaned[old_name] = cleaned[new_name]
            # Derive canonical_tag from template fields if UC1 columns are empty.
            # Priority depends on template type (sheet_name).
            if not clean_cell(cleaned.get("canonical_tag", "")):
                inst_id = clean_cell(cleaned.get("instrument_id", "") or cleaned.get("Instrument_ID", ""))
                tag = clean_cell(cleaned.get("tag", "") or cleaned.get("Tag", ""))
                obj_id = clean_cell(cleaned.get("object_id", "") or cleaned.get("Object_ID", ""))
                term_id = clean_cell(cleaned.get("terminal_id", "") or cleaned.get("Terminal_ID", ""))
                if sheet_name == "Instrument_Data":
                    if inst_id:
                        cleaned["canonical_tag"] = normalize_identifier(inst_id.replace("_", "."))
                        cleaned["device_id"] = "urn:ievpi:device:" + normalize_identifier(inst_id)
                        cleaned["tag"] = inst_id.replace("_", ".")
                    elif tag:
                        cleaned["canonical_tag"] = normalize_identifier(tag)
                        cleaned["device_id"] = "urn:ievpi:device:" + normalize_identifier(tag)
                elif sheet_name == "Terminal_ID":
                    # Use Object_ID as device context, Terminal_ID as entry
                    obj = obj_id or "unknown"
                    cleaned["device_id"] = "urn:ievpi:device:" + normalize_identifier(obj)
                    cleaned["canonical_tag"] = normalize_identifier(obj)
                    if term_id:
                        cleaned["plt_stelle"] = normalize_identifier(term_id)
                elif sheet_name == "Device_ID":
                    # Datasheet template: use device_id or tag from row
                    dev_id = clean_cell(cleaned.get("device_id", "") or cleaned.get("Device_ID", ""))
                    if tag and not dev_id:
                        cleaned["canonical_tag"] = normalize_identifier(tag)
                        cleaned["device_id"] = "urn:ievpi:device:" + normalize_identifier(tag)
            if not clean_cell(cleaned.get("entry_id", "")):
                cleaned["entry_id"] = "entry_" + (normalize_identifier(cleaned.get("canonical_tag", "") or cleaned.get("terminal_id", "") or "unknown"))
            if sheet_name == "Terminal_ID" and not clean_cell(cleaned.get("plt_stelle", "")):
                cleaned["plt_stelle"] = normalize_identifier(clean_cell(cleaned.get("terminal_id", "") or cleaned.get("Terminal_ID", "") or "unknown"))
            # Derive device_information from Manufacturer + Model if missing
            if not clean_cell(cleaned.get("device_information", "")):
                mfr = clean_cell(cleaned.get("Manufacturer", "") or cleaned.get("manufacturer", ""))
                mdl = clean_cell(cleaned.get("Model", "") or cleaned.get("model", ""))
                if mfr or mdl:
                    cleaned["device_information"] = f"{mfr} {mdl}".strip()
            if not clean_cell(cleaned.get("presence_status", "")):
                if clean_cell(cleaned.get("canonical_tag", "")) or clean_cell(cleaned.get("tag", "")):
                    cleaned["presence_status"] = "present"
            result.append(cleaned)
        return result

    def generate_use_case_1_aas_models(
        self,
        progress: ProgressCallback | None = None,
        *,
        output_dir: Path | None = None,
        target_formats: list[str] | None = None,
        tx_rule_paths: dict[str, str] | None = None,
        standardized_workbook_paths: dict[str, str] | None = None,
    ) -> dict[str, list[str]]:
        """Generate AAS models for all UC1 source types.

        * **pid**: generates directly from in-memory DEXPI data (no Excel).
        * **instrument_list / wiring / datasheet**: reads from filled
          workbooks in ``data/filled_templates/`` (falls back to blank
          templates in ``data/templates/`` if no filled workbook exists).
        * **piping**: reads from the filled PID template (Piping sheet).
        """
        from iev4pi_transformation_tool.core.standardized_templates import (
            FAMILY_TO_STANDARDIZED_TEMPLATE,
            FILLED_TEMPLATES_DIR,
            STANDARDIZED_TEMPLATE_DIR,
        )

        formats = target_formats or ["json", "xml"]
        aas_root = ensure_dir(output_dir or (self.resolve_results_export_dir() / "AAS"))
        results: dict[str, list[str]] = {
            "pid": [], "instrument_list": [], "wiring": [], "datasheet": [],
            "piping": [], "stromlaufplan": [],
        }

        # ── pid: from aggregated workbook if available, else in-memory DEXPI ──
        source_rows_map: dict[str, list[dict[str, str]]] = {}
        pid_rows: list[dict[str, str]] = []
        agg_path = (standardized_workbook_paths or {}).get("aggregated_workbook_path", "")
        if agg_path and Path(agg_path).is_file():
            try:
                import pandas as pd
                agg_wb = pd.ExcelFile(agg_path)
                if "ri_devices" in agg_wb.sheet_names:
                    pid_rows = agg_wb.parse("ri_devices").fillna("").to_dict(orient="records")
            except Exception:
                pid_rows = []
        if not pid_rows:
            try:
                src = self._uc1_build_source_rows()
                pid_rows = src.get("pid", [])
            except Exception:
                pid_rows = []
        source_rows_map["pid"] = pid_rows

        # ── instrument_list / wiring / datasheet / piping / stromlaufplan: from filled templates ──
        _TEMPLATE_BASED_SOURCE_TYPES = {
            "instrument_list": ("Stellenplan_template.xlsx", "Instrument_Data"),
            "wiring": ("Klemmenplan_template.xlsx", "Terminal_ID"),
            "datasheet": ("Datasheet_template.xlsx", "Device_ID"),
            "piping": ("PID_template.xlsx", "Piping"),
            "stromlaufplan": ("Stromlaufplan_template.xlsx", "Object_ID"),
        }
        # Build completion context from in-memory records (no Excel needed)
        completion_map: dict[str, dict[str, str]] = {}
        try:
            sheet_rows: list[dict[str, str]] = []
            for record in self.records:
                row: dict[str, str] = {}
                for res in record.results:
                    row[clean_cell(res.field_name or "")] = clean_cell(res.value or "")
                if row:
                    sheet_rows.append(row)
            completion_map = self._uc1_completion_map(sheet_rows)
        except Exception:
            pass

        for source_type, (tpl_name, sheet_name) in _TEMPLATE_BASED_SOURCE_TYPES.items():
            # Prefer standardized_workbook_paths if provided
            swb_path = (standardized_workbook_paths or {}).get(source_type)
            if swb_path and Path(swb_path).is_file():
                rows = self._uc1_rows_from_standardized_workbook(source_type, swb_path)
            else:
                # Prefer filled_templates, fallback to blank templates
                tpl_path = FILLED_TEMPLATES_DIR / tpl_name
                if not tpl_path.is_file():
                    tpl_path = STANDARDIZED_TEMPLATE_DIR / tpl_name
                if not tpl_path.is_file():
                    source_rows_map[source_type] = []
                    continue

                if source_type == "stromlaufplan":
                    rows = self._stromlaufplan_rows_for_tx(tpl_path)
                else:
                    rows = self._uc1_rows_from_template(tpl_path, sheet_name)

            source_rows_map[source_type] = [
                self._uc1_with_completion_context(source_type, row, completion_map)
                for row in rows
            ]

        # ── Generate AAS ──
        total_units = 0
        completed_units = 0
        for source_type in list(results.keys()):
            grouped = self._uc1_group_rows_by_identity(source_rows_map.get(source_type, []))
            total_units += len(grouped) or 1
        total_units = max(1, total_units)

        for source_type in list(results.keys()):
            rows = source_rows_map.get(source_type, [])
            if not rows:
                continue
            grouped_rows = self._uc1_group_rows_by_identity(rows)
            for identity_key, identity_rows in grouped_rows.items():
                payload, traces, issues = self._uc1_build_tx_payload(
                    source_type,
                    identity_rows,
                    identity_key,
                    tx_rule_path=Path(tx_rule_paths[source_type]) if tx_rule_paths and tx_rule_paths.get(source_type) else None,
                )
                for path in self._uc1_write_payload_formats(
                    payload,
                    aas_root / source_type,
                    self._uc1_payload_identity(payload),
                    formats,
                ):
                    results[source_type].append(str(path))
                completed_units += 1
                self._report_progress(
                    progress,
                    round(10 + completed_units * 90 / total_units),
                    f"Generated {source_type} AAS for {identity_key}",
                )

        # Warn if template-based sources had no filled data available —
        # stored on the instance so UI / callers can read it without
        # polluting the dict[str, list[str]] return contract.
        _tpl_sources = {"instrument_list", "wiring", "datasheet", "stromlaufplan"}
        _tpl_empty = {
            src for src in _tpl_sources
            if not results.get(src) and not (standardized_workbook_paths or {}).get(src)
        }
        if _tpl_empty and not any(results.get(src) for src in _tpl_sources):
            self._last_uc1_generation_warnings.append(
                "No filled templates found in data/filled_templates/. "
                "Run extraction + fill standardized templates first. "
                f"Missing source types: {', '.join(sorted(_tpl_empty))}."
            )

        self._report_progress(progress, 100, f"Generated UC1 AAS models in {aas_root.name}")
        return results

    def export_use_case_1_source_ontologies(
        self,
        progress: ProgressCallback | None = None,
        *,
        output_dir: Path | None = None,
        standardized_workbook_paths: dict[str, str] | None = None,
    ) -> dict[str, object]:
        _temp_aas_dir = tempfile.mkdtemp(prefix="uc1_aas_")
        try:
            generated_aas = self.generate_use_case_1_aas_models(
                self._sub_progress(progress, 0, 70) if progress else None,
                output_dir=Path(_temp_aas_dir),
                target_formats=["json"],
                standardized_workbook_paths=standardized_workbook_paths,
            )
            ontology_dir = ensure_dir(output_dir or (self.resolve_results_export_dir() / "OWL"))
            file_map = {
                "pid": "t6_pid.owl",
                "instrument_list": "t7_instrument_list.owl",
                "wiring": "t8_wiring.owl",
                "datasheet": "t9_datasheet.owl",
                "piping": "t10_piping.owl",
                "stromlaufplan": "t11_stromlaufplan.owl",
            }
            ontology_paths: dict[str, str] = {}
            total = max(1, len(file_map))
            for index, (source_type, file_name) in enumerate(file_map.items(), start=1):
                json_paths = [Path(path) for path in generated_aas.get(source_type, []) if Path(path).suffix.lower() == ".json"]
                if not json_paths:
                    continue
                path = self.ontology_export_service.export_from_aas_json(
                    json_paths,
                    ontology_dir / file_name,
                    source_type=source_type,
                )
                ontology_paths[source_type] = str(path)
                self._report_progress(progress, 70 + round(index * 30 / total), f"Exported ontology {path.name}")

            return {
                "aas_paths_by_source": {},
                "ontology_paths": ontology_paths,
            }
        finally:
            shutil.rmtree(_temp_aas_dir, ignore_errors=True)

    def export_use_case_1_transformation_bundle(
        self,
        progress: ProgressCallback | None = None,
    ) -> dict[str, object]:
        aas_paths_by_source = self.generate_use_case_1_aas_models(
            self._sub_progress(progress, 0, 60) if progress else None,
            target_formats=["json"],
        )
        ontology_bundle = self.export_use_case_1_source_ontologies(
            self._sub_progress(progress, 60, 100) if progress else None,
        )
        return {
            "aas_paths_by_source": aas_paths_by_source,
            "ontology_paths": ontology_bundle.get("ontology_paths", {}),
        }

    def _uc1_device_id(self, canonical_tag: str) -> str:
        normalized = normalize_identifier(canonical_tag) or "device"
        return f"urn:ievpi:device:{normalized}"

    def _uc1_document_id(self, source_path: str, source_root: str = "") -> str:
        cleaned = clean_cell(source_path)
        normalized_root = normalize_identifier(source_root or self._uc1_source_root_from_path(cleaned)) or "source"
        digest = sha1(cleaned.encode("utf-8")).hexdigest()[:12] if cleaned else "unknown"
        return f"urn:ievpi:document:{normalized_root}:{digest}"

    def _uc1_relation_id(self, *parts: str) -> str:
        joined = "::".join(clean_cell(part) for part in parts if clean_cell(part))
        digest = sha1(joined.encode("utf-8")).hexdigest()[:16] if joined else "unknown"
        return f"urn:ievpi:relation:{digest}"

    def _uc1_source_root_from_path(self, source_path: str) -> str:
        for root_name in ("R&I-Fließbild", "Stellenplaene", "Verschaltungslisten", "Diagram-PDF", "IFC"):
            if root_name in source_path:
                return root_name
        return ""

    def _uc1_evidence_locator(self, evidence_refs: list[EvidenceRef]) -> str:
        if not evidence_refs:
            return ""
        first_evidence = evidence_refs[0]
        return clean_cell(first_evidence.cell_range_or_bbox)

    def _uc1_first_value(self, *values: str) -> str:
        for value in values:
            cleaned = clean_cell(value)
            if cleaned:
                return cleaned
        return ""

    @staticmethod
    def _uc1_clean_siemens_mlfb(value: str) -> str:
        """Fix common OCR errors in Siemens MLFB part numbers (e.g. Q→0)."""
        import re
        # Fix trailing Q that should be 0 in patterns like 0AA0
        # Fix common OCR error: trailing Q in Siemens MLFB pattern (e.g. 0AAQ → 0AA0)
        value = re.sub(r"0AAQ$", "0AA0", value)
        return value

    def _uc1_attr_value(self, attributes: dict[str, str], *keys: str) -> str:
        normalized_map = {
            normalize_identifier(key): clean_cell(value)
            for key, value in attributes.items()
            if clean_cell(key) and clean_cell(value)
        }
        for key in keys:
            direct = clean_cell(attributes.get(key, ""))
            if direct:
                return direct
            normalized = normalized_map.get(normalize_identifier(key), "")
            if normalized:
                return normalized
        return ""

    def _uc1_load_excel_rows(self, workbook_path: Path) -> dict[str, list[dict[str, str]]]:
        sheets: dict[str, list[dict[str, str]]] = {}
        with pd.ExcelFile(workbook_path) as workbook:
            for sheet_name in workbook.sheet_names:
                frame = workbook.parse(sheet_name, dtype=object)
                frame = frame.where(pd.notna(frame), "")
                sheets[sheet_name] = [
                    {clean_cell(key): clean_cell(value) for key, value in row.items()}
                    for row in frame.astype(str).replace({"nan": ""}).to_dict(orient="records")
                ]
        return sheets

    def _uc1_identity_key(self, row: dict[str, str]) -> str:
        for key in (
            # template column names (current)
            "AKZ_Canonical", "Instrument_ID", "Device_ID", "Terminal_ID",
            "CanonicalTag", "GlobalId", "Tag", "PLTStelle", "EntryId",
            # old standardized workbook column names (legacy Tx rules)
            "device_id", "canonical_tag", "global_id", "tag", "plt_stelle", "entry_id",
            # Stromlaufplan identity
            "object_id", "document_id",
        ):
            normalized = normalize_identifier(clean_cell(row.get(key, "")))
            if normalized:
                return normalized
        return ""

    def _uc1_external_identity_keys(self, row: dict[str, str]) -> list[str]:
        keys: list[str] = []
        for field_name in (
            # template column names
            "AKZ_Canonical", "CanonicalTag",
            "ProcessInstrumentationFunctionNumber",
            "HasInstrumentationLoopFunctionNumber",
            "Tag", "TagName",
            # old column names (legacy)
            "canonical_tag",
            "process_instrumentation_function_number",
            "has_instrumentation_loop_function_number",
            "tag",
            "logical_tag",
            "messstelle",
            "signal_tag",
            "plt_stelle",
            "display_name",
        ):
            normalized = normalize_identifier(clean_cell(row.get(field_name, "")))
            if normalized and normalized not in keys:
                keys.append(normalized)
        return keys

    def _uc1_completion_map(self, rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
        mapped: dict[str, dict[str, str]] = {}
        for row in rows:
            key = self._uc1_identity_key(row)
            if key:
                mapped[key] = row
        return mapped

    def _uc1_boolish(self, value: str) -> bool:
        return clean_cell(value).strip().lower() in {"1", "true", "yes", "present", "complete", "matched"}

    def _uc1_score_text(self, value: str | float | int | None, default: float = 0.0) -> str:
        candidate = clean_cell(value)
        if candidate:
            try:
                return f"{float(candidate):.4f}".rstrip("0").rstrip(".")
            except ValueError:
                return candidate
        return f"{default:.4f}".rstrip("0").rstrip(".")

    def _uc1_decision_trace_json(self, decision_trace: dict[str, ConsistencyDecision] | dict[str, object] | None) -> str:
        if not decision_trace:
            return ""
        payload = {
            clean_cell(source_key): (
                decision.model_dump(mode="json") if isinstance(decision, ConsistencyDecision) else decision
            )
            for source_key, decision in decision_trace.items()
            if clean_cell(source_key)
        }
        return json.dumps(payload, ensure_ascii=False, sort_keys=True)

    def _uc1_constraint_violations(
        self,
        source_type: str,
        row: dict[str, str],
        completion: dict[str, str],
    ) -> str:
        violations: list[str] = []
        if source_type == "piping":
            presence = clean_cell(row.get("presence_status", ""))
            match_method = clean_cell(row.get("match_method", ""))
            if presence not in ("not_required", "missing") and match_method != "assembly_mapping_evidence":
                flange_complete = clean_cell(row.get("flange_complete", "") or completion.get("flange_complete", ""))
                if flange_complete and not self._uc1_boolish(flange_complete):
                    violations.append("ifc_flange_incomplete")
        missing_targets = clean_cell(completion.get("missing_targets", ""))
        if missing_targets:
            violations.append(f"missing_targets:{missing_targets}")
        return " | ".join(item for item in violations if item)

    def _uc1_finalize_standardized_row(
        self,
        source_type: str,
        row: dict[str, str],
        completion: dict[str, str] | None = None,
        *,
        canonical_entity_id: str = "",
        match_confidence: str | float | int | None = None,
        match_method: str = "",
        needs_review_reason: str = "",
        constraint_violations: str = "",
    ) -> dict[str, str]:
        merged = {clean_cell(key): clean_cell(value) for key, value in row.items()}
        completion_row = completion or {}
        merged["canonical_entity_id"] = clean_cell(canonical_entity_id) or clean_cell(merged.get("device_id", "")) or clean_cell(merged.get("canonical_tag", ""))
        presence_status = clean_cell(merged.get("presence_status", ""))
        default_score = 1.0 if presence_status == "present" else 0.0
        merged["match_confidence"] = self._uc1_score_text(match_confidence if match_confidence is not None else merged.get("confidence", ""), default_score)
        if clean_cell(match_method):
            merged["match_method"] = clean_cell(match_method)
        elif not clean_cell(merged.get("match_method", "")):
            merged["match_method"] = "missing_placeholder" if presence_status == "missing" else "strict_source_link"
        if clean_cell(needs_review_reason):
            merged["needs_review_reason"] = clean_cell(needs_review_reason)
        elif not clean_cell(merged.get("needs_review_reason", "")):
            if clean_cell(merged.get("needs_review", "")) == "true":
                if presence_status == "missing":
                    merged["needs_review_reason"] = "source_missing"
                elif float(self._uc1_score_text(merged.get("match_confidence", ""), default_score) or "0") < 0.95:
                    merged["needs_review_reason"] = "low_confidence_alignment"
                else:
                    merged["needs_review_reason"] = "manual_verification_required"
            else:
                merged["needs_review_reason"] = ""
        if clean_cell(constraint_violations):
            merged["constraint_violations"] = clean_cell(constraint_violations)
        elif not clean_cell(merged.get("constraint_violations", "")):
            merged["constraint_violations"] = self._uc1_constraint_violations(source_type, merged, completion_row)
        if not clean_cell(merged.get("decision_confidence", "")):
            merged["decision_confidence"] = self._uc1_score_text(
                merged.get("match_confidence", "") or merged.get("confidence", ""),
                default_score,
            )
        if not clean_cell(merged.get("evidence_bundle_id", "")):
            merged["evidence_bundle_id"] = clean_cell(completion_row.get("evidence_bundle_id", ""))
        if not clean_cell(merged.get("uncertainty_reason", "")):
            merged["uncertainty_reason"] = clean_cell(completion_row.get("uncertainty_reason", ""))
        if not clean_cell(merged.get("llm_verification_status", "")):
            merged["llm_verification_status"] = clean_cell(completion_row.get("llm_verification_status", ""))
        if not clean_cell(merged.get("rule_support", "")):
            merged["rule_support"] = clean_cell(completion_row.get("rule_support", ""))
        if not clean_cell(merged.get("review_feedback_status", "")):
            merged["review_feedback_status"] = clean_cell(completion_row.get("review_feedback_status", ""))
        if not clean_cell(merged.get("decision_trace_json", "")):
            merged["decision_trace_json"] = clean_cell(completion_row.get("decision_trace_json", ""))
        if not clean_cell(merged.get("SemanticID", "")):
            merged["SemanticID"] = self._uc1_semantic_id(source_type)
        return merged

    def _uc1_semantic_id(self, source_type: str) -> str:
        """Return a standard IRDI for the given source type."""
        _SOURCE_SEMANTIC = {
            "instrument_list": ("SM_InstrumentListEntry", "canonicalTag"),
            "wiring": ("SM_WiringEntry", "DeviceId"),
            "datasheet": ("SM_DatasheetEntry", "canonicalTag"),
            "pid": ("SM_CoreIdentity", "canonicalTag"),
            "piping": ("SM_ActuationAndPiping", "LineNumber"),
        }
        submodel, prop = _SOURCE_SEMANTIC.get(source_type, ("SM_CoreIdentity", "canonicalTag"))
        return get_irdi(submodel, prop) or f"urn:ievpi:semantic:{source_type}"

    def _uc1_with_completion_context(
        self,
        source_type: str,
        row: dict[str, str],
        completion_map: dict[str, dict[str, str]],
    ) -> dict[str, str]:
        merged = {clean_cell(key): clean_cell(value) for key, value in row.items()}
        completion = completion_map.get(self._uc1_identity_key(merged), {})
        for key in (
            "recommended_action",
            "proposal_status",
            "missing_targets",
            "decision_confidence",
            "evidence_bundle_id",
            "uncertainty_reason",
            "llm_verification_status",
            "rule_support",
            "review_feedback_status",
            "decision_trace_json",
        ):
            if clean_cell(completion.get(key, "")) and not clean_cell(merged.get(key, "")):
                merged[key] = clean_cell(completion.get(key, ""))
        if not clean_cell(merged.get("confidence", "")) and clean_cell(merged.get("presence_status", "")) == "present":
            merged["confidence"] = "1.0"
        if source_type == "piping" and clean_cell(completion.get("flange_complete", "")):
            merged["flange_complete"] = clean_cell(completion.get("flange_complete", ""))
        present_field_map = {
            "instrument_list": "present_in_stellenplan",
            "wiring": "present_in_wiring",
            "datasheet": "present_in_datasheet",
            "piping": "present_in_ifc",
        }
        present_field = present_field_map.get(source_type, "")
        present_value = clean_cell(completion.get(present_field, "")) if present_field else ""
        needs_review = clean_cell(merged.get("presence_status", "")) != "present"
        if present_value:
            needs_review = needs_review or not self._uc1_boolish(present_value)
        if source_type == "piping":
            flange_complete = clean_cell(completion.get("flange_complete", "") or merged.get("flange_complete", ""))
            if flange_complete:
                needs_review = needs_review or not self._uc1_boolish(flange_complete)
        merged["needs_review"] = "true" if needs_review else "false"
        return self._uc1_finalize_standardized_row(
            source_type,
            merged,
            completion,
        )

    def _uc1_documents_for_source(
        self,
        document_rows: list[dict[str, str]],
        source_type: str,
    ) -> list[dict[str, str]]:
        def belongs(row: dict[str, str]) -> bool:
            source_family = clean_cell(row.get("source_family", "")).lower()
            original_path = clean_cell(row.get("original_path", ""))
            if source_type == "instrument_list":
                return source_family == SourceDocumentKind.STELLEN_TU.value or original_path.startswith("Stellenplaene/")
            if source_type == "datasheet":
                return source_family == SourceDocumentKind.STELLEN_TU.value or "Stellenplaene/" in original_path
            if source_type == "wiring":
                return source_family in {
                    SourceDocumentKind.STROMLAUFPLAN.value,
                    SourceDocumentKind.VERSCHALTUNGSLISTE.value,
                    SourceDocumentKind.CABINET_REFERENCE.value,
                    SourceDocumentKind.KLEMMENPLAN.value,
                } or original_path.startswith("Verschaltungslisten/") or original_path.startswith("Stellenplaene/")
            if source_type == "piping":
                return source_family == SourceDocumentKind.IFC_MODEL.value or original_path.startswith("IFC/")
            return False

        return [row for row in document_rows if belongs(row)]

    def _uc1_coverage_for_source(
        self,
        coverage_rows: list[dict[str, str]],
        source_type: str,
    ) -> list[dict[str, str]]:
        selected: list[dict[str, str]] = []
        for row in coverage_rows:
            matched_field = normalize_identifier(clean_cell(row.get("matched_field", "")))
            document = normalize_identifier(clean_cell(row.get("document", "")))
            class_name = normalize_identifier(clean_cell(row.get("class_name", "")))
            data_property = normalize_identifier(clean_cell(row.get("data_property", "")))

            if source_type == "instrument_list":
                keep = matched_field.startswith("stellenplanentries") or "stellenplan" in document or class_name == "stellenplanentry"
            elif source_type == "wiring":
                keep = matched_field.startswith("wiringentries") or "wiring" in class_name or "wiring" in data_property
            elif source_type == "datasheet":
                keep = (
                    matched_field.startswith("datasheetentries")
                    or class_name == "datasheetentry"
                    or "spezifikation" in document
                    or data_property in {"vendorcompanyname", "safetyrelevanceclass", "deviceinformation"}
                )
            elif source_type == "piping":
                keep = (
                    matched_field.startswith("ifcentries")
                    or matched_field.startswith("relations")
                    or "ifc" in document
                    or class_name in {"ifcpipesegment", "ifcvalve", "ifcactuator", "pipingnetworksegment"}
                )
            else:
                keep = False
            if keep:
                selected.append(row)
        return selected

    def _uc1_datasheet_source_rows(
        self,
        ri_device_rows: list[dict[str, str]],
        completion_map: dict[str, dict[str, str]],
    ) -> list[dict[str, str]]:
        datasheet_index: dict[str, list[ExtractedRecord]] = defaultdict(list)
        all_datasheet_records: list[ExtractedRecord] = []
        datasheet_candidate_rows: list[dict[str, str]] = []
        for record in self.records:
            if record.family != DocumentFamily.STELLEN_TU_DATASHEET:
                continue
            all_datasheet_records.append(record)
            result_map = self._record_result_map(record)
            tag_value = self._uc1_first_value(
                result_map.get("tag", ""),
                result_map.get("logical_tag", ""),
                result_map.get("messstelle", ""),
                result_map.get("signal_tag", ""),
                clean_cell(record.display_name),
            )
            candidate_row = {
                "record_key": record.record_key,
                "canonical_tag": tag_value,
                "tag": tag_value,
                "device_information": self._uc1_clean_siemens_mlfb(self._uc1_first_value(
                    result_map.get("device_information", ""),
                    result_map.get("device", ""),
                )),
                "art": clean_cell(result_map.get("art", "")),
                "kanal": clean_cell(result_map.get("kanal", "")),
                "yp": clean_cell(result_map.get("yp", "")),
                "position": clean_cell(result_map.get("position", "")),
                "address": self._uc1_first_value(
                    result_map.get("address", ""),
                    result_map.get("adresse", ""),
                    result_map.get("dresse", ""),
                ),
                "project": self._uc1_first_value(
                    result_map.get("project", ""),
                    result_map.get("projekt", ""),
                ),
                "display_name": clean_cell(record.display_name),
            }
            datasheet_candidate_rows.append(candidate_row)
            for candidate in (
                self._record_value(record, "tag", "logical_tag", "messstelle", "signal_tag"),
                clean_cell(record.display_name),
            ):
                key = normalize_identifier(candidate)
                if key and record not in datasheet_index[key]:
                    datasheet_index[key].append(record)

        datasheet_resolution_map = self.entity_resolver.resolve(datasheet_candidate_rows, ri_device_rows)

        rows: list[dict[str, str]] = []
        used_record_keys: set[str] = set()
        for device_row in ri_device_rows:
            device_id = clean_cell(device_row.get("device_id", ""))
            canonical_tag = clean_cell(device_row.get("canonical_tag", ""))
            matches: list[ExtractedRecord] = []
            seen_record_keys: set[str] = set()
            for key in self._uc1_external_identity_keys(device_row):
                for record in datasheet_index.get(key, []):
                    if record.record_key in seen_record_keys:
                        continue
                    seen_record_keys.add(record.record_key)
                    matches.append(record)
            if not matches:
                rows.append(
                    self._uc1_with_completion_context(
                        "datasheet",
                        {
                            "entry_id": self._uc1_relation_id(device_id, "datasheet"),
                            "device_id": device_id,
                            "canonical_tag": canonical_tag,
                            "tag": canonical_tag,
                            "device_information": "",
                            "art": "",
                            "kanal": "",
                            "yp": "",
                            "position": "",
                            "address": "",
                            "project": "",
                            "source_doc_id": "",
                            "source_locator": "",
                            "confidence": "",
                            "presence_status": "missing",
                            "record_key": "",
                            "display_name": "",
                            "match_method": "missing_placeholder",
                            "match_confidence": "0.0",
                            "canonical_entity_id": device_id,
                            "needs_review_reason": "datasheet_missing",
                        },
                        completion_map,
                    )
                )
                continue

            for record in matches:
                used_record_keys.add(record.record_key)
                result_map = self._record_result_map(record)
                trace_metadata = self._record_trace_metadata(record)
                first_evidence = next((evidence for result in record.results for evidence in result.evidence_refs), None)
                source_root = record.source_root or self.family_source_root(record.family)
                resolution = datasheet_resolution_map.get(record.record_key)
                constraint_violations = ""
                if resolution is not None and resolution.target_key and resolution.target_key != device_id:
                    constraint_violations = "resolver_disagrees_with_strict_match"
                rows.append(
                    self._uc1_with_completion_context(
                        "datasheet",
                        {
                            "entry_id": self._uc1_relation_id(device_id, record.record_key, "datasheet"),
                            "device_id": device_id,
                            "canonical_tag": canonical_tag,
                            "tag": self._uc1_first_value(
                                result_map.get("tag", ""),
                                result_map.get("logical_tag", ""),
                                result_map.get("messstelle", ""),
                                result_map.get("signal_tag", ""),
                                canonical_tag,
                            ),
                            "device_information": self._uc1_clean_siemens_mlfb(self._uc1_first_value(
                                result_map.get("device_information", ""),
                                result_map.get("device", ""),
                            )),
                            "art": clean_cell(result_map.get("art", "")),
                            "kanal": clean_cell(result_map.get("kanal", "")),
                            "yp": clean_cell(result_map.get("yp", "")),
                            "position": clean_cell(result_map.get("position", "")),
                            "address": self._uc1_first_value(
                                result_map.get("address", ""),
                                result_map.get("adresse", ""),
                                result_map.get("dresse", ""),
                            ),
                            "project": self._uc1_first_value(
                                result_map.get("project", ""),
                                result_map.get("projekt", ""),
                            ),
                            "source_doc_id": self._uc1_document_id(clean_cell(record.source_path), source_root),
                            "source_locator": clean_cell(first_evidence.cell_range_or_bbox) if first_evidence is not None else "",
                            "confidence": "1.0",
                            "presence_status": "present",
                            "record_key": record.record_key,
                            "display_name": record.display_name,
                            "match_method": "strict_external_match",
                            "match_confidence": "1.0",
                            "canonical_entity_id": device_id,
                            "needs_review_reason": "",
                            "constraint_violations": constraint_violations,
                            **trace_metadata,
                        },
                        completion_map,
                    )
                )
        for record in all_datasheet_records:
            if record.record_key in used_record_keys:
                continue
            result_map = self._record_result_map(record)
            tag_value = self._uc1_first_value(
                result_map.get("tag", ""),
                result_map.get("logical_tag", ""),
                result_map.get("messstelle", ""),
                result_map.get("signal_tag", ""),
                clean_cell(record.display_name),
            )
            identity_key = normalize_identifier(tag_value) or normalize_identifier(record.record_key) or "datasheet"
            device_id = f"urn:ievpi:device:{identity_key}"
            first_evidence = next((evidence for result in record.results for evidence in result.evidence_refs), None)
            source_root = record.source_root or self.family_source_root(record.family)
            resolution = datasheet_resolution_map.get(record.record_key)
            trace_metadata = self._record_trace_metadata(record)
            canonical_entity_id = clean_cell(resolution.target_key) if resolution is not None else ""
            match_confidence = resolution.score if resolution is not None else 0.0
            match_method = resolution.method if resolution is not None else "unmatched"
            needs_review_reason = resolution.needs_review_reason if resolution is not None else "datasheet_only_candidate"
            if not canonical_entity_id:
                canonical_entity_id = device_id
            rows.append(
                self._uc1_with_completion_context(
                    "datasheet",
                    {
                        "entry_id": self._uc1_relation_id(device_id, record.record_key, "datasheet"),
                        "device_id": device_id,
                        "canonical_tag": tag_value,
                        "tag": tag_value,
                        "device_information": self._uc1_first_value(
                            result_map.get("device_information", ""),
                            result_map.get("device", ""),
                        ),
                        "art": clean_cell(result_map.get("art", "")),
                        "kanal": clean_cell(result_map.get("kanal", "")),
                        "yp": clean_cell(result_map.get("yp", "")),
                        "position": clean_cell(result_map.get("position", "")),
                        "address": self._uc1_first_value(
                            result_map.get("address", ""),
                            result_map.get("adresse", ""),
                            result_map.get("dresse", ""),
                        ),
                        "project": self._uc1_first_value(
                            result_map.get("project", ""),
                            result_map.get("projekt", ""),
                        ),
                        "source_doc_id": self._uc1_document_id(clean_cell(record.source_path), source_root),
                        "source_locator": clean_cell(first_evidence.cell_range_or_bbox) if first_evidence is not None else "",
                        "confidence": "1.0",
                        "presence_status": "present",
                        "record_key": record.record_key,
                        "display_name": record.display_name,
                        "match_method": match_method,
                        "match_confidence": self._uc1_score_text(match_confidence, 0.0),
                        "canonical_entity_id": canonical_entity_id,
                        "needs_review_reason": needs_review_reason,
                        **trace_metadata,
                    },
                    completion_map,
                )
            )
        return rows

    def _uc1_primary_sheet_name(self, source_type: str) -> str:
        return {
            "pid": "ri_devices",
            "instrument_list": "instrument_list_entries",
            "wiring": "wiring_entries",
            "datasheet": "datasheet_entries",
            "piping": "piping_entries",
        }[source_type]

    def _stromlaufplan_rows_for_tx(self, template_path: Path) -> list[dict[str, str]]:
        """Flatten the Stromlaufplan workbook into TX-consumable rows.

        Merges Object_ID + Element_ID + Element_Data + Document_Data into
        one row per object with embedded element/attribute/connection data.
        """
        import openpyxl
        wb = openpyxl.load_workbook(str(template_path), read_only=True, data_only=True)
        try:
            # Read all sheets into memory
            def _read_sheet(name: str) -> list[dict[str, str]]:
                if name not in wb.sheetnames:
                    return []
                ws = wb[name]
                headers = []
                rows = []
                for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    vals = [str(v) if v is not None else "" for v in row]
                    if r_idx == 1:
                        headers = [clean_cell(h) for h in vals]
                        continue
                    if r_idx == 2:
                        continue  # skip legend row
                    if not any(v.strip() for v in vals):
                        continue
                    row_dict = {}
                    for i, h in enumerate(headers):
                        if i < len(vals):
                            row_dict[h] = vals[i]
                    rows.append(row_dict)
                return rows

            doc_data = {r.get("Document_ID", ""): r for r in _read_sheet("Document_Data")}
            obj_rows = _read_sheet("Object_ID")
            elem_rows = _read_sheet("Element_ID")
            ed_rows = _read_sheet("Element_Data")
            conn_rows = _read_sheet("Connection_Data")

            # Index elements by Element_ID
            elem_by_id: dict[str, dict] = {}
            for er in elem_rows:
                eid = er.get("Element_ID", "")
                if eid:
                    elem_by_id[eid] = er

            # Index element data by Element_ID
            ed_by_elem: dict[str, list[dict]] = {}
            for ed in ed_rows:
                eid = ed.get("Element_ID", "")
                if eid:
                    ed_by_elem.setdefault(eid, []).append(ed)

            result: list[dict[str, str]] = []
            for obj in obj_rows:
                doc_id = obj.get("Document_ID", "")
                oid = obj.get("Object_ID", "")
                ref_data = obj.get("Object_Reference_Data", "")
                dd = doc_data.get(doc_id, {})
                # Generate stable identity for AAS grouping
                entry_id = f"{doc_id}_{oid}"
                canonical_tag = ref_data.lstrip("-") if ref_data.startswith("-") else ref_data
                row = {
                    "document_id": doc_id,
                    "object_id": oid,
                    "entry_id": entry_id,
                    "canonical_tag": canonical_tag,
                    "device_id": oid,
                    "object_reference_data": ref_data,
                    "sheet_number": dd.get("Sheet_Number", ""),
                    "sheet_name": dd.get("Sheet_Name", ""),
                    "project": dd.get("Project_Entry", ""),
                    "project_nr": dd.get("Project_Nr_Entry", ""),
                    "location": dd.get("Location_Entry", ""),
                    "date": dd.get("Date_Entry", ""),
                    "author": dd.get("Author_Entry", ""),
                }
                # Count elements for this object
                obj_elements = [e for e in elem_rows if e.get("Object_ID", "") == oid]
                row["element_count"] = str(len(obj_elements))
                # Collect element IDs
                row["element_ids"] = " | ".join(
                    e.get("Element_ID", "") for e in obj_elements[:10]
                )
                # Collect attribute names
                obj_ed = [ed for ed in ed_rows if ed.get("Object_ID", "") == oid
                          or any(e.get("Element_ID", "") == ed.get("Element_ID", "")
                                for e in obj_elements)]
                attrs = list(dict.fromkeys(ed.get("Attribute_Name", "") for ed in obj_ed if ed.get("Attribute_Name")))
                row["attribute_names"] = " | ".join(attrs[:10])
                # Connection count
                obj_conns = [c for c in conn_rows if c.get("From_Element_ID", "").
                             startswith(oid) or c.get("To_Element_ID", "").startswith(oid)]
                row["connection_count"] = str(len(obj_conns))
                result.append(row)
            return result
        finally:
            wb.close()

    def _uc1_group_rows_by_identity(self, rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
        grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
        for row in rows:
            key = self._uc1_identity_key(row)
            if key:
                grouped[key].append(row)
        return grouped

    def _uc1_join_values(self, rows: list[dict[str, str]], *field_names: str) -> str:
        values: list[str] = []
        for row in rows:
            for field_name in field_names:
                value = clean_cell(row.get(field_name, ""))
                if value and value not in values:
                    values.append(value)
        return " | ".join(values)

    def _uc1_make_aas_payload(
        self,
        source_type: str,
        identity_value: str,
        submodel_specs: list[tuple[str, dict[str, str]]],
    ) -> dict[str, object]:
        identity = normalize_identifier(identity_value) or "entry"
        shell_id = f"urn:ievpi:aas:{source_type}:{identity}"
        asset_id = f"urn:ievpi:asset:{source_type}:{identity}"
        submodels: list[dict[str, object]] = []
        submodel_refs: list[dict[str, object]] = []
        for submodel_name, properties in submodel_specs:
            submodel_id = f"{shell_id}:{normalize_identifier(submodel_name)}"
            submodels.append(
                {
                    "modelType": "Submodel",
                    "id": submodel_id,
                    "idShort": submodel_name,
                    "submodelElements": [
                        {
                            "modelType": "Property",
                            "idShort": property_name,
                            "valueType": "xs:boolean" if clean_cell(property_value).lower() in {"true", "false"} else "xs:string",
                            "value": clean_cell(property_value),
                        }
                        for property_name, property_value in properties.items()
                    ],
                }
            )
            submodel_refs.append({"type": "ModelReference", "keys": [{"type": "Submodel", "value": submodel_id}]})

        return {
            "x-ievpi-source_type": source_type,
            "assetAdministrationShells": [
                {
                    "modelType": "AssetAdministrationShell",
                    "id": shell_id,
                    "idShort": identity_value.replace(".", "_").replace("-", "_"),
                    "assetInformation": {"assetKind": "Instance", "globalAssetId": asset_id},
                    "submodels": submodel_refs,
                }
            ],
            "submodels": submodels,
            "conceptDescriptions": [],
        }

    def _uc1_build_tx_payload(
        self,
        source_type: str,
        rows: list[dict[str, str]],
        identity_key: str,
        *,
        workbook_path: Path | None = None,
        tx_rule_path: Path | None = None,
        tx_rule_set_id: str = "",
        rule_payload: dict[str, object] | None = None,
    ) -> tuple[dict[str, object], list[object], list[TxValidationIssue]]:
        rule_set = TxRuleSet.model_validate(rule_payload) if rule_payload else self.load_tx_rule_set(
            source_type,
            tx_rule_path=tx_rule_path,
            tx_rule_set_id=tx_rule_set_id,
            allow_saved_rules=self._use_saved_tx_rules(),
        )
        issues = self.tx_executor.validate(rule_set)
        if any(issue.severity == "error" for issue in issues):
            fallback = self._uc1_build_pid_aas_payload(rows[0], rows[0]) if source_type == "pid" else self._uc1_build_source_aas_payload(source_type, rows, identity_key)
            fallback["x-ievpi-tx-fallback"] = True
            fallback["x-ievpi-tx-validation_issues"] = [issue.model_dump(mode="json") for issue in issues]
            return fallback, [], issues
        try:
            identity_value = identity_key
            if source_type == "pid" and rows:
                identity_value = clean_cell(rows[0].get("canonical_tag", "") or rows[0].get("device_id", "") or identity_key)
            payload, traces = self.tx_executor.execute(
                rule_set,
                rows,
                identity_value=identity_value,
                workbook_path=workbook_path,
                source_type=source_type,
            )
            payload["x-ievpi-tx-traces"] = [trace.model_dump(mode="json") for trace in traces]
            if tx_rule_path is not None:
                payload["x-ievpi-tx-rule_path"] = str(tx_rule_path)
            elif self.tx_rule_store.exists(source_type=source_type, rule_set_id=tx_rule_set_id):
                payload["x-ievpi-tx-rule_path"] = str(self.tx_rule_store.path_for(source_type, tx_rule_set_id))
            return payload, traces, issues
        except Exception as exc:
            failure_issue = TxValidationIssue(
                severity="warning",
                code="tx_execution_failed",
                message=str(exc),
            )
            issues = [*issues, failure_issue]
            fallback = self._uc1_build_pid_aas_payload(rows[0], rows[0]) if source_type == "pid" else self._uc1_build_source_aas_payload(source_type, rows, identity_key)
            fallback["x-ievpi-tx-fallback"] = True
            fallback["x-ievpi-tx-validation_issues"] = [issue.model_dump(mode="json") for issue in issues]
            return fallback, [], issues

    def _uc1_build_pid_aas_payload(
        self,
        row: dict[str, str],
        completion: dict[str, str],
    ) -> dict[str, object]:
        canonical_tag = clean_cell(row.get("canonical_tag", ""))
        identity_value = canonical_tag or clean_cell(row.get("device_id", "")) or "pid"
        return self._uc1_make_aas_payload(
            "pid",
            identity_value,
            [
                (
                    "SM_CoreIdentity",
                    {
                        "canonicalTag": canonical_tag,
                        "DeviceId": clean_cell(row.get("device_id", "")),
                        "hasInstrumentationLoopFunctionNumber": clean_cell(row.get("has_instrumentation_loop_function_number", "")),
                        "ProcessInstrumentationFunctionNumber": clean_cell(row.get("process_instrumentation_function_number", "")),
                        "ProcessInstrumentationFunctionCategory": clean_cell(row.get("process_instrumentation_function_category", "")),
                        "ProcessInstrumentationFunctionModifier": clean_cell(row.get("process_instrumentation_function_modifier", "")),
                        "ProcessInstrumentationFunctions": clean_cell(row.get("process_instrumentation_functions", "")),
                        "ContextSummary": clean_cell(row.get("context_summary", "")),
                        "XSDStatus": clean_cell(row.get("xsd_status", "")),
                    },
                ),
                (
                    "SM_FunctionAndVendor",
                    {
                        "DeviceInformation": clean_cell(row.get("device_information", "")),
                        "VendorCompanyName": clean_cell(row.get("vendor_company_name", "")),
                        "SafetyRelevanceClass": clean_cell(row.get("safety_relevance_class", "")),
                        "LabelText": clean_cell(row.get("label_text", "")),
                        "FunctionCode": clean_cell(row.get("function_code", "")),
                        "SourceDocument": clean_cell(row.get("source_doc_id", "")),
                        "SourceLocator": clean_cell(row.get("source_locator", "")),
                    },
                ),
                (
                    "SM_ActuationAndPiping",
                    {
                        "ActuatingFunctionNumber": clean_cell(row.get("actuating_function_number", "")),
                        "ActuatingLocation": clean_cell(row.get("actuating_location", "")),
                        "ActuatingSystemNumber": clean_cell(row.get("actuating_system_number", "")),
                        "OperatedValveReference": clean_cell(row.get("operated_valve_reference", "")),
                        "FlowDirection": clean_cell(row.get("flow_direction", "")),
                        "NominalDiameterNumericalValueRepresentation": clean_cell(row.get("nominal_diameter_numerical_value_representation", "")),
                        "NominalDiameterRepresentation": clean_cell(row.get("nominal_diameter_representation", "")),
                        "NominalDiameterStandard": clean_cell(row.get("nominal_diameter_standard", "")),
                        "NominalDiameterTypeRepresentation": clean_cell(row.get("nominal_diameter_type_representation", "")),
                        "LineNumber": clean_cell(row.get("line_number", "")),
                        "PipingComponentName": clean_cell(row.get("piping_component_name", "")),
                        "FromEquipmentId": clean_cell(row.get("from_equipment_id", "")),
                        "ToEquipmentId": clean_cell(row.get("to_equipment_id", "")),
                        "PipingAnchorId": clean_cell(row.get("piping_anchor_id", "")),
                        "RecommendedAction": clean_cell(completion.get("recommended_action", "")),
                    },
                ),
                (
                    "SM_Traceability",
                    {
                        "SourceDocument": clean_cell(row.get("source_doc_id", "")),
                        "SourceLocator": clean_cell(row.get("source_locator", "")),
                        "Confidence": clean_cell(row.get("confidence", "")),
                        "NeedsReview": clean_cell(row.get("needs_review", "")),
                        "RecommendedAction": clean_cell(completion.get("recommended_action", "")),
                        "ProposalStatus": clean_cell(completion.get("proposal_status", "")),
                        "MissingTargets": clean_cell(completion.get("missing_targets", "")),
                        "EvidenceBundleId": clean_cell(row.get("evidence_bundle_id", "") or completion.get("evidence_bundle_id", "")),
                        "DecisionConfidence": clean_cell(row.get("decision_confidence", "") or completion.get("decision_confidence", "")),
                        "UncertaintyReason": clean_cell(row.get("uncertainty_reason", "") or completion.get("uncertainty_reason", "")),
                        "LLMVerificationStatus": clean_cell(
                            row.get("llm_verification_status", "") or completion.get("llm_verification_status", "")
                        ),
                        "RuleSupport": clean_cell(row.get("rule_support", "") or completion.get("rule_support", "")),
                        "ReviewFeedbackStatus": clean_cell(
                            row.get("review_feedback_status", "") or completion.get("review_feedback_status", "")
                        ),
                        "DecisionTrace": clean_cell(row.get("decision_trace_json", "") or completion.get("decision_trace_json", "")),
                    },
                ),
            ],
        )

    def _uc1_build_source_aas_payload(
        self,
        source_type: str,
        rows: list[dict[str, str]],
        identity_key: str,
    ) -> dict[str, object]:
        identity_value = self._uc1_join_values(rows, "canonical_tag", "device_id", "tag", "plt_stelle", "global_id") or identity_key
        common_trace = {
            "SourceDocument": self._uc1_join_values(rows, "source_doc_id"),
            "SourceLocator": self._uc1_join_values(rows, "source_locator"),
            "Confidence": self._uc1_join_values(rows, "confidence"),
            "NeedsReview": self._uc1_join_values(rows, "needs_review"),
            "RecommendedAction": self._uc1_join_values(rows, "recommended_action"),
            "ProposalStatus": self._uc1_join_values(rows, "proposal_status"),
            "MissingTargets": self._uc1_join_values(rows, "missing_targets"),
            "EvidenceBundleId": self._uc1_join_values(rows, "evidence_bundle_id"),
            "DecisionConfidence": self._uc1_join_values(rows, "decision_confidence"),
            "UncertaintyReason": self._uc1_join_values(rows, "uncertainty_reason"),
            "LLMVerificationStatus": self._uc1_join_values(rows, "llm_verification_status"),
            "RuleSupport": self._uc1_join_values(rows, "rule_support"),
            "ReviewFeedbackStatus": self._uc1_join_values(rows, "review_feedback_status"),
            "DecisionTrace": self._uc1_join_values(rows, "decision_trace_json"),
        }
        completion_trace = {
            "SourceDocument": common_trace["SourceDocument"],
            "SourceLocator": common_trace["SourceLocator"],
            "Confidence": common_trace["Confidence"],
            "NeedsReview": common_trace["NeedsReview"],
            "RecommendedAction": common_trace["RecommendedAction"],
            "ProposalStatus": common_trace["ProposalStatus"],
            "MissingTargets": common_trace["MissingTargets"],
        }
        if source_type == "instrument_list":
            submodels = [
                (
                    "SM_InstrumentListEntry",
                    {
                        "EntryId": self._uc1_join_values(rows, "entry_id"),
                        "DeviceId": self._uc1_join_values(rows, "device_id"),
                        "canonicalTag": self._uc1_join_values(rows, "canonical_tag"),
                        "Tag": self._uc1_join_values(rows, "tag"),
                        "DeviceInformation": self._uc1_join_values(rows, "device_information"),
                        "PresenceStatus": self._uc1_join_values(rows, "presence_status"),
                        "DisplayName": self._uc1_join_values(rows, "display_name"),
                    },
                ),
                ("SM_Traceability", common_trace),
            ]
        elif source_type == "wiring":
            submodels = [
                (
                    "SM_WiringEntry",
                    {
                        "EntryId": self._uc1_join_values(rows, "entry_id"),
                        "DeviceId": self._uc1_join_values(rows, "device_id"),
                        "canonicalTag": self._uc1_join_values(rows, "canonical_tag"),
                        "PLTStelle": self._uc1_join_values(rows, "plt_stelle"),
                        "Funktion": self._uc1_join_values(rows, "funktion"),
                        "Beschreibung": self._uc1_join_values(rows, "beschreibung"),
                        "ESchrank": self._uc1_join_values(rows, "e_schrank"),
                        "WireLabel": self._uc1_join_values(rows, "wire_label"),
                        "PresenceStatus": self._uc1_join_values(rows, "presence_status"),
                    },
                ),
                ("SM_Traceability", common_trace),
            ]
        elif source_type == "datasheet":
            submodels = [
                (
                    "SM_DatasheetEntry",
                    {
                        "EntryId": self._uc1_join_values(rows, "entry_id"),
                        "DeviceId": self._uc1_join_values(rows, "device_id"),
                        "canonicalTag": self._uc1_join_values(rows, "canonical_tag"),
                        "Tag": self._uc1_join_values(rows, "tag"),
                        "DeviceInformation": self._uc1_join_values(rows, "device_information"),
                        "Art": self._uc1_join_values(rows, "art"),
                        "Kanal": self._uc1_join_values(rows, "kanal"),
                        "YP": self._uc1_join_values(rows, "yp"),
                        "Position": self._uc1_join_values(rows, "position"),
                        "Address": self._uc1_join_values(rows, "address"),
                        "Project": self._uc1_join_values(rows, "project"),
                        "PresenceStatus": self._uc1_join_values(rows, "presence_status"),
                    },
                ),
                ("SM_Traceability", common_trace),
            ]
        else:
            submodels = [
                (
                    "SM_IFCConnectivity",
                    {
                        "EntryId": self._uc1_join_values(rows, "entry_id"),
                        "DeviceId": self._uc1_join_values(rows, "device_id"),
                        "canonicalTag": self._uc1_join_values(rows, "canonical_tag"),
                        "IFCClass": self._uc1_join_values(rows, "ifc_class"),
                        "GlobalId": self._uc1_join_values(rows, "global_id"),
                        "Tag": self._uc1_join_values(rows, "tag"),
                        "HasPorts": self._uc1_join_values(rows, "has_ports"),
                        "ConnectedTo": self._uc1_join_values(rows, "connected_to"),
                        "ConnectedFrom": self._uc1_join_values(rows, "connected_from"),
                        "HasControlElements": self._uc1_join_values(rows, "has_control_elements"),
                        "PredefinedType": self._uc1_join_values(rows, "predefined_type"),
                        "Size": self._uc1_join_values(rows, "size"),
                        "ValveMechanism": self._uc1_join_values(rows, "valve_mechanism"),
                        "FlowCoefficient": self._uc1_join_values(rows, "flow_coefficient"),
                        "FailPosition": self._uc1_join_values(rows, "fail_position"),
                        "ManualOverride": self._uc1_join_values(rows, "manual_override"),
                        "ActuatorApplication": self._uc1_join_values(rows, "actuator_application"),
                        "FlangeComplete": self._uc1_join_values(rows, "flange_complete"),
                        "PresenceStatus": self._uc1_join_values(rows, "presence_status"),
                    },
                ),
                ("SM_Traceability", common_trace),
            ]
        if clean_cell(common_trace.get("RecommendedAction", "")) or clean_cell(common_trace.get("ProposalStatus", "")):
            submodels.append(("SM_CompletionProposal", completion_trace))
        return self._uc1_make_aas_payload(source_type, identity_value, submodels)

    def _uc1_payload_identity(self, payload: dict[str, object]) -> str:
        shells = payload.get("assetAdministrationShells", [])
        if isinstance(shells, list) and shells:
            shell = shells[0]
            if isinstance(shell, dict):
                identity = clean_cell(shell.get("idShort", "")) or clean_cell(shell.get("id", ""))
                normalized = normalize_identifier(identity)
                if normalized:
                    return normalized
        return "entry"

    def _uc1_write_payload_formats(
        self,
        payload: dict[str, object],
        output_dir: Path,
        identity: str,
        formats: list[str],
    ) -> list[Path]:
        ensure_dir(output_dir)
        generated: list[Path] = []
        for target_format in formats:
            if target_format == "json":
                target_path = output_dir / f"{identity}.json"
                target_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            elif target_format == "xml":
                target_path = output_dir / f"{identity}.xml"
                target_path.write_text(self.aas_generation_service._standardized_payload_to_xml(payload), encoding="utf-8")
            elif target_format == "aasx":
                target_path = output_dir / f"{identity}.aasx"
                temp_json_path = output_dir / f".{identity}.intermediate.json"
                temp_json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
                try:
                    self.aas_generation_service._write_aasx(temp_json_path, target_path)
                finally:
                    if temp_json_path.exists():
                        temp_json_path.unlink()
            else:
                raise ValueError(f"Unsupported AAS target format: {target_format}")
            generated.append(target_path)
        return generated

    def _uc1_append_standardized_sheet_rows(
        self,
        target_rows: list[dict[str, object]],
        *,
        sheet_type: str,
        device_id: str,
        canonical_tag: str,
        record_keys: list[str],
        record_lookup: dict[str, ExtractedRecord],
    ) -> None:
        if not record_keys:
            if sheet_type == "stellenplan":
                target_rows.append(
                    self._uc1_finalize_standardized_row(
                        "instrument_list",
                        {
                            "entry_id": self._uc1_relation_id(device_id, sheet_type),
                            "device_id": device_id,
                            "canonical_tag": canonical_tag,
                            "tag": "",
                            "device_information": "",
                            "source_doc_id": "",
                            "source_locator": "",
                            "confidence": "",
                            "presence_status": "missing",
                            "record_key": "",
                            "display_name": "",
                        },
                        {},
                        canonical_entity_id=device_id,
                        match_confidence="0.0",
                        match_method="missing_placeholder",
                        needs_review_reason="instrument_list_missing",
                    )
                )
            else:
                target_rows.append(
                    self._uc1_finalize_standardized_row(
                        "wiring",
                        {
                            "entry_id": self._uc1_relation_id(device_id, sheet_type),
                            "device_id": device_id,
                            "canonical_tag": canonical_tag,
                            "plt_stelle": "",
                            "funktion": "",
                            "beschreibung": "",
                            "e_schrank": "",
                            "wire_label": "",
                            "source_doc_id": "",
                            "source_locator": "",
                            "confidence": "",
                            "presence_status": "missing",
                            "record_key": "",
                            "display_name": "",
                        },
                        {},
                        canonical_entity_id=device_id,
                        match_confidence="0.0",
                        match_method="missing_placeholder",
                        needs_review_reason="wiring_missing",
                    )
                )
            return

        for record_key in record_keys:
            record = record_lookup.get(record_key)
            if record is None:
                continue
            trace_metadata = self._record_trace_metadata(record)
            first_evidence = next(
                (
                    evidence
                    for result in record.results
                    for evidence in result.evidence_refs
                ),
                None,
            )
            result_map = self._record_result_map(record)
            source_path = clean_cell(record.source_path)
            source_root = record.source_root or self.family_source_root(record.family)
            row_base = {
                "entry_id": self._uc1_relation_id(device_id, record.record_key, sheet_type),
                "device_id": device_id,
                "canonical_tag": canonical_tag,
                "source_doc_id": self._uc1_document_id(source_path, source_root),
                "source_locator": clean_cell(first_evidence.cell_range_or_bbox) if first_evidence is not None else "",
                "confidence": 1.0,
                "presence_status": "present",
                "record_key": record.record_key,
                "display_name": record.display_name,
                **trace_metadata,
            }
            if sheet_type == "stellenplan":
                target_rows.append(
                    self._uc1_finalize_standardized_row(
                        "instrument_list",
                        {
                            **row_base,
                            "tag": clean_cell(result_map.get("tag", "") or result_map.get("messstelle", "")),
                            "device_information": clean_cell(result_map.get("device", "") or result_map.get("art", "")),
                        },
                        {},
                        canonical_entity_id=device_id,
                        match_confidence="1.0",
                        match_method="strict_external_match",
                    )
                )
            else:
                target_rows.append(
                    self._uc1_finalize_standardized_row(
                        "wiring",
                        {
                            **row_base,
                            "plt_stelle": clean_cell(result_map.get("plt_stelle", "")),
                            "funktion": clean_cell(result_map.get("funktion", "")),
                            "beschreibung": clean_cell(result_map.get("beschreibung", "")),
                            "e_schrank": clean_cell(result_map.get("e_schrank", "")),
                            "wire_label": clean_cell(result_map.get("wire_label", "")),
                        },
                        {},
                        canonical_entity_id=device_id,
                        match_confidence="1.0",
                        match_method="strict_external_match",
                    )
                )

    def _uc1_ifc_details_index(self) -> dict[str, list[dict[str, str]]]:
        if not self.documents:
            return {}
        details: dict[str, list[dict[str, str]]] = defaultdict(list)
        for document in self.documents:
            if document.source_kind != SourceDocumentKind.IFC_MODEL:
                continue
            parsed = self._parse_document(document)
            package = parsed.ifc_package
            if package is None:
                continue
            outgoing: dict[str, list[str]] = defaultdict(list)
            incoming: dict[str, list[str]] = defaultdict(list)
            for edge in package.ifc_edges:
                outgoing[edge.from_id].append(edge.to_id)
                incoming[edge.to_id].append(edge.from_id)
            for node in package.ifc_nodes:
                item = {
                    "ifc_class": node.ifc_class,
                    "global_id": node.node_id,
                    "tag": node.tag,
                    "has_ports": " | ".join(sorted(set(outgoing.get(node.node_id, [])))),
                    "connected_to": " | ".join(sorted(set(outgoing.get(node.node_id, [])))),
                    "connected_from": " | ".join(sorted(set(incoming.get(node.node_id, [])))),
                    "has_control_elements": " | ".join(
                        sorted(
                            {
                                edge.to_id
                                for edge in package.ifc_edges
                                if edge.relation_type == "IfcRelFlowControlElements" and edge.from_id == node.node_id
                            }
                        )
                    ),
                    "predefined_type": node.predefined_type,
                    "size": self._uc1_attr_value(node.attributes, "size", "nominal_diameter", "pset_valvetypecommon_size"),
                    "valve_mechanism": self._uc1_attr_value(
                        node.attributes,
                        "valve_mechanism",
                        "pset_valvetypecommon_valvepattern",
                    ),
                    "flow_coefficient": self._uc1_attr_value(
                        node.attributes,
                        "flow_coefficient",
                        "pset_valvetypecommon_flowcoefficient",
                    ),
                    "fail_position": self._uc1_attr_value(node.attributes, "fail_position"),
                    "manual_override": self._uc1_attr_value(node.attributes, "manual_override"),
                    "actuator_application": self._uc1_attr_value(node.attributes, "actuator_application"),
                    "source_doc_id": self._uc1_document_id(document.relative_path, document.source_root),
                    "source_locator": node.locator,
                }
                for match_key in node.match_keys:
                    if match_key:
                        details[match_key].append(item)
        return details

    def _is_ri_family(self, family: DocumentFamily) -> bool:
        return family in {
            DocumentFamily.RI_EQUIPMENT_ROW,
            DocumentFamily.RI_INSTRUMENT_FUNCTION_ROW,
            DocumentFamily.RI_PIPING_COMPONENT_ROW,
            DocumentFamily.RI_CONNECTION_ROW,
        }

    def _bundle_by_id(self, bundle_id: str) -> RiBundle | None:
        return next((bundle for bundle in self.ri_bundles if bundle.bundle_id == bundle_id), None)

    def _ri_sheet_name(self, family: DocumentFamily) -> str:
        mapping = {
            DocumentFamily.RI_EQUIPMENT_ROW: "equipment",
            DocumentFamily.RI_INSTRUMENT_FUNCTION_ROW: "instrument_functions",
            DocumentFamily.RI_PIPING_COMPONENT_ROW: "piping_components",
            DocumentFamily.RI_CONNECTION_ROW: "connections",
        }
        return mapping.get(family, family.value)

    def _family_source_root(self, family: DocumentFamily | str) -> str:
        family_value = family.value if isinstance(family, DocumentFamily) else family
        if family_value in {
            DocumentFamily.STELLEN_OVERVIEW_RECORD.value,
            DocumentFamily.STELLEN_TU_DATASHEET.value,
        }:
            return "Stellenplaene"
        if family_value in {
            DocumentFamily.STROMLAUF_COMPONENT_GROUP.value,
            DocumentFamily.STROMLAUF_COMPONENT.value,
            DocumentFamily.STROMLAUF_CONNECTION.value,
        }:
            return "Diagram-PDF"
        if family_value in {
            DocumentFamily.RI_EQUIPMENT_ROW.value,
            DocumentFamily.RI_INSTRUMENT_FUNCTION_ROW.value,
            DocumentFamily.RI_PIPING_COMPONENT_ROW.value,
            DocumentFamily.RI_CONNECTION_ROW.value,
        }:
            return "R&I-Fließbild"
        if family_value in {
            DocumentFamily.IFC_PIPING_ITEM_ROW.value,
            DocumentFamily.IFC_CONNECTION_ROW.value,
        }:
            return "IFC"
        return "Verschaltungslisten"

    def family_source_root(self, family: DocumentFamily | str) -> str:
        return self._family_source_root(family)

    def bundle_name_for_scope(self, scope_id: str) -> str:
        return self.bundle_display_name_for_scope(scope_id)

    def bundle_display_name_for_scope(self, scope_id: str) -> str:
        bundle = self._bundle_by_id(scope_id)
        if bundle is not None:
            return bundle.display_name or bundle.drawing_name or bundle.bundle_id
        first_schema = next(iter(self.ri_bundle_schemas.get(scope_id, {}).values()), None)
        if first_schema is not None and first_schema.bundle_name:
            return first_schema.bundle_name
        return scope_id

    def bundle_export_name_for_scope(self, scope_id: str) -> str:
        return self._sanitize_export_filename(self.bundle_display_name_for_scope(scope_id))

    def _result_family_dir(self, base_dir: Path, family_name: str) -> Path:
        return ensure_dir(base_dir / self._family_source_root(family_name))

    def _ri_result_dir(self, base_dir: Path, bundle_id: str) -> Path:
        bundle = self._bundle_by_id(bundle_id)
        source_root = bundle.source_root if bundle else "R&I-Fließbild"
        if bundle is None:
            first_schema = next(iter(self.ri_bundle_schemas.get(bundle_id, {}).values()), None)
            if first_schema is not None and first_schema.source_root:
                source_root = first_schema.source_root
        return ensure_dir(base_dir / source_root)

    def _ri_bundle_names(self) -> dict[str, str]:
        names = {
            bundle.bundle_id: self.bundle_display_name_for_scope(bundle.bundle_id)
            for bundle in self.ri_bundles
        }
        for bundle_id, bundle_schemas in self.ri_bundle_schemas.items():
            if bundle_id in names:
                continue
            first_schema = next(iter(bundle_schemas.values()), None)
            if first_schema is not None:
                names[bundle_id] = self.bundle_display_name_for_scope(bundle_id)
        return names

    def _ri_record_groups(self) -> dict[str, dict]:
        grouped: dict[str, dict[DocumentFamily, list]] = {}
        names: dict[str, str] = {}
        for record in self.records:
            if not self._is_ri_family(record.family):
                continue
            scope_id = record.scope_id or ""
            if not scope_id:
                continue
            grouped.setdefault(scope_id, {}).setdefault(record.family, []).append(record)
            if scope_id not in names:
                names[scope_id] = self.bundle_display_name_for_scope(scope_id)
        return {"records": grouped, "names": names}

    def _sanitize_export_filename(self, name: str) -> str:
        sanitized = re.sub(r'[<>:"/\\|?*]+', "_", name or "")
        sanitized = re.sub(r"_+", "_", sanitized).strip(" ._")
        return sanitized or "bundle"

    def _pid_component_families(self) -> set[DocumentFamily]:
        return {
            DocumentFamily.RI_EQUIPMENT_ROW,
            DocumentFamily.RI_INSTRUMENT_FUNCTION_ROW,
            DocumentFamily.RI_PIPING_COMPONENT_ROW,
        }

    def _pid_bundle_contexts(self, progress: ProgressCallback | None = None) -> dict[str, dict[str, object]]:
        contexts: dict[str, dict[str, object]] = {}
        if not self.ri_bundles:
            return contexts
        active_bundles = [bundle for bundle in self.ri_bundles if bundle.pairing_status in {"paired", "missing_xsd"}]
        total = max(1, len(active_bundles))
        for index, bundle in enumerate(active_bundles, start=1):
            parsed = self._parse_ri_bundle(bundle, self.reader)
            pdf_tokens: dict[str, str] = {}
            pdf_identifiers: dict[str, str] = {}
            if parsed is not None:
                for page in parsed.pages:
                    for block in page.blocks:
                        for token in extract_component_tokens(block.text):
                            normalized_key = normalize_label(token)
                            if normalized_key and normalized_key not in pdf_tokens:
                                pdf_tokens[normalized_key] = token
                            identifier_key = normalize_identifier(token)
                            if identifier_key and identifier_key not in pdf_identifiers:
                                pdf_identifiers[identifier_key] = token
            contexts[bundle.bundle_id] = {
                "bundle": bundle,
                "parsed": parsed,
                "pdf_tokens": pdf_tokens,
                "pdf_identifiers": pdf_identifiers,
            }
            self._report_progress(
                progress,
                20 + round(index * 20 / total),
                f"Read R&I bundle {bundle.display_name or bundle.bundle_id}",
            )
        return contexts

    def _pid_external_match_index(self) -> dict[str, dict[str, set[str]]]:
        index: dict[str, dict[str, set[str]]] = {
            "Stellenplaene": defaultdict(set),
            "Verschaltungslisten": defaultdict(set),
        }
        for record in self.records:
            if record.family in self._pid_component_families() or record.family == DocumentFamily.RI_CONNECTION_ROW:
                continue
            source_root = record.source_root or self.family_source_root(record.family)
            if source_root not in index:
                continue
            candidates = {record.display_name, record.record_key}
            for result in record.results:
                if result.value:
                    candidates.add(result.value)
            for candidate in candidates:
                normalized_key = normalize_label(candidate)
                if normalized_key:
                    index[source_root][normalized_key].add(record.record_key)
        return index

    def _ensure_pid_bucket(
        self,
        buckets: dict[tuple[str, str], dict[str, object]],
        scope_id: str,
        normalized_key: str,
        display_name: str,
    ) -> dict[str, object]:
        return buckets.setdefault(
            (scope_id, normalized_key),
            {
                "display_name": display_name,
                "source_root": "R&I-Fließbild",
                "pdf_present": False,
                "xml_present": False,
                "categories": set(),
                "type_candidates": [],
                "ri_record_keys": set(),
                "pdf_record_keys": set(),
                "xml_record_keys": set(),
            },
        )

    def _pid_record_display_name(self, record: ExtractedRecord) -> str:
        if record.display_name:
            return record.display_name
        for field_name in ("tag_name", "node_id", "from_id"):
            for result in record.results:
                if result.field_name == field_name and result.value:
                    return result.value
        return record.record_key

    def _pid_record_primary_type(self, record: ExtractedRecord) -> str:
        for field_name in ("normalized_type", "class_name", "sub_class", "edge_type"):
            for result in record.results:
                if result.field_name == field_name and result.value:
                    return result.value
        return self._pid_category_for_family(record.family)

    def _pid_primary_type(self, candidates: list[str]) -> tuple[str, bool]:
        ordered = [candidate for candidate in candidates if candidate]
        if not ordered:
            return "", False
        counts = Counter(ordered)
        primary_type, _ = counts.most_common(1)[0]
        normalized = {normalize_label(candidate) for candidate in ordered if normalize_label(candidate)}
        return primary_type, len(normalized) > 1

    def _pid_category_for_family(self, family: DocumentFamily) -> str:
        mapping = {
            DocumentFamily.RI_EQUIPMENT_ROW: "equipment",
            DocumentFamily.RI_INSTRUMENT_FUNCTION_ROW: "instrument_function",
            DocumentFamily.RI_PIPING_COMPONENT_ROW: "piping_component",
            DocumentFamily.RI_CONNECTION_ROW: "connection",
        }
        return mapping.get(family, family.value)

    def _pid_xsd_status(self, scope_id: str, bucket: dict[str, object], context: dict[str, object] | None) -> str:
        if not bucket["xml_present"]:
            return "missing"
        if context is None:
            return "present" if scope_id in self.ri_bundle_schemas else "missing"
        parsed = context.get("parsed")
        if not isinstance(parsed, ParsedDocument) or parsed.ri_package is None:
            return "missing"
        package = parsed.ri_package
        if package.validation_errors:
            return "conflict"
        categories = {str(category) for category in bucket["categories"] if category}
        available = {
            field.category
            for field in package.xsd_field_defs
            if field.category
        }
        if not available:
            return "missing"
        if categories and any(category in available for category in categories):
            return "present"
        if "common" in available:
            return "present"
        return "missing"

    def pid_source_preview_context(
        self,
        row: PidInconsistencyRow,
        source_column: str,
    ) -> dict[str, object] | None:
        if source_column == "xsd":
            return self._pid_xsd_preview_context(row)

        # VV* components: disable Stellenplaene preview entirely
        if source_column == "stellenplaene" and (row.canonical_tag or "").startswith("VV"):
            return None

        # Component rows: missing sources do NOT fallback to wrong previews
        is_component = (row.canonical_tag or "").startswith(("VV", "PL", "HE"))
        source_status = {
            "pdf": row.pdf_status, "xml": row.xml_status,
            "stellenplaene": row.stellenplaene_status,
            "verschaltungslisten": row.verschaltungslisten_status,
        }.get(source_column, "missing")
        if is_component and source_status != "present":
            return None

        target = row.jump_targets.get(source_column)
        if is_component and (target is None or not target.matching_record_keys):
            return None
        if target is not None:
            records = self._records_for_keys(target.matching_record_keys)
            if records:
                preferred = self._preferred_pid_record(records, target.preferred_record_key)
                ordered_records = records
                if preferred is not None:
                    ordered_records = [preferred, *[record for record in records if record.record_key != preferred.record_key]]
                for record in ordered_records:
                    preview = self._pid_record_preview_context(record, row, source_column)
                    if preview is not None:
                        # Use jump target keyword instead of canonical_tag for target_value
                        if target.keyword and normalize_identifier(target.keyword) != normalize_identifier(row.canonical_tag):
                            preview["target_value"] = target.keyword
                        return preview
        if is_component:
            return None  # no fallback for component rows
        return self._pid_source_preview_fallback(row, source_column)

    def _pid_record_preview_context(
        self,
        record: ExtractedRecord,
        row: PidInconsistencyRow,
        source_column: str,
    ) -> dict[str, object] | None:
        normalized_tag = normalize_identifier(row.canonical_tag or row.component_key)
        preferred_fields = {
            "pdf": ("canonical_tag", "label_text", "tag_name", "name", "label"),
            "xml": ("canonical_tag", "label_text", "tag_name", "name", "label"),
            "stellenplaene": ("tag", "messstelle", "canonical_tag"),
            "verschaltungslisten": ("plt_stelle", "canonical_tag", "funktion"),
        }.get(source_column, ())
        scored_results: list[tuple[int, object, list[EvidenceRef]]] = []
        for result in record.results:
            evidences = [
                evidence
                for evidence in result.evidence_refs
                if self._evidence_matches_pid_source(evidence, source_column)
            ]
            if not evidences:
                continue
            score = 0
            normalized_value = normalize_identifier(result.normalized_value or result.value)
            if result.field_name in preferred_fields:
                score += 10
            if normalized_tag and normalized_value == normalized_tag:
                score += 15
            if row.canonical_tag and row.canonical_tag.casefold() in clean_cell(result.value).casefold():
                score += 8
            if result.value:
                score += 1
            score += self._preview_evidence_precision_score(evidences)
            scored_results.append((score, result, evidences))
        if scored_results:
            scored_results.sort(key=lambda item: item[0], reverse=True)
            _score, result, evidences = scored_results[0]
        else:
            fallback = self._pid_fallback_preview_candidate(record, row, source_column, preferred_fields)
            if fallback is None:
                return None
            result, evidences = fallback
        resolved_source_path = evidences[0].source_path if evidences else record.source_path
        return {
            "source_path": resolved_source_path,
            "record_display_name": record.display_name,
            "field_name": result.field_name,
            "target_value": row.canonical_tag or result.value or row.display_name,
            "evidences": [evidence.model_dump(mode="json") for evidence in evidences],
        }

    def _pid_fallback_preview_candidate(
        self,
        record: ExtractedRecord,
        row: PidInconsistencyRow,
        source_column: str,
        preferred_fields: tuple[str, ...],
    ) -> tuple[object, list[EvidenceRef]] | None:
        source_matches = self._source_path_matches_pid_source(record.source_path, source_column)
        if not source_matches:
            return None
        fallback_result = None
        fallback_score = -1
        fallback_evidences: list[EvidenceRef] = []
        normalized_tag = normalize_identifier(row.canonical_tag or row.component_key)
        for result in record.results:
            score = 0
            if result.field_name in preferred_fields:
                score += 10
            if normalized_tag and normalize_identifier(result.normalized_value or result.value) == normalized_tag:
                score += 15
            if row.canonical_tag and row.canonical_tag.casefold() in clean_cell(result.value).casefold():
                score += 8
            if result.value:
                score += 1
            candidate_evidences = [
                evidence
                for evidence in result.evidence_refs
                if self._source_path_matches_pid_source(evidence.source_path, source_column)
            ]
            if candidate_evidences:
                score += self._preview_evidence_precision_score(candidate_evidences)
            if score > fallback_score:
                fallback_score = score
                fallback_result = result
                fallback_evidences = candidate_evidences
        if fallback_result is None:
            return None
        if not fallback_evidences:
            fallback_evidences = [self._synthetic_preview_evidence(record, row, source_column, fallback_result.value)]
        return fallback_result, fallback_evidences

    def _pid_xsd_preview_context(self, row: PidInconsistencyRow) -> dict[str, object] | None:
        context = self._pid_bundle_contexts().get(row.scope_id)
        if context is None:
            return None
        bundle = context.get("bundle")
        parsed = context.get("parsed")
        if not isinstance(bundle, RiBundle) or not isinstance(parsed, ParsedDocument) or parsed.ri_package is None:
            return None
        if bundle.xsd_path is None or not bundle.xsd_path.exists():
            return None
        candidate = next(
            (
                field
                for field in parsed.ri_package.xsd_field_defs
                if field.category in {"instrument_function", "common"}
            ),
            None,
        )
        snippet = (
            f"{candidate.xml_name}: {candidate.description}"
            if candidate is not None
            else f"XSD support for {row.canonical_tag}"
        )
        locator = (
            f"xsd::{candidate.category}::{candidate.xml_name}"
            if candidate is not None
            else "xsd::instrument_function"
        )
        evidence = EvidenceRef(
            source_path=bundle.xsd_path.as_posix(),
            page_or_sheet="XSD",
            cell_range_or_bbox=locator,
            snippet=snippet[:240],
            score=1.0,
            evidence_type="xsd",
            engine="xsd",
        )
        return {
            "source_path": bundle.xsd_path.as_posix(),
            "record_display_name": row.display_name,
            "field_name": "xsd_schema",
            "target_value": row.canonical_tag or row.display_name,
            "evidences": [evidence.model_dump(mode="json")],
        }

    def _records_for_keys(self, record_keys: list[str]) -> list[ExtractedRecord]:
        wanted = {clean_cell(record_key) for record_key in record_keys if clean_cell(record_key)}
        if not wanted:
            return []
        return [record for record in self.records if record.record_key in wanted]

    def _preferred_pid_record(
        self,
        records: list[ExtractedRecord],
        preferred_record_key: str,
    ) -> ExtractedRecord | None:
        if preferred_record_key:
            for record in records:
                if record.record_key == preferred_record_key:
                    return record
        return records[0] if records else None

    def _evidence_matches_pid_source(self, evidence: EvidenceRef, source_column: str) -> bool:
        source_path = clean_cell(evidence.source_path).lower()
        page_label = clean_cell(evidence.page_or_sheet).lower()
        evidence_type = clean_cell(evidence.evidence_type).lower()
        if source_column == "pdf":
            return evidence_type in {"native_text", "ocr_text", "table_cell", "kv_pair"} or source_path.endswith(".pdf")
        if source_column == "xml":
            return evidence_type.startswith("dexpi_") or source_path.endswith(".xml") or "xml" in page_label
        if source_column == "stellenplaene":
            if self._source_path_matches_pid_source(source_path, source_column):
                return True
            return "stellen" in page_label
        if source_column == "verschaltungslisten":
            if self._source_path_matches_pid_source(source_path, source_column):
                return True
            return "verschaltung" in page_label or "wiring" in page_label
        return False

    def _source_path_matches_pid_source(self, source_path: str, source_column: str) -> bool:
        lowered = clean_cell(source_path).lower()
        if source_column == "pdf":
            return lowered.endswith(".pdf")
        if source_column == "xml":
            return lowered.endswith(".xml")
        if source_column == "stellenplaene":
            return "stellenplaene" in lowered and lowered.endswith((".pdf", ".xls", ".xlsx"))
        if source_column == "verschaltungslisten":
            return "verschaltungslisten" in lowered and lowered.endswith((".pdf", ".xls", ".xlsx"))
        return False

    def _synthetic_preview_evidence(
        self,
        record: ExtractedRecord,
        row: PidInconsistencyRow,
        source_column: str,
        fallback_value: str,
    ) -> EvidenceRef:
        page_hint = ""
        if source_column in {"pdf", "stellenplaene"} and record.source_path.lower().endswith(".pdf"):
            page_hint = "Page 1"
        if source_column == "xml":
            page_hint = "DEXPI XML"
        if source_column == "stellenplaene" and record.source_path.lower().endswith((".xls", ".xlsx")):
            page_hint = "Sheet1"
        if source_column == "verschaltungslisten" and record.source_path.lower().endswith((".xls", ".xlsx")):
            page_hint = "Tabelle1"
        snippet = clean_cell(fallback_value) or clean_cell(row.canonical_tag) or clean_cell(row.display_name)
        return EvidenceRef(
            source_path=record.source_path,
            page_or_sheet=page_hint,
            cell_range_or_bbox="",
            snippet=snippet[:240],
            score=0.5,
            evidence_type=f"pid_{source_column}_fallback",
            engine="pid-fallback",
        )

    def _preview_evidence_precision_score(self, evidences: list[EvidenceRef]) -> int:
        best = 0
        for evidence in evidences:
            locator = clean_cell(evidence.cell_range_or_bbox).lower()
            page_label = clean_cell(evidence.page_or_sheet).lower()
            if locator and locator != "filename":
                best = max(best, 20)
            elif page_label and page_label != "filename":
                best = max(best, 10)
            else:
                best = max(best, -5)
        return best

    def _pid_source_preview_fallback(
        self,
        row: PidInconsistencyRow,
        source_column: str,
    ) -> dict[str, object] | None:
        fallback_record = self._find_pid_source_record(row, source_column)
        if fallback_record is not None:
            evidence = self._synthetic_preview_evidence(
                fallback_record,
                row,
                source_column,
                row.canonical_tag or fallback_record.display_name,
            )
            return {
                "source_path": fallback_record.source_path,
                "record_display_name": fallback_record.display_name,
                "field_name": "canonical_tag",
                "target_value": row.canonical_tag or row.display_name,
                "evidences": [evidence.model_dump(mode="json")],
            }

        bundle = self._bundle_by_id(row.scope_id)
        if source_column == "pdf" and bundle is not None and bundle.pdf_path is not None:
            return self._bundle_source_preview_context(
                source_path=bundle.pdf_path,
                row=row,
                source_column=source_column,
                page_hint="Page 1",
            )
        if source_column == "xml" and bundle is not None and bundle.xml_path is not None:
            return self._bundle_source_preview_context(
                source_path=bundle.xml_path,
                row=row,
                source_column=source_column,
                page_hint="DEXPI XML",
            )
        return None

    def _bundle_source_preview_context(
        self,
        *,
        source_path: Path,
        row: PidInconsistencyRow,
        source_column: str,
        page_hint: str,
    ) -> dict[str, object]:
        evidence = EvidenceRef(
            source_path=source_path.as_posix(),
            page_or_sheet=page_hint,
            cell_range_or_bbox="",
            snippet=(row.canonical_tag or row.display_name)[:240],
            score=0.5,
            evidence_type=f"pid_{source_column}_bundle_fallback",
            engine="pid-fallback",
        )
        return {
            "source_path": source_path.as_posix(),
            "record_display_name": row.display_name,
            "field_name": "canonical_tag",
            "target_value": row.canonical_tag or row.display_name,
            "evidences": [evidence.model_dump(mode="json")],
        }

    def _find_pid_source_record(
        self,
        row: PidInconsistencyRow,
        source_column: str,
    ) -> ExtractedRecord | None:
        normalized_tag = normalize_identifier(row.canonical_tag or row.component_key)
        desired_root = {
            "stellenplaene": "Stellenplaene",
            "verschaltungslisten": "Verschaltungslisten",
        }.get(source_column, "")
        if not desired_root:
            return None
        best_record: ExtractedRecord | None = None
        best_score = -1
        for record in self.records:
            source_root = record.source_root or self.family_source_root(record.family)
            if source_root != desired_root:
                continue
            score = 0
            if self._source_path_matches_pid_source(record.source_path, source_column):
                score += 10
            if normalized_tag and normalize_identifier(record.display_name) == normalized_tag:
                score += 12
            for result in record.results:
                normalized_value = normalize_identifier(result.normalized_value or result.value)
                if normalized_tag and normalized_value == normalized_tag:
                    score += 15
                elif row.canonical_tag and row.canonical_tag.casefold() in clean_cell(result.value).casefold():
                    score += 6
            if score > best_score:
                best_score = score
                best_record = record
        return best_record if best_score > 0 else None

    def _ui_language_code(self) -> str:
        value = clean_cell(getattr(self.settings, "ui_language", "en")).lower()
        if value.startswith("de"):
            return "de"
        if value.startswith("zh"):
            return "zh"
        return "en"

    def _pid_jump_target(
        self,
        keyword: str,
        *,
        preferred_source_root: str = "",
        preferred_scope_id: str = "",
        matching_record_keys: list[str] | None = None,
        preferred_record_key: str = "",
    ) -> PidJumpTarget:
        return PidJumpTarget(
            keyword=keyword,
            preferred_source_root=preferred_source_root,
            preferred_scope_id=preferred_scope_id,
            matching_record_keys=list(matching_record_keys or []),
            preferred_record_key=preferred_record_key,
        )

    def _first_record_key(self, values: list[str] | set[str]) -> str:
        if isinstance(values, set):
            values = sorted(values)
        return values[0] if values else ""
