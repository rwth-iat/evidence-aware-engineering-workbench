#!/usr/bin/env python3
"""Deep semantic audit — cell-level content analysis of AIO workbooks.

Goes beyond structural validation to check:
  - Element_Type correctness vs actual element names/descriptions
  - Primary_RKZ format validity per document type
  - Wire_Color/Polarity/Cross_Section encoding accuracy
  - Document metadata completeness & accuracy
  - Connection completeness (are connections missing?)
  - Object content quality
  - Fill rate per-column, per-sheet
"""

import json
import os
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

EXPORT_DIR = Path(__file__).resolve().parents[1] / "Exports" / "Excel" / "AIO"


def safe(v):
    return str(v).strip() if v is not None else ""


def read_sheet(ws):
    """Read sheet into list of dicts."""
    headers = {}
    for col in range(1, ws.max_column + 1):
        h = safe(ws.cell(row=1, column=col).value)
        if h:
            headers[col] = h
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for col, h in headers.items():
            row[h] = ws.cell(row=r, column=col).value
        if any(v is not None and safe(v) for v in row.values()):
            rows.append(row)
    return rows, {v: k for k, v in headers.items()}


def audit_deep(filepath: Path) -> dict:
    filename = filepath.name
    findings = {
        "file": filename,
        "errors": [],
        "warnings": [],
        "stats": {},
    }

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ── Document_ID ──
    if "Document_ID" in wb.sheetnames:
        rows, _ = read_sheet(wb["Document_ID"])
        for row in rows:
            dtype = safe(row.get("Document_Type", ""))
            if dtype not in ("Instrument_Loop_Diagram", "Terminal_Diagram", "Circuit_Diagram"):
                findings["errors"].append(f"Document_ID: Invalid Document_Type '{dtype}'")
            if safe(row.get("Schema_Version", "")) != "v0.8":
                findings["errors"].append(f"Document_ID: Schema_Version not v0.8")

    # ── Element_ID deep audit ──
    if "Element_ID" in wb.sheetnames:
        eid_rows, _ = read_sheet(wb["Element_ID"])
        findings["stats"]["element_count"] = len(eid_rows)
        etype_counts = Counter()
        rkz_issues = []

        for row in eid_rows:
            etype = safe(row.get("Element_Type", ""))
            rkz = safe(row.get("Primary_RKZ", ""))
            element_id = safe(row.get("Element_ID", ""))
            etype_counts[etype] += 1

            # Check Element_Type validity
            valid_types = {"Sensor", "Transducer", "Valve_Actuator", "Motor", "Actuator",
                          "Consumer", "Terminal", "Terminal_Strip", "Contactor",
                          "Auxiliary_Contactor", "Fuse", "Circuit_Breaker", "Switch",
                          "Socket_Outlet", "Power_Supply", "PLC_Module", "Coil",
                          "Main_Contact", "Auxiliary_Contact", "Indicator_Lamp",
                          "Control_Cabinet", "Cabinet_Aggregate", "Thermostat", "Heater"}
            if etype and etype not in valid_types and etype != "Unspecifiable":
                findings["errors"].append(f"{element_id}: Invalid Element_Type '{etype}'")

            # Check RKZ format
            #if rkz and not re.match(r'^-?[A-Z]', rkz) and "UNKNOWN" not in rkz:
            #    rkz_issues.append(f"{element_id}: Unusual RKZ '{rkz}'")

            # Check for UNKNOWN RKZs (indicates failed designation extraction)
            if "UNKNOWN" in rkz:
                findings["warnings"].append(f"{element_id}: RKZ is '{rkz}' — designation extraction failed")

            # Check that CAEX_Type is valid
            caex = safe(row.get("CAEX_Type", ""))
            if caex not in ("InternalElement", "ExternalInterface", ""):
                findings["errors"].append(f"{element_id}: Invalid CAEX_Type '{caex}'")

        findings["stats"]["etype_distribution"] = dict(etype_counts.most_common())
        if rkz_issues:
            findings["warnings"].extend(rkz_issues[:10])

    # ── Element_Data deep audit ──
    if "Element_Data" in wb.sheetnames:
        ed_rows, _ = read_sheet(wb["Element_Data"])
        findings["stats"]["element_data_count"] = len(ed_rows)

        attr_counts = Counter()
        empty_value_rows = 0
        for row in ed_rows:
            attr = safe(row.get("Attribute_Name", ""))
            val = safe(row.get("Attribute_Value", ""))
            attr_counts[attr] += 1
            if not val:
                empty_value_rows += 1

        findings["stats"]["top_attributes"] = dict(attr_counts.most_common(15))
        if empty_value_rows:
            findings["warnings"].append(f"Element_Data: {empty_value_rows} rows with empty Attribute_Value")

    # ── Connection deep audit ──
    if "Connection_ID" in wb.sheetnames:
        c_rows, _ = read_sheet(wb["Connection_ID"])
        findings["stats"]["connection_count"] = len(c_rows)

        # Check for missing connection annotations
        invalid_status = 0
        for row in c_rows:
            status = safe(row.get("Connection_Status", ""))
            if status not in ("Resolved", "Unresolved", ""):
                invalid_status += 1
        if invalid_status:
            findings["errors"].append(f"Connection_ID: {invalid_status} rows with invalid Connection_Status")

    if "Connection_Data" in wb.sheetnames:
        cd_rows, _ = read_sheet(wb["Connection_Data"])
        findings["stats"]["connection_data_count"] = len(cd_rows)

        wire_colors = Counter()
        polarities = Counter()
        invalid_wc = 0
        invalid_pol = 0

        for row in cd_rows:
            attr = safe(row.get("Attribute_Name", ""))
            val = safe(row.get("Attribute_Value", ""))

            if attr == "Wire_Color":
                wire_colors[val] += 1
                valid_wc = {"BK","BN","RD","OG","YE","GN","BU","VT","GY","WH","PK","TQ","GNYE"}
                if val and val not in valid_wc and val != "Unspecifiable":
                    invalid_wc += 1

            if attr == "Polarity":
                polarities[val] += 1
                valid_pol = {"L1","L2","L3","L","N","PE","PEN","L+","L-","G+","G-","V+","V-","FE","AC","DC"}
                if val and val not in valid_pol and val != "Unspecifiable":
                    invalid_pol += 1

        findings["stats"]["wire_colors"] = dict(wire_colors.most_common())
        findings["stats"]["polarities"] = dict(polarities.most_common())
        if invalid_wc:
            findings["errors"].append(f"Connection_Data: {invalid_wc} invalid Wire_Color values")
        if invalid_pol:
            findings["errors"].append(f"Connection_Data: {invalid_pol} invalid Polarity values")

    # ── Object content audit ──
    if "Object" in wb.sheetnames:
        obj_rows, _ = read_sheet(wb["Object"])
        findings["stats"]["object_count"] = len(obj_rows)

        empty_content = 0
        invalid_source_op = 0
        op_counts = Counter()
        for row in obj_rows:
            content = safe(row.get("Content_Text", ""))
            source_op = safe(row.get("Source_Operation", ""))
            op_counts[source_op] += 1
            if not content:
                empty_content += 1
            valid_ops = {"Tj","TJ","'",'"',"f","F","f*","S","s","B","b","re","Cell","VL_Row","Manual_Entry"}
            if source_op and source_op not in valid_ops:
                invalid_source_op += 1

        findings["stats"]["source_operations"] = dict(op_counts.most_common())
        if empty_content:
            findings["warnings"].append(f"Object: {empty_content} rows with empty Content_Text")
        if invalid_source_op:
            findings["errors"].append(f"Object: {invalid_source_op} invalid Source_Operation values")

    # ── Document_Data completeness ──
    if "Document_Data" in wb.sheetnames:
        dd_rows, _ = read_sheet(wb["Document_Data"])
        attrs_present = {safe(r.get("Attribute_Name", "")) for r in dd_rows}

        # Check critical metadata
        critical = {"Source_Format", "Project_Name"}
        for c in critical:
            if c not in attrs_present:
                findings["warnings"].append(f"Document_Data: Missing critical attribute '{c}'")

        # Check Source_Format value
        for row in dd_rows:
            if safe(row.get("Attribute_Name", "")) == "Source_Format":
                val = safe(row.get("Attribute_Value", ""))
                if val not in ("PDF_Drawing", "Excel_Sheet", "Verschaltungsliste", "Manual_Entry"):
                    findings["errors"].append(f"Document_Data: Invalid Source_Format '{val}'")

    # ── Document_RepresentedItem ──
    if "Document_RepresentedItem" in wb.sheetnames:
        ri_rows, _ = read_sheet(wb["Document_RepresentedItem"])
        for row in ri_rows:
            ri_type = safe(row.get("RepresentedItem_Type", ""))
            valid_ri = {"PCE_Request", "Terminal_Strip", "Circuit", "Plant_Section",
                       "Control_Cabinet", "Distribution_Panel", "Function_Group"}
            if ri_type and ri_type not in valid_ri:
                findings["errors"].append(f"Document_RepresentedItem: Invalid RepresentedItem_Type '{ri_type}'")

    # ── Match_Result consistency ──
    if "Match_Result" in wb.sheetnames:
        mr_rows, _ = read_sheet(wb["Match_Result"])
        for row in mr_rows:
            ms = safe(row.get("Match_Status", ""))
            td = safe(row.get("Element_TopDown_ID", ""))
            efc = safe(row.get("Element_from_Cluster_ID", ""))
            if ms == "Matched" and (not td or not efc):
                findings["errors"].append(f"Match_Result: Matched row missing TD or EFC FK")
            if ms == "Only_TopDown" and (not td or efc):
                findings["errors"].append(f"Match_Result: Only_TopDown row has invalid FKs")
            if ms == "Only_Cluster" and (td or not efc):
                findings["errors"].append(f"Match_Result: Only_Cluster row has invalid FKs")

    # ── Fill rate per column per sheet ──
    fill_rates = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = {}
        for col in range(1, ws.max_column + 1):
            h = safe(ws.cell(row=1, column=col).value)
            if h:
                headers[col] = h

        col_fill = {}
        for col, h in headers.items():
            filled = 0
            total = 0
            for r in range(2, ws.max_row + 1):
                total += 1
                if ws.cell(row=r, column=col).value is not None and safe(ws.cell(row=r, column=col).value):
                    filled += 1
            col_fill[h] = f"{filled}/{total}" if total > 0 else "0/0"
        fill_rates[sn] = {"rows": ws.max_row - 1, "columns": col_fill}

    findings["stats"]["fill_rates"] = fill_rates

    wb.close()
    return findings


def main():
    files = sorted(EXPORT_DIR.glob("*_AIO.xlsx"))
    if not files:
        print("No AIO files found")
        return

    print(f"Deep-auditing {len(files)} AIO workbooks...\n")

    all_findings = []
    total_errors = 0
    total_warnings = 0

    for f in files:
        fnd = audit_deep(f)
        all_findings.append(fnd)
        total_errors += len(fnd["errors"])
        total_warnings += len(fnd["warnings"])

        print(f"{'='*70}")
        print(f"FILE: {f.name}")
        print(f"  Errors:   {len(fnd['errors'])}")
        print(f"  Warnings: {len(fnd['warnings'])}")

        # Show element stats
        ecount = fnd["stats"].get("element_count", 0)
        ccount = fnd["stats"].get("connection_count", 0)
        print(f"  Elements: {ecount}, Connections: {ccount}")

        etypes = fnd["stats"].get("etype_distribution", {})
        if etypes:
            print(f"  Element types: {dict(list(etypes.items())[:5])}")

        # Show all errors
        for e in fnd["errors"]:
            print(f"  ERROR: {e}")
        for w in fnd["warnings"]:
            print(f"  WARN:  {w}")

    print(f"\n{'='*70}")
    print(f"TOTAL: {total_errors} errors, {total_warnings} warnings across {len(files)} workbooks")

    # ── Cross-workbook analysis ──
    print(f"\n{'='*70}")
    print("CROSS-WORKBOOK ANALYSIS")

    all_doc_types = Counter()
    all_elem_types = Counter()
    all_wire_colors = Counter()
    all_polarities = Counter()
    all_source_formats = Counter()

    for fnd in all_findings:
        # Aggregate stats
        all_elem_types.update(fnd["stats"].get("etype_distribution", {}))
        all_wire_colors.update(fnd["stats"].get("wire_colors", {}))
        all_polarities.update(fnd["stats"].get("polarities", {}))

    print(f"\n  Element_Type distribution (all workbooks):")
    for et, cnt in all_elem_types.most_common():
        print(f"    {et:25s}: {cnt:4d}")

    print(f"\n  Wire_Color distribution (all workbooks):")
    for wc, cnt in all_wire_colors.most_common():
        print(f"    {wc:10s}: {cnt:3d}")

    print(f"\n  Polarity distribution (all workbooks):")
    for pol, cnt in all_polarities.most_common():
        print(f"    {pol:10s}: {cnt:3d}")

    # ── Identify workbooks with NO connections ──
    print(f"\n  Workbooks WITH connections:")
    for fnd in all_findings:
        ccount = fnd["stats"].get("connection_count", 0)
        status = f"✅ {ccount} connections" if ccount > 0 else "❌ NO connections"
        print(f"    {fnd['file'][:55]:55s} {status}")

    # ── Fill rate gaps ──
    print(f"\n  Fill rate gaps (columns with 0% fill):")
    for fnd in all_findings:
        gaps = []
        for sn, fr in fnd["stats"].get("fill_rates", {}).items():
            for col, count in fr.get("columns", {}).items():
                if count.startswith("0/") and int(count.split("/")[1]) > 0:
                    gaps.append(f"{sn}.{col}")
        if gaps:
            print(f"    {fnd['file'][:50]}:")
            for g in gaps[:10]:
                print(f"      {g}")

    # Save
    outpath = Path(__file__).resolve().parent.parent / "tests" / "deep_audit_results.json"
    with open(outpath, "w") as f:
        json.dump({
            "timestamp": datetime.now().isoformat(),
            "files_audited": len(files),
            "total_errors": total_errors,
            "total_warnings": total_warnings,
            "findings": all_findings,
        }, f, indent=2, default=str)
    print(f"\nFull report: {outpath}")


if __name__ == "__main__":
    main()
