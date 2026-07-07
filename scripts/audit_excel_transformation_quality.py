#!/usr/bin/env python3
"""Audit standardized Excel transformation quality against golden references.

Compares current export output against the golden samples and produces a
Markdown report + JSON summary.

Usage:
  python scripts/audit_excel_transformation_quality.py
  python scripts/audit_excel_transformation_quality.py --fail-on key,header,empty-required
  python scripts/audit_excel_transformation_quality.py --current-root /tmp/test_exports
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
_DEFAULT_CURRENT = REPO / "Exports" / "Excel"
_DEFAULT_GOLDEN = REPO / "data" / "examples"
_DEFAULT_OUTPUT = REPO / "Exports" / "audit"

TEMPLATE_MAP = {
    "instrument_list": "Standardized_Stellenplan.xlsx",
    "wiring": "Standardized_Klemmenplan.xlsx",
    "datasheet": "Standardized_Datasheet.xlsx",
    "stromlaufplan": "Standardized_Stromlaufplan.xlsx",
    "3d_assembly": "Assembly_3D_template.xlsx",
}

# Composite key rules: template → sheet → list of header-name groups
# Each group is a tuple of Row-1 header names that must be unique together.
COMPOSITE_KEY_SPEC: dict[str, dict[str, list[tuple[str, ...]]]] = {
    "instrument_list": {
        "Document_ID": [("Document_ID",)],
        "Document_Data": [("Document_ID", "Instrument_Sheet_ID")],
        "Object_ID": [("Document_ID", "Instrument_Sheet_ID", "Object_ID")],
    },
    "wiring": {
        "Document_ID": [("Document_ID",)],
        "Dokument_Data": [("Document_ID", "Documentblatt_ID")],
        "Object_ID": [("Document_ID", "Documentblatt_ID", "Object_ID")],
        "Terminal_ID": [("Document_ID", "Object_ID", "Terminal_ID")],
    },
    "datasheet": {
        "Device_ID": [("Document_ID", "Device_ID")],
        "Process_Attributes": [("Document_ID", "Device_ID", "Attribute_Key")],
        "Technical_Attributes": [("Document_ID", "Device_ID", "Attribute_Key")],
        "Geometric_Attributes": [("Document_ID", "Device_ID", "Attribute_Key")],
        "Manufacturer_Specific": [("Document_ID", "Device_ID", "Attribute_Key")],
    },
    "stromlaufplan": {
        "Document_ID": [("Document_ID",)],
        "Object_ID": [("Document_ID", "Object_ID")],
        "Element_ID": [("Document_ID", "Element_ID")],
        "Connection_Data": [("Document_ID", "Connection_Key")],
    },
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _header(ws) -> list[str | None]:
    """Return header row (row 1) — the stable field-key row."""
    return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]


def _data_rows(ws) -> list[tuple]:
    """Return data rows (row 3+) as tuples of cell values."""
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if any(v is not None and str(v).strip() != "" for v in row):
            rows.append(tuple(row))
    return rows


def _col_index(header: list[str | None], name: str) -> int | None:
    """Find 0-based column index for a header name (Row 1 key)."""
    for i, h in enumerate(header):
        if h is not None and str(h).strip() == name:
            return i
    return None


def _resolve_key_specs(
    header: list[str | None],
    key_specs: list[tuple[str, ...]],
) -> list[tuple[int, ...]]:
    """Convert header-name key specs to 0-based column-index specs."""
    result: list[tuple[int, ...]] = []
    for group in key_specs:
        indices: list[int] = []
        for name in group:
            ci = _col_index(header, name)
            if ci is not None:
                indices.append(ci)
        if len(indices) == len(group):
            result.append(tuple(indices))
    return result


def _check_composite_keys(
    rows: list[tuple],
    key_specs: list[tuple[str, ...]],
    header: list[str | None],
) -> list[str]:
    """Return issues for violated composite keys (by header name)."""
    issues = []
    index_specs = _resolve_key_specs(header, key_specs)
    for key_cols in index_specs:
        seen: dict[tuple, int] = {}
        for idx, row in enumerate(rows, start=3):
            key = tuple(
                str(row[c]).strip()
                if c < len(row) and row[c] is not None
                else ""
                for c in key_cols
            )
            if all(v == "" for v in key):
                continue
            if key in seen:
                issues.append(
                    f"  Duplicate key {key} at rows {seen[key]} and {idx} "
                    f"(cols {key_cols})"
                )
            else:
                seen[key] = idx
    return issues


def _load_source_manifest(path: Path | None) -> dict:
    """Load source manifest JSON, or return empty dict."""
    if path is None or not path.is_file():
        return {}
    import json as _json
    with open(path, encoding="utf-8") as f:
        return _json.load(f)


_SOURCE_KIND_TO_EXPECTED: dict[str, str] = {
    "device_datasheet": "datasheet",
    "stellen_tu": "instrument_list",
    "stellen_overview": "instrument_list",
    "klemmenplan": "wiring",
    "verschaltungsliste": "wiring",
    "cabinet_reference": "wiring",
    "stromlaufplan": "stromlaufplan",
    "ri_flowsheet": "instrument_list",
}

# Sheets that are expected to be empty when no source exists for the template
_TYPED_SHEETS_BY_TEMPLATE: dict[str, set[str]] = {
    "datasheet": {
        "Device_Classification", "Process_Attributes", "Technical_Attributes",
        "Geometric_Attributes", "Connection_Attributes",
    },
    "stromlaufplan": {"Connection_Data"},
    "wiring": set(),
    "instrument_list": set(),
}


def _source_aware_empty_reason(
    sheet_name: str,
    c_rows: list[tuple],
    g_rows: list[tuple],
    template: str,
    source_manifest: dict,
) -> str:
    """Categorize why a sheet is empty, considering source availability."""
    if c_rows:
        return ""

    # Check if the template has any source documents at all
    has_source = False
    for kind, tmpl in _SOURCE_KIND_TO_EXPECTED.items():
        if tmpl == template:
            count = source_manifest.get("source_counts", {}).get(kind, 0)
            if count > 0:
                has_source = True
                break

    # If the template has zero source documents of the relevant kind,
    # ALL sheets in that template are expected to be empty / no-source.
    if not has_source:
        typed_sheets = _TYPED_SHEETS_BY_TEMPLATE.get(template, set())
        if sheet_name in typed_sheets or g_rows:
            return "no source document"
        return "no evidence (also blank in golden)"

    # Template has sources — check typed sheets specifically
    typed_sheets = _TYPED_SHEETS_BY_TEMPLATE.get(template, set())
    if sheet_name in typed_sheets:
        if g_rows:
            return "missing extraction"
        return "no source document"

    # Generic: golden has data but current doesn't
    if g_rows:
        return "missing extraction"

    return "no evidence (also blank in golden)"


# ---------------------------------------------------------------------------
# Audit
# ---------------------------------------------------------------------------


def audit(current_root: Path, golden_root: Path,
          source_manifest_path: Path | None = None) -> dict:
    """Run audit and return structured results."""
    source_manifest = _load_source_manifest(source_manifest_path)
    results: dict[str, dict] = {}
    results["source_counts"] = source_manifest.get("source_counts", {})

    # Check canonical templates for stale data (row 3+ should be empty)
    _TEMPLATES_DIR = REPO / "data" / "templates"
    stale_templates: list[str] = []
    for tpl_path in sorted(_TEMPLATES_DIR.glob("*_template.xlsx")):
        try:
            wb = openpyxl.load_workbook(str(tpl_path), data_only=True)
            has_data = False
            for ws in wb.worksheets:
                if ws.max_row and ws.max_row >= 3:
                    for row in ws.iter_rows(min_row=3, max_row=min(ws.max_row, 5), values_only=True):
                        if any(v is not None for v in row):
                            has_data = True
                            break
                if has_data:
                    break
            wb.close()
            if has_data:
                stale_templates.append(tpl_path.name)
        except Exception:
            pass
    results["_stale_templates"] = stale_templates

    for subdir, golden_name in TEMPLATE_MAP.items():
        golden_path = golden_root / golden_name
        current_dir = current_root / subdir
        if not current_dir.exists():
            results[subdir] = {"error": f"Export dir not found: {current_dir}"}
            continue
        current_files = list(current_dir.glob("*.xlsx"))
        if not current_files:
            results[subdir] = {"error": "No .xlsx export found"}
            continue
        current_path = current_files[0]

        if not golden_path.is_file():
            # No golden reference — still check current export structure.
            results[subdir] = {
                "golden": str(golden_path.name),
                "current": str(current_path.name),
                "note": "no golden reference available",
                "sheets": {},
            }
            continue

        golden_wb = openpyxl.load_workbook(str(golden_path), data_only=True)
        current_wb = openpyxl.load_workbook(str(current_path), data_only=True)

        entry: dict = {
            "golden": str(golden_path.name),
            "current": str(current_path.name),
            "sheets": {},
        }

        all_sheets = sorted(set(golden_wb.sheetnames) | set(current_wb.sheetnames))
        for sheet_name in all_sheets:
            g_ws = golden_wb[sheet_name] if sheet_name in golden_wb.sheetnames else None
            c_ws = current_wb[sheet_name] if sheet_name in current_wb.sheetnames else None

            g_header = _header(g_ws) if g_ws else []
            c_header = _header(c_ws) if c_ws else []
            g_rows = _data_rows(g_ws) if g_ws else []
            c_rows = _data_rows(c_ws) if c_ws else []

            # Compare header keys (Row 1 field names).
            # Use intersection-based matching: columns that exist in both
            # must agree on position.  New columns in the current template
            # (not in golden) are acceptable — the template is the schema
            # authority, golden is a content reference only.
            g_keys = [str(h).strip() if h else "" for h in g_header]
            c_keys = [str(h).strip() if h else "" for h in c_header]
            _common = set(g_keys) & set(c_keys)
            _common.discard("")  # ignore empty column names
            _mismatch = any(
                g_keys.index(k) != c_keys.index(k)
                for k in _common
                if k in g_keys and k in c_keys
            )
            header_match = not _mismatch

            sheet_result: dict = {
                "golden_rows": len(g_rows),
                "current_rows": len(c_rows),
                "delta": len(c_rows) - len(g_rows),
                "header_match": header_match,
            }

            # Composite key check by header name
            key_specs = COMPOSITE_KEY_SPEC.get(subdir, {}).get(sheet_name, [])
            if key_specs and c_rows:
                key_issues = _check_composite_keys(c_rows, key_specs, c_header)
                if key_issues:
                    sheet_result["key_issues"] = key_issues

            # Empty sheet categorization (source-aware)
            empty_reason = _source_aware_empty_reason(
                sheet_name, c_rows, g_rows, subdir, source_manifest,
            )
            if empty_reason:
                sheet_result["empty_reason"] = empty_reason

            # Column-level analysis
            if c_rows and c_header:
                non_empty_cols = 0
                total_cols = len(c_rows[0]) if c_rows else 0
                blank_cols: list[str] = []
                for col_idx in range(total_cols):
                    vals = [
                        str(r[col_idx]).strip()
                        if col_idx < len(r) and r[col_idx] is not None
                        else ""
                        for r in c_rows
                    ]
                    if any(v for v in vals):
                        non_empty_cols += 1
                    else:
                        col_name = (
                            c_header[col_idx]
                            if col_idx < len(c_header) and c_header[col_idx]
                            else f"col_{col_idx}"
                        )
                        # Reason classification
                        if g_rows and col_idx < len(g_rows[0]) if g_rows else False:
                            g_has = any(
                                str(r[col_idx]).strip()
                                if col_idx < len(r) and r[col_idx] is not None
                                else ""
                                for r in g_rows
                            )
                            reason = "missing extraction" if g_has else "no evidence"
                        else:
                            reason = "no source evidence"
                        blank_cols.append(f"{col_name}: {reason}")
                sheet_result["non_empty_cols"] = non_empty_cols
                sheet_result["total_cols"] = total_cols
                if blank_cols:
                    sheet_result["blank_columns"] = blank_cols

            if not header_match and (g_keys or c_keys):
                sheet_result["golden_header"] = g_keys[:12]
                sheet_result["current_header"] = c_keys[:12]

            entry["sheets"][sheet_name] = sheet_result

        golden_wb.close()
        current_wb.close()
        results[subdir] = entry

    return results


# ---------------------------------------------------------------------------
# Report writing
# ---------------------------------------------------------------------------


def write_report(results: dict, output_dir: Path) -> Path:
    """Write Markdown + JSON audit reports."""
    os.makedirs(output_dir, exist_ok=True)

    json_path = output_dir / "excel_transformation_quality.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False, default=str)

    md_path = output_dir / "excel_transformation_quality.md"
    lines = [
        "# Excel Transformation Quality Audit",
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        "",
        "## Summary",
        "",
        "| Template | Sheet | Golden | Current | Δ | Header | Notes |",
        "|----------|-------|--------|---------|---|--------|-------|",
    ]

    for template, entry in sorted(results.items()):
        if template.startswith("_"):
            continue  # skip metadata keys
        if not isinstance(entry, dict):
            continue
        if "error" in entry:
            lines.append(
                f"| **{template}** | — | — | — | — | — | {entry['error']} |"
            )
            continue
        sheets = entry.get("sheets", {})
        first = True
        for sheet_name, info in sorted(sheets.items()):
            notes_parts = []
            if info.get("key_issues"):
                notes_parts.append(f"{len(info['key_issues'])} key violations")
            if info.get("blank_columns"):
                notes_parts.append(f"{len(info['blank_columns'])} blank cols")
            if info.get("empty_reason"):
                notes_parts.append(info["empty_reason"])
            if not info.get("header_match"):
                notes_parts.append("HEADER MISMATCH")
            notes = "; ".join(notes_parts) if notes_parts else "✓"
            header_ok = "✓" if info.get("header_match") else "✗"

            label = f"**{template}**" if first else ""
            first = False
            lines.append(
                f"| {label} | {sheet_name} | {info['golden_rows']} | "
                f"{info['current_rows']} | {info['delta']:+d} | {header_ok} | {notes} |"
            )

    # Key violations
    lines.append("")
    lines.append("## Key Violations")
    lines.append("")
    has_keys = False
    for template, entry in sorted(results.items()):
        if template.startswith("_") or not isinstance(entry, dict):
            continue
        if "error" in entry:
            continue
        for sheet_name, info in sorted(entry.get("sheets", {}).items()):
            if info.get("key_issues"):
                has_keys = True
                lines.append(f"### {template} → {sheet_name}")
                for issue in info["key_issues"]:
                    lines.append(f"- {issue}")
                lines.append("")
    if not has_keys:
        lines.append("No composite-key violations found.")

    # Blank columns
    lines.append("")
    lines.append("## Blank Business Columns")
    lines.append("")
    has_blanks = False
    for template, entry in sorted(results.items()):
        if template.startswith("_") or not isinstance(entry, dict):
            continue
        if "error" in entry:
            continue
        for sheet_name, info in sorted(entry.get("sheets", {}).items()):
            if info.get("blank_columns"):
                has_blanks = True
                lines.append(f"### {template} → {sheet_name}")
                for bc in info["blank_columns"]:
                    lines.append(f"- {bc}")
                lines.append("")
    if not has_blanks:
        lines.append("No blank business columns found.")

    # Empty sheets
    lines.append("")
    lines.append("## Empty Sheets")
    lines.append("")
    has_empty = False
    for template, entry in sorted(results.items()):
        if template.startswith("_") or not isinstance(entry, dict):
            continue
        if "error" in entry:
            continue
        for sheet_name, info in sorted(entry.get("sheets", {}).items()):
            if info.get("empty_reason"):
                has_empty = True
                lines.append(
                    f"- **{template} → {sheet_name}**: {info['empty_reason']}"
                )
    if not has_empty:
        lines.append("No empty sheets found.")
    lines.append("")

    lines.append("")
    lines.append("---")
    lines.append(
        f"*Report generated by `scripts/audit_excel_transformation_quality.py`*"
    )

    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return md_path


# ---------------------------------------------------------------------------
# Fail-on logic
# ---------------------------------------------------------------------------


def check_fail_on(results: dict, fail_criteria: set[str]) -> list[str]:
    """Return a list of failure messages for the given criteria."""
    failures: list[str] = []

    # stale-data: check once per dirty template (outside sheet loop to avoid duplicates)
    if "stale-data" in fail_criteria:
        for st in results.get("_stale_templates", []):
            failures.append(f"[stale-data] template: {st}")

    for template, entry in results.items():
        if template.startswith("_") or not isinstance(entry, dict):
            continue
        if "error" in entry:
            if "missing-export" in fail_criteria:
                failures.append(
                    f"[missing-export] {template}: {entry['error']}"
                )
            continue
        for sheet_name, info in entry.get("sheets", {}).items():
            if "key" in fail_criteria and info.get("key_issues"):
                failures.append(
                    f"[key] {template}/{sheet_name}: "
                    f"{len(info['key_issues'])} violations"
                )
            if "header" in fail_criteria and not info.get("header_match"):
                failures.append(
                    f"[header] {template}/{sheet_name}: header mismatch"
                )
            if "empty-required" in fail_criteria and info.get("empty_reason"):
                reason = info["empty_reason"]
                if reason == "missing extraction":
                    failures.append(
                        f"[empty-required] {template}/{sheet_name}: {reason}"
                    )
            if "no-source" in fail_criteria and info.get("empty_reason") == "no source document":
                failures.append(
                    f"[no-source] {template}/{sheet_name}: no source document"
                )

    return failures


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Audit Excel transformation quality against golden samples."
    )
    parser.add_argument(
        "--current-root",
        type=Path,
        default=_DEFAULT_CURRENT,
        help=f"Root of current export files (default: {_DEFAULT_CURRENT})",
    )
    parser.add_argument(
        "--golden-root",
        type=Path,
        default=_DEFAULT_GOLDEN,
        help=f"Root of golden sample files (default: {_DEFAULT_GOLDEN})",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=_DEFAULT_OUTPUT,
        help=f"Output directory for reports (default: {_DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--source-manifest",
        type=Path,
        default=None,
        help="Path to source_manifest.json for source-aware audit",
    )
    parser.add_argument(
        "--fail-on",
        type=str,
        default="",
        help="Comma-separated criteria to fail on: key,header,empty-required,missing-export,no-source,stale-data",
    )
    args = parser.parse_args()

    fail_criteria = set(
        c.strip() for c in args.fail_on.split(",") if c.strip()
    )

    print(f"Running Excel transformation quality audit...")
    print(f"  Current: {args.current_root}")
    print(f"  Golden:  {args.golden_root}")

    results = audit(args.current_root, args.golden_root, args.source_manifest)
    md_path = write_report(results, args.output_dir)
    print(f"Audit complete:")
    print(f"  Markdown: {md_path}")
    print(f"  JSON:     {args.output_dir / 'excel_transformation_quality.json'}")

    if fail_criteria:
        failures = check_fail_on(results, fail_criteria)
        if failures:
            print(f"\nFAILURES ({len(failures)}):")
            for f in failures:
                print(f"  {f}")
            return 1
        print("\nAll fail-on criteria passed.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
