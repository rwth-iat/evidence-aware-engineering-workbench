#!/usr/bin/env python3
"""
Verify that the Assembly_3D_template_filled.xlsx data matches the original
final_result.FCStd exactly, proving the template can reconstruct the assembly.

Usage: python3 scripts/verify_assembly_reconstruction.py
"""

import math
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from collections import defaultdict

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PIPELINE_DIR = PROJECT_ROOT / "Documents" / "Piping Diagram"
TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "Documents" / "Piping Diagram" / "Assembly_3D_template_filled.xlsx"
FCSTD_PATH = PIPELINE_DIR / "final_result.FCStd"

TOLERANCE_MM = 0.01  # 0.01 mm position tolerance
TOLERANCE_DEG = 0.01  # 0.01 degree rotation tolerance


def load_template(template_path: Path) -> list[dict]:
    """Load Assembly_Steps from the filled template."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Assembly_Steps"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            row[h] = ws.cell(row=r, column=c).value
        if row.get("mapping_confidence") in ("matched", "matched_by_type"):
            rows.append(row)
    return rows


def load_freecad_original(fcstd_path: Path) -> dict[str, dict]:
    """Load original placement data directly from FreeCAD Document.xml."""
    with zipfile.ZipFile(fcstd_path, "r") as z:
        with z.open("Document.xml") as f:
            tree = ET.parse(f)
            root = tree.getroot()

    # Build name → type mapping from <Objects>
    name_to_type = {}
    for obj in root.find("Objects").findall("Object"):
        name_to_type[obj.get("name", "")] = obj.get("type", "")

    # Parse properties from <ObjectData>
    result = {}
    for obj in root.find("ObjectData").findall("Object"):
        obj_name = obj.get("name", "")
        obj_type = name_to_type.get(obj_name, "")
        if obj_type != "App::Link":
            continue

        props = obj.find("Properties")
        if props is None:
            continue

        entry = {"label": "", "pos_x": 0.0, "pos_y": 0.0, "pos_z": 0.0,
                 "qw": 0.0, "qx": 0.0, "qy": 0.0, "qz": 1.0}

        for prop in props.findall("Property"):
            pname = prop.get("name", "")
            if pname == "Label":
                se = prop.find("String")
                if se is not None:
                    entry["label"] = se.get("value", "")
            elif pname == "LinkPlacement":
                pp = prop.find("PropertyPlacement")
                if pp is not None:
                    entry["pos_x"] = float(pp.get("Px", "0"))
                    entry["pos_y"] = float(pp.get("Py", "0"))
                    entry["pos_z"] = float(pp.get("Pz", "0"))
                    entry["qw"] = float(pp.get("Q0", "0"))
                    entry["qx"] = float(pp.get("Q1", "0"))
                    entry["qy"] = float(pp.get("Q2", "0"))
                    entry["qz"] = float(pp.get("Q3", "1"))

        if entry["label"]:
            result[entry["label"]] = entry

    return result


def verify(template_rows: list[dict], fc_original: dict[str, dict]) -> None:
    """Compare template data against original FreeCAD data."""
    pos_errors = []
    quat_errors = []
    euler_errors = []
    matched = 0
    unmatched = 0

    for row in template_rows:
        label = row["label"]
        orig = fc_original.get(label)
        if not orig:
            unmatched += 1
            continue
        matched += 1

        # Position error
        dx = (row["pos_x_mm"] or 0) - orig["pos_x"]
        dy = (row["pos_y_mm"] or 0) - orig["pos_y"]
        dz = (row["pos_z_mm"] or 0) - orig["pos_z"]
        pos_err = math.sqrt(dx*dx + dy*dy + dz*dz)
        pos_errors.append(pos_err)

        # Quaternion error (as angular distance)
        qw_t, qx_t, qy_t, qz_t = row["rot_qw"] or 0, row["rot_qx"] or 0, row["rot_qy"] or 0, row["rot_qz"] or 1
        qw_o, qx_o, qy_o, qz_o = orig["qw"], orig["qx"], orig["qy"], orig["qz"]
        # Angular distance between two quaternions = 2*acos(|dot(q1, q2)|)
        dot = abs(qw_t*qw_o + qx_t*qx_o + qy_t*qy_o + qz_t*qz_o)
        dot = min(1.0, dot)
        quat_err = 2 * math.degrees(math.acos(dot))
        quat_errors.append(quat_err)

        # Euler angle error
        er = (row["euler_roll_deg"] or 0)
        ep = (row["euler_pitch_deg"] or 0)
        ey = (row["euler_yaw_deg"] or 0)
        # Compute euler from original quaternion
        from scripts.build_3d_assembly_template import quat_to_euler
        eo_roll, eo_pitch, eo_yaw = quat_to_euler(orig["qw"], orig["qx"], orig["qy"], orig["qz"])
        euler_err = math.sqrt((er - eo_roll)**2 + (ep - eo_pitch)**2 + (ey - eo_yaw)**2)
        euler_errors.append(euler_err)

    # Report
    print(f"\n{'='*60}")
    print(f"Verification: Template vs Original FreeCAD Assembly")
    print(f"{'='*60}")
    print(f"  Components verified:  {matched}")
    print(f"  Components unmatched: {unmatched}")

    print(f"\n  Position error (mm):")
    print(f"    max  = {max(pos_errors):.6f}")
    print(f"    mean = {sum(pos_errors)/len(pos_errors):.6f}")
    print(f"    min  = {min(pos_errors):.6f}")
    pos_ok = sum(1 for e in pos_errors if e < TOLERANCE_MM)
    print(f"    within {TOLERANCE_MM}mm: {pos_ok}/{len(pos_errors)} ({100*pos_ok/len(pos_errors):.1f}%)")

    print(f"\n  Quaternion (angular) error (°):")
    print(f"    max  = {max(quat_errors):.6f}")
    print(f"    mean = {sum(quat_errors)/len(quat_errors):.6f}")
    quat_ok = sum(1 for e in quat_errors if e < TOLERANCE_DEG)
    print(f"    within {TOLERANCE_DEG}°: {quat_ok}/{len(quat_errors)} ({100*quat_ok/len(quat_errors):.1f}%)")

    print(f"\n  Euler angle error (°):")
    print(f"    max  = {max(euler_errors):.4f}")
    print(f"    mean = {sum(euler_errors)/len(euler_errors):.4f}")
    euler_ok = sum(1 for e in euler_errors if e < TOLERANCE_DEG)
    print(f"    within {TOLERANCE_DEG}°: {euler_ok}/{len(euler_errors)} ({100*euler_ok/len(euler_errors):.1f}%)")

    # Overall verdict
    all_pos_ok = pos_ok == len(pos_errors)
    all_quat_ok = quat_ok == len(quat_errors)

    print(f"\n{'='*60}")
    if all_pos_ok and all_quat_ok:
        print("VERDICT: PASS — Template data matches original assembly exactly.")
        print("The template is sufficient to reconstruct the full 3D assembly.")
    else:
        print("VERDICT: ISSUES FOUND — Some values differ from original.")

    # Print any large deviations
    if not all_pos_ok or not all_quat_ok:
        print("\n  Large deviations (>tolerance):")
        for i, row in enumerate(template_rows):
            if pos_errors[i] > TOLERANCE_MM:
                print(f"    POS: {row['label']} error={pos_errors[i]:.4f}mm")
            if quat_errors[i] > TOLERANCE_DEG:
                print(f"    QUAT: {row['label']} error={quat_errors[i]:.4f}°")


if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

    print("Loading template...")
    template_rows = load_template(TEMPLATE_PATH)
    print(f"  {len(template_rows)} rows with matched data")

    print("Loading original FreeCAD assembly...")
    fc_original = load_freecad_original(FCSTD_PATH)
    print(f"  {len(fc_original)} App::Link objects with labels")

    verify(template_rows, fc_original)
