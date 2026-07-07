#!/usr/bin/env python3
"""
Build a mapping table: P&ID instruments → 3D pipeline components.

Reads the inconsistency report and the assembly template, produces an Excel
mapping sheet that the user can fill in to link instruments to pipeline parts.
"""

import sys
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = Path(__file__).resolve().parents[1]

UC1_PATH = PROJECT_ROOT / "Exports/UseCase1/use_case_1_standardized_transformation.xlsx"
TEMPLATE_PATH = PROJECT_ROOT / "Documents/Piping Diagram/Assembly_3D_template_filled.xlsx"
OUTPUT_PATH = PROJECT_ROOT / "Documents/Piping Diagram/_legacy/instrument_pipe_mapping.xlsx"

HEADER_FONT = Font(name="Calibri", size=11, bold=True)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def load_instruments():
    """Load instrument list from UC1 report."""
    wb = openpyxl.load_workbook(UC1_PATH)
    ws = wb["ri_devices"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    instruments = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=c + 1).value for c, h in enumerate(headers)}
        instruments.append(row)
    return instruments


def load_pipeline_components():
    """Load all pipeline components from assembly template, grouped by type."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Assembly_Steps"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    components = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=c + 1).value for c, h in enumerate(headers)}
        if row.get("mapping_confidence") in ("matched", "matched_by_type"):
            components.append(row)
    return components


def build_mapping_sheet(instruments, pipeline_comps):
    """Create the mapping Excel."""

    # Group pipeline components by type for reference
    by_type = defaultdict(list)
    for c in pipeline_comps:
        by_type[c["type"]].append(c)

    # Group instruments by anchor component
    by_comp = defaultdict(list)
    no_anchor = []
    for inst in instruments:
        anchor = (inst.get("piping_anchor_id") or "").strip()
        comp = (inst.get("piping_component_name") or "").strip()
        if comp:
            by_comp[comp].append(inst)
        else:
            no_anchor.append(inst)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Mapping (user fills this in) ──
    ws1 = wb.create_sheet("Instrument_Pipe_Mapping")
    map_headers = [
        "pand_id_component", "pand_id_anchor", "instrument_count",
        "instrument_tags",
        "suggested_3d_type",
        "mapped_3d_label",          # ← user fills
        "mapped_3d_fcstd_name",     # ← auto-filled
        "pos_x_mm", "pos_y_mm", "pos_z_mm",  # ← auto-filled
        "confidence", "notes",
    ]
    for ci, h in enumerate(map_headers, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    row_idx = 2
    for comp_name, insts in sorted(by_comp.items()):
        tags = ", ".join(i.get("canonical_tag", "") for i in insts)
        anchor = insts[0].get("piping_anchor_id", "")
        # Suggest 3D type based on component naming convention
        if comp_name.startswith("R"):
            suggested = "tank"
        elif comp_name.startswith("VV"):
            suggested = "blackbox"
        elif comp_name.startswith("DM"):
            suggested = "blackbox"
        elif comp_name.startswith("B"):
            suggested = "blackbox"
        else:
            suggested = "?"

        ws1.cell(row=row_idx, column=1, value=comp_name).border = THIN_BORDER
        ws1.cell(row=row_idx, column=2, value=anchor).border = THIN_BORDER
        ws1.cell(row=row_idx, column=3, value=len(insts)).border = THIN_BORDER
        ws1.cell(row=row_idx, column=4, value=tags).border = THIN_BORDER
        ws1.cell(row=row_idx, column=5, value=suggested).border = THIN_BORDER
        # Highlight the user-fill column
        ws1.cell(row=row_idx, column=6).border = THIN_BORDER
        ws1.cell(row=row_idx, column=6).fill = YELLOW_FILL
        ws1.cell(row=row_idx, column=10, value="to be filled").border = THIN_BORDER

        # List available 3D components of suggested type
        candidates = by_type.get(suggested, [])
        if candidates:
            names = ", ".join(c["label"] for c in candidates[:8])
            ws1.cell(row=row_idx, column=11, value=f"candidates: {names}").border = THIN_BORDER

        row_idx += 1

    # No-anchor instruments
    if no_anchor:
        ws1.cell(row=row_idx, column=1, value="(no anchor)").border = THIN_BORDER
        tags = ", ".join(i.get("canonical_tag", "") for i in no_anchor)
        ws1.cell(row=row_idx, column=4, value=tags).border = THIN_BORDER
        ws1.cell(row=row_idx, column=6).fill = YELLOW_FILL
        ws1.cell(row=row_idx, column=6).border = THIN_BORDER

    # Auto-width
    for ci in range(1, 12):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22

    # ── Sheet 2: 3D Pipeline Components (reference) ──
    ws2 = wb.create_sheet("3D_Pipeline_Components")
    ref_headers = ["label", "type", "fcstd_link_name", "ifc_part_name",
                   "pos_x_mm", "pos_y_mm", "pos_z_mm", "ifc_parts_library"]
    for ci, h in enumerate(ref_headers, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for ri, c in enumerate(pipeline_comps, 2):
        for ci, h in enumerate(ref_headers, 1):
            cell = ws2.cell(row=ri, column=ci, value=c.get(h, ""))
            cell.border = THIN_BORDER

    for ci in range(1, len(ref_headers) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22

    # ── Sheet 3: Instrument List (reference) ──
    ws3 = wb.create_sheet("Instruments_from_PID")
    inst_headers = ["canonical_tag", "function_code", "piping_anchor_id",
                    "piping_component_name", "actuating_location", "device_information"]
    for ci, h in enumerate(inst_headers, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for ri, inst in enumerate(instruments, 2):
        for ci, h in enumerate(inst_headers, 1):
            cell = ws3.cell(row=ri, column=ci, value=inst.get(h, ""))
            cell.border = THIN_BORDER

    for ci in range(1, len(inst_headers) + 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22

    wb.save(str(OUTPUT_PATH))
    print(f"Mapping table saved to: {OUTPUT_PATH}")


def main():
    print("Loading instruments...")
    instruments = load_instruments()
    print(f"  {len(instruments)} instruments from P&ID")

    print("Loading pipeline components...")
    pipeline_comps = load_pipeline_components()
    print(f"  {len(pipeline_comps)} pipeline components from 3D assembly")

    # Summary
    by_type = defaultdict(list)
    for c in pipeline_comps:
        by_type[c["type"]].append(c)
    print(f"\n  3D components by type:")
    for t, comps in sorted(by_type.items()):
        print(f"    {t}: {len(comps)} ({comps[0]['label']} ~ {comps[-1]['label']})")

    print("\nBuilding mapping table...")
    build_mapping_sheet(instruments, pipeline_comps)

    print("\nNext step: open the Excel, fill in the yellow 'mapped_3d_label' column.")
    print("For each P&ID component (R003, B1, VV001...), pick the matching 3D pipeline label.")
    print("Use sheet '3D_Pipeline_Components' as reference.")


if __name__ == "__main__":
    main()
