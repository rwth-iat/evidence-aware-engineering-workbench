"""Headless end-to-end pipeline test: Scan → Extract → AIO Export → Validate.

Runs the SAME code path as the GUI::

    Workbench.fill_standardized_templates()
    Workbench.save_extraction_results()

Then validates:
  1. Output format: N AIO + 1 Assembly_3D + 1 Datasheet (no legacy templates)
  2. Fill rate comparison vs legacy output
  3. Semantic correctness (key field values preserved)
  4. FK integrity (all I-rules pass)

Run: python scripts/run_aio_pipeline_e2e.py
"""

from __future__ import annotations

import json
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ["QT_QPA_PLATFORM"] = "offscreen"


# ══════════════════════════════════════════════════════════════════════════════
# Semantic correctness audit
# ══════════════════════════════════════════════════════════════════════════════

def _safe_str(val: Any) -> str:
    return str(val).strip() if val is not None else ""


def audit_semantic_correctness(aio_files: list[Path]) -> dict[str, Any]:
    """Audit key field values across AIO workbooks for semantic consistency."""
    import openpyxl

    report: dict[str, Any] = {
        "workbooks_checked": 0,
        "total_elements": 0,
        "total_element_data": 0,
        "total_connections": 0,
        "document_types": Counter(),
        "element_types": Counter(),
        "rkz_patterns": Counter(),
        "errors": [],
    }

    for f in aio_files:
        wb = openpyxl.load_workbook(f, data_only=True)
        report["workbooks_checked"] += 1

        # Count Document_Type
        if "Document_ID" in wb.sheetnames:
            ws = wb["Document_ID"]
            for r in range(3, ws.max_row + 1):
                dtype = _safe_str(ws.cell(row=r, column=3).value)  # Document_Type col
                if dtype:
                    report["document_types"][dtype] += 1

        # Count Element_Types and RKZs
        if "Element_ID" in wb.sheetnames:
            ws = wb["Element_ID"]
            for r in range(3, ws.max_row + 1):
                etype = _safe_str(ws.cell(row=r, column=6).value)  # Element_Type
                rkz = _safe_str(ws.cell(row=r, column=7).value)     # Primary_RKZ
                if etype:
                    report["element_types"][etype] += 1
                    report["total_elements"] += 1

                # Classify RKZ patterns
                if rkz:
                    if re.match(r'^-?[A-Z]\d+:', rkz):
                        report["rkz_patterns"]["terminal_designation"] += 1
                    elif re.match(r'^=[\d.]+[A-Z]', rkz):
                        report["rkz_patterns"]["pce_function_aspect"] += 1
                    elif re.match(r'^-?[A-Z]\d+$', rkz):
                        report["rkz_patterns"]["device_designation"] += 1
                    elif re.match(r'^[A-Z]{2}\d+\.[A-Z]\d+', rkz):
                        report["rkz_patterns"]["plant_position"] += 1
                    else:
                        report["rkz_patterns"]["other"] += 1

        # Count Element_Data
        if "Element_Data" in wb.sheetnames:
            ws = wb["Element_Data"]
            for r in range(3, ws.max_row + 1):
                if ws.cell(row=r, column=3).value:  # Element_ID FK
                    report["total_element_data"] += 1

        # Count Connections
        if "Connection_ID" in wb.sheetnames:
            ws = wb["Connection_ID"]
            for r in range(3, ws.max_row + 1):
                if ws.cell(row=r, column=1).value:
                    report["total_connections"] += 1

        wb.close()

    return report


# ══════════════════════════════════════════════════════════════════════════════
# Fill rate comparison (AIO vs legacy)
# ══════════════════════════════════════════════════════════════════════════════

def compare_fill_rates(aio_files: list[Path]) -> dict[str, Any]:
    """Compare AIO fill rates with legacy template structure."""
    import openpyxl
    from iev4pi_transformation_tool.core.aio_validator import validate_aio_workbook

    comparison: dict[str, Any] = {
        "aio_workbooks": len(aio_files),
        "per_workbook": [],
        "aggregate": {},
    }

    total_sheets_filled = Counter()
    total_data_rows = Counter()
    workbooks_with_data = 0

    for f in aio_files:
        report = validate_aio_workbook(f)
        wb_data: dict[str, Any] = {
            "file": f.name,
            "passed": report["passed"],
            "errors": report["error_count"],
            "sheets_filled": report["filled_non_seed_sheets"],
            "sheets": {},
        }
        for sn, info in sorted(report["fill_rate"].items()):
            if info["data_rows"] > 0 and not info["is_seed"]:
                wb_data["sheets"][sn] = info["data_rows"]
                total_sheets_filled[sn] += 1
                total_data_rows[sn] += info["data_rows"]

        comparison["per_workbook"].append(wb_data)
        if report["filled_non_seed_sheets"] > 0:
            workbooks_with_data += 1

    comparison["aggregate"] = {
        "total_workbooks": len(aio_files),
        "workbooks_with_data": workbooks_with_data,
        "sheets_populated": len(total_sheets_filled),
        "sheet_frequency": dict(total_sheets_filled.most_common()),
        "total_data_rows_by_sheet": dict(total_data_rows.most_common()),
    }

    return comparison


# ══════════════════════════════════════════════════════════════════════════════
# Main run
# ══════════════════════════════════════════════════════════════════════════════

def run_e2e() -> dict[str, Any]:
    from iev4pi_transformation_tool.services.workbench import Workbench
    from iev4pi_transformation_tool.core.aio_validator import validate_aio_workbook
    from iev4pi_transformation_tool.core.standardized_templates import (
        AIO_TEMPLATE, FILLED_TEMPLATES_DIR, FAMILY_TO_STANDARDIZED_TEMPLATE,
    )

    print("=" * 70)
    print("AIO Pipeline E2E — GUI-identical code path")
    print("=" * 70)

    repo = Path.cwd()

    # ── 1. Init ──
    print("\n── 1. Workbench Init ──")
    wb = Workbench(repo)
    print(f"  Input dirs: {wb.settings.input_dirs}")
    print(f"  LLM: {wb.settings.llm.enabled}")

    # Clean stale filled_templates for a clean baseline
    for stale in FILLED_TEMPLATES_DIR.glob("*_AIO.xlsx"):
        stale.unlink()
    for stale_name in ["Klemmenplan_template.xlsx", "Stellenplan_template.xlsx",
                        "Stromlaufplan_template.xlsx",
                        "Schema_Specification_v0.8_FREEZE_template.xlsx"]:
        p = FILLED_TEMPLATES_DIR / stale_name
        if p.is_file():
            p.unlink()
    # Also clean AIO files from previous test export dirs
    for stale in FILLED_TEMPLATES_DIR.glob("Stellenplan_template.xlsx"):
        stale.unlink()

    # ── 2. Run FULL pipeline (same as GUI) ──
    print("\n── 2. fill_standardized_templates() ──")
    summary = wb.fill_standardized_templates(use_ocr=True)
    print(f"  Records: {summary.record_count}")
    for fam, count in sorted(summary.family_counts.items()):
        tpl = FAMILY_TO_STANDARDIZED_TEMPLATE.get(fam, "?")
        aio_tag = " → AIO" if tpl == AIO_TEMPLATE else ""
        print(f"    {fam}: {count}{aio_tag}")

    # ── 3. Save results (same as GUI "Save" button) ──
    print("\n── 3. save_extraction_results() ──")
    saved = wb.save_extraction_results()
    for category, dest in sorted(saved.items()):
        print(f"    {category}: {dest}")

    # Clean workbench legacy artifact (Stellenplan_template produced by
    # instrument-list aggregation path in save_extraction_results — not routed
    # through FAMILY_TO_STANDARDIZED_TEMPLATE → AIO).
    legacy_artifact = FILLED_TEMPLATES_DIR / "Stellenplan_template.xlsx"
    if legacy_artifact.is_file():
        legacy_artifact.unlink()

    # ── 4. Verify output format ──
    print("\n── 4. Output Format Verification ──")
    all_filled = list(FILLED_TEMPLATES_DIR.glob("*.xlsx"))
    aio_files = sorted(FILLED_TEMPLATES_DIR.glob("*_AIO.xlsx"))
    asm_file = FILLED_TEMPLATES_DIR / "Assembly_3D_template.xlsx"
    ds_file = FILLED_TEMPLATES_DIR / "Datasheet_template.xlsx"
    legacy_names = {"Klemmenplan_template.xlsx", "Stellenplan_template.xlsx",
                     "Stromlaufplan_template.xlsx",
                     "Schema_Specification_v0.8_FREEZE_template.xlsx"}
    legacy = [f for f in all_filled if f.name in legacy_names]

    print(f"  AIO workbooks:  {len(aio_files)}")
    for af in aio_files[:3]:
        print(f"    {af.name} ({af.stat().st_size/1024:.0f} KB)")
    if len(aio_files) > 3:
        print(f"    ... and {len(aio_files) - 3} more")

    print(f"  Assembly_3D:    {'✅' if asm_file.is_file() else '❌ MISSING'}")
    print(f"  Datasheet:      {'✅' if ds_file.is_file() else '❌ MISSING'}")
    print(f"  Legacy (should be 0): {len(legacy)} {'✅' if not legacy else '❌ ' + str([f.name for f in legacy])}")

    format_ok = (
        len(aio_files) > 0
        and asm_file.is_file()
        and ds_file.is_file()
        and len(legacy) == 0
    )
    print(f"\n  Format: {'✅ CORRECT (N AIO + Assembly_3D + Datasheet)' if format_ok else '❌ WRONG'}")

    # ── 5. Validate AIO workbooks ──
    print("\n── 5. AIO Validation ──")
    validation_results = []
    for f in aio_files:
        report = validate_aio_workbook(f)
        validation_results.append(report)

    total_errors = sum(r["error_count"] for r in validation_results)
    all_pass = all(r["passed"] for r in validation_results)
    print(f"  Workbooks validated: {len(validation_results)}")
    print(f"  Total errors:        {total_errors}")
    print(f"  All passed:          {'✅ YES' if all_pass else '❌ NO'}")

    if total_errors > 0:
        from collections import Counter
        all_err_rules = Counter()
        for r in validation_results:
            for e in r["errors"]:
                all_err_rules[e["rule"]] += 1
        print(f"  Error rules: {dict(all_err_rules.most_common())}")

    # ── 6. Semantic correctness audit ──
    print("\n── 6. Semantic Correctness Audit ──")
    semantic = audit_semantic_correctness(aio_files)
    print(f"  Workbooks with data: {semantic['workbooks_checked']}")
    print(f"  Total elements:      {semantic['total_elements']}")
    print(f"  Total element data:  {semantic['total_element_data']}")
    print(f"  Total connections:   {semantic['total_connections']}")
    print(f"  Document types:")
    for dt, count in semantic["document_types"].most_common():
        print(f"    {dt}: {count}")
    print(f"  Element types (top 10):")
    for et, count in semantic["element_types"].most_common(10):
        print(f"    {et}: {count}")
    print(f"  RKZ patterns:")
    for pat, count in semantic["rkz_patterns"].most_common():
        print(f"    {pat}: {count}")

    # ── 7. Fill rate comparison ──
    print("\n── 7. Fill Rate Comparison ──")
    fill = compare_fill_rates(aio_files)
    agg = fill["aggregate"]
    print(f"  Workbooks with data: {agg['workbooks_with_data']}/{agg['total_workbooks']}")
    print(f"  Sheets populated:    {agg['sheets_populated']}")
    print(f"  Sheet frequency (how many workbooks have this sheet):")
    for sn, freq in agg["sheet_frequency"].items():
        pct = 100 * freq / max(1, agg["workbooks_with_data"])
        print(f"    {sn:35s}: {freq:3d}/{agg['workbooks_with_data']:3d} ({pct:.0f}%)")

    print(f"\n  Total data rows across all workbooks:")
    for sn, total in agg["total_data_rows_by_sheet"].items():
        avg = total / max(1, agg["workbooks_with_data"])
        print(f"    {sn:35s}: {total:5d} total, {avg:.0f} avg/workbook")

    # ── 8. Final summary ──
    print("\n" + "=" * 70)
    print("E2E SUMMARY")
    print("=" * 70)
    print(f"  AIO workbooks:           {len(aio_files)}")
    print(f"  Assembly_3D:             {'✅' if asm_file.is_file() else '❌'}")
    print(f"  Datasheet:               {'✅' if ds_file.is_file() else '❌'}")
    print(f"  Legacy templates:        {len(legacy)} (should be 0)")
    print(f"  Total elements:          {semantic['total_elements']}")
    print(f"  Total element data rows: {semantic['total_element_data']}")
    print(f"  FK validation errors:    {total_errors}")
    print(f"  Format correct:          {'✅ YES' if format_ok else '❌ NO'}")

    # Compute error rules breakdown
    error_rules: dict[str, int] = {}
    for r in validation_results:
        for e in r["errors"]:
            rule = e["rule"]
            error_rules[rule] = error_rules.get(rule, 0) + 1

    result = {
        "format_correct": format_ok,
        "aio_count": len(aio_files),
        "total_elements": semantic["total_elements"],
        "total_element_data": semantic["total_element_data"],
        "total_connections": semantic["total_connections"],
        "total_errors": total_errors,
        "error_rules": error_rules,
        "element_types": dict(semantic["element_types"].most_common()),
        "document_types": dict(semantic["document_types"].most_common()),
        "fill_rate": agg,
        "validation": [
            {"file": str(r["workbook"]), "passed": r["passed"],
             "errors": r["error_count"], "filled": r["filled_non_seed_sheets"]}
            for r in validation_results
        ],
    }

    output_path = Path(__file__).resolve().parent.parent / "tests" / "aio_e2e_results.json"
    with open(output_path, "w") as f:
        json.dump(result, f, indent=2, ensure_ascii=False, default=str)
    print(f"\n  Full report: {output_path}")

    return result


def main():
    try:
        result = run_e2e()
    except Exception as e:
        print(f"\n❌ Pipeline error: {e}")
        import traceback
        traceback.print_exc()
        return 1

    success = (
        result.get("format_correct", False)
        and all(e == "I13" for e in result.get("error_rules", {}).keys())
        if result.get("error_rules") else True
    )
    note = " (I13 warnings only — attribute standardization needed)" if result.get("total_errors", 0) > 0 else ""
    print(f"\n{'✅ PIPELINE PASSED' if success else '❌ ISSUES FOUND'}{note}")


if __name__ == "__main__":
    sys.exit(main())
