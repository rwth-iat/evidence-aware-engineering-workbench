"""Convert filled Standardized_*.xlsx files into blank templates.

Keeps row 1 (English column header) and row 2 (German legend / FK description)
together with all styling, column widths, frozen panes, and merged-cell ranges
that intersect those rows. All data rows (3..max_row) are removed.

Run once after the data/examples/ files change:

    python scripts/build_standardized_blank_templates.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = REPO_ROOT / "data" / "examples"
TARGET_DIR = REPO_ROOT / "data" / "templates"

JOBS = [
    ("Standardized_Stellenplan.xlsx", "Stellenplan_template.xlsx"),
    ("Standardized_Klemmenplan.xlsx", "Klemmenplan_template.xlsx"),
]


def blank_workbook(source: Path, target: Path) -> None:
    wb = openpyxl.load_workbook(source)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 2:
            ws.delete_rows(3, ws.max_row - 2)
        # Ensure freeze panes start at row 3 so legend stays visible.
        ws.freeze_panes = f"{get_column_letter(1)}3"
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    print(f"wrote {target.relative_to(REPO_ROOT)}")


def main() -> None:
    for src_name, dst_name in JOBS:
        blank_workbook(SOURCE_DIR / src_name, TARGET_DIR / dst_name)


if __name__ == "__main__":
    main()
