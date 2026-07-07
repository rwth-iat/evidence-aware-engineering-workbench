#!/usr/bin/env python3
"""
Rebuild Assembly_3D_template_filled.xlsx — 6-sheet standard format.
Parameterized: accepts V1_DIR, PROJECT_DIR, OUTPUT as CLI args or env vars.
"""
import csv, math, os, sys, xml.etree.ElementTree as ET, zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

import ifcopenshell, openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════
# CONFIG (override via env or CLI)
# ══════════════════════════════════════════════════════════════════════════
PROJECT = Path(os.environ.get("PROJECT_DIR", Path.cwd()))
V1_DIR = Path(os.environ.get("V1_DIR", str(PROJECT / "Documents" / "Piping Diagram")))
V0_DIR = Path(os.environ.get("V0_DIR", str(PROJECT / "Documents" / "Piping Diagram")))
OUT = Path(os.environ.get("OUTPUT", str(PROJECT / "Documents" / "Piping Diagram" / "Assembly_3D_template_filled.xlsx")))

FCSTD_PATH = V1_DIR / "final_result.FCStd"
CSV_PATH = V1_DIR / "final_result_components.csv"

# ══════════════════════════════════════════════════════════════════════════
# FINAL AKZ MAPPING (2025-05-15)
# ══════════════════════════════════════════════════════════════════════════
AKZ_RAW = """
tank_001, TU10.B1
tank_002, TU20.B2
tank_003, TU30.B3
blackbox_001, TU10.N18, PL002
blackbox_002, TU10.N13, PL001
blackbox_003, TU10.P14
blackbox_004, TU10.P19
blackbox_005, TU30.F40
blackbox_007, TU30.Y37, VV009
blackbox_008, TU30.Y39, VV010
blackbox_011, TU20.F31
blackbox_012, TU20.Y30, VV006
blackbox_013, TU20.N29, PL003
blackbox_014, HE002
blackbox_015, TU10.T15
blackbox_016, TU10.T20
blackbox_017, TU10.Y21, VV002
blackbox_018, TU10.F22
blackbox_019, TU10.Y6, VV001
blackbox_020, TU10.F17
blackbox_021, TU10.T23
blackbox_022, TU10.Y25
blackbox_023, TU10.Y24
blackbox_024, TU10.L10
blackbox_025, TU30.L35
blackbox_026, TU30.T33
blackbox_027, TU30.L34
blackbox_028, TU20.L26
blackbox_029, TU20.T27
blackbox_030, TU20.Q28
blackbox_031, TU30.L32
blackbox_032, TU30.N38
blackbox_033, TU30.N36
blackbox_034, VV007
blackbox_035, VV008
"""

# UC1 explicitly-missing items (for testing UC1 detection)
UC1_MISSING_AKZ = {"TU10.Y6", "VV005"}
UC1_SOFTWARE_CONTROL = {"TU10.U41", "TU20.U42"}

# ══════════════════════════════════════════════════════════════════════════
# STYLES
# ══════════════════════════════════════════════════════════════════════════
HF = Font(name="Calibri", size=11, bold=True)
HFILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
GFILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RFILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BD = Border(left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))

# ══════════════════════════════════════════════════════════════════════════
# SHEET SCHEMAS (from blank template)
# ══════════════════════════════════════════════════════════════════════════
SHEET_PART_LIBRARY = ["part_id", "part_ref_name", "part_type", "fcstd_link_name",
    "source_fcstd_file", "ifc_parts_library", "ifc_part_name", "instance_count", "ifc_coverage", "description"]
SHEET_ASSEMBLY_STEPS = ["assembly_no", "label", "type", "akz_tag", "vv_tag",
    "fcstd_link_name", "source_fcstd_file", "ifc_parts_library", "ifc_part_name", "ifc_global_id",
    "pos_x_mm", "pos_y_mm", "pos_z_mm",
    "rot_qw", "rot_qx", "rot_qy", "rot_qz",
    "rot_axis_x", "rot_axis_y", "rot_axis_z", "rot_angle_deg",
    "euler_roll_deg", "euler_pitch_deg", "euler_yaw_deg",
    "forward_x", "forward_y", "forward_z",
    "mapping_confidence", "ifc_coverage", "notes"]
SHEET_CONN_TOPOLOGY = ["connection_id", "from_assembly_no", "from_label", "from_port",
    "to_assembly_no", "to_label", "to_port", "connection_type",
    "from_pos_x", "from_pos_y", "from_pos_z",
    "to_pos_x", "to_pos_y", "to_pos_z", "euclidean_distance_mm"]
SHEET_CONN_ANGLES = ["connection_id", "from_label", "to_label",
    "from_forward_x", "from_forward_y", "from_forward_z",
    "to_forward_x", "to_forward_y", "to_forward_z",
    "connection_vector_x", "connection_vector_y", "connection_vector_z",
    "bend_angle_deg", "torsion_angle_deg", "from_to_distance_mm"]
SHEET_POSITION = ["fcstd_link_name", "label", "type", "source_fcstd_file", "ifc_parts_library",
    "pos_x_mm", "pos_y_mm", "pos_z_mm",
    "qw", "qx", "qy", "qz", "axis_x", "axis_y", "axis_z", "angle_deg",
    "euler_roll_deg", "euler_pitch_deg", "euler_yaw_deg"]
SHEET_TREE = ["tree_id", "parent_label", "child_label", "aggregation_type", "assembly_level", "notes"]

# ══════════════════════════════════════════════════════════════════════════
# PARSE
# ══════════════════════════════════════════════════════════════════════════
def parse_akz(raw):
    akz_m, vv_m = {}, {}
    for line in raw.strip().split("\n"):
        parts = [p.strip() for p in line.split(",") if p.strip()]
        if len(parts) < 2: continue
        label = parts[0]; akz = ""; vv = ""
        for p in parts[1:]:
            if p.startswith(("VV", "PL", "HE")): vv = p
            else: akz = p
        if akz: akz_m[label] = akz
        if vv: vv_m[label] = vv
    return akz_m, vv_m

def parse_fcstd(path):
    with zipfile.ZipFile(path, "r") as z:
        with z.open("Document.xml") as f:
            tree = ET.parse(f); root = tree.getroot()
    ntt = {}
    for obj in root.find("Objects").findall("Object"):
        ntt[obj.get("name", "")] = obj.get("type", "")
    comps = []
    for obj in root.find("ObjectData").findall("Object"):
        if ntt.get(obj.get("name", "")) != "App::Link": continue
        props = obj.find("Properties")
        if props is None: continue
        c = {"name": obj.get("name", ""), "label": "", "src": "", "internal": "",
             "px": 0.0, "py": 0.0, "pz": 0.0,
             "qw": 0.0, "qx": 0.0, "qy": 0.0, "qz": 1.0,
             "ax": 0.0, "ay": 0.0, "az": 1.0, "angle": 0.0}
        for prop in props.findall("Property"):
            pn = prop.get("name", "")
            if pn == "Label":
                se = prop.find("String")
                if se is not None: c["label"] = se.get("value", "")
            elif pn == "LinkedObject":
                xl = prop.find("XLink")
                if xl is not None:
                    c["src"] = xl.get("file", "")
                    c["internal"] = xl.get("name", "")
            elif pn == "LinkPlacement":
                pp = prop.find("PropertyPlacement")
                if pp is not None:
                    c["px"] = float(pp.get("Px", "0"))
                    c["py"] = float(pp.get("Py", "0"))
                    c["pz"] = float(pp.get("Pz", "0"))
                    c["qw"] = float(pp.get("Q0", "0"))
                    c["qx"] = float(pp.get("Q1", "0"))
                    c["qy"] = float(pp.get("Q2", "0"))
                    c["qz"] = float(pp.get("Q3", "1"))
                    c["ax"] = float(pp.get("Ox", "0"))
                    c["ay"] = float(pp.get("Oy", "0"))
                    c["az"] = float(pp.get("Oz", "1"))
                    c["angle"] = float(pp.get("A", "0"))
        if c["label"]: comps.append(c)
    return comps

def build_ifc_idx():
    idx = {}
    for d in [V1_DIR, V0_DIR]:
        for ip in sorted(d.glob("*.ifc")):
            if any(x in ip.name for x in ["FLAT_EXPORT", "from_step", "geometry"]): continue
            try:
                f = ifcopenshell.open(str(ip))
                for e in f.by_type("IfcBuildingElementProxy"):
                    if e.Name and e.Name not in idx:
                        idx[e.Name] = (ip.name, e.GlobalId)
            except: pass
    return idx

def build_xlink_map(comps):
    xm = {}
    for sf in set(c["src"] for c in comps):
        fpath = V1_DIR / sf
        if not fpath.exists() or sf == "Unnamed9.FCStd": continue
        try:
            with zipfile.ZipFile(fpath, "r") as z:
                with z.open("Document.xml") as f:
                    t = ET.parse(f); r = t.getroot()
            ntt = {}
            for obj in r.find("Objects").findall("Object"):
                ntt[obj.get("name", "")] = obj.get("type", "")
            for obj in r.find("ObjectData").findall("Object"):
                if ntt.get(obj.get("name", "")) != "App::Link": continue
                pr = obj.find("Properties")
                if pr is None: continue
                for prop in pr.findall("Property"):
                    if prop.get("name") == "LinkedObject":
                        xl = prop.find("XLink")
                        if xl is not None:
                            xm[f"{sf}::{obj.get('name','')}"] = xl.get("name", "")
        except: pass
    return xm

# ══════════════════════════════════════════════════════════════════════════
# MATH
# ══════════════════════════════════════════════════════════════════════════
def q2e(qw, qx, qy, qz):
    n = math.sqrt(qw*qw+qx*qx+qy*qy+qz*qz)
    if n < 1e-15: return (0,0,0)
    qw,qx,qy,qz = qw/n,qx/n,qy/n,qz/n
    r20=2*(qx*qz-qw*qy); r21=2*(qy*qz+qw*qx); r22=1-2*(qx*qx+qy*qy)
    r10=2*(qx*qy+qw*qz); r00=1-2*(qy*qy+qz*qz)
    p=-math.asin(max(-1,min(1,r20))); cp=math.cos(p)
    if abs(cp)>1e-10: return (math.degrees(math.atan2(r21/cp,r22/cp)), math.degrees(p), math.degrees(math.atan2(r10/cp,r00/cp)))
    return (0, math.degrees(p), math.degrees(math.atan2(-2*(qy*qz-qw*qx),1-2*(qx*qx+qz*qz))))

def qrot(qw,qx,qy,qz, vx,vy,vz):
    tw=-qx*vx-qy*vy-qz*vz; tx=qw*vx+qy*vz-qz*vy
    ty=qw*vy+qz*vx-qx*vz; tz=qw*vz+qx*vy-qy*vx
    return (tx*qw+tw*(-qx)+ty*(-qz)-tz*(-qy),
            ty*qw+tw*(-qy)+tz*(-qx)-tx*(-qz),
            tz*qw+tw*(-qz)+tx*(-qy)-ty*(-qx))

def classify(l):
    if l.startswith("tank"): return "tank"
    if l.startswith("blackbox"): return "blackbox"
    if l.startswith("elbow"): return "elbow"
    if l.startswith("pipeline"): return "pipeline"
    return "other"

def guess_connection_type(frm, to):
    ft, tt = frm["type"], to["type"]
    if "tank" in (ft, tt): return "flange"
    if "elbow" in (ft, tt) and "pipeline" in (ft, tt): return "weld_or_threaded"
    if "blackbox" in (ft, tt): return "flange_or_threaded"
    if ft == "elbow" and tt == "elbow": return "direct_fitting"
    if ft == "pipeline" and tt == "pipeline": return "weld_or_coupling"
    return "unknown"

# ══════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════
def write_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h); c.font = HF; c.fill = HFILL; c.border = BD
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            v = row.get(h, "")
            cell = ws.cell(row=ri, column=ci, value=v); cell.border = BD
            if h == "ifc_coverage" and v == "NO": cell.fill = RFILL
            if h == "akz_tag" and v: cell.fill = GFILL
    for ci in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(ci)].width = min(22, max(len(str(headers[ci-1])), 8))
    return ws

# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════
def main():
    akz_map, vv_map = parse_akz(AKZ_RAW)
    comps = parse_fcstd(FCSTD_PATH)
    ifc_idx = build_ifc_idx()
    xlink_map = build_xlink_map(comps)

    # ── Generate assembly_no: single stable sequence 1..302 ──
    # Strategy: use CSV as a sorting hint (preserve user's intended order),
    # then assign consecutive 1..302 numbers to the final sorted list.
    # This guarantees uniqueness and no gaps regardless of CSV completeness.
    csv_order = {}
    if CSV_PATH.exists():
        with open(CSV_PATH, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                csv_order[row["label"]] = int(row["assembly_no"])

    # Sort comps: CSV-labeled first (by their CSV order), then new comps (by FCStd position)
    comps_with_csv = [c for c in comps if c["label"] in csv_order]
    comps_without_csv = [c for c in comps if c["label"] not in csv_order]
    comps_with_csv.sort(key=lambda c: csv_order[c["label"]])
    # comps_without_csv already ordered by FCStd creation index
    sorted_comps = comps_with_csv + comps_without_csv

    # Assign final unique assembly_no 1..302
    label_to_ano = {c["label"]: i + 1 for i, c in enumerate(sorted_comps)}

    # ── Build steps ──
    steps = []
    for c in comps:
        label = c["label"]; ctype = classify(label)
        euler = q2e(c["qw"], c["qx"], c["qy"], c["qz"])
        fwd = qrot(c["qw"], c["qx"], c["qy"], c["qz"], 0, 0, 1)
        ano = label_to_ano.get(label, "")

        # IFC lookup
        ifc_file = ""; ifc_part = ""; ifc_gid = ""
        for try_name in (c["name"], c["internal"]):
            if try_name and try_name in ifc_idx:
                ifc_file, ifc_gid = ifc_idx[try_name]
                ifc_part = try_name; break
        if not ifc_file:
            key = f"{c['src']}::{c['name']}"
            resolved = xlink_map.get(key, "")
            if resolved and resolved in ifc_idx:
                ifc_file, ifc_gid = ifc_idx[resolved]
                ifc_part = resolved

        akz = akz_map.get(label, "")
        vv = vv_map.get(label, "")
        notes = ""
        if akz in UC1_SOFTWARE_CONTROL:
            notes = "software_control_no_ifc_required"
        elif akz in UC1_MISSING_AKZ:
            notes = "uc1_expected_missing"

        steps.append({
            "assembly_no": ano,
            "label": label, "type": ctype,
            "akz_tag": akz, "vv_tag": vv,
            "fcstd_link_name": c["name"],
            "source_fcstd_file": c["src"],
            "ifc_parts_library": ifc_file, "ifc_part_name": ifc_part, "ifc_global_id": ifc_gid,
            "pos_x_mm": round(c["px"], 4), "pos_y_mm": round(c["py"], 4), "pos_z_mm": round(c["pz"], 4),
            "rot_qw": round(c["qw"], 9), "rot_qx": round(c["qx"], 9),
            "rot_qy": round(c["qy"], 9), "rot_qz": round(c["qz"], 9),
            "rot_axis_x": round(c["ax"], 6), "rot_axis_y": round(c["ay"], 6), "rot_axis_z": round(c["az"], 6),
            "rot_angle_deg": round(math.degrees(c["angle"]), 4),
            "euler_roll_deg": round(euler[0], 4), "euler_pitch_deg": round(euler[1], 4), "euler_yaw_deg": round(euler[2], 4),
            "forward_x": round(fwd[0], 4), "forward_y": round(fwd[1], 4), "forward_z": round(fwd[2], 4),
            "mapping_confidence": "matched",
            "ifc_coverage": "YES" if ifc_file else "NO",
            "notes": notes,
        })

    steps.sort(key=lambda s: s["assembly_no"] if isinstance(s["assembly_no"], int) else 99999)

    # ── Build Part_Library ──
    part_lib = {}
    for s in steps:
        fn = s["fcstd_link_name"]
        if fn not in part_lib:
            part_lib[fn] = {"part_id": f"PART_{len(part_lib)+1:04d}", "part_ref_name": fn,
                "part_type": s["type"], "fcstd_link_name": fn,
                "source_fcstd_file": s["source_fcstd_file"],
                "ifc_parts_library": s["ifc_parts_library"], "ifc_part_name": s["ifc_part_name"],
                "instance_count": 0, "ifc_coverage": s["ifc_coverage"], "description": ""}
        part_lib[fn]["instance_count"] += 1

    # ── Build Connection_Topology & Connection_Angles ──
    valid = [s for s in steps if isinstance(s["assembly_no"], int)]
    connections = []
    angles = []
    for i in range(len(valid)-1):
        frm, to = valid[i], valid[i+1]
        cid = i + 1
        dist = round(math.sqrt(
            (frm["pos_x_mm"]-to["pos_x_mm"])**2 +
            (frm["pos_y_mm"]-to["pos_y_mm"])**2 +
            (frm["pos_z_mm"]-to["pos_z_mm"])**2), 2) if all(v is not None for v in [frm["pos_x_mm"], to["pos_x_mm"]]) else None

        connections.append({
            "connection_id": cid,
            "from_assembly_no": frm["assembly_no"], "from_label": frm["label"], "from_port": "out",
            "to_assembly_no": to["assembly_no"], "to_label": to["label"], "to_port": "in",
            "connection_type": guess_connection_type(frm, to),
            "from_pos_x": frm["pos_x_mm"], "from_pos_y": frm["pos_y_mm"], "from_pos_z": frm["pos_z_mm"],
            "to_pos_x": to["pos_x_mm"], "to_pos_y": to["pos_y_mm"], "to_pos_z": to["pos_z_mm"],
            "euclidean_distance_mm": dist,
        })

        fwd_f = (frm["forward_x"], frm["forward_y"], frm["forward_z"])
        fwd_t = (to["forward_x"], to["forward_y"], to["forward_z"])
        cv = (0.0, 0.0, 0.0); n_cv = 0.0
        if dist and dist > 0:
            cv = ((to["pos_x_mm"]-frm["pos_x_mm"])/dist, (to["pos_y_mm"]-frm["pos_y_mm"])/dist, (to["pos_z_mm"]-frm["pos_z_mm"])/dist)
            n_cv = dist

        bend = 0.0
        if all(v is not None for v in fwd_f + fwd_t):
            dot = fwd_f[0]*fwd_t[0] + fwd_f[1]*fwd_t[1] + fwd_f[2]*fwd_t[2]
            bend = math.degrees(math.acos(max(-1, min(1, dot))))
        torsion = 0.0
        if n_cv > 1e-12 and fwd_f[0] is not None:
            dot_cv = abs(fwd_f[0]*cv[0] + fwd_f[1]*cv[1] + fwd_f[2]*cv[2])
            torsion = math.degrees(math.acos(max(-1, min(1, dot_cv))))

        angles.append({
            "connection_id": cid, "from_label": frm["label"], "to_label": to["label"],
            "from_forward_x": round(fwd_f[0], 6) if fwd_f[0] is not None else None,
            "from_forward_y": round(fwd_f[1], 6) if fwd_f[1] is not None else None,
            "from_forward_z": round(fwd_f[2], 6) if fwd_f[2] is not None else None,
            "to_forward_x": round(fwd_t[0], 6) if fwd_t[0] is not None else None,
            "to_forward_y": round(fwd_t[1], 6) if fwd_t[1] is not None else None,
            "to_forward_z": round(fwd_t[2], 6) if fwd_t[2] is not None else None,
            "connection_vector_x": round(cv[0], 6), "connection_vector_y": round(cv[1], 6), "connection_vector_z": round(cv[2], 6),
            "bend_angle_deg": round(bend, 2), "torsion_angle_deg": round(torsion, 2),
            "from_to_distance_mm": round(n_cv, 2),
        })

    # ── Build Position_Data ──
    positions = []
    for s in steps:
        if s["ifc_coverage"] == "YES":
            positions.append({
                "fcstd_link_name": s["fcstd_link_name"], "label": s["label"], "type": s["type"],
                "source_fcstd_file": s["source_fcstd_file"], "ifc_parts_library": s["ifc_parts_library"],
                "pos_x_mm": s["pos_x_mm"], "pos_y_mm": s["pos_y_mm"], "pos_z_mm": s["pos_z_mm"],
                "qw": s["rot_qw"], "qx": s["rot_qx"], "qy": s["rot_qy"], "qz": s["rot_qz"],
                "axis_x": s["rot_axis_x"], "axis_y": s["rot_axis_y"], "axis_z": s["rot_axis_z"],
                "angle_deg": s["rot_angle_deg"],
                "euler_roll_deg": s["euler_roll_deg"], "euler_pitch_deg": s["euler_pitch_deg"], "euler_yaw_deg": s["euler_yaw_deg"],
            })

    # ── Build Assembly_Tree ──
    tree = []
    tid = 0; current_tank = None
    for s in steps:
        if s["type"] == "tank":
            tid += 1; current_tank = s["label"]
            tree.append({"tree_id": tid, "parent_label": None, "child_label": s["label"],
                "aggregation_type": "root_assembly", "assembly_level": "assembly", "notes": "tank root"})
        elif current_tank and s["type"] in ("pipeline", "elbow", "blackbox"):
            tid += 1
            tree.append({"tree_id": tid, "parent_label": current_tank, "child_label": s["label"],
                "aggregation_type": "connected_to", "assembly_level": "part", "notes": f"part of {current_tank} run"})

    # ── AKZ_Summary ──
    akz_summary = [s for s in steps if s["akz_tag"] or s["vv_tag"]]

    # ── IFC_Gaps ──
    ifc_gaps = [s for s in steps if s["ifc_coverage"] == "NO"]

    # ══════════════════════════════════════════════════════════════════════
    # WRITE
    # ══════════════════════════════════════════════════════════════════════
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    write_sheet(wb, "Part_Library", SHEET_PART_LIBRARY, list(part_lib.values()))
    write_sheet(wb, "Assembly_Steps", SHEET_ASSEMBLY_STEPS, steps)
    write_sheet(wb, "Connection_Topology", SHEET_CONN_TOPOLOGY, connections)
    write_sheet(wb, "Connection_Angles", SHEET_CONN_ANGLES, angles)
    write_sheet(wb, "Position_Data", SHEET_POSITION, positions)
    write_sheet(wb, "Assembly_Tree", SHEET_TREE, tree)
    write_sheet(wb, "AKZ_Summary", ["label","type","akz_tag","vv_tag","pos_x_mm","pos_y_mm","pos_z_mm","ifc_parts_library","ifc_part_name","ifc_global_id"],
                [{k: s[k] for k in ["label","type","akz_tag","vv_tag","pos_x_mm","pos_y_mm","pos_z_mm","ifc_parts_library","ifc_part_name","ifc_global_id"]} for s in akz_summary])
    write_sheet(wb, "IFC_Gaps", ["label","type","fcstd_link_name","source_fcstd_file","pos_x_mm","pos_y_mm","pos_z_mm"],
                [{k: s[k] for k in ["label","type","fcstd_link_name","source_fcstd_file","pos_x_mm","pos_y_mm","pos_z_mm"]} for s in ifc_gaps])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT))

    # ── Summary ──
    empty_ano = sum(1 for s in steps if not isinstance(s["assembly_no"], int) or s["assembly_no"] == "")
    print(f"Saved: {OUT}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Assembly_Steps: {len(steps)} rows ({len(steps)-empty_ano} with assembly_no, {empty_ano} empty)")
    print(f"Part_Library: {len(part_lib)} unique parts")
    print(f"Connection_Topology: {len(connections)} connections")
    print(f"Connection_Angles: {len(angles)} angle entries")
    print(f"Position_Data: {len(positions)} positions")
    print(f"Assembly_Tree: {len(tree)} tree nodes")
    print(f"AKZ_Summary: {len(akz_summary)} entries")
    print(f"IFC_Gaps: {len(ifc_gaps)} gaps")
    print(f"IFC coverage: {sum(1 for s in steps if s['ifc_coverage']=='YES')}/{len(steps)}")
    tanks = [s for s in steps if s["type"]=="tank"]
    bboxes = [s for s in steps if s["type"]=="blackbox"]
    print(f"Types: {len(tanks)} tanks, {len(bboxes)} blackboxes, {sum(1 for s in steps if s['type']=='elbow')} elbows, {sum(1 for s in steps if s['type']=='pipeline')} pipelines")

if __name__ == "__main__":
    main()
