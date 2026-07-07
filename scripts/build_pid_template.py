"""Build PID_template.xlsx with standardized column headers.

Run once:
    python scripts/build_pid_template.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
TARGET_DIR = REPO_ROOT / "data" / "templates"
TARGET_FILE = TARGET_DIR / "PID_template.xlsx"

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
LEGEND_FONT = Font(italic=True, size=10, color="555555")
LEGEND_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

SHEETS: dict[str, list[tuple[str, str]]] = {
    "Document_ID": [
        ("Index", "1-based row index within this sheet"),
        ("Document_ID", "Unique document identifier"),
        ("Document", "Source file name"),
        ("SemanticID", "Semantic identifier for ontology linking"),
    ],
    "Document_Data": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Project_Name", "Project name from R&I title block"),
        ("Project_ID", "Project identifier"),
        ("Plant", "Plant / facility name"),
        ("Drawing_Number", "Drawing or document number"),
        ("Revision", "Document revision"),
        ("Date", "Document date (YYYY-MM-DD)"),
        ("Norm", "Applicable standard (e.g. DIN EN 62424)"),
        ("Software", "Authoring software if known"),
        ("SemanticID", "Semantic identifier"),
    ],
    "Equipment": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Equipment_ID", "Unique equipment instance identifier"),
        ("AKZ", "Original AKZ string as it appears in source"),
        ("AKZ_Canonical", "Normalised canonical AKZ (no separators, uppercase)"),
        ("Equipment_Class", "E.g. Vessel, Pump, HeatExchanger, Valve"),
        ("Equipment_Subclass", "E.g. CentrifugalPump, GateValve"),
        ("Description", "Free-text description from source"),
        ("TagName", "DEXPI TagName if available"),
        ("SemanticID", "Semantic identifier"),
        ("Source_Vendor", "Detected vendor or empty if unknown"),
        ("Source_File", "Source document path"),
        ("Confidence", "Extraction confidence 0.0–1.0"),
        ("LLM_Reasoning", "LLM reasoning if LLM was used for extraction"),
    ],
    "Equipment_Data": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Equipment_ID", "FK → Equipment"),
        ("Attribute_Name", "Standard attribute name"),
        ("Attribute_Value", "Attribute value"),
        ("Attribute_Unit", "Unit of measurement"),
        ("Attribute_Source", "Source text span"),
        ("SemanticID_Attribute", "Semantic identifier for the attribute"),
        ("Confidence", "Extraction confidence 0.0–1.0"),
    ],
    "Piping": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Pipe_ID", "Unique pipe identifier"),
        ("AKZ", "Original AKZ string"),
        ("AKZ_Canonical", "Normalised canonical AKZ"),
        ("From_Equipment_ID", "FK → source Equipment"),
        ("To_Equipment_ID", "FK → target Equipment"),
        ("Pipe_Class", "Pipe class / specification"),
        ("Nominal_Diameter", "Nominal diameter (DN)"),
        ("Nominal_Pressure", "Nominal pressure (PN)"),
        ("Medium", "Flowing medium"),
        ("SemanticID", "Semantic identifier"),
        ("Confidence", "Extraction confidence 0.0–1.0"),
        # IFC match details
        ("IfcClass", "IFC class name"),
        ("GlobalId", "IFC GlobalId"),
        ("Tag", "IFC tag name"),
        ("HasPorts", "Whether IFC element has ports"),
        ("ConnectedTo", "Connected-to IFC reference"),
        ("ConnectedFrom", "Connected-from IFC reference"),
        ("HasControlElements", "Whether control elements are present"),
        ("PredefinedType", "IFC predefined type"),
        ("Size", "Nominal size from IFC"),
        ("ValveMechanism", "Valve mechanism type from IFC"),
        ("FlowCoefficient", "Flow coefficient from IFC"),
        ("FailPosition", "Fail position from IFC"),
        ("ManualOverride", "Manual override capability"),
        ("ActuatorApplication", "Actuator application type"),
        ("PresenceStatus", "IFC match presence status"),
        ("FlangeComplete", "Whether flange data is complete"),
    ],
    "Instrumentation": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Instrument_ID", "Unique instrument instance identifier"),
        ("AKZ", "Original AKZ string"),
        ("AKZ_Canonical", "Normalised canonical AKZ"),
        ("ProcessInstrumentationFunctionCategory", "IEC 81346 function category, e.g. TI"),
        ("ProcessInstrumentationFunctionModifier", "Function modifier letter"),
        ("ProcessInstrumentationFunctionNumber", "Sequential number, e.g. T41"),
        ("TagName", "Full DEXPI TagName"),
        ("Function", "Measured or controlled variable description"),
        ("Loop_ID", "Control loop identifier"),
        ("Connected_Equipment_ID", "FK → connected Equipment"),
        ("Connected_Pipe_ID", "FK → connected Piping"),
        ("SemanticID", "Semantic identifier"),
        ("Source_Vendor", "Detected vendor"),
        ("Source_File", "Source document path"),
        ("Confidence", "Extraction confidence 0.0–1.0"),
        ("LLM_Reasoning", "LLM reasoning if used"),
        # SM_Nameplate
        ("ManufacturerName", "Manufacturer company name"),
        ("ManufacturerProductDesignation", "Product designation from manufacturer"),
        ("OrderCode", "Order code / part number"),
        # SM_ProcessFunction
        ("ProcessInstrumentationFunctions", "Full function string"),
        ("HasInstrumentationLoopFunctionNumber", "Loop function number if present"),
        # SM_ActuationAndPiping
        ("ActuatingFunctionNumber", "Actuating function number from DEXPI"),
        ("ActuatingLocation", "Actuating location / anchor reference"),
        ("ActuatingSystemNumber", "Actuating system number"),
        ("OperatedValveRef", "Reference to operated valve"),
        ("PipingAnchorId", "Piping anchor identifier from DEXPI"),
        ("PipingComponentName", "Piping component name"),
        ("FlowDirection", "Flow direction"),
        ("LineNumber", "Line number"),
        ("NominalDiameterValue", "Nominal diameter numeric value"),
        ("NominalDiameterRepr", "Nominal diameter representation string"),
        ("NominalDiameterStandard", "Nominal diameter standard"),
        ("NominalDiameterType", "Nominal diameter type"),
        ("FromEquipmentId", "Source equipment identifier from DEXPI"),
        ("ToEquipmentId", "Target equipment identifier from DEXPI"),
        # SM_FunctionalSafety
        ("SafetyRelevanceClass", "Safety relevance classification"),
        # SM_TechnicalData
        ("DeviceInformation", "Device description / information"),
        ("LabelText", "DEXPI LabelText"),
        ("FunctionCode", "Function code from DEXPI"),
        # SM_SourceDocumentation
        ("SourceDocId", "Source document identifier"),
        ("SourceLocator", "Source document locator / page reference"),
        # Completion
        ("ContextSummary", "Cross-document context summary"),
        ("XSDStatus", "XSD validation status"),
        ("RecommendedAction", "Recommended action for missing data"),
        ("ProposalStatus", "Proposal status"),
        ("MissingTargets", "Missing target documents"),
        ("NeedsReview", "Whether manual review is needed"),
        ("DecisionConfidence", "Decision confidence score"),
        ("EvidenceBundleId", "Evidence bundle identifier"),
        ("UncertaintyReason", "Reason for uncertainty"),
        ("LLMVerificationStatus", "LLM verification status"),
        ("RuleSupport", "Supporting rules"),
        ("ReviewFeedbackStatus", "Review feedback status"),
        ("DecisionTraceJson", "Decision trace as JSON"),
    ],
    "Instrumentation_Data": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Instrument_ID", "FK → Instrumentation"),
        ("Attribute_Name", "Standard attribute name"),
        ("Attribute_Value", "Attribute value"),
        ("Attribute_Unit", "Unit of measurement"),
        ("Attribute_Source", "Source text span"),
        ("SemanticID_Attribute", "Semantic identifier"),
        ("Confidence", "Extraction confidence 0.0–1.0"),
    ],
    "Connection": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Connection_ID", "Unique connection identifier"),
        ("From_ID", "FK → source node (Equipment / Instrument / Pipe)"),
        ("From_Type", "Type of source node"),
        ("To_ID", "FK → target node"),
        ("To_Type", "Type of target node"),
        ("Connection_Class", "E.g. PipingNetworkSegment, SignalLine"),
        ("SemanticID", "Semantic identifier"),
    ],
    "AKZ_Canonical_Map": [
        ("Index", "1-based row index"),
        ("Canonical_AKZ", "Normalised canonical AKZ"),
        ("Document_ID", "FK → Document_ID"),
        ("Original_AKZ", "Original AKZ string in this document"),
        ("Source_Sheet", "Sheet name where found"),
        ("Source_Row", "Row number where found"),
        ("Match_Confidence", "Confidence of cross-document match"),
        ("Match_Method", "exact | fuzzy_edit_distance | llm_judge"),
        ("LLM_Reasoning", "LLM reasoning if LLM was used"),
    ],
    "Inconsistency_Report": [
        ("Index", "1-based row index"),
        ("Inc_ID", "Unique inconsistency identifier"),
        ("Canonical_AKZ", "Affected canonical AKZ"),
        ("Rule", "UC1 | UC2 | cardinality | value_conflict"),
        ("Severity", "critical | warning | info"),
        ("Missing_In", "Comma-separated list of documents missing this AKZ"),
        ("Present_In", "Comma-separated list of documents containing this AKZ"),
        ("LLM_Verdict", "LLM consistency verdict if used"),
        ("LLM_Reasoning", "LLM reasoning"),
        ("Confidence", "Detection confidence 0.0–1.0"),
        ("Detected_At", "ISO timestamp of detection"),
        ("Reviewed_By", "User who reviewed"),
        ("Review_Status", "auto | needs_review | reviewed_ok | reviewed_overruled"),
    ],
}


def build_template() -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, columns in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        # Row 1: English column headers
        for col_idx, (header, _legend) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Row 2: German legend / description
        for col_idx, (_header, legend) in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col_idx, value=legend)
            cell.font = LEGEND_FONT
            cell.fill = LEGEND_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Column widths
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        # Freeze panes below headers
        ws.freeze_panes = f"{get_column_letter(1)}3"

    TARGET_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(TARGET_FILE)
    print(f"wrote {TARGET_FILE.relative_to(REPO_ROOT)} ({len(SHEETS)} sheets)")


if __name__ == "__main__":
    build_template()
