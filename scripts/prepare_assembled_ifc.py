#!/usr/bin/env python3
"""Prepare an assembled IFC for use with the simplified pipeline.

1. Reads Assembly_3D_template_filled.xlsx to get label → internal_name mapping
2. Renames IfcBuildingElementProxy entities in the assembled IFC from
   internal FreeCAD names (Body, Body001...) to semantic labels (tank_001, ...)
3. Optionally extracts individual part IFCs

Usage:
  python scripts/prepare_assembled_ifc.py \\
    --ifc Documents/Piping\\ Diagram/assembly.ifc \\
    --template Documents/Piping\\ Diagram/Assembly_3D_template_filled.xlsx \\
    [--extract-parts] [--output Documents/Piping\\ Diagram/assembly_prepared.ifc]
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import ifcopenshell
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_IFC = PROJECT_ROOT / "Documents" / "Piping Diagram" / "assembly.ifc"
DEFAULT_TEMPLATE = PROJECT_ROOT / "Documents" / "Piping Diagram" / "Assembly_3D_template_filled.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "Documents" / "Piping Diagram" / "assembly_prepared.ifc"
PARTS_DIR = PROJECT_ROOT / "Documents" / "Piping Diagram" / "individual_parts"


def _get_entity_position(entity) -> tuple[float, float, float] | None:
    """Extract (x, y, z) from an IFC entity's ObjectPlacement, or None."""
    placement = entity.ObjectPlacement
    if placement is None:
        return None
    rel = placement.RelativePlacement
    if rel is None or rel.Location is None:
        return None
    c = rel.Location.Coordinates
    return (c[0], c[1], c[2])


def _dist(a: tuple[float, float, float],
          b: tuple[float, float, float]) -> float:
    return ((a[0]-b[0])**2 + (a[1]-b[1])**2 + (a[2]-b[2])**2) ** 0.5


def build_rename_map(
    template_path: Path, ifc_path: Path | None = None, tolerance_mm: float = 5.0
) -> dict[str, str]:
    """Build GlobalId → semantic_label mapping from the template + IFC.

    When *ifc_path* is provided and entities carry ObjectPlacement, matches
    by position proximity (handles name collisions like multiple ``Body``
    instances).  Otherwise falls back to a simple ifc_part_name → label map.
    """
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb["Assembly_Steps"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    ci = {h: i + 1 for i, h in enumerate(headers)}

    # Collect template entries
    tmpl: list[dict] = []
    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=ci.get("label", 1)).value or ""
        pname = ws.cell(row=r, column=ci.get("ifc_part_name", 1)).value or ""
        gid = ws.cell(row=r, column=ci.get("ifc_global_id", 1)).value or ""
        try:
            px = float(ws.cell(row=r, column=ci.get("pos_x_mm", 1)).value or 0)
            py = float(ws.cell(row=r, column=ci.get("pos_y_mm", 1)).value or 0)
            pz = float(ws.cell(row=r, column=ci.get("pos_z_mm", 1)).value or 0)
        except (ValueError, TypeError):
            px = py = pz = 0.0
        if label and pname:
            tmpl.append({"label": label, "pname": pname, "gid": gid,
                         "px": px, "py": py, "pz": pz})
    wb.close()

    rename: dict[str, str] = {}

    if ifc_path and ifc_path.exists():
        f = ifcopenshell.open(str(ifc_path))
        entity_types = ("IfcBuildingElementProxy", "IfcPipeSegment",
                        "IfcPipeFitting", "IfcFlowSegment", "IfcFlowFitting")
        entities: list = []
        for etype in entity_types:
            entities.extend(f.by_type(etype))
        # Check if entities have placements
        with_pos = [(e, _get_entity_position(e)) for e in entities
                    if _get_entity_position(e) is not None]

        if with_pos:
            # Position-based matching: for each template entry, find the
            # closest IFC entity by position.  Unique per match.
            used: set[int] = set()
            for t in tmpl:
                best_idx, best_dist = -1, float("inf")
                for i, (ent, epos) in enumerate(with_pos):
                    if i in used:
                        continue
                    d = _dist((t["px"], t["py"], t["pz"]), epos)
                    if d < best_dist:
                        best_dist, best_idx = d, i
                if best_idx >= 0 and best_dist <= tolerance_mm:
                    used.add(best_idx)
                    ent = with_pos[best_idx][0]
                    rename[ent.GlobalId] = t["label"]
            return rename

    # Fallback: simple part_name → label (no position data available)
    seen: set[str] = set()
    for t in tmpl:
        if t["pname"] not in seen:
            rename[t["pname"]] = t["label"]
            seen.add(t["pname"])
    return rename


def rename_ifc_entities(
    ifc_path: Path, rename_map: dict[str, str], output_path: Path
) -> int:
    """Rename entities.  Keys are either GlobalIds or old entity Names.

    Also handles ``_flat`` suffix stripping (FreeCAD FLAT_EXPORT naming)."""
    f = ifcopenshell.open(str(ifc_path))
    count = 0
    entity_types = ("IfcBuildingElementProxy", "IfcPipeSegment",
                    "IfcPipeFitting", "IfcFlowSegment", "IfcFlowFitting")
    for etype in entity_types:
        for entity in f.by_type(etype):
            # Try exact match first (GlobalId or Name), then strip _flat suffix
            new_name = rename_map.get(entity.GlobalId) or rename_map.get(entity.Name or "")
            if not new_name:
                base = (entity.Name or "").replace("_flat", "")
                new_name = rename_map.get(base) or rename_map.get(entity.GlobalId)
            if new_name:
                entity.Name = new_name
                count += 1
    f.write(str(output_path))
    return count


def collect_tank_blackbox_names(template_path: Path) -> set[str]:
    """Get the set of label names that are tanks or blackboxes."""
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb["Assembly_Steps"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    ci = {h: i + 1 for i, h in enumerate(headers)}

    names: set[str] = set()
    for r in range(2, ws.max_row + 1):
        ctype = ws.cell(row=r, column=ci.get("type", 1)).value
        if ctype not in ("tank", "blackbox"):
            continue
        label = ws.cell(row=r, column=ci.get("label", 1)).value or ""
        if label:
            names.add(label)
    wb.close()
    return names


def extract_individual_parts(ifc_path: Path, labels: set[str],
                              output_dir: Path) -> tuple[int, int]:
    """Extract tank/blackbox entities into individual IFC files."""
    output_dir.mkdir(parents=True, exist_ok=True)
    f = ifcopenshell.open(str(ifc_path))
    schema = f.schema

    success, total = 0, 0
    index_rows: list[str] = ["label,type,ifc_part_name,ifc_source_file"]
    seen: set[str] = set()

    for entity in f.by_type("IfcBuildingElementProxy"):
        name = entity.Name or ""
        if name not in labels:
            continue
        total += 1

        out_path = output_dir / f"{name}.ifc"
        # Skip duplicates (keep first occurrence)
        if name in seen:
            continue
        seen.add(name)

        try:
            new_f = ifcopenshell.file(schema=schema)
            for proj in f.by_type("IfcProject"):
                new_f.add(proj)
                break
            new_f.add(entity)
            new_f.write(str(out_path))
            success += 1
            ctype = "tank" if name.startswith("tank") else "blackbox"
            index_rows.append(f"{name},{ctype},{name},{ifc_path.name}")
        except Exception as exc:
            print(f"  FAIL {name}: {exc}", file=sys.stderr)

    # Write index
    (output_dir / "_index.csv").write_text("\n".join(index_rows) + "\n")
    return success, total


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--ifc", type=Path, default=DEFAULT_IFC,
                        help="Path to the assembled IFC file")
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE,
                        help="Path to Assembly_3D_template_filled.xlsx")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
                        help="Output path for the renamed IFC")
    parser.add_argument("--extract-parts", action="store_true",
                        help="Also extract individual tank/blackbox IFC files")
    parser.add_argument("--parts-dir", type=Path, default=PARTS_DIR,
                        help="Output directory for individual part IFCs")
    args = parser.parse_args()

    if not args.ifc.exists():
        print(f"ERROR: IFC file not found: {args.ifc}", file=sys.stderr)
        print("Export the assembled IFC from FreeCAD first, then run this script.",
              file=sys.stderr)
        return 1

    if not args.template.exists():
        print(f"ERROR: Template not found: {args.template}", file=sys.stderr)
        print("Run the assembly pipeline first to generate the template.",
              file=sys.stderr)
        return 1

    # Step 1: Build rename map (position-aware when IFC has placements)
    rename_map = build_rename_map(args.template, ifc_path=args.ifc)
    print(f"Built rename map: {len(rename_map)} entries")

    # Step 2: Rename entities
    renamed = rename_ifc_entities(args.ifc, rename_map, args.output)
    print(f"Renamed {renamed} entities → {args.output}")

    # Step 3: Optionally extract individual parts
    if args.extract_parts:
        labels = collect_tank_blackbox_names(args.template)
        success, total = extract_individual_parts(args.output, labels, args.parts_dir)
        print(f"Extracted {success}/{total} individual parts → {args.parts_dir}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
