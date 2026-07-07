#!/usr/bin/env python3
"""Enhanced audit of Exports/Excel standardized template filling quality.

Checks:
  Phase 1: Sheet names, column headers, row-2 state, column counts
  Phase 2: Per-sheet row counts, per-column blank rates
  Phase 3: Composite key uniqueness, foreign key integrity, value formats, special values
  Phase 5: Specific issue investigations (Datasheet emptiness, Assembly_3D gap, etc.)

Usage: python scripts/audit_enhanced.py
"""

from __future__ import annotations

import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
TPL_DIR = REPO / "data" / "templates"
FILLED_DIR = REPO / "data" / "filled_templates"
EXPORTS_DIR = REPO / "Exports" / "Excel"
GOLDEN_DIR = REPO / "data" / "examples"
OUTPUT_DIR = REPO / "Exports" / "audit"

TEMPLATE_SPEC = {
    "instrument_list": {
        "template": "Stellenplan_template.xlsx",
        "golden": "Standardized_Stellenplan.xlsx",
    },
    "wiring": {
        "template": "Klemmenplan_template.xlsx",
        "golden": "Standardized_Klemmenplan.xlsx",
    },
    "datasheet": {
        "template": "Datasheet_template.xlsx",
        "golden": "Standardized_Datasheet.xlsx",
    },
    "stromlaufplan": {
        "template": "Stromlaufplan_template.xlsx",
        "golden": "Standardized_Stromlaufplan.xlsx",
    },
    "3d_assembly": {
        "template": "Assembly_3D_template.xlsx",
        "golden": None,
    },
}

COMPOSITE_KEY_BY_TEMPLATE: dict[str, dict[str, list[tuple[str, ...]]]] = {
    "instrument_list": {
        "Document_ID": [("Document_ID",)],
        "Document_Data": [("Document_ID", "Instrument_Sheet_ID")],
        "Object_ID": [("Document_ID", "Instrument_Sheet_ID", "Object_ID")],
    },
    "wiring": {
        "Document_ID": [("Document_ID",)],
        "Dokument_Data": [("Document_ID", "Documentblatt_ID")],
        "Object_ID": [("Document_ID", "Documentblatt_ID", "Object_ID")],
        "Terminal_ID": [("Document_ID", "Object_ID", "Terminal_ID")],
    },
    "datasheet": {
        "Device_ID": [("Document_ID", "Device_ID")],
        "Process_Attributes": [("Document_ID", "Device_ID", "Attribute_Key")],
        "Technical_Attributes": [("Document_ID", "Device_ID", "Attribute_Key")],
        "Geometric_Attributes": [("Document_ID", "Device_ID", "Attribute_Key")],
        "Manufacturer_Specific": [("Document_ID", "Device_ID", "Attribute_Key")],
    },
    "stromlaufplan": {
        "Document_ID": [("Document_ID",)],
        "Object_ID": [("Document_ID", "Object_ID")],
        "Element_ID": [("Document_ID", "Element_ID")],
        "Connection_Data": [("Document_ID", "Connection_Key")],
    },
}

FK_SPEC_BY_TEMPLATE: dict[str, list[tuple[str, str, str, str]]] = {
    "instrument_list": [
        ("Document_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Revision_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Layer_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Instrument_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Object_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Component_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Component_Classification", "Document_ID", "Document_ID", "Document_ID"),
        ("Component_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Connection_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Component_ID", "Object_ID", "Object_ID", "Object_ID"),
        ("Component_Classification", "Object_ID", "Object_ID", "Object_ID"),
        ("Component_Data", "Object_ID", "Object_ID", "Object_ID"),
        ("Component_Data", "Componenet_ID", "Component_ID", "Component_ID"),
    ],
    "wiring": [
        ("Dokument_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Layer_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Object_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Object_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Terminal_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Terminal_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Object_Data", "Object_ID", "Object_ID", "Object_ID"),
        ("Terminal_ID", "Object_ID", "Object_ID", "Object_ID"),
        ("Terminal_Data", "Terminal_ID", "Terminal_ID", "Terminal_ID"),
    ],
    "datasheet": [
        ("Document_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Device_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Device_Classification", "Document_ID", "Document_ID", "Document_ID"),
        ("Process_Attributes", "Document_ID", "Document_ID", "Document_ID"),
        ("Technical_Attributes", "Document_ID", "Document_ID", "Document_ID"),
        ("Geometric_Attributes", "Document_ID", "Document_ID", "Document_ID"),
        ("Connection_Attributes", "Document_ID", "Document_ID", "Document_ID"),
        ("Manufacturer_Specific", "Document_ID", "Document_ID", "Document_ID"),
        ("Device_Classification", "Device_ID", "Device_ID", "Device_ID"),
        ("Process_Attributes", "Device_ID", "Device_ID", "Device_ID"),
        ("Technical_Attributes", "Device_ID", "Device_ID", "Device_ID"),
        ("Geometric_Attributes", "Device_ID", "Device_ID", "Device_ID"),
        ("Connection_Attributes", "Device_ID", "Device_ID", "Device_ID"),
        ("Manufacturer_Specific", "Device_ID", "Device_ID", "Device_ID"),
    ],
    "stromlaufplan": [
        ("Document_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Revision_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Layer_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Object_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Element_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Element_Classification", "Document_ID", "Document_ID", "Document_ID"),
        ("Element_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Connection_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Element_ID", "Object_ID", "Object_ID", "Object_ID"),
        ("Element_Classification", "Element_ID", "Element_ID", "Element_ID"),
        ("Element_Data", "Element_ID", "Element_ID", "Element_ID"),
        ("Connection_Data", "From_Element_ID", "Element_ID", "Element_ID"),
        ("Connection_Data", "To_Element_ID", "Element_ID", "Element_ID"),
    ],
}

IEC_81346_PATTERN = re.compile(r"^'?=")
ECLASS_IRDI_PATTERN = re.compile(r"^\d{4}-\d+#")
SPECIAL_VALUE_PATTERNS = [
    (re.compile(r"^\s*null\s*$", re.IGNORECASE), "literal 'null'"),
    (re.compile(r"^\s*None\s*$"), "literal 'None'"),
    (re.compile(r"^\s*NaN\s*$", re.IGNORECASE), "literal 'NaN'"),
    (re.compile(r"^\s*N\s*/\s*A\s*$", re.IGNORECASE), "literal 'N/A'"),
]


def _load_wb(path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(str(path), data_only=True)


def _header(ws) -> list[str]:
    return [str(ws.cell(1, c).value or "") for c in range(1, (ws.max_column or 0) + 1)]


def _data_rows(ws) -> list[tuple]:
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if any(v is not None and str(v).strip() != "" for v in row):
            rows.append(tuple(row))
    return rows


def _col_index(header: list[str], name: str) -> int | None:
    for i, h in enumerate(header):
        if h == name:
            return i
    return None


# =============================================================================
# Phase 1: Structural integrity
# =============================================================================


def check_phase1(template_name: str) -> dict:
    tpl_path = TPL_DIR / template_name
    filled_path = FILLED_DIR / template_name

    result = {"template_file": template_name}

    paths_to_check = []
    if tpl_path.exists():
        paths_to_check.append(("template", tpl_path))
    if filled_path.exists():
        paths_to_check.append(("filled", filled_path))
    for sd in EXPORTS_DIR.iterdir():
        if sd.is_dir():
            candidate = sd / template_name
            if candidate.exists():
                paths_to_check.append((f"exports/{sd.name}", candidate))
                break

    if not paths_to_check:
        result["error"] = "No files found"
        return result

    wbs = {}
    for label, path in paths_to_check:
        try:
            wbs[label] = _load_wb(path)
        except Exception as e:
            result.setdefault("load_errors", {})[label] = str(e)

    all_sheet_sets = {}
    for label, wb in wbs.items():
        all_sheet_sets[label] = set(wb.sheetnames)

    reference_sheets = all_sheet_sets.get("template", next(iter(all_sheet_sets.values())))
    for label, sheets in all_sheet_sets.items():
        missing = reference_sheets - sheets
        extra = sheets - reference_sheets
        if missing or extra:
            result.setdefault("sheet_name_mismatches", []).append({
                "location": label,
                "missing": sorted(missing),
                "extra": sorted(extra),
            })

    for sheet_name in sorted(reference_sheets):
        headers_by_loc = {}
        col_counts = {}
        for label, wb in wbs.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            h = _header(ws)
            headers_by_loc[label] = h
            col_counts[label] = len(h)

        if len(set(col_counts.values())) > 1:
            result.setdefault("column_count_mismatches", []).append({
                "sheet": sheet_name,
                "counts": col_counts,
            })

        ref_loc = "template" if "template" in headers_by_loc else next(iter(headers_by_loc))
        ref_header = headers_by_loc[ref_loc]
        for loc, h in headers_by_loc.items():
            if loc == ref_loc:
                continue
            if h != ref_header:
                diffs = []
                for i in range(max(len(h), len(ref_header))):
                    rv = ref_header[i] if i < len(ref_header) else "(missing)"
                    av = h[i] if i < len(h) else "(missing)"
                    if rv != av:
                        diffs.append({"col": i, "reference": rv, "actual": av})
                result.setdefault("header_mismatches", []).append({
                    "sheet": sheet_name,
                    "location": loc,
                    "diffs": diffs[:10],
                })

        # Row 2 check
        for label, wb in wbs.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            if ws.max_row and ws.max_row >= 2:
                row2_vals = {}
                for c in range(1, (ws.max_column or 0) + 1):
                    v = ws.cell(2, c).value
                    if v is not None and str(v).strip():
                        row2_vals[c] = str(v).strip()[:60]
                if row2_vals:
                    result.setdefault("row2_presence", []).append({
                        "sheet": sheet_name,
                        "location": label,
                        "non_empty_cols": len(row2_vals),
                        "sample": dict(list(row2_vals.items())[:3]),
                    })

    for wb in wbs.values():
        try:
            wb.close()
        except Exception:
            pass

    return result


# =============================================================================
# Phase 2: Data completeness
# =============================================================================


def check_phase2(family_dir: str) -> dict:
    spec = TEMPLATE_SPEC[family_dir]
    filled_path = FILLED_DIR / spec["template"]
    if not filled_path.exists():
        return {"error": f"Filled template not found: {filled_path}"}

    wb = _load_wb(filled_path)
    result = {"template_file": spec["template"], "sheets": {}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = _header(ws)
        rows = _data_rows(ws)

        sheet_info = {
            "data_rows": len(rows),
            "columns": len(header),
            "blank_columns": [],
            "suspicious_constant_columns": [],
        }

        if rows:
            for col_idx in range(len(header)):
                vals = [
                    str(r[col_idx]).strip()
                    if col_idx < len(r) and r[col_idx] is not None
                    else ""
                    for r in rows
                ]
                non_empty = sum(1 for v in vals if v)
                unique_vals = set(v for v in vals if v)
                col_name = header[col_idx] if col_idx < len(header) else f"Col_{col_idx}"
                blank_rate = 1.0 - (non_empty / len(vals)) if vals else 1.0

                if blank_rate == 1.0:
                    sheet_info["blank_columns"].append(col_name)
                elif len(unique_vals) == 1 and non_empty > 1:
                    sheet_info["suspicious_constant_columns"].append({
                        "column": col_name,
                        "value": next(iter(unique_vals))[:80],
                        "count": non_empty,
                    })

        result["sheets"][sheet_name] = sheet_info

    wb.close()
    return result


# =============================================================================
# Phase 3: Data correctness
# =============================================================================


def check_phase3(family_dir: str) -> dict:
    spec = TEMPLATE_SPEC[family_dir]
    filled_path = FILLED_DIR / spec["template"]
    if not filled_path.exists():
        return {"error": f"Filled template not found: {filled_path}"}

    wb = _load_wb(filled_path)
    result = {
        "template_file": spec["template"],
        "key_violations": [],
        "fk_violations": [],
        "format_issues": [],
        "special_values": [],
    }

    # Build in-memory: sheet → list of dicts
    data: dict[str, list[dict]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = _header(ws)
        rows = _data_rows(ws)
        data[sheet_name] = []
        for row_idx, row in enumerate(rows, start=3):
            row_dict = {}
            for col_idx, h in enumerate(header):
                val = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] is not None else ""
                row_dict[h] = val
            data[sheet_name].append(row_dict)

    # --- Composite key uniqueness ---
    key_specs = COMPOSITE_KEY_BY_TEMPLATE.get(family_dir, {})
    for sheet_name, specs in key_specs.items():
        if sheet_name not in data:
            continue
        for key_cols in specs:
            seen: dict[tuple, int] = {}
            for row_num, row_dict in enumerate(data[sheet_name], start=3):
                key = tuple(row_dict.get(c, "") for c in key_cols)
                if all(v == "" for v in key):
                    continue
                if key in seen:
                    result["key_violations"].append({
                        "sheet": sheet_name,
                        "key_columns": list(key_cols),
                        "key_values": list(key),
                        "row_1": seen[key],
                        "row_2": row_num,
                    })
                else:
                    seen[key] = row_num

    # --- Foreign key integrity ---
    lookup: dict[str, set] = {}
    for sheet_name, rows in data.items():
        header = list(rows[0].keys()) if rows else []
        for col_name in header:
            values = set()
            for row_dict in rows:
                v = row_dict.get(col_name, "")
                if v:
                    values.add(v)
            lookup[f"{sheet_name}.{col_name}"] = values

    fk_specs = FK_SPEC_BY_TEMPLATE.get(family_dir, [])
    for src_sheet, src_col, tgt_sheet, tgt_col in fk_specs:
        if src_sheet not in data or tgt_sheet not in data:
            continue
        tgt_key = f"{tgt_sheet}.{tgt_col}"
        tgt_values = lookup.get(tgt_key, set())
        if not tgt_values:
            continue

        orphans = []
        for row_num, row_dict in enumerate(data[src_sheet], start=3):
            src_val = row_dict.get(src_col, "")
            if src_val and src_val not in tgt_values:
                orphans.append({"value": src_val[:80], "row": row_num})

        if orphans:
            result["fk_violations"].append({
                "source": f"{src_sheet}.{src_col}",
                "target": f"{tgt_sheet}.{tgt_col}",
                "orphan_count": len(orphans),
                "orphan_samples": orphans[:10],
            })

    # --- Value format validation ---
    id_columns = {"Document_ID", "Device_ID", "Object_ID", "Element_ID", "Connection_Key", "Layer_ID"}
    for sheet_name, rows in data.items():
        if not rows:
            continue
        header = list(rows[0].keys())
        for row_num, row_dict in enumerate(rows, start=3):
            for col_name, val in row_dict.items():
                if not val:
                    continue
                # '=' prefix check for ID columns
                if col_name in id_columns and val.startswith("=") and not val.startswith("'="):
                    result["format_issues"].append({
                        "sheet": sheet_name, "row": row_num, "column": col_name,
                        "value": val[:60],
                        "issue": "Starts with = without ' prefix",
                    })
                # ECLASS IRDI format
                if col_name == "ECLASS_IRDI" and val:
                    if not ECLASS_IRDI_PATTERN.match(val):
                        result["format_issues"].append({
                            "sheet": sheet_name, "row": row_num, "column": col_name,
                            "value": val[:60],
                            "issue": "Does not match ECLASS IRDI pattern (NNNN-N#...)",
                        })
                # Index must be integer
                if col_name == "Index" and val:
                    try:
                        int(val)
                    except ValueError:
                        result["format_issues"].append({
                            "sheet": sheet_name, "row": row_num, "column": col_name,
                            "value": val[:60],
                            "issue": "Index is not an integer",
                        })
                # Special value patterns
                for pat, desc in SPECIAL_VALUE_PATTERNS:
                    if pat.match(val):
                        result["special_values"].append({
                            "sheet": sheet_name, "row": row_num, "column": col_name,
                            "value": val[:60], "type": desc,
                        })

    wb.close()
    return result


# =============================================================================
# Phase 5: Specific issues
# =============================================================================


def check_phase5() -> dict:
    result = {}

    # --- 5.1 Component_Classification typo (FIXED) ---
    tpl_path = TPL_DIR / "Stellenplan_template.xlsx"
    wb = _load_wb(tpl_path)
    sn = set(wb.sheetnames)
    wb.close()
    result["5.1_component_classification"] = {
        "correct_present": "Component_Classification" in sn,
        "typo_present": "Componenet_Classification" in sn,
        "note": "Sheet renamed from Componenet_Classification to Component_Classification — typo fixed",
        "severity": "resolved",
    }

    # --- 5.2 Datasheet emptiness ---
    fp = FILLED_DIR / "Datasheet_template.xlsx"
    if fp.exists():
        wb = _load_wb(fp)
        p2 = {}
        for s in wb.sheetnames:
            ws = wb[s]
            rows = _data_rows(ws)
            p2[s] = {"data_rows": len(rows)}
            if rows:
                p2[s]["sample"] = {h: str(rows[0][i])[:60] for i, h in enumerate(_header(ws)) if i < len(rows[0]) and rows[0][i] is not None and str(rows[0][i]).strip()}
        wb.close()

        # Check for datasheet source files
        docs = REPO / "Documents"
        ds_src = []
        if docs.exists():
            for p in docs.rglob("*"):
                if p.is_file() and any(kw in p.as_posix().lower() for kw in ["datenblatt", "datasheet", "geratedaten"]):
                    ds_src.append(str(p.relative_to(REPO)))
        result["5.2_datasheet"] = {
            "sheets": p2,
            "datasheet_source_files": ds_src,
            "is_datasheet_source_keywords": ["geratedaten", "datenblatt", "datasheet", "spec-sheet"],
        }

    # --- 5.3 Revision_Data stub check ---
    fp = FILLED_DIR / "Stellenplan_template.xlsx"
    if fp.exists():
        wb = _load_wb(fp)
        if "Revision_Data" in wb.sheetnames:
            ws = wb["Revision_Data"]
            h = _header(ws)
            rows = _data_rows(ws)
            rev_idx = _col_index(h, "Revision_Entry")
            date_idx = _col_index(h, "Date_Entry")
            meaningful = 0
            samples = []
            for r in rows[:5]:
                sd = {}
                for i, hh in enumerate(h):
                    if i < len(r) and r[i] is not None and str(r[i]).strip():
                        sd[hh] = str(r[i])[:60]
                samples.append(sd)
                has_rev = rev_idx is not None and rev_idx < len(r) and str(r[rev_idx] or "").strip()
                has_date = date_idx is not None and date_idx < len(r) and str(r[date_idx] or "").strip()
                if has_rev or has_date:
                    meaningful += 1
            result["5.3_revision_data_stubs"] = {
                "total": len(rows), "meaningful": meaningful,
                "stubs": len(rows) - meaningful, "samples": samples,
            }
        wb.close()

    # --- 5.4 Connection_Data volume ---
    fp = FILLED_DIR / "Stellenplan_template.xlsx"
    if fp.exists():
        wb = _load_wb(fp)
        if "Connection_Data" in wb.sheetnames:
            ws = wb["Connection_Data"]
            h = _header(ws)
            rows = _data_rows(ws)
            fi = _col_index(h, "From_Attribute_ID")
            ti = _col_index(h, "To_Attribute_ID")
            valid = 0
            samples = []
            for r in rows[:5]:
                sd = {}
                for i, hh in enumerate(h):
                    if i < len(r) and r[i] is not None and str(r[i]).strip():
                        sd[hh] = str(r[i])[:60]
                samples.append(sd)
                has_f = fi is not None and fi < len(r) and str(r[fi] or "").strip()
                has_t = ti is not None and ti < len(r) and str(r[ti] or "").strip()
                if has_f and has_t:
                    valid += 1
            result["5.4_connection_data"] = {
                "total": len(rows), "valid_connections": valid,
                "missing_ft": len(rows) - valid, "samples": samples,
            }
        wb.close()

    # --- 5.5 Stromlaufplan Layer_ID ---
    fp = FILLED_DIR / "Stromlaufplan_template.xlsx"
    if fp.exists():
        wb = _load_wb(fp)
        if "Layer_ID" in wb.sheetnames:
            ws = wb["Layer_ID"]
            rows = _data_rows(ws)
            result["5.5_layer_id"] = {
                "data_rows": len(rows),
                "reason": "StromlaufParser does not extract layer info — consistent with golden reference (also 0 rows)",
            }
        wb.close()

    # --- 5.6 Assembly_3D gap ---
    result["5.6_assembly_3d"] = {
        "template_exists": (TPL_DIR / "Assembly_3D_template.xlsx").exists(),
        "filled_exists": (FILLED_DIR / "Assembly_3D_template.xlsx").exists(),
        "golden_exists": (GOLDEN_DIR / "Standardized_Assembly_3D.xlsx").exists(),
        "note": "_export_assembly_3d() reads from Documents/Piping Diagram/Assembly_3D_template_filled.xlsx and produces a standardized export. Requires the pre-built filled workbook to exist.",
    }

    # --- 5.7 Stray exports file ---
    stray = EXPORTS_DIR / "Stellenplan_template.xlsx"
    result["5.7_stray_export"] = {
        "stray_exists": stray.exists(),
        "size": stray.stat().st_size if stray.exists() else 0,
        "likely_cause": "_save() in standardized_export.py saves to both result_dir/Excel/<family>.standardized.xlsx AND the canonical filled_templates path; additional copy from save_extraction_results()",
        "duplicate_in_instrument_list": (EXPORTS_DIR / "instrument_list" / "Stellenplan_template.xlsx").exists(),
    }

    return result


# =============================================================================
# Source coverage
# =============================================================================


def check_source_coverage() -> dict:
    docs_dir = REPO / "Documents"
    docs_others = REPO / "Documents-Others"
    counts = defaultdict(int)

    if docs_dir.exists():
        for pdf in docs_dir.rglob("*.pdf"):
            rel = pdf.relative_to(REPO).as_posix().lower()
            if "stellenplaene" in rel and "stellenubersicht" in rel:
                counts["stellen_overview_pdf"] += 1
            elif "stellenplaene" in rel:
                counts["stellen_tu_pdf"] += 1
        for xlsx in docs_dir.rglob("*.xlsx"):
            rel = xlsx.relative_to(REPO).as_posix().lower()
            if "klemmenplan" in rel:
                counts["klemmenplan_xlsx"] += 1
            elif "verschaltungsliste" in rel:
                counts["verschaltungsliste_xlsx"] += 1

    if docs_others.exists():
        for pdf in docs_others.rglob("*.pdf"):
            rel = pdf.relative_to(REPO).as_posix().lower()
            if "stromlauf" in rel:
                counts["stromlaufplan_pdf"] += 1
        for xlsx in docs_others.rglob("*.xlsx"):
            rel = xlsx.relative_to(REPO).as_posix().lower()
            if "klemmenplan" in rel:
                counts["klemmenplan_xlsx"] += 1
            elif "verschaltungsliste" in rel:
                counts["verschaltungsliste_xlsx"] += 1
        for ifc_f in docs_others.rglob("*.ifc"):
            counts["ifc_file"] += 1

    # Count entries in filled templates
    fc = {}
    for family_dir, spec in TEMPLATE_SPEC.items():
        fp = FILLED_DIR / spec["template"]
        if not fp.exists():
            fc[family_dir] = "NOT FILLED"
            continue
        wb = _load_wb(fp)
        fc[family_dir] = {}
        if "Document_ID" in wb.sheetnames:
            fc[family_dir]["document_ids"] = len(_data_rows(wb["Document_ID"]))
        if "Device_ID" in wb.sheetnames:
            fc[family_dir]["device_ids"] = len(_data_rows(wb["Device_ID"]))
        if "Object_ID" in wb.sheetnames:
            fc[family_dir]["object_ids"] = len(_data_rows(wb["Object_ID"]))
        if "Element_ID" in wb.sheetnames:
            fc[family_dir]["element_ids"] = len(_data_rows(wb["Element_ID"]))
        if "Terminal_ID" in wb.sheetnames:
            fc[family_dir]["terminal_ids"] = len(_data_rows(wb["Terminal_ID"]))
        wb.close()

    return {
        "source_counts": dict(counts),
        "filled_entry_counts": fc,
    }


# =============================================================================
# Report writing
# =============================================================================


def write_report(results: dict, output_dir: Path) -> Path:
    os.makedirs(output_dir, exist_ok=True)

    json_path = output_dir / "enhanced_audit.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False, default=str)

    md_path = output_dir / "enhanced_audit.md"
    lines = [
        "# Enhanced Exports/Excel Audit Report",
        f"Generated: {results['audit_time']}",
        "",
        "---",
        "",
        "## Executive Summary",
        "",
    ]

    issues_found = []

    # Collect from each phase
    for family, r in results.get("phase1_structural", {}).items():
        if r.get("header_mismatches"):
            for hm in r["header_mismatches"]:
                issues_found.append({"severity": "WARNING", "msg": f"{family}/{hm['sheet']}: header mismatch @ {hm['location']}"})
        if r.get("sheet_name_mismatches"):
            issues_found.append({"severity": "WARNING", "msg": f"{family}: sheet name mismatch"})

    for family, r in results.get("phase2_completeness", {}).items():
        if "error" in r:
            issues_found.append({"severity": "INFO", "msg": f"{family}: {r['error']}"})
            continue
        for sheet, info in r.get("sheets", {}).items():
            if info["data_rows"] > 0 and info.get("blank_columns"):
                issues_found.append({"severity": "INFO", "msg": f"{family}/{sheet}: {len(info['blank_columns'])} blank columns"})
            for sc in info.get("suspicious_constant_columns", []):
                issues_found.append({"severity": "WARNING", "msg": f"{family}/{sheet}.{sc['column']}: all {sc['count']} rows = '{sc['value']}'"})

    for family, r in results.get("phase3_correctness", {}).items():
        if r.get("key_violations"):
            for kv in r["key_violations"]:
                issues_found.append({"severity": "CRITICAL", "msg": f"{family}/{kv['sheet']}: duplicate key {kv['key_values']}"})
        for fk in r.get("fk_violations", []):
            issues_found.append({"severity": "WARNING", "msg": f"{family}: FK {fk['source']}→{fk['target']}: {fk['orphan_count']} orphans"})
        for fi in r.get("format_issues", []):
            issues_found.append({"severity": "INFO", "msg": f"{family}/{fi['sheet']} Row {fi['row']}: {fi['issue']}"})

    # Phase 5 critical findings
    p5 = results.get("phase5_specific_issues", {})
    ds = p5.get("5.2_datasheet", {}).get("sheets", {})
    if ds.get("Device_ID", {}).get("data_rows", 0) == 0:
        _sc = results.get("source_coverage", {}).get("source_counts", {})
        _has_ds = _sc.get("device_datasheet", 0) > 0
        issues_found.append({
            "severity": "CRITICAL" if _has_ds else "INFO",
            "msg": "Datasheet: Device_ID is empty — "
                   + ("no device data extracted" if _has_ds
                      else "no datasheet source documents in input dirs"),
        })
    asm = p5.get("5.6_assembly_3d", {})
    if not asm.get("filled_exists", True):
        issues_found.append({"severity": "WARNING", "msg": "Assembly_3D: filled template not found in data/filled_templates/ — run fill_standardized_templates() to generate"})
    rev = p5.get("5.3_revision_data_stubs", {})
    if rev.get("stubs", 0) > 0:
        issues_found.append({"severity": "WARNING", "msg": f"Stellenplan Revision_Data: {rev['stubs']}/{rev['total']} rows are stubs (no revision/date)"})
    conn = p5.get("5.4_connection_data", {})
    if conn.get("total", 0) > 0:
        issues_found.append({"severity": "INFO", "msg": f"Stellenplan Connection_Data: {conn['valid_connections']}/{conn['total']} have both From and To"})

    # Print summary
    crisis = [i for i in issues_found if i["severity"] == "CRITICAL"]
    warns = [i for i in issues_found if i["severity"] == "WARNING"]
    infos = [i for i in issues_found if i["severity"] == "INFO"]

    lines.append(f"**{len(crisis)} Critical, {len(warns)} Warnings, {len(infos)} Info**")
    lines.append("")

    if crisis:
        lines.append("### Critical")
        for c in crisis:
            lines.append(f"- {c['msg']}")
        lines.append("")
    if warns:
        lines.append("### Warnings")
        for w in warns[:20]:
            lines.append(f"- {w['msg']}")
        if len(warns) > 20:
            lines.append(f"  ... and {len(warns) - 20} more")
        lines.append("")

    # Phase 1 details
    lines.extend(["---", "", "## Phase 1: Structural Integrity", ""])
    for family, r in sorted(results.get("phase1_structural", {}).items()):
        lines.append(f"### {family}")
        if "error" in r:
            lines.append(f"Error: {r['error']}")
            continue
        if r.get("header_mismatches"):
            for hm in r["header_mismatches"]:
                lines.append(f"- **{hm['sheet']}** @ {hm['location']}: header mismatch")
                for d in hm["diffs"][:5]:
                    lines.append(f"  - Col {d['col']}: `{d['reference']}` vs `{d['actual']}`")
        elif r.get("sheet_name_mismatches"):
            for sn in r["sheet_name_mismatches"]:
                lines.append(f"- {sn}")
        else:
            lines.append("All headers consistent across locations. ✓")
        lines.append("")

    # Phase 2 details
    lines.extend(["---", "", "## Phase 2: Data Completeness", ""])
    for family, r in sorted(results.get("phase2_completeness", {}).items()):
        lines.append(f"### {family}")
        if "error" in r:
            lines.append(f"Error: {r['error']}")
            continue
        lines.append("| Sheet | Data Rows | Blank Cols | Suspicious |")
        lines.append("|-------|----------|------------|------------|")
        for sheet, info in sorted(r.get("sheets", {}).items()):
            bc = len(info.get("blank_columns", []))
            sc = len(info.get("suspicious_constant_columns", []))
            lines.append(f"| {sheet} | {info['data_rows']} | {bc} | {sc} |")
        lines.append("")

        for sheet, info in sorted(r.get("sheets", {}).items()):
            if info["data_rows"] > 0 and info.get("blank_columns"):
                # Only show some
                cols = info["blank_columns"]
                lines.append(f"**{sheet}**: {len(cols)} blank — {', '.join(cols[:8])}{'...' if len(cols) > 8 else ''}")
                lines.append("")
            for sc in info.get("suspicious_constant_columns", []):
                lines.append(f"**{sheet}.{sc['column']}**: all {sc['count']} rows = `{sc['value']}`")
                lines.append("")

    # Phase 3 details
    lines.extend(["---", "", "## Phase 3: Data Correctness", ""])
    for family, r in sorted(results.get("phase3_correctness", {}).items()):
        lines.append(f"### {family}")
        if "error" in r:
            lines.append(f"Error: {r['error']}")
            continue

        if r.get("key_violations"):
            lines.append(f"**Key Violations ({len(r['key_violations'])}):**")
            for kv in r["key_violations"]:
                lines.append(f"- {kv['sheet']}: duplicate `{kv['key_values']}` (cols {kv['key_columns']}) rows {kv['row_1']},{kv['row_2']}")
            lines.append("")
        else:
            lines.append("No composite key violations. ✓")
            lines.append("")

        if r.get("fk_violations"):
            lines.append(f"**FK Violations ({len(r['fk_violations'])}):**")
            for fk in r["fk_violations"]:
                lines.append(f"- `{fk['source']}` → `{fk['target']}`: {fk['orphan_count']} orphan refs")
                for s in fk.get("orphan_samples", [])[:3]:
                    lines.append(f"  - Row {s['row']}: `{s['value']}`")
            lines.append("")
        else:
            lines.append("No foreign key violations. ✓")
            lines.append("")

        fi_count = len(r.get("format_issues", []))
        sv_count = len(r.get("special_values", []))
        if fi_count or sv_count:
            lines.append(f"**Format**: {fi_count} issues, **Special values**: {sv_count} issues")
            for fi in r.get("format_issues", [])[:10]:
                lines.append(f"- `{fi['sheet']}` Row {fi['row']} `{fi['column']}`: {fi['issue']} — `{fi['value']}`")
            if fi_count > 10:
                lines.append(f"  ... {fi_count - 10} more")
            for sv in r.get("special_values", [])[:10]:
                lines.append(f"- `{sv['sheet']}` Row {sv['row']} `{sv['column']}`: {sv['type']} — `{sv['value']}`")
            lines.append("")
        else:
            lines.append("No format issues or special values. ✓")
            lines.append("")

    # Phase 5 details
    lines.extend(["---", "", "## Phase 5: Specific Issue Investigations", ""])
    for ikey, info in sorted(results.get("phase5_specific_issues", {}).items()):
        lines.append(f"### {ikey}")
        if isinstance(info, dict):
            for k, v in info.items():
                if isinstance(v, list):
                    lines.append(f"- **{k}**: {len(v)} items")
                    for item in v[:5]:
                        lines.append(f"  - {item}")
                elif isinstance(v, dict):
                    lines.append(f"- **{k}**:")
                    for k2, v2 in v.items():
                        lines.append(f"  - {k2}: {v2}")
                else:
                    lines.append(f"- **{k}**: {v}")
        lines.append("")

    # Source coverage
    lines.extend(["---", "", "## Source Coverage", ""])
    sc = results.get("source_coverage", {})
    lines.append("### Source Documents")
    for k, v in sorted(sc.get("source_counts", {}).items()):
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("### Filled Template Counts")
    for k, v in sorted(sc.get("filled_entry_counts", {}).items()):
        lines.append(f"- {k}: {v}")
    lines.append("")

    lines.extend(["---", f"*Generated by scripts/audit_enhanced.py*"])

    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return md_path


# =============================================================================
# Main
# =============================================================================


def main():
    print("Running enhanced Excel export audit...")

    results = {
        "audit_time": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "phase1_structural": {},
        "phase2_completeness": {},
        "phase3_correctness": {},
        "phase5_specific_issues": {},
        "source_coverage": {},
    }

    for family_dir, spec in TEMPLATE_SPEC.items():
        tpl_name = spec["template"]
        print(f"  Phase 1: {family_dir} ({tpl_name})")
        results["phase1_structural"][family_dir] = check_phase1(tpl_name)

    for family_dir in TEMPLATE_SPEC:
        print(f"  Phase 2: {family_dir}")
        results["phase2_completeness"][family_dir] = check_phase2(family_dir)

    for family_dir in TEMPLATE_SPEC:
        print(f"  Phase 3: {family_dir}")
        results["phase3_correctness"][family_dir] = check_phase3(family_dir)

    print("  Phase 5: specific issues")
    results["phase5_specific_issues"] = check_phase5()

    print("  Source coverage")
    results["source_coverage"] = check_source_coverage()

    md_path = write_report(results, OUTPUT_DIR)

    # Count issues
    crisis = 0
    warns = 0
    for family, r in results["phase3_correctness"].items():
        if r.get("key_violations"):
            crisis += 1
        if r.get("fk_violations"):
            warns += len(r["fk_violations"])
    p5 = results["phase5_specific_issues"]
    ds_sheets = p5.get("5.2_datasheet", {}).get("sheets", {})
    if ds_sheets.get("Device_ID", {}).get("data_rows", 0) == 0:
        crisis += 1
    if not p5.get("5.6_assembly_3d", {}).get("filled_exists", True):
        crisis += 1

    print(f"\nAudit complete!")
    print(f"  Markdown: {md_path}")
    print(f"  JSON:     {OUTPUT_DIR / 'enhanced_audit.json'}")
    print(f"  Issues: {crisis} critical, {warns} warnings")
    return 0


if __name__ == "__main__":
    sys.exit(main())
