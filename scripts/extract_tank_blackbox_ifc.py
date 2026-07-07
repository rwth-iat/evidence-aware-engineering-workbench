#!/usr/bin/env python3
"""Extract tank and blackbox parts from an assembled IFC into individual IFC files.

Can use either:
- A single assembled IFC (recommended): ``--ifc assembly.ifc``
- The legacy multi-file approach (scans from Assembly_3D_template_filled.xlsx)
"""

import argparse
import sys
from pathlib import Path

import ifcopenshell
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PIPING_DIR = PROJECT_ROOT / "Documents" / "Piping Diagram"
DEFAULT_IFC = PIPING_DIR / "assembly_prepared.ifc"
TEMPLATE_PATH = PIPING_DIR / "Assembly_3D_template_filled.xlsx"
OUTPUT_DIR = PIPING_DIR / "individual_parts"
OUTPUT_DIR.mkdir(exist_ok=True)


def collect_parts_from_template():
    """Get unique (ifc_file, part_name, label) triples from the template."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Assembly_Steps"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    ci = {h: i + 1 for i, h in enumerate(headers)}

    parts = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=ci["type"]).value
        if t not in ("tank", "blackbox"):
            continue
        lib = ws.cell(row=r, column=ci["ifc_parts_library"]).value
        pname = ws.cell(row=r, column=ci["ifc_part_name"]).value
        label = ws.cell(row=r, column=ci["label"]).value
        key = (lib, pname)
        if key not in seen:
            seen.add(key)
            parts.append((lib, pname, label))
    return parts


def create_single_part_ifc(src, part_name, output_path):
    """Extract a single named entity from source IFC into a standalone IFC file."""
    target = None
    for entity in src.by_type("IfcBuildingElementProxy"):
        if entity.Name == part_name:
            target = entity
            break

    if target is None:
        return False

    new_file = ifcopenshell.file(schema=src.schema)
    for proj in src.by_type("IfcProject"):
        new_file.add(proj)
        break
    try:
        new_file.add(target)
    except Exception:
        return False

    new_file.write(str(output_path))
    return True


def extract_from_single_ifc(ifc_path: Path, parts: list, output_dir: Path) -> tuple[int, int]:
    """Extract parts from a single assembled IFC (after rename, entity Name == label)."""
    if not ifc_path.exists():
        print(f"IFC not found: {ifc_path}", file=sys.stderr)
        return 0, len(parts)

    src = ifcopenshell.open(str(ifc_path))
    print(f"Extracting from: {ifc_path}")

    success = 0
    seen: set[str] = set()
    index_rows = ["label,type,ifc_part_name,ifc_source_file"]

    for lib, pname, label in parts:
        if label in seen:
            continue
        seen.add(label)

        output_path = output_dir / f"{label}.ifc"
        print(f"  {label:25s} → {label}.ifc ... ", end="", flush=True)

        # After rename: entity.Name == label; fall back to original part name
        if create_single_part_ifc(src, label, output_path):
            print("OK")
            success += 1
        elif create_single_part_ifc(src, pname, output_path):
            print("OK (by part_name)")
            success += 1
        else:
            print("FAIL")
            continue

        ctype = "tank" if label.startswith("tank") else "blackbox"
        index_rows.append(f"{label},{ctype},{label},{ifc_path.name}")

    # Write index
    (output_dir / "_index.csv").write_text("\n".join(index_rows) + "\n")
    return success, len(parts)


def extract_legacy(parts: list, output_dir: Path) -> tuple[int, int]:
    """Legacy: extract from multiple IFC files referenced in the template."""
    print(f"Extracting {len(parts)} parts (legacy multi-file mode)")
    success = 0
    for lib, pname, label in parts:
        ifc_path = PIPING_DIR / lib
        if not ifc_path.exists():
            print(f"  SKIP: {lib} not found")
            continue
        output_path = output_dir / f"{label}.ifc"
        print(f"  {pname:25s} → {label}.ifc ... ", end="", flush=True)
        src = ifcopenshell.open(str(ifc_path))
        if create_single_part_ifc(src, pname, output_path):
            print("OK")
            success += 1
        else:
            print("FAIL")

    index_path = output_dir / "_index.csv"
    with open(index_path, "w") as f:
        f.write("filename,label,type,ifc_part_name,ifc_source_file\n")
        for lib, pname, label in parts:
            ctype = "tank" if label.startswith("tank") else "blackbox"
            f.write(f"{label}.ifc,{label},{ctype},{pname},{lib}\n")
    return success, len(parts)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--ifc", type=Path, default=None,
                        help="Path to a single assembled IFC (prepared with semantic names)")
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR,
                        help="Output directory for individual part IFCs")
    args = parser.parse_args()

    parts = collect_parts_from_template()
    if not parts:
        print("No tank/blackbox parts found in template.", file=sys.stderr)
        return 1

    args.output_dir.mkdir(parents=True, exist_ok=True)

    if args.ifc and args.ifc.exists():
        success, total = extract_from_single_ifc(args.ifc, parts, args.output_dir)
    else:
        if args.ifc:
            print(f"IFC not found: {args.ifc}, falling back to legacy mode")
        success, total = extract_legacy(parts, args.output_dir)

    print(f"\nDone: {success}/{total} extracted → {args.output_dir}/")
    if success < total:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
