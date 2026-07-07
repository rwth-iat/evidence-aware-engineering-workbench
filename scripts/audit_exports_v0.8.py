#!/usr/bin/env python3
"""
Comprehensive audit of Exports/Excel against the v0.8 AIO workbook schema.
Checks every cell, row, column for semantic correctness, fill rate, and schema compliance.
"""
import json
import os
import re
import sys
from datetime import datetime
from collections import defaultdict
import openpyxl

EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Exports", "Excel")

# ── 28 Mandatory Sheets in exact order (§3.0) ──────────────────────
MANDATORY_SHEETS = [
    "Rules", "Schema_Metadata", "Document_ID", "Document_Data",
    "Revision_Data", "Document_RepresentedItem", "Object", "Cluster",
    "Object_Cluster", "Elements_TopDown", "Elements_from_Cluster",
    "Match_Result", "Element_ID", "Element_RepresentedItem_Mapping",
    "Element_Data", "Element_Data_Source", "RepresentedItem_Data",
    "RepresentedItem_Data_Source", "Element_Classification",
    "Connection_ID", "Connection_Data", "Connection_Data_Source",
    "Layer_ID", "Attribute_Lookup", "Enum_Lookup",
    "Document_Data_Source", "Revision_Data_Source",
    "Element_Classification_Source",
]

# ── Expected column headers per sheet (from Appendix A) ─────────────
EXPECTED_HEADERS = {
    "Rules": ["Rule_ID", "Rule_Category", "Rule_Name", "Rule_Description", "Rule_Reference", "Schema_Version_Introduced"],
    "Schema_Metadata": ["Index", "Metadata_Key", "Metadata_Value", "Description"],
    "Document_ID": ["Index", "Document_ID", "Document_Type", "Document_Filename", "Page_Count", "Schema_Version", "Lookup_Version", "Created_Timestamp", "Created_By", "SemanticID"],
    "Document_Data": ["Index", "Document_Data_ID", "Document_ID", "Attribute_Name", "Attribute_Value", "Raw_Value", "Normalized_Value", "Unit", "Quantity_Qualifier", "Parsing_Status", "SemanticID"],
    "Revision_Data": ["Index", "Revision_ID", "Document_ID", "Revision_Index", "Revision_Date", "Revision_Author", "Revision_Description", "SemanticID"],
    "Document_RepresentedItem": ["Index", "RepresentedItem_ID", "Document_ID", "RepresentedItem_Type", "Primary_RKZ", "Parent_RepresentedItem_ID", "Topic_Identification_Status", "CAEX_RoleClass_Path", "CAEX_SystemUnitClass_Path", "SemanticID"],
    "Object": ["Index", "Object_ID", "Document_ID", "Page_Number", "Object_Type", "Source_Operation", "BBox_X1", "BBox_Y1", "BBox_X2", "BBox_Y2", "Content_Text", "Content_Font_Size", "Geometry_Type", "Geometry_Closed", "Topology_From_Object_ID", "Topology_To_Object_ID", "Topology_Validation_Status", "Object_Role", "SemanticID"],
    "Cluster": ["Index", "Cluster_ID", "Document_ID", "Parent_Cluster_ID", "Container_Object_ID", "Cluster_Type", "Cluster_BBox_X1", "Cluster_BBox_Y1", "Cluster_BBox_X2", "Cluster_BBox_Y2", "Cluster_Method", "Cluster_Parameter_Set", "SemanticID"],
    "Object_Cluster": ["Index", "Object_ID", "Cluster_ID", "Membership_Reason"],
    "Elements_TopDown": ["Index", "Element_TopDown_ID", "Document_ID", "Element_Name", "Primary_RKZ", "Element_Type", "Parent_Element_TopDown_ID", "SemanticID"],
    "Elements_from_Cluster": ["Index", "Element_from_Cluster_ID", "Document_ID", "Source_Cluster_ID", "Element_Name", "Primary_RKZ_Extracted", "Element_Type_Inferred", "Derivation_Status", "SemanticID"],
    "Match_Result": ["Index", "Match_ID", "Document_ID", "Element_TopDown_ID", "Element_from_Cluster_ID", "Match_Status", "Match_Rule", "Resolution_Note", "Resolution_Status", "Reviewed_By", "Review_Status", "Correction_Reason", "Review_Timestamp", "SemanticID"],
    "Element_ID": ["Index", "Element_ID", "Document_ID", "Source_Match_ID", "Source", "Element_Type", "Primary_RKZ", "Parent_Element_ID", "Layer_ID", "CAEX_Type", "CAEX_RoleClass_Path", "CAEX_SystemUnitClass_Path", "CAEX_InterfaceClass_Path", "SemanticID"],
    "Element_RepresentedItem_Mapping": ["Index", "Mapping_ID", "Element_ID", "RepresentedItem_ID", "Relationship_Type", "SemanticID"],
    "Element_Data": ["Index", "Element_Data_ID", "Element_ID", "Attribute_Name", "Attribute_Value", "Raw_Value", "Normalized_Value", "Unit", "Quantity_Qualifier", "Parsing_Status", "SemanticID"],
    "Element_Data_Source": ["Index", "Element_Data_ID", "Source_Object_ID", "Source_Role", "Extraction_Method", "Confidence", "Reviewed_By", "Review_Status", "Correction_Reason", "Extraction_Timestamp", "SemanticID"],
    "RepresentedItem_Data": ["Index", "RepresentedItem_Data_ID", "RepresentedItem_ID", "Attribute_Name", "Attribute_Value", "Raw_Value", "Normalized_Value", "Unit", "Quantity_Qualifier", "Parsing_Status", "SemanticID"],
    "RepresentedItem_Data_Source": ["Index", "RepresentedItem_Data_ID", "Source_Object_ID", "Source_Role", "Extraction_Method", "Confidence", "Reviewed_By", "Review_Status", "Correction_Reason", "Extraction_Timestamp", "SemanticID"],
    "Element_Classification": ["Index", "Classification_ID", "Document_ID", "Classified_Object_Type", "Classified_Object_ID", "Classification_System", "Classification_Code", "Classification_Description", "Source_Symbol_Reference", "SemanticID"],
    "Connection_ID": ["Index", "Connection_ID", "Document_ID", "From_Element_ID", "To_Element_ID", "Source_Topology_Object_ID", "Connection_Status", "Cable_Data_ID", "SemanticID"],
    "Connection_Data": ["Index", "Connection_Data_ID", "Connection_ID", "Attribute_Name", "Attribute_Value", "Raw_Value", "Normalized_Value", "Unit", "Quantity_Qualifier", "Parsing_Status", "SemanticID"],
    "Connection_Data_Source": ["Index", "Connection_Data_ID", "Source_Object_ID", "Source_Role", "Extraction_Method", "Confidence", "Reviewed_By", "Review_Status", "Correction_Reason", "Extraction_Timestamp", "SemanticID"],
    "Layer_ID": ["Index", "Layer_ID", "Document_ID", "Layer_Description", "Layer_Type", "Voltage_Level", "SemanticID"],
    "Attribute_Lookup": ["Index", "Lookup_ID", "Scope", "Type_Constraint", "Attribute_Name", "Required", "Data_Type", "Allowed_Values_Enum_Field", "Normative_Reference", "Description", "Schema_Version_Introduced"],
    "Enum_Lookup": ["Index", "Enum_Lookup_ID", "Field_Name", "Allowed_Value", "Description", "Normative_Reference"],
    "Document_Data_Source": ["Index", "Document_Data_ID", "Source_Object_ID", "Source_Role", "Extraction_Method", "Confidence", "Reviewed_By", "Review_Status", "Correction_Reason", "Extraction_Timestamp", "SemanticID"],
    "Revision_Data_Source": ["Index", "Revision_ID", "Source_Object_ID", "Source_Role", "Extraction_Method", "Confidence", "Reviewed_By", "Review_Status", "Correction_Reason", "Extraction_Timestamp", "SemanticID"],
    "Element_Classification_Source": ["Index", "Classification_ID", "Source_Object_ID", "Source_Role", "Extraction_Method", "Confidence", "Reviewed_By", "Review_Status", "Correction_Reason", "Extraction_Timestamp", "SemanticID"],
}

# ── ID prefix per sheet (§3.7.2) ────────────────────────────────────
ID_PREFIXES = {
    "Document_ID": "D.",
    "Document_Data": "DD.",
    "Revision_Data": "R.",
    "Document_RepresentedItem": "RI.",
    "RepresentedItem_Data": "RID.",
    "Object": "O.",
    "Cluster": "CL.",
    "Elements_TopDown": "ETD.",
    "Elements_from_Cluster": "EFC.",
    "Match_Result": "M.",
    "Element_ID": "E.",
    "Element_RepresentedItem_Mapping": "MAP.",
    "Element_Data": "ED.",
    "Element_Classification": "EC.",
    "Connection_ID": "C.",
    "Connection_Data": "CD.",
    "Layer_ID": None,  # project-specific
    "Attribute_Lookup": "AL.",
    "Enum_Lookup": "EL.",
}

# PK column per sheet
PK_COLUMN = {
    "Rules": "Rule_ID",
    "Schema_Metadata": "Metadata_Key",
    "Document_ID": "Document_ID",
    "Document_Data": "Document_Data_ID",
    "Revision_Data": "Revision_ID",
    "Document_RepresentedItem": "RepresentedItem_ID",
    "Object": "Object_ID",
    "Cluster": "Cluster_ID",
    "Object_Cluster": None,  # composite key
    "Elements_TopDown": "Element_TopDown_ID",
    "Elements_from_Cluster": "Element_from_Cluster_ID",
    "Match_Result": "Match_ID",
    "Element_ID": "Element_ID",
    "Element_RepresentedItem_Mapping": "Mapping_ID",
    "Element_Data": "Element_Data_ID",
    "Element_Data_Source": None,  # FK-based
    "RepresentedItem_Data": "RepresentedItem_Data_ID",
    "RepresentedItem_Data_Source": None,
    "Element_Classification": "Classification_ID",
    "Connection_ID": "Connection_ID",
    "Connection_Data": "Connection_Data_ID",
    "Connection_Data_Source": None,
    "Layer_ID": "Layer_ID",
    "Attribute_Lookup": "Lookup_ID",
    "Enum_Lookup": "Enum_Lookup_ID",
    "Document_Data_Source": None,
    "Revision_Data_Source": None,
    "Element_Classification_Source": None,
}

# ── Enum fields and their expected value domains (from §9.3) ─────────
ENUM_DOMAINS = {
    "Document_Type": {"Instrument_Loop_Diagram", "Terminal_Diagram", "Circuit_Diagram"},
    "Document_Subtype": {"Primary", "Secondary"},
    "RepresentedItem_Type": {"PCE_Request", "Terminal_Strip", "Circuit", "Plant_Section", "Control_Cabinet", "Distribution_Panel", "Function_Group"},
    "Element_Type": {"Sensor", "Transducer", "Valve_Actuator", "Motor", "Actuator", "Consumer", "Terminal", "Terminal_Strip", "Contactor", "Auxiliary_Contactor", "Fuse", "Circuit_Breaker", "Switch", "Socket_Outlet", "Power_Supply", "PLC_Module", "Coil", "Main_Contact", "Auxiliary_Contact", "Indicator_Lamp", "Control_Cabinet", "Cabinet_Aggregate", "Thermostat", "Heater"},
    "Classification_System": {"IEC 81346-2", "IEC 62424", "IEC 60617-2", "IEC 60617-3", "IEC 60617-6", "IEC 60617-7", "IEC 60617-8", "DIN 19227-2"},
    "Source_Format": {"PDF_Drawing", "Excel_Sheet", "Verschaltungsliste", "Manual_Entry"},
    "Scope": {"Document", "Element", "RepresentedItem", "Connection"},
    "Match_Status": {"Matched", "Only_TopDown", "Only_Cluster"},
    "Match_Rule": {"M1_Primary_RKZ", "M2_Spatial", "Manual_Resolution", "Not_Applicable"},
    "Resolution_Status": {"Open", "Resolved_AutoMatch", "Resolved_KeepBoth", "Resolved_TopDown_Valid", "Resolved_Cluster_Valid", "Rejected"},
    "Connection_Type": {"Wire", "Bridge_Longitudinal", "Bridge_Cross_Fixed", "Bridge_Cross_Pluggable", "Bridge_Insulated"},
    "Connection_Status": {"Resolved", "Unresolved"},
    "Object_Type": {"Text", "Graphic", "Topology"},
    "Object_Role": {"Connection_Point", "Label", "Symbol", "Border", "Annotation", "Topology"},
    "Geometry_Type": {"rect", "line", "path"},
    "Source_Operation": {"Tj", "TJ", "'", '"', "f", "F", "f*", "S", "s", "B", "b", "re", "Cell", "VL_Row", "Manual_Entry"},
    "Source_Role": {"Label", "Value", "Symbol"},
    "Cluster_Type": {"Containment", "Proximity", "Topology", "Pre_Existing_Structural"},
    "Membership_Reason": {"Containment", "Proximity", "Pre_Existing_Structural"},
    "Topology_Validation_Status": {"Valid_Connection", "Unresolved"},
    "Rule_Category": {"A", "C", "M", "P", "E", "K", "I", "S", "AG"},
    "Data_Type": {"String", "Enum", "Integer", "Float", "Boolean", "FK", "Date", "DateTime"},
    "Source": {"TopDown", "Cluster", "Matched", "Manual_Entry", "Rule_Derived"},
    "CAEX_Type": {"InternalElement", "ExternalInterface"},
    "Relationship_Type": {"Primary", "Shared", "Secondary"},
    "Derivation_Status": {"Element_Derived", "No_Element_Derivable", "Ambiguous", "Failed"},
    "Extraction_Method": {"OCR", "Native_Text", "LLM_Classification", "Manual_Entry", "Rule_Based_Parser"},
    "Classified_Object_Type": {"Element", "RepresentedItem", "Connection", "Document"},
    "Topic_Identification_Status": {"Confirmed", "Inferred", "Ambiguous", "Failed"},
    "Layer_Type": {"Voltage_Level", "Functional_Section", "Signal_Group", "Protection_Group"},
    "Cable_Modeling_Profile": {"Core", "Asset"},
    "Review_Status": {"Unreviewed", "Requires_Review", "Auto_Approved", "Manually_Reviewed", "Manually_Corrected", "Rejected"},
    "Voltage_Level": {"230V_AC", "400V_AC", "24V_DC", "Signal_4_20mA", "Signal_0_10V", "Bus_Signal"},
    "Wire_Color": {"BK", "BN", "RD", "OG", "YE", "GN", "BU", "VT", "GY", "WH", "PK", "TQ", "GNYE"},
    "Polarity": {"L1", "L2", "L3", "L", "N", "PE", "PEN", "L+", "L-", "G+", "G-", "V+", "V-", "FE", "AC", "DC"},
    "Parsing_Status": {"Parsed_OK", "Parsed_Ambiguous", "Parsed_Failed"},
}

# ── Column-to-enum mapping (which columns should contain enum values) ─
COLUMN_ENUM_MAP = {
    "Document_ID.Document_Type": "Document_Type",
    "Document_Data.Parsing_Status": "Parsing_Status",
    "Document_RepresentedItem.RepresentedItem_Type": "RepresentedItem_Type",
    "Document_RepresentedItem.Topic_Identification_Status": "Topic_Identification_Status",
    "Object.Object_Type": "Object_Type",
    "Object.Source_Operation": "Source_Operation",
    "Object.Object_Role": "Object_Role",
    "Object.Geometry_Type": "Geometry_Type",
    "Object.Topology_Validation_Status": "Topology_Validation_Status",
    "Cluster.Cluster_Type": "Cluster_Type",
    "Object_Cluster.Membership_Reason": "Membership_Reason",
    "Elements_TopDown.Element_Type": "Element_Type",
    "Elements_from_Cluster.Derivation_Status": "Derivation_Status",
    "Elements_from_Cluster.Element_Type_Inferred": "Element_Type",
    "Match_Result.Match_Status": "Match_Status",
    "Match_Result.Match_Rule": "Match_Rule",
    "Match_Result.Resolution_Status": "Resolution_Status",
    "Match_Result.Review_Status": "Review_Status",
    "Element_ID.Source": "Source",
    "Element_ID.Element_Type": "Element_Type",
    "Element_ID.CAEX_Type": "CAEX_Type",
    "Element_RepresentedItem_Mapping.Relationship_Type": "Relationship_Type",
    "Element_Data.Parsing_Status": "Parsing_Status",
    "Element_Data_Source.Source_Role": "Source_Role",
    "Element_Data_Source.Extraction_Method": "Extraction_Method",
    "Element_Data_Source.Review_Status": "Review_Status",
    "RepresentedItem_Data.Parsing_Status": "Parsing_Status",
    "RepresentedItem_Data_Source.Source_Role": "Source_Role",
    "RepresentedItem_Data_Source.Extraction_Method": "Extraction_Method",
    "RepresentedItem_Data_Source.Review_Status": "Review_Status",
    "Element_Classification.Classified_Object_Type": "Classified_Object_Type",
    "Element_Classification.Classification_System": "Classification_System",
    "Connection_ID.Connection_Status": "Connection_Status",
    "Connection_Data.Parsing_Status": "Parsing_Status",
    "Connection_Data.Attribute_Name": None,  # validated against Attribute_Lookup
    "Connection_Data_Source.Source_Role": "Source_Role",
    "Connection_Data_Source.Extraction_Method": "Extraction_Method",
    "Connection_Data_Source.Review_Status": "Review_Status",
    "Layer_ID.Layer_Type": "Layer_Type",
    "Layer_ID.Voltage_Level": "Voltage_Level",
    "Attribute_Lookup.Scope": "Scope",
    "Attribute_Lookup.Data_Type": "Data_Type",
    "Document_Data_Source.Source_Role": "Source_Role",
    "Document_Data_Source.Extraction_Method": "Extraction_Method",
    "Document_Data_Source.Review_Status": "Review_Status",
    "Revision_Data_Source.Source_Role": "Source_Role",
    "Revision_Data_Source.Extraction_Method": "Extraction_Method",
    "Revision_Data_Source.Review_Status": "Review_Status",
    "Element_Classification_Source.Source_Role": "Source_Role",
    "Element_Classification_Source.Extraction_Method": "Extraction_Method",
    "Element_Classification_Source.Review_Status": "Review_Status",
}

# Also check Enum_Lookup sheet values for connection data attributes
CONNECTION_ATTR_ENUMS = {
    "Wire_Color": "Wire_Color",
    "Wire_Color_Secondary": "Wire_Color",
    "Polarity": "Polarity",
    "Cable_Type": None,  # open
    "Connection_Type": "Connection_Type",
    "Shielding": None,  # open
    "Voltage_Level": "Voltage_Level",
}


def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


def is_empty(v):
    return v is None or str(v).strip() == ""


def load_workbook_safe(path):
    """Load workbook, handling issues."""
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return None


def get_headers(ws):
    """Get header row as dict: col_idx → header_name."""
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        val = safe_str(cell.value)
        if val:
            headers[col_idx] = val
    return headers


def get_column_values(ws, col_names, start_row=2):
    """Get list of values for given column names."""
    headers = get_headers(ws)
    # Build name→idx map
    name_to_idx = {}
    for idx, name in headers.items():
        name_to_idx[name] = idx

    result = {}
    for cn in col_names:
        if cn not in name_to_idx:
            result[cn] = []
            continue
        idx = name_to_idx[cn]
        vals = []
        for row in ws.iter_rows(min_row=start_row, max_col=idx, max_row=ws.max_row):
            cell = row[idx - 1]
            vals.append(cell.value)
        result[cn] = vals
    return result


def read_sheet_as_dicts(ws, start_row=2):
    """Read sheet rows as list of dicts."""
    headers = get_headers(ws)
    if not headers:
        return [], {}
    name_to_idx = {v: k for k, v in headers.items()}
    rows = []
    for r in range(start_row, ws.max_row + 1):
        row_dict = {}
        for name, idx in name_to_idx.items():
            cell = ws.cell(row=r, column=idx)
            row_dict[name] = cell.value
        rows.append(row_dict)
    return rows, name_to_idx


class AuditResult:
    def __init__(self, filename):
        self.filename = filename
        self.errors = []       # spec violations, must fix
        self.warnings = []     # potential issues, review
        self.info = []         # observations
        self.fill_stats = {}   # sheet → fill rate info
        self.sheet_stats = {}  # sheet → {rows, cols, empty_cells, total_cells}

    def add_error(self, sheet, detail):
        self.errors.append(f"[{sheet}] {detail}")

    def add_warning(self, sheet, detail):
        self.warnings.append(f"[{sheet}] {detail}")

    def add_info(self, sheet, detail):
        self.info.append(f"[{sheet}] {detail}")

    def summary(self):
        lines = []
        lines.append(f"\n{'='*80}")
        lines.append(f"FILE: {self.filename}")
        lines.append(f"{'='*80}")
        lines.append(f"  ERRORS:   {len(self.errors)}")
        lines.append(f"  WARNINGS: {len(self.warnings)}")
        lines.append(f"  INFO:     {len(self.info)}")
        return "\n".join(lines)


def audit_sheet_structure(wb, result):
    """Check sheet count, names, and order (§3.0, D2)."""
    actual = wb.sheetnames

    # Check for unexpected sheets
    mandatory_found = [s for s in actual if s in MANDATORY_SHEETS]
    optional_found = [s for s in actual if s not in MANDATORY_SHEETS]

    # Check mandatory sheets presence
    missing = set(MANDATORY_SHEETS) - set(actual)
    for s in missing:
        result.add_error("STRUCTURE", f"Missing mandatory sheet: {s}")

    # Check order of mandatory sheets
    expected_order = [s for s in MANDATORY_SHEETS if s in actual]
    actual_mandatory_order = [s for s in actual if s in MANDATORY_SHEETS]
    if expected_order != actual_mandatory_order:
        for i, (exp, act) in enumerate(zip(expected_order, actual_mandatory_order)):
            if exp != act:
                result.add_error("STRUCTURE", f"Sheet order wrong at position {i+1}: expected '{exp}', got '{act}'")
                break

    if optional_found:
        result.add_info("STRUCTURE", f"Optional sheets present: {optional_found}")


def audit_headers(ws_name, ws, result):
    """Check column headers match expected for the sheet."""
    if ws_name not in EXPECTED_HEADERS:
        return  # optional sheet, skip

    expected = EXPECTED_HEADERS[ws_name]
    actual_headers = []
    for cell in ws[1]:
        val = safe_str(cell.value)
        if val:
            actual_headers.append(val)

    if actual_headers != expected:
        missing = set(expected) - set(actual_headers)
        extra = set(actual_headers) - set(expected)
        if missing:
            result.add_error(ws_name, f"Missing columns: {missing}")
        if extra:
            result.add_warning(ws_name, f"Extra columns: {extra}")
        # Check order
        if not missing and not extra:
            for i, (exp, act) in enumerate(zip(expected, actual_headers)):
                if exp != act:
                    result.add_error(ws_name, f"Column order wrong at pos {i+1}: expected '{exp}', got '{act}'")
                    break


def audit_id_formats(ws_name, ws, result):
    """Check ID formats per §3.7 (prefix.integer)."""
    if ws_name not in ID_PREFIXES or ID_PREFIXES[ws_name] is None:
        return

    pk_col = PK_COLUMN.get(ws_name)
    if pk_col is None:
        return

    prefix = ID_PREFIXES[ws_name]
    headers = get_headers(ws)
    if pk_col not in headers.values():
        return

    pk_idx = None
    for idx, name in headers.items():
        if name == pk_col:
            pk_idx = idx
            break

    ids = []
    for row in ws.iter_rows(min_row=2, max_col=pk_idx, max_row=ws.max_row):
        cell = row[pk_idx - 1]
        if cell.value is not None:
            ids.append(safe_str(cell.value))

    expected_pattern = re.escape(prefix) + r"\d+$"

    for id_val in ids:
        if not id_val:
            result.add_error(ws_name, f"Empty {pk_col} value")
            continue
        if not id_val.startswith(prefix):
            result.add_error(ws_name, f"ID '{id_val}' does not start with prefix '{prefix}'")
        elif not re.match(r"^" + re.escape(prefix) + r"\d+$", id_val):
            # Check for sub-element IDs (E.1.signal_out pattern)
            if ws_name == "Element_ID" and re.match(r"^E\.\d+\.[a-z_]+$", id_val):
                continue  # valid sub-element ID
            result.add_error(ws_name, f"ID '{id_val}' does not match pattern '{prefix}<integer>'")

    # Check sequential numbering (no gaps)
    if ids:
        nums = []
        for id_val in ids:
            match = re.match(r"^" + re.escape(prefix) + r"(\d+)$", id_val)
            if match:
                nums.append(int(match.group(1)))
            elif ws_name == "Element_ID":
                # Try sub-element pattern
                match2 = re.match(r"^E\.(\d+)\.[a-z_]+$", id_val)
                if match2:
                    nums.append(int(match2.group(1)))
        if nums:
            nums = sorted(set(nums))
            if nums[0] != 1:
                result.add_warning(ws_name, f"Sequential IDs do not start at 1 (first={nums[0]})")
            # Check for gaps
            for i, num in enumerate(nums):
                expected_num = i + 1
                if num != expected_num:
                    result.add_warning(ws_name, f"Gap in sequential IDs: expected {expected_num}, got {num}")
                    break


def audit_enum_values(ws_name, ws, result, enum_lookup_values=None):
    """Check enum columns contain valid values."""
    # Build enum lookup from Enum_Lookup sheet if provided
    if enum_lookup_values is None:
        enum_lookup_values = {}

    headers = get_headers(ws)
    name_to_idx = {v: k for k, v in headers.items()}

    for col_name, enum_field in COLUMN_ENUM_MAP.items():
        sheet, col = col_name.split(".", 1)
        if sheet != ws_name:
            continue
        if col not in name_to_idx:
            continue

        idx = name_to_idx[col]
        valid_set = ENUM_DOMAINS.get(enum_field, set())

        # Also check Enum_Lookup sheet
        if enum_field in enum_lookup_values:
            valid_set = valid_set | set(enum_lookup_values[enum_field])

        if not valid_set:
            continue

        for row_num in range(2, ws.max_row + 1):
            val = ws.cell(row=row_num, column=idx).value
            if val is None or safe_str(val) == "":
                continue
            val_str = safe_str(val)

            # Check if value is valid
            if val_str not in valid_set and val_str != "Unspecifiable":
                result.add_error(ws_name,
                    f"Row {row_num}: Invalid enum '{val_str}' in column '{col}' "
                    f"(expected one of {sorted(valid_set)[:20]}{'...' if len(valid_set)>20 else ''})")


def audit_enum_lookup_sheet(ws_name, ws, result):
    """Extract enum values from Enum_Lookup sheet and validate them."""
    if ws_name != "Enum_Lookup":
        return {}

    rows, mapping = read_sheet_as_dicts(ws)
    enum_map = defaultdict(set)

    for row in rows:
        field_name = safe_str(row.get("Field_Name", ""))
        allowed_value = safe_str(row.get("Allowed_Value", ""))
        if field_name and allowed_value:
            enum_map[field_name].add(allowed_value)

    # Audit Enum_Lookup structure
    if "Field_Name" not in mapping:
        result.add_error(ws_name, "Missing 'Field_Name' column")
    if "Allowed_Value" not in mapping:
        result.add_error(ws_name, "Missing 'Allowed_Value' column")

    # Check required Field_Names are present
    required_fields = [
        "Document_Type", "Element_Type", "RepresentedItem_Type",
        "Classification_System", "Source_Format", "Match_Status",
        "Connection_Type", "Object_Type", "Rule_Category",
        "Relationship_Type", "Connection_Status", "CAEX_Type",
        "Source", "Scope", "Cluster_Type", "Membership_Reason",
        "Voltage_Level", "Wire_Color", "Polarity", "Data_Type",
        "Review_Status", "Extraction_Method", "Source_Role",
        "Topology_Validation_Status", "Topic_Identification_Status",
        "Layer_Type", "Cable_Modeling_Profile", "Classified_Object_Type",
        "Derivation_Status", "Parsing_Status", "Match_Rule",
        "Resolution_Status", "Object_Role", "Geometry_Type",
        "Source_Operation", "IEC_81346_2_Class",
    ]

    for field in required_fields:
        if field not in enum_map:
            result.add_warning(ws_name, f"Required Field_Name '{field}' not found in Enum_Lookup")

    # Check that IDs follow EL. prefix
    for row in rows:
        el_id = safe_str(row.get("Enum_Lookup_ID", ""))
        if el_id and not el_id.startswith("EL."):
            result.add_error(ws_name, f"Enum_Lookup_ID '{el_id}' does not start with 'EL.'")

    return dict(enum_map)


def audit_fk_integrity(wb, result):
    """Check foreign key references are valid (I1-I22 subset)."""
    # Build ID registries from each sheet
    id_registry = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        pk = PK_COLUMN.get(sheet_name)
        if pk is None:
            continue
        headers = get_headers(ws)
        if pk not in headers.values():
            continue
        pk_idx = None
        for idx, name in headers.items():
            if name == pk:
                pk_idx = idx
                break
        ids = set()
        for row in ws.iter_rows(min_row=2, max_col=pk_idx, max_row=ws.max_row):
            cell = row[pk_idx - 1]
            if cell.value is not None and safe_str(cell.value):
                ids.add(safe_str(cell.value))
        id_registry[sheet_name] = ids

    # FK checks
    fk_checks = [
        # (source_sheet, fk_col, target_sheet, target_pk)
        ("Document_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Revision_Data", "Document_ID", "Document_ID", "Document_ID"),
        ("Document_RepresentedItem", "Document_ID", "Document_ID", "Document_ID"),
        ("Object", "Document_ID", "Document_ID", "Document_ID"),
        ("Cluster", "Document_ID", "Document_ID", "Document_ID"),
        ("Elements_TopDown", "Document_ID", "Document_ID", "Document_ID"),
        ("Elements_from_Cluster", "Document_ID", "Document_ID", "Document_ID"),
        ("Match_Result", "Document_ID", "Document_ID", "Document_ID"),
        ("Element_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Connection_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Element_Classification", "Document_ID", "Document_ID", "Document_ID"),
        ("Layer_ID", "Document_ID", "Document_ID", "Document_ID"),
        ("Match_Result", "Element_TopDown_ID", "Elements_TopDown", "Element_TopDown_ID"),
        ("Match_Result", "Element_from_Cluster_ID", "Elements_from_Cluster", "Element_from_Cluster_ID"),
        ("Element_ID", "Source_Match_ID", "Match_Result", "Match_ID"),
        ("Elements_from_Cluster", "Source_Cluster_ID", "Cluster", "Cluster_ID"),
        ("Element_RepresentedItem_Mapping", "Element_ID", "Element_ID", "Element_ID"),
        ("Element_RepresentedItem_Mapping", "RepresentedItem_ID", "Document_RepresentedItem", "RepresentedItem_ID"),
        ("Element_Data", "Element_ID", "Element_ID", "Element_ID"),
        ("RepresentedItem_Data", "RepresentedItem_ID", "Document_RepresentedItem", "RepresentedItem_ID"),
        ("Connection_ID", "From_Element_ID", "Element_ID", "Element_ID"),
        ("Connection_ID", "To_Element_ID", "Element_ID", "Element_ID"),
        ("Connection_Data", "Connection_ID", "Connection_ID", "Connection_ID"),
        ("Object_Cluster", "Object_ID", "Object", "Object_ID"),
        ("Object_Cluster", "Cluster_ID", "Cluster", "Cluster_ID"),
    ]

    for src_sheet, fk_col, tgt_sheet, tgt_pk in fk_checks:
        if src_sheet not in wb.sheetnames or tgt_sheet not in wb.sheetnames:
            continue
        ws = wb[src_sheet]
        headers = get_headers(ws)
        if fk_col not in headers.values():
            continue
        fk_idx = None
        for idx, name in headers.items():
            if name == fk_col:
                fk_idx = idx
                break

        target_ids = id_registry.get(tgt_sheet, set())

        for row_num in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=fk_idx)
            val = safe_str(cell.value)
            if not val:
                continue
            if val not in target_ids:
                result.add_error(src_sheet,
                    f"Row {row_num}: FK {fk_col}='{val}' not found in {tgt_sheet}.{tgt_pk}")


def audit_fill_rate(ws_name, ws, result):
    """Calculate fill rate for each sheet."""
    total_cells = 0
    filled_cells = 0
    col_empty_counts = defaultdict(int)
    row_empty_counts = defaultdict(int)

    headers = get_headers(ws)
    if not headers:
        result.fill_stats[ws_name] = {"rows": 0, "cols": 0, "fill_rate": 0, "note": "empty"}
        return

    for row_num in range(2, ws.max_row + 1):
        row_total = 0
        row_filled = 0
        for col_num, header_name in headers.items():
            cell = ws.cell(row=row_num, column=col_num)
            total_cells += 1
            row_total += 1
            if not is_empty(cell.value):
                filled_cells += 1
                row_filled += 1
            else:
                col_empty_counts[header_name] += 1
        row_empty_counts[row_num] = row_total - row_filled

    n_rows = ws.max_row - 1  # exclude header
    n_cols = len(headers)
    fill_rate = (filled_cells / total_cells * 100) if total_cells > 0 else 0

    result.fill_stats[ws_name] = {
        "rows": n_rows,
        "cols": n_cols,
        "filled": filled_cells,
        "total": total_cells,
        "fill_rate": round(fill_rate, 1),
        "empty_columns": {k: v for k, v in col_empty_counts.items() if v == n_rows and n_rows > 0},
        "empty_rows": [k for k, v in row_empty_counts.items() if v == n_cols and n_cols > 0],
    }

    # Flag low fill rate
    if n_rows > 0 and fill_rate < 50:
        result.add_warning(ws_name, f"Low fill rate: {fill_rate:.1f}% ({filled_cells}/{total_cells} cells)")


def _count_data_rows(ws):
    """Count rows with at least one non-empty cell (skip phantom rows)."""
    count = 0
    for row_num in range(2, ws.max_row + 1):
        row_empty = True
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value is not None and str(cell.value).strip():
                row_empty = False
                break
        if not row_empty:
            count += 1
    return count


def audit_document_id_uniqueness(wb, result):
    """Check Document_ID sheet has exactly 1 row (§1.1 principle 6)."""
    ws = wb["Document_ID"]
    n_rows = _count_data_rows(ws)
    if n_rows == 0:
        result.add_error("Document_ID", "No Document_ID rows (must have exactly 1)")
    elif n_rows > 1:
        result.add_error("Document_ID", f"Multiple Document_ID rows ({n_rows}) (must have exactly 1 per §1.1 principle 6)")


def audit_schema_metadata(wb, result):
    """Check Schema_Metadata for required keys."""
    if "Schema_Metadata" not in wb.sheetnames:
        return
    ws = wb["Schema_Metadata"]
    rows, mapping = read_sheet_as_dicts(ws)

    required_keys = {
        "Schema_Version": "v0.8",
        "Lookup_Version": "v0.8.0",
        "Cable_Modeling_Profile": None,  # must exist
    }

    found_keys = {}
    for row in rows:
        key = safe_str(row.get("Metadata_Key", ""))
        val = safe_str(row.get("Metadata_Value", ""))
        if key:
            found_keys[key] = val

    for key, expected_val in required_keys.items():
        if key not in found_keys:
            result.add_error("Schema_Metadata", f"Missing required Metadata_Key: '{key}'")
        elif expected_val is not None and found_keys[key] != expected_val:
            result.add_error("Schema_Metadata", f"Metadata '{key}': expected '{expected_val}', got '{found_keys[key]}'")

    # Check Cable_Modeling_Profile is valid
    if "Cable_Modeling_Profile" in found_keys:
        val = found_keys["Cable_Modeling_Profile"]
        if val not in ("Core", "Asset"):
            result.add_error("Schema_Metadata", f"Cable_Modeling_Profile='{val}' not in {{Core, Asset}}")


def audit_source_format(wb, result):
    """Check I29: exactly one Source_Format Document_Data row."""
    if "Document_Data" not in wb.sheetnames:
        return
    ws = wb["Document_Data"]
    rows, mapping = read_sheet_as_dicts(ws)

    source_format_rows = []
    for row in rows:
        attr = safe_str(row.get("Attribute_Name", ""))
        val = safe_str(row.get("Attribute_Value", ""))
        if attr == "Source_Format":
            source_format_rows.append(val)

    if len(source_format_rows) == 0:
        result.add_error("Document_Data", "I29: No Source_Format row (must have exactly 1)")
    elif len(source_format_rows) > 1:
        result.add_error("Document_Data", f"I29: Multiple Source_Format rows ({len(source_format_rows)}), must have exactly 1")

    for val in source_format_rows:
        if val and val not in ENUM_DOMAINS["Source_Format"]:
            result.add_error("Document_Data", f"I29: Invalid Source_Format value '{val}'")


def audit_i12_source_completeness(wb, result):
    """Check each data row has at least one source row (I12)."""
    checks = [
        ("Document_Data", "Document_Data_ID", "Document_Data_Source", "Document_Data_ID"),
        ("Element_Data", "Element_Data_ID", "Element_Data_Source", "Element_Data_ID"),
        ("RepresentedItem_Data", "RepresentedItem_Data_ID", "RepresentedItem_Data_Source", "RepresentedItem_Data_ID"),
        ("Connection_Data", "Connection_Data_ID", "Connection_Data_Source", "Connection_Data_ID"),
        ("Revision_Data", "Revision_ID", "Revision_Data_Source", "Revision_ID"),
        ("Element_Classification", "Classification_ID", "Element_Classification_Source", "Classification_ID"),
    ]

    for data_sheet, id_col, src_sheet, src_id_col in checks:
        if data_sheet not in wb.sheetnames or src_sheet not in wb.sheetnames:
            continue

        # Get all data IDs
        ws_data = wb[data_sheet]
        data_rows, _ = read_sheet_as_dicts(ws_data)
        if not data_rows:
            continue

        data_ids = set()
        for row in data_rows:
            val = safe_str(row.get(id_col, ""))
            if val:
                data_ids.add(val)

        # Get all source IDs
        ws_src = wb[src_sheet]
        src_rows, _ = read_sheet_as_dicts(ws_src)
        src_ids = set()
        for row in src_rows:
            val = safe_str(row.get(src_id_col, ""))
            if val:
                src_ids.add(val)

        # Check coverage
        missing = data_ids - src_ids
        if missing:
            result.add_error(data_sheet,
                f"I12: {len(missing)}/{len(data_ids)} rows lack {src_sheet} entry "
                f"(missing IDs: {sorted(missing)[:10]}{'...' if len(missing)>10 else ''})")


def audit_semantic_correctness(wb, result):
    """Check semantic correctness of key content fields."""
    # Check I24: Wire_Color/Polarity consistency
    if "Connection_Data" not in wb.sheetnames:
        return

    ws = wb["Connection_Data"]
    rows, _ = read_sheet_as_dicts(ws)

    # Group by Connection_ID
    conn_attrs = defaultdict(dict)
    for row in rows:
        cid = safe_str(row.get("Connection_ID", ""))
        attr = safe_str(row.get("Attribute_Name", ""))
        val = safe_str(row.get("Attribute_Value", ""))
        if cid and attr:
            conn_attrs[cid][attr] = val

    for cid, attrs in conn_attrs.items():
        wire_color = attrs.get("Wire_Color", "")
        polarity = attrs.get("Polarity", "")

        # I24a: GNYE → PE
        if wire_color == "GNYE" and polarity and polarity != "PE":
            result.add_error("Connection_Data",
                f"I24: Connection '{cid}': Wire_Color=GNYE requires Polarity=PE, got '{polarity}'")

        # I24b: BU + (230V_AC or 400V_AC) → N
        if wire_color == "BU":
            voltage_level = attrs.get("Voltage_Level", "")
            if voltage_level in ("230V_AC", "400V_AC"):
                if polarity and polarity != "N":
                    result.add_error("Connection_Data",
                        f"I24: Connection '{cid}': Wire_Color=BU with {voltage_level} requires Polarity=N, got '{polarity}'")


def audit_connection_data_attributes(wb, result):
    """Check Connection_Data.Attribute_Name values are valid per Attribute_Lookup (§3.4, I13)."""
    if "Attribute_Lookup" not in wb.sheetnames:
        return
    if "Connection_Data" not in wb.sheetnames:
        return

    # Build allowed attribute names per scope
    al_ws = wb["Attribute_Lookup"]
    al_rows, _ = read_sheet_as_dicts(al_ws)
    allowed_by_scope = defaultdict(set)
    for row in al_rows:
        scope = safe_str(row.get("Scope", ""))
        attr = safe_str(row.get("Attribute_Name", ""))
        if scope and attr:
            allowed_by_scope[scope].add(attr)

    # Check Document_Data
    for sheet_name, scope in [("Document_Data", "Document"), ("Element_Data", "Element"),
                               ("RepresentedItem_Data", "RepresentedItem"), ("Connection_Data", "Connection")]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows, _ = read_sheet_as_dicts(ws)
        allowed = allowed_by_scope.get(scope, set())

        if not allowed:
            continue  # skip if Attribute_Lookup is incomplete

        for row in rows:
            attr = safe_str(row.get("Attribute_Name", ""))
            if attr and allowed and attr not in allowed:
                result.add_warning(sheet_name,
                    f"I13: Attribute_Name '{attr}' not in Attribute_Lookup for Scope={scope}")


def audit_document_data_required(wb, result):
    """Check document-level required attributes are present."""
    if "Document_Data" not in wb.sheetnames:
        return

    ws = wb["Document_Data"]
    rows, _ = read_sheet_as_dicts(ws)

    attrs_found = set()
    for row in rows:
        attr = safe_str(row.get("Attribute_Name", ""))
        val = safe_str(row.get("Attribute_Value", ""))
        if attr and val:
            attrs_found.add(attr)

    # Source_Format is mandatory (I29)
    if "Source_Format" not in attrs_found:
        result.add_error("Document_Data", "Missing required attribute: Source_Format")

    # Project_Name is required for AG-qualification
    if "Project_Name" not in attrs_found:
        result.add_warning("Document_Data", "Project_Name attribute missing (required for AG1 aggregation)")


def audit_empty_sheets(wb, result):
    """Flag sheets that are completely empty (no data rows)."""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        n_rows = _count_data_rows(ws)
        if n_rows == 0 and sheet_name in MANDATORY_SHEETS:
            # Some sheets may legitimately have no data (e.g. Layer_ID for simple docs)
            if sheet_name in ("Layer_ID",):
                result.add_info(sheet_name, "No data rows (may be expected for simple documents)")
            else:
                result.add_warning(sheet_name, f"No data rows in mandatory sheet")


def audit_fully(filepath):
    """Full audit of one exported file."""
    filename = os.path.basename(filepath)
    result = AuditResult(filename)

    wb = load_workbook_safe(filepath)
    if wb is None:
        result.add_error("LOAD", f"Failed to load workbook: {filepath}")
        return result

    # 1. Sheet structure
    audit_sheet_structure(wb, result)

    # 2. Document_ID uniqueness
    if "Document_ID" in wb.sheetnames:
        audit_document_id_uniqueness(wb, result)

    # 3. Schema_Metadata check
    audit_schema_metadata(wb, result)

    # 4. Header and ID format checks
    enum_values = {}
    for sheet_name in MANDATORY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        audit_headers(sheet_name, ws, result)
        audit_id_formats(sheet_name, ws, result)

        # Extract Enum_Lookup values for enum validation
        if sheet_name == "Enum_Lookup":
            enum_values = audit_enum_lookup_sheet(sheet_name, ws, result)

    # 5. Enum value validation (need Enum_Lookup first)
    for sheet_name in MANDATORY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        audit_enum_values(sheet_name, ws, result, enum_values)

    # 6. FK integrity
    audit_fk_integrity(wb, result)

    # 7. Fill rates
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        audit_fill_rate(sheet_name, ws, result)

    # 8. Source format (I29)
    audit_source_format(wb, result)

    # 9. Source completeness (I12)
    audit_i12_source_completeness(wb, result)

    # 10. Semantic checks
    audit_semantic_correctness(wb, result)

    # 11. Connection data attribute validation (I13)
    audit_connection_data_attributes(wb, result)

    # 12. Required document data
    audit_document_data_required(wb, result)

    # 13. Empty mandatory sheets
    audit_empty_sheets(wb, result)

    wb.close()
    return result


def main():
    all_files = []
    for root, dirs, files in os.walk(EXPORT_DIR):
        for f in files:
            if f.endswith('.xlsx') and not f.startswith('~$'):
                all_files.append(os.path.join(root, f))

    print(f"Found {len(all_files)} Excel files to audit in {EXPORT_DIR}")
    print()

    all_results = []
    for filepath in sorted(all_files):
        result = audit_fully(filepath)
        all_results.append(result)
        print(result.summary())

    # Overall summary
    total_errors = sum(len(r.errors) for r in all_results)
    total_warnings = sum(len(r.warnings) for r in all_results)

    print(f"\n{'='*80}")
    print(f"OVERALL SUMMARY")
    print(f"{'='*80}")
    print(f"  Files audited: {len(all_results)}")
    print(f"  Total ERRORS:  {total_errors}")
    print(f"  Total WARNINGS:{total_warnings}")

    # Group errors by type
    error_types = defaultdict(list)
    for r in all_results:
        for e in r.errors:
            # Extract sheet name
            match = re.match(r"\[(.*?)\] (.*)", e)
            sheet = match.group(1) if match else "UNKNOWN"
            desc = match.group(2) if match else e
            error_types[sheet].append((r.filename, desc))

    print(f"\nERRORS BY SHEET:")
    for sheet in sorted(error_types.keys()):
        items = error_types[sheet]
        print(f"  {sheet}: {len(items)} errors")
        for fname, desc in items[:5]:
            print(f"    [{fname[:50]}] {desc[:100]}")
        if len(items) > 5:
            print(f"    ... and {len(items)-5} more")

    # Fill rate summary
    print(f"\nFILL RATE SUMMARY (showing low-fill sheets < 80%):")
    for r in all_results:
        print(f"\n  {r.filename[:60]}:")
        for sheet, stats in sorted(r.fill_stats.items()):
            if stats.get("fill_rate", 100) < 80:
                print(f"    {sheet}: {stats['fill_rate']}% ({stats.get('rows',0)} rows × {stats.get('cols',0)} cols)")

    # Save detailed results
    output = {
        "audit_timestamp": datetime.now().isoformat(),
        "files_audited": len(all_results),
        "total_errors": total_errors,
        "total_warnings": total_warnings,
        "results": []
    }
    for r in all_results:
        output["results"].append({
            "filename": r.filename,
            "errors": r.errors,
            "warnings": r.warnings,
            "info": r.info,
            "fill_stats": {k: v for k, v in r.fill_stats.items()},
        })

    outpath = os.path.join(os.path.dirname(__file__), "..", "tests", "aio_spec_audit.json")
    with open(outpath, "w") as f:
        json.dump(output, f, indent=2, default=str)
    print(f"\nDetailed audit saved to: {outpath}")

    # Return non-zero if errors found
    return 1 if total_errors > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
