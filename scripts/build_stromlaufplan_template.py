"""Build empty Stromlaufplan_template.xlsx from the golden sample.

Copies sheet structure, headers (rows 1-2), column widths, and base styles
from Standardized_Stromlaufplan.xlsx, clears all data rows (row 3+).
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
GOLDEN = REPO_ROOT / "data" / "examples" / "Standardized_Stromlaufplan.xlsx"
TEMPLATE = REPO_ROOT / "data" / "templates" / "Stromlaufplan_template.xlsx"

EXPECTED_SHEETS = [
    "Document_ID",
    "Document_Data",
    "Revision_Data",
    "Layer_ID",
    "Object_ID",
    "Element_ID",
    "Element_Classification",
    "Element_Data",
    "Connection_Data",
]


def build() -> Path:
    if not GOLDEN.is_file():
        raise FileNotFoundError(f"Golden sample not found: {GOLDEN}")

    TEMPLATE.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(GOLDEN), str(TEMPLATE))

    wb = openpyxl.load_workbook(str(TEMPLATE))

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row is None or ws.max_row < 3:
            continue
        for row_idx in range(3, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).value = None

    wb.save(str(TEMPLATE))
    print(f"Built empty Stromlaufplan template: {TEMPLATE}")

    # Verify
    wb2 = openpyxl.load_workbook(str(TEMPLATE))
    actual_sheets = wb2.sheetnames
    assert actual_sheets == EXPECTED_SHEETS, f"Sheet mismatch: {actual_sheets}"

    for name in EXPECTED_SHEETS:
        ws = wb2[name]
        assert ws.max_row is not None and ws.max_row >= 2, f"{name}: missing header rows"
        for row_idx in range(3, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                assert val is None, f"{name} row {row_idx} col {col_idx}: expected None, got {val!r}"

    wb2.close()
    print("Verification passed: 9 sheets, headers intact, data rows empty")
    return TEMPLATE


if __name__ == "__main__":
    build()
