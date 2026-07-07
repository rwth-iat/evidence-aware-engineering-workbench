"""Build Datasheet_template.xlsx with standardized column headers.

Run once:
    python scripts/build_datasheet_template.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
TARGET_DIR = REPO_ROOT / "data" / "templates"
TARGET_FILE = TARGET_DIR / "Datasheet_template.xlsx"

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
LEGEND_FONT = Font(italic=True, size=10, color="555555")
LEGEND_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

SHEETS: dict[str, list[tuple[str, str]]] = {
    "Document_ID": [
        ("Index", "1-based row index within this sheet"),
        ("Document_ID", "Unique document identifier"),
        ("Document", "Source file name"),
        ("SemanticID", "Semantic identifier"),
    ],
    "Document_Data": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Document_Type", "e.g. Stellgeraetedatenblatt, Typenblatt"),
        ("Vendor", "Manufacturer name (SAMSON, E+H, ...)"),
        ("Vendor_Document_Number", "Vendor's own document number"),
        ("Revision", "Document revision"),
        ("Date", "Document date (YYYY-MM-DD)"),
        ("Norm", "Applicable standard (e.g. IEC 60534-7)"),
        ("SemanticID", "Semantic identifier"),
    ],
    "Device_ID": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Device_ID", "Unique device instance identifier"),
        ("AKZ", "Original AKZ string"),
        ("AKZ_Canonical", "Normalised canonical AKZ"),
        ("TagName", "Device tag name"),
        ("Manufacturer", "Manufacturer name"),
        ("Model", "Model / type designation"),
        ("Serial_Number", "Serial number"),
        ("ECLASS_IRDI", "ECLASS IRDI if available"),
        ("SemanticID", "Semantic identifier"),
        ("Source_Vendor", "Detected vendor"),
        ("Source_File", "Source document path"),
        ("Confidence", "Extraction confidence 0.0–1.0"),
        ("LLM_Reasoning", "LLM reasoning if used"),
        # SM_Nameplate
        ("OrderCode", "Order code / part number"),
        ("UniqueFacilityId", "Unique facility identifier per ESPR"),
        # SM_InstrumentListEntry
        ("EntryId", "Entry identifier for cross-document matching"),
        ("CanonicalTag", "Canonical device tag"),
        ("PresenceStatus", "Presence status in source documents"),
        # SM_Classification
        ("ClassSystem", "Classification system name"),
        ("ClassCode", "Classification code"),
        ("ClassName", "Human-readable class name"),
        # SM_SourceDocumentation
        ("SourceDocId", "Source document identifier"),
        ("SourceLocator", "Source document locator / page reference"),
    ],
    "Device_Classification": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Device_ID", "FK → Device_ID"),
        ("Class_System", "ECLASS | ETIM | proprietary"),
        ("Class_Code", "Classification code"),
        ("Class_Name", "Human-readable class name"),
        ("SemanticID_Class", "Semantic identifier for the class"),
        ("Confidence", "Extraction confidence 0.0–1.0"),
    ],
    "Process_Attributes": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Device_ID", "FK → Device_ID"),
        ("Attribute_Key", "Standard attribute key"),
        ("Attribute_Name", "Standard attribute name"),
        ("Attribute_Value", "Attribute value"),
        ("Attribute_Unit", "Unit of measurement"),
        ("Attribute_Source", "Source text span"),
        ("SemanticID_Attribute", "Semantic identifier"),
        ("Mapping_Confidence", "Field-to-column mapping confidence 0.0–1.0"),
        ("LLM_Reasoning", "LLM reasoning if LLM mapped this field"),
    ],
    "Technical_Attributes": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Device_ID", "FK → Device_ID"),
        ("Attribute_Key", "Standard attribute key"),
        ("Attribute_Name", "Standard attribute name"),
        ("Attribute_Value", "Attribute value"),
        ("Attribute_Unit", "Unit of measurement"),
        ("Attribute_Source", "Source text span"),
        ("SemanticID_Attribute", "Semantic identifier"),
        ("Mapping_Confidence", "Field-to-column mapping confidence 0.0–1.0"),
        ("LLM_Reasoning", "LLM reasoning"),
    ],
    "Geometric_Attributes": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Device_ID", "FK → Device_ID"),
        ("Attribute_Key", "Standard attribute key"),
        ("Attribute_Name", "Standard attribute name"),
        ("Attribute_Value", "Attribute value"),
        ("Attribute_Unit", "Unit of measurement"),
        ("Attribute_Source", "Source text span"),
        ("SemanticID_Attribute", "Semantic identifier"),
        ("Mapping_Confidence", "Field-to-column mapping confidence 0.0–1.0"),
        ("LLM_Reasoning", "LLM reasoning"),
    ],
    "Connection_Attributes": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Device_ID", "FK → Device_ID"),
        ("Connection_Type", "e.g. process, electrical, pneumatic"),
        ("Signal_Type", "4-20mA | HART | Profibus | ..."),
        ("Power_Supply", "24VDC | 230VAC | loop-powered | ..."),
        ("Bus_Protocol", "Fieldbus protocol if applicable"),
        ("Attribute_Source", "Source text span"),
        ("SemanticID", "Semantic identifier"),
        ("Confidence", "Extraction confidence 0.0–1.0"),
    ],
    "Manufacturer_Specific": [
        ("Index", "1-based row index"),
        ("Document_ID", "FK → Document_ID"),
        ("Device_ID", "FK → Device_ID"),
        ("Attribute_Key", "Original manufacturer field name"),
        ("Attribute_Name", "Translated / standardized name"),
        ("Attribute_Value", "Attribute value"),
        ("Attribute_Source", "Source text span"),
    ],
}


def build_template() -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, columns in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, (header, _legend) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, (_header, legend) in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col_idx, value=legend)
            cell.font = LEGEND_FONT
            cell.fill = LEGEND_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 24
        ws.freeze_panes = f"{get_column_letter(1)}3"

    TARGET_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(TARGET_FILE)
    print(f"wrote {TARGET_FILE.relative_to(REPO_ROOT)} ({len(SHEETS)} sheets)")


if __name__ == "__main__":
    build_template()
