#!/usr/bin/env python3
"""
Build 3D pipeline assembly template from CSV + FreeCAD Document.xml.

Primary data source:
  - final_result.FCStd / Document.xml → Link objects with Label, LinkPlacement (pos + quaternion), LinkedObject
  - final_result_components.csv → assembly sequence (BOM order)

Outputs:
  - data/templates/Assembly_3D_template.xlsx         (blank template)
  - data/templates/Assembly_3D_template_filled.xlsx  (populated with current data)
"""

from __future__ import annotations

import csv
import math
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────
REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = REPO_ROOT / "data" / "templates"
PIPELINE_DIR = REPO_ROOT / "Documents" / "Piping Diagram"
CSV_PATH = PIPELINE_DIR / "final_result_components.csv"
FCSTD_PATH = PIPELINE_DIR / "final_result.FCStd"
IFC_FLAT_PATH = PIPELINE_DIR / "IFC_FLAT_EXPORT.ifc"

# ── Header style ───────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", size=11, bold=True)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── Sheet definitions ──────────────────────────────────────────────────────
SHEET_DEFS: dict[str, list[str]] = {
    "Part_Library": [
        "part_id",
        "part_ref_name",
        "part_type",
        "fcstd_link_name",
        "source_fcstd_file",
        "ifc_parts_library",
        "ifc_part_name",
        "instance_count",
        "ifc_coverage",
        "description",
    ],
    "Assembly_Steps": [
        "assembly_no",
        "label",
        "type",
        "fcstd_link_name",
        "source_fcstd_file",
        "ifc_parts_library",
        "ifc_part_name",
        "pos_x_mm",
        "pos_y_mm",
        "pos_z_mm",
        "rot_qw",
        "rot_qx",
        "rot_qy",
        "rot_qz",
        "rot_axis_x",
        "rot_axis_y",
        "rot_axis_z",
        "rot_angle_deg",
        "euler_roll_deg",
        "euler_pitch_deg",
        "euler_yaw_deg",
        "mapping_confidence",
        "notes",
    ],
    "Connection_Topology": [
        "connection_id",
        "from_assembly_no",
        "from_label",
        "from_port",
        "to_assembly_no",
        "to_label",
        "to_port",
        "connection_type",
        "from_pos_x",
        "from_pos_y",
        "from_pos_z",
        "to_pos_x",
        "to_pos_y",
        "to_pos_z",
        "euclidean_distance_mm",
    ],
    "Connection_Angles": [
        "connection_id",
        "from_label",
        "to_label",
        "from_forward_x",
        "from_forward_y",
        "from_forward_z",
        "to_forward_x",
        "to_forward_y",
        "to_forward_z",
        "connection_vector_x",
        "connection_vector_y",
        "connection_vector_z",
        "bend_angle_deg",
        "torsion_angle_deg",
        "from_to_distance_mm",
    ],
    "Position_Data": [
        "fcstd_link_name",
        "label",
        "type",
        "source_fcstd_file",
        "ifc_parts_library",
        "pos_x_mm",
        "pos_y_mm",
        "pos_z_mm",
        "qw", "qx", "qy", "qz",
        "axis_x", "axis_y", "axis_z",
        "angle_deg",
        "euler_roll_deg",
        "euler_pitch_deg",
        "euler_yaw_deg",
    ],
    "Assembly_Tree": [
        "tree_id",
        "parent_label",
        "child_label",
        "aggregation_type",
        "assembly_level",
        "notes",
    ],
}


# ═══════════════════════════════════════════════════════════════════════════
# FreeCAD Document.xml parser
# ═══════════════════════════════════════════════════════════════════════════

def parse_freecad_links(fcstd_path: Path) -> list[dict[str, Any]]:
    """Extract all App::Link objects with Label, LinkPlacement, and LinkedObject.

    FreeCAD Document.xml structure:
    1. <Objects> — self-closing tags (name + type + id). Type is ONLY here.
    2. <ObjectData> — full property details per object. Properties are ONLY here.
    We cross-reference: Objects[name] → type, ObjectData[name] → properties.
    """
    with zipfile.ZipFile(fcstd_path, "r") as z:
        with z.open("Document.xml") as f:
            tree = ET.parse(f)
            root = tree.getroot()

    # Step 1: Build name → type mapping from <Objects>
    name_to_type: dict[str, str] = {}
    objects_section = root.find("Objects")
    if objects_section is not None:
        for obj in objects_section.findall("Object"):
            name = obj.get("name", "")
            obj_type = obj.get("type", "")
            if name:
                name_to_type[name] = obj_type

    # Step 2: Parse <ObjectData> for properties, cross-ref type from name_to_type
    links: list[dict[str, Any]] = []
    object_data = root.find("ObjectData")
    if object_data is None:
        print("WARNING: No ObjectData section found in Document.xml")
        return links

    for obj in object_data.findall("Object"):
        obj_name = obj.get("name", "")
        obj_type = name_to_type.get(obj_name, "")
        if obj_type != "App::Link":
            continue

        props = obj.find("Properties")
        if props is None:
            continue

        entry: dict[str, Any] = {
            "fcstd_link_name": obj_name,
            "label": "",
            "source_model_file": "",
            "xlink_internal_name": "",  # the base part name inside the linked file
            "pos_x": 0.0,
            "pos_y": 0.0,
            "pos_z": 0.0,
            "qw": 0.0, "qx": 0.0, "qy": 0.0, "qz": 1.0,
            "axis_x": 0.0, "axis_y": 0.0, "axis_z": 1.0,
            "angle_rad": 0.0,
        }

        for prop in props.findall("Property"):
            pname = prop.get("name", "")

            if pname == "Label":
                string_el = prop.find("String")
                if string_el is not None:
                    entry["label"] = string_el.get("value", "")

            elif pname == "LinkedObject":
                xlink = prop.find("XLink")
                if xlink is not None:
                    entry["source_model_file"] = xlink.get("file", "")
                    entry["xlink_internal_name"] = xlink.get("name", "")

            elif pname == "LinkPlacement":
                placement = prop.find("PropertyPlacement")
                if placement is not None:
                    entry["pos_x"] = float(placement.get("Px", "0"))
                    entry["pos_y"] = float(placement.get("Py", "0"))
                    entry["pos_z"] = float(placement.get("Pz", "0"))
                    entry["qw"] = float(placement.get("Q0", "0"))
                    entry["qx"] = float(placement.get("Q1", "0"))
                    entry["qy"] = float(placement.get("Q2", "0"))
                    entry["qz"] = float(placement.get("Q3", "1"))
                    entry["axis_x"] = float(placement.get("Ox", "0"))
                    entry["axis_y"] = float(placement.get("Oy", "0"))
                    entry["axis_z"] = float(placement.get("Oz", "1"))
                    entry["angle_rad"] = float(placement.get("A", "0"))

        if entry["label"]:
            links.append(entry)

    return links


def build_xlink_resolve_map(fcstd_path: Path) -> dict[str, str]:
    """Build a map: link_name@file → ultimate_base_part_name.

    For parts from 3D_pipeline.FCStd, the Link points to another file (Unnamed8/Unnamed9).
    We need to resolve this chain to find the actual base part name for IFC lookup.
    """
    resolve: dict[str, str] = {}

    with zipfile.ZipFile(fcstd_path, "r") as z:
        with z.open("Document.xml") as f:
            tree = ET.parse(f)
            root = tree.getroot()

    name_to_type: dict[str, str] = {}
    objects_section = root.find("Objects")
    if objects_section is not None:
        for obj in objects_section.findall("Object"):
            name = obj.get("name", "")
            if name:
                name_to_type[name] = obj.get("type", "")

    object_data = root.find("ObjectData")
    if object_data is None:
        return resolve

    for obj in object_data.findall("Object"):
        obj_type = name_to_type.get(obj.get("name", ""), "")
        if obj_type != "App::Link":
            continue
        props = obj.find("Properties")
        if props is None:
            continue
        for prop in props.findall("Property"):
            if prop.get("name") == "LinkedObject":
                xlink = prop.find("XLink")
                if xlink is not None:
                    resolve[obj.get("name", "")] = xlink.get("name", "")

    return resolve


# ═══════════════════════════════════════════════════════════════════════════
# Quaternion utilities
# ═══════════════════════════════════════════════════════════════════════════

def quat_to_euler(qw: float, qx: float, qy: float, qz: float) -> tuple[float, float, float]:
    """Convert quaternion to intrinsic ZYX Euler angles (roll, pitch, yaw) in degrees."""
    # Normalize
    norm = math.sqrt(qw * qw + qx * qx + qy * qy + qz * qz)
    if norm < 1e-15:
        return (0.0, 0.0, 0.0)
    qw, qx, qy, qz = qw / norm, qx / norm, qy / norm, qz / norm

    # Rotation matrix from quaternion
    # R = [[1-2(qy²+qz²), 2(qxqy-qwqz), 2(qxqz+qwqy)],
    #      [2(qxqy+qwqz), 1-2(qx²+qz²), 2(qyqz-qwqx)],
    #      [2(qxqz-qwqy), 2(qyqz+qwqx), 1-2(qx²+qy²)]]

    # Extract Euler (ZYX intrinsic = roll around X, pitch around Y, yaw around Z)
    # R = Rz(yaw) * Ry(pitch) * Rx(roll)
    r20 = 2 * (qx * qz - qw * qy)
    r21 = 2 * (qy * qz + qw * qx)
    r22 = 1 - 2 * (qx * qx + qy * qy)
    r10 = 2 * (qx * qy + qw * qz)
    r00 = 1 - 2 * (qy * qy + qz * qz)

    pitch = -math.asin(max(-1.0, min(1.0, r20)))
    cos_pitch = math.cos(pitch)

    if abs(cos_pitch) > 1e-10:
        roll = math.atan2(r21 / cos_pitch, r22 / cos_pitch)
        yaw = math.atan2(r10 / cos_pitch, r00 / cos_pitch)
    else:
        # Gimbal lock
        roll = 0.0
        yaw = math.atan2(-(2 * (qy * qz - qw * qx)), 1 - 2 * (qx * qx + qz * qz))

    return (math.degrees(roll), math.degrees(pitch), math.degrees(yaw))


def quat_rotate_vector(
    qw: float, qx: float, qy: float, qz: float,
    vx: float, vy: float, vz: float,
) -> tuple[float, float, float]:
    """Rotate vector v by quaternion q. Returns q * v * q_conjugate."""
    # q * v (quaternion multiply where v is pure quaternion (0, vx, vy, vz))
    tw = -qx * vx - qy * vy - qz * vz
    tx = qw * vx + qy * vz - qz * vy
    ty = qw * vy + qz * vx - qx * vz
    tz = qw * vz + qx * vy - qy * vx
    # q * v * q_conjugate (conjugate = (qw, -qx, -qy, -qz))
    rx = tx * qw + tw * (-qx) + ty * (-qz) - tz * (-qy)
    ry = ty * qw + tw * (-qy) + tz * (-qx) - tx * (-qz)
    rz = tz * qw + tw * (-qz) + tx * (-qy) - ty * (-qx)
    return (rx, ry, rz)


def compute_forward_direction(
    qw: float, qx: float, qy: float, qz: float,
    axis_x: float, axis_y: float, axis_z: float,
) -> tuple[float, float, float]:
    """Compute the component's forward direction (local Z axis rotated by quaternion).
    If axis-angle rotation is non-zero, rotate local Z=(0,0,1) by that.
    In FreeCAD, Placement rotates around the given axis by the given angle.
    """
    angle = 2 * math.acos(max(-1.0, min(1.0, qw)))
    sin_half = math.sqrt(1 - qw * qw)

    # The axis of rotation is encoded in the quaternion
    if sin_half > 1e-15:
        ax = qx / sin_half
        ay = qy / sin_half
        az = qz / sin_half
    else:
        ax, ay, az = 0.0, 0.0, 1.0

    # Forward direction = local Z axis = (0, 0, 1) rotated by quaternion
    fwd = quat_rotate_vector(qw, qx, qy, qz, 0.0, 0.0, 1.0)
    return fwd


# ═══════════════════════════════════════════════════════════════════════════
# IFC parts index
# ═══════════════════════════════════════════════════════════════════════════

def build_ifc_parts_index(pipeline_dir: Path) -> dict[str, tuple[str, str]]:
    """Build index: fcstd_part_name → (ifc_file_name, ifc_global_id).

    Scans all small IFC files in the pipeline directory.
    """
    index: dict[str, tuple[str, str]] = {}

    for ifc_path in sorted(pipeline_dir.glob("*.ifc")):
        # Skip the huge flat/geometry exports
        if "FLAT_EXPORT" in ifc_path.name or "from_step" in ifc_path.name or "geometry" in ifc_path.name:
            continue
        try:
            import ifcopenshell
            f = ifcopenshell.open(str(ifc_path))
            for entity in f.by_type("IfcBuildingElementProxy"):
                name = entity.Name
                if name and name not in index:
                    index[name] = (ifc_path.name, entity.GlobalId)
        except Exception:
            pass

    return index


# ═══════════════════════════════════════════════════════════════════════════
# Assembly build logic
# ═══════════════════════════════════════════════════════════════════════════

def classify_link_type(label: str) -> str:
    """Classify a FreeCAD link label into tank / elbow / pipeline / blackbox."""
    if label.startswith("tank"):
        return "tank"
    if label.startswith("elbow"):
        return "elbow"
    if label.startswith("pipeline"):
        return "pipeline"
    if label.startswith("blackbox"):
        return "blackbox"
    # Fallback: try to match from name
    if "tank" in label.lower():
        return "tank"
    if "elbow" in label.lower():
        return "elbow"
    if "pipeline" in label.lower() or "pipe" in label.lower():
        return "pipeline"
    return "body"


def build_assembly_steps(
    csv_rows: list[dict[str, str]],
    fc_links: list[dict[str, Any]],
    ifc_index: dict[str, tuple[str, str]] | None = None,
) -> list[dict[str, Any]]:
    """Match CSV rows to FreeCAD Link objects by Label.

    The FreeCAD Label property corresponds directly to CSV label.
    """
    # Build label → FreeCAD link lookup
    by_label: dict[str, dict[str, Any]] = {}
    for link in fc_links:
        lbl = link["label"]
        if lbl and lbl not in by_label:
            by_label[lbl] = link
        elif lbl:
            # Duplicate label — keep the one with non-zero placement
            existing = by_label[lbl]
            if abs(existing["pos_x"]) < 0.001 and abs(existing["pos_y"]) < 0.001 and abs(existing["pos_z"]) < 0.001:
                by_label[lbl] = link

    steps: list[dict[str, Any]] = []
    unmatched_fc = set(by_label.keys())
    unmatched_csv: list[int] = []  # indices in csv_rows

    ifc_idx = ifc_index or {}
    _xlink_cache: dict[str, dict[str, str]] = {}

    def _get_xlink_map(src_file: str) -> dict[str, str]:
        """Get the xlink resolve map for a source FCStd file (cached)."""
        if src_file not in _xlink_cache:
            fpath = PIPELINE_DIR / src_file
            if fpath.exists() and src_file != "Unnamed9.FCStd":
                _xlink_cache[src_file] = build_xlink_resolve_map(fpath)
            else:
                _xlink_cache[src_file] = {}
        return _xlink_cache[src_file]

    def _lookup_ifc(fcstd_name: str, xlink_internal: str, source_file: str) -> tuple[str, str, str]:
        """Look up a part in the IFC index. Try multiple name variants.
        Returns (ifc_file, ifc_part_name, note).
        """
        # Try 1: exact fcstd_link_name
        if fcstd_name in ifc_idx:
            return (ifc_idx[fcstd_name][0], fcstd_name, "")

        # Try 2: xlink internal name (base part name in linked file)
        if xlink_internal and xlink_internal in ifc_idx:
            return (ifc_idx[xlink_internal][0], xlink_internal, "")

        # Try 3: resolve through 3D_pipeline link chain
        xlink_map = _get_xlink_map(source_file)
        base_name = xlink_map.get(fcstd_name, "")
        if base_name and base_name in ifc_idx:
            return (ifc_idx[base_name][0], base_name, "")

        # Not found
        return ("", "", f"! NO IFC — source: {source_file}")

    def _make_matched_step(row: dict[str, str], link: dict[str, Any], extra_notes: str = "") -> dict[str, Any]:
        euler = quat_to_euler(link["qw"], link["qx"], link["qy"], link["qz"])
        fwd = compute_forward_direction(
            link["qw"], link["qx"], link["qy"], link["qz"],
            link["axis_x"], link["axis_y"], link["axis_z"],
        )
        fcstd_name = link["fcstd_link_name"]
        xlink_internal = link.get("xlink_internal_name", "")
        source_file = link.get("source_model_file", "")

        ifc_file, ifc_part, ifc_note = _lookup_ifc(fcstd_name, xlink_internal, source_file)

        notes = extra_notes + ("; " + ifc_note if ifc_note and extra_notes else ifc_note)

        return {
            "assembly_no": row["assembly_no"],
            "label": row["label"],
            "type": row["type"],
            "fcstd_link_name": fcstd_name,
            "source_fcstd_file": link["source_model_file"],
            "ifc_parts_library": ifc_file,
            "ifc_part_name": ifc_part,
            "pos_x_mm": round(link["pos_x"], 6),
            "pos_y_mm": round(link["pos_y"], 6),
            "pos_z_mm": round(link["pos_z"], 6),
            "rot_qw": round(link["qw"], 9),
            "rot_qx": round(link["qx"], 9),
            "rot_qy": round(link["qy"], 9),
            "rot_qz": round(link["qz"], 9),
            "rot_axis_x": round(link["axis_x"], 6),
            "rot_axis_y": round(link["axis_y"], 6),
            "rot_axis_z": round(link["axis_z"], 6),
            "rot_angle_deg": round(math.degrees(link["angle_rad"]), 4),
            "euler_roll_deg": round(euler[0], 4),
            "euler_pitch_deg": round(euler[1], 4),
            "euler_yaw_deg": round(euler[2], 4),
            "forward_x": round(fwd[0], 6),
            "forward_y": round(fwd[1], 6),
            "forward_z": round(fwd[2], 6),
            "mapping_confidence": "matched",
            "notes": notes.strip(),
        }

    def _make_unmatched_step(row: dict[str, str]) -> dict[str, Any]:
        return {
            "assembly_no": row["assembly_no"],
            "label": row["label"],
            "type": row["type"],
            "fcstd_link_name": "",
            "source_fcstd_file": "",
            "ifc_parts_library": "",
            "ifc_part_name": "",
            "pos_x_mm": None, "pos_y_mm": None, "pos_z_mm": None,
            "rot_qw": None, "rot_qx": None, "rot_qy": None, "rot_qz": None,
            "rot_axis_x": None, "rot_axis_y": None, "rot_axis_z": None,
            "rot_angle_deg": None,
            "euler_roll_deg": None, "euler_pitch_deg": None, "euler_yaw_deg": None,
            "forward_x": None, "forward_y": None, "forward_z": None,
            "mapping_confidence": "unmatched",
            "notes": "no FreeCAD Link with matching label",
        }

    # Pass 1: direct label match
    for csv_idx, row in enumerate(csv_rows):
        label = row["label"]
        link = by_label.get(label)
        if link:
            unmatched_fc.discard(label)
            steps.append(_make_matched_step(row, link))
        else:
            unmatched_csv.append(csv_idx)
            steps.append(None)  # placeholder

    # Pass 2: match remaining by type + assembly position
    # Group unmatched FC links by type
    fc_by_type: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for lbl in sorted(unmatched_fc):
        link = by_label[lbl]
        ctype = classify_link_type(lbl)
        fc_by_type[ctype].append(link)

    for idx in unmatched_csv:
        row = csv_rows[idx]
        ctype = row["type"]
        pool = fc_by_type.get(ctype, [])
        if pool:
            link = pool.pop(0)
            # Remove from unmatched_fc so it won't be appended later
            unmatched_fc.discard(link["label"])
            steps[idx] = _make_matched_step(row, link, extra_notes="matched by type+position (label numbering differs)")
            steps[idx]["mapping_confidence"] = "matched_by_type"
        else:
            steps[idx] = _make_unmatched_step(row)

    # Append remaining unmatched FreeCAD objects
    for lbl in sorted(unmatched_fc):
        link = by_label[lbl]
        euler = quat_to_euler(link["qw"], link["qx"], link["qy"], link["qz"])
        fcstd_name = link["fcstd_link_name"]
        ifc_info = ifc_idx.get(fcstd_name)
        steps.append({
            "assembly_no": "",
            "label": lbl,
            "type": classify_link_type(lbl),
            "fcstd_link_name": fcstd_name,
            "source_fcstd_file": link["source_model_file"],
            "ifc_parts_library": ifc_info[0] if ifc_info else "",
            "ifc_part_name": fcstd_name,
            "pos_x_mm": round(link["pos_x"], 6),
            "pos_y_mm": round(link["pos_y"], 6),
            "pos_z_mm": round(link["pos_z"], 6),
            "rot_qw": round(link["qw"], 9),
            "rot_qx": round(link["qx"], 9),
            "rot_qy": round(link["qy"], 9),
            "rot_qz": round(link["qz"], 9),
            "rot_axis_x": round(link["axis_x"], 6),
            "rot_axis_y": round(link["axis_y"], 6),
            "rot_axis_z": round(link["axis_z"], 6),
            "rot_angle_deg": round(math.degrees(link["angle_rad"]), 4),
            "euler_roll_deg": round(euler[0], 4),
            "euler_pitch_deg": round(euler[1], 4),
            "euler_yaw_deg": round(euler[2], 4),
            "forward_x": None, "forward_y": None, "forward_z": None,
            "mapping_confidence": "unassigned_fc",
            "notes": "FreeCAD Link with no CSV match",
        })

    return steps


def build_connections(
    assembly_steps: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build connection topology from assembly sequence."""
    connections: list[dict[str, Any]] = []
    conn_id = 0

    valid = [e for e in assembly_steps if e["assembly_no"] != "" and e["assembly_no"] is not None]
    valid.sort(key=lambda e: int(e["assembly_no"]))

    for i in range(len(valid) - 1):
        conn_id += 1
        frm = valid[i]
        to = valid[i + 1]

        dist = None
        if frm["pos_x_mm"] is not None and to["pos_x_mm"] is not None:
            dist = round(math.sqrt(
                (frm["pos_x_mm"] - to["pos_x_mm"]) ** 2
                + (frm["pos_y_mm"] - to["pos_y_mm"]) ** 2
                + (frm["pos_z_mm"] - to["pos_z_mm"]) ** 2
            ), 2)

        connections.append({
            "connection_id": conn_id,
            "from_assembly_no": frm["assembly_no"],
            "from_label": frm["label"],
            "from_port": "out",
            "to_assembly_no": to["assembly_no"],
            "to_label": to["label"],
            "to_port": "in",
            "connection_type": _guess_connection_type(frm, to),
            "from_pos_x": frm["pos_x_mm"],
            "from_pos_y": frm["pos_y_mm"],
            "from_pos_z": frm["pos_z_mm"],
            "to_pos_x": to["pos_x_mm"],
            "to_pos_y": to["pos_y_mm"],
            "to_pos_z": to["pos_z_mm"],
            "euclidean_distance_mm": dist,
        })

    return connections


def _guess_connection_type(frm: dict[str, Any], to: dict[str, Any]) -> str:
    f_type = frm["type"]
    t_type = to["type"]
    if "tank" in (f_type, t_type):
        return "flange"
    if "elbow" in (f_type, t_type) and "pipeline" in (f_type, t_type):
        return "weld_or_threaded"
    if "blackbox" in (f_type, t_type):
        return "flange_or_threaded"
    if f_type == "elbow" and t_type == "elbow":
        return "direct_fitting"
    if f_type == "pipeline" and t_type == "pipeline":
        return "weld_or_coupling"
    return "unknown"


def build_connection_angles(
    connections: list[dict[str, Any]],
    assembly_steps: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Calculate 3D angles using quaternion-derived forward directions."""
    by_label: dict[str, dict[str, Any]] = {}
    for e in assembly_steps:
        if e["label"]:
            by_label[e["label"]] = e

    angles: list[dict[str, Any]] = []
    for conn in connections:
        frm = by_label.get(conn["from_label"])
        to = by_label.get(conn["to_label"])
        if not frm or not to:
            continue

        # Get forward directions (rotated local Z)
        fwd_f = (frm.get("forward_x"), frm.get("forward_y"), frm.get("forward_z"))
        fwd_t = (to.get("forward_x"), to.get("forward_y"), to.get("forward_z"))

        # Connection vector
        cv = (0.0, 0.0, 0.0)
        n_cv = 0.0
        if frm.get("pos_x_mm") is not None and to.get("pos_x_mm") is not None:
            cv = (
                to["pos_x_mm"] - frm["pos_x_mm"],
                to["pos_y_mm"] - frm["pos_y_mm"],
                to["pos_z_mm"] - frm["pos_z_mm"],
            )
            n_cv = math.sqrt(sum(v * v for v in cv))
            if n_cv > 1e-12:
                cv = (cv[0] / n_cv, cv[1] / n_cv, cv[2] / n_cv)

        # Bend angle between forward directions
        bend = 0.0
        if fwd_f[0] is not None and fwd_t[0] is not None:
            dot = fwd_f[0] * fwd_t[0] + fwd_f[1] * fwd_t[1] + fwd_f[2] * fwd_t[2]
            dot = max(-1.0, min(1.0, dot))
            bend = math.degrees(math.acos(dot))

        # Torsion between connection vector and from_forward
        torsion = 0.0
        if n_cv > 1e-12 and fwd_f[0] is not None:
            dot_cv = fwd_f[0] * cv[0] + fwd_f[1] * cv[1] + fwd_f[2] * cv[2]
            dot_cv = max(-1.0, min(1.0, abs(dot_cv)))
            torsion = math.degrees(math.acos(dot_cv))

        angles.append({
            "connection_id": conn["connection_id"],
            "from_label": conn["from_label"],
            "to_label": conn["to_label"],
            "from_forward_x": round(fwd_f[0], 6) if fwd_f[0] is not None else None,
            "from_forward_y": round(fwd_f[1], 6) if fwd_f[1] is not None else None,
            "from_forward_z": round(fwd_f[2], 6) if fwd_f[2] is not None else None,
            "to_forward_x": round(fwd_t[0], 6) if fwd_t[0] is not None else None,
            "to_forward_y": round(fwd_t[1], 6) if fwd_t[1] is not None else None,
            "to_forward_z": round(fwd_t[2], 6) if fwd_t[2] is not None else None,
            "connection_vector_x": round(cv[0], 6),
            "connection_vector_y": round(cv[1], 6),
            "connection_vector_z": round(cv[2], 6),
            "bend_angle_deg": round(bend, 2),
            "torsion_angle_deg": round(torsion, 2),
            "from_to_distance_mm": round(n_cv, 2),
        })

    return angles


def build_part_library(
    assembly_steps: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build unique part library."""
    parts: dict[str, dict[str, Any]] = {}
    for step in assembly_steps:
        fc_name = step.get("fcstd_link_name") or ""
        if not fc_name or step["mapping_confidence"] == "unmatched":
            continue
        if fc_name not in parts:
            ifc_file = step.get("ifc_parts_library", "")
            parts[fc_name] = {
                "part_id": f"PART_{len(parts) + 1:04d}",
                "part_ref_name": fc_name,
                "part_type": step["type"],
                "fcstd_link_name": fc_name,
                "source_fcstd_file": step.get("source_fcstd_file", ""),
                "ifc_parts_library": ifc_file,
                "ifc_part_name": fc_name if ifc_file else "",
                "instance_count": 0,
                "ifc_coverage": "YES" if ifc_file else "NO — needs IFC export",
                "description": "",
            }
        parts[fc_name]["instance_count"] += 1
    return list(parts.values())


def build_assembly_tree(
    assembly_steps: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build assembly hierarchy tree."""
    tree: list[dict[str, Any]] = []
    tree_id = 0
    current_tank = None

    for step in assembly_steps:
        if step.get("type") == "tank":
            tree_id += 1
            current_tank = step["label"]
            tree.append({
                "tree_id": tree_id,
                "parent_label": None,
                "child_label": step["label"],
                "aggregation_type": "root_assembly",
                "assembly_level": "assembly",
                "notes": "tank assembly root",
            })
        elif current_tank and step.get("type") in ("pipeline", "elbow", "blackbox"):
            tree_id += 1
            tree.append({
                "tree_id": tree_id,
                "parent_label": current_tank,
                "child_label": step["label"],
                "aggregation_type": "connected_to",
                "assembly_level": "part",
                "notes": f"part of {current_tank} run",
            })

    return tree


def build_position_data(
    assembly_steps: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Extract position data for all matched components."""
    result = []
    for step in assembly_steps:
        if step["mapping_confidence"] not in ("matched", "unassigned_fc"):
            continue
        result.append({
            "fcstd_link_name": step.get("fcstd_link_name", ""),
            "label": step.get("label", ""),
            "type": step.get("type", ""),
            "source_fcstd_file": step.get("source_fcstd_file", ""),
            "ifc_parts_library": step.get("ifc_parts_library", ""),
            "pos_x_mm": step.get("pos_x_mm"),
            "pos_y_mm": step.get("pos_y_mm"),
            "pos_z_mm": step.get("pos_z_mm"),
            "qw": step.get("rot_qw"),
            "qx": step.get("rot_qx"),
            "qy": step.get("rot_qy"),
            "qz": step.get("rot_qz"),
            "axis_x": step.get("rot_axis_x"),
            "axis_y": step.get("rot_axis_y"),
            "axis_z": step.get("rot_axis_z"),
            "angle_deg": step.get("rot_angle_deg"),
            "euler_roll_deg": step.get("euler_roll_deg"),
            "euler_pitch_deg": step.get("euler_pitch_deg"),
            "euler_yaw_deg": step.get("euler_yaw_deg"),
        })
    return result


# ═══════════════════════════════════════════════════════════════════════════
# Excel Output
# ═══════════════════════════════════════════════════════════════════════════

def apply_header_style(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    headers: list[str],
) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_width(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, min(ws.max_row + 1, 50)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 30)


def write_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(title=sheet_name)
    apply_header_style(ws, headers)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
    auto_width(ws)


def create_blank_template(output_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, headers in SHEET_DEFS.items():
        ws = wb.create_sheet(title=sheet_name)
        apply_header_style(ws, headers)
        auto_width(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Blank template saved to: {output_path}")


def create_filled_template(output_path: Path, **sheets: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        headers = SHEET_DEFS.get(sheet_name, [])
        write_sheet(wb, sheet_name, headers, rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Filled template saved to: {output_path}")


# ═══════════════════════════════════════════════════════════════════════════
# Summary
# ═══════════════════════════════════════════════════════════════════════════

def print_summary(
    assembly_steps: list[dict[str, Any]],
    connections: list[dict[str, Any]],
    part_library: list[dict[str, Any]],
) -> None:
    matched = sum(1 for e in assembly_steps if e["mapping_confidence"] == "matched")
    unmatched = sum(1 for e in assembly_steps if e["mapping_confidence"] == "unmatched")
    unassigned = sum(1 for e in assembly_steps if e["mapping_confidence"] == "unassigned_fc")

    print(f"\n{'='*60}")
    print("Summary")
    print(f"{'='*60}")
    print(f"  Total assembly steps:       {len(assembly_steps)}")
    print(f"  Matched (label→Link):       {matched}")
    print(f"  CSV-unmatched (no Link):    {unmatched}")
    print(f"  FC-unassigned (no CSV):     {unassigned}")
    print(f"  Total connections derived:  {len(connections)}")
    print(f"  Unique parts in library:    {len(part_library)}")

    type_counts = defaultdict(int)
    for e in assembly_steps:
        type_counts[e["type"]] += 1
    print(f"\n  Component types:")
    for t, c in sorted(type_counts.items()):
        print(f"    {t}: {c}")

    # Distribution of bend angles
    all_bends = []
    for e in assembly_steps:
        if e.get("rot_angle_deg") is not None and e.get("rot_angle_deg", 0) > 0.1:
            all_bends.append(e["rot_angle_deg"])
    if all_bends:
        print(f"\n  Rotation angle (°) distribution:")
        print(f"    count={len(all_bends)}, min={min(all_bends):.1f}, "
              f"max={max(all_bends):.1f}, median={sorted(all_bends)[len(all_bends)//2]:.1f}")

    # Elbow connection distances
    elbow_dists = [
        c.get("euclidean_distance_mm") for c in connections
        if c.get("euclidean_distance_mm") is not None
    ]
    if elbow_dists:
        valid = [d for d in elbow_dists if d is not None]
        print(f"\n  Connection distances (mm):")
        print(f"    min={min(valid):.1f}, max={max(valid):.1f}, "
              f"median={sorted(valid)[len(valid)//2]:.1f}")


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════

def main() -> None:
    if not CSV_PATH.exists():
        print(f"ERROR: CSV not found: {CSV_PATH}")
        sys.exit(1)
    if not FCSTD_PATH.exists():
        print(f"ERROR: FreeCAD file not found: {FCSTD_PATH}")
        sys.exit(1)

    print("Loading CSV assembly data...")
    csv_rows = []
    with open(CSV_PATH, newline="", encoding="utf-8") as f:
        csv_rows = list(csv.DictReader(f))
    print(f"  Loaded {len(csv_rows)} CSV rows")

    print("Parsing FreeCAD Document.xml...")
    fc_links = parse_freecad_links(FCSTD_PATH)
    print(f"  Found {len(fc_links)} App::Link objects with labels")

    print("Building IFC parts index...")
    ifc_index = build_ifc_parts_index(PIPELINE_DIR)
    print(f"  Indexed {len(ifc_index)} parts from IFC files")

    print("Matching CSV labels to FreeCAD Links...")
    assembly_steps = build_assembly_steps(csv_rows, fc_links, ifc_index)

    print("Building part library...")
    part_library = build_part_library(assembly_steps)

    print("Building connection topology...")
    connections = build_connections(assembly_steps)

    print("Calculating connection angles...")
    connection_angles = build_connection_angles(connections, assembly_steps)

    print("Building position data...")
    position_data = build_position_data(assembly_steps)

    print("Building assembly tree...")
    assembly_tree = build_assembly_tree(assembly_steps)

    # Create blank template
    blank_path = DATA_DIR / "Assembly_3D_template.xlsx"
    create_blank_template(blank_path)

    # Create filled template
    filled_path = DATA_DIR / "Assembly_3D_template_filled.xlsx"
    create_filled_template(
        filled_path,
        Part_Library=part_library,
        Assembly_Steps=assembly_steps,
        Connection_Topology=connections,
        Connection_Angles=connection_angles,
        Position_Data=position_data,
        Assembly_Tree=assembly_tree,
    )

    print_summary(assembly_steps, connections, part_library)


if __name__ == "__main__":
    main()
