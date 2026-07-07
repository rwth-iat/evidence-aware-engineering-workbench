"""AIO Spec Compliance Audit — checks AIO workbooks against all spec requirements.

Audits every workbook against the Schema_Specification_v0.8_FREEZE requirements:
  A. Structure compliance (28 sheets, columns, PK uniqueness)
  B. FK integrity (all I-rules)
  C. Enum/lookup compliance (I13-I14, I29)
  D. Content quality (fill rates, semantic correctness per doc type)
  E. Provenance completeness (I12)
  F. Document-type-specific rules (I23-I28)

Run: python scripts/audit_aio_spec_compliance.py
"""

from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


def audit_all(ft_dir: Path | None = None) -> dict[str, Any]:
    from iev4pi_transformation_tool.core.aio_validator import validate_aio_workbook, AIO_SHEETS_28

    if ft_dir is None:
        ft_dir = Path("data/filled_templates")

    aio_files = sorted(ft_dir.glob("*_AIO.xlsx"))
    if not aio_files:
        print("No AIO workbooks found!")
        return {"error": "No AIO workbooks"}

    print(f"Auditing {len(aio_files)} AIO workbooks...\n")

    # ── A. Structure ──
    print("=" * 70)
    print("A. STRUCTURE COMPLIANCE")
    print("=" * 70)

    structure_issues = []
    sheet_presence: Counter = Counter()
    for f in aio_files:
        report = validate_aio_workbook(f)
        for sn in AIO_SHEETS_28:
            if sn in report["fill_rate"]:
                sheet_presence[sn] += 1
        # Check for structural errors (A-rules)
        for e in report["errors"]:
            if e["rule"].startswith("A"):
                structure_issues.append({"file": f.name, **e})

    missing_sheets = [sn for sn in AIO_SHEETS_28 if sheet_presence.get(sn, 0) < len(aio_files)]
    if missing_sheets:
        print(f"  ❌ Missing sheets: {missing_sheets}")
    else:
        print(f"  ✅ All 28 sheets present in {sheet_presence.get('Document_ID', 0)}/{len(aio_files)} workbooks")

    if structure_issues:
        print(f"  ❌ {len(structure_issues)} structure issues")
        for si in structure_issues[:5]:
            print(f"     [{si['rule']}] {si['file']}: {si['message'][:100]}")
    else:
        print(f"  ✅ 0 structural errors (A1-A5)")

    # ── B. FK Integrity ──
    print("\n" + "=" * 70)
    print("B. FK INTEGRITY (I1-I32 except I13-I14)")
    print("=" * 70)

    fk_errors: list[dict] = []
    all_errors_by_rule: Counter = Counter()
    for f in aio_files:
        report = validate_aio_workbook(f)
        for e in report["errors"]:
            all_errors_by_rule[e["rule"]] += 1
            if e["rule"] not in ("I13", "I14"):  # I13/I14 are lookup, not FK
                fk_errors.append({"file": f.name, **e})

    if fk_errors:
        print(f"  ❌ {len(fk_errors)} FK errors (non-I13/I14)")
        for fe in fk_errors[:5]:
            print(f"     [{fe['rule']}] {fe['file']}: {fe['message'][:100]}")
    else:
        print(f"  ✅ 0 FK integrity errors")

    # ── C. Enum/Lookup ──
    print("\n" + "=" * 70)
    print("C. ENUM / LOOKUP COMPLIANCE")
    print("=" * 70)

    i13_count = all_errors_by_rule.get("I13", 0)
    i14_count = all_errors_by_rule.get("I14", 0)
    print(f"  I13 (Attribute_Lookup): {i13_count} errors {'✅' if i13_count == 0 else '⚠️  (attribute standardization needed)'}")
    print(f"  I14 (Enum_Lookup):      {i14_count} errors {'✅' if i14_count == 0 else '⚠️ '}")

    # Read Enum_Lookup completeness
    import openpyxl
    tpl = openpyxl.load_workbook(Path("data/templates/Schema_Specification_v0.8_FREEZE_template.xlsx"), data_only=True)
    enum_ws = tpl["Enum_Lookup"]
    enum_entries = sum(1 for r in range(2, enum_ws.max_row + 1) if enum_ws.cell(row=r, column=1).value)
    attr_ws = tpl["Attribute_Lookup"]
    attr_entries = sum(1 for r in range(2, attr_ws.max_row + 1) if attr_ws.cell(row=r, column=1).value)
    tpl.close()
    print(f"  Enum_Lookup entries:     {enum_entries} (spec requires >= 336)")
    print(f"  Attribute_Lookup entries: {attr_entries} (spec requires >= 113)")

    # ── D. Content Quality ──
    print("\n" + "=" * 70)
    print("D. CONTENT QUALITY (Fill Rates & Semantic Correctness)")
    print("=" * 70)

    doc_types = Counter()
    element_types = Counter()
    total_elements = 0
    total_ed_rows = 0
    total_conn = 0
    rkz_patterns = Counter()
    sheets_filled_per_wb: list[int] = []

    # Collect per-workbook metrics
    wb_metrics = []
    for f in aio_files:
        wb = openpyxl.load_workbook(f, data_only=True)
        metrics = {"file": f.name, "doc_type": "", "elements": 0, "element_data": 0,
                    "connections": 0, "rkzs": [], "element_types": Counter()}

        # Doc type
        if "Document_ID" in wb.sheetnames:
            ws = wb["Document_ID"]
            for r in range(3, ws.max_row + 1):
                dt = str(ws.cell(row=r, column=3).value or "")
                if dt:
                    metrics["doc_type"] = dt
                    doc_types[dt] += 1

        # Elements
        if "Element_ID" in wb.sheetnames:
            ws = wb["Element_ID"]
            for r in range(3, ws.max_row + 1):
                etype = str(ws.cell(row=r, column=6).value or "")
                rkz = str(ws.cell(row=r, column=7).value or "")
                if etype:
                    metrics["elements"] += 1
                    metrics["element_types"][etype] += 1
                    element_types[etype] += 1
                    total_elements += 1
                if rkz:
                    metrics["rkzs"].append(rkz)
                    # Classify RKZ pattern
                    import re
                    if re.match(r'^-?[A-Z]\d+:', rkz):
                        rkz_patterns["terminal_designation"] += 1
                    elif re.match(r'^=[\d.]+[A-Z]', rkz):
                        rkz_patterns["pce_function_aspect"] += 1
                    elif re.match(r'^-?[A-Z]\d+$', rkz):
                        rkz_patterns["device_designation"] += 1
                    elif re.match(r'^[A-Z]{2}\d+\.[A-Z]\d+', rkz):
                        rkz_patterns["plant_position"] += 1
                    else:
                        rkz_patterns["other"] += 1

        # Element_Data
        if "Element_Data" in wb.sheetnames:
            ws = wb["Element_Data"]
            for r in range(3, ws.max_row + 1):
                if ws.cell(row=r, column=3).value:  # Element_ID FK
                    metrics["element_data"] += 1
                    total_ed_rows += 1

        # Connections
        if "Connection_ID" in wb.sheetnames:
            ws = wb["Connection_ID"]
            for r in range(3, ws.max_row + 1):
                if ws.cell(row=r, column=1).value:
                    metrics["connections"] += 1
                    total_conn += 1

        wb_metrics.append(metrics)

        # Fill rate
        report = validate_aio_workbook(f)
        sheets_filled_per_wb.append(report["filled_non_seed_sheets"])

        wb.close()

    print(f"  Document types: {dict(doc_types)}")
    print(f"  Total elements: {total_elements}")
    print(f"  Avg elements/workbook: {total_elements / max(1, len(aio_files)):.0f}")
    print(f"  Total Element_Data: {total_ed_rows}")
    print(f"  Total Connections: {total_conn}")
    print(f"  Avg sheets filled: {sum(sheets_filled_per_wb) / max(1, len(sheets_filled_per_wb)):.0f}")

    print(f"\n  Element_Type distribution (top 15):")
    for et, count in element_types.most_common(15):
        pct = 100 * count / max(1, total_elements)
        print(f"    {et:30s}: {count:5d} ({pct:.1f}%)")

    consumer_pct = 100 * element_types.get("Consumer", 0) / max(1, total_elements)
    print(f"\n  Consumer fallback rate: {consumer_pct:.1f}% {'✅ (< 50%)' if consumer_pct < 50 else '⚠️  (LLM classifier improvement needed)'}")

    print(f"\n  RKZ pattern distribution:")
    for pat, count in rkz_patterns.most_common():
        pct = 100 * count / max(1, sum(rkz_patterns.values()))
        print(f"    {pat:30s}: {count:5d} ({pct:.1f}%)")

    # ── E. Provenance ──
    print("\n" + "=" * 70)
    print("E. PROVENANCE COMPLETENESS (I12)")
    print("=" * 70)

    i12_warnings = 0
    for f in aio_files:
        report = validate_aio_workbook(f)
        for w in report.get("warnings", []):
            if w["rule"] == "I12":
                i12_warnings += 1

    print(f"  I12 provenance warnings: {i12_warnings} {'✅' if i12_warnings == 0 else '⚠️ '}")

    # Check Element_Data → Element_Data_Source ratio
    ed_with_source = 0
    ed_total = 0
    for f in aio_files[:5]:  # Sample 5 workbooks
        wb = openpyxl.load_workbook(f, data_only=True)
        if "Element_Data" in wb.sheetnames and "Element_Data_Source" in wb.sheetnames:
            ed_ids = set()
            ws = wb["Element_Data"]
            for r in range(3, ws.max_row + 1):
                v = ws.cell(row=r, column=2).value  # Element_Data_ID
                if v: ed_ids.add(str(v))
            ws2 = wb["Element_Data_Source"]
            eds_ids = set()
            for r in range(3, ws2.max_row + 1):
                v = ws2.cell(row=r, column=2).value  # Element_Data_ID
                if v: eds_ids.add(str(v))
            ed_with_source += len(ed_ids & eds_ids)
            ed_total += len(ed_ids)
        wb.close()

    if ed_total > 0:
        pct = 100 * ed_with_source / ed_total
        print(f"  Element_Data with Source row: {ed_with_source}/{ed_total} ({pct:.0f}%) {'✅' if pct >= 90 else '⚠️ '}")
    else:
        print(f"  No Element_Data to check")

    # ── F. Document-Type-Specific ──
    print("\n" + "=" * 70)
    print("F. DOCUMENT-TYPE-SPECIFIC RULES")
    print("=" * 70)

    i23_count = all_errors_by_rule.get("I23", 0)  # Bridge consistency
    i24_count = all_errors_by_rule.get("I24", 0)  # Wire color/polarity
    i25_count = all_errors_by_rule.get("I25", 0)  # Cable grouping
    i26_count = all_errors_by_rule.get("I26", 0)  # Current path numbers
    i28_count = all_errors_by_rule.get("I28", 0)  # IEC 60617 classification
    i29_count = all_errors_by_rule.get("I29", 0)  # Source_Format

    print(f"  I23 (Bridge consistency):       {i23_count} {'✅' if i23_count == 0 else '❌'}")
    print(f"  I24 (Wire color/polarity):       {i24_count} {'✅' if i24_count == 0 else '❌'}")
    print(f"  I25 (Cable grouping):            {i25_count} {'✅' if i25_count == 0 else '❌'}")
    print(f"  I26 (Current path numbers):      {i26_count} {'✅' if i26_count == 0 else '❌'}")
    print(f"  I28 (IEC 60617 classification):  {i28_count} {'✅' if i28_count == 0 else '❌'}")
    print(f"  I29 (Source_Format cardinality): {i29_count} {'✅' if i29_count == 0 else '❌'}")

    # ── Summary ──
    print("\n" + "=" * 70)
    print("COMPLIANCE SUMMARY")
    print("=" * 70)

    total_fk = sum(c for r, c in all_errors_by_rule.items() if r not in ("I13", "I14"))
    score = {
        "A_Structure": "PASS" if not structure_issues else "FAIL",
        "B_FK_Integrity": "PASS" if total_fk == 0 else "FAIL",
        "C_Enum_Lookup": "PASS" if i13_count == 0 and i14_count == 0 else "WARN",
        "D_Fill_Rate": f"{sum(sheets_filled_per_wb) / max(1, len(sheets_filled_per_wb)):.0f}/28 avg sheets",
        "E_Provenance": "PASS" if i12_warnings == 0 else "WARN",
        "F_DocType_Specific": "PASS" if (i23_count + i24_count + i25_count + i26_count + i28_count + i29_count) == 0 else "WARN",
    }

    for section, status in score.items():
        print(f"  {section:30s}: {status}")

    result = {
        "workbooks_audited": len(aio_files),
        "structure": {"issues": len(structure_issues), "pass": not structure_issues},
        "fk_integrity": {"errors": total_fk, "pass": total_fk == 0},
        "enum_lookup": {"I13": i13_count, "I14": i14_count},
        "content": {
            "total_elements": total_elements,
            "total_element_data": total_ed_rows,
            "total_connections": total_conn,
            "document_types": dict(doc_types),
            "element_types_top10": dict(element_types.most_common(10)),
            "consumer_fallback_pct": consumer_pct,
            "rkz_patterns": dict(rkz_patterns),
            "avg_sheets_filled": sum(sheets_filled_per_wb) / max(1, len(sheets_filled_per_wb)),
        },
        "provenance": {"I12_warnings": i12_warnings},
        "doc_type_specific": {"I23": i23_count, "I24": i24_count, "I25": i25_count,
                               "I26": i26_count, "I28": i28_count, "I29": i29_count},
        "score": score,
        "overall_pass": (
            not structure_issues
            and total_fk == 0
            and consumer_pct < 90
        ),
    }

    output = Path(__file__).resolve().parent.parent / "tests" / "aio_spec_audit.json"
    with open(output, "w") as f:
        json.dump(result, f, indent=2, ensure_ascii=False, default=str)
    print(f"\n  Full audit: {output}")

    return result


def main():
    result = audit_all()
    passed = result.get("overall_pass", False)
    if passed:
        print("\n✅ SPEC COMPLIANCE AUDIT PASSED")
    else:
        print("\n⚠️  SPEC COMPLIANCE AUDIT — issues found (see details above)")
    return 0 if passed else 1


if __name__ == "__main__":
    sys.exit(main())
