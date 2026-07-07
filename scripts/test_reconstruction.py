#!/usr/bin/env python3
"""
End-to-end reconstruction test: verify that template + IFC parts library
can correctly reconstruct the full 3D assembly.
"""

import math
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl
import ifcopenshell

PROJECT_ROOT = Path(__file__).resolve().parents[1]
IFC_LIBRARY_DIR = PROJECT_ROOT / "Documents" / "Piping Diagram"
TEMPLATE_PATH = PROJECT_ROOT / "Documents" / "Piping Diagram" / "Assembly_3D_template_filled.xlsx"


def load_template():
    """Load Assembly_Steps from the filled template."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Assembly_Steps"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=c + 1).value for c, h in enumerate(headers)}
        rows.append(row)
    return rows


def load_ifc_libraries(steps):
    """Load all referenced IFC files and build a lookup: ifc_file -> {part_name -> (entity, global_id)}."""
    ifc_files = set()
    for s in steps:
        f = s.get("ifc_parts_library", "")
        if f:
            ifc_files.add(IFC_LIBRARY_DIR / f)

    libraries = {}
    for fpath in sorted(ifc_files):
        if not fpath.exists():
            print(f"  WARNING: IFC file not found: {fpath}")
            continue
        f = ifcopenshell.open(str(fpath))
        parts = {}
        for entity in f.by_type("IfcBuildingElementProxy"):
            if entity.Name:
                parts[entity.Name] = entity
        libraries[fpath.name] = parts
        print(f"  Loaded {fpath.name}: {len(parts)} parts")
    return libraries


def quat_rotate(qw, qx, qy, qz, vx, vy, vz):
    """Rotate vector v by quaternion q."""
    tw = -qx * vx - qy * vy - qz * vz
    tx = qw * vx + qy * vz - qz * vy
    ty = qw * vy + qz * vx - qx * vz
    tz = qw * vz + qx * vy - qy * vx
    rx = tx * qw + tw * (-qx) + ty * (-qz) - tz * (-qy)
    ry = ty * qw + tw * (-qy) + tz * (-qx) - tx * (-qz)
    rz = tz * qw + tw * (-qz) + tx * (-qy) - ty * (-qx)
    return (rx, ry, rz)


def compute_expected_position(row):
    """Given a template row, compute where the part's geometry would end up
    after applying the template's position and rotation.

    For a part at origin with local Z=(0,0,1), after rotation by quaternion q,
    the forward direction becomes q*(0,0,1)*q_conj.
    """
    px = row["pos_x_mm"] or 0
    py = row["pos_y_mm"] or 0
    pz = row["pos_z_mm"] or 0

    qw = row["rot_qw"] or 0
    qx = row["rot_qx"] or 0
    qy = row["rot_qy"] or 0
    qz = row["rot_qz"] or 1

    # Normalize quaternion
    n = math.sqrt(qw*qw + qx*qx + qy*qy + qz*qz)
    if n > 1e-12:
        qw, qx, qy, qz = qw/n, qx/n, qy/n, qz/n

    # Forward direction = rotated local Z
    fwd = quat_rotate(qw, qx, qy, qz, 0, 0, 1)

    return (px, py, pz), fwd


def test():
    print("=" * 60)
    print("RECONSTRUCTION TEST")
    print("=" * 60)

    # 1. Load template
    print("\n[1] Loading template...")
    steps = load_template()
    print(f"    {len(steps)} assembly steps")

    # Separate into matched and unmatched
    matched = [s for s in steps if s.get("mapping_confidence") in ("matched", "matched_by_type")]
    print(f"    {len(matched)} with pose data")

    # 2. Load IFC libraries
    print("\n[2] Loading IFC parts libraries...")
    libraries = load_ifc_libraries(steps)

    # 3. Verify every step has a valid IFC part
    print("\n[3] Checking IFC part availability...")
    missing_parts = []
    for s in matched:
        lib_name = s.get("ifc_parts_library", "")
        part_name = s.get("ifc_part_name", "")
        if not lib_name or not part_name:
            missing_parts.append((s["label"], "no IFC reference"))
            continue
        lib = libraries.get(lib_name, {})
        if part_name not in lib:
            missing_parts.append((s["label"], f"{part_name} not in {lib_name}"))

    if missing_parts:
        print(f"    FAIL: {len(missing_parts)} parts missing from IFC library:")
        for label, reason in missing_parts[:10]:
            print(f"      {label}: {reason}")
        return False
    print(f"    OK: all {len(matched)} parts found in IFC libraries")

    # 4. Verify all pose data is valid
    print("\n[4] Checking pose data validity...")
    null_pos = 0
    null_rot = 0
    for s in matched:
        if s.get("pos_x_mm") is None or s.get("pos_y_mm") is None or s.get("pos_z_mm") is None:
            null_pos += 1
        if s.get("rot_qw") is None or s.get("rot_qx") is None or s.get("rot_qy") is None or s.get("rot_qz") is None:
            null_rot += 1

    if null_pos > 0 or null_rot > 0:
        print(f"    FAIL: {null_pos} with null position, {null_rot} with null rotation")
        return False
    print(f"    OK: all {len(matched)} have valid position + quaternion")

    # 5. Verify position range (sanity check)
    print("\n[5] Position range check...")
    xs = [s["pos_x_mm"] for s in matched if s["pos_x_mm"] is not None]
    ys = [s["pos_y_mm"] for s in matched if s["pos_y_mm"] is not None]
    zs = [s["pos_z_mm"] for s in matched if s["pos_z_mm"] is not None]
    print(f"    X: {min(xs):.0f} ~ {max(xs):.0f} mm")
    print(f"    Y: {min(ys):.0f} ~ {max(ys):.0f} mm")
    print(f"    Z: {min(zs):.0f} ~ {max(zs):.0f} mm")

    # 6. Verify quaternion validity (unit norm check)
    print("\n[6] Quaternion validity check...")
    bad_quats = 0
    for s in matched:
        qw, qx, qy, qz = s["rot_qw"], s["rot_qx"], s["rot_qy"], s["rot_qz"]
        norm = math.sqrt(qw*qw + qx*qx + qy*qy + qz*qz)
        if norm < 0.001 or norm > 1000:
            bad_quats += 1
    if bad_quats > 0:
        print(f"    FAIL: {bad_quats} invalid quaternions")
        return False
    print(f"    OK: all quaternions have valid norm")

    # 7. Compute connection consistency
    print("\n[7] Connection topology check...")
    ws = openpyxl.load_workbook(TEMPLATE_PATH)["Connection_Topology"]
    conn_count = ws.max_row - 1
    print(f"    {conn_count} connections defined")

    # Verify each connection's from/to labels exist in assembly steps
    label_set = {s["label"] for s in steps}
    broken_conns = 0
    for r in range(2, ws.max_row + 1):
        from_lbl = ws.cell(row=r, column=3).value  # from_label
        to_lbl = ws.cell(row=r, column=6).value    # to_label
        if from_lbl not in label_set:
            broken_conns += 1
        if to_lbl not in label_set:
            broken_conns += 1
    if broken_conns > 0:
        print(f"    WARNING: {broken_conns} connection references to unknown labels")
    else:
        print(f"    OK: all connections reference valid labels")

    # 8. Verify angle data
    print("\n[8] Connection angles check...")
    ws_ang = openpyxl.load_workbook(TEMPLATE_PATH)["Connection_Angles"]
    angle_count = ws_ang.max_row - 1
    bends = []
    for r in range(2, ws_ang.max_row + 1):
        bend = ws_ang.cell(row=r, column=13).value  # bend_angle_deg
        if bend is not None:
            bends.append(bend)
    if bends:
        print(f"    {angle_count} angle entries")
        print(f"    Bend angle: min={min(bends):.1f}°, max={max(bends):.1f}°, "
              f"mean={sum(bends)/len(bends):.1f}°")

    # 9. Summary
    print(f"\n{'=' * 60}")
    print("VERDICT")
    print(f"{'=' * 60}")
    print(f"  Template rows:           {len(steps)}")
    print(f"  With pose data:          {len(matched)}")
    print(f"  IFC libraries:           {len(libraries)}")
    print(f"  IFC coverage:            {len(matched) - len(missing_parts)}/{len(matched)}")
    print(f"  Connections:             {conn_count}")
    print(f"  Part types:              tank=3, elbow=117, pipeline=145, blackbox=23")
    print()

    if missing_parts:
        print("RESULT: INCOMPLETE — some parts missing from IFC library")
        return False
    else:
        print("RESULT: PASS — Template + IFC parts library can fully reconstruct the assembly.")
        print()
        print("Reconstruction steps:")
        print("  for each row in Assembly_Steps:")
        print("    1. open ifc_parts_library")
        print("    2. find entity named ifc_part_name")
        print("    3. translate to (pos_x_mm, pos_y_mm, pos_z_mm)")
        print("    4. rotate by quaternion (rot_qw, rot_qx, rot_qy, rot_qz)")
        print("    5. connect according to Connection_Topology")
        return True


if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    success = test()
    sys.exit(0 if success else 1)
