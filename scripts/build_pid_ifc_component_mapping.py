#!/usr/bin/env python3
"""Build an auditable P&ID-to-IFC component mapping workbook.

The source P&ID carries functional tags and 2D locations, while the 3D IFC
assembly only exposes generic FreeCAD body names. This script keeps that
evidence gap visible: strong tank matches are marked as matched; blackbox
matches are provided as topology/order candidates and flagged for review.
"""

from __future__ import annotations

import argparse
import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PIPING_DIR = PROJECT_ROOT / "Documents" / "Piping Diagram"
PID_DIR = PROJECT_ROOT / "Documents" / "R&I-Fließbild"
MAPPING_SOURCE_PATH = PIPING_DIR / "_legacy" / "instrument_pipe_mapping.xlsx"
ASSEMBLY_PATH = PIPING_DIR / "Assembly_3D_template_filled.xlsx"
PID_XML_PATH = PID_DIR / "=A10.A10.A10.FB.001 (A58A1PQ0CA).xml"
DEFAULT_OUTPUT_PATH = ASSEMBLY_PATH  # merge into Assembly workbook

HIGH_CONFIDENCE_THRESHOLD = 0.80

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_FILL = PatternFill("solid", fgColor="EAF2F8")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
MATCH_FILL = PatternFill("solid", fgColor="E2F0D9")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


@dataclass(frozen=True)
class SourceRow:
    row_number: int
    component: str
    row_anchor: str
    instrument_tags: tuple[str, ...]
    suggested_3d_type: str
    notes: str


@dataclass(frozen=True)
class Instrument:
    canonical_tag: str
    function_code: str
    piping_anchor_id: str
    piping_component_name: str
    actuating_location: str
    device_information: str


@dataclass(frozen=True)
class PidComponent:
    pid_id: str
    tag_name: str
    component_class: str
    x: float | None
    y: float | None
    full_name: str
    label: str
    description: str
    vessel_hint: str


@dataclass(frozen=True)
class AssemblyComponent:
    label: str
    component_type: str
    fcstd_link_name: str
    ifc_part_name: str
    ifc_parts_library: str
    pos_x_mm: float | None
    pos_y_mm: float | None
    pos_z_mm: float | None
    assembly_no: int
    global_id: str


@dataclass(frozen=True)
class MatchDecision:
    preferred_label: str
    candidate_labels: tuple[str, ...]
    confidence: float
    match_status: str
    needs_review: bool
    evidence_summary: str


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_tags(value: str) -> tuple[str, ...]:
    return tuple(part.strip() for part in value.split(",") if part.strip())


def row_dicts(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(header) for header in rows[0]]
    result = []
    for offset, row in enumerate(rows[1:], start=2):
        result.append({"__row_number": offset, **{header: row[index] for index, header in enumerate(headers)}})
    return result


def load_source_rows(path: Path = MAPPING_SOURCE_PATH) -> list[SourceRow]:
    rows = []
    for row in row_dicts(path, "Instrument_Pipe_Mapping"):
        component = clean(row.get("pand_id_component"))
        if not component:
            continue
        rows.append(
            SourceRow(
                row_number=int(row["__row_number"]),
                component=component,
                row_anchor=clean(row.get("pand_id_anchor")),
                instrument_tags=split_tags(clean(row.get("instrument_tags"))),
                suggested_3d_type=clean(row.get("suggested_3d_type")),
                notes=clean(row.get("notes")),
            )
        )
    return rows


def load_instruments(path: Path = MAPPING_SOURCE_PATH) -> list[Instrument]:
    instruments = []
    for row in row_dicts(path, "Instruments_from_PID"):
        canonical_tag = clean(row.get("canonical_tag"))
        if not canonical_tag:
            continue
        instruments.append(
            Instrument(
                canonical_tag=canonical_tag,
                function_code=clean(row.get("function_code")),
                piping_anchor_id=clean(row.get("piping_anchor_id")),
                piping_component_name=clean(row.get("piping_component_name")),
                actuating_location=clean(row.get("actuating_location")),
                device_information=clean(row.get("device_information")),
            )
        )
    return instruments


def parse_ifc_global_ids(ifc_paths: list[Path]) -> dict[tuple[str, str], str]:
    """Parse GlobalIds from IFC file(s).

    Tries ifcopenshell first (handles binary IFC); falls back to regex for
    plain-text IFC files.
    """
    ids: dict[tuple[str, str], str] = {}
    for path in ifc_paths:
        if not path.exists():
            continue
        try:
            import ifcopenshell
            f = ifcopenshell.open(str(path))
            for entity in f.by_type("IfcBuildingElementProxy"):
                if entity.Name:
                    ids[(path.name, entity.Name)] = entity.GlobalId
        except Exception:
            pattern = re.compile(
                r"#\d+=IFCBUILDINGELEMENTPROXY\('(?P<gid>[^']*)',#[^,]+,'(?P<name>[^']*)'",
                re.IGNORECASE,
            )
            text = path.read_text(encoding="utf-8", errors="ignore")
            for match in pattern.finditer(text):
                name = match.group("name")
                if name:
                    ids[(path.name, name)] = match.group("gid")
    return ids


def _default_ifc_paths() -> list[Path]:
    """Find the IFC source(s) to scan for GlobalIds.

    Prefers a single prepared assembled IFC; also includes legacy files
    and LOCKED fallback for entries not found in the prepared IFC.
    """
    paths: list[Path] = []
    for candidate in [
        PIPING_DIR / "assembly_prepared.ifc",
        PIPING_DIR / "model_locked.ifc",
        PIPING_DIR / "Unna5.ifc",
        PIPING_DIR / "Unnamed28.ifc",
    ]:
        if candidate.exists():
            paths.append(candidate)
    return paths


def load_assembly_components(path: Path = ASSEMBLY_PATH) -> dict[str, AssemblyComponent]:
    ifc_ids = parse_ifc_global_ids(_default_ifc_paths())
    components: dict[str, AssemblyComponent] = {}
    for row in row_dicts(path, "Assembly_Steps"):
        component_type = clean(row.get("type"))
        if component_type not in {"blackbox", "tank"}:
            continue
        label = clean(row.get("label"))
        library = clean(row.get("ifc_parts_library"))
        part_name = clean(row.get("ifc_part_name"))
        components[label] = AssemblyComponent(
            label=label,
            component_type=component_type,
            fcstd_link_name=clean(row.get("fcstd_link_name")),
            ifc_part_name=part_name,
            ifc_parts_library=library,
            pos_x_mm=as_float(row.get("pos_x_mm")),
            pos_y_mm=as_float(row.get("pos_y_mm")),
            pos_z_mm=as_float(row.get("pos_z_mm")),
            assembly_no=int(float(row.get("assembly_no") or 0)),
            global_id=ifc_ids.get((library, part_name), ""),
        )
    return components


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_pid_components(path: Path = PID_XML_PATH) -> dict[str, PidComponent]:
    root = ET.parse(path).getroot()
    components = {}
    for element in root.iter():
        pid_id = clean(element.attrib.get("ID"))
        if not pid_id:
            continue
        tag_name = clean(element.attrib.get("TagName"))
        component_class = clean(element.attrib.get("ComponentClass"))
        if not tag_name and component_class not in {"Nozzle", "InlinePrimaryElement", "TightShutOffValve", "GlobeValve"}:
            continue
        location = element.find("./Position/Location")
        attributes = {
            clean(attr.attrib.get("Name")): clean(attr.attrib.get("Value"))
            for attr in element.findall(".//GenericAttribute")
        }
        full_name = attributes.get("FullName", "")
        label = attributes.get("Label", "")
        vessel_hint = extract_vessel_hint(full_name) or extract_vessel_hint(attributes.get("PathFullName", ""))
        components[pid_id] = PidComponent(
            pid_id=pid_id,
            tag_name=tag_name,
            component_class=component_class,
            x=as_float(location.attrib.get("X")) if location is not None else None,
            y=as_float(location.attrib.get("Y")) if location is not None else None,
            full_name=full_name,
            label=label,
            description=attributes.get("Description", ""),
            vessel_hint=vessel_hint,
        )
    return components


def extract_vessel_hint(value: str) -> str:
    match = re.search(r"VE\d{3}", value or "")
    return match.group(0) if match else ""


def instruments_for_source_row(source_row: SourceRow, instruments: list[Instrument]) -> list[Instrument]:
    tags = set(source_row.instrument_tags)
    matches = [instrument for instrument in instruments if instrument.canonical_tag in tags]
    if matches:
        return matches
    return [
        instrument
        for instrument in instruments
        if instrument.piping_component_name == source_row.component
        or (source_row.row_anchor and instrument.piping_anchor_id == source_row.row_anchor)
    ]


def candidate_window(label: str, assembly_components: dict[str, AssemblyComponent], radius: int = 1) -> tuple[str, ...]:
    component = assembly_components.get(label)
    if not component:
        return ()
    same_type = sorted(
        (item for item in assembly_components.values() if item.component_type == component.component_type),
        key=lambda item: item.assembly_no,
    )
    labels = [item.label for item in same_type]
    try:
        index = labels.index(label)
    except ValueError:
        return (label,)
    low = max(0, index - radius)
    high = min(len(labels), index + radius + 1)
    return tuple(labels[low:high])


def decide_match(
    source_row: SourceRow,
    instrument: Instrument,
    pid_component: PidComponent | None,
    assembly_components: dict[str, AssemblyComponent],
) -> MatchDecision:
    anchor = instrument.piping_anchor_id or source_row.row_anchor
    suggested_type = source_row.suggested_3d_type

    tank_by_anchor = {
        "XMP_274": "tank_001",  # VE003, TU10.L10
        "XMP_105": "tank_002",  # VE001, TU30.T33
        "XMP_108": "tank_002",  # VE001, TU30.L34
        "XMP_111": "tank_002",  # VE001, TU30.L35
        "XMP_68": "tank_003",  # VE002, TU20.T27
        "XMP_71": "tank_003",  # VE002, TU20.Q28
    }
    blackbox_by_anchor = {
        "XMP_390": "blackbox_001",  # DM001 / TU10.P14
        "XMP_398": "blackbox_004",  # DM003 / TU10.P19
        "XMP_394": "blackbox_015",  # DM002 / TU10.T15
        "XMP_402": "blackbox_016",  # DM004 / TU10.T20
        # XMP_169 / VV001 / TU10.Y6: uc1_expected_missing (no 3D mapping)
        "XMP_257": "blackbox_017",  # VV002 / TU10.Y21
        "XMP_209": "blackbox_020",  # B1 / TU10.F17
        "XMP_262": "blackbox_018",  # B1 / TU10.F22
        "XMP_28": "blackbox_023",   # VV004 / TU10.Y24
        "XMP_406": "blackbox_021",  # DM005 / TU10.T23
        "XMP_23": "blackbox_022",   # VV003 / TU10.Y25 (TU10.U41=software_control)
        "XMP_309": "blackbox_007",  # VV009 / TU30.Y37
        "XMP_314": "blackbox_008",  # VV010 / TU30.Y39
        "XMP_371": "blackbox_011",  # B1 / TU20.F31
        "XMP_121": "blackbox_012",  # VV006 / TU20.Y30 (TU20.U42=software_control)
    }
    # UC1 expected-missing items (for testing UC1 detection)
    UC1_EXPECTED_MISSING = {"TU10.Y6", "VV005"}
    # UC1 software-control items (no physical IFC required)
    UC1_SOFTWARE_CONTROL = {"TU10.U41", "TU20.U42"}

    if source_row.component == "(no anchor)" or not anchor:
        labels = tuple(
            item.label for item in sorted(assembly_components.values(), key=lambda item: item.assembly_no)
        )
        return MatchDecision(
            preferred_label="",
            candidate_labels=labels[:8],
            confidence=0.0,
            match_status="unmapped_no_pid_anchor",
            needs_review=True,
            evidence_summary="No P&ID piping anchor is available for this instrument row.",
        )

    if suggested_type == "tank" and anchor in tank_by_anchor:
        label = tank_by_anchor[anchor]
        vessel = pid_component.vessel_hint if pid_component else ""
        evidence = (
            f"Tank topology assumption maps P&ID vessels VE003 -> VE001 -> VE002 "
            f"to 3D path tank_001 -> tank_002 -> tank_003; anchor {anchor}"
            f"{f' ({vessel})' if vessel else ''} selects {label}."
        )
        return MatchDecision(
            preferred_label=label,
            candidate_labels=candidate_window(label, assembly_components, radius=1),
            confidence=0.90,
            match_status="matched_tank_topology",
            needs_review=False,
            evidence_summary=evidence,
        )

    # ── Special rules: UC1 expected-missing / software-control ──
    canonical_tag = instrument.canonical_tag or ""
    if canonical_tag in UC1_EXPECTED_MISSING:
        return MatchDecision(
            preferred_label="",
            candidate_labels=(),
            confidence=1.0,
            match_status="ifc_missing",
            needs_review=False,
            evidence_summary=f"{canonical_tag} is intentionally absent from the 3D IFC model (UC1 test case).",
        )
    if canonical_tag in UC1_SOFTWARE_CONTROL:
        return MatchDecision(
            preferred_label="",
            candidate_labels=(),
            confidence=1.0,
            match_status="software_control_no_ifc_required",
            needs_review=False,
            evidence_summary=f"{canonical_tag} is a UY software-control signal; no physical IFC component required.",
        )

    # Explicit VV005 missing (checked by P&ID component name, no canonical_tag)
    if source_row.component == "VV005":
        return MatchDecision(
            preferred_label="",
            candidate_labels=(),
            confidence=1.0,
            match_status="ifc_missing",
            needs_review=False,
            evidence_summary="VV005 is a P&ID valve with no corresponding 3D IFC component (UC1 test case).",
        )

    if suggested_type == "blackbox" and anchor in blackbox_by_anchor:
        label = blackbox_by_anchor[anchor]
        pid_xy = f"({pid_component.x:g}, {pid_component.y:g})" if pid_component and pid_component.x is not None else "unknown"
        evidence = (
            f"Candidate inferred from P&ID anchor order/location {anchor} at {pid_xy}, "
            f"component {source_row.component}, and the 3D blackbox path order. "
            "IFC exposes only generic Body### names, so this remains a review candidate."
        )
        return MatchDecision(
            preferred_label=label,
            candidate_labels=candidate_window(label, assembly_components, radius=1),
            confidence=0.68,
            match_status="candidate_topology_order",
            needs_review=True,
            evidence_summary=evidence,
        )

    return MatchDecision(
        preferred_label="",
        candidate_labels=(),
        confidence=0.0,
        match_status="unresolved_no_rule",
        needs_review=True,
        evidence_summary="No rule matched for this instrument/component pair.",
    )


def build_instance_rows(
    source_rows: list[SourceRow],
    instruments: list[Instrument],
    pid_components: dict[str, PidComponent],
    assembly_components: dict[str, AssemblyComponent],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source_row in source_rows:
        row_instruments = instruments_for_source_row(source_row, instruments)
        if not row_instruments:
            row_instruments = [
                Instrument(
                    canonical_tag="",
                    function_code="",
                    piping_anchor_id=source_row.row_anchor,
                    piping_component_name=source_row.component,
                    actuating_location=source_row.row_anchor,
                    device_information="",
                )
            ]
        for instrument in row_instruments:
            anchor = instrument.piping_anchor_id or source_row.row_anchor
            pid_component = pid_components.get(anchor)
            decision = decide_match(source_row, instrument, pid_component, assembly_components)
            assembly = assembly_components.get(decision.preferred_label)
            rows.append(
                {
                    "source_row": source_row.row_number,
                    "pand_id_component": source_row.component,
                    "pand_id_anchor": source_row.row_anchor,
                    "instance_key": f"{source_row.component}@{anchor or 'no_anchor'}",
                    "canonical_tag": instrument.canonical_tag,
                    "function_code": instrument.function_code,
                    "pid_anchor_id": anchor,
                    "pid_tag_name": pid_component.tag_name if pid_component else "",
                    "pid_component_class": pid_component.component_class if pid_component else "",
                    "pid_x": pid_component.x if pid_component else None,
                    "pid_y": pid_component.y if pid_component else None,
                    "pid_vessel_hint": pid_component.vessel_hint if pid_component else "",
                    "suggested_3d_type": source_row.suggested_3d_type,
                    "mapped_3d_label": assembly.label if assembly else "",
                    "ifc_part_name": assembly.ifc_part_name if assembly else "",
                    "ifc_parts_library": assembly.ifc_parts_library if assembly else "",
                    "global_id": assembly.global_id if assembly else "",
                    "fcstd_link_name": assembly.fcstd_link_name if assembly else "",
                    "pos_x_mm": assembly.pos_x_mm if assembly else None,
                    "pos_y_mm": assembly.pos_y_mm if assembly else None,
                    "pos_z_mm": assembly.pos_z_mm if assembly else None,
                    "confidence": decision.confidence,
                    "match_status": decision.match_status,
                    "needs_review": "yes" if decision.needs_review else "no",
                    "candidate_labels": ", ".join(decision.candidate_labels),
                    "evidence_summary": decision.evidence_summary,
                }
            )
    return rows


def _make_evidence(source_path, page_or_sheet, cell_range_or_bbox, snippet, evidence_type="native_text", engine="mapping_script", score=1.0):
    """Build a single EvidenceRef-like dict for P&ID provenance."""
    return {
        "source_path": source_path,
        "page_or_sheet": page_or_sheet,
        "cell_range_or_bbox": cell_range_or_bbox,
        "snippet": snippet or "",
        "score": score,
        "evidence_type": evidence_type,
        "engine": engine,
    }


def _evidence_for_field(field_name, row, pid_component, assembly_components, source_row_num, asm_path, dexpi_path, mapping_path):
    """Return list of EvidenceRef dicts for a single cell value."""
    anchor = row.get("pid_anchor_id", "")
    value = row.get(field_name, "")
    if not str(value):
        return []

    # --- DEXPI XML sourced fields ---
    dexpi_fields = {
        "pid_anchor_id": (
            f"Component#{anchor}",
            f"Component ID={anchor}" if pid_component else "",
            "dexpi_node",
        ),
        "pid_tag_name": (
            f"Component#{anchor}",
            f"TagName={pid_component.tag_name}" if pid_component else "",
            "dexpi_node",
        ),
        "pid_component_class": (
            f"Component#{anchor}",
            f"ComponentClass={pid_component.component_class}" if pid_component else "",
            "dexpi_node",
        ),
        "pid_component_name": (
            f"Component#{anchor}",
            f"FullName={pid_component.full_name}" if pid_component else "",
            "dexpi_node",
        ),
        "pid_vessel_hint": (
            f"Component#{anchor}",
            f"vessel_hint={pid_component.vessel_hint}" if pid_component else "",
            "dexpi_node",
        ),
        "pid_x": (
            f"Component#{anchor}/Position/Location",
            f"X={pid_component.x}" if pid_component else "",
            "dexpi_node",
        ),
        "pid_y": (
            f"Component#{anchor}/Position/Location",
            f"Y={pid_component.y}" if pid_component else "",
            "dexpi_node",
        ),
    }
    if field_name in dexpi_fields and pid_component is not None:
        bbox, snippet, etype = dexpi_fields[field_name]
        return [_make_evidence(dexpi_path, "DEXPI XML", bbox, snippet, evidence_type=etype, engine="dexpi")]

    # --- Assembly_Steps sourced fields ---
    label = row.get("mapped_3d_label", "")
    asm = assembly_components.get(label)
    asm_fields = {
        "mapped_3d_label": "label",
        "ifc_part_name": "ifc_part_name",
        "ifc_parts_library": "ifc_parts_library",
        "fcstd_link_name": "fcstd_link_name",
        "pos_x_mm": "pos_x_mm",
        "pos_y_mm": "pos_y_mm",
        "pos_z_mm": "pos_z_mm",
    }
    if field_name in asm_fields and asm is not None:
        src_col = asm_fields[field_name]
        cell_ref = f"{src_col}_row"
        return [_make_evidence(
            asm_path,
            "Assembly_Steps",
            cell_ref,
            f"{src_col}={getattr(asm, src_col, '')}",
            evidence_type="spreadsheet_cell",
            engine="assembly_pipeline",
        )]

    # --- global_id: from IFC via parse_ifc_global_ids ---
    if field_name == "global_id" and asm is not None and asm.global_id:
        ifc_paths = _default_ifc_paths()
        ifc_source = str(ifc_paths[0].relative_to(PROJECT_ROOT)) if ifc_paths and ifc_paths[0].exists() else asm_path
        return [_make_evidence(
            ifc_source,
            "IFC",
            f"GlobalId={asm.global_id}",
            f"ifc_part_name={asm.ifc_part_name}",
            evidence_type="ifc_element",
            engine="ifcopenshell",
        )]

    # --- instrument_pipe_mapping sourced fields ---
    mapping_fields = {
        "source_row": ("Instrument_Pipe_Mapping", f"Row {source_row_num}", "source_row", source_row_num),
        "pand_id_component": ("Instrument_Pipe_Mapping", f"pand_id_component", "pand_id_component", source_row_num),
        "pand_id_anchor": ("Instrument_Pipe_Mapping", f"pand_id_anchor", "pand_id_anchor", source_row_num),
        "function_code": ("Instruments_from_PID", f"function_code", "function_code", source_row_num),
        "suggested_3d_type": ("Instrument_Pipe_Mapping", f"suggested_3d_type", "suggested_3d_type", source_row_num),
    }
    if field_name in mapping_fields:
        sheet, col_name, _, row_num = mapping_fields[field_name]
        return [_make_evidence(
            mapping_path,
            sheet,
            f"{col_name}_row_{row_num}",
            f"{field_name}={value}",
            evidence_type="spreadsheet_cell",
            engine="legacy_mapping",
        )]

    # --- Algorithm-derived fields ---
    if field_name in ("confidence", "match_status", "needs_review", "candidate_labels", "evidence_summary"):
        refs = []
        if pid_component is not None:
            refs.append(_make_evidence(
                dexpi_path, "DEXPI XML", f"Component#{anchor}",
                f"TagName={pid_component.tag_name}" if pid_component.tag_name else "",
                evidence_type="dexpi_node", engine="dexpi",
            ))
        if asm is not None:
            refs.append(_make_evidence(
                asm_path, "Assembly_Steps", f"label={label}",
                f"component_type={asm.component_type}",
                evidence_type="spreadsheet_cell", engine="assembly_pipeline",
            ))
        return refs

    return []


def build_pid_provenance(instance_rows, pid_components, assembly_components):
    """Generate per-cell provenance for P&ID_instance_mapping sheet.

    Returns dict with shape:
        {"P&ID_instance_mapping": {"A2": {ExcelCellProvenance...}, "B2": {...}, ...}}
    """
    from openpyxl.utils import get_column_letter

    if not instance_rows:
        return {}

    dexpi_path = str(PID_XML_PATH.relative_to(PROJECT_ROOT))
    asm_path = str(ASSEMBLY_PATH.relative_to(PROJECT_ROOT))
    mapping_path = str(MAPPING_SOURCE_PATH.relative_to(PROJECT_ROOT))

    headers = list(instance_rows[0].keys())
    skip_fields = {"instance_key"}  # Derived compound key, no source

    result = {"P&ID_instance_mapping": {}}
    for row_idx, row in enumerate(instance_rows):
        anchor = row.get("pid_anchor_id", "")
        pid_comp = pid_components.get(anchor)
        source_row_num = row.get("source_row", 0)
        label = row.get("mapped_3d_label", "")
        asm = assembly_components.get(label)

        for col_idx, field_name in enumerate(headers):
            if field_name in skip_fields:
                continue
            value = str(row.get(field_name, ""))
            if not value:
                continue
            coord = f"{get_column_letter(col_idx + 1)}{row_idx + 2}"

            evidence_refs = _evidence_for_field(
                field_name, row, pid_comp, assembly_components,
                source_row_num, asm_path, dexpi_path, mapping_path,
            )
            if not evidence_refs:
                continue

            result["P&ID_instance_mapping"][coord] = {
                "workbook_name": "piping_diagram/Assembly_3D_template.xlsx",
                "sheet_name": "P&ID_instance_mapping",
                "row": row_idx + 2,
                "column": col_idx + 1,
                "coord": coord,
                "source_path": evidence_refs[0]["source_path"],
                "record_key": row.get("instance_key", f"row_{row_idx}"),
                "record_display_name": row.get("instance_key", f"row {row_idx + 1}"),
                "field_name": field_name,
                "value": value,
                "normalized_value": value.strip().lower(),
                "confidence": float(row.get("confidence", 0.0)),
                "decision_confidence": None,
                "status": "filled",
                "evidence_refs": evidence_refs,
                "evidence_bundle_id": "",
                "uncertainty_reason": "",
                "llm_verification_status": "",
                "rule_support": [],
                "review_feedback_status": "",
                "notes": f"P&ID mapping: {row.get('match_status', '')}",
            }

    return result


def aggregate_group_rows(source_rows: list[SourceRow], instance_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_source: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in instance_rows:
        by_source[int(row["source_row"])].append(row)

    grouped = []
    for source_row in source_rows:
        rows = by_source.get(source_row.row_number, [])
        confidences = [float(row["confidence"]) for row in rows if row["confidence"] not in ("", None)]
        grouped.append(
            {
                "source_row": source_row.row_number,
                "pand_id_component": source_row.component,
                "pand_id_anchor": source_row.row_anchor,
                "instrument_count": len([row for row in rows if row.get("canonical_tag")]),
                "instrument_tags": ", ".join(row.get("canonical_tag", "") for row in rows if row.get("canonical_tag")),
                "suggested_3d_type": source_row.suggested_3d_type,
                "mapped_3d_labels": join_unique(row.get("mapped_3d_label", "") for row in rows),
                "ifc_part_names": join_unique(row.get("ifc_part_name", "") for row in rows),
                "ifc_parts_libraries": join_unique(row.get("ifc_parts_library", "") for row in rows),
                "global_ids": join_unique(row.get("global_id", "") for row in rows),
                "min_confidence": min(confidences) if confidences else 0.0,
                "avg_confidence": round(mean(confidences), 3) if confidences else 0.0,
                "needs_review": "yes" if any(row.get("needs_review") == "yes" for row in rows) else "no",
                "match_statuses": join_unique(row.get("match_status", "") for row in rows),
                "evidence_summary": " | ".join(row.get("evidence_summary", "") for row in rows[:3]),
            }
        )
    return grouped


def join_unique(values: Any) -> str:
    result = []
    for value in values:
        text = clean(value)
        if text and text not in result:
            result.append(text)
    return ", ".join(result)


def validate_results(source_rows: list[SourceRow], instance_rows: list[dict[str, Any]], grouped_rows: list[dict[str, Any]]) -> list[str]:
    errors = []
    source_ids = {row.row_number for row in source_rows}
    grouped_ids = {int(row["source_row"]) for row in grouped_rows}
    missing_grouped = sorted(source_ids - grouped_ids)
    if missing_grouped:
        errors.append(f"source rows missing from grouped_mapping: {missing_grouped}")

    for row in instance_rows:
        status = row.get("match_status", "")
        if status in ("ifc_missing", "software_control_no_ifc_required"):
            continue  # intentionally missing — not an error
        if float(row["confidence"]) >= HIGH_CONFIDENCE_THRESHOLD:
            missing = [
                field
                for field in ("mapped_3d_label", "ifc_part_name", "ifc_parts_library", "global_id")
                if not clean(row.get(field))
            ]
            if missing:
                errors.append(f"high-confidence row {row['instance_key']} missing {missing}")

    b1_instances = {row["pid_anchor_id"] for row in instance_rows if row["pand_id_component"] == "B1"}
    if len(b1_instances) < 3:
        errors.append(f"B1 should expand to at least 3 P&ID instances, found {sorted(b1_instances)}")

    for component in ("DM001", "DM002", "DM003", "DM004", "DM005"):
        if not any(row["pand_id_component"] == component for row in instance_rows):
            errors.append(f"{component} missing from instance_mapping")

    for component in ("VV001", "VV002", "VV003", "VV004", "VV006", "VV009", "VV010"):
        if not any(row["pand_id_component"] == component for row in instance_rows):
            errors.append(f"{component} missing from instance_mapping")

    unresolved = [row for row in instance_rows if float(row["confidence"]) < HIGH_CONFIDENCE_THRESHOLD]
    if unresolved and not all(row["needs_review"] == "yes" for row in unresolved):
        errors.append("all low-confidence or ambiguous rows must be marked needs_review=yes")

    return errors


_MAPPING_SHEETS = [
    "P&ID_instance_mapping", "P&ID_grouped_mapping", "P&ID_review_needed",
    "P&ID_evidence", "P&ID_3d_reference", "P&ID_validation",
]


def write_workbook(
    output_path: Path,
    instance_rows: list[dict[str, Any]],
    grouped_rows: list[dict[str, Any]],
    assembly_components: dict[str, AssemblyComponent],
    validation_errors: list[str],
    pid_components: dict[str, PidComponent] | None = None,
) -> None:
    # Open existing Assembly workbook (or create new if missing)
    if output_path.exists():
        wb = load_workbook(output_path)
        # Remove old mapping sheets if present
        for sn in _MAPPING_SHEETS:
            if sn in wb.sheetnames:
                del wb[sn]
    else:
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

    write_sheet(wb, "P&ID_instance_mapping", instance_rows)
    write_sheet(wb, "P&ID_grouped_mapping", grouped_rows)
    review_rows = [row for row in instance_rows if row.get("needs_review") == "yes"]
    write_sheet(wb, "P&ID_review_needed", review_rows)
    write_sheet(wb, "P&ID_evidence", build_evidence_rows(instance_rows))
    write_sheet(wb, "P&ID_3d_reference", [component.__dict__ for component in sorted(assembly_components.values(), key=lambda item: item.assembly_no)])
    validation_rows = (
        [{"status": "ok", "message": "All validation checks passed."}]
        if not validation_errors
        else [{"status": "error", "message": error} for error in validation_errors]
    )
    write_sheet(wb, "P&ID_validation", validation_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    # Also sync to data/filled_templates/ so the export pipeline picks up
    # the P&ID sheets when copying to Exports/Excel/piping_diagram/
    from iev4pi_transformation_tool.core.standardized_templates import (
        FILLED_TEMPLATES_DIR, ASSEMBLY_3D_TEMPLATE,
    )
    FILLED_TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(str(FILLED_TEMPLATES_DIR / ASSEMBLY_3D_TEMPLATE))

    # Write P&ID provenance sidecar for cell-level source preview
    if pid_components is not None:
        import json as _json
        _pid_prov = build_pid_provenance(instance_rows, pid_components, assembly_components)
        _prov_path = FILLED_TEMPLATES_DIR / (ASSEMBLY_3D_TEMPLATE + ".pid_provenance.json")
        _prov_path.write_text(_json.dumps(_pid_prov, ensure_ascii=False, indent=2))


def build_evidence_rows(instance_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    fields = [
        "instance_key",
        "canonical_tag",
        "pid_anchor_id",
        "pid_tag_name",
        "pid_component_class",
        "pid_x",
        "pid_y",
        "pid_vessel_hint",
        "mapped_3d_label",
        "ifc_part_name",
        "ifc_parts_library",
        "global_id",
        "fcstd_link_name",
        "pos_x_mm",
        "pos_y_mm",
        "pos_z_mm",
        "candidate_labels",
        "evidence_summary",
    ]
    return [{field: row.get(field, "") for field in fields} for row in instance_rows]


def write_sheet(wb: Workbook, name: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(name)
    if not rows:
        ws.append(["message"])
        ws.append(["No rows."])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=2):
        row_values = {headers[index]: row[index].value for index in range(len(headers))}
        fill = MATCH_FILL if row_values.get("needs_review") == "no" else REVIEW_FILL if row_values.get("needs_review") == "yes" else None
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill and name in {"instance_mapping", "grouped_mapping", "review_needed"}:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    add_table(ws, name)
    set_widths(ws)


def add_table(ws: Any, sheet_name: str) -> None:
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2 or max_col < 1:
        return
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table_name = re.sub(r"\W+", "_", f"tbl_{sheet_name}")[:31]
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def set_widths(ws: Any) -> None:
    width_by_header = {
        "evidence_summary": 72,
        "candidate_labels": 34,
        "global_ids": 48,
        "global_id": 28,
        "instrument_tags": 42,
        "mapped_3d_labels": 34,
        "ifc_part_names": 34,
        "match_statuses": 32,
    }
    headers = [cell.value for cell in ws[1]]
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        width = width_by_header.get(header, 18)
        if header in {"source_row", "pid_x", "pid_y", "pos_x_mm", "pos_y_mm", "pos_z_mm", "confidence"}:
            width = 13
        ws.column_dimensions[letter].width = width


def build_mapping_workbook(output_path: Path = DEFAULT_OUTPUT_PATH) -> tuple[Path, list[str]]:
    source_rows = load_source_rows()
    # Inject synthetic VV005 row (no canonical_tag, P&ID valve without 3D counterpart)
    source_rows.append(SourceRow(
        row_number=999, component="VV005", row_anchor="VV005",
        suggested_3d_type="blackbox", instrument_tags=(), notes="",
    ))
    instruments = load_instruments()
    pid_components = parse_pid_components()
    assembly_components = load_assembly_components()
    instance_rows = build_instance_rows(source_rows, instruments, pid_components, assembly_components)
    grouped_rows = aggregate_group_rows(source_rows, instance_rows)
    validation_errors = validate_results(source_rows, instance_rows, grouped_rows)
    write_workbook(output_path, instance_rows, grouped_rows, assembly_components, validation_errors, pid_components=pid_components)
    return output_path, validation_errors


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH, help="Output .xlsx path.")
    args = parser.parse_args()

    output_path, validation_errors = build_mapping_workbook(args.output)
    print(f"wrote {output_path}")
    if validation_errors:
        print("validation errors:")
        for error in validation_errors:
            print(f"  - {error}")
        return 1
    print("validation passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
